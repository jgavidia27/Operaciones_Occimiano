"""
excel_reincidencias_eds.py — Genera Excel de reincidencias EDS del mes.

Replica el formato del archivo modelo:
    HHEE semanal - GPS / Resumen reincidencias por EDS (3).xlsx

- Hoja "Ranking EDS": una fila por EDS con >=3 correctivos MTD.
    Columnas: Cód. Occim | Nombre / Dirección | Cliente | Comuna |
              Llamados | % Cumpl. SLA | Último Llamado | Último Técnico
- Hoja "Detalle OTs": todas las OTs correctivas de esas EDS.
    Columnas: OS Fracttal | N° Aviso | Fecha llamado | Fecha atención |
              Cód. EDS | EDS | Cliente | Técnico | Prioridad
"""

from __future__ import annotations

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


UMBRAL_REINCIDENCIA = 3   # EDS con "Llamados" >= 3 aparecen en el resumen


# ── Estilos ──────────────────────────────────────────────────────────────
_HDR_FILL   = PatternFill("solid", fgColor="1F4E78")
_HDR_FONT   = Font(name="Arial", size=11, bold=True, color="FFFFFF")
_HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_FONT  = Font(name="Arial", size=10)
_CELL_ALIGN = Alignment(vertical="center")
_ALERT_FILL = PatternFill("solid", fgColor="FFF3CD")  # amarillo suave (>=5 llamados)
_BORDER_THIN = Border(
    left=Side(style="thin",  color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin",   color="D9D9D9"),
    bottom=Side(style="thin",color="D9D9D9"),
)


def _clean_str(v) -> str:
    """Convierte a str seguro, sin 'nan'/'None'/'NaT'."""
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    s = str(v).strip()
    if s.lower() in {"nan", "none", "nat", "<na>"}:
        return ""
    return s


def _fecha_str(v) -> str:
    """Formato yyyy-mm-dd, o '' si nulo."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    try:
        ts = pd.to_datetime(v, errors="coerce")
        if pd.isna(ts):
            return ""
        return ts.strftime("%Y-%m-%d")
    except Exception:
        return ""


def _write_headers(ws, headers: list[str], widths: list[int]):
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = _HDR_ALIGN
        cell.border = _BORDER_THIN
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


def _write_row(ws, row_idx: int, values: list, alert: bool = False):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=c, value=v)
        cell.font = _CELL_FONT
        cell.alignment = _CELL_ALIGN
        cell.border = _BORDER_THIN
        if alert:
            cell.fill = _ALERT_FILL


def build_excel_reincidencias(
    df_llamados: pd.DataFrame,
    mes_yyyy_mm: str,
    umbral: int = UMBRAL_REINCIDENCIA,
) -> tuple[bytes, dict]:
    """
    Genera el Excel de reincidencias del mes indicado.

    Retorna (bytes_xlsx, stats) donde stats = {
        "eds_reincidentes":  int,
        "ots_totales":       int,
        "top_eds":           list[(eds_cod, eds_nombre, count)]  # top 5 para el correo
    }
    """
    if df_llamados is None or df_llamados.empty:
        raise ValueError("df_llamados está vacío — no hay datos para generar el reporte.")

    # ── Filtrar mes objetivo ──
    df = df_llamados.copy()
    df["fecha_llamado"] = pd.to_datetime(df["fecha_llamado"], errors="coerce")
    df = df[df["fecha_llamado"].dt.strftime("%Y-%m") == mes_yyyy_mm]

    if df.empty:
        # Excel vacío pero válido (con encabezados y un aviso)
        wb = Workbook()
        ws = wb.active
        ws.title = "Ranking EDS"
        ws["A1"] = f"Sin datos para el mes {mes_yyyy_mm}."
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), {"eds_reincidentes": 0, "ots_totales": 0, "top_eds": []}

    # ── Agrupar por EDS ──
    df["eds_occim"] = df["eds_occim"].fillna("").astype(str)
    df["eds_nombre"] = df["eds_nombre"].fillna("").astype(str)

    grp = df.groupby("eds_occim", dropna=False)
    ranking_rows = []
    for eds_cod, gg in grp:
        if not eds_cod:
            continue
        n_llamados = len(gg)
        if n_llamados < umbral:
            continue
        # % Cumplimiento SLA: 'CUMPLE' y 'EXCEPCION' cuentan como OK. Ignora 'SIN DATOS'.
        cumpl_col = gg.get("cumplimiento", pd.Series([], dtype=str)).fillna("").astype(str)
        cumpl_up = cumpl_col.str.strip().str.upper()
        ok_mask = cumpl_up.isin({"CUMPLE", "EXCEPCION", "EXCEPCIÓN"})
        eval_mask = cumpl_up.isin({"CUMPLE", "NO CUMPLE", "EXCEPCION", "EXCEPCIÓN"})
        n_eval = int(eval_mask.sum())
        pct_sla = round(100.0 * ok_mask.sum() / n_eval, 1) if n_eval else 0.0
        # Último llamado y último técnico
        ult = gg.sort_values("fecha_llamado", ascending=False).iloc[0]
        ranking_rows.append({
            "cod":      _clean_str(eds_cod),
            "nombre":   _clean_str(ult.get("eds_nombre")),
            "cliente":  _clean_str(ult.get("cliente")),
            "comuna":   _clean_str(ult.get("comuna")),
            "llamados": n_llamados,
            "pct_sla":  pct_sla,
            "ultimo":   _fecha_str(ult.get("fecha_llamado")),
            "tecnico":  _clean_str(ult.get("tecnico")),
        })

    # Ordenar por llamados desc, luego por último llamado desc
    ranking_rows.sort(key=lambda r: (-r["llamados"], r["ultimo"]), reverse=False)
    ranking_rows.sort(key=lambda r: (-r["llamados"]))

    eds_reincidentes_codes = {r["cod"] for r in ranking_rows}

    # ── Detalle de OTs (solo de EDS reincidentes) ──
    df_det = df[df["eds_occim"].isin(eds_reincidentes_codes)].copy()
    df_det = df_det.sort_values(["eds_occim", "fecha_llamado"], ascending=[True, True])

    # ── Escribir Excel ──
    wb = Workbook()

    # Hoja 1: Ranking EDS
    ws1 = wb.active
    ws1.title = "Ranking EDS"
    headers1 = ["Cód. Occim", "Nombre / Dirección", "Cliente", "Comuna",
                "Llamados", "% Cumpl. SLA", "Último Llamado", "Último Técnico"]
    widths1 = [14, 45, 16, 20, 10, 14, 15, 32]
    _write_headers(ws1, headers1, widths1)

    for i, r in enumerate(ranking_rows, start=2):
        alert = r["llamados"] >= 5   # amarillo suave para casos críticos
        _write_row(ws1, i, [
            r["cod"], r["nombre"], r["cliente"], r["comuna"],
            r["llamados"], r["pct_sla"], r["ultimo"], r["tecnico"],
        ], alert=alert)
        # Formato numérico columna E (Llamados) y F (% SLA)
        ws1.cell(row=i, column=5).number_format = "0"
        ws1.cell(row=i, column=6).number_format = '0.0"%"'

    # ── Cross-fetch a ordenes_trabajo: causa/tipo/observación técnico ──
    # Fuente única para (a) enriquecer detalle Excel y (b) desglose HTML.
    import os as _os
    import requests as _rq
    ot_info: dict[str, dict] = {}
    _sb_url = _os.getenv("SUPABASE_URL", "")
    _sb_key = _os.getenv("SUPABASE_KEY", "")
    if _sb_url and _sb_key and not df_det.empty and "os_fracttal" in df_det.columns:
        os_list = [str(x) for x in df_det["os_fracttal"].dropna().unique().tolist()]
        for i in range(0, len(os_list), 80):
            chunk = os_list[i:i+80]
            try:
                r = _rq.get(f"{_sb_url}/rest/v1/ordenes_trabajo",
                    params={"select": "id_ot,causa_raiz,tipo_falla,comentario_tecnico",
                            "id_ot": f"in.({','.join(chunk)})", "limit": 200},
                    headers={"apikey": _sb_key,
                             "Authorization": f"Bearer {_sb_key}"},
                    timeout=20)
                if r.status_code == 200:
                    for row in (r.json() or []):
                        ot_info[row["id_ot"]] = row
            except Exception:
                pass

    # Hoja 2: Detalle OTs (con causa/tipo/observación agregados)
    ws2 = wb.create_sheet("Detalle OTs")
    headers2 = ["OS Fracttal", "N° Aviso", "Fecha llamado", "Fecha atención",
                "Cód. EDS", "EDS", "Cliente", "Técnico", "Prioridad",
                "Origen falla", "Tipo falla", "Observación técnico"]
    widths2 = [14, 15, 15, 15, 14, 45, 16, 32, 10, 32, 22, 60]
    _write_headers(ws2, headers2, widths2)

    for i, (_, r) in enumerate(df_det.iterrows(), start=2):
        os_f = _clean_str(r.get("os_fracttal"))
        info = ot_info.get(os_f, {})
        _write_row(ws2, i, [
            os_f,
            _clean_str(r.get("n_llamado")),
            _fecha_str(r.get("fecha_llamado")),
            _fecha_str(r.get("fecha_atencion")),
            _clean_str(r.get("eds_occim")),
            _clean_str(r.get("eds_nombre")),
            _clean_str(r.get("cliente")),
            _clean_str(r.get("tecnico")),
            _clean_str(r.get("prioridad")),
            _clean_str(info.get("causa_raiz")),
            _clean_str(info.get("tipo_falla")),
            _clean_str(info.get("comentario_tecnico")),
        ])

    buf = io.BytesIO()
    wb.save(buf)

    # ── Enriquecer TOP 5 con origen/tipo (moda) + desglose OT por EDS ──
    top5_rows = ranking_rows[:5]
    top5_codes = {r["cod"] for r in top5_rows}
    df_top = df_det[df_det["eds_occim"].isin(top5_codes)].copy() if not df_det.empty else df_det

    from collections import Counter
    causa_por_eds: dict[str, str] = {}
    tipo_por_eds: dict[str, str] = {}
    desglose_por_eds: dict[str, list[dict]] = {}   # para el HTML del correo

    if not df_top.empty:
        df_top = df_top.sort_values("fecha_llamado", ascending=True)
        for eds in top5_codes:
            df_eds = df_top[df_top["eds_occim"] == eds]
            causas, tipos = [], []
            ots_lst = []
            for _, row in df_eds.iterrows():
                os_f = _clean_str(row.get("os_fracttal"))
                info = ot_info.get(os_f, {})
                c = _clean_str(info.get("causa_raiz"))
                t = _clean_str(info.get("tipo_falla"))
                if c and c != "—": causas.append(c)
                if t and t != "—": tipos.append(t)
                ots_lst.append({
                    "os":         os_f,
                    "fecha_llam": _fecha_str(row.get("fecha_llamado")),
                    "fecha_atn":  _fecha_str(row.get("fecha_atencion")),
                    "tecnico":    _clean_str(row.get("tecnico")),
                    "prioridad":  _clean_str(row.get("prioridad")),
                    "origen":     c,
                    "tipo":       t,
                    "obs":        _clean_str(info.get("comentario_tecnico")),
                })
            causa_por_eds[eds] = Counter(causas).most_common(1)[0][0] if causas else "—"
            tipo_por_eds[eds]  = Counter(tipos).most_common(1)[0][0] if tipos else "—"
            desglose_por_eds[eds] = ots_lst

    # top_eds: (cod, nombre, llamados, origen_moda, tipo_moda)
    top_eds = [(r["cod"], r["nombre"], r["llamados"],
                causa_por_eds.get(r["cod"], "—"),
                tipo_por_eds.get(r["cod"],  "—"))
               for r in top5_rows]
    stats = {
        "eds_reincidentes": len(ranking_rows),
        "ots_totales":      len(df_det),
        "top_eds":          top_eds,
        "desglose_por_eds": desglose_por_eds,
    }
    return buf.getvalue(), stats
