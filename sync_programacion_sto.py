"""
sync_programacion_sto.py
========================
Lee el grid de técnicos del Excel '2026 UTILIZACIÓN DE TIEMPO.xlsx'
(hojas mensuales, MAYO 2026 en adelante) y lo sube a Supabase
(tabla programacion_sto) para que el dashboard lo lea rápido sin
depender del Excel de Google Drive.

Cada celda del grid = un registro (fecha, técnico, actividad, color).

Programado en GitHub Actions 2x/día (el Excel se llena a diario;
con 2 syncs es suficiente para esta vista poco demandante).

Uso:
    python sync_programacion_sto.py                 # mayo en adelante
    python sync_programacion_sto.py --desde-mes 3   # desde marzo
"""

import argparse
import json
import sys
from datetime import datetime, timezone

import requests
from openpyxl import load_workbook

# ── Rutas y credenciales ────────────────────────────────────────────────────
_PATH_XLSX = ("G:/.shortcut-targets-by-id/15zHnoU5VZlkOwYc6EBziNcnS-sAAtwtk/"
              "OPERACIONES/OPERACIONES/2026 UTILIZACIÓN DE TIEMPO.xlsx")

import os
SUPABASE_URL = os.getenv("SUPABASE_URL", "https://puefgkyjghwwgdfxbrex.supabase.co")
SUPABASE_KEY = os.getenv("SUPABASE_KEY",
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6In"
    "B1ZWZna3lqZ2h3d2dkZnhicmV4Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4"
    "MDcxMTk0OCwiZXhwIjoyMDk2Mjg3OTQ4fQ.keB15jRQ7ahuXiDHktFC_Yi000XUlExjDMqTuC6VLgw")
SUPABASE_TABLE = "programacion_sto"

_MESES = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
          "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]
_EXCLUIR_COL = {"ALEXIS"}   # columnas que no son técnicos operativos


def log(msg, lvl="INFO"):
    ts = datetime.now().strftime("%H:%M:%S")
    tags = {"INFO":"    ","OK":"[OK]","WARN":"[!] ","ERR":"[X] ","PROG":"--> "}
    print(f"[{ts}] {tags.get(lvl,'    ')} {msg}", flush=True)


def parse_hoja(ws, mes_hoja):
    """Extrae registros (fecha, técnico, actividad, color) de una hoja mensual."""
    # Encabezados de técnicos (fila 1, cols B..R)
    tecnicos = {}
    for c in range(2, min(19, ws.max_column + 1)):
        v = ws.cell(1, c).value
        if v and str(v).strip():
            nom = str(v).strip()
            if nom.upper() in _EXCLUIR_COL:
                continue
            tecnicos[c] = nom
    if not tecnicos:
        return []

    registros = []
    for r in range(2, ws.max_row + 1):
        fecha = ws.cell(r, 1).value
        if not fecha or not hasattr(fecha, "year"):
            continue
        for c, tec in tecnicos.items():
            cell = ws.cell(r, c)
            val = cell.value
            actividad = str(val).strip() if val is not None else ""
            # Color de fondo
            color = None
            if cell.fill and cell.fill.start_color:
                try:
                    rgb = cell.fill.start_color.rgb
                    if rgb and isinstance(rgb, str) and rgb not in ("00000000", "FFFFFFFF"):
                        color = "#" + rgb[-6:]
                except Exception:
                    pass
            # Solo guardar filas con algún contenido (actividad o color)
            if not actividad and not color:
                continue
            registros.append({
                "fecha":       fecha.date().isoformat(),
                "tecnico":     tec,
                "actividad":   actividad or None,
                "color_excel": color,
                "mes_hoja":    mes_hoja,
                "synced_at":   datetime.now(timezone.utc).isoformat(),
            })
    return registros


def upsert(registros):
    if not registros:
        return 0
    headers = {
        "apikey":        SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type":  "application/json",
        "Prefer":        "resolution=merge-duplicates,return=minimal",
    }
    total = 0
    for i in range(0, len(registros), 500):
        batch = registros[i:i+500]
        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/{SUPABASE_TABLE}",
            headers=headers,
            data=json.dumps(batch, ensure_ascii=False).encode("utf-8"),
            timeout=30,
        )
        if r.status_code not in (200, 201, 204):
            log(f"Error Supabase {r.status_code}: {r.text[:250]}", "ERR")
            r.raise_for_status()
        total += len(batch)
    return total


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--desde-mes", type=int, default=4,
                    help="Mes desde el cual sincronizar (1-12). Default 4=abril.")
    ap.add_argument("--anio", type=int, default=datetime.now().year)
    args = ap.parse_args()

    log(f"Sync Programación STO — desde {_MESES[args.desde_mes-1]} {args.anio}", "PROG")
    try:
        wb = load_workbook(_PATH_XLSX, data_only=True)
    except Exception as e:
        log(f"No se pudo abrir el Excel: {e}", "ERR")
        sys.exit(1)

    mes_actual = datetime.now().month
    hojas_objetivo = [f"{_MESES[m-1]} {args.anio}"
                      for m in range(args.desde_mes, mes_actual + 1)]

    total_general = 0
    for hoja in hojas_objetivo:
        if hoja not in wb.sheetnames:
            log(f"Hoja '{hoja}' no existe, saltando", "WARN")
            continue
        regs = parse_hoja(wb[hoja], hoja)
        n = upsert(regs)
        total_general += n
        log(f"{hoja}: {n} registros sincronizados", "OK")

    log(f"COMPLETADO — {total_general:,} registros en '{SUPABASE_TABLE}'", "OK")


if __name__ == "__main__":
    main()
