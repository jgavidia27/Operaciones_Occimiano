import re
import os
import json
import base64
import unicodedata
from pathlib import Path
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime

from auth import (
    init_cookie_manager, is_authenticated, try_login, logout,
    request_password_reset, consume_reset_token,
    get_pending_reset_token_from_url, clear_reset_token_from_url,
)

from api import (
    load_work_orders, load_third_parties,
    load_work_orders_subtasks, load_wo_resources,
    load_meters, load_meters_reading,
    load_work_requests, load_items_catalog,
)
from data import (
    build_work_orders_df, build_third_parties_df, station_summary, CLIENT_COLORS,
    build_kpi_llenado_df, score_llenado_por_ot, score_llenado_por_tecnico,
    build_reincidencias, build_numeral_historial, analizar_secuencias, CAT_LABEL,
    NUMERAL_MOTIVO_LABEL, aplicar_numerales_subtarea,
    GRUPOS_TERRENO, get_grupo_tecnico, TECNICOS_NO_APLICA,
    SENIORS, get_senior_team_members,
    build_meters_fichas_df, enrich_fichas_with_readings,
)
from gdrive import (
    load_listado_eds, load_all_llamados, load_llamados_fracttal, kpis_por_eds,
    load_llamados_copec, load_llamados_esmax, load_llamados_shell,
    resolve_llamados_eds_codes,
    load_utilizacion_tiempo, list_utilizacion_sheets,
    load_mtto_realizados_planilla,
    load_base_tecnicos, build_tech_name_maps,
    CATEGORY_COLORS_UTIL, classify_task_line,
    SLA_HOURS, SLA_DEFAULT, TECNICOS_OCCIMIANO_FULL, TECH_NAME_MAP,
)

# ── SUPABASE — nueva capa de datos (reemplaza Fracttal API y Excel) ───────────
from supabase_client import (
    load_work_orders_supabase,
    load_numerales_subtarea_supabase,
    load_listado_eds_supabase,
    load_tecnicos_supabase,
    load_equipos_supabase,
    load_preventivas_supabase,
    load_correctivas_supabase,
    load_all_llamados_supabase,
    load_sla_umbrales_supabase,
    load_cotalker_index_supabase,
    load_notas_tarea_index,
    load_ots_en_vivo_supabase,
)
_USE_SUPABASE = True   # ← cambiar a False para volver a Fracttal/Excel

# ── Caché en disco para build_kpi_llenado_df (≈9s sin caché) ────────────────
_KPI_CACHE_VERSION = "v23-equipo-nombre"  # bump para invalidar disco al cambiar data.py


def _filtro_ot_input(key: str, columna_ot: str = "OT") -> str:
    """Caja de texto compacta para filtrar tablas por número de OT.
    Devuelve el query normalizado en MAYÚSCULAS sin espacios; "" si no filtra.
    Acepta entradas como "37852", "os-37852", "OS-37852, OS-38066" (lista).
    """
    raw = st.text_input(
        f"🔎 Filtrar por OT", "", key=key,
        placeholder="OS-37852  ·  ó solo el número (37852)  ·  separar varias con coma",
        help="Filtra esta tabla por una o varias OTs. Vacío = todas.",
    )
    return raw.strip().upper()


def _aplicar_filtro_ot(df, query: str, col: str = "OT"):
    """Aplica el filtro de OT al DataFrame. Soporta lista separada por comas
    y matching por substring (37852 encuentra OS-37852)."""
    if not query or col not in df.columns:
        return df
    tokens = [t.strip() for t in query.replace(";", ",").split(",") if t.strip()]
    if not tokens:
        return df
    serie = df[col].astype(str).str.upper()
    mask = pd.Series(False, index=df.index)
    for t in tokens:
        mask |= serie.str.contains(t, regex=False, na=False)
    return df[mask]


# Limpia los encabezados verbosos del comentario consolidado del técnico (data
# antigua del sync trae "DESCRIPCIÓN DE LA FALLA ENCONTRADA: xxx | TRABAJO
# REALIZADO ...: yyy"). Devuelve solo el contenido, separado por " | ".
import re as _re
_COMENT_HEADER_RE = _re.compile(
    r"(?:DESCRIPCI[ÓO]N DE LA FALLA[^:|]*|"
    r"TRABAJO REALIZADO[^:|]*|"
    r"OBSERVACI[ÓO]NES?|"
    r"PENDIENTES?)\s*:\s*",
    flags=_re.IGNORECASE,
)

def _strip_comentario_headers(txt) -> str:
    s = str(txt or "").strip()
    if not s or s == "—":
        return "—"
    return _COMENT_HEADER_RE.sub("", s).strip(" |") or "—"

@st.cache_data(ttl=1800, show_spinner=False, persist="disk")
def _cached_build_kpi_llenado(raw_wo: list, cache_v: str = _KPI_CACHE_VERSION) -> pd.DataFrame:
    """Wrapper con caché persistente en disco. Sobrevive reinicios de Streamlit.
    cache_v: bumpearlo fuerza invalidación del caché persistente."""
    return build_kpi_llenado_df(raw_wo)

# ── Session-state cache helper ───────────────────────────────────────────────
def _sc(key: str, sig: str, builder):
    """
    Guarda en st.session_state[key] el resultado de builder().
    Solo recalcula cuando `sig` cambia (nueva carga de datos).
    Evita recomputar DataFrames pesados en cada rerun por filtros.
    """
    if st.session_state.get(f"_sc_sig_{key}") != sig or key not in st.session_state:
        st.session_state[key]              = builder()
        st.session_state[f"_sc_sig_{key}"] = sig
    return st.session_state[key]


# ── Station name helper ───────────────────────────────────────────────────────
def _get_addr_numbers(text: str) -> set:
    """Extract 3-or-more-digit numbers from an address string."""
    return set(re.findall(r'\b\d{3,}\b', str(text).upper()))


@st.cache_data(show_spinner=False)
def build_station_name_map(df_tp_: pd.DataFrame, df_eds_: pd.DataFrame) -> dict:
    """
    Returns {eds_occim: station_name} by matching Fracttal third-parties
    to the EDS listado via street number in the address field.
    Only produces a mapping when a single candidate is found (safe matches only).
    """
    if df_tp_.empty or df_eds_.empty:
        return {}

    # Index EDS by address numbers for O(1) lookup
    eds_by_num: dict = {}
    for _, row in df_eds_.iterrows():
        for n in _get_addr_numbers(row.get("direccion", "")):
            eds_by_num.setdefault(n, []).append(row)

    result: dict = {}
    for _, tp in df_tp_.iterrows():
        tp_name = str(tp.get("name", "")).strip()
        if not tp_name:
            continue
        tp_nums = _get_addr_numbers(str(tp.get("address", "") or ""))
        if not tp_nums:
            continue

        # Collect EDS candidates that share a street number
        candidates: list = []
        for n in tp_nums:
            for c in eds_by_num.get(n, []):
                if str(c["eds_occim"]) not in {str(x["eds_occim"]) for x in candidates}:
                    candidates.append(c)

        # Narrow by client prefix (COPEC, SHELL, ESMAX …)
        prefix = tp_name.split()[0].upper()
        filtered = [c for c in candidates
                    if str(c.get("cliente", "")).upper().startswith(prefix[:4])]
        final = filtered if filtered else candidates

        if len(final) == 1:
            result[str(final[0]["eds_occim"])] = tp_name

    return result


# ── Wallpaper / Theme ─────────────────────────────────────────────────────────
_APP_DIR = os.path.dirname(os.path.abspath(__file__))

@st.cache_resource(show_spinner=False)
def _load_wallpaper_b64(theme: str) -> str:
    """Carga wallpaper_{theme}.jpg/.png como data-URI base64 (cacheado)."""
    for ext in ("jpg", "jpeg", "png"):
        path = os.path.join(_APP_DIR, f"wallpaper_{theme}.{ext}")
        if os.path.exists(path):
            with open(path, "rb") as fh:
                raw = fh.read()
            mime = "image/jpeg" if ext in ("jpg", "jpeg") else "image/png"
            return f"data:{mime};base64,{base64.b64encode(raw).decode()}"
    return ""


@st.cache_resource(show_spinner=False)
def _load_logo_b64() -> str:
    """Carga logo.png/.jpg como data-URI base64 (cacheado). PNG tiene prioridad."""
    for name in ("logo.png", "logo_occim.png", "logo_login.png", "logo_dashboard.jpg", "logo_occim.jpg", "logo.jpg", "logo.jpeg"):
        path = os.path.join(_APP_DIR, name)
        if os.path.exists(path):
            with open(path, "rb") as fh:
                raw = fh.read()
            mime = "image/jpeg" if name.endswith((".jpg", ".jpeg")) else "image/png"
            return f"data:{mime};base64,{base64.b64encode(raw).decode()}"
    return ""


def _inject_theme(theme: str) -> None:
    """
    Inyecta CSS completo de tema oscuro/claro con efecto marca de agua.
    El wallpaper aparece muy difuminado (8-12 % de opacidad) bajo el color
    de fondo del tema, exactamente como el cambio de tema en Google Gemini.
    """
    uri = _load_wallpaper_b64(theme)
    wp_layer  = f'url("{uri}")' if uri else "none"

    # ── Paleta según tema ─────────────────────────────────────────────────────
    if theme == "dark":
        bg_rgba       = "10, 18, 38,  0.93"   # overlay 93 % → wallpaper al 7 %
        app_bg_solid  = "#0a1226"
        card_bg       = "#111f38"
        card_border   = "#1e3356"
        text_color    = "#e2e8f0"
        text_muted    = "#94a3b8"
        sidebar_bg    = "#0d1427"
        input_bg      = "#111f38"
        input_border  = "#1e3356"
        tab_bg        = "#111f38"
        tab_color     = "#94a3b8"
        tab_active_c  = "#60a5fa"
        tab_active_bd = "#60a5fa"
        divider_c     = "#1e3356"
        section_color = "#cbd5e1"
        btn_bg        = "#1e3356"
        btn_hover     = "#2d4a73"
        btn_color     = "#e2e8f0"
        metric_bg     = "#111f38"
        toggle_bg     = "rgba(10,18,38,0.75)"
        toggle_active = "rgba(255,255,255,0.18)"
        toggle_icon   = "☀️"           # ícono para pasar a Light
        toggle_href   = "?theme=light"
    else:  # light
        bg_rgba       = "248, 250, 252, 0.88"  # overlay 88 % → wallpaper al 12 %
        app_bg_solid  = "#f8fafc"
        card_bg       = "#ffffff"
        card_border   = "#e2e8f0"
        text_color    = "#1e293b"
        text_muted    = "#64748b"
        sidebar_bg    = "#0d1427"
        input_bg      = "#ffffff"
        input_border  = "#cbd5e1"
        tab_bg        = "#f1f5f9"
        tab_color     = "#64748b"
        tab_active_c  = "#3b82f6"
        tab_active_bd = "#3b82f6"
        divider_c     = "#e2e8f0"
        section_color = "#1e293b"
        btn_bg        = "#f1f5f9"
        btn_hover     = "#e2e8f0"
        btn_color     = "#1e293b"
        metric_bg     = "#ffffff"
        toggle_bg     = "rgba(15,32,68,0.72)"
        toggle_active = "rgba(255,255,255,0.22)"
        toggle_icon   = "🌙"           # ícono para pasar a Dark
        toggle_href   = "?theme=dark"

    st.markdown(
        f"""
        <style>
        /* ══════════════════════════════════════════════════════
           FONDO — color sólido en html/body como base firme
           ══════════════════════════════════════════════════════ */
        html, body {{
            background-color: {app_bg_solid} !important;
        }}

        /* Wallpaper como marca de agua sobre el color sólido */
        .stApp {{
            background-color: transparent !important;
            background-image:
                linear-gradient(rgba({bg_rgba}), rgba({bg_rgba})),
                {wp_layer} !important;
            background-size: cover !important;
            background-position: center top !important;
            background-attachment: fixed !important;
            background-repeat: no-repeat !important;
            min-height: 100vh !important;
        }}

        /* ══════════════════════════════════════════════════════
           CONTENIDO PRINCIPAL
           ══════════════════════════════════════════════════════ */
        [data-testid="stAppViewContainer"] {{
            background-color: transparent !important;
            background-image: none !important;
        }}
        [data-testid="stAppViewContainer"] > section.main {{
            background-color: transparent !important;
            background-image: none !important;
        }}
        [data-testid="stAppViewContainer"] > section.main > div.block-container {{
            background-color: transparent !important;
            background-image: none !important;
            padding-top: 0.5rem !important;
        }}
        /* Selector adicional para Streamlit ≥1.32 */
        [data-testid="stMainBlockContainer"] {{
            padding-top: 0.5rem !important;
        }}
        .main .block-container, div.block-container {{
            padding-top: 0.5rem !important;
        }}

        /* Texto general */
        .stApp, .stApp p, .stApp div, .stApp span,
        .stApp li, .stApp h1, .stApp h2, .stApp h3 {{
            color: {text_color} !important;
        }}
        .stApp .stCaption, .stApp small {{
            color: {text_muted} !important;
        }}

        /* Dividers */
        hr {{ border-color: {divider_c} !important; }}

        /* Section headers del dashboard */
        .section-header {{
            color: {section_color} !important;
            border-bottom-color: {divider_c} !important;
        }}

        /* ── Métricas ────────────────────────────────────────── */
        [data-testid="metric-container"] {{
            background: {metric_bg} !important;
            border: 1px solid {card_border} !important;
            border-radius: 8px !important;
            padding: 14px !important;
        }}
        [data-testid="stMetricValue"] {{ color: {text_color} !important; }}
        [data-testid="stMetricLabel"] {{ color: {text_muted} !important; }}

        /* ── Tabs — colores de tema (tamaño y hover en CSS estático) ─── */
        [data-baseweb="tab-list"] {{
            background: {tab_bg} !important;
            border-radius: 10px 10px 0 0 !important;
        }}
        /* Color base del texto según tema */
        [data-baseweb="tab"] {{
            color: {tab_color} !important;
            background: transparent !important;
        }}
        /* Tab inactivo hover — color de texto según tema */
        [data-testid="stTabs"] button[data-baseweb="tab"]:not([aria-selected="true"]):hover {{
            color: {tab_active_c} !important;
        }}
        /* Tab activo — color y borde según tema */
        [data-baseweb="tab"][aria-selected="true"] {{
            color: {tab_active_c} !important;
            border-bottom: 3px solid {tab_active_bd} !important;
        }}

        /* ── Selectboxes / inputs ─────────────────────────────── */
        [data-baseweb="select"] > div,
        [data-baseweb="input"] > div,
        [data-testid="stTextInput"] > div > div {{
            background: {input_bg} !important;
            border-color: {input_border} !important;
            color: {text_color} !important;
        }}
        [data-baseweb="select"] * {{ color: {text_color} !important; }}

        /* ── Botones ──────────────────────────────────────────── */
        .stButton > button {{
            background: {btn_bg} !important;
            color: {btn_color} !important;
            border: 1px solid {card_border} !important;
        }}
        .stButton > button:hover {{
            background: {btn_hover} !important;
            border-color: {tab_active_bd} !important;
        }}

        /* ── Expanders ────────────────────────────────────────── */
        [data-testid="stExpander"] {{
            background: {card_bg} !important;
            border: 1px solid {card_border} !important;
        }}

        /* ── DataFrames ───────────────────────────────────────── */
        [data-testid="stDataFrame"],
        [data-testid="stDataFrame"] > div,
        [data-testid="stDataFrameResizable"],
        [data-testid="stDataFrameResizable"] > div {{
            background: {card_bg} !important;
            border-radius: 8px !important;
        }}
        /* Forzar colores de celda y encabezado dentro del grid */
        [data-testid="stDataFrame"] [role="gridcell"] {{
            background-color: {card_bg} !important;
            color: {text_color} !important;
        }}
        [data-testid="stDataFrame"] [role="columnheader"],
        [data-testid="stDataFrame"] [role="rowheader"] {{
            background-color: #0C2540 !important;
            color: #ffffff !important;
        }}
        /* Scrollbars en dark mode */
        [data-testid="stDataFrame"] ::-webkit-scrollbar-thumb {{
            background: {card_border} !important;
            border-radius: 4px !important;
        }}
        [data-testid="stDataFrame"] ::-webkit-scrollbar-track {{
            background: {card_bg} !important;
        }}

        /* ── Spinner / info boxes ─────────────────────────────── */
        [data-testid="stAlert"] {{
            background: {card_bg} !important;
            border-color: {card_border} !important;
            color: {text_color} !important;
        }}

        /* ══════════════════════════════════════════════════════
           SIDEBAR
           ══════════════════════════════════════════════════════ */
        [data-testid="stSidebar"],
        [data-testid="stSidebar"] > div,
        [data-testid="stSidebar"] > div > div,
        [data-testid="stSidebar"] section,
        [data-testid="stSidebarContent"] {{
            background: {sidebar_bg} !important;
        }}
        /* Sidebar: texto siempre blanco (anula .stApp span con misma especificidad pero regla posterior) */
        [data-testid="stSidebar"] span,
        [data-testid="stSidebar"] p,
        [data-testid="stSidebar"] div,
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] li,
        [data-testid="stSidebar"] h1,
        [data-testid="stSidebar"] h2,
        [data-testid="stSidebar"] h3 {{
            color: #f1f5f9 !important;
        }}

        /* Desktop: sidebar siempre visible */
        @media screen and (min-width: 768px) {{
            section[data-testid="stSidebar"] {{
                transform: translateX(0) !important;
                visibility: visible !important;
            }}
            [data-testid="stSidebarCollapseButton"] {{
                display: none !important;
            }}
        }}
        /* Mobile: botón hamburguesa siempre visible para abrir/cerrar */
        @media screen and (max-width: 767px) {{
            [data-testid="stSidebarCollapseButton"] {{
                display: flex !important;
            }}
        }}

        button[title="View fullscreen"] {{ display: none !important; }}

        </style>
        """,
        unsafe_allow_html=True,
    )


def _inject_toggle(theme: str) -> None:
    """
    Inyecta el toggle ☀️/🌙 en el <body> del parent via iframe JS.
    st.components.v1.html sí ejecuta scripts (a diferencia de st.markdown).
    """
    if theme == "dark":
        bg_css     = "rgba(10,18,38,0.85)"
        active_css = "rgba(255,255,255,0.18)"
    else:
        bg_css     = "rgba(15,32,68,0.75)"
        active_css = "rgba(255,255,255,0.22)"

    sun_active  = "active-theme" if theme == "light" else ""
    moon_active = "active-theme" if theme == "dark"  else ""

    components.html(
        f"""
        <script>
        (function() {{
            var p = window.parent;
            if (!p || !p.document) return;

            // Remove previous toggle if re-render
            var old = p.document.getElementById('occ-theme-toggle');
            if (old) old.remove();

            // Build toggle
            var d = p.document.createElement('div');
            d.id = 'occ-theme-toggle';
            d.style.cssText =
                'position:fixed;top:8px;right:80px;z-index:2147483647;' +
                'display:flex;align-items:center;gap:2px;' +
                'background:{bg_css};border-radius:20px;padding:4px 10px;' +
                'backdrop-filter:blur(8px);-webkit-backdrop-filter:blur(8px);' +
                'box-shadow:0 2px 12px rgba(0,0,0,0.35);';

            function makeLink(href, emoji, title, activeClass) {{
                var a = p.document.createElement('a');
                a.href = href;
                a.title = title;
                a.textContent = emoji;
                a.style.cssText =
                    'text-decoration:none;font-size:1.2rem;padding:3px 8px;' +
                    'border-radius:14px;transition:background 0.15s;';
                if (activeClass) a.style.background = '{active_css}';
                return a;
            }}

            d.appendChild(makeLink('?theme=light', '☀️', 'Tema claro',  '{sun_active}'));
            d.appendChild(makeLink('?theme=dark',  '🌙', 'Tema oscuro', '{moon_active}'));
            p.document.body.appendChild(d);
        }})();
        </script>
        """,
        height=0,
        scrolling=False,
    )


# ── Pantalla de inicio de sesión ──────────────────────────────────────────────
def _get_logo_path() -> str:
    """Retorna la ruta al archivo de logo."""
    for name in ("logo.png", "logo_occim.png", "logo_login.png", "logo_occim.jpg", "logo.jpg", "logo.jpeg"):
        path = os.path.join(_APP_DIR, name)
        if os.path.exists(path):
            return path
    return ""


def _show_login_page() -> None:
    """Pantalla de login: imagen hero a la izquierda + card de login a la derecha."""
    import base64 as _b64

    # ── Cargar imagen hero como base64 ─────────────────────────────────────
    _hero_data = ""
    for _hname in ("imagen_inicio.jpg", "imagen_inicio.png"):
        _hpath = os.path.join(_APP_DIR, _hname)
        if os.path.exists(_hpath):
            _mime = "jpeg" if _hname.endswith(".jpg") else "png"
            with open(_hpath, "rb") as _hf:
                _hero_data = f"data:image/{_mime};base64,{_b64.b64encode(_hf.read()).decode()}"
            break

    # Fondo: imagen real + gradiente oscuro hacia la derecha
    if _hero_data:
        _bg = (
            "background-image: "
            "linear-gradient(to right, transparent 0%, transparent 50%, "
            "rgba(13,20,39,0.45) 63%, rgba(13,20,39,0.95) 74%, #0d1427 84%), "
            f'url("{_hero_data}"); '
            "background-size: cover; background-position: center; "
            "background-repeat: no-repeat;"
        )
    else:
        _bg = "background: #0d1427;"

    st.markdown(f"""
    <style>
    /* ── Ocultar chrome de Streamlit ── */
    [data-testid="stSidebar"], [data-testid="stHeader"],
    [data-testid="stToolbar"], [data-testid="stMainMenu"],
    footer, .stDeployButton {{ display: none !important; }}

    /* ── Imagen hero pantalla completa ── */
    .stApp {{
        {_bg}
        background-color: #0d1427 !important;
    }}
    [data-testid="stAppViewContainer"] {{
        background: transparent !important;
    }}

    /* ── Centrado vertical del form ── */
    [data-testid="stMain"] > div {{
        padding-top: 0 !important;
        padding-bottom: 0 !important;
    }}
    .block-container {{
        padding-top: 0 !important;
        padding-bottom: 0 !important;
        max-width: 100% !important;
    }}

    /* ── Card blanco ── */
    [data-testid="stForm"] {{
        background: #ffffff !important;
        border-radius: 16px !important;
        border: none !important;
        padding: 2rem 2rem 1.75rem !important;
        box-shadow: 0 20px 70px rgba(0,0,0,0.65) !important;
        max-width: 360px !important;
        margin: 0 auto !important;
    }}

    /* ── Logo centrado dentro del card ── */
    [data-testid="stForm"] [data-testid="stImage"] {{
        display: flex !important;
        justify-content: center !important;
        margin-bottom: 0.25rem !important;
    }}
    [data-testid="stForm"] [data-testid="stImage"] img {{
        max-height: 56px !important;
        width: auto !important;
        object-fit: contain !important;
    }}

    /* ── Labels ── */
    [data-testid="stForm"] label {{
        color: #374151 !important;
        font-size: 0.8rem !important;
        font-weight: 600 !important;
        letter-spacing: 0.4px !important;
        text-transform: uppercase !important;
    }}

    /* ── Inputs ── */
    [data-testid="stForm"] input {{
        background: #f9fafb !important;
        border: 1.5px solid #e5e7eb !important;
        color: #111827 !important;
        border-radius: 8px !important;
        font-size: 0.95rem !important;
    }}
    [data-testid="stForm"] input:focus {{
        border-color: #0d1427 !important;
        box-shadow: 0 0 0 3px rgba(13,26,39,0.1) !important;
    }}
    [data-testid="stForm"] input::placeholder {{ color: #9ca3af !important; }}

    /* ── Botón submit ── */
    [data-testid="stForm"] button[kind="primaryFormSubmit"],
    [data-testid="stForm"] button[data-testid="baseButton-primaryFormSubmit"] {{
        background: #0d1427 !important;
        border: none !important;
        color: #fff !important;
        font-weight: 600 !important;
        font-size: 0.95rem !important;
        height: 2.85rem !important;
        border-radius: 9px !important;
        letter-spacing: 0.3px !important;
        transition: background 0.2s !important;
    }}
    [data-testid="stForm"] button[kind="primaryFormSubmit"]:hover {{
        background: #1a3066 !important;
    }}

    /* ── Error dentro del card ── */
    [data-testid="stForm"] [data-testid="stAlert"] {{
        background: #fef2f2 !important;
        border: 1px solid #fecaca !important;
        border-radius: 8px !important;
        color: #991b1b !important;
    }}
    </style>
    """, unsafe_allow_html=True)

    # ── Layout: espacio izquierdo (imagen) | form | padding derecho ────────
    _, _col, _ = st.columns([3.2, 1, 0.05])
    with _col:
        # Espaciador para centrado vertical (~25% desde arriba)
        st.markdown("<div style='height:22vh; min-height:60px;'></div>", unsafe_allow_html=True)

        # Logo de login: preferir logo_login.png, luego los genéricos
        _login_logo = os.path.join(_APP_DIR, "logo_login.png")
        _logo_path = _login_logo if os.path.exists(_login_logo) else _get_logo_path()

        with st.form("occim_login"):

            # Logo centrado
            _lc1, _lc2, _lc3 = st.columns([1, 2, 1])
            with _lc2:
                if _logo_path:
                    st.image(_logo_path, use_container_width=True)
                else:
                    st.markdown(
                        '<p style="text-align:center;font-size:2rem;font-weight:900;'
                        'color:#0d1427;letter-spacing:3px;margin:0 0 0.5rem;">OCCIM</p>',
                        unsafe_allow_html=True,
                    )

            # Subtítulo bajo el logo
            st.markdown("""
            <p style="text-align:center;color:#1a3066;font-size:0.78rem;font-weight:600;
                letter-spacing:0.6px;text-transform:uppercase;margin:0.1rem 0 0.2rem;">
                Indicadores Operacionales</p>
            """, unsafe_allow_html=True)

            # Título
            st.markdown("""
            <div style="text-align:center;padding:0.75rem 0 1.4rem;">
                <p style="color:#0d1427;font-size:1.45rem;font-weight:700;
                    margin:0;line-height:1.2;">Iniciar sesión</p>
            </div>
            """, unsafe_allow_html=True)

            # Error
            if st.session_state.get("_login_failed"):
                st.error("Correo no autorizado o contraseña incorrecta.", icon="⚠️")

            _login_email = st.text_input(
                "Correo electrónico",
                placeholder="nombre@occimiano.cl",
                key="_lf_email",
            )
            _login_pw = st.text_input(
                "Contraseña",
                placeholder="••••••••",
                type="password",
                key="_lf_pw",
            )
            st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
            _submitted = st.form_submit_button(
                "Ingresar  →",
                use_container_width=True,
                type="primary",
            )

        if _submitted:
            if try_login(_login_email, _login_pw):
                st.session_state.pop("_login_failed", None)
                st.rerun()
            else:
                st.session_state["_login_failed"] = True
                st.rerun()

        # ── Link "Olvidé mi contraseña" ─────────────────────────────────
        _, _link_col, _ = st.columns([1, 2, 1])
        with _link_col:
            if st.button("¿Olvidó su contraseña?",
                         key="_btn_forgot", use_container_width=True):
                st.session_state["_auth_view"] = "forgot"
                st.rerun()


def _show_forgot_page() -> None:
    """Pantalla 'Ingresa tu correo' para iniciar recuperación de contraseña."""
    _ = _login_page_wrapper_start("Recuperar contraseña")
    with st.form("occim_forgot"):
        st.caption("Ingresa el correo asociado a tu cuenta. Te enviaremos un "
                   "enlace para crear una nueva contraseña.")
        _email = st.text_input("Correo electrónico",
                               placeholder="nombre@occimiano.cl",
                               key="_lf_forgot_email")
        _send = st.form_submit_button("Enviar enlace  →",
                                      use_container_width=True, type="primary")
    if _send:
        ok, msg = request_password_reset(_email, proposito="reset")
        if ok:
            st.success(msg)
        else:
            st.error(msg)
    _, _back_col, _ = st.columns([1, 2, 1])
    with _back_col:
        if st.button("← Volver al login", key="_btn_forgot_back",
                     use_container_width=True):
            st.session_state["_auth_view"] = "login"
            st.rerun()


def _show_reset_page(token: str) -> None:
    """Pantalla 'Setea tu nueva contraseña' a la que se llega vía link del correo."""
    _ = _login_page_wrapper_start("Define tu nueva contraseña")
    with st.form("occim_reset"):
        st.caption("Elige una contraseña de al menos 8 caracteres. Confírmala "
                   "abajo y haz clic en Guardar.")
        _pw1 = st.text_input("Nueva contraseña", type="password",
                             key="_lf_reset_pw1")
        _pw2 = st.text_input("Confirmar contraseña", type="password",
                             key="_lf_reset_pw2")
        _save = st.form_submit_button("Guardar  →",
                                      use_container_width=True, type="primary")
    if _save:
        if _pw1 != _pw2:
            st.error("Las contraseñas no coinciden.")
        elif len(_pw1) < 8:
            st.error("La contraseña debe tener al menos 8 caracteres.")
        else:
            ok, msg = consume_reset_token(token, _pw1)
            if ok:
                st.success(msg + " Redirigiendo al login…")
                clear_reset_token_from_url()
                st.session_state["_auth_view"] = "login"
                import time as _t; _t.sleep(2)
                st.rerun()
            else:
                st.error(msg)


def _login_page_wrapper_start(title: str) -> None:
    """Header visual común a las pantallas pre-login (login/forgot/reset)."""
    _logo_b64 = _load_logo_b64()
    _, _center, _ = st.columns([1, 2, 1])
    with _center:
        st.markdown(f"""
        <div style="text-align:center;margin:42px 0 8px 0;">
            {f'<img src="data:image/png;base64,{_logo_b64}" style="height:60px;">' if _logo_b64 else ''}
            <p style="color:#01798A;letter-spacing:0.18em;font-weight:700;
                margin:14px 0 4px 0;font-size:0.78rem;">INDICADORES OPERACIONALES</p>
            <p style="color:#0d1427;font-size:1.45rem;font-weight:700;
                margin:0;line-height:1.2;">{title}</p>
        </div>
        """, unsafe_allow_html=True)


st.set_page_config(
    page_title="Indicadores Operacionales - Occim",
    page_icon="🎖️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Cookie manager — debe inicializarse ANTES de cualquier chequeo de sesión ─
# init_cookie_manager() retorna False en el primer render (cookies aún cargando)
# En ese caso detenemos la ejecución y esperamos el re-run automático del componente
if not init_cookie_manager():
    st.stop()

# ── Logout via query param (?_lo=1) ──────────────────────────────────────────
if st.query_params.get("_lo") == "1":
    logout()
    try:
        del st.query_params["_lo"]
    except Exception:
        pass
    st.rerun()

# ── Autenticación — mostrar login y detener si no hay sesión ─────────────────
if not is_authenticated():
    # Router de pantallas pre-login:
    #   1. ?rst=<token>  → setear nueva contraseña (link del correo)
    #   2. _auth_view = "forgot" → ingresar correo para recuperar
    #   3. (default)     → login normal
    _reset_token = get_pending_reset_token_from_url()
    _auth_view   = st.session_state.get("_auth_view", "login")
    if _reset_token:
        _show_reset_page(_reset_token)
    elif _auth_view == "forgot":
        _show_forgot_page()
    else:
        _show_login_page()
    st.stop()

# ── Tema (dark / light) — leer desde URL query param ─────────────────────────
# Limpiar query param heredado del sistema anterior (URL ?theme=dark/light)
# ya no se usa — el tema vive solo en session_state
if "theme" in st.query_params:
    try:
        del st.query_params["theme"]
    except Exception:
        pass
_current_theme = st.session_state.get("_theme", "light")
_inject_theme(_current_theme)

# ── Indicador discreto de capacidad OTs (esquina sup. derecha) ───────────────
_wo_2026 = st.session_state.get("_wo_2026_count", 0)
if _wo_2026:
    _wo_pct = min(round(_wo_2026 / 20_000 * 100), 100)
    _wo_clr = "#22c55e" if _wo_pct < 70 else ("#f59e0b" if _wo_pct < 90 else "#ef4444")
    components.html(f"""<script>
(function(){{
    var p=window.parent; if(!p||!p.document) return;
    var old=p.document.getElementById('occ-ot-ctr'); if(old) old.remove();
    var d=p.document.createElement('div'); d.id='occ-ot-ctr';
    d.style.cssText='position:fixed;top:9px;right:200px;z-index:2147483646;'
        +'font-size:0.68rem;color:rgba(160,160,160,0.7);'
        +'font-family:monospace;pointer-events:none;letter-spacing:0.02em;';
    d.innerHTML='<span style="color:{_wo_clr};font-weight:600;">{_wo_2026:,}</span>'
        +'<span style="opacity:.55;"> / 20K OTs 2026 &nbsp;·&nbsp; {_wo_pct}%</span>';
    p.document.body.appendChild(d);
}})();
</script>""", height=0)

# ── Badge usuario + Salir inyectado en el body del parent (evita clip) ───────
_auth_email_badge = st.session_state.get("_auth_email", "")
components.html(f"""<script>
(function(){{
    var p = window.parent; if(!p||!p.document) return;
    var old = p.document.getElementById('occ-user-badge'); if(old) old.remove();
    var d = p.document.createElement('div'); d.id='occ-user-badge';
    d.style.cssText='position:fixed;top:7px;right:0.9rem;z-index:2147483640;'
        +'display:flex;align-items:center;'
        +'background:rgba(13,20,39,0.90);border:1px solid rgba(255,255,255,0.15);'
        +'border-radius:20px;padding:4px 14px;'
        +'font-family:system-ui,sans-serif;font-size:0.75rem;'
        +'color:rgba(255,255,255,0.85);white-space:nowrap;'
        +'box-shadow:0 2px 8px rgba(0,0,0,0.35);';
    var sp = p.document.createElement('span');
    sp.textContent = '👤 {_auth_email_badge}';
    d.appendChild(sp);
    p.document.body.appendChild(d);
}})();
</script>""", height=0)

# Colores de tema para HTML inline — evita hardcodear colores claros en dark mode
_t = {
    "card":     "#111f38"               if _current_theme == "dark" else "#f8fafc",
    "border":   "#1e3356"               if _current_theme == "dark" else "#e2e8f0",
    "text":     "#e2e8f0"               if _current_theme == "dark" else "#1e293b",
    "muted":    "#94a3b8"               if _current_theme == "dark" else "#64748b",
    # Fondos de tarjetas informativas — paleta Occimiano (#0C2540 / #01798A)
    "info_bg":  "rgba(1,121,138,0.15)"  if _current_theme == "dark" else "rgba(1,121,138,0.08)",
    "warn_bg":  "rgba(1,121,138,0.18)"  if _current_theme == "dark" else "rgba(1,121,138,0.10)",
    "err_bg":   "rgba(239,68,68,0.12)"  if _current_theme == "dark" else "#fef2f2",
    "tbl_head": "#0C2540",                                                              # navy Occimiano
    "tbl_head_text": "#ffffff",                                                          # texto blanco sobre navy
    "tbl_alt":  "rgba(1,121,138,0.10)"  if _current_theme == "dark" else "rgba(1,121,138,0.07)",  # teal suave
    "prog_bg":  "#1e3356"               if _current_theme == "dark" else "#e2e8f0",
    "orange_bg":"rgba(249,115,22,0.12)" if _current_theme == "dark" else "#fff7ed",
}


# ── Helper: st.dataframe con colores de tema ──────────────────────────────────
_st_df = st.dataframe  # alias interno — no tocar

def _show_df(df, **kwargs):
    """
    Wrapper de st.dataframe que aplica estilos oscuros cuando el tema es dark.
    Usa pandas Styler para colorear celdas/encabezados — compatible con column_config.
    """
    if _current_theme == "dark":
        try:
            _s = df.style.set_properties(**{
                "background-color": "#111f38",
                "color": "#e2e8f0",
                "border-color": "#1e3356",
            }).set_table_styles([
                {"selector": "th",
                 "props": [("background-color", "#0C2540"),
                           ("color", "#ffffff"),
                           ("font-weight", "700")]},
                {"selector": "tr:nth-child(even) td",
                 "props": [("background-color", "rgba(1,121,138,0.10)")]},
            ])
            _st_df(_s, **kwargs)
        except Exception:
            # fallback si el Styler falla (p.ej. MultiIndex)
            _st_df(df, **kwargs)
    else:
        _st_df(df, **kwargs)


# ── Helper: gráficos Plotly con tema oscuro/claro ─────────────────────────────
def _apply_plot_theme(fig) -> None:
    """Aplica colores del tema actual a un objeto Figure Plotly (in-place).
    También añade bordes oscuros a barras, pies y sunbursts para consistencia visual."""
    _fc  = _t["text"]
    _gc  = _t["border"]
    _pbg = "rgba(255,255,255,0.04)" if _current_theme == "dark" else "rgba(0,0,0,0)"
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor=_pbg,
        font=dict(color=_fc),
        legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color=_fc)),
    )
    fig.update_xaxes(gridcolor=_gc, zerolinecolor=_gc,
                     tickfont=dict(color=_fc), title_font=dict(color=_fc))
    fig.update_yaxes(gridcolor=_gc, zerolinecolor=_gc,
                     tickfont=dict(color=_fc), title_font=dict(color=_fc))
    # Bordes finos uniformes en barras, tortas y sunbursts
    fig.update_traces(selector={"type": "bar"},
                      marker_line_color="#2d3436", marker_line_width=0.8)
    fig.update_traces(selector={"type": "pie"},
                      marker=dict(line=dict(color="#2d3436", width=1.5)))
    fig.update_traces(selector={"type": "sunburst"},
                      marker=dict(line=dict(color="#2d3436", width=0.8)))


def _plot(fig, **kw) -> None:
    """Aplica colores de tema al fondo/ejes/fuente del chart y lo renderiza."""
    _apply_plot_theme(fig)
    st.plotly_chart(fig, **kw)


def _plot_cached(cache_key: str, sig: str, builder, **kw) -> None:
    """
    Crea, aplica tema y cachea gráficos Plotly en session_state.
    `builder` (callable sin args) solo se invoca en cache miss.
    La clave incluye tema y sig → rerenders son instantáneos si nada cambió.
    """
    _k = f"_fig_{cache_key}_{_current_theme}_{sig}"
    if _k not in st.session_state:
        _fig = builder()
        _apply_plot_theme(_fig)
        st.session_state[_k] = _fig
    st.plotly_chart(st.session_state[_k], **kw)

# ── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    /* Desktop: ocultar header completo (barra Deploy) */
    @media screen and (min-width: 768px) {
        [data-testid="stHeader"] {
            display: none !important;
        }
        [data-testid="stToolbar"] {
            display: none !important;
        }
    }
    /* Mobile: mostrar header con botón ☰ nativo de Streamlit */
    @media screen and (max-width: 767px) {
        [data-testid="stHeader"] {
            display: flex !important;
            background: #0d1427 !important;
            height: 2.8rem !important;
            padding: 0 0.5rem !important;
            align-items: center !important;
            border-bottom: 1px solid rgba(255,255,255,0.12) !important;
        }
        /* Botón ☰ — ícono blanco sobre fondo oscuro */
        [data-testid="stHeader"] button {
            color: #f1f5f9 !important;
            background: rgba(255,255,255,0.08) !important;
            border-radius: 6px !important;
            padding: 6px !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            min-width: 36px !important;
            min-height: 32px !important;
        }
        [data-testid="stHeader"] button svg {
            fill: #f1f5f9 !important;
            color: #f1f5f9 !important;
            width: 20px !important;
            height: 20px !important;
        }
        [data-testid="stHeader"] button svg path {
            fill: #f1f5f9 !important;
            stroke: #f1f5f9 !important;
        }
        /* Ocultar Deploy/Share incluso en móvil */
        [data-testid="stToolbar"],
        [data-testid="stMainMenu"] {
            display: none !important;
        }
    }

    .section-header {
        font-size: 1.1rem; font-weight: 700; color: #1e293b;
        margin: 1rem 0 0.5rem; padding-bottom: 4px;
        border-bottom: 2px solid #e2e8f0;
    }
    [data-testid="stSidebar"],
    [data-testid="stSidebar"] > div,
    [data-testid="stSidebar"] > div > div,
    [data-testid="stSidebar"] section,
    [data-testid="stSidebarContent"] { background: #0d1427; }
    [data-testid="stSidebar"] * { color: #f1f5f9 !important; }
    /* Botones del sidebar — fondo y texto siempre visibles sobre azul oscuro */
    [data-testid="stSidebar"] .stButton > button {
        color: #f1f5f9 !important;
        background: rgba(255,255,255,0.10) !important;
        border: 1px solid rgba(255,255,255,0.22) !important;
        font-weight: 500 !important;
        transition: background 0.18s !important;
    }
    [data-testid="stSidebar"] .stButton > button:hover {
        background: rgba(255,255,255,0.20) !important;
        border-color: rgba(255,255,255,0.38) !important;
    }
    /* ── Sidebar: ancho controlado por _sb_open en session_state ────────────── */
    section[data-testid="stSidebar"] {
        transform: translateX(0) !important;
        visibility: visible !important;
        display: flex !important;
        overflow-x: hidden !important;
        overflow-y: auto !important;
        transition: min-width 0.28s ease, max-width 0.28s ease;
    }
    [data-testid="stSidebarCollapseButton"] { display: none !important; }
    [data-testid="stSidebarContent"] {
        overflow-x: hidden !important;
        overflow-y: auto !important;
        width: 100% !important;
        height: 100% !important;
        max-height: 100vh !important;
        padding-top: 0.5rem !important;
    }
    [data-testid="stSidebarUserContent"] {
        padding-top: 0 !important;
    }

    /* ── Navigation radio buttons styled as menu items ─────────────────────── */
    [data-testid="stSidebar"] [data-testid="stRadio"] > label { display: none; }
    [data-testid="stSidebar"] [data-testid="stRadio"] div[role="radiogroup"] {
        gap: 2px !important;
    }
    [data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"] {
        border-radius: 8px !important;
        transition: background 0.15s, box-shadow 0.15s !important;
        width: 100%;
        white-space: nowrap !important;
        overflow: hidden !important;
        cursor: pointer !important;
    }
    [data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"]:hover {
        background: rgba(59,130,246,0.18) !important;
        box-shadow: inset 3px 0 0 #3b82f6 !important;
    }
    [data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"][aria-checked="true"] {
        background: rgba(59,130,246,0.35) !important;
        box-shadow: inset 3px 0 0 #60a5fa !important;
    }

    /* ── Botones del sidebar: base ──────────────────────────────────────────── */
    [data-testid="stSidebar"] .stButton > button {
        white-space: nowrap !important;
        overflow: hidden !important;
    }

    /* ── Logo: transición de opacidad ─────────────────────────────────────── */
    section[data-testid="stSidebar"] img {
        transition: opacity 0.25s !important;
    }
    /* Multiselect "Select all" → ocultar el inglés y simular "Seleccionar todos" */
    [data-testid="stMultiSelect"] li[role="option"]:first-child > div > span {
        visibility: hidden;
        position: relative;
    }
    [data-testid="stMultiSelect"] li[role="option"]:first-child > div > span::after {
        content: "Seleccionar todos";
        visibility: visible;
        position: absolute;
        left: 0;
    }

    /* ── Multiselect / Selectbox — items del dropdown más nítidos ──────────
       Por defecto Streamlit renderiza las opciones no-hover con un color
       secundario grisáceo que se ve "borroso". Forzamos color pleno y
       antialiasing para mayor legibilidad. */
    [data-testid="stMultiSelect"] li[role="option"],
    [data-testid="stMultiSelect"] li[role="option"] *,
    [data-baseweb="popover"] li[role="option"],
    [data-baseweb="popover"] li[role="option"] *,
    [data-baseweb="menu"] li[role="option"],
    [data-baseweb="menu"] li[role="option"] * {
        color: #1f2937 !important;
        opacity: 1 !important;
        -webkit-font-smoothing: antialiased !important;
        -moz-osx-font-smoothing: grayscale !important;
        text-rendering: optimizeLegibility !important;
    }
    /* Hover un poco más marcado */
    [data-baseweb="popover"] li[role="option"]:hover,
    [data-baseweb="menu"] li[role="option"]:hover {
        background: rgba(1,121,138,0.10) !important;
    }

    /* Ocultar el iframe del toggle JS — solo sirve para ejecutar el script */
    [data-testid="stCustomComponentV1"],
    [data-testid="stCustomComponentV1"] iframe {
        display: none !important;
        height: 0 !important;
        min-height: 0 !important;
        margin: 0 !important;
        padding: 0 !important;
        border: none !important;
    }

    /* ── Títulos de página — más compactos que el default de Streamlit ─────── */
    h1,
    [data-testid="stHeading"],
    [data-testid="stHeadingWithActionElements"] h1 {
        font-size: 1.9rem !important;
        font-weight: 800 !important;
        line-height: 1.2 !important;
    }

    /* ══════════════════════════════════════════════════════
       TABS — Tamaño grande + hover luminoso + activo claro
       ══════════════════════════════════════════════════════ */

    /* Contenedor de la lista de tabs */
    [data-testid="stTabs"] [data-baseweb="tab-list"] {
        gap: 4px !important;
        padding: 0 4px !important;
    }

    /* Botón base — mucho más grande */
    [data-testid="stTabs"] button[data-baseweb="tab"] {
        font-size: 1.10rem !important;
        font-weight: 700 !important;
        padding: 16px 46px !important;
        border-radius: 10px 10px 0 0 !important;
        letter-spacing: 0.03em !important;
        position: relative !important;
        /* Transición suave para TODOS los efectos */
        transition:
            background   0.20s ease,
            box-shadow   0.20s ease,
            color        0.18s ease,
            transform    0.15s ease,
            border-color 0.18s ease !important;
    }

    /* ── HOVER — iluminación al pasar el mouse (sin hacer clic) ── */
    [data-testid="stTabs"] button[data-baseweb="tab"]:not([aria-selected="true"]):hover {
        background: rgba(59, 130, 246, 0.13) !important;
        /* Glow exterior azul + borde inferior iluminado */
        box-shadow:
            0  -3px 18px  4px rgba(59, 130, 246, 0.35),
            0  -2px  0    0   rgba(99, 162, 255, 0.55) inset !important;
        color: #2563eb !important;
        transform: translateY(-3px) !important;
    }

    /* ── ACTIVO — tab seleccionado ── */
    [data-testid="stTabs"] button[data-baseweb="tab"][aria-selected="true"] {
        background: rgba(37, 99, 235, 0.12) !important;
        box-shadow:
            0  -4px 22px  6px rgba(37, 99, 235, 0.45),
            0  -3px  0    0   #3b82f6 inset !important;
        color: #1d4ed8 !important;
        border-bottom: 3px solid #3b82f6 !important;
        transform: translateY(-1px) !important;
    }

    /* Dark mode — ajuste de colores para fondo oscuro */
    .stApp[data-theme="dark"] [data-testid="stTabs"] button[data-baseweb="tab"]:not([aria-selected="true"]):hover,
    [data-theme="dark"] [data-testid="stTabs"] button[data-baseweb="tab"]:not([aria-selected="true"]):hover {
        color: #93c5fd !important;
    }
    .stApp[data-theme="dark"] [data-testid="stTabs"] button[data-baseweb="tab"][aria-selected="true"],
    [data-theme="dark"] [data-testid="stTabs"] button[data-baseweb="tab"][aria-selected="true"] {
        color: #60a5fa !important;
    }
</style>
""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
_NAV_PAGES_BASE = [
    "🥇  Indicadores STO",
    "✅  Cumplimiento SLA",
    "🛠️  Mantenciones Preventivas",
    "⛽  Estaciones de Servicio",
    "⌛  Utilización del Tiempo",
]
# La página Admin solo aparece para usuarios con rol admin
_is_admin   = st.session_state.get("_auth_rol", "usuario") == "admin"
_NAV_PAGES  = _NAV_PAGES_BASE + (["🔐  Administración"] if _is_admin else [])

# ── CSS del sidebar: colapsado / expandido (controlado por _sb_open) ─────────
_sb_open = st.session_state.get("_sb_open", True)
if _sb_open:
    st.markdown("""<style>
section[data-testid="stSidebar"] {
    min-width: 16rem !important; max-width: 16rem !important;
}
[data-testid="stSidebar"] [data-testid="stSidebarContent"],
[data-testid="stSidebar"] [data-testid="stSidebarContent"] > div,
[data-testid="stSidebar"] [data-testid="stSidebarContent"] > div > div {
    padding-left: 0.5rem !important; padding-right: 0.5rem !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"] {
    padding: 10px 14px !important; font-size: 1.1rem !important;
    white-space: normal !important; overflow: visible !important;
    height: auto !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"] > div:first-child {
    display: flex !important; flex-shrink: 0 !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"] > div:last-child {
    max-width: none !important; overflow: visible !important;
    white-space: normal !important; word-break: break-word !important;
}
[data-testid="stSidebar"] .stButton > button {
    max-width: none !important; font-size: 1.0rem !important;
    justify-content: flex-start !important; text-align: left !important;
    padding-left: 14px !important; padding-top: 0.55rem !important;
    padding-bottom: 0.55rem !important;
}
section[data-testid="stSidebar"] img { opacity: 1 !important; }
[data-testid="stSidebar"] small,
[data-testid="stSidebar"] [data-testid="stCaptionContainer"] {
    opacity: 1 !important; font-size: inherit !important; line-height: inherit !important;
}
</style>""", unsafe_allow_html=True)
else:
    st.markdown("""<style>
section[data-testid="stSidebar"] {
    min-width: 6.0rem !important; max-width: 6.0rem !important;
}
[data-testid="stSidebar"] [data-testid="stSidebarContent"],
[data-testid="stSidebar"] [data-testid="stSidebarContent"] > div,
[data-testid="stSidebar"] [data-testid="stSidebarContent"] > div > div {
    padding-left: 0.1rem !important; padding-right: 0.1rem !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"] > div:first-child {
    display: none !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"] {
    padding: 6px 0 !important; font-size: 1.5rem !important;
    justify-content: center !important; line-height: 1.5 !important;
    height: auto !important; overflow: visible !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"] > div:last-child {
    font-size: 1.5rem !important; line-height: 1.5 !important;
    max-width: 1.8em !important; overflow: hidden !important; display: block !important;
    white-space: nowrap !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"] > div:last-child * {
    font-size: 1.5rem !important; line-height: 1.5 !important;
}
[data-testid="stSidebar"] .stButton > button {
    font-size: 1.25rem !important; justify-content: center !important;
    padding: 0.35rem 0 !important; max-width: none !important;
}
section[data-testid="stSidebar"] img { opacity: 0 !important; }
[data-testid="stSidebar"] small,
[data-testid="stSidebar"] [data-testid="stCaptionContainer"] {
    opacity: 0 !important; font-size: 0 !important;
    line-height: 0 !important; pointer-events: none !important;
}
</style>""", unsafe_allow_html=True)

with st.sidebar:
    # Logo Occimiano (solo en modo expandido)
    if _sb_open:
        _sidebar_logo = os.path.join(_APP_DIR, "logo_sidebar.jpg")
        if not os.path.exists(_sidebar_logo):
            _sidebar_logo = os.path.join(_APP_DIR, "logo_dashboard.jpg")
        if os.path.exists(_sidebar_logo):
            st.image(_sidebar_logo, use_container_width=True)
        else:
            st.markdown(
                '<div style="text-align:center;padding:12px 0;">'
                '<span style="font-size:1.6rem;font-weight:900;letter-spacing:2px;color:#fff;">OCCIM</span>'
                '</div>',
                unsafe_allow_html=True,
            )
        st.caption("Indicadores Operacionales")
    st.divider()

    _page = st.radio("Navegación", _NAV_PAGES, label_visibility="collapsed", key="_nav_radio")

# ── Tracking de actividad de sesión (debounce 5 min) ─────────────────────────
_session_id_track = st.session_state.get("_session_id", "")
if _session_id_track:
    import time as _time_mod
    _now_ts = _time_mod.time()
    _last_track = st.session_state.get("_last_activity_track", 0)
    _last_page  = st.session_state.get("_last_activity_page", "")
    _cur_page   = _page if "_page" in dir() else ""
    # Actualizar si cambió de página O si pasaron más de 5 min
    if _cur_page != _last_page or (_now_ts - _last_track) > 300:
        try:
            from supabase_client import update_session_activity
            update_session_activity(_session_id_track, _cur_page)
        except Exception:
            pass
        st.session_state["_last_activity_track"] = _now_ts
        st.session_state["_last_activity_page"]  = _cur_page

with st.sidebar:
    st.divider()
    if st.button("↺  Actualizar datos" if _sb_open else "↺", use_container_width=True):  # noqa
        # load_work_orders se limpia junto con todo lo demás.
        # Es necesario para que Calidad, Precisión y reincidencias muestren
        # los meses más recientes (junio, etc.) — sin esto, el caché de disco
        # puede tener datos viejos aunque haya OTs nuevas en Fracttal.
        # Tarda ~30-45s en recargarse, pero es necesario para datos actualizados.
        for fn in [load_work_orders,                # ← OTs Fracttal (Calidad + Precisión)
                   load_third_parties,
                   load_work_orders_subtasks, load_wo_resources,
                   load_meters, load_meters_reading, load_work_requests,
                   load_listado_eds,
                   load_llamados_fracttal,
                   load_llamados_copec,
                   load_llamados_shell,
                   load_llamados_esmax,
                   load_all_llamados,
                   load_utilizacion_tiempo, list_utilizacion_sheets,
                   load_mtto_realizados_planilla, load_base_tecnicos,
                   # Supabase (solo las que tienen @st.cache_data)
                   load_work_orders_supabase, load_listado_eds_supabase,
                   load_tecnicos_supabase, load_equipos_supabase,
                   load_sla_umbrales_supabase, load_preventivas_supabase]:
            fn.clear()
        # Limpiar TODOS los caches derivados de session_state
        # Incluye df_kpi_raw y df_ot_all_scores — crítico para que junio aparezca:
        # _wo_sig = str(len(raw_wo)) = "20000" siempre, por lo que sin esta limpieza
        # el caché nunca se invalida y no muestra los meses nuevos.
        for _k in ["_base_sig", "_df_llamados", "_df_tp", "_station_name_map",
                   "eds_fracttal_items", "df_llamados_supa", "df_llamados_supa_v3",
                   "_sc_sig_df_llamados_supa", "_sc_sig_df_llamados_supa_v3"]:
            st.session_state.pop(_k, None)
        # Limpiar caches _sc (df_kpi_raw, df_ot_all_scores, df_wo_base, etc.)
        # También eds_enrich y eds_fracttal_items para que Listado de EDS se recalcule
        for _k in list(st.session_state.keys()):
            if str(_k).startswith(("_sla_proc", "_fig_", "_sc_sig_", "df_kpi",
                                   "df_ot_all", "df_ot_bono", "df_wo", "df_reinc",
                                   "eds_enrich", "eds_fracttal")):
                st.session_state.pop(_k, None)
        st.rerun()
    # ── Modo oscuro / claro ───────────────────────────────────────────────────
    _toggle_lbl_full = "☀️  Modo claro" if _current_theme == "dark" else "🌙  Modo oscuro"
    _toggle_lbl_icon = "☀️" if _current_theme == "dark" else "🌙"
    _toggle_lbl = _toggle_lbl_full if _sb_open else _toggle_lbl_icon
    if st.button(_toggle_lbl, use_container_width=True, key="theme_toggle"):
        _new_theme = "light" if _current_theme == "dark" else "dark"
        st.session_state["_theme"] = _new_theme
        import plotly.graph_objects as _pgo
        _old_pfx = f"_{_current_theme}_"
        _new_pfx = f"_{_new_theme}_"
        _figs_to_rename = {
            k: v for k, v in st.session_state.items()
            if k.startswith("_fig_") and isinstance(v, _pgo.Figure)
        }
        _nt = {
            "text":   "#e2e8f0" if _new_theme == "dark" else "#1e293b",
            "border": "#1e3356" if _new_theme == "dark" else "#e2e8f0",
            "pbg":    "rgba(255,255,255,0.04)" if _new_theme == "dark" else "rgba(0,0,0,0)",
        }
        for _fk, _fv in _figs_to_rename.items():
            _fv.update_layout(
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor=_nt["pbg"],
                font=dict(color=_nt["text"]),
                legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color=_nt["text"])),
            )
            _fv.update_xaxes(gridcolor=_nt["border"], zerolinecolor=_nt["border"],
                             tickfont=dict(color=_nt["text"]), title_font=dict(color=_nt["text"]))
            _fv.update_yaxes(gridcolor=_nt["border"], zerolinecolor=_nt["border"],
                             tickfont=dict(color=_nt["text"]), title_font=dict(color=_nt["text"]))
            _new_fk = _fk.replace(_old_pfx, _new_pfx, 1)
            if _new_fk != _fk:
                st.session_state.pop(_fk, None)
                st.session_state[_new_fk] = _fv
        st.rerun()

    # ── Sesión activa: usuario y cierre de sesión ─────────────────────────────
    _auth_email = st.session_state.get("_auth_email", "")
    _auth_user  = _auth_email.split("@")[0] if _auth_email else "usuario"
    if _sb_open:
        st.markdown(
            f'<div style="font-size:0.7rem;color:rgba(255,255,255,0.35);'
            f'text-align:center;padding:2px 0 4px;">👤 {_auth_user}</div>',
            unsafe_allow_html=True,
        )
    if st.button("⎋  Cerrar sesión" if _sb_open else "⎋",
                 use_container_width=True, key="logout_btn"):
        logout()
        st.rerun()
    if _sb_open:
        st.caption(f"Cache: 30 min · disco  |  {datetime.now().strftime('%H:%M:%S')}")

    if _sb_open:
        st.divider()
        # ── Auditor Dash 1.0: última revisión de calidad de datos ────────────────
        _alerta_json = Path(__file__).parent / "alertas_resultado.json"
        try:
            if _alerta_json.exists():
                _ar = json.loads(_alerta_json.read_text(encoding="utf-8"))
                _ar_estado   = _ar.get("estado", "OK")
                _ar_fecha    = (_ar.get("fecha_ejecucion") or "")[:16].replace("T", " ")
                _ar_nc       = _ar.get("total_criticos", 0)
                _ar_na       = _ar.get("total_advertencias", 0)
                if _ar_estado == "CRÍTICO":
                    _ar_ico = "🔴"
                    _ar_txt = f"{_ar_nc} crítica(s)"
                    _ar_col = "#ef4444"
                elif _ar_estado == "ADVERTENCIA":
                    _ar_ico = "🟡"
                    _ar_txt = f"{_ar_na} advertencia(s)"
                    _ar_col = "#f59e0b"
                else:
                    _ar_ico = "✅"
                    _ar_txt = "Datos OK"
                    _ar_col = "#22c55e"
                st.markdown(
                    f'<div style="font-size:0.68rem;color:#64748b;text-align:center;'
                    f'padding:2px 0 0 0;line-height:1.5;">'
                    f'<span style="color:{_ar_col};">{_ar_ico} {_ar_txt}</span>'
                    f'&nbsp;·&nbsp;{_ar_fecha}</div>',
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    '<div style="font-size:0.68rem;color:#475569;text-align:center;'
                    'padding:2px 0 0 0;">⚪ Sin revisión de datos</div>',
                    unsafe_allow_html=True,
                )
        except Exception:
            pass   # nunca romper el sidebar por Auditor Dash 1.0

    _sidebar_load_slot = st.empty()   # placeholder para la barra de carga

# (bloque de toggle móvil eliminado — se usa el botón nativo de Streamlit)

# ── Carga de datos base (barra de progreso en sidebar, no en área principal) ──
# Páginas que necesitan los llamados de Excel (COPEC / ESMAX / Shell).
# En las demás páginas se omite esa carga para no bloquear el arranque.
_PAGES_NEED_LLAMADOS = {_NAV_PAGES_BASE[0], _NAV_PAGES_BASE[1], _NAV_PAGES_BASE[3]}

with _sidebar_load_slot.container():
    _base_prog = st.progress(0, text="📂 Cargando…")
raw_tp = load_third_parties()
with _sidebar_load_slot.container():
    _base_prog = st.progress(50, text="📂 EDS…")

# ── EDS: Supabase en vez de Excel ────────────────────────────────────────────
if _USE_SUPABASE:
    df_eds = load_listado_eds_supabase()
else:
    df_eds = load_listado_eds()

if _page in _PAGES_NEED_LLAMADOS:
    with _sidebar_load_slot.container():
        _base_prog = st.progress(75, text="📂 Llamados…")
    # ── Llamados SLA: Supabase (v_llamados_sla) ──────────────────────────────
    # Bump 'supa_v2' → 'supa_v3_exc' fuerza invalidacion del session_state cache
    # para que todas las sesiones abiertas recarguen y reflejen las excepciones
    # SLA agregadas en runtime (OS-37080 etc.).
    if _USE_SUPABASE:
        _ll_refresh = pd.Timestamp.now().floor("30min").strftime("%Y%m%d%H%M")
        df_llamados = _sc(
            "df_llamados_supa_v3", f"supa_v3_exc_{_ll_refresh}",
            lambda: load_all_llamados_supabase("2026-01-01")
        )
    else:
        df_llamados = load_all_llamados("2026-01-01")
else:
    df_llamados = st.session_state.get("_df_llamados", pd.DataFrame())

_sidebar_load_slot.empty()   # desaparece al terminar

# ── Excepciones SLA globales — aplicadas a df_llamados antes de cualquier cálculo
# OTs justificadas que el Excel marca NO CUMPLE por causas externas a Occimiano.
# Se aplican aquí para que afecten TODA la app (Cumplimiento de SLA + Desempeño Terreno).
if not df_llamados.empty and "cumplimiento" in df_llamados.columns:
    _SLA_EXC_OS  = {"OS-37055", "OS-37448", "OS-37547"}   # folios OS Fracttal
    _SLA_EXC_NUM = {"140785", "143926", "145331"}           # N° llamado ESMAX
    _exc_mask = pd.Series(False, index=df_llamados.index)
    for _col in ["os_fracttal", "OS FRACTTAL"]:
        if _col in df_llamados.columns:
            _exc_mask |= df_llamados[_col].astype(str).str.strip().str.upper().isin(
                {s.upper() for s in _SLA_EXC_OS})
            break
    for _col in ["n_llamado", "LLAMADO", "N° llamado"]:
        if _col in df_llamados.columns:
            _exc_mask |= df_llamados[_col].astype(str).str.strip().isin(_SLA_EXC_NUM)
            break
    if _exc_mask.any():
        df_llamados.loc[_exc_mask, "cumplimiento"] = "CUMPLE"

# ── DataFrames derivados — cacheados en session_state para no recalcular
# en reruns de solo tema (theme toggle). Se invalidan cuando se pulsa "Actualizar".
_base_sig = f"{len(raw_tp)}_{len(df_eds)}_{len(df_llamados)}"
if st.session_state.get("_base_sig") != _base_sig:
    # Datos cambiaron → recalcular
    if not df_llamados.empty and not df_eds.empty:
        df_llamados = resolve_llamados_eds_codes(df_llamados, df_eds)
    df_tp           = build_third_parties_df(raw_tp)
    station_name_map = build_station_name_map(df_tp, df_eds)
    st.session_state["_base_sig"]        = _base_sig
    st.session_state["_df_llamados"]     = df_llamados
    st.session_state["_df_tp"]           = df_tp
    st.session_state["_station_name_map"]= station_name_map
else:
    # Solo rerun de tema → usar caché de session_state
    df_llamados      = st.session_state["_df_llamados"]
    df_tp            = st.session_state["_df_tp"]
    station_name_map = st.session_state["_station_name_map"]

# ── Títulos por página ────────────────────────────────────────────────────────
_PAGE_TITLE = {
    _NAV_PAGES_BASE[0]: "Indicadores STO",
    _NAV_PAGES_BASE[1]: "Cumplimiento SLA",
    _NAV_PAGES_BASE[2]: "Mantenciones Preventivas",
    _NAV_PAGES_BASE[3]: "Estaciones de Servicio",
    _NAV_PAGES_BASE[4]: "Utilización del Tiempo",
    "🔐  Administración":  "Administración",
}
_n_ll = f"{len(df_llamados):,}" if not df_llamados.empty else "–"
_CAPTION = (
    f"Llamados 2026: **{_n_ll}**  |  "
    f"EDS activas: **{int(df_eds['activa'].sum()) if not df_eds.empty else '?'}**"
)


def _hdr(title: str, caption=None) -> None:
    """Renderiza título de página con botón toggle del sidebar al lado derecho."""
    _cap = caption if caption is not None else _CAPTION
    _h_c, _t_c = st.columns([11, 1])
    with _h_c:
        st.title(title)
        if _cap:
            st.caption(_cap)
    with _t_c:
        _open = st.session_state.get("_sb_open", True)
        st.markdown("<div style='padding-top:20px;'>", unsafe_allow_html=True)
        if st.button("◀" if _open else "▶", key="_sb_toggle_main",
                     help="Colapsar/expandir menú lateral"):
            st.session_state["_sb_open"] = not _open
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)


# ── Helper: carga OTs con spinner nativo ────────────────────────────────────
def _load_wo_con_progreso(label: str = "órdenes de trabajo") -> list:
    """
    Carga work orders desde Fracttal.
    - Caché 8 h en disco: si el caché es válido, retorna en < 1 s sin UI.
    - Primera carga o caché expirado: muestra spinner mientras descarga ~30-45 s.
    """
    # Verificar si ya está en caché de disco (el st.cache_data lo gestiona internamente)
    # Usar un st.spinner nativo es más robusto que un thread de progreso animado
    _ph = st.empty()
    with _ph.container():
        with st.spinner(f"⏳ Cargando {label} desde Fracttal… (primera carga puede tardar 1-2 min)"):
            _raw = load_work_orders()
    _ph.empty()
    # Guardar conteo en session_state para el indicador discreto global
    st.session_state["_wo_loaded_count"] = len(_raw)
    st.session_state["_wo_2026_count"] = sum(
        1 for r in _raw
        if str(r.get("creation_date") or "").startswith("2026")
    )
    return _raw

# ══════════════════════════════════════════════════════════════════════════════
# NAVEGACIÓN LATERAL → CONTENIDO CONDICIONAL
# ══════════════════════════════════════════════════════════════════════════════

# ─────────────────────────────────────────────────────────────────────────────
# PÁGINA 1: HISTORIAL FRACTTAL
# ─────────────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
# PÁGINA 2: CUMPLIMIENTO DE SLA 2026
# ─────────────────────────────────────────────────────────────────────────────
if _page == _NAV_PAGES[1]:
    _hdr(_PAGE_TITLE[_NAV_PAGES[1]])
    st.divider()
    if df_llamados.empty:
        st.warning("No se encontraron llamados 2026. Verifica que Google Drive está sincronizado (G:).")
    else:
        # ── Helpers compartidos de fecha ──────────────────────────────────────
        _MESES_ES_CHART = {
            "01":"Enero","02":"Febrero","03":"Marzo","04":"Abril","05":"Mayo","06":"Junio",
            "07":"Julio","08":"Agosto","09":"Septiembre","10":"Octubre","11":"Noviembre","12":"Diciembre",
        }
        def _ym_a_lbl(ym: str) -> str:
            p = str(ym).split("-")
            return f"{_MESES_ES_CHART.get(p[1], p[1])} {p[0][2:]}" if len(p) == 2 else ym

        _TRIMESTRES_DEF = {
            "T1 · Ene–Mar": [1, 2, 3],
            "T2 · Abr–Jun": [4, 5, 6],
            "T3 · Jul–Sep": [7, 8, 9],
            "T4 · Oct–Dic": [10, 11, 12],
        }
        # Inverso de _MESES_ES_CHART: abreviatura → número de mes (para filtrar por trimestre)
        _MES_ABR_NUM_LL = {v: int(k) for k, v in _MESES_ES_CHART.items()}
        _fl_norm = df_llamados["fecha_llamado"]
        if _fl_norm.dt.tz is not None:
            _fl_norm = _fl_norm.dt.tz_convert(None)
        _meses_con_datos = set(_fl_norm.dt.month.dropna().astype(int).unique())
        _trim_opts = ["Todos"] + [
            k for k, v in _TRIMESTRES_DEF.items()
            if any(m in _meses_con_datos for m in v)
        ]
        _mes_raw = _fl_norm.dt.to_period("M").astype(str)
        _meses_disp_lbl = ["Todos"] + [_ym_a_lbl(m) for m in sorted(_mes_raw.dropna().unique(), reverse=True)]
        _lbl_to_period  = {_ym_a_lbl(m): m for m in sorted(_mes_raw.dropna().unique(), reverse=True)}

        _equipo_to_full: dict[str, list[str]] = {
            grp: [TECH_NAME_MAP.get(m, m) for m in info["miembros"]]
            for grp, info in GRUPOS_TERRENO.items()
        }
        _equipo_opts = ["Todos"] + list(GRUPOS_TERRENO.keys())
        prio_colors = {"P1":"#ef4444","P2":"#f97316","P3":"#eab308","P4":"#22c55e"}
        _cu_colors  = {"CUMPLE":"#22c55e","NO CUMPLE":"#ef4444","SIN DATOS":"#94a3b8","NAN":"#94a3b8"}

        # Helper SLA compartido
        def _get_sla_h(cliente: str, prioridad: str, zona_key: str):
            return (SLA_HOURS.get(str(cliente), SLA_DEFAULT)
                    .get(str(prioridad).upper(), {})
                    .get(str(zona_key), None))

        # ── Sub-pestañas ──────────────────────────────────────────────────────
        _tab_cli, _tab_tec, _tab_en_curso, _tab_uptime_sla = st.tabs(
            ["👤  Clientes", "🔧  Servicio Técnico", "⏳  SLA en curso", "⏱️  Uptime General"]
        )

        # ══════════════════════════════════════════════════════════════════════
        # SUB-PESTAÑA: CLIENTES
        # ══════════════════════════════════════════════════════════════════════
        with _tab_cli:
            # ── Clasificación de macrozona (Norte / Santiago / Sur) ────────────
            _NORTE_CIUDADES = {"IQUIQUE","ARICA","ANTOFAGASTA","CALAMA","COPIAPO","COPIAPÓ",
                               "OVALLE","LA SERENA","COQUIMBO","VALLENAR","ILLAPEL","HUASCO",
                               "ALTO HOSPICIO","TOCOPILLA","MEJILLONES"}
            _SUR_CIUDADES   = {"CONCEPCION","CONCEPCIÓN","OSORNO","TEMUCO","VALDIVIA",
                               "PUERTO MONTT","CHILLAN","CHILLÁN","LOS ANGELES","LOS ÁNGELES",
                               "LINARES","TALCA","CURICO","CURICÓ","RANCAGUA","FRUTILLAR",
                               "VICTORIA","RENGO","SAN FERNANDO","PUERTO VARAS","ANCUD",
                               "CASTRO","COIHAIQUE","LA UNION","LA UNIÓN"}
            def _macrozona_ll(eds_nombre, comuna):
                s = (str(eds_nombre or "") + " " + str(comuna or "")).upper()
                if any(c in s for c in _NORTE_CIUDADES): return "Norte"
                if any(c in s for c in _SUR_CIUDADES):   return "Sur"
                return "Centro (Santiago)"

            # ── Filtros ───────────────────────────────────────────────────────
            cf1, cf2, cf3, cf4, cf5, cf6 = st.columns([1.4, 1.2, 1.4, 1.2, 1.4, 1.4])
            with cf1:
                sel_trim_c = st.selectbox("Período", _trim_opts, key="cl_trim")
            with cf2:
                if sel_trim_c != "Todos":
                    _trim_m_c = _TRIMESTRES_DEF[sel_trim_c]
                    _meses_c_disp = [l for l in _meses_disp_lbl[1:]  # skip "Todos"
                                     if _MES_ABR_NUM_LL.get(l.split(" ")[0], 0) in _trim_m_c]
                else:
                    _meses_c_disp = _meses_disp_lbl[1:]
                sel_mes_c = st.multiselect("Mes", _meses_c_disp, key="cl_mes",
                                           placeholder="Todos los meses")
            with cf3:
                _cl_opts = ["Todos"] + sorted(df_llamados["cliente"].dropna().unique().tolist())
                sel_cl_c = st.selectbox("Cliente", _cl_opts, key="cl_cli")
            with cf4:
                _pr_opts_c = ["Todas"] + sorted(df_llamados["prioridad"].dropna().unique().tolist())
                sel_pr_c = st.selectbox("Prioridad", _pr_opts_c, key="cl_pr")
            with cf5:
                sel_cu_c = st.selectbox("Cumplimiento SLA", ["Todos","CUMPLE","NO CUMPLE"], key="cl_cu")
            with cf6:
                # Multiselect (permite combinar Centro + Norte + Sur).
                # Vacío = todas (equivalente a "Todas").
                sel_zona_c = st.multiselect(
                    "Zona", ["Centro (Santiago)", "Norte", "Sur"],
                    key="cl_zona", placeholder="Todas las zonas",
                )

            # ── Aplicar filtros ───────────────────────────────────────────────
            df_ll = df_llamados.copy()
            _fl2 = df_ll["fecha_llamado"]
            if _fl2.dt.tz is not None:
                _fl2 = _fl2.dt.tz_convert(None)
            df_ll["_mes"]   = _fl2.dt.to_period("M").astype(str)
            df_ll["_month"] = _fl2.dt.month.astype("Int64")
            df_ll["_macrozona"] = [
                _macrozona_ll(e, c)
                for e, c in zip(
                    df_ll.get("eds_nombre", pd.Series("", index=df_ll.index)).fillna(""),
                    df_ll.get("comuna", pd.Series("", index=df_ll.index)).fillna(""),
                )
            ]
            if sel_trim_c != "Todos":
                df_ll = df_ll[df_ll["_month"].isin(_TRIMESTRES_DEF[sel_trim_c])]
            if sel_mes_c:
                _periods_c = [_lbl_to_period[l] for l in sel_mes_c if l in _lbl_to_period]
                if _periods_c: df_ll = df_ll[df_ll["_mes"].isin(_periods_c)]
            if sel_cl_c != "Todos":  df_ll = df_ll[df_ll["cliente"] == sel_cl_c]
            if sel_pr_c != "Todas":  df_ll = df_ll[df_ll["prioridad"].str.upper() == sel_pr_c.upper()]
            if sel_cu_c != "Todos":  df_ll = df_ll[df_ll["cumplimiento"] == sel_cu_c]
            if sel_zona_c: df_ll = df_ll[df_ll["_macrozona"].isin(sel_zona_c)]

            # ── KPIs ──────────────────────────────────────────────────────────
            _mes_lbl_tit = f" — {', '.join(sel_mes_c)}" if sel_mes_c else ""
            _cumple_c    = (df_ll["cumplimiento"] == "CUMPLE").sum()
            _nocumple_c  = (df_ll["cumplimiento"] == "NO CUMPLE").sum()
            _pct_c       = round(_cumple_c/(_cumple_c+_nocumple_c)*100,1) if (_cumple_c+_nocumple_c)>0 else 0
            lk1, lk2, lk3, lk4, lk5 = st.columns(5)
            lk1.metric(f"Total llamados{_mes_lbl_tit}", f"{len(df_ll):,}")
            lk2.metric("P1 (máquina detenida)", f"{(df_ll['prioridad'].str.upper()=='P1').sum():,}")
            lk3.metric("Cumple SLA", f"{_cumple_c:,}", delta=f"{_pct_c}%")
            lk4.metric("No cumple SLA", f"{_nocumple_c:,}",
                       delta=f"-{100-_pct_c:.1f}%" if (_cumple_c+_nocumple_c)>0 else None,
                       delta_color="inverse")
            # ── KPI semana actual (Domingo→Sábado, renovación automática) ─────
            _ts_hoy     = pd.Timestamp.now().normalize()
            _dow_s      = (_ts_hoy.dayofweek + 1) % 7          # 0=Dom … 6=Sáb
            _sem_ts     = _ts_hoy - pd.Timedelta(days=_dow_s)  # domingo de esta semana
            _sem_ini    = _sem_ts.date()
            _sem_fin_d  = (_sem_ts + pd.Timedelta(days=6)).date()
            _sem_num    = (_sem_ts + pd.Timedelta(days=1)).isocalendar()[1]
            _meses_abr  = ["Ene","Feb","Mar","Abr","May","Jun",
                           "Jul","Ago","Sep","Oct","Nov","Dic"]
            _sem_lbl    = (f"{_sem_ini.day}–{_sem_fin_d.day} "
                           f"{_meses_abr[_sem_ini.month-1]}")
            _fl_sem     = df_ll["fecha_llamado"]
            if _fl_sem.dt.tz is not None:
                _fl_sem = _fl_sem.dt.tz_convert(None)
            _mask_sem   = (
                (_fl_sem.dt.date >= _sem_ini) &
                (_fl_sem.dt.date <= _sem_fin_d) &
                (df_ll["cumplimiento"].isin(["CUMPLE","NO CUMPLE"]))
            )
            _df_sem     = df_ll[_mask_sem]
            _sem_total  = len(_df_sem)
            _sem_cumple = (_df_sem["cumplimiento"] == "CUMPLE").sum()
            _sem_pct    = round(_sem_cumple / _sem_total * 100, 1) if _sem_total else 0.0
            lk5.metric(
                f"Sem. {_sem_num}  ({_sem_lbl})",
                f"{_sem_pct}%",
                delta=f"{_sem_total} llamados",
                delta_color="off",
            )

            st.divider()
            st.markdown('<div class="section-header">Distribución por prioridad y cliente</div>', unsafe_allow_html=True)
            _ll_sig_c = f"{len(df_ll)}_{sel_trim_c}_{sel_mes_c}_{sel_cl_c}_{sel_pr_c}_{sel_cu_c}_{tuple(sorted(sel_zona_c))}"
            gc1, gc2, gc3 = st.columns([2, 1.1, 2.7])

            with gc1:
                _k = f"_fig_ll_prio_{_current_theme}_{_ll_sig_c}"
                if _k not in st.session_state:
                    _d = df_ll["prioridad"].value_counts().reset_index()
                    _d.columns = ["Prioridad","Llamados"]
                    _f = px.pie(_d, values="Llamados", names="Prioridad", color="Prioridad",
                                color_discrete_map=prio_colors, title="Por prioridad", hole=0.45,
                                category_orders={"Prioridad": ["P1", "P2", "P3", "P4"]})
                    _f.update_traces(
                        marker=dict(line=dict(color="#2d3436", width=1.5)),
                        textfont=dict(size=13),
                    )
                    _f.update_layout(height=400, margin=dict(t=40,b=5,l=5,r=5))
                    _apply_plot_theme(_f); st.session_state[_k] = _f
                st.plotly_chart(st.session_state[_k], width="stretch")

            with gc2:
                _k = f"_fig_ll_cli_{_current_theme}_{_ll_sig_c}"
                if _k not in st.session_state:
                    _d = df_ll["cliente"].value_counts().reset_index()
                    _d.columns = ["Cliente","Llamados"]
                    _f = px.bar(_d, x="Cliente", y="Llamados", color="Cliente",
                                color_discrete_map=CLIENT_COLORS, title="Por cliente", text_auto=True)
                    _f.update_traces(marker_line_color="#2d3436", marker_line_width=0.8,
                                     textfont=dict(size=13))
                    _f.update_layout(showlegend=False, xaxis_title="",
                                     height=400, margin=dict(t=40, b=5))
                    _apply_plot_theme(_f); st.session_state[_k] = _f
                st.plotly_chart(st.session_state[_k], width="stretch")

            with gc3:
                _k = f"_fig_ll_gauge_{_current_theme}_{_ll_sig_c}"
                if _k not in st.session_state:
                    import math as _math

                    _g_clr = (
                        "#22c55e" if _pct_c >= 90 else
                        "#84cc16" if _pct_c >= 75 else
                        "#f97316" if _pct_c >= 60 else "#ef4444"
                    )

                    # ── Gradiente suave rojo→verde (10 zonas sin bordes visibles) ─
                    def _arc(vmin, vmax, r_in, r_out, n=80):
                        a0 = _math.pi*(1 - vmax/100)
                        a1 = _math.pi*(1 - vmin/100)
                        t  = [a0+(a1-a0)*i/(n-1) for i in range(n)]
                        xo = [r_out*_math.cos(a) for a in t]
                        yo = [r_out*_math.sin(a) for a in t]
                        xi = [r_in *_math.cos(a) for a in reversed(t)]
                        yi = [r_in *_math.sin(a) for a in reversed(t)]
                        return xo+xi+[xo[0]], yo+yi+[yo[0]]

                    # Arco exterior gris oscuro (borde)
                    _R_OUT, _R_IN = 1.0, 0.52
                    _zones = [
                        (0,  10, "#c0392b"),
                        (10, 20, "#e74c3c"),
                        (20, 35, "#e67e22"),
                        (35, 50, "#f39c12"),
                        (50, 65, "#f1c40f"),
                        (65, 75, "#d4e157"),
                        (75, 85, "#a5d63a"),
                        (85, 92, "#66bb6a"),
                        (92, 97, "#43a047"),
                        (97,100, "#2e7d32"),
                    ]
                    _fig_gauge = go.Figure()
                    for _vn, _vx, _clr in _zones:
                        _xz, _yz = _arc(_vn, _vx, _R_IN, _R_OUT)
                        _fig_gauge.add_trace(go.Scatter(
                            x=_xz, y=_yz, fill="toself",
                            fillcolor=_clr,
                            line=dict(color=_clr, width=0.5),
                            mode="lines", showlegend=False, hoverinfo="skip"
                        ))

                    # Aro exterior oscuro (delgado)
                    _xout, _yout = _arc(0, 100, _R_OUT, _R_OUT+0.025)
                    _fig_gauge.add_trace(go.Scatter(
                        x=_xout, y=_yout, fill="toself",
                        fillcolor="#2d3436", line=dict(color="#2d3436", width=0),
                        mode="lines", showlegend=False, hoverinfo="skip"
                    ))
                    # Aro interior oscuro (delgado)
                    _xinn, _yinn = _arc(0, 100, _R_IN-0.025, _R_IN)
                    _fig_gauge.add_trace(go.Scatter(
                        x=_xinn, y=_yinn, fill="toself",
                        fillcolor="#2d3436", line=dict(color="#2d3436", width=0),
                        mode="lines", showlegend=False, hoverinfo="skip"
                    ))

                    # ── Aguja delgada y profesional ───────────────────────────
                    _ang = _math.pi*(1 - _pct_c/100)
                    _nl  = 0.78   # largo aguja
                    _bw  = 0.025  # ancho base
                    _al, _ar = _ang+_math.pi/2, _ang-_math.pi/2
                    # Aguja principal
                    _fig_gauge.add_trace(go.Scatter(
                        x=[_bw*_math.cos(_al), _nl*_math.cos(_ang),
                           _bw*_math.cos(_ar), _bw*_math.cos(_al)],
                        y=[_bw*_math.sin(_al), _nl*_math.sin(_ang),
                           _bw*_math.sin(_ar), _bw*_math.sin(_al)],
                        fill="toself", fillcolor="#2d3436",
                        line=dict(color="#2d3436", width=0.5),
                        mode="lines", showlegend=False, hoverinfo="skip"
                    ))
                    # Contrapeso (cola corta)
                    _tail = 0.18
                    _fig_gauge.add_trace(go.Scatter(
                        x=[_bw*_math.cos(_al),
                           -_tail*_math.cos(_ang),
                           _bw*_math.cos(_ar),
                           _bw*_math.cos(_al)],
                        y=[_bw*_math.sin(_al),
                           -_tail*_math.sin(_ang),
                           _bw*_math.sin(_ar),
                           _bw*_math.sin(_al)],
                        fill="toself", fillcolor="#555",
                        line=dict(color="#555", width=0.5),
                        mode="lines", showlegend=False, hoverinfo="skip"
                    ))
                    # Hub: círculo blanco con borde oscuro
                    _hn, _hr_o, _hr_i = 40, 0.10, 0.06
                    for _hr, _fc in [(_hr_o,"#2d3436"),(_hr_i,"#ffffff")]:
                        _fig_gauge.add_trace(go.Scatter(
                            x=[_hr*_math.cos(2*_math.pi*i/_hn) for i in range(_hn+1)],
                            y=[_hr*_math.sin(2*_math.pi*i/_hn) for i in range(_hn+1)],
                            fill="toself", fillcolor=_fc,
                            line=dict(color=_fc), mode="lines",
                            showlegend=False, hoverinfo="skip"
                        ))

                    # ── Ticks ─────────────────────────────────────────────────
                    for _tv in [0, 25, 50, 75, 90, 100]:
                        _ta    = _math.pi*(1 - _tv/100)
                        _is_90 = _tv == 90
                        _tc    = "#e74c3c" if _is_90 else "#ecf0f1"
                        _tw    = 3 if _is_90 else 2
                        _tl    = 1.18 if _is_90 else 1.13
                        _fig_gauge.add_trace(go.Scatter(
                            x=[(_R_OUT+0.07)*_math.cos(_ta), _tl*_math.cos(_ta)],
                            y=[(_R_OUT+0.07)*_math.sin(_ta), _tl*_math.sin(_ta)],
                            line=dict(color=_tc, width=_tw),
                            mode="lines", showlegend=False, hoverinfo="skip"
                        ))
                        _lbl = f"<b>{_tv}%✓</b>" if _is_90 else f"{_tv}%"
                        _fig_gauge.add_annotation(
                            x=1.30*_math.cos(_ta), y=1.30*_math.sin(_ta),
                            text=_lbl, showarrow=False,
                            font=dict(size=10 if _is_90 else 9,
                                      color=_tc)
                        )

                    # ── Textos ────────────────────────────────────────────────
                    _fig_gauge.add_annotation(
                        x=0, y=-0.15,
                        text=f"<b>{_pct_c:.1f}%</b>",
                        showarrow=False, font=dict(size=36, color=_g_clr)
                    )
                    _fig_gauge.add_annotation(
                        x=0, y=-0.38,
                        text=f"{int(_cumple_c)} cumple · {int(_nocumple_c)} no cumple",
                        showarrow=False, font=dict(size=11, color=_t["muted"])
                    )
                    _fig_gauge.add_annotation(
                        x=0, y=1.30,
                        text="<b>Cumplimiento SLA</b>",
                        showarrow=False, font=dict(size=14, color=_t["text"])
                    )

                    _fig_gauge.update_layout(
                        xaxis=dict(range=[-1.2, 1.2], visible=False, scaleanchor="y"),
                        yaxis=dict(range=[-0.50, 1.42], visible=False),
                        height=400,
                        margin=dict(t=5, b=5, l=5, r=5),
                        paper_bgcolor="rgba(0,0,0,0)",
                        plot_bgcolor="rgba(0,0,0,0)",
                    )
                    st.session_state[_k] = _fig_gauge
                st.plotly_chart(st.session_state[_k], width="stretch")

            # ── Evolución SLA por mes ─────────────────────────────────────────
            st.divider()
            _cli_lbl_evol = sel_cl_c if sel_cl_c != "Todos" else "todos los clientes"
            st.markdown(
                f'<div class="section-header">📈 Evolución SLA Occimiano por mes — {_cli_lbl_evol}</div>',
                unsafe_allow_html=True,
            )

            _sla_evol_sig = f"{len(df_llamados)}_{sel_trim_c}_{sel_cl_c}_{sel_pr_c}"
            _sla_evol_k   = f"_fig_ll_sla_evol_{_current_theme}_{_sla_evol_sig}"
            if _sla_evol_k not in st.session_state:
                # Base: filtrar por período y cliente pero NO por cumplimiento
                # para medir la tasa SLA real mes a mes
                _ev_base = df_llamados.copy()
                _fl_ev = _ev_base["fecha_llamado"]
                if _fl_ev.dt.tz is not None:
                    _fl_ev = _fl_ev.dt.tz_convert(None)
                _ev_base["_mes"]   = _fl_ev.dt.to_period("M").astype(str)
                _ev_base["_month"] = _fl_ev.dt.month.astype("Int64")
                if sel_trim_c != "Todos":
                    _ev_base = _ev_base[_ev_base["_month"].isin(_TRIMESTRES_DEF[sel_trim_c])]
                if sel_cl_c  != "Todos":
                    _ev_base = _ev_base[_ev_base["cliente"] == sel_cl_c]
                if sel_pr_c  != "Todas":
                    _ev_base = _ev_base[_ev_base["prioridad"].str.upper() == sel_pr_c.upper()]

                _ev_grp = (
                    _ev_base.groupby("_mes").agg(
                        llamados =("_mes", "count"),
                        cumple   =("cumplimiento", lambda x: (x == "CUMPLE").sum()),
                        no_cumple=("cumplimiento", lambda x: (x == "NO CUMPLE").sum()),
                    ).reset_index().sort_values("_mes")
                )
                _ev_grp["pct_sla"] = (
                    (_ev_grp["cumple"] / (_ev_grp["cumple"] + _ev_grp["no_cumple"]) * 100)
                    .where((_ev_grp["cumple"] + _ev_grp["no_cumple"]) > 0, 0)
                    .astype(float).round(1)
                )
                _ev_grp["mes_lbl"] = _ev_grp["_mes"].apply(_ym_a_lbl)

                if not _ev_grp.empty:
                    _fig_sla_evol = make_subplots(specs=[[{"secondary_y": True}]])

                    # Color de barra según cliente seleccionado
                    _bar_color_evol = CLIENT_COLORS.get(sel_cl_c, "#94a3b8")
                    # Texto oscuro sobre amarillo (Shell), blanco sobre el resto
                    _bar_txt_evol = "#1e293b" if sel_cl_c == "SHELL (Enex)" else "#ffffff"

                    # Barras: llamados por mes
                    _fig_sla_evol.add_trace(
                        go.Bar(
                            x=_ev_grp["mes_lbl"], y=_ev_grp["llamados"],
                            name="Llamados correctivos",
                            marker_color=_bar_color_evol, opacity=0.9,
                            text=_ev_grp["llamados"],
                            textposition="inside",
                            insidetextanchor="start",
                            textfont=dict(size=13, color=_bar_txt_evol, family="Arial"),
                        ),
                        secondary_y=False,
                    )

                    # Línea: tendencia de llamados — naranja
                    _fig_sla_evol.add_trace(
                        go.Scatter(
                            x=_ev_grp["mes_lbl"], y=_ev_grp["llamados"],
                            name="Tendencia llamados",
                            mode="lines+markers",
                            line=dict(color="#f97316", width=2.5, dash="dot"),
                            marker=dict(size=7, color="#f97316",
                                        line=dict(color="#ffffff", width=1)),
                        ),
                        secondary_y=False,
                    )

                    # Línea: % SLA cumplimiento — verde (eje derecho)
                    _fig_sla_evol.add_trace(
                        go.Scatter(
                            x=_ev_grp["mes_lbl"], y=_ev_grp["pct_sla"],
                            name="% Cumplimiento SLA",
                            mode="lines+markers",
                            line=dict(color="#22c55e", width=3),
                            marker=dict(size=11, color="#22c55e",
                                        line=dict(color="#ffffff", width=2)),
                            customdata=list(zip(
                                _ev_grp["cumple"], _ev_grp["no_cumple"], _ev_grp["llamados"]
                            )),
                            hovertemplate=(
                                "<b>%{x}</b><br>"
                                "% SLA: %{y:.1f}%<br>"
                                "Cumple: %{customdata[0]}<br>"
                                "No cumple: %{customdata[1]}<br>"
                                "Total llamados: %{customdata[2]}<br>"
                                "<extra></extra>"
                            ),
                        ),
                        secondary_y=True,
                    )
                    # Anotaciones con fondo blanco bordeado en verde — debajo del punto
                    for _, _ann_row in _ev_grp.iterrows():
                        _fig_sla_evol.add_annotation(
                            x=_ann_row["mes_lbl"],
                            y=_ann_row["pct_sla"],
                            yref="y2",
                            text=f"<b>{_ann_row['pct_sla']:.1f}%</b>  {int(_ann_row['cumple'])} cumple",
                            showarrow=False,
                            yanchor="top",
                            yshift=-8,
                            font=dict(size=11, color="#16a34a", family="Arial"),
                            bgcolor="rgba(255,255,255,0.88)",
                            bordercolor="#22c55e",
                            borderwidth=1.5,
                            borderpad=4,
                        )

                    _fig_sla_evol.update_layout(
                        title=f"Evolución mensual — Llamados correctivos vs % SLA",
                        height=430,
                        legend=dict(orientation="h", y=1.08, x=0),
                        margin=dict(t=60, b=20),
                        bargap=0.3,
                    )
                    _fig_sla_evol.update_yaxes(title_text="Llamados correctivos", secondary_y=False)
                    _fig_sla_evol.update_yaxes(
                        title_text="% Cumplimiento SLA",
                        secondary_y=True,
                        tickformat=".1f", ticksuffix="%",
                        range=[0, 110],
                    )
                    _apply_plot_theme(_fig_sla_evol)
                    st.session_state[_sla_evol_k] = _fig_sla_evol
                else:
                    st.session_state[_sla_evol_k] = None

            _fig_sla_show = st.session_state.get(_sla_evol_k)
            if _fig_sla_show is not None:
                st.plotly_chart(_fig_sla_show, width="stretch")
            else:
                st.info("Sin datos para el período seleccionado.")

            st.divider()
            st.markdown('<div class="section-header">Evolución mensual de llamados</div>', unsafe_allow_html=True)
            _df_llm = df_ll.dropna(subset=["fecha_llamado"]).copy()
            _df_llm["mes_lbl"] = _df_llm["_mes"].apply(_ym_a_lbl)
            _monthly_c = _df_llm.groupby(["mes_lbl","prioridad"]).size().reset_index(name="llamados")
            _mes_ord_c = [_ym_a_lbl(m) for m in sorted(_df_llm["_mes"].unique())]
            if not _monthly_c.empty:
                _k = f"_fig_ll_monthly_{_current_theme}_{_ll_sig_c}"
                if _k not in st.session_state:
                    _f = px.bar(_monthly_c, x="mes_lbl", y="llamados", color="prioridad",
                                color_discrete_map=prio_colors, title="Llamados por mes y prioridad",
                                barmode="stack", category_orders={"mes_lbl": _mes_ord_c})
                    _f.update_layout(xaxis_title="", yaxis_title="Llamados",
                                     legend_title="Prioridad", xaxis_type="category")
                    _apply_plot_theme(_f); st.session_state[_k] = _f
                st.plotly_chart(st.session_state[_k], width="stretch")

            st.markdown('<div class="section-header">Ranking de EDS por llamados (2026)</div>', unsafe_allow_html=True)

            # Pre-calcular % uso SLA por OT (tiempo real / umbral × 100)
            # para tener un indicador real incluso cuando cumplimiento=0%
            def _td_h_c(v):
                try: return pd.to_timedelta(v).total_seconds()/3600
                except: return float("nan")
            def _zona_c(z):
                z=str(z).upper().strip()
                if any(k in z for k in ["SANTIAGO","METRO","RM","R.M."]): return "Santiago"
                return "Regiones"
            _df_sla_uso = df_ll.copy()
            _df_sla_uso["_h"] = _df_sla_uso.get(
                "tiempo_resp_real", pd.Series(dtype=object, index=_df_sla_uso.index)
            ).apply(_td_h_c)
            _df_sla_uso["_zona"] = _df_sla_uso.get(
                "zona", pd.Series("", index=_df_sla_uso.index)
            ).apply(_zona_c)
            _df_sla_uso["_sla"] = [
                _get_sla_h(c, p, z)
                for c, p, z in zip(
                    _df_sla_uso.get("cliente",  pd.Series("", index=_df_sla_uso.index)),
                    _df_sla_uso.get("prioridad", pd.Series("", index=_df_sla_uso.index)),
                    _df_sla_uso["_zona"],
                )
            ]
            _df_sla_uso["_pct"] = (
                (_df_sla_uso["_h"] / _df_sla_uso["_sla"] * 100)
                .where(
                    _df_sla_uso["_h"].notna() &
                    pd.Series([pd.notna(x) for x in _df_sla_uso["_sla"]], index=_df_sla_uso.index)
                )
            )
            _eds_avg_pct = (
                _df_sla_uso.groupby("eds_occim")["_pct"].mean().round(1)
                .reset_index().rename(columns={"_pct": "pct_sla_uso"})
            )

            _kpis_raw_c = kpis_por_eds(df_ll)
            if not df_eds.empty:
                _eds_act = df_eds[df_eds["activa"]].copy()
                if sel_cl_c != "Todos":
                    _eds_act = _eds_act[_eds_act["cliente"] == sel_cl_c]
                _bc = [c for c in ["eds_occim","eds_cliente","cliente","direccion","comuna","zona_occim","region"]
                       if c in _eds_act.columns]
                _eds_base = _eds_act[_bc].drop_duplicates("eds_occim")
                if not _kpis_raw_c.empty:
                    # Partir desde los llamados reales (no desde el listado completo de EDS)
                    # → garantiza que aparecen todos los llamados aunque su código no esté en df_eds
                    kpis_ll = _kpis_raw_c.merge(
                        _eds_base.drop(columns=["cliente"], errors="ignore"),
                        on="eds_occim", how="left")
                else:
                    kpis_ll = pd.DataFrame()
                for _col in ["total_llamados","p1","p2","p3","p4","cumple","no_cumple"]:
                    if _col in kpis_ll.columns:
                        kpis_ll[_col] = kpis_ll[_col].fillna(0).astype(int)
                if "pct_cumplimiento" in kpis_ll.columns:
                    kpis_ll["pct_cumplimiento"] = kpis_ll["pct_cumplimiento"].fillna(0.0)
                # Solo EDS que han tenido al menos 1 llamado
                if not kpis_ll.empty and "total_llamados" in kpis_ll.columns:
                    kpis_ll = kpis_ll[kpis_ll["total_llamados"] > 0]
                # Unir % promedio de uso SLA (tiempo real / umbral × 100)
                if not kpis_ll.empty:
                    kpis_ll = kpis_ll.merge(_eds_avg_pct, on="eds_occim", how="left")
            else:
                kpis_ll = _kpis_raw_c
                kpis_ll = kpis_ll.merge(_eds_avg_pct, on="eds_occim", how="left")
            if not kpis_ll.empty:
                # Nombre para mostrar: nombre comercial → dirección → código
                kpis_ll["nombre"] = kpis_ll["eds_occim"].astype(str).map(station_name_map)
                kpis_ll["nombre"] = kpis_ll.apply(
                    lambda r: r["nombre"] if pd.notna(r.get("nombre")) and str(r.get("nombre","")).strip()
                    else r.get("direccion",""), axis=1)
                if "ultimo_llamado" in kpis_ll.columns:
                    kpis_ll["ultimo_llamado"] = (pd.to_datetime(kpis_ll["ultimo_llamado"],errors="coerce")
                                                 .dt.strftime("%d/%m/%Y"))
                # Filtro definitivo: solo EDS con al menos 1 llamado real
                # (doble garantía: el filtro previo pudo no alcanzar si total_llamados llegó como NaN)
                if "total_llamados" in kpis_ll.columns:
                    kpis_ll = kpis_ll[kpis_ll["total_llamados"].fillna(0).astype(int) > 0]

                _ll_buscar = st.text_input("🔍 Buscar estación (código, dirección, comuna)",
                                           key="ll_buscar", placeholder="ej: Talagante, 60783, SH_647")
                if _ll_buscar:
                    _q = _ll_buscar.lower()
                    _mask = (
                        kpis_ll["eds_occim"].astype(str).str.lower().str.contains(_q, na=False)
                        | kpis_ll.get("eds_cliente",pd.Series(dtype=str)).fillna("").str.lower().str.contains(_q,na=False)
                        | kpis_ll.get("nombre",pd.Series(dtype=str)).fillna("").str.lower().str.contains(_q,na=False)
                        | kpis_ll.get("direccion",pd.Series(dtype=str)).fillna("").str.lower().str.contains(_q,na=False)
                        | kpis_ll.get("comuna",pd.Series(dtype=str)).fillna("").str.lower().str.contains(_q,na=False)
                    )
                    kpis_ll = kpis_ll[_mask]

                # Columnas a mostrar: solo las relevantes (sin desglose de prioridad)
                _dc = [c for c in ["eds_occim", "nombre", "cliente", "comuna",
                                   "total_llamados", "pct_cumplimiento",
                                   "ultimo_llamado", "ultimo_tecnico"]
                       if c in kpis_ll.columns]
                _kpis_disp = (
                    kpis_ll[_dc]
                    .sort_values("total_llamados", ascending=False)
                    .rename(columns={
                        "eds_occim":        "Cód. Occim",
                        "nombre":           "Nombre / Dirección",
                        "cliente":          "Cliente",
                        "comuna":           "Comuna",
                        "total_llamados":   "Llamados",
                        "pct_cumplimiento": "% Cumpl. SLA",
                        "ultimo_llamado":   "Último Llamado",
                        "ultimo_tecnico":   "Último Técnico",
                    })
                )
                if not _kpis_disp.empty:
                    _show_df(_kpis_disp, width="stretch", hide_index=True,
                        column_config={
                            "Cód. Occim":         st.column_config.TextColumn(width=100),
                            "Nombre / Dirección": st.column_config.TextColumn(width=280),
                            "Cliente":            st.column_config.TextColumn(width=110),
                            "Comuna":             st.column_config.TextColumn(width=120),
                            "Llamados":           st.column_config.NumberColumn(
                                                    format="%d", width=90),
                            "% Cumpl. SLA":       st.column_config.ProgressColumn(
                                                    label="% Cumpl. SLA",
                                                    min_value=0, max_value=100,
                                                    format="%.1f%%"),
                            "Último Llamado":     st.column_config.TextColumn(width=110),
                            "Último Técnico":     st.column_config.TextColumn(width=140),
                        })
                else:
                    st.info("Sin llamados en el período seleccionado.")

            st.divider()
            st.markdown('<div class="section-header">⏱ Cumplimiento de SLA por OT</div>', unsafe_allow_html=True)
            st.caption("Tiempo de resolución real por cada llamado cerrado, comparado con el umbral de SLA según prioridad y zona.")

            def _fmt_horas_ot(h: float) -> str:
                if pd.isna(h) or h < 0: return "—"
                hh, mm = int(h), int((h % 1) * 60)
                return f"{hh}h {mm:02d}m"

            _df_sla_ot = df_ll[df_ll["fecha_llamado"].notna() & df_ll["fecha_atencion"].notna()].copy()
            if not _df_sla_ot.empty:
                def _merge_dt(fecha_ser: pd.Series, hora_col: str) -> pd.Series:
                    fechas = pd.to_datetime(fecha_ser, errors="coerce")
                    if hora_col not in _df_sla_ot.columns: return fechas
                    def _to_hms(h):
                        if pd.isna(h): return ""
                        if hasattr(h, "strftime"): return h.strftime("%H:%M:%S")
                        s = str(h).strip()
                        return "" if s in ("nan","NaT","None","") else s
                    hora_str = _df_sla_ot[hora_col].apply(_to_hms)
                    date_str = fechas.dt.strftime("%Y-%m-%d").fillna("")
                    combined = pd.to_datetime((date_str + " " + hora_str).str.strip(), errors="coerce")
                    return combined.fillna(fechas)

                _dt_ll  = _merge_dt(_df_sla_ot["fecha_llamado"], "hora_llamado")
                _dt_at  = _merge_dt(_df_sla_ot["fecha_atencion"], "hora_fin")
                _df_sla_ot["horas_res"] = ((_dt_at - _dt_ll).dt.total_seconds() / 3600).clip(lower=0)

                # ── Zona derivada de la comuna (autoritativa) ───────────────
                # Comunas de la Región Metropolitana de Santiago.
                # Si la comuna está en esta lista → Santiago, si no → Regiones.
                # Fallback: usar el campo "zona" del llamado solo si la comuna no se conoce.
                import unicodedata as _ud
                _RM_COMUNAS = {
                    "SANTIAGO","LAS CONDES","VITACURA","PROVIDENCIA","NUNOA","LA REINA","MACUL",
                    "PENALOLEN","LA FLORIDA","PUENTE ALTO","MAIPU","ESTACION CENTRAL","CERRILLOS",
                    "PUDAHUEL","QUILICURA","RENCA","CONCHALI","INDEPENDENCIA","RECOLETA","HUECHURABA",
                    "LO BARNECHEA","SAN MIGUEL","SAN JOAQUIN","LA GRANJA","LA CISTERNA","EL BOSQUE",
                    "SAN BERNARDO","LO ESPEJO","PEDRO AGUIRRE CERDA","P.A. CERDA","SAN RAMON",
                    "LA PINTANA","CERRO NAVIA","LO PRADO","QUINTA NORMAL","BUIN","CALERA DE TANGO",
                    "COLINA","LAMPA","TALAGANTE","PENAFLOR","EL MONTE","PADRE HURTADO","MELIPILLA",
                    "CURACAVI","MARIA PINTO","ISLA DE MAIPO","SAN PEDRO","ALHUE","PIRQUE","TILTIL",
                    "BATUCO","PAINE",
                }
                def _norm_com(s):
                    s = _ud.normalize("NFD", str(s or "").upper().strip()).encode("ascii","ignore").decode()
                    return s
                def _zona_por_comuna(com, zona_orig):
                    cn = _norm_com(com)
                    if cn and cn in _RM_COMUNAS:
                        return "Santiago"
                    if cn:
                        return "Regiones"
                    # Comuna desconocida → fallback al campo zona del llamado
                    z = str(zona_orig or "").upper().strip()
                    if z in ("RM","R.M.") or any(k in z for k in ["SANTIAGO","METRO"]): return "Santiago"
                    return "Regiones"

                # Inicializamos zona_ot temporal con el campo zona original; lo
                # recalcularemos abajo después de resolver la ciudad final.
                _df_sla_ot["zona_ot"] = (
                    _df_sla_ot["zona"].fillna("").astype(str).str.strip()
                    if "zona" in _df_sla_ot.columns
                    else pd.Series([""] * len(_df_sla_ot), index=_df_sla_ot.index)
                )
                _df_sla_ot["umbral_h"] = [
                    _get_sla_h(c, p, z)
                    for c, p, z in zip(
                        _df_sla_ot.get("cliente", pd.Series([""] * len(_df_sla_ot), index=_df_sla_ot.index)),
                        _df_sla_ot.get("prioridad", pd.Series([""] * len(_df_sla_ot), index=_df_sla_ot.index)),
                        _df_sla_ot["zona_ot"],
                    )
                ]
                _df_sla_ot["umbral_lbl"] = [
                    f"{int(u)}h" if (u is not None and pd.notna(u)) else "—"
                    for u in _df_sla_ot["umbral_h"]
                ]
                _df_sla_ot["tiempo_res"] = [_fmt_horas_ot(h) for h in _df_sla_ot["horas_res"]]
                # % de uso SLA real (tiempo / umbral × 100)
                _df_sla_ot["pct_sla_ot"] = [
                    round(h / u * 100, 1) if (u is not None and pd.notna(u) and u > 0 and pd.notna(h)) else None
                    for h, u in zip(_df_sla_ot["horas_res"], _df_sla_ot["umbral_h"])
                ]
                # estado_sla: usar cumplimiento del view como fuente de verdad.
                # Si la OT tiene excepcion_motivo (registrada en sla_excepciones),
                # se muestra 'ℹ️ Excepción' en lugar de '✅ Cumple' — pero cuenta
                # como cumple igual en los %.
                def _estado_sla_row(row):
                    cumpl = str(row.get("cumplimiento") or "").upper()
                    exc   = row.get("excepcion_motivo")
                    tiene_exc = pd.notna(exc) and str(exc).strip() != ""
                    if cumpl == "CUMPLE":
                        return "ℹ️ Excepción" if tiene_exc else "✅ Cumple"
                    if cumpl == "NO CUMPLE":
                        return "❌ No cumple"
                    return "⚪ Sin datos"
                if "cumplimiento" in _df_sla_ot.columns:
                    _df_sla_ot["estado_sla"] = _df_sla_ot.apply(_estado_sla_row, axis=1)
                else:
                    _df_sla_ot["estado_sla"] = [
                        ("✅ Cumple" if (u is not None and pd.notna(u)) and h <= u else
                         "❌ No cumple" if (u is not None and pd.notna(u)) else "⚪ Sin prioridad")
                        for h, u in zip(_df_sla_ot["horas_res"], _df_sla_ot["umbral_h"])
                    ]
                # Agregar ciudad: usar comuna del propio llamado, completar vacíos
                # con el catálogo de EDS. Para Aramco el eds_occim viene como
                # "EE_S016" (formato Fracttal) → matchear con eds_occim_raw del catálogo.
                _df_sla_ot["ciudad"] = _df_sla_ot["comuna"].fillna("") if "comuna" in _df_sla_ot.columns else ""
                if not df_eds.empty and "comuna" in df_eds.columns and "eds_occim" in _df_sla_ot.columns:
                    _eds_clean = df_eds.dropna(subset=["comuna"])
                    # 1) Match por eds_occim directo (PBR-XX, etc.)
                    _eds_comuna_map = _eds_clean.drop_duplicates("eds_occim").set_index("eds_occim")["comuna"]
                    _fill1 = _df_sla_ot["eds_occim"].map(_eds_comuna_map).fillna("")
                    # 2) Match por eds_occim_raw (EE_S### que usa Aramco)
                    if "eds_occim_raw" in _eds_clean.columns:
                        _eds_raw_map = (_eds_clean.dropna(subset=["eds_occim_raw"])
                                                  .drop_duplicates("eds_occim_raw")
                                                  .assign(_k=lambda d: d["eds_occim_raw"].str.upper().str.strip())
                                                  .set_index("_k")["comuna"])
                        _fill2 = _df_sla_ot["eds_occim"].astype(str).str.upper().str.strip().map(_eds_raw_map).fillna("")
                    else:
                        _fill2 = pd.Series([""] * len(_df_sla_ot), index=_df_sla_ot.index)
                    # Aplicar fallback en cascada
                    _df_sla_ot["ciudad"] = _df_sla_ot["ciudad"].where(_df_sla_ot["ciudad"].str.strip() != "", _fill1)
                    _df_sla_ot["ciudad"] = _df_sla_ot["ciudad"].where(_df_sla_ot["ciudad"].str.strip() != "", _fill2)

                # Recalcular zona_ot ahora que tenemos la comuna real
                _df_sla_ot["zona_ot"] = [
                    _zona_por_comuna(c, z) for c, z in zip(_df_sla_ot["ciudad"], _df_sla_ot["zona_ot"])
                ]
                _df_sla_ot["ciudad"] = _df_sla_ot["ciudad"].replace("", "—")

                # Recalcular umbral SLA usando la zona corregida
                _df_sla_ot["umbral_h"] = [
                    _get_sla_h(c, p, z)
                    for c, p, z in zip(
                        _df_sla_ot.get("cliente", pd.Series([""] * len(_df_sla_ot), index=_df_sla_ot.index)),
                        _df_sla_ot.get("prioridad", pd.Series([""] * len(_df_sla_ot), index=_df_sla_ot.index)),
                        _df_sla_ot["zona_ot"],
                    )
                ]
                _df_sla_ot["umbral_lbl"] = [
                    f"{int(u)}h" if (u is not None and pd.notna(u)) else "—"
                    for u in _df_sla_ot["umbral_h"]
                ]
                _df_sla_ot["pct_sla_ot"] = [
                    round(h / u * 100, 1) if (u is not None and pd.notna(u) and u > 0 and pd.notna(h)) else None
                    for h, u in zip(_df_sla_ot["horas_res"], _df_sla_ot["umbral_h"])
                ]

                # N° Aviso cliente — Aramco: N° Cotalker / COPEC: N° Aviso del email
                _cotalker_idx = load_cotalker_index_supabase()
                if _cotalker_idx and "os_fracttal" in _df_sla_ot.columns:
                    _df_sla_ot["n_cotalker"] = (
                        _df_sla_ot["os_fracttal"]
                        .map(_cotalker_idx)
                        .apply(lambda v: str(v) if pd.notna(v) and str(v) not in ("", "nan") else "")
                    )
                else:
                    _df_sla_ot["n_cotalker"] = ""

                # Reporte de falla — extraído de nota_tarea:
                #   Aramco: texto después de "Detalles del incidente:"  (Cotalker)
                #   COPEC:  texto después de "Falla reportada:"        (sistema COPEC)
                #   SHELL:  texto entre comillas tras                  (correo SHELL)
                #           "Descripción del Requerimiento:"
                _df_sla_ot["reporte"] = ""
                if "os_fracttal" in _df_sla_ot.columns:
                    try:
                        _notas_idx = load_notas_tarea_index(tuple(sorted(_df_sla_ot["os_fracttal"].dropna().unique().tolist())))
                        import re as _re_rep
                        _pat_aramco = _re_rep.compile(r"Detalles\s*del\s*incidente\s*:\s*(.+)", _re_rep.IGNORECASE | _re_rep.DOTALL)
                        _pat_copec  = _re_rep.compile(r"Falla\s*reportada\s*:\s*(.+?)(?:\n|Tiempo\s*de\s*respuesta|Contacto|Direcc|$)", _re_rep.IGNORECASE | _re_rep.DOTALL)
                        _pat_shell  = _re_rep.compile(r'Descripci[óo]n\s+del\s+Requerimiento\s*:\s*["“](.+?)["”]', _re_rep.IGNORECASE | _re_rep.DOTALL)

                        def _extraer_reporte(ot):
                            nota = _notas_idx.get(ot, "")
                            if not nota:
                                return ""
                            nota_str = str(nota).replace("\xa0", " ")
                            # Aramco
                            m = _pat_aramco.search(nota_str)
                            if m:
                                return m.group(1).strip()
                            # COPEC
                            m = _pat_copec.search(nota_str)
                            if m:
                                return m.group(1).strip().rstrip(":").strip()
                            # SHELL
                            m = _pat_shell.search(nota_str)
                            if m:
                                return m.group(1).strip()
                            # Fallback Aramco legacy: "151022 - 169357 - ee_s268 - EDS:... - <texto>"
                            parts = nota_str.split(" - ")
                            return parts[-1].strip() if len(parts) >= 5 else ""
                        _df_sla_ot["reporte"] = _df_sla_ot["os_fracttal"].apply(_extraer_reporte)
                    except Exception:
                        pass

                # Barra dividida: '% Uso SLA' se muestra como DOS columnas paralelas:
                #   - "Uso SLA" (barra hasta 100%): tiempo consumido dentro del umbral.
                #     Si la OT excede, esta barra queda LLENA (100%) — señal de tope.
                #   - "Exceso" (barra 0-100%): sobretiempo. VACÍA si cumple; si excedió,
                #     muestra cuánto se pasó (en % del umbral).
                #
                # Lectura visual: "barra Uso llena + barra Exceso llena" = OT muy pasada.
                # "barra Uso parcial + Exceso vacía" = OT que cumplió.
                _df_sla_ot["_uso_pct"] = _df_sla_ot["pct_sla_ot"].apply(
                    lambda v: round(min(float(v), 100.0), 1) if pd.notna(v) else None)
                _df_sla_ot["_exc_pct"] = _df_sla_ot["pct_sla_ot"].apply(
                    lambda v: round(max(float(v) - 100.0, 0.0), 1) if pd.notna(v) else None)

                # Orden solicitado por operaciones:
                # OS Fracttal | N° Aviso | Fecha llamado | Fecha atención | Cód. EDS |
                # EDS | Cliente | Técnico | Prioridad | Ciudad | Zona |
                # Tiempo resolución | Umbral SLA | Uso SLA | Exceso | Estado SLA |
                # Reporte de falla | Motivo excepción
                # (Se quitó 'Cierre completo OT' por pedido explícito.)
                _sla_ot_base = [c for c in ["os_fracttal","n_cotalker","fecha_llamado","fecha_atencion",
                                            "eds_occim","eds_nombre","cliente","tecnico",
                                            "prioridad","ciudad","zona_ot"] if c in _df_sla_ot.columns]
                _extra = ["tiempo_res","umbral_lbl","_uso_pct","_exc_pct","estado_sla"]
                # Reporte de falla ANTES de Motivo excepción (orden pedido)
                if "reporte" in _df_sla_ot.columns:
                    _extra.append("reporte")
                _hay_exc = ("excepcion_motivo" in _df_sla_ot.columns and
                            _df_sla_ot["excepcion_motivo"].notna().any())
                if _hay_exc:
                    _extra.append("excepcion_motivo")
                _df_sla_ot_disp = _df_sla_ot[_sla_ot_base + _extra].copy()
                if _hay_exc:
                    _df_sla_ot_disp["excepcion_motivo"] = _df_sla_ot_disp["excepcion_motivo"].fillna("")
                _df_sla_ot_disp["fecha_llamado"]  = pd.to_datetime(_df_sla_ot_disp["fecha_llamado"],  errors="coerce").dt.strftime("%d/%m/%Y")
                _df_sla_ot_disp["fecha_atencion"] = pd.to_datetime(_df_sla_ot_disp["fecha_atencion"], errors="coerce").dt.strftime("%d/%m/%Y")
                _df_sla_ot_disp = _df_sla_ot_disp.sort_values("fecha_llamado", ascending=False).rename(
                    columns={"os_fracttal":"OS Fracttal",
                             "fecha_llamado":"Fecha llamado",
                             "fecha_atencion":"Fecha atención",
                             "eds_occim":"Cód. EDS",
                             "eds_nombre":"EDS","cliente":"Cliente","tecnico":"Técnico",
                             "prioridad":"Prioridad","ciudad":"Ciudad","zona_ot":"Zona",
                             "tiempo_res":"Tiempo resolución","umbral_lbl":"Umbral SLA",
                             "_uso_pct":"Uso SLA","_exc_pct":"Exceso",
                             "estado_sla":"Estado SLA",
                             "excepcion_motivo":"Motivo excepción",
                             "n_cotalker":"N° Aviso","reporte":"Reporte de falla"})
                _buscar_ot = st.text_input(
                    "Buscar OT", placeholder="Ej: OS-33894",
                    key="buscar_ot_sla",
                    help="Escribe un número de OT (OS-XXXXX) y presiona Enter para filtrar",
                )
                if _buscar_ot.strip():
                    _buscar_ot_clean = _buscar_ot.strip().upper()
                    _df_sla_ot_disp = _df_sla_ot_disp[
                        _df_sla_ot_disp["OS Fracttal"].str.upper().str.contains(_buscar_ot_clean, na=False)
                    ]
                st.caption(f"**{len(_df_sla_ot_disp):,}** OTs con fechas de apertura y cierre registradas")
                _show_df(_df_sla_ot_disp, width="stretch", hide_index=True,
                    column_config={
                        "OS Fracttal":          st.column_config.TextColumn(width=110),
                        "N° Aviso":             st.column_config.TextColumn(width=105,
                            help="N° de referencia del cliente: 'No. Aviso' para COPEC / N° Cotalker para ESMAX-Aramco. Vacío = sin referencia registrada."),
                        "Fecha llamado":        st.column_config.TextColumn(width=110),
                        "Fecha atención":       st.column_config.TextColumn(width=110),
                        "Cód. EDS":          st.column_config.TextColumn(width=100),
                        "EDS":               st.column_config.TextColumn(width=210),
                        "Cliente":           st.column_config.TextColumn(width=90),
                        "Técnico":           st.column_config.TextColumn(width=155),
                        "Prioridad":         st.column_config.TextColumn(width=80),
                        "Ciudad":            st.column_config.TextColumn(width=110),
                        "Zona":              st.column_config.TextColumn(width=85),
                        "Tiempo resolución": st.column_config.TextColumn(width=120),
                        "Umbral SLA":        st.column_config.TextColumn(width=85),
                        "Uso SLA":           st.column_config.ProgressColumn(
                            label="Uso SLA (0–100%)",
                            min_value=0, max_value=100, format="%.1f%%",
                            help="Porcentaje del umbral SLA consumido, tope 100%. "
                                 "Barra completa = tiempo agotado. Si la OT excede el "
                                 "umbral, el excedente se muestra en la columna 'Exceso'."),
                        "Exceso":            st.column_config.ProgressColumn(
                            label="Exceso (>100%)",
                            min_value=0, max_value=100, format="%.1f%%",
                            help="Cuánto se excedió el umbral SLA (0 si cumple). "
                                 "Ej. si tiempo real = 125% del umbral, aquí ves 25%. "
                                 "Tope visual 100%; en el tooltip está el valor real."),
                        "Estado SLA":        st.column_config.TextColumn(width=110,
                            help="✅ Cumple = dentro del umbral SLA · ❌ No cumple = excedió el umbral · ℹ️ Excepción = excedió el umbral pero está validado por operaciones como caso ajeno a Occimiano (cuenta como cumple en los %)"),
                        "Motivo excepción":  st.column_config.TextColumn(
                            width=300,
                            help="Razón por la que la OT fue marcada como Excepción (registrada en sla_excepciones por operaciones).",
                        ),
                        "Reporte de falla":  st.column_config.TextColumn(
                            width=320,
                            help="Descripción del problema reportado por el cliente (extraído de nota_tarea Fracttal / 'Detalles del incidente' de Cotalker)",
                        ),
                    })
            else:
                st.info("No hay llamados con fechas de apertura y cierre registradas en el período seleccionado.")

        # ══════════════════════════════════════════════════════════════════════
        # SUB-PESTAÑA: SERVICIO TÉCNICO
        # ══════════════════════════════════════════════════════════════════════
        with _tab_tec:
            # ── Filtros fila 1: Empresa / Cliente / Período / Mes / Prioridad / Cumplimiento ──
            ef1, ef_cli, ef2, ef3, ef4, ef5 = st.columns([1.1, 1.3, 1.1, 1.6, 1.0, 1.3])
            with ef1:
                sel_emp_t = st.selectbox("Empresa", ["Occimiano", "Elecons", "AUTEC"],
                                         key="tec_emp")
            with ef_cli:
                # Cliente: permite cruzar Empresa (quién ejecuta) × Cliente (para quién),
                # ej. ¿cómo se desempeña AUTEC atendiendo Aramco?
                _CLI_PREF_T = ["COPEC", "Aramco (Esmax)", "ESMAX (Aramco)", "SHELL (Enex)"]
                _cli_present_t = set(df_llamados["cliente"].dropna().unique())
                _cli_opts_t = ["Todos"] + [c for c in _CLI_PREF_T if c in _cli_present_t] + \
                    sorted([c for c in _cli_present_t if c not in _CLI_PREF_T])
                sel_cli_t = st.selectbox("Cliente", _cli_opts_t, key="tec_cli")
            with ef2:
                sel_trim_t = st.selectbox("Período", _trim_opts, key="tec_trim")
            with ef3:
                if sel_trim_t != "Todos":
                    _trim_m_t = _TRIMESTRES_DEF[sel_trim_t]
                    _meses_t_disp = [l for l in _meses_disp_lbl[1:]
                                     if _MES_ABR_NUM_LL.get(l.split(" ")[0], 0) in _trim_m_t]
                else:
                    _meses_t_disp = _meses_disp_lbl[1:]
                sel_mes_t = st.multiselect("Mes", _meses_t_disp, key="tec_mes",
                                           placeholder="Todos los meses")
            with ef4:
                _pr_opts_t = ["Todas"] + sorted(df_llamados["prioridad"].dropna().unique().tolist())
                sel_pr_t = st.selectbox("Prioridad", _pr_opts_t, key="tec_pr")
            with ef5:
                sel_cu_t = st.selectbox("Cumplimiento SLA", ["Todos","CUMPLE","NO CUMPLE"], key="tec_cu")

            # ── Filtros fila 2: Equipo / Técnico (solo Occimiano) ──────────────
            if sel_emp_t == "Occimiano":
                ef6, ef7, _, _, _ = st.columns([1.5, 1.5, 1.2, 1.2, 1.5])
                with ef6:
                    sel_eq_t = st.selectbox("Equipo", _equipo_opts, key="tec_eq")
                with ef7:
                    _tec_pool_t = sorted(TECNICOS_OCCIMIANO_FULL) if sel_eq_t == "Todos" \
                                  else sorted(_equipo_to_full.get(sel_eq_t, []))
                    sel_tec_t = st.selectbox("Técnico", ["Todos"] + _tec_pool_t,
                                             key=f"tec_tec_{sel_eq_t}")
            else:
                sel_eq_t  = "Todos"
                sel_tec_t = "Todos"

            # ── Aplicar filtros ───────────────────────────────────────────────
            df_lt = df_llamados.copy()
            _fl3 = df_lt["fecha_llamado"]
            if _fl3.dt.tz is not None:
                _fl3 = _fl3.dt.tz_convert(None)
            df_lt["_mes"]   = _fl3.dt.to_period("M").astype(str)
            df_lt["_month"] = _fl3.dt.month.astype("Int64")
            # Empresa primero
            _occa_members_t = {m for g in _equipo_to_full.values() for m in g}
            if sel_emp_t == "Occimiano":
                df_lt = df_lt[df_lt["tecnico"].isin(_occa_members_t)]
            elif sel_emp_t == "Elecons":
                df_lt = df_lt[df_lt["tecnico"].str.contains("ocampo", case=False, na=False)]
            elif sel_emp_t == "AUTEC":
                df_lt = df_lt[df_lt["tecnico"].str.contains("autec", case=False, na=False)]
            if sel_trim_t != "Todos":
                df_lt = df_lt[df_lt["_month"].isin(_TRIMESTRES_DEF[sel_trim_t])]
            if sel_mes_t:
                _periods_t = [_lbl_to_period[l] for l in sel_mes_t if l in _lbl_to_period]
                if _periods_t: df_lt = df_lt[df_lt["_mes"].isin(_periods_t)]
            if sel_eq_t  != "Todos": df_lt = df_lt[df_lt["tecnico"].isin(_equipo_to_full[sel_eq_t])]
            if sel_tec_t != "Todos": df_lt = df_lt[df_lt["tecnico"] == sel_tec_t]
            if sel_pr_t  != "Todas": df_lt = df_lt[df_lt["prioridad"].str.upper() == sel_pr_t.upper()]
            if sel_cu_t  != "Todos": df_lt = df_lt[df_lt["cumplimiento"] == sel_cu_t]
            # Filtro Cliente (normalizado para que Aramco/ESMAX en cualquier variante
            # matchee — ej. 'Aramco (Esmax)' vs 'ESMAX (Aramco)').
            if sel_cli_t != "Todos":
                def _norm_cli_t(s):
                    s = str(s or "").upper()
                    if "COPEC" in s: return "COPEC"
                    if "ARAMCO" in s or "ESMAX" in s: return "ARAMCO"
                    if "SHELL" in s or "ENEX" in s: return "SHELL"
                    return s.strip()
                _sel_key = _norm_cli_t(sel_cli_t)
                df_lt = df_lt[df_lt["cliente"].apply(_norm_cli_t) == _sel_key]

            # ── KPIs ──────────────────────────────────────────────────────────
            _cumple_t   = (df_lt["cumplimiento"] == "CUMPLE").sum()
            _nocumple_t = (df_lt["cumplimiento"] == "NO CUMPLE").sum()
            _pct_t      = round(_cumple_t/(_cumple_t+_nocumple_t)*100,1) if (_cumple_t+_nocumple_t)>0 else 0
            if sel_emp_t == "Occimiano":
                if sel_tec_t != "Todos":
                    _tec_activos_cnt = 1
                elif sel_eq_t != "Todos":
                    _tec_activos_cnt = len(_equipo_to_full.get(sel_eq_t, []))
                else:
                    _tec_activos_cnt = sum(len(v) for v in _equipo_to_full.values())
            else:
                _tec_activos_cnt = int(df_lt["tecnico"].nunique()) if not df_lt.empty else 0
            tk1, tk2, tk3, tk4, tk5 = st.columns(5)
            tk1.metric("Total llamados", f"{len(df_lt):,}")
            tk2.metric("P1 (máquina detenida)", f"{(df_lt['prioridad'].str.upper()=='P1').sum():,}")
            tk3.metric("Cumple SLA", f"{_cumple_t:,}", delta=f"{_pct_t}%")
            tk4.metric("No cumple SLA", f"{_nocumple_t:,}",
                       delta=f"-{100-_pct_t:.1f}%" if (_cumple_t+_nocumple_t)>0 else None,
                       delta_color="inverse")
            tk5.metric("Técnicos activos", str(_tec_activos_cnt))

            st.divider()
            _ll_sig_t = f"{len(df_llamados)}_{sel_emp_t}_{sel_cli_t}_{sel_trim_t}_{sel_mes_t}_{sel_eq_t}_{sel_tec_t}_{sel_pr_t}_{sel_cu_t}"

            # ── SECCIÓN: Cumplimiento SLA ─────────────────────────────────────
            st.markdown('<div class="section-header">📊  Cumplimiento SLA</div>', unsafe_allow_html=True)

            # ── Velocímetro (común a todas las empresas) ──────────────────────
            _tgauge_k = f"_fig_tec_gauge_{_current_theme}_{_ll_sig_t}"
            if _tgauge_k not in st.session_state:
                import math as _math_t
                _tg_clr = (
                    "#22c55e" if _pct_t >= 90 else
                    "#84cc16" if _pct_t >= 75 else
                    "#f97316" if _pct_t >= 60 else "#ef4444"
                )
                def _arc_t(vmin, vmax, r_in, r_out, n=80):
                    a0 = _math_t.pi*(1 - vmax/100)
                    a1 = _math_t.pi*(1 - vmin/100)
                    t  = [a0+(a1-a0)*i/(n-1) for i in range(n)]
                    xo = [r_out*_math_t.cos(a) for a in t]
                    yo = [r_out*_math_t.sin(a) for a in t]
                    xi = [r_in *_math_t.cos(a) for a in reversed(t)]
                    yi = [r_in *_math_t.sin(a) for a in reversed(t)]
                    return xo+xi+[xo[0]], yo+yi+[yo[0]]
                _R_OUT_T, _R_IN_T = 1.0, 0.52
                _zones_t = [
                    (0,  10, "#c0392b"),(10, 20, "#e74c3c"),(20, 35, "#e67e22"),
                    (35, 50, "#f39c12"),(50, 65, "#f1c40f"),(65, 75, "#d4e157"),
                    (75, 85, "#a5d63a"),(85, 92, "#66bb6a"),(92, 97, "#43a047"),
                    (97,100, "#2e7d32"),
                ]
                _fig_tgauge = go.Figure()
                for _vn, _vx, _clr in _zones_t:
                    _xz, _yz = _arc_t(_vn, _vx, _R_IN_T, _R_OUT_T)
                    _fig_tgauge.add_trace(go.Scatter(x=_xz, y=_yz, fill="toself",
                        fillcolor=_clr, line=dict(color=_clr, width=0.5),
                        mode="lines", showlegend=False, hoverinfo="skip"))
                for _rb0, _rb1, _rbc in [(_R_OUT_T, _R_OUT_T+0.025,"#2d3436"),
                                          (_R_IN_T-0.025, _R_IN_T,  "#2d3436")]:
                    _xb, _yb = _arc_t(0, 100, _rb0, _rb1)
                    _fig_tgauge.add_trace(go.Scatter(x=_xb, y=_yb, fill="toself",
                        fillcolor=_rbc, line=dict(color=_rbc, width=0),
                        mode="lines", showlegend=False, hoverinfo="skip"))
                _ang_t = _math_t.pi*(1 - _pct_t/100)
                _nl_t, _bw_t = 0.78, 0.025
                _al_t, _ar_t = _ang_t+_math_t.pi/2, _ang_t-_math_t.pi/2
                _fig_tgauge.add_trace(go.Scatter(
                    x=[_bw_t*_math_t.cos(_al_t), _nl_t*_math_t.cos(_ang_t),
                       _bw_t*_math_t.cos(_ar_t), _bw_t*_math_t.cos(_al_t)],
                    y=[_bw_t*_math_t.sin(_al_t), _nl_t*_math_t.sin(_ang_t),
                       _bw_t*_math_t.sin(_ar_t), _bw_t*_math_t.sin(_al_t)],
                    fill="toself", fillcolor="#2d3436", line=dict(color="#2d3436", width=0.5),
                    mode="lines", showlegend=False, hoverinfo="skip"))
                _fig_tgauge.add_trace(go.Scatter(
                    x=[_bw_t*_math_t.cos(_al_t), -0.18*_math_t.cos(_ang_t),
                       _bw_t*_math_t.cos(_ar_t), _bw_t*_math_t.cos(_al_t)],
                    y=[_bw_t*_math_t.sin(_al_t), -0.18*_math_t.sin(_ang_t),
                       _bw_t*_math_t.sin(_ar_t), _bw_t*_math_t.sin(_al_t)],
                    fill="toself", fillcolor="#555", line=dict(color="#555", width=0.5),
                    mode="lines", showlegend=False, hoverinfo="skip"))
                for _hr_t, _fc_t in [(0.10,"#2d3436"),(0.06,"#ffffff")]:
                    _fig_tgauge.add_trace(go.Scatter(
                        x=[_hr_t*_math_t.cos(2*_math_t.pi*i/40) for i in range(41)],
                        y=[_hr_t*_math_t.sin(2*_math_t.pi*i/40) for i in range(41)],
                        fill="toself", fillcolor=_fc_t, line=dict(color=_fc_t),
                        mode="lines", showlegend=False, hoverinfo="skip"))
                for _tv in [0, 25, 50, 75, 90, 100]:
                    _ta_t  = _math_t.pi*(1 - _tv/100)
                    _is90  = _tv == 90
                    _tc_t  = "#e74c3c" if _is90 else "#ecf0f1"
                    _fig_tgauge.add_trace(go.Scatter(
                        x=[(_R_OUT_T+0.07)*_math_t.cos(_ta_t),
                           (1.18 if _is90 else 1.13)*_math_t.cos(_ta_t)],
                        y=[(_R_OUT_T+0.07)*_math_t.sin(_ta_t),
                           (1.18 if _is90 else 1.13)*_math_t.sin(_ta_t)],
                        line=dict(color=_tc_t, width=3 if _is90 else 2),
                        mode="lines", showlegend=False, hoverinfo="skip"))
                    _fig_tgauge.add_annotation(
                        x=1.30*_math_t.cos(_ta_t), y=1.30*_math_t.sin(_ta_t),
                        text=f"<b>{_tv}%✓</b>" if _is90 else f"{_tv}%",
                        showarrow=False,
                        font=dict(size=10 if _is90 else 9, color=_tc_t))
                # Título dinámico según contexto
                if sel_emp_t == "Occimiano":
                    _tg_lbl = (sel_tec_t if sel_tec_t != "Todos"
                               else (sel_eq_t if sel_eq_t != "Todos"
                                     else "Occimiano — todos los equipos"))
                else:
                    _tg_lbl = sel_emp_t
                _fig_tgauge.add_annotation(x=0, y=-0.15,
                    text=f"<b>{_pct_t:.1f}%</b>",
                    showarrow=False, font=dict(size=36, color=_tg_clr))
                _fig_tgauge.add_annotation(x=0, y=-0.38,
                    text=f"{int(_cumple_t)} cumple · {int(_nocumple_t)} no cumple",
                    showarrow=False, font=dict(size=11, color=_t["muted"]))
                _fig_tgauge.add_annotation(x=0, y=1.30,
                    text=f"<b>Cumplimiento SLA</b>",
                    showarrow=False, font=dict(size=14, color=_t["text"]))
                _fig_tgauge.update_layout(
                    xaxis=dict(range=[-1.2, 1.2], visible=False, scaleanchor="y"),
                    yaxis=dict(range=[-0.50, 1.42], visible=False),
                    height=400, margin=dict(t=5, b=5, l=5, r=5),
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                )
                st.session_state[_tgauge_k] = _fig_tgauge
            _fig_tgauge_show = st.session_state.get(_tgauge_k)

            # ── Comparativo por empresa (tabla resumen) ───────────────────────
            _df_comp = df_llamados.copy()
            _fl_comp = _df_comp["fecha_llamado"]
            if _fl_comp.dt.tz is not None:
                _fl_comp = _fl_comp.dt.tz_convert(None)
            _df_comp["_month"] = _fl_comp.dt.month.astype("Int64")
            _df_comp["_mes"]   = _fl_comp.dt.to_period("M").astype(str)
            if sel_trim_t != "Todos":
                _df_comp = _df_comp[_df_comp["_month"].isin(_TRIMESTRES_DEF[sel_trim_t])]
            if sel_mes_t:
                _periods_comp = [_lbl_to_period[l] for l in sel_mes_t if l in _lbl_to_period]
                if _periods_comp:
                    _df_comp = _df_comp[_df_comp["_mes"].isin(_periods_comp)]
            if sel_pr_t != "Todas":
                _df_comp = _df_comp[_df_comp["prioridad"].str.upper() == sel_pr_t.upper()]
            _df_occ_c  = _df_comp[_df_comp["tecnico"].isin(_occa_members_t)]
            _df_autec_c = _df_comp[_df_comp["tecnico"].str.contains("autec", case=False, na=False)]
            _df_elec_c  = _df_comp[_df_comp["tecnico"].str.contains("ocampo", case=False, na=False)]
            _comp_rows = []
            for _emp_lbl, _df_e in [("Occimiano", _df_occ_c), ("Elecons", _df_elec_c), ("AUTEC", _df_autec_c)]:
                _cc  = int((_df_e["cumplimiento"] == "CUMPLE").sum())
                _cnc = int((_df_e["cumplimiento"] == "NO CUMPLE").sum())
                _ct  = _cc + _cnc
                _cpct = round(_cc / _ct * 100, 1) if _ct > 0 else 0.0
                _comp_rows.append({
                    "Empresa":    _emp_lbl,
                    "Total":      _ct,
                    "Cumple":     _cc,
                    "No Cumple":  _cnc,
                    "% SLA":      f"{_cpct}%" if _ct > 0 else "—",
                })
            _comp_df = pd.DataFrame(_comp_rows)

            # ── Layout: velocímetro + tabla comparativa empresa ───────────────
            _tc0, _tc1 = st.columns([0.9, 1.6])
            with _tc0:
                if _fig_tgauge_show:
                    st.plotly_chart(_fig_tgauge_show, width="stretch")
                else:
                    st.info("Sin datos para el período seleccionado.")
            with _tc1:
                if not _comp_df.empty:
                    _muted_c = _t["muted"]
                    st.markdown(
                        f"<p style='margin:0 0 6px 0; font-size:13px; font-weight:600;"
                        f" color:{_muted_c};'>Resumen por empresa</p>",
                        unsafe_allow_html=True,
                    )
                    _, _ctbl, _ = st.columns([0.05, 0.9, 0.05])
                    with _ctbl:
                        _show_df(_comp_df, hide_index=True, use_container_width=True,
                            column_config={
                                "Empresa":    st.column_config.TextColumn(width=85),
                                "Total":      st.column_config.NumberColumn(format="%d", width=50),
                                "Cumple":     st.column_config.NumberColumn(format="%d", width=55),
                                "No Cumple":  st.column_config.NumberColumn(format="%d", width=72),
                                "% SLA":      st.column_config.TextColumn(width=55),
                            })

            # ── Mini-tabla resumen (solo Occimiano) ───────────────────────────
            if sel_emp_t == "Occimiano":
                _all_mbs_t2 = {m for g in _equipo_to_full.values() for m in g}
                _df_lt_rs   = df_lt[df_lt["tecnico"].isin(_all_mbs_t2)]
                if sel_tec_t != "Todos":
                    _rs_grp   = [sel_tec_t]
                    _rs_label = "Técnico"
                    _rs_mask  = lambda n: _df_lt_rs[_df_lt_rs["tecnico"] == n]
                elif sel_eq_t != "Todos":
                    _rs_grp   = _equipo_to_full.get(sel_eq_t, [])
                    _rs_label = "Técnico"
                    _rs_mask  = lambda n: _df_lt_rs[_df_lt_rs["tecnico"] == n]
                else:
                    _rs_grp   = list(_equipo_to_full.keys())
                    _rs_label = "Equipo"
                    _rs_mask  = lambda n: _df_lt_rs[_df_lt_rs["tecnico"].isin(_equipo_to_full[n])]
                _rs_rows = []
                for _rn in _rs_grp:
                    _drs = _rs_mask(_rn)
                    _rc  = (_drs["cumplimiento"] == "CUMPLE").sum()
                    _rnc = (_drs["cumplimiento"] == "NO CUMPLE").sum()
                    _rt  = _rc + _rnc
                    _rs_rows.append({
                        _rs_label: _rn,
                        "Total": len(_drs),
                        "Cumple": int(_rc),
                        "No Cumple": int(_rnc),
                        "% SLA": f"{round(_rc/_rt*100,1)}%" if _rt > 0 else "—",
                    })
                if _rs_rows:
                    _show_df(pd.DataFrame(_rs_rows), hide_index=True, width="stretch",
                        column_config={
                            _rs_label:   st.column_config.TextColumn(width=140),
                            "Total":     st.column_config.NumberColumn(format="%d", width=60),
                            "Cumple":    st.column_config.NumberColumn(format="%d", width=70),
                            "No Cumple": st.column_config.NumberColumn(format="%d", width=80),
                            "% SLA":     st.column_config.TextColumn(width=65),
                        })

            # ── GRÁFICO 2: Llamados por técnico (stacked prioridad) ───────────
            st.divider()
            st.markdown('<div class="section-header">👷  Llamados correctivos por técnico</div>', unsafe_allow_html=True)
            _tbar_k = f"_fig_tec_bar_tec_{_current_theme}_{_ll_sig_t}"
            if _tbar_k not in st.session_state:
                _all_mbs3 = {m for g in _equipo_to_full.values() for m in g}
                _df_bar = df_lt[df_lt["tecnico"].isin(_all_mbs3)]
                if not _df_bar.empty:
                    _tec_prio = _df_bar.groupby(["tecnico","prioridad"]).size().reset_index(name="llamados")
                    _tec_ord  = (_tec_prio.groupby("tecnico")["llamados"].sum()
                                 .sort_values(ascending=False).index.tolist())
                    _fig_bar = px.bar(
                        _tec_prio, x="tecnico", y="llamados", color="prioridad",
                        color_discrete_map=prio_colors,
                        title="Llamados correctivos por técnico",
                        barmode="stack",
                        category_orders={"tecnico": _tec_ord},
                        labels={"tecnico":"Técnico","llamados":"Llamados","prioridad":"Prioridad"},
                        text_auto=True,
                    )
                    _fig_bar.update_layout(xaxis_tickangle=-30, legend_title="Prioridad",
                                           xaxis_title="", height=430, margin=dict(t=50,b=80))
                    _apply_plot_theme(_fig_bar)
                    st.session_state[_tbar_k] = _fig_bar
                else:
                    st.session_state[_tbar_k] = None
            _fig_bar_show = st.session_state.get(_tbar_k)
            if _fig_bar_show:
                st.plotly_chart(_fig_bar_show, width="stretch")
            else:
                st.info("Sin datos para el período seleccionado.")

            # ── GRÁFICO 3: Evolución temporal ─────────────────────────────────
            st.divider()
            st.markdown('<div class="section-header">📈  Evolución temporal</div>', unsafe_allow_html=True)
            _tev_grain = st.radio("Granularidad", ["Mensual","Trimestral"],
                                  horizontal=True, key="tec_grain")
            _tev_k = f"_fig_tec_evol_{_current_theme}_{_ll_sig_t}_{_tev_grain}"
            if _tev_k not in st.session_state:
                _all_mbs4 = {m for g in _equipo_to_full.values() for m in g}
                _df_ev = df_lt[df_lt["tecnico"].isin(_all_mbs4)].copy()

                if _tev_grain == "Mensual":
                    _df_ev["periodo_lbl"] = _df_ev["_mes"].apply(_ym_a_lbl)
                    _periodo_ord = [_ym_a_lbl(m) for m in sorted(_df_ev["_mes"].dropna().unique())]
                else:
                    def _trim_lbl(m):
                        try:
                            mn = int(str(m).split("-")[1])
                            t  = (mn - 1) // 3 + 1
                            yr = str(m).split("-")[0][2:]
                            return f"T{t} '{yr}"
                        except Exception:
                            return str(m)
                    _df_ev["periodo_lbl"] = _df_ev["_mes"].apply(_trim_lbl)
                    _uniq_p = sorted(_df_ev["_mes"].dropna().unique())
                    _seen_p = []
                    _periodo_ord = []
                    for _mp in _uniq_p:
                        _lp = _trim_lbl(_mp)
                        if _lp not in _seen_p:
                            _seen_p.append(_lp)
                            _periodo_ord.append(_lp)

                if sel_tec_t != "Todos":
                    _df_ev["_grp_col"] = _df_ev["tecnico"]
                    _grp_title = "Técnico"
                elif sel_eq_t != "Todos":
                    _df_ev["_grp_col"] = _df_ev["tecnico"]
                    _grp_title = "Técnico"
                else:
                    def _tec_to_eq(tec):
                        for _eq, _mbs in _equipo_to_full.items():
                            if tec in _mbs: return _eq
                        return "Otros"
                    _df_ev["_grp_col"] = _df_ev["tecnico"].apply(_tec_to_eq)
                    _grp_title = "Equipo"

                if not _df_ev.empty:
                    _monthly_t = (_df_ev.groupby(["periodo_lbl","_grp_col"])
                                  .size().reset_index(name="llamados")
                                  .rename(columns={"_grp_col": _grp_title}))
                    _fig_ev = px.bar(
                        _monthly_t, x="periodo_lbl", y="llamados", color=_grp_title,
                        title=f"Llamados por {'mes' if _tev_grain == 'Mensual' else 'trimestre'} "
                              f"y {_grp_title.lower()}",
                        barmode="group",
                        category_orders={"periodo_lbl": _periodo_ord},
                        labels={"periodo_lbl": "Período", "llamados": "Llamados"},
                    )
                    _fig_ev.update_layout(xaxis_title="", yaxis_title="Llamados",
                                          xaxis_type="category", height=400,
                                          margin=dict(t=50,b=10))
                    _apply_plot_theme(_fig_ev)
                    st.session_state[_tev_k] = _fig_ev
                else:
                    st.session_state[_tev_k] = None
            _fig_ev_show = st.session_state.get(_tev_k)
            if _fig_ev_show:
                st.plotly_chart(_fig_ev_show, width="stretch")
            else:
                st.info("Sin datos para el período seleccionado.")

            # ── Detalle de llamados ───────────────────────────────────────────
            st.divider()
            st.markdown('<div class="section-header">Detalle de llamados</div>', unsafe_allow_html=True)
            df_det = df_lt.copy()
            if TECNICOS_OCCIMIANO_FULL:
                df_det = df_det[df_det["tecnico"].isin(TECNICOS_OCCIMIANO_FULL)]

            def _td_horas(v):
                if v is None or (hasattr(v,"__class__") and v.__class__.__name__ == "NaTType"):
                    return float("nan")
                try:
                    return pd.to_timedelta(v).total_seconds() / 3600
                except Exception:
                    return float("nan")

            def _fmt_h_min(h) -> str:
                if pd.isna(h) or h < 0: return "⏳ Pendiente"
                hh = int(h); mm = int(round((h % 1) * 60))
                if hh == 0: return f"{mm}min"
                if mm == 0: return f"{hh}h"
                return f"{hh}h {mm:02d}min"

            def _zona_key_det(z: str) -> str:
                z = str(z).upper().strip()
                if z in ("RM","R.M.") or any(k in z for k in ["SANTIAGO","METRO"]): return "Santiago"
                if any(k in z for k in ["NORTE","SUR","CENTRO","REGION","REGIONES","RG"]): return "Regiones"
                return "Santiago"

            df_det["_hrs_real"] = df_det.get("tiempo_resp_real", pd.Series(dtype=object,index=df_det.index)).apply(_td_horas)
            df_det["_zona_key"] = df_det.get("zona", pd.Series("",index=df_det.index)).apply(_zona_key_det)
            df_det["_sla_h"]    = [
                (SLA_HOURS.get(str(c), SLA_DEFAULT).get(str(p).upper(), {}).get(str(z), None))
                for c, p, z in zip(
                    df_det.get("cliente",  pd.Series("", index=df_det.index)),
                    df_det.get("prioridad",pd.Series("", index=df_det.index)),
                    df_det["_zona_key"],
                )
            ]
            df_det["_pct_sla"] = (df_det["_hrs_real"] / df_det["_sla_h"] * 100).where(
                df_det["_hrs_real"].notna() & pd.Series([pd.notna(x) for x in df_det["_sla_h"]], index=df_det.index)
            )
            df_det["Fecha Llamado"]    = pd.to_datetime(df_det["fecha_llamado"], errors="coerce").dt.strftime("%d/%m/%Y")
            df_det["Tiempo Respuesta"] = df_det["_hrs_real"].apply(_fmt_h_min)
            df_det["SLA"]              = [f"{int(h)}h" if pd.notna(h) else "—" for h in df_det["_sla_h"]]
            df_det["% SLA"]            = df_det["_pct_sla"]
            df_det["Cumplimiento"]     = df_det.get("cumplimiento", pd.Series("—", index=df_det.index)).apply(
                lambda v: "✅" if str(v).upper() == "CUMPLE" else ("❌" if str(v).upper() == "NO CUMPLE" else "—")
            )
            _det_cols = [c for c in ["Fecha Llamado","eds_occim","eds_nombre","cliente","prioridad",
                                     "tecnico","Tiempo Respuesta","SLA","% SLA","Cumplimiento","os_fracttal"]
                         if c in df_det.columns]
            df_det_disp = (df_det[_det_cols]
                           .rename(columns={"eds_occim":"Cód. EDS","eds_nombre":"EDS","cliente":"Cliente",
                                            "prioridad":"Prioridad","tecnico":"Técnico","os_fracttal":"OS Fracttal"})
                           .sort_values("Fecha Llamado", ascending=False))
            st.caption(f"**{len(df_det_disp):,}** llamados atendidos por técnicos Occimiano")
            _show_df(df_det_disp, width="stretch", hide_index=True,
                column_config={
                    "Fecha Llamado":    st.column_config.TextColumn(width=105),
                    "Cód. EDS":         st.column_config.TextColumn(width=90),
                    "EDS":              st.column_config.TextColumn(width=220),
                    "Cliente":          st.column_config.TextColumn(width=90),
                    "Prioridad":        st.column_config.TextColumn(width=75),
                    "Técnico":          st.column_config.TextColumn(width=180),
                    "Tiempo Respuesta": st.column_config.TextColumn(width=120),
                    "SLA":              st.column_config.TextColumn(width=65),
                    "% SLA":            st.column_config.ProgressColumn(
                        label="% SLA usado", min_value=0, max_value=200, format="%.1f%%"),
                    "Cumplimiento":     st.column_config.TextColumn(width=90),
                    "OS Fracttal":      st.column_config.TextColumn(width=105),
                })

        # ══════════════════════════════════════════════════════════════════════
        # SUB-PESTAÑA: SLA EN CURSO  (radar de correctivas activas)
        # Muestra las correctivas ABIERTAS (fecha_finalizacion IS NULL) con su
        # cronómetro hasta el vencimiento del SLA. Distingue:
        #   - Sin iniciar (fecha_inicio IS NULL) = nadie ha visto la OT
        #   - En curso (fecha_inicio IS NOT NULL) = técnico ya trabajando
        # Colores por urgencia:
        #   verde >50% del umbral restante · amarillo 25-50% · rojo <25% o vencido
        # ══════════════════════════════════════════════════════════════════════
        with _tab_en_curso:
            st.markdown(
                '<div class="section-header">⏳  SLA en curso — correctivas activas</div>',
                unsafe_allow_html=True,
            )
            st.caption(
                "Monitor en tiempo real de correctivas NO cerradas. Cada OT tiene su "
                "cronómetro descontando desde `fecha_llamado + umbral SLA`. "
                "**🟢 verde** = >50% del umbral restante · **🟡 amarillo** = 25–50% · "
                "**🔴 rojo** = <25% o vencido · **⚫ negro** = sin iniciar (nadie la ha tomado)."
            )

            # ── Filtros (Cliente + Mes) + Botón + timestamp ────────────────
            _c_cli, _c_mes, _c_refresh, _c_ts = st.columns([1.6, 1.6, 1, 3])
            with _c_cli:
                _CLI_PREF_LIVE = ["COPEC", "Aramco (Esmax)", "ESMAX (Aramco)",
                                  "SHELL (Enex)", "OCCIMIANO", "ABASTIBLE"]
                sel_cli_live = st.selectbox(
                    "Cliente", ["Todos"] + _CLI_PREF_LIVE,
                    key="sla_curso_cli",
                )
            with _c_mes:
                # Meses disponibles (año actual hasta el mes actual)
                _MES_ES_LIVE = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
                                7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}
                _year = pd.Timestamp.now().year
                _cur_month = pd.Timestamp.now().month
                _meses_opts = ["Todos"] + [f"{_MES_ES_LIVE[m]} {_year}"
                                           for m in range(_cur_month, 0, -1)]
                sel_mes_live = st.selectbox("Mes (por T0)", _meses_opts,
                                            key="sla_curso_mes")
            with _c_refresh:
                st.write("")   # spacer para alinear con selectbox
                if st.button("🔄 Actualizar", key="sla_curso_refresh",
                             use_container_width=True):
                    st.rerun()
            _hoy_now = pd.Timestamp.now()
            with _c_ts:
                st.markdown(
                    f'<div style="padding:32px 0 0 0;color:{_t["muted"]};font-size:0.82rem;">'
                    f'Última actualización: <b>{_hoy_now.strftime("%d/%m/%Y %H:%M:%S")}</b> · '
                    f'usa el botón para refrescar cronómetros.'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            # ── Traer correctivas ABIERTAS desde ordenes_trabajo ─────────
            # Excluir estados basura/anulados que no cuentan como OT operativa:
            # ERROR DE INGRESO, DUPLICADO, DE PRUEBA, etc. + cualquier variante
            # de Cancelada. Esto quita ~80% del ruido en la vista.
            _ESTADOS_EXCLUIR = (
                "Canceladas,Cancelada,Cancelado,ERROR DE INGRESO,DUPLICADO,"
                "Duplicidad,DE PRUEBA,PRUEBA ROBOT,EQUIPO CON RECAMBIO,"
                "FUE REPETIDA EN OTRA OS,PLAN INCOMPLETO,PENDIENTE,Finalizadas"
            )
            try:
                from supabase_client import _query as _sc_query
                _rows_ab = _sc_query(
                    "ordenes_trabajo",
                    "select=id_ot,cliente,estacion,codigo_eds,responsable,"
                    "prioridad_calc,fecha_creacion,fecha_incidente,fecha_inicio,"
                    "fecha_finalizacion,estado,estado_tarea,nombre_activo,nota_tarea"
                    "&tipo_tarea=ilike.*CORRECTIV*"
                    "&fecha_finalizacion=is.null"
                    f"&estado=not.in.({_ESTADOS_EXCLUIR})"
                    "&fecha_creacion=gte.2026-01-01",
                    limit=2000,
                )
            except Exception as _e_ab:
                st.error(f"No se pudo cargar correctivas abiertas: {_e_ab}")
                _rows_ab = []

            _df_ab = pd.DataFrame(_rows_ab) if _rows_ab else pd.DataFrame()

            # Helpers globales del tab (se usan en KPIs de abiertas + historico + ranking).
            def _norm_cli_live(s):
                s = str(s or "").upper()
                if "COPEC" in s: return "COPEC"
                if "ARAMCO" in s or "ESMAX" in s: return "ARAMCO"
                if "SHELL" in s or "ENEX" in s: return "SHELL"
                if "OCCIMIANO" in s: return "OCCIMIANO"
                if "ABASTIBLE" in s: return "ABASTIBLE"
                return s.strip()
            _sel_norm = _norm_cli_live(sel_cli_live) if sel_cli_live != "Todos" else ""

            if not _rows_ab:
                st.success("✅ ¡Excelente! No hay correctivas abiertas en este momento.")
            elif _df_ab.empty:
                pass  # imposible ya que _rows_ab es truthy → _df_ab tendría filas
            else:
                # Aplicar filtro de Cliente (normalizando Aramco/ESMAX)
                if sel_cli_live != "Todos":
                    _df_ab = _df_ab[_df_ab["cliente"].apply(_norm_cli_live) == _sel_norm]
                # Aplicar filtro de Mes por fecha_creacion (T0 se calcula despues)
                if sel_mes_live != "Todos":
                    _mes_num = {v: k for k, v in _MES_ES_LIVE.items()}.get(
                        sel_mes_live.split(" ")[0])
                    if _mes_num:
                        _fc = pd.to_datetime(_df_ab["fecha_creacion"],
                                             errors="coerce", utc=True)
                        _df_ab = _df_ab[
                            (_fc.dt.year == _year) & (_fc.dt.month == _mes_num)
                        ]
                if _df_ab.empty:
                    _ctx_bits = [sel_cli_live] if sel_cli_live != "Todos" else []
                    if sel_mes_live != "Todos": _ctx_bits.append(sel_mes_live)
                    _ctx = ", ".join(_ctx_bits) or "el filtro actual"
                    st.info(f"No hay correctivas abiertas para **{_ctx}**.")

            if not _df_ab.empty:
                # ── N° Aviso (Cotalker para Aramco / No. Aviso para COPEC) ─
                # load_cotalker_index_supabase() cachea por 30min, es barato.
                try:
                    _n_aviso_idx = load_cotalker_index_supabase()
                except Exception:
                    _n_aviso_idx = {}
                _df_ab["_n_aviso"] = _df_ab["id_ot"].astype(str).map(_n_aviso_idx).fillna("—")

                # ── Enrich: umbral SLA desde sla_umbrales_horas ─────────
                # cliente + prioridad_calc + zona_sla (default Regiones)
                _sla_map = load_sla_umbrales_supabase()   # {cliente:{prio:{zona:h}}}

                # Determinar zona SLA por EDS (Santiago vs Regiones)
                _eds_zonas = {}
                try:
                    _eds_rows = _sc_query("estaciones_servicio",
                        "select=eds_occim,zona_sla", limit=2000)
                    _eds_zonas = {r.get("eds_occim"): (r.get("zona_sla") or "Regiones")
                                  for r in _eds_rows}
                except Exception:
                    pass

                def _umbral_horas(row):
                    cli = str(row.get("cliente") or "")
                    prio = str(row.get("prioridad_calc") or "").upper()
                    zona = _eds_zonas.get(row.get("codigo_eds"), "Regiones")
                    try:
                        return _sla_map.get(cli, {}).get(prio, {}).get(zona)
                    except Exception:
                        return None
                _df_ab["_umbral_h"] = _df_ab.apply(_umbral_horas, axis=1)

                # T0 = fecha_incidente si existe, sino fecha_creacion
                def _t0(row):
                    for c in ("fecha_incidente","fecha_creacion"):
                        v = row.get(c)
                        if v:
                            ts = pd.to_datetime(v, errors="coerce", utc=True)
                            if pd.notna(ts):
                                return ts.tz_convert("America/Santiago").tz_localize(None)
                    return pd.NaT
                _df_ab["_t0"]       = _df_ab.apply(_t0, axis=1)
                _df_ab["_deadline"] = _df_ab.apply(
                    lambda r: r["_t0"] + pd.Timedelta(hours=r["_umbral_h"])
                    if pd.notna(r["_t0"]) and pd.notna(r["_umbral_h"]) else pd.NaT,
                    axis=1,
                )
                _df_ab["_restante_min"] = _df_ab["_deadline"].apply(
                    lambda dl: (dl - _hoy_now).total_seconds() / 60
                    if pd.notna(dl) else None
                )
                _df_ab["_pct_restante"] = _df_ab.apply(
                    lambda r: (r["_restante_min"] / (r["_umbral_h"] * 60) * 100)
                    if pd.notna(r["_restante_min"]) and pd.notna(r["_umbral_h"]) and r["_umbral_h"] > 0
                    else None,
                    axis=1,
                )
                # Estado
                def _estado_ot(row):
                    if pd.isna(row.get("fecha_inicio")):
                        return "sin_iniciar"
                    return "en_curso"
                _df_ab["_estado_ot"] = _df_ab.apply(_estado_ot, axis=1)
                _df_ab["_vencida"]   = _df_ab["_restante_min"].apply(
                    lambda m: bool(pd.notna(m) and m < 0))

                # Color por urgencia
                def _color_urg(pct, vencida):
                    if vencida: return "🔴"
                    if pd.isna(pct): return "⚪"
                    if pct <= 25:  return "🔴"
                    if pct <= 50:  return "🟡"
                    return "🟢"
                _df_ab["_semaf"] = _df_ab.apply(
                    lambda r: _color_urg(r["_pct_restante"], r["_vencida"]), axis=1)

                # Formatear cronómetro humano. Si está vencida, mostrar
                # '🔴 VENCIDA' en vez del negativo (más claro visualmente).
                def _fmt_crono(min_):
                    if pd.isna(min_):
                        return "—"
                    if min_ < 0:
                        return "🔴 VENCIDA"
                    m = int(min_)
                    dias = m // (60*24)
                    horas = (m % (60*24)) // 60
                    mins = m % 60
                    partes = []
                    if dias:  partes.append(f"{dias}d")
                    if horas: partes.append(f"{horas}h")
                    if mins:  partes.append(f"{mins}m")
                    return " ".join(partes) or "0m"
                _df_ab["_crono"] = _df_ab["_restante_min"].apply(_fmt_crono)

                # Extraer descripción del cliente desde nota_tarea
                _pat_det_ar = re.compile(
                    r"Detalles del incidente:\s*(.+?)(?:\n\n|$)", re.IGNORECASE | re.DOTALL)
                _pat_desc_copec = re.compile(
                    r"Descripci[oó]n:\s*(.+?)(?:\n\n|$)", re.IGNORECASE | re.DOTALL)
                def _extraer_req(nota):
                    txt = str(nota or "")
                    for pat in (_pat_det_ar, _pat_desc_copec):
                        m = pat.search(txt)
                        if m:
                            return m.group(1).strip()[:120]
                    # Fallback: última parte separada por " - "
                    parts = [p.strip() for p in txt.split(" - ") if p.strip()]
                    if len(parts) >= 5:
                        return parts[-1][:120]
                    return "—"
                _df_ab["_req"] = _df_ab["nota_tarea"].apply(_extraer_req)

                # Tipo de equipo desde nombre_activo
                _EQUIPO_TIPOS_LIVE = [
                    (r"HIDROLAVAD", "💦 Hidrolavadora"),
                    (r"LAVAINT",    "🧼 Lavainteriores"),
                    (r"LAVAD",      "🚿 Lavadora"),
                    (r"ASPIRA",     "🧹 Aspiradora"),
                    (r"ABLAND",     "💧 Ablandador"),
                    (r"HIDROPACK",  "🛢️ Hidropack"),
                    (r"BOMBA",      "⚙️ Bomba"),
                    (r"TERMO",      "🔥 Termo"),
                    (r"COMPRESOR",  "🌬️ Compresor"),
                ]
                def _tipo_eq(nom):
                    s = str(nom or "").upper()
                    if not s: return "🔧 Otro"
                    tipos = [lbl for pat, lbl in _EQUIPO_TIPOS_LIVE if re.search(pat, s)]
                    return " + ".join(tipos) if tipos else "🔧 Otro"
                _df_ab["_tipo_eq"] = _df_ab["nombre_activo"].apply(_tipo_eq)

                # Formatear T0 y umbral legible
                _df_ab["_t0_lbl"] = _df_ab["_t0"].apply(
                    lambda t: t.strftime("%d/%m/%Y %H:%M") if pd.notna(t) else "—")
                _df_ab["_umbral_lbl"] = _df_ab["_umbral_h"].apply(
                    lambda h: f"{int(h)}h" if pd.notna(h) else "—")
                _df_ab["_estado_lbl"] = _df_ab["_estado_ot"].map({
                    "sin_iniciar": "⚫ Sin iniciar",
                    "en_curso":    "🟢 En curso",
                })

                # ── KPIs superiores ────────────────────────────────────
                _tot = len(_df_ab)
                _venc = int(_df_ab["_vencida"].sum())
                _sin_ini = int((_df_ab["_estado_ot"] == "sin_iniciar").sum())
                _rojas = int((_df_ab["_semaf"] == "🔴").sum())
                _criticas = _df_ab[(_df_ab["_semaf"] == "🔴") &
                                   (_df_ab["_estado_ot"] == "sin_iniciar")]
                _n_criticas = len(_criticas)

                _k1, _k2, _k3, _k4, _k5 = st.columns(5)
                _k1.metric("Correctivas abiertas", f"{_tot:,}")
                _k2.metric("Sin iniciar", f"{_sin_ini:,}",
                           delta="nadie las ha tomado", delta_color="off")
                _k3.metric("En zona roja", f"{_rojas:,}",
                           delta="<25% umbral o vencidas", delta_color="inverse")
                _k4.metric("Ya vencidas", f"{_venc:,}",
                           delta="excedieron el SLA", delta_color="inverse")
                _k5.metric("🚨 Críticas olvidadas", f"{_n_criticas:,}",
                           delta="rojas + sin iniciar", delta_color="inverse")

                st.divider()

                # ── 🚨 OTs CRÍTICAS OLVIDADAS ──────────────────────────
                # Rojas + sin iniciar = a punto de vencer o vencidas y nadie las tomó
                if _n_criticas > 0:
                    st.markdown(
                        f'<div style="background:rgba(239,68,68,0.10);'
                        f'border-left:5px solid #ef4444;border-radius:8px;'
                        f'padding:12px 16px;margin-bottom:14px;">'
                        f'<div style="font-size:1.05rem;font-weight:800;color:#ef4444;'
                        f'margin-bottom:4px;">🚨 CRÍTICAS OLVIDADAS ({_n_criticas})</div>'
                        f'<div style="font-size:0.85rem;color:{_t["text"]};">'
                        f'OTs a punto de vencer o ya vencidas donde NINGÚN técnico ha '
                        f'iniciado labores. Requieren asignación urgente.</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                    _crit_show = _criticas.sort_values("_restante_min", ascending=True)[[
                        "_semaf","id_ot","_n_aviso","cliente","prioridad_calc","_tipo_eq",
                        "estacion","_t0_lbl","_umbral_lbl","_crono","_req","responsable",
                    ]].rename(columns={
                        "_semaf":"🚦", "id_ot":"OT", "_n_aviso":"N° Aviso",
                        "cliente":"Cliente",
                        "prioridad_calc":"Prio", "_tipo_eq":"Equipo",
                        "estacion":"Estación", "_t0_lbl":"Llamado (T0)",
                        "_umbral_lbl":"Umbral", "_crono":"⏳ Restante",
                        "_req":"Requerimiento", "responsable":"Asignado a",
                    })
                    _show_df(_crit_show, hide_index=True, width="stretch",
                        column_config={
                            "🚦":            st.column_config.TextColumn(width=45),
                            "OT":            st.column_config.TextColumn(width=95),
                            "N° Aviso":      st.column_config.TextColumn(width=100,
                                help="N° Cotalker (Aramco) / N° Aviso del email (COPEC/Shell)."),
                            "Cliente":       st.column_config.TextColumn(width=110),
                            "Prio":          st.column_config.TextColumn(width=55),
                            "Equipo":        st.column_config.TextColumn(width=140),
                            "Estación":      st.column_config.TextColumn(width=180),
                            "Llamado (T0)":  st.column_config.TextColumn(width=125),
                            "Umbral":        st.column_config.TextColumn(width=60),
                            "⏳ Restante":   st.column_config.TextColumn(width=105,
                                help="Tiempo hasta vencer el SLA. Negativo = ya vencido."),
                            "Requerimiento": st.column_config.TextColumn(width=280),
                            "Asignado a":    st.column_config.TextColumn(width=160,
                                help="Aparece pero NADIE ha iniciado la OT en Fracttal."),
                        })
                    st.divider()
                else:
                    st.success(
                        "✅ Sin críticas olvidadas: todas las OTs a punto de vencer "
                        "tienen al menos un técnico trabajando en ellas."
                    )
                    st.divider()

                # ── Secciones por prioridad ────────────────────────────
                def _render_bloque(titulo, df_blk, color_hdr):
                    if df_blk.empty:
                        st.caption(f"— Sin correctivas activas en {titulo} —")
                        return
                    n_sin_ini = int((df_blk["_estado_ot"] == "sin_iniciar").sum())
                    n_rojas = int((df_blk["_semaf"] == "🔴").sum())
                    st.markdown(
                        f'<div style="background:{color_hdr}18;border-left:4px solid {color_hdr};'
                        f'border-radius:6px;padding:8px 14px;margin:8px 0;">'
                        f'<b style="color:{color_hdr};font-size:1.02rem;">{titulo}</b>'
                        f'<span style="color:{_t["muted"]};font-size:0.82rem;margin-left:12px;">'
                        f'{len(df_blk)} OTs · {n_sin_ini} sin iniciar · {n_rojas} en rojo'
                        f'</span></div>',
                        unsafe_allow_html=True,
                    )
                    _blk = df_blk.sort_values("_restante_min", ascending=True)[[
                        "_semaf","id_ot","_n_aviso","_estado_lbl","cliente","prioridad_calc","_tipo_eq",
                        "estacion","_t0_lbl","_umbral_lbl","_pct_restante","_crono",
                        "_req","responsable",
                    ]].rename(columns={
                        "_semaf":"🚦","id_ot":"OT","_n_aviso":"N° Aviso",
                        "_estado_lbl":"Estado",
                        "cliente":"Cliente","prioridad_calc":"Prio","_tipo_eq":"Equipo",
                        "estacion":"Estación","_t0_lbl":"Llamado (T0)","_umbral_lbl":"Umbral",
                        "_pct_restante":"% restante","_crono":"⏳ Restante",
                        "_req":"Requerimiento","responsable":"Asignado a",
                    })
                    _blk["% restante"] = _blk["% restante"].apply(
                        lambda v: max(0, min(round(v, 1), 100)) if pd.notna(v) else 0)
                    _show_df(_blk, hide_index=True, width="stretch",
                        column_config={
                            "🚦":            st.column_config.TextColumn(width=45),
                            "OT":            st.column_config.TextColumn(width=95),
                            "N° Aviso":      st.column_config.TextColumn(width=100,
                                help="N° Cotalker (Aramco) / N° Aviso del email (COPEC/Shell)."),
                            "Estado":        st.column_config.TextColumn(width=110),
                            "Cliente":       st.column_config.TextColumn(width=110),
                            "Prio":          st.column_config.TextColumn(width=55),
                            "Equipo":        st.column_config.TextColumn(width=140),
                            "Estación":      st.column_config.TextColumn(width=170),
                            "Llamado (T0)":  st.column_config.TextColumn(width=125),
                            "Umbral":        st.column_config.TextColumn(width=60),
                            "% restante":    st.column_config.ProgressColumn(
                                min_value=0, max_value=100, format="%.0f%%",
                                help="% del umbral SLA aún disponible. 0% = vencido."),
                            "⏳ Restante":   st.column_config.TextColumn(width=100,
                                help="Tiempo hasta vencer. Negativo = ya vencido."),
                            "Requerimiento": st.column_config.TextColumn(width=260),
                            "Asignado a":    st.column_config.TextColumn(width=150),
                        })

                # P1
                _p1 = _df_ab[_df_ab["prioridad_calc"].str.upper() == "P1"]
                _render_bloque(f"🔴 P1 · Máquina detenida (más críticas)",
                               _p1, "#ef4444")
                # P2
                _p2 = _df_ab[_df_ab["prioridad_calc"].str.upper() == "P2"]
                _render_bloque(f"🟡 P2 · Falla operativa importante",
                               _p2, "#f59e0b")
                # P3 + P4
                _p34 = _df_ab[_df_ab["prioridad_calc"].str.upper().isin(["P3","P4"])]
                _render_bloque(f"🟢 P3 + P4 · Prioridad menor",
                               _p34, "#22c55e")

                st.divider()

            # ══════════════════════════════════════════════════════════════════
            # HISTÓRICO DEL MES + RANKING (fuera del if _df_ab.empty)
            # Estas secciones SIEMPRE aparecen para dar contexto operativo,
            # aunque no haya correctivas abiertas en el filtro actual.
            # Nivel de indentación: 12 espacios (dentro de with _tab_en_curso).
            # ══════════════════════════════════════════════════════════════════
            if sel_mes_live != "Todos":
                st.markdown(
                    f'<div class="section-header">'
                    f'📚  Histórico correctivas — {sel_mes_live}'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                _mes_num_h = {v: k for k, v in _MES_ES_LIVE.items()}.get(
                    sel_mes_live.split(" ")[0])
                if _mes_num_h and not df_llamados.empty:
                    _hist = df_llamados.copy()
                    _fl_h = pd.to_datetime(_hist["fecha_llamado"],
                                           errors="coerce", utc=True)
                    _hist = _hist[
                        (_fl_h.dt.year == _year) &
                        (_fl_h.dt.month == _mes_num_h)
                    ]
                    if sel_cli_live != "Todos":
                        _hist = _hist[
                            _hist["cliente"].apply(_norm_cli_live) == _sel_norm
                        ]
                    _hist = _hist[_hist["fecha_atencion"].notna()]

                    if _hist.empty:
                        st.info(
                            f"No hay correctivas cerradas registradas para "
                            f"**{sel_mes_live}** con el filtro actual."
                        )
                    else:
                        _hist["_tiene_exc"] = (
                            _hist["excepcion_motivo"].notna()
                            & (_hist["excepcion_motivo"].astype(str).str.strip() != "")
                        ) if "excepcion_motivo" in _hist.columns else False

                        _cumple_n   = int((_hist["cumplimiento"] == "CUMPLE").sum())
                        _nocumple_n = int((_hist["cumplimiento"] == "NO CUMPLE").sum())
                        _exc_n      = int(_hist["_tiene_exc"].sum())
                        _tot_h      = _cumple_n + _nocumple_n
                        _pct_ok     = round(_cumple_n / _tot_h * 100, 1) if _tot_h else 0

                        _hk1,_hk2,_hk3,_hk4,_hk5 = st.columns(5)
                        _hk1.metric("Total atendidas", f"{_tot_h:,}")
                        _hk2.metric("✅ Cumplen SLA", f"{_cumple_n:,}",
                                    delta=f"{_pct_ok}%")
                        _hk3.metric("❌ No cumplen SLA", f"{_nocumple_n:,}",
                                    delta=f"{100-_pct_ok:.1f}%",
                                    delta_color="inverse")
                        _hk4.metric("ℹ️ Con excepción", f"{_exc_n:,}",
                                    delta="ya cuentan como cumple",
                                    delta_color="off")
                        _hk5.metric("Cumplimiento", f"{_pct_ok}%")

                        # Tabla resumen de OTs cerradas del mes
                        _hist_disp = _hist.copy()
                        def _estado_h(row):
                            cumpl = str(row.get("cumplimiento","")).upper()
                            if bool(row.get("_tiene_exc")): return "ℹ️ Excepción"
                            if cumpl == "CUMPLE":    return "✅ Cumple"
                            if cumpl == "NO CUMPLE": return "❌ No cumple"
                            return "⚪ Sin datos"
                        _hist_disp["_estado"] = _hist_disp.apply(_estado_h, axis=1)
                        _hist_disp["_fecha"] = pd.to_datetime(
                            _hist_disp["fecha_llamado"], errors="coerce"
                        ).dt.strftime("%d/%m/%Y")
                        _hist_disp["_atencion"] = pd.to_datetime(
                            _hist_disp["fecha_atencion"], errors="coerce"
                        ).dt.strftime("%d/%m/%Y")

                        _hist_cols = [c for c in [
                            "os_fracttal","_fecha","_atencion","cliente","eds_nombre",
                            "tecnico","prioridad","horas_resolucion","tiempo_resp_esp",
                            "_estado",
                        ] if c in _hist_disp.columns]
                        _hist_show = _hist_disp[_hist_cols].rename(columns={
                            "os_fracttal":"OS Fracttal",
                            "_fecha":"Fecha llamado",
                            "_atencion":"Fecha atención",
                            "cliente":"Cliente","eds_nombre":"EDS",
                            "tecnico":"Técnico","prioridad":"Prio",
                            "horas_resolucion":"Horas",
                            "tiempo_resp_esp":"Umbral (h)",
                            "_estado":"Estado SLA",
                        })
                        # Orden: NO CUMPLE primero (para ver problemas de un vistazo)
                        _orden_estado = {"❌ No cumple":0, "ℹ️ Excepción":1,
                                         "✅ Cumple":2, "⚪ Sin datos":3}
                        _hist_show = _hist_show.assign(
                            _ord=_hist_show["Estado SLA"].map(_orden_estado).fillna(9)
                        ).sort_values(["_ord","Fecha llamado"], ascending=[True,False]) \
                         .drop(columns="_ord")

                        _show_df(_hist_show.reset_index(drop=True),
                            hide_index=True, width="stretch",
                            column_config={
                                "OS Fracttal":    st.column_config.TextColumn(width=105),
                                "Fecha llamado":  st.column_config.TextColumn(width=105),
                                "Fecha atención": st.column_config.TextColumn(width=110),
                                "Cliente":        st.column_config.TextColumn(width=110),
                                "EDS":            st.column_config.TextColumn(width=220),
                                "Técnico":        st.column_config.TextColumn(width=170),
                                "Prio":           st.column_config.TextColumn(width=55),
                                "Horas":          st.column_config.NumberColumn(
                                    format="%.1f h", width=80),
                                "Umbral (h)":     st.column_config.NumberColumn(
                                    format="%d h", width=80),
                                "Estado SLA":     st.column_config.TextColumn(width=130,
                                    help="Ordenadas: No cumple → Excepción → Cumple."),
                            })
                st.divider()

            # ══════════════════════════════════════════════════════════════════
            # RANKING: Técnicos que MÁS dejan vencer SLA
            # ══════════════════════════════════════════════════════════════════
            st.markdown(
                '<div class="section-header">'
                '📊  Ranking · Técnicos con más SLA vencidos'
                '</div>',
                unsafe_allow_html=True,
            )
            st.caption(
                "Consolidado histórico 2026 de OTs vencidas por técnico. "
                "Suma **OTs cerradas fuera de SLA** + **OTs actualmente vencidas y aún abiertas**. "
                "Incluye Occimiano, AUTEC, Elecons y externos."
            )

            # (1) OTs cerradas que NO cumplieron
            _venc_cerradas = pd.DataFrame()
            if not df_llamados.empty and "cumplimiento" in df_llamados.columns:
                _venc_cerradas = df_llamados[
                    (df_llamados["cumplimiento"] == "NO CUMPLE") &
                    (df_llamados["tecnico"].notna())
                ].copy()
                if sel_cli_live != "Todos":
                    _venc_cerradas = _venc_cerradas[
                        _venc_cerradas["cliente"].apply(_norm_cli_live) == _sel_norm
                    ]
                if sel_mes_live != "Todos":
                    _mes_num_v = {v: k for k, v in _MES_ES_LIVE.items()}.get(
                        sel_mes_live.split(" ")[0])
                    if _mes_num_v:
                        _fl_v = pd.to_datetime(_venc_cerradas["fecha_llamado"],
                                               errors="coerce", utc=True)
                        _venc_cerradas = _venc_cerradas[
                            (_fl_v.dt.year == _year) &
                            (_fl_v.dt.month == _mes_num_v)
                        ]

            # (2) OTs abiertas ya vencidas — protegido si _df_ab está vacío
            _venc_abiertas = pd.DataFrame()
            if not _df_ab.empty and "_vencida" in _df_ab.columns:
                _venc_abiertas = _df_ab[_df_ab["_vencida"]].copy()

            def _norm_tec(t):
                s = str(t or "").strip()
                if not s: return ""
                su = s.upper()
                if "AUTEC" in su: return "AUTEC"
                if "OCAMPO" in su: return "Jaime Ocampo (Elecons)"
                return s

            # Excluir técnicos "no aplica" (personas que no son técnicos
            # operativos, dueños, administrativos o similares).
            # Se preserva 'Jaime Ocampo' porque lo tratamos como Elecons
            # (contratista externo válido).
            def _es_no_aplica_ranking(tec_norm):
                s = str(tec_norm or "").upper()
                for excluido in ("WALTER SOTO", "ANA GUZMAN", "JUAN VALLE",
                                 "WALTER MAURICIO", "ANA MARIA GUZMAN",
                                 "JUAN PABLO VALLE", "ERIC DAYLLER",
                                 "ERIC ESTEBAN", "DAYLLER MESA"):
                    if excluido in s: return True
                return False

            _rk_data = []
            if not _venc_cerradas.empty:
                _venc_cerradas = _venc_cerradas.assign(
                    _tec_norm=_venc_cerradas["tecnico"].apply(_norm_tec))
                _rk_data.append(_venc_cerradas.groupby("_tec_norm").agg(
                    cerradas=("os_fracttal", "count"),
                ).reset_index().rename(columns={"_tec_norm": "_tec"}))
            if not _venc_abiertas.empty:
                _venc_abiertas = _venc_abiertas.assign(
                    _tec_norm=_venc_abiertas["responsable"].apply(_norm_tec))
                _rk_data.append(_venc_abiertas.groupby("_tec_norm").agg(
                    abiertas=("id_ot", "count"),
                ).reset_index().rename(columns={"_tec_norm": "_tec"}))

            if not _rk_data:
                st.success("✅ No hay OTs vencidas en el filtro seleccionado.")
            else:
                _rk = _rk_data[0]
                for _extra in _rk_data[1:]:
                    _rk = _rk.merge(_extra, on="_tec", how="outer")
                for _c in ("cerradas", "abiertas"):
                    if _c not in _rk.columns:
                        _rk[_c] = 0
                _rk["cerradas"] = _rk["cerradas"].fillna(0).astype(int)
                _rk["abiertas"] = _rk["abiertas"].fillna(0).astype(int)
                _rk["total"]    = _rk["cerradas"] + _rk["abiertas"]
                _rk = _rk[_rk["_tec"].astype(str).str.strip() != ""]
                _rk = _rk[~_rk["_tec"].apply(_es_no_aplica_ranking)]
                _rk = _rk.sort_values("total", ascending=False)

                def _empresa(tec):
                    s = str(tec or "").upper()
                    if "AUTEC" in s: return "🔵 AUTEC"
                    if "OCAMPO" in s or "ELECONS" in s: return "🟣 Elecons"
                    return "🟢 Occimiano"
                _rk["Empresa"] = _rk["_tec"].apply(_empresa)

                _rk_show = _rk.head(20).rename(columns={
                    "_tec":     "Técnico",
                    "cerradas": "Vencidas cerradas",
                    "abiertas": "Vencidas abiertas ahora",
                })[["Técnico","Empresa","Vencidas cerradas","Vencidas abiertas ahora"]]

                _show_df(_rk_show.reset_index(drop=True), hide_index=True,
                    width="stretch",
                    column_config={
                        "Técnico":                st.column_config.TextColumn(width=240),
                        "Empresa":                st.column_config.TextColumn(width=120),
                        "Vencidas cerradas":      st.column_config.NumberColumn(
                            format="%d", width=160,
                            help="OTs ya cerradas que NO cumplieron SLA."),
                        "Vencidas abiertas ahora":st.column_config.NumberColumn(
                            format="%d", width=200,
                            help="OTs actualmente abiertas y ya vencidas (del filtro actual)."),
                    })
                _n_show = min(20, len(_rk))
                st.caption(f"Top {_n_show} de {len(_rk)} responsables con al menos 1 SLA vencido.")

                # ── Gráfico de barras: TOP responsables con más vencimientos ──
                import plotly.graph_objects as _pgo_v
                _bar_df = _rk.head(10).copy()
                if not _bar_df.empty:
                    _color_map = {"🔵 AUTEC": "#3b82f6",
                                  "🟣 Elecons": "#a855f7",
                                  "🟢 Occimiano": "#22c55e"}
                    _colors = [_color_map.get(_empresa(t), "#94a3b8")
                               for t in _bar_df["_tec"]]
                    _fig_v = _pgo_v.Figure(_pgo_v.Bar(
                        x=_bar_df["total"], y=_bar_df["_tec"],
                        orientation="h",
                        marker_color=_colors,
                        text=_bar_df["total"], textposition="outside",
                        hovertemplate="<b>%{y}</b><br>SLA vencidos: %{x}<extra></extra>",
                    ))
                    _ctx_titulo = sel_cli_live if sel_cli_live != "Todos" else "Todos los clientes"
                    _ctx_mes    = sel_mes_live if sel_mes_live != "Todos" else "todo 2026"
                    _fig_v.update_layout(
                        title=dict(
                            text=f"<b>Top 10 · Mayores responsables de vencimientos</b>"
                                 f"<br><span style='font-size:0.72rem;color:#94a3b8;font-weight:400'>"
                                 f"{_ctx_titulo} · {_ctx_mes}</span>",
                            font=dict(size=13)),
                        height=max(220, 32 * len(_bar_df) + 90),
                        margin=dict(l=10, r=40, t=70, b=10),
                        yaxis=dict(autorange="reversed", title=""),
                        xaxis=dict(title="OTs con SLA vencido", tickformat="d"),
                        paper_bgcolor="rgba(0,0,0,0)",
                        plot_bgcolor="rgba(0,0,0,0)",
                        showlegend=False,
                    )
                    st.plotly_chart(_fig_v, use_container_width=True,
                                    key=f"bar_venc_{sel_cli_live}_{sel_mes_live}")

            st.divider()

            # ══════════════════════════════════════════════════════════════════
            # RANKING: Técnicos MÁS EFICIENTES (OTs CUMPLIDAS en tiempo)
            # ══════════════════════════════════════════════════════════════════
            st.markdown(
                '<div class="section-header">'
                '⭐  Ranking · Técnicos más eficientes (SLA cumplidos)'
                '</div>',
                unsafe_allow_html=True,
            )
            st.caption(
                "OTs correctivas **CERRADAS EN TIEMPO** por técnico. "
                "Complementa el ranking de vencimientos — permite ver quién sostiene "
                "el desempeño operativo. Respeta filtros de Cliente y Mes."
            )

            _cumpl_hist = pd.DataFrame()
            if not df_llamados.empty and "cumplimiento" in df_llamados.columns:
                _cumpl_hist = df_llamados[
                    (df_llamados["cumplimiento"] == "CUMPLE") &
                    (df_llamados["tecnico"].notna())
                ].copy()
                if sel_cli_live != "Todos":
                    _cumpl_hist = _cumpl_hist[
                        _cumpl_hist["cliente"].apply(_norm_cli_live) == _sel_norm
                    ]
                if sel_mes_live != "Todos":
                    _mes_num_e = {v: k for k, v in _MES_ES_LIVE.items()}.get(
                        sel_mes_live.split(" ")[0])
                    if _mes_num_e:
                        _fl_e = pd.to_datetime(_cumpl_hist["fecha_llamado"],
                                               errors="coerce", utc=True)
                        _cumpl_hist = _cumpl_hist[
                            (_fl_e.dt.year == _year) &
                            (_fl_e.dt.month == _mes_num_e)
                        ]

            if _cumpl_hist.empty:
                st.info("No hay OTs cumplidas en el filtro seleccionado.")
            else:
                _cumpl_hist = _cumpl_hist.assign(
                    _tec_norm=_cumpl_hist["tecnico"].apply(_norm_tec))
                _ef = _cumpl_hist.groupby("_tec_norm").agg(
                    cumplen=("os_fracttal", "count"),
                ).reset_index().rename(columns={"_tec_norm": "_tec"})
                _ef = _ef[_ef["_tec"].astype(str).str.strip() != ""]
                _ef = _ef[~_ef["_tec"].apply(_es_no_aplica_ranking)]

                # Enriquecer con vencidas cerradas del mismo scope
                # para poder mostrar total atendidas y % efectividad
                _venc_scope = pd.DataFrame()
                if not df_llamados.empty:
                    _venc_scope = df_llamados[
                        (df_llamados["cumplimiento"] == "NO CUMPLE") &
                        (df_llamados["tecnico"].notna())
                    ].copy()
                    if sel_cli_live != "Todos":
                        _venc_scope = _venc_scope[
                            _venc_scope["cliente"].apply(_norm_cli_live) == _sel_norm
                        ]
                    if sel_mes_live != "Todos" and _mes_num_e:
                        _fl_s = pd.to_datetime(_venc_scope["fecha_llamado"],
                                               errors="coerce", utc=True)
                        _venc_scope = _venc_scope[
                            (_fl_s.dt.year == _year) &
                            (_fl_s.dt.month == _mes_num_e)
                        ]
                    _venc_scope = _venc_scope.assign(
                        _tec_norm=_venc_scope["tecnico"].apply(_norm_tec))
                    _venc_ct = _venc_scope.groupby("_tec_norm").agg(
                        vencen=("os_fracttal", "count"),
                    ).reset_index().rename(columns={"_tec_norm": "_tec"})
                    _ef = _ef.merge(_venc_ct, on="_tec", how="left")
                if "vencen" not in _ef.columns:
                    _ef["vencen"] = 0
                _ef["vencen"] = _ef["vencen"].fillna(0).astype(int)
                _ef["total"]  = _ef["cumplen"] + _ef["vencen"]
                _ef["_pct"]   = (_ef["cumplen"] / _ef["total"].clip(lower=1) * 100).round(1)
                _ef = _ef.sort_values(["cumplen","_pct"], ascending=[False, False])

                _ef["Empresa"] = _ef["_tec"].apply(_empresa)
                _ef_show = _ef.head(20).rename(columns={
                    "_tec":    "Técnico",
                    "cumplen": "SLA cumplidos",
                    "vencen":  "SLA vencidos",
                    "total":   "Total atendidos",
                    "_pct":    "% Efectividad",
                })[["Técnico","Empresa","SLA cumplidos","SLA vencidos",
                    "Total atendidos","% Efectividad"]]

                _show_df(_ef_show.reset_index(drop=True), hide_index=True,
                    width="stretch",
                    column_config={
                        "Técnico":         st.column_config.TextColumn(width=240),
                        "Empresa":         st.column_config.TextColumn(width=120),
                        "SLA cumplidos":   st.column_config.NumberColumn(format="%d", width=130,
                            help="OTs cerradas dentro del umbral SLA (incluye excepciones)."),
                        "SLA vencidos":    st.column_config.NumberColumn(format="%d", width=120),
                        "Total atendidos": st.column_config.NumberColumn(format="%d", width=130),
                        "% Efectividad":   st.column_config.ProgressColumn(
                            format="%.1f%%", min_value=0, max_value=100, width=160,
                            help="Cumplidos / Total. 100% = ningún vencimiento."),
                    })
                _n_ef = min(20, len(_ef))
                st.caption(f"Top {_n_ef} de {len(_ef)} técnicos con al menos 1 SLA cumplido, ordenado por volumen.")

            st.divider()
            st.caption(
                "**Fuentes**: `ordenes_trabajo` con `tipo_tarea=CORRECTIVA` y sin "
                "`fecha_finalizacion` · Umbral SLA desde `sla_umbrales_horas` por "
                "(cliente, prioridad, zona) · T0 = `fecha_incidente` (o `fecha_creacion` "
                "como fallback) · Estado 'en curso' = existe `fecha_inicio` (técnico "
                "abrió Fracttal para trabajar). Se excluyen Canceladas."
            )

        # ══════════════════════════════════════════════════════════════════════
        # SUB-PESTAÑA: UPTIME (llamados de emergencia / correctivas)
        # Tiempo detenido = (fecha_atencion - fecha_llamado) por cada llamado.
        # Granularidad: EDS (df_llamados no tiene codigo_activo individual).
        # ══════════════════════════════════════════════════════════════════════
        with _tab_uptime_sla:
            st.markdown(
                '<div class="section-header">⏱️  Uptime General (correctivas + preventivas)</div>',
                unsafe_allow_html=True,
            )

            _df_up = df_llamados.copy()
            _df_up = _df_up[_df_up["fecha_llamado"].notna() & _df_up["fecha_atencion"].notna()]

            if _df_up.empty:
                st.info("Sin llamados atendidos en el período.")
            else:
                # ── Filtros de fecha (Trimestre + Mes) ─────────────────
                _fl_up_norm = _df_up["fecha_llamado"]
                if _fl_up_norm.dt.tz is not None:
                    _fl_up_norm = _fl_up_norm.dt.tz_convert(None)
                _df_up["_mes_p"]   = _fl_up_norm.dt.to_period("M").astype(str)
                _df_up["_month_n"] = _fl_up_norm.dt.month.astype("Int64")

                _ff1, _ff2, _ff3 = st.columns([1.2, 1.8, 1.3])
                with _ff1:
                    _up_trim = st.selectbox("Período", _trim_opts, key="upsla_trim")
                with _ff2:
                    if _up_trim != "Todos":
                        _trim_m_up = _TRIMESTRES_DEF[_up_trim]
                        _meses_up_disp = [l for l in _meses_disp_lbl[1:]
                                          if _MES_ABR_NUM_LL.get(l.split(" ")[0], 0) in _trim_m_up]
                    else:
                        _meses_up_disp = _meses_disp_lbl[1:]
                    _up_meses = st.multiselect("Mes", _meses_up_disp, key="upsla_mes",
                                               placeholder="Todos los meses")
                with _ff3:
                    _up_prio = st.selectbox(
                        "Prioridad a contar",
                        ["Solo P1 (máquina detenida)", "P1 + P2", "Todas"],
                        key="upsla_prio",
                        help="Solo las P1 implican máquina detenida en el negocio. "
                             "Las P2/P3/P4 son fallas donde la máquina sigue operativa."
                    )

                if _up_trim != "Todos":
                    _df_up = _df_up[_df_up["_month_n"].isin(_TRIMESTRES_DEF[_up_trim])]
                if _up_meses:
                    _meses_periodo = [_lbl_to_period.get(l) for l in _up_meses if _lbl_to_period.get(l)]
                    _df_up = _df_up[_df_up["_mes_p"].isin(_meses_periodo)]
                # Filtro de prioridad — para uptime real solo cuentan llamados con máquina detenida
                _prio_norm = _df_up["prioridad"].astype(str).str.upper().str.strip()
                if _up_prio == "Solo P1 (máquina detenida)":
                    _df_up = _df_up[_prio_norm == "P1"]
                elif _up_prio == "P1 + P2":
                    _df_up = _df_up[_prio_norm.isin(["P1", "P2"])]

                # ── Período evaluado dinámico según filtros de fecha ───
                _hoy_up = pd.Timestamp.today().normalize()
                if _up_meses:
                    _meses_efectivos = [_lbl_to_period.get(l) for l in _up_meses if _lbl_to_period.get(l)]
                elif _up_trim != "Todos":
                    _trim_m_up = _TRIMESTRES_DEF[_up_trim]
                    _meses_efectivos = [m for m in sorted(_lbl_to_period.values())
                                        if int(m.split("-")[1]) in _trim_m_up]
                else:
                    _meses_efectivos = sorted(_lbl_to_period.values())

                if _meses_efectivos:
                    _meses_ord = sorted(_meses_efectivos)
                    _rango_inicio_up = pd.Timestamp(_meses_ord[0] + "-01")
                    _last = pd.Timestamp(_meses_ord[-1] + "-01") + pd.offsets.MonthEnd(1)
                    _rango_fin_up = min(_last.normalize(), _hoy_up)
                else:
                    _rango_inicio_up = pd.Timestamp("2026-01-01")
                    _rango_fin_up = _hoy_up

                # Suma de días sobre los meses seleccionados (no es contiguo si saltas meses)
                _total_dias = 0
                for _ym in _meses_efectivos:
                    _y, _m = map(int, _ym.split("-"))
                    _first = pd.Timestamp(_y, _m, 1)
                    if _first > _hoy_up:
                        continue
                    _last_m = (_first + pd.offsets.MonthEnd(1)).normalize()
                    _last_m = min(_last_m, _hoy_up)
                    _total_dias += (_last_m - _first).days + 1
                _rango_dias_up = max(_total_dias, 1)
                _seg_equipo_up = _rango_dias_up * 24 * 3600

                # ── Cruce con OTs correctivas para traer equipo + cierre Fracttal ──
                # df_llamados solo tiene EDS y fecha_atencion (cierre SLA en terreno).
                # ordenes_trabajo tiene codigo_activo, nombre_activo y fecha_finalizacion
                # (cierre administrativo en Fracttal, puede demorar días/semanas).
                # IMPORTANTE: el merge resetea índices, por eso lo hacemos ANTES de
                # calcular fechas (si no, las Series no se alinean en la resta).
                with st.spinner("Cargando datos de OTs correctivas (equipos)…"):
                    _raw_corr_up = load_correctivas_supabase()
                if _raw_corr_up:
                    _df_corr_up = pd.DataFrame(_raw_corr_up)[
                        ["id_ot","codigo_activo","nombre_activo",
                         "fecha_finalizacion","duracion_real_seg"]
                    ].copy()
                    _df_corr_up["duracion_real_seg"] = pd.to_numeric(
                        _df_corr_up["duracion_real_seg"], errors="coerce").fillna(0)
                    _df_corr_up["fecha_finalizacion"] = pd.to_datetime(
                        _df_corr_up["fecha_finalizacion"], errors="coerce", utc=True)
                    # Una OT puede tener varios equipos concatenados — los unimos como vinieron
                    _df_corr_up = _df_corr_up.groupby("id_ot", as_index=False).agg(
                        codigo_activo=("codigo_activo",
                            lambda s: ", ".join(sorted({str(x).strip() for x in s if pd.notna(x)}))),
                        nombre_activo=("nombre_activo",
                            lambda s: " · ".join(sorted({str(x).strip() for x in s if pd.notna(x)}))),
                        fecha_finalizacion=("fecha_finalizacion", "max"),
                        duracion_real_seg=("duracion_real_seg", "sum"),
                    )
                    _df_up = _df_up.merge(
                        _df_corr_up, how="left",
                        left_on="os_fracttal", right_on="id_ot"
                    ).reset_index(drop=True)
                    _df_up["duracion_real_seg"] = _df_up["duracion_real_seg"].fillna(0)
                else:
                    _df_up = _df_up.reset_index(drop=True)
                    _df_up["codigo_activo"] = ""
                    _df_up["nombre_activo"] = ""
                    _df_up["fecha_finalizacion"] = pd.NaT
                    _df_up["duracion_real_seg"] = 0

                # Tiempo detenido por llamado (en segundos), calculado DESPUÉS del merge
                def _merge_dt_up(d, h):
                    d_ts = pd.to_datetime(d, errors="coerce")
                    h_str = h.astype(str).fillna("00:00:00").str[:8].replace(
                        {"nan": "00:00:00", "NaT": "00:00:00", "":"00:00:00"})
                    return pd.to_datetime(
                        d_ts.dt.strftime("%Y-%m-%d") + " " + h_str,
                        errors="coerce")

                _dt_ll_up = _merge_dt_up(_df_up["fecha_llamado"],
                                         _df_up.get("hora_llamado", pd.Series("", index=_df_up.index)))
                _dt_at_up = _merge_dt_up(_df_up["fecha_atencion"],
                                         _df_up.get("hora_fin", pd.Series("", index=_df_up.index)))

                # Cierre real = MIN(fecha_atencion, fecha_finalizacion)
                #   - fecha_atencion (df_llamados): cierre técnico en terreno (cuando
                #     el técnico cerró el SLA con el cliente)
                #   - fecha_finalizacion (ordenes_trabajo): cierre administrativo
                #     Fracttal — puede demorar días/semanas
                _fin_tec = pd.to_datetime(
                    _df_up["fecha_finalizacion"], errors="coerce", utc=True)
                if hasattr(_fin_tec, "dt") and _fin_tec.dt.tz is not None:
                    _fin_tec = _fin_tec.dt.tz_convert(None)
                _cierre = pd.DataFrame({"a": _dt_at_up, "b": _fin_tec}).min(axis=1)
                _df_up["_paro_seg"] = (_cierre - _dt_ll_up).dt.total_seconds().clip(lower=0).fillna(0)

                st.markdown(
                    f"""<div style="background:rgba(1,121,138,0.10);border-left:3px solid #01798A;
                         padding:10px 16px;border-radius:6px;margin-bottom:14px;">
                      <span style="color:#01798A;font-weight:700;font-size:0.95rem;">📅 Período evaluado:</span>
                      <span style="color:var(--text-color, #475569);font-size:0.9rem;margin-left:8px;">
                        <b>{_rango_inicio_up.strftime('%d/%m/%Y')}</b> → <b>{_rango_fin_up.strftime('%d/%m/%Y')}</b>
                        ({_rango_dias_up} días · {_rango_dias_up * 24:,} horas por EDS).</span>
                    </div>""",
                    unsafe_allow_html=True,
                )
                st.caption(
                    "**Fórmula Uptime General**: 1 − (Σ horas detenidas por CORRECTIVAS + Σ horas detenidas por PREVENTIVAS) ÷ (N EDS × horas del período) · "
                    "**Correctivas**: paro = desde `fecha_llamado` hasta el primer cierre real (mínimo entre `fecha_atención` y `fecha_finalización`). "
                    "Solo cuentan según el filtro de prioridad (default: P1 = máquina detenida). "
                    "**Preventivas**: paro = `tiempo_paro_real_seg` de OTs con `¿Paro de equipo? = SÍ` y `fecha_finalizacion` dentro del período. "
                    "El % refleja el tiempo REAL en que la flota estuvo operativa considerando AMBOS mundos."
                )

                # ── Clasificación regional por eds_nombre / comuna ────
                def _region_ll(s):
                    s = str(s or "").upper()
                    _NORTE = ["IQUIQUE","ARICA","ANTOFAGASTA","CALAMA","COPIAPÓ","COPIAPO",
                              "OVALLE","LA SERENA","COQUIMBO","VALLENAR","ILLAPEL","HUASCO",
                              "ALTO HOSPICIO","TOCOPILLA","MEJILLONES"]
                    _SUR   = ["CONCEPCIÓN","CONCEPCION","OSORNO","TEMUCO","VALDIVIA",
                              "PUERTO MONTT","CHILLÁN","CHILLAN","LOS ANGELES","LOS ÁNGELES",
                              "LINARES","TALCA","CURICÓ","CURICO","RANCAGUA","FRUTILLAR",
                              "VICTORIA","RENGO","SAN FERNANDO","PUERTO VARAS","ANCUD",
                              "CASTRO","COIHAIQUE","LA UNION","LA UNIÓN"]
                    if any(c in s for c in _NORTE): return "Norte"
                    if any(c in s for c in _SUR):   return "Sur"
                    return "Santiago"

                _src_reg = _df_up.get("eds_nombre",
                    pd.Series("", index=_df_up.index)).fillna("").astype(str)
                if "comuna" in _df_up.columns:
                    _src_reg = _src_reg + " " + _df_up["comuna"].fillna("").astype(str)
                _df_up["_region"] = _src_reg.apply(_region_ll)

                # ── Filtros: Cliente + Región + EDS ────────────────────
                _fu1s, _fu2s, _fu3s = st.columns(3)
                with _fu1s:
                    _CLI_PREF = ["COPEC", "Aramco (Esmax)", "SHELL (Enex)"]
                    _cli_present = set(_df_up["cliente"].dropna().unique())
                    _cli_up_opts = ["Todos"] + [c for c in _CLI_PREF if c in _cli_present] + \
                        sorted([c for c in _cli_present if c not in _CLI_PREF])
                    _up_sla_cli = st.selectbox("Cliente", _cli_up_opts, key="upsla_cli")
                with _fu2s:
                    _up_sla_reg = st.selectbox("Región",
                        ["Todas", "Santiago", "Norte", "Sur"], key="upsla_reg")
                with _fu3s:
                    _eds_pool = sorted(
                        _df_up["eds_nombre"].dropna().unique().tolist()
                    ) if "eds_nombre" in _df_up.columns else []
                    _up_sla_eds = st.selectbox("Estación / EDS",
                        ["Todas"] + _eds_pool, key="upsla_eds")

                if _up_sla_cli != "Todos":
                    _df_up = _df_up[_df_up["cliente"] == _up_sla_cli]
                if _up_sla_reg != "Todas":
                    _df_up = _df_up[_df_up["_region"] == _up_sla_reg]
                if _up_sla_eds != "Todas":
                    _df_up = _df_up[_df_up["eds_nombre"] == _up_sla_eds]

                # ── Cálculo principal: paro por CORRECTIVAS ────────────
                _paro_corr_seg = _df_up["_paro_seg"].sum()
                _eds_unicas = _df_up["eds_occim"].dropna().astype(str).str.strip()
                _eds_unicas = set(_eds_unicas[_eds_unicas != ""].unique())
                _n_eds_up = max(len(_eds_unicas), 1)
                _total_seg_up = _seg_equipo_up * _n_eds_up

                # ── Cálculo del paro por PREVENTIVAS (UPTIME GENERAL) ──
                # El uptime real de una máquina incluye TAMBIÉN el tiempo
                # detenido por mantenciones planificadas (MP). Aquí sumamos
                # las MP que:
                #   - tienen paro_equipo = True
                #   - fueron finalizadas dentro del período evaluado
                #   - cumplen los filtros de cliente / región / EDS
                # Usamos tiempo_paro_real_seg (con fallback a estim).
                _paro_prev_seg = 0.0
                _prev_ots_con_paro = 0
                _df_prev_up = pd.DataFrame()  # placeholder para ranking/tabla
                try:
                    with st.spinner("Cargando preventivas para uptime general…"):
                        _raw_prev_up = load_preventivas_supabase()
                    if _raw_prev_up:
                        _df_prev_up = pd.DataFrame(_raw_prev_up)
                        # Normalizar fechas y flags
                        _df_prev_up["fecha_finalizacion"] = pd.to_datetime(
                            _df_prev_up.get("fecha_finalizacion"),
                            errors="coerce", utc=True)
                        if _df_prev_up["fecha_finalizacion"].dt.tz is not None:
                            _df_prev_up["fecha_finalizacion"] = \
                                _df_prev_up["fecha_finalizacion"].dt.tz_convert(None)
                        _df_prev_up["paro_equipo"] = _df_prev_up.get("paro_equipo", False)
                        _df_prev_up["tiempo_paro_real_seg"] = pd.to_numeric(
                            _df_prev_up.get("tiempo_paro_real_seg"), errors="coerce").fillna(0)
                        _df_prev_up["tiempo_paro_estim_seg"] = pd.to_numeric(
                            _df_prev_up.get("tiempo_paro_estim_seg"), errors="coerce").fillna(0)

                        # Filtro 1: dentro del período evaluado (por fecha_finalizacion)
                        _mask_periodo = (
                            (_df_prev_up["fecha_finalizacion"] >= _rango_inicio_up) &
                            (_df_prev_up["fecha_finalizacion"] <= _rango_fin_up + pd.Timedelta(days=1))
                        )
                        _df_prev_up = _df_prev_up[_mask_periodo & (_df_prev_up["paro_equipo"] == True)].copy()

                        # Filtro 2: cliente (normalizando ESMAX vs "ESMAX (Aramco)" vs "Aramco (Esmax)")
                        def _norm_cli(s):
                            s = str(s or "").upper()
                            if "COPEC" in s: return "COPEC"
                            if "ESMAX" in s or "ARAMCO" in s: return "ARAMCO"
                            if "SHELL" in s or "ENEX" in s: return "SHELL"
                            return s
                        if _up_sla_cli != "Todos":
                            _df_prev_up = _df_prev_up[
                                _df_prev_up["cliente"].apply(_norm_cli) == _norm_cli(_up_sla_cli)
                            ]

                        # Filtro 3: EDS (por codigo_eds que es el estándar)
                        if _up_sla_eds != "Todas" and not _df_prev_up.empty:
                            # Buscar codigo_eds correspondiente al eds_nombre elegido en correctivas
                            _cod_eds_sel = (
                                df_llamados[df_llamados["eds_nombre"] == _up_sla_eds]
                                ["eds_occim"].dropna().unique().tolist()
                            )
                            if _cod_eds_sel:
                                _df_prev_up = _df_prev_up[
                                    _df_prev_up["codigo_eds"].isin(_cod_eds_sel)
                                ]

                        # Filtro 4: región (por estacion/ubicacion, misma lógica del uptime MP)
                        if _up_sla_reg != "Todas" and not _df_prev_up.empty:
                            _src_reg_prev = (
                                _df_prev_up["estacion"].fillna(_df_prev_up.get("ubicacion",""))
                                .fillna("").astype(str)
                            )
                            _df_prev_up = _df_prev_up[
                                _src_reg_prev.apply(_region_ll) == _up_sla_reg
                            ]

                        # Suma de paro: real con fallback a estim
                        if not _df_prev_up.empty:
                            _pr = _df_prev_up["tiempo_paro_real_seg"]
                            _pe = _df_prev_up["tiempo_paro_estim_seg"]
                            _df_prev_up["_paro_calc_seg"] = _pr.where(_pr > 0, _pe).clip(lower=0)
                            _paro_prev_seg = float(_df_prev_up["_paro_calc_seg"].sum())
                            _prev_ots_con_paro = int(len(_df_prev_up))
                except Exception as _e:
                    st.warning(f"Preventivas no disponibles para el uptime: {_e}")

                # ── TOTAL GENERAL = correctivas + preventivas ──────────
                _paro_total_seg = _paro_corr_seg + _paro_prev_seg

                # ── KPIs ───────────────────────────────────────────────
                _k1s, _k2s, _k3s, _k4s, _k5s = st.columns(5)
                _prio_lbl = {"Solo P1 (máquina detenida)": "P1", "P1 + P2": "P1+P2", "Todas": "Todas"}[_up_prio]
                _k1s.metric("Llamados contados", f"{len(_df_up):,}",
                            delta=f"prioridad: {_prio_lbl}")
                _k2s.metric("Tiempo detenido (Correctivas)",
                            f"{int(_paro_corr_seg // 3600):,}h "
                            f"{int((_paro_corr_seg % 3600) // 60):02d}m",
                            help="Suma del tiempo entre fecha_llamado y primer cierre real, "
                                 "sobre los llamados que pasan el filtro de prioridad.")
                _k3s.metric("Tiempo detenido (Preventivas)",
                            f"{int(_paro_prev_seg // 3600):,}h "
                            f"{int((_paro_prev_seg % 3600) // 60):02d}m",
                            delta=f"{_prev_ots_con_paro} MPs c/paro",
                            help="Suma de tiempo_paro_real_seg de las MP con "
                                 "'¿Paro de equipo? = SÍ' finalizadas dentro del período "
                                 "y que cumplen los filtros de cliente/región/EDS.")
                _k4s.metric("EDS evaluadas", f"{_n_eds_up:,}",
                            delta=f"{_rango_dias_up} días")
                _upt_pct = max(0.0, round(
                    (1 - _paro_total_seg / _total_seg_up) * 100, 3))
                _k5s.metric("Uptime General", f"{_upt_pct}%",
                            help=f"Denominador: {_n_eds_up:,} EDS × {_rango_dias_up} días × 24h · "
                                 f"Numerador: {int(_paro_corr_seg//3600):,}h correctivas "
                                 f"+ {int(_paro_prev_seg//3600):,}h preventivas "
                                 f"= {int(_paro_total_seg//3600):,}h detenidas totales.")

                st.divider()

                # ── Ranking: 5 EDS con mayor paro (correctivas + preventivas) ──
                st.markdown(
                    '<div class="section-header">'
                    '🏆  Ranking · 5 estaciones con mayor tiempo detenido (general)'
                    '</div>',
                    unsafe_allow_html=True,
                )
                if _df_up.empty or _paro_total_seg == 0:
                    st.info("Sin tiempo detenido en el filtro.")
                else:
                    # Paro por EDS: correctivas
                    _rk_corr = (_df_up.groupby(
                        ["eds_nombre", "eds_occim"], dropna=False, as_index=False
                    ).agg(
                        Llamados=("os_fracttal", "count"),
                        Paro_corr_seg=("_paro_seg", "sum"),
                    ))
                    # Paro por EDS: preventivas (sumar al mismo eds_occim)
                    if not _df_prev_up.empty:
                        _rk_prev = (_df_prev_up.groupby(
                            "codigo_eds", dropna=False, as_index=False
                        ).agg(
                            MPs_con_paro=("id_ot", "count"),
                            Paro_prev_seg=("_paro_calc_seg", "sum"),
                        )).rename(columns={"codigo_eds": "eds_occim"})
                    else:
                        _rk_prev = pd.DataFrame(columns=["eds_occim","MPs_con_paro","Paro_prev_seg"])
                    _rk = _rk_corr.merge(_rk_prev, on="eds_occim", how="outer")
                    _rk["Llamados"]      = _rk["Llamados"].fillna(0).astype(int)
                    _rk["MPs_con_paro"]  = _rk["MPs_con_paro"].fillna(0).astype(int)
                    _rk["Paro_corr_seg"] = _rk["Paro_corr_seg"].fillna(0)
                    _rk["Paro_prev_seg"] = _rk["Paro_prev_seg"].fillna(0)
                    _rk["Tiempo_paro_seg"] = _rk["Paro_corr_seg"] + _rk["Paro_prev_seg"]
                    _rk["Tiempo detenido (h:mm)"] = _rk["Tiempo_paro_seg"].apply(
                        lambda s: f"{int(s // 3600):,}h {int((s % 3600) // 60):02d}m"
                        if s else "—")
                    _rk["Tiempo detenido (h)"] = (_rk["Tiempo_paro_seg"] / 3600).round(1)
                    _rk["Correctivas (h)"] = (_rk["Paro_corr_seg"] / 3600).round(1)
                    _rk["Preventivas (h)"] = (_rk["Paro_prev_seg"] / 3600).round(1)
                    _rk = _rk.sort_values("Tiempo_paro_seg", ascending=False).head(5)
                    _rk["#"] = ["🥇","🥈","🥉","4️⃣","5️⃣"][:len(_rk)]

                    _rk_show = _rk[[
                        "#", "eds_nombre", "eds_occim", "Llamados", "MPs_con_paro",
                        "Correctivas (h)", "Preventivas (h)",
                        "Tiempo detenido (h:mm)", "Tiempo detenido (h)"
                    ]].rename(columns={
                        "eds_nombre":    "Estación",
                        "eds_occim":     "Cód. EDS",
                        "MPs_con_paro":  "MPs c/paro",
                    })
                    _show_df(_rk_show.reset_index(drop=True), hide_index=True,
                        use_container_width=True,
                        column_config={
                            "#":                       st.column_config.TextColumn(width=50),
                            "Estación":                st.column_config.TextColumn(width=220),
                            "Cód. EDS":                st.column_config.TextColumn(width=90),
                            "Llamados":                st.column_config.NumberColumn(format="%d", width=80),
                            "MPs c/paro":              st.column_config.NumberColumn(format="%d", width=80,
                                help="Mantenciones preventivas con paro de equipo"),
                            "Correctivas (h)":         st.column_config.NumberColumn(format="%.1f", width=100),
                            "Preventivas (h)":         st.column_config.NumberColumn(format="%.1f", width=100),
                            "Tiempo detenido (h:mm)":  st.column_config.TextColumn(width=130),
                            "Tiempo detenido (h)":     st.column_config.ProgressColumn(
                                format="%.1f h", min_value=0,
                                max_value=float(_rk["Tiempo detenido (h)"].max()) if not _rk.empty else 1,
                                width=180),
                        })

                st.divider()

                # ── Uptime por EDS (general) ───────────────────────────
                st.markdown(
                    '<div class="section-header">🏭  Uptime por EDS (general)</div>',
                    unsafe_allow_html=True,
                )
                st.caption(
                    "Cada fila es una EDS. **Uptime %** = % del tiempo del período "
                    "en que la EDS NO tuvo emergencias activas ni mantenciones preventivas con paro. "
                    "Suma correctivas + preventivas."
                )
                if _df_up.empty or _paro_total_seg == 0:
                    st.info("Sin tiempo detenido en el filtro.")
                else:
                    _gpe_corr = (_df_up.groupby(
                        ["eds_occim", "eds_nombre", "cliente"],
                        dropna=False, as_index=False,
                    ).agg(
                        Llamados=("os_fracttal", "count"),
                        Paro_corr_seg=("_paro_seg", "sum"),
                        Equipos_codigos=("codigo_activo",
                            lambda s: ", ".join(sorted({c for x in s if pd.notna(x) and str(x)
                                                        for c in str(x).split(",") if c.strip()}))),
                        Equipos_nombres=("nombre_activo",
                            lambda s: " · ".join(sorted({n.strip() for x in s if pd.notna(x) and str(x)
                                                        for n in str(x).split(" · ") if n.strip()}))),
                        _ultimo_llamado=("fecha_llamado", "max"),
                    ))
                    # Preventivas agregadas por EDS
                    if not _df_prev_up.empty:
                        _gpe_prev = (_df_prev_up.groupby(
                            "codigo_eds", dropna=False, as_index=False
                        ).agg(
                            MPs=("id_ot", "count"),
                            Paro_prev_seg=("_paro_calc_seg", "sum"),
                        )).rename(columns={"codigo_eds": "eds_occim"})
                    else:
                        _gpe_prev = pd.DataFrame(columns=["eds_occim","MPs","Paro_prev_seg"])
                    _gpe = _gpe_corr.merge(_gpe_prev, on="eds_occim", how="outer")
                    _gpe["Llamados"]      = _gpe["Llamados"].fillna(0).astype(int)
                    _gpe["MPs"]           = _gpe["MPs"].fillna(0).astype(int)
                    _gpe["Paro_corr_seg"] = _gpe["Paro_corr_seg"].fillna(0)
                    _gpe["Paro_prev_seg"] = _gpe["Paro_prev_seg"].fillna(0)
                    _gpe["Tiempo_paro_seg"] = _gpe["Paro_corr_seg"] + _gpe["Paro_prev_seg"]
                    _gpe["Última emergencia"] = pd.to_datetime(
                        _gpe["_ultimo_llamado"], errors="coerce"
                    ).dt.strftime("%d/%m/%Y").fillna("—")
                    _gpe["Tiempo detenido (h:mm)"] = _gpe["Tiempo_paro_seg"].apply(
                        lambda s: f"{int(s // 3600):,}h {int((s % 3600) // 60):02d}m"
                        if s else "—")
                    _gpe["Correctivas (h)"] = (_gpe["Paro_corr_seg"] / 3600).round(1)
                    _gpe["Preventivas (h)"] = (_gpe["Paro_prev_seg"] / 3600).round(1)
                    _gpe["Uptime %"] = (
                        (1 - _gpe["Tiempo_paro_seg"] / _seg_equipo_up).clip(lower=0) * 100
                    ).round(3)
                    _gpe = _gpe.sort_values("Tiempo_paro_seg", ascending=False).head(30)

                    _gpe_show = _gpe[[
                        "eds_occim","eds_nombre","cliente","Equipos_codigos","Equipos_nombres",
                        "Llamados","MPs","Correctivas (h)","Preventivas (h)",
                        "Tiempo detenido (h:mm)","Última emergencia","Uptime %"
                    ]].rename(columns={
                        "eds_occim":       "Cód. EDS",
                        "eds_nombre":      "Estación",
                        "cliente":         "Cliente",
                        "Equipos_codigos": "Equipo(s)",
                        "Equipos_nombres": "Tipo equipo",
                        "MPs":             "MPs c/paro",
                    })
                    _show_df(_gpe_show.reset_index(drop=True), hide_index=True,
                             use_container_width=True,
                             column_config={
                                 "Cód. EDS":                st.column_config.TextColumn(width=90),
                                 "Estación":                st.column_config.TextColumn(width=200),
                                 "Cliente":                 st.column_config.TextColumn(width=110),
                                 "Equipo(s)":               st.column_config.TextColumn(width=120),
                                 "Tipo equipo":             st.column_config.TextColumn(width=180),
                                 "Llamados":                st.column_config.NumberColumn(
                                     format="%d", width=70),
                                 "MPs c/paro":              st.column_config.NumberColumn(
                                     format="%d", width=70,
                                     help="MPs preventivas con paro de equipo"),
                                 "Correctivas (h)":         st.column_config.NumberColumn(format="%.1f", width=95),
                                 "Preventivas (h)":         st.column_config.NumberColumn(format="%.1f", width=95),
                                 "Tiempo detenido (h:mm)":  st.column_config.TextColumn(width=125,
                                     help="Correctivas + Preventivas"),
                                 "Última emergencia":       st.column_config.TextColumn(width=105),
                                 "Uptime %":                st.column_config.NumberColumn(
                                     format="%.3f%%", width=95),
                             })
                    st.caption(
                        f"Top 30 EDS ordenadas por tiempo total detenido · "
                        f"período: {_rango_inicio_up.strftime('%d/%m/%Y')}"
                        f" → {_rango_fin_up.strftime('%d/%m/%Y')}"
                    )


# ─────────────────────────────────────────────────────────────────────────────
# PÁGINA 3: ESTACIONES DE SERVICIO
# ─────────────────────────────────────────────────────────────────────────────
elif _page == _NAV_PAGES[3]:

    # ── Datos base ──────────────────────────────────────────────────────────
    import datetime as _dt
    import plotly.graph_objects as go
    import plotly.express as px
    _hoy        = pd.Timestamp.now()
    _cur_year   = _hoy.year
    _cur_month  = _hoy.month
    _MES_ES     = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
                   7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}
    _MESES_DISP = [f"{_cur_year}-{m:02d}" for m in range(1, _cur_month + 1)]

    # Work orders con eds_occim ──────────────────────────────────────────────
    raw_wo_eds = load_work_orders_supabase()
    _wo_eds_sig = str(len(raw_wo_eds))

    def _build_wo_eds():
        _df = build_work_orders_df(raw_wo_eds)
        _fmap = {
            str(wo.get("wo_folio")): wo.get("groups_2_description")
            for wo in raw_wo_eds
            if wo.get("groups_2_description")
        }
        _df["eds_occim"] = _df["folio"].astype(str).map(_fmap)
        _cd = _df["creation_date"].dt.tz_convert(None) if _df["creation_date"].dt.tz is not None else _df["creation_date"]
        _df["mes_str"] = _cd.dt.to_period("M").astype(str)
        _df["mes_num"] = _cd.dt.month
        _df["year"]    = _cd.dt.year
        return _df

    df_wo_eds_full = _sc("df_wo_eds_v4_comentario", _wo_eds_sig, _build_wo_eds)

    # Filtrar a clientes EDS y año actual ────────────────────────────────────
    _EDS_CLIENTS = {"COPEC", "Aramco (Esmax)", "ESMAX (Aramco)", "SHELL (Enex)"}
    df_wo_cur = df_wo_eds_full[
        df_wo_eds_full["client"].isin(_EDS_CLIENTS) &
        (df_wo_eds_full["year"] == _cur_year)
    ].copy()

    # Llamados correctivos ───────────────────────────────────────────────────
    if not df_llamados.empty:
        df_ll_cur = df_llamados[
            df_llamados["cliente"].isin(_EDS_CLIENTS) &
            (df_llamados.get("Año", pd.Series(dtype=int)) == _cur_year
             if "Año" in df_llamados.columns else
             pd.to_datetime(df_llamados["fecha_llamado"], errors="coerce").dt.year == _cur_year)
        ].copy()
        if "Mes" not in df_ll_cur.columns:
            df_ll_cur["Mes"] = pd.to_datetime(df_ll_cur["fecha_llamado"], errors="coerce").dt.month
    else:
        df_ll_cur = pd.DataFrame()

    # EDS master ─────────────────────────────────────────────────────────────
    _CLIENTES_EXCLUIR = {"BLANDFORD","OCCIMIANO","PARTICULAR","TERPEL COL","TERPEL","ABASTIBLE"}
    df_eds_activas = df_eds[df_eds["activa"]].copy()
    df_eds_activas = df_eds_activas[
        ~df_eds_activas["cliente"].str.strip().str.upper().isin(_CLIENTES_EXCLUIR)
    ].copy()

    # Numerales por subtarea (para detalle preventivas Shell)
    df_num_sub_eds = _sc("df_num_sub_eds_v2", _wo_eds_sig, load_numerales_subtarea_supabase)

    # ── Override per-asset: el numeral correcto vive en numerales_subtarea,
    # NO en ordenes_trabajo (que es OT-level y se duplica al expandir activos).
    # Si (folio, equipment_code) tiene fila en numerales_subtarea, sobrescribir.
    if not df_num_sub_eds.empty and "equipment_code" in df_wo_eds_full.columns:
        _ns_map = {
            (str(r["id_ot"]), str(r["codigo_activo"])): (
                r.get("numeral_inicial"),
                r.get("numeral_final"),
            )
            for _, r in df_num_sub_eds.iterrows()
            if pd.notna(r.get("id_ot")) and pd.notna(r.get("codigo_activo"))
        }
        def _override_ni(row):
            k = (str(row.get("folio", "")), str(row.get("equipment_code", "")))
            v = _ns_map.get(k)
            return v[0] if v and v[0] is not None and str(v[0]).strip() not in ("", "nan", "None") else row.get("numeral_inicial")
        def _override_nf(row):
            k = (str(row.get("folio", "")), str(row.get("equipment_code", "")))
            v = _ns_map.get(k)
            return v[1] if v and v[1] is not None and str(v[1]).strip() not in ("", "nan", "None") else row.get("numeral_final")
        df_wo_eds_full = df_wo_eds_full.copy()
        df_wo_eds_full["numeral_inicial"] = df_wo_eds_full.apply(_override_ni, axis=1)
        df_wo_eds_full["numeral_final"]   = df_wo_eds_full.apply(_override_nf, axis=1)
        df_wo_cur = df_wo_eds_full[
            df_wo_eds_full["client"].isin(_EDS_CLIENTS) &
            (df_wo_eds_full["year"] == _cur_year)
        ].copy()

    # Colores empresa ────────────────────────────────────────────────────────
    _CL_COLORS = {
        "COPEC":          {"tab":"🔴","pm":"#CC0000","cm":"#F4A7A9","accent":"#CC0000","label":"COPEC"},
        "SHELL (Enex)":   {"tab":"🟡","pm":"#D4A800","cm":"#FAE57A","accent":"#D4A800","label":"Shell (Enex)"},
        "Aramco (Esmax)": {"tab":"🟢","pm":"#16A34A","cm":"#86EFAC","accent":"#16A34A","label":"Aramco (Esmax)"},
        "ESMAX (Aramco)": {"tab":"🟢","pm":"#16A34A","cm":"#86EFAC","accent":"#16A34A","label":"Aramco (Esmax)"},  # alias legacy
    }

    # ── Título y tabs ────────────────────────────────────────────────────────
    _hdr("Estaciones de Servicio")
    st.divider()

    _tabs_eds = st.tabs([
        "🔴  COPEC",
        "🟢  Aramco (Esmax)",
        "🟡  Shell (Enex)",
        "🔢  Historial de Numerales",
    ])

    # ════════════════════════════════════════════════════════════════════════
    # HELPER: dona de tipos de falla + desglose de causas
    # ════════════════════════════════════════════════════════════════════════
    _FALLA_PAL = ["#EF4444","#F59E0B","#6B7280","#3B82F6","#8B5CF6","#10B981","#EC4899","#14B8A6"]

    def _limpiar_tipo(t: str) -> str:
        """Quita prefijos numéricos: '01.- F.N.A.O.' → 'F.N.A.O.'"""
        import re as _re2
        return _re2.sub(r"^\d+\.\-?\s*", "", str(t)).strip()

    def _render_fallas_panel(df_src: "pd.DataFrame", key_sfx: str):
        """Dona con tipos de falla + panel de causas por tipo."""
        if df_src.empty or "failure_type" not in df_src.columns:
            st.caption("Sin datos de tipo de falla disponibles.")
            return
        _df_ft = df_src[df_src["failure_type"].str.strip() != ""].copy()
        if _df_ft.empty:
            st.caption("Sin datos de tipo de falla disponibles.")
            return

        _fallas_cnt = _df_ft["failure_type"].value_counts().head(6)
        _total      = int(_fallas_cnt.sum())
        _has_cause  = "failure_cause" in _df_ft.columns

        _dona_col, _det_col = st.columns([2, 3])

        with _dona_col:
            _labels_short = [_limpiar_tipo(f) for f in _fallas_cnt.index]
            _fig_dona = go.Figure(go.Pie(
                labels=_labels_short,
                values=_fallas_cnt.values,
                hole=0.52,
                marker=dict(colors=_FALLA_PAL[:len(_fallas_cnt)],
                            line=dict(color="rgba(0,0,0,0.08)", width=1)),
                textinfo="percent",
                textposition="outside",
                hovertemplate="%{label}<br>%{value} OTs · %{percent}<extra></extra>",
                sort=False,
            ))
            _fig_dona.update_layout(
                height=300,
                margin=dict(l=0, r=0, t=10, b=10),
                showlegend=True,
                legend=dict(orientation="v", x=1.0, y=0.5,
                            font=dict(size=11), bgcolor="rgba(0,0,0,0)"),
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
                font=dict(color=_t["text"]),
                annotations=[dict(
                    text=f"<b>{_total}</b><br><span style='font-size:11px'>OTs</span>",
                    x=0.5, y=0.5, showarrow=False,
                    font=dict(size=16, color=_t["text"]),
                )],
            )
            st.plotly_chart(_fig_dona, use_container_width=True, key=f"dona_{key_sfx}")

        with _det_col:
            st.markdown(
                f'<div style="font-size:0.78rem;font-weight:700;color:{_t["muted"]};'
                f'letter-spacing:0.05em;margin-bottom:8px;">DESGLOSE POR TIPO</div>',
                unsafe_allow_html=True,
            )
            for _i, (_ftype, _cnt) in enumerate(_fallas_cnt.items()):
                _pct   = _cnt / _total * 100
                _color = _FALLA_PAL[_i % len(_FALLA_PAL)]
                _short = _limpiar_tipo(_ftype)

                with st.expander(f"{_short}  —  {_cnt} OTs · {_pct:.1f}%"):
                    if _has_cause:
                        _df_tipo  = _df_ft[_df_ft["failure_type"] == _ftype]
                        _causas_v = (
                            _df_tipo[_df_tipo["failure_cause"].str.strip() != ""]
                            ["failure_cause"].value_counts().head(4)
                        )
                        if not _causas_v.empty:
                            for _cn, _cc in _causas_v.items():
                                _cp = _cc / _cnt * 100
                                _bar_w = int(_cp)
                                st.markdown(
                                    f'<div style="padding:3px 6px 3px 10px;">'
                                    f'<div style="display:flex;justify-content:space-between;'
                                    f'font-size:0.76rem;color:{_t["muted"]};">'
                                    f'<span>↳ {_cn}</span>'
                                    f'<span style="font-weight:600;color:{_t["text"]};">'
                                    f'{_cc} ({_cp:.0f}%)</span></div>'
                                    f'<div style="height:3px;border-radius:2px;margin-top:2px;'
                                    f'background:{_color};opacity:0.35;width:{_bar_w}%;"></div>'
                                    f'</div>',
                                    unsafe_allow_html=True,
                                )
                        else:
                            st.caption("Sin causa registrada")

    # ════════════════════════════════════════════════════════════════════════
    # FUNCIÓN: renderizar pestaña por empresa
    # ════════════════════════════════════════════════════════════════════════
    def _render_eds_tab(company: str, tab_idx: int):
        _col = _CL_COLORS[company]
        _ck  = company.replace(" ","_").replace("(","").replace(")","")  # key prefix

        df_ll_c  = df_ll_cur[df_ll_cur["cliente"] == company].copy()  if not df_ll_cur.empty  else pd.DataFrame()
        df_wo_c  = df_wo_cur[df_wo_cur["client"]  == company].copy()
        df_eds_c = df_eds_activas[df_eds_activas["cliente"].str.contains(
            company.split()[0], case=False, na=False)].copy()

        # ── Filtros ──────────────────────────────────────────────────────────
        _frow1, _frow2 = st.columns([2, 3])
        with _frow1:
            _mes_sel = st.multiselect(
                "Mes (dejar vacío = año completo)",
                options=_MESES_DISP,
                default=[],
                key=f"eds_mes_{_ck}",
                format_func=lambda m: f"{_MES_ES[int(m.split('-')[1])]} {m.split('-')[0]}",
            )
        _meses_activos = _mes_sel if _mes_sel else _MESES_DISP

        # Opciones EDS desde df_llamados (tienen eds_occim y eds_nombre)
        if not df_ll_c.empty and "eds_occim" in df_ll_c.columns and "eds_nombre" in df_ll_c.columns:
            _eds_opts_df = (
                df_ll_c[["eds_occim","eds_nombre"]]
                .dropna(subset=["eds_occim","eds_nombre"])
                .drop_duplicates()
                .sort_values("eds_nombre")
            )
            _eds_nombre_to_code = dict(zip(_eds_opts_df["eds_nombre"], _eds_opts_df["eds_occim"]))
            _eds_opts_list = ["Todas"] + list(_eds_opts_df["eds_nombre"])
        else:
            _eds_nombre_to_code = {}
            _eds_opts_list = ["Todas"]
        with _frow2:
            _eds_sel_nombre = st.selectbox(
                "EDS específica",
                _eds_opts_list,
                key=f"eds_sel_{_ck}",
            )
        _eds_sel_code = _eds_nombre_to_code.get(_eds_sel_nombre) if _eds_sel_nombre != "Todas" else None

        # ── Aplicar filtros ──────────────────────────────────────────────────
        def _filtrar_ll(df):
            if df.empty: return df
            df = df[df["mes_str"].isin(_meses_activos)] if "mes_str" in df.columns else df[
                pd.to_datetime(df["fecha_llamado"], errors="coerce").dt.to_period("M").astype(str).isin(_meses_activos)]
            if _eds_sel_code:
                df = df[df["eds_occim"] == _eds_sel_code]
            return df

        def _filtrar_wo(df):
            if df.empty: return df
            df = df[df["mes_str"].isin(_meses_activos)]
            if _eds_sel_code:
                df = df[df["eds_occim"] == _eds_sel_code]
            return df

        # Enrich df_ll_c with mes_str if missing
        if not df_ll_c.empty and "mes_str" not in df_ll_c.columns:
            df_ll_c["mes_str"] = pd.to_datetime(df_ll_c["fecha_llamado"], errors="coerce").dt.to_period("M").astype(str)

        df_ll_f  = _filtrar_ll(df_ll_c)
        df_pm_f  = _filtrar_wo(df_wo_c[df_wo_c["maint_type"] == "Preventiva"])
        df_cm_f  = _filtrar_wo(df_wo_c[df_wo_c["maint_type"] == "Correctiva"])

        # ── KPI cards ────────────────────────────────────────────────────────
        _n_eds_atend   = df_ll_f["eds_occim"].nunique()    if not df_ll_f.empty and "eds_occim" in df_ll_f.columns else 0
        _n_llamados    = df_ll_f["n_llamado"].nunique()    if not df_ll_f.empty and "n_llamado" in df_ll_f.columns else len(df_ll_f)
        _n_pms         = df_pm_f["folio"].nunique()        if not df_pm_f.empty else 0
        _pct_sla       = None
        if not df_ll_f.empty and "cumplimiento" in df_ll_f.columns:
            _tot_sla   = (df_ll_f["cumplimiento"] != "SIN DATOS").sum()
            _cumple    = (df_ll_f["cumplimiento"] == "CUMPLE").sum()
            _pct_sla   = f"{_cumple/_tot_sla*100:.1f}%" if _tot_sla > 0 else "—"
        _ultima_fecha  = None
        if not df_ll_f.empty and "fecha_llamado" in df_ll_f.columns:
            _ult = pd.to_datetime(df_ll_f["fecha_llamado"], errors="coerce").max()
            _ultima_fecha = _ult.strftime("%d/%m/%Y") if pd.notna(_ult) else "—"

        _kc1, _kc2, _kc3, _kc4, _kc5 = st.columns(5)
        _kc1.metric("EDS atendidas",        _n_eds_atend)
        _kc2.metric("Llamados correctivos", _n_llamados)
        _kc3.metric("PMs realizados",       _n_pms)
        _kc4.metric("% SLA cumplido",       _pct_sla if _pct_sla else "—")
        _kc5.metric("Última atención",      _ultima_fecha if _ultima_fecha else "—")

        # ── Gráfico PM vs CM mensual ─────────────────────────────────────────
        # Filtrar por EDS seleccionada (fix: antes usaba datos sin filtrar)
        _pm_chart_src = df_wo_c[df_wo_c["maint_type"] == "Preventiva"].copy()
        _cm_chart_src = df_ll_c.copy()
        if _eds_sel_code:
            if not _pm_chart_src.empty and "eds_occim" in _pm_chart_src.columns:
                _pm_chart_src = _pm_chart_src[_pm_chart_src["eds_occim"] == _eds_sel_code]
            if not _cm_chart_src.empty and "eds_occim" in _cm_chart_src.columns:
                _cm_chart_src = _cm_chart_src[_cm_chart_src["eds_occim"] == _eds_sel_code]

        _pm_by_m = (
            _pm_chart_src[_pm_chart_src["mes_str"].isin(_MESES_DISP) if "mes_str" in _pm_chart_src.columns else pd.Series(True, index=_pm_chart_src.index)]
            .groupby("mes_num")["folio"].nunique()
            .reset_index()
            .rename(columns={"mes_num":"mes","folio":"PM"})
        ) if not _pm_chart_src.empty else pd.DataFrame(columns=["mes","PM"])

        _cm_by_m = (
            _cm_chart_src[_cm_chart_src["Mes"].isin(range(1, _cur_month+1))]
            .groupby("Mes")["n_llamado"].count()
            .reset_index()
            .rename(columns={"Mes":"mes","n_llamado":"CM"})
        ) if not _cm_chart_src.empty and "Mes" in _cm_chart_src.columns else pd.DataFrame(columns=["mes","CM"])

        _chart_base = pd.DataFrame({"mes": list(range(1, _cur_month+1))})
        _chart_df = _chart_base.merge(_pm_by_m, on="mes", how="left").merge(_cm_by_m, on="mes", how="left").fillna(0)
        _chart_df["mes_lbl"] = _chart_df["mes"].map(_MES_ES)

        import plotly.graph_objects as go
        # Etiquetas: mostrar count solo si > 0
        _pm_txt = [str(int(v)) if v > 0 else "" for v in _chart_df["PM"]]
        _cm_txt = [str(int(v)) if v > 0 else "" for v in _chart_df["CM"]]
        _fig_main = go.Figure()
        _fig_main.add_trace(go.Bar(
            x=_chart_df["mes_lbl"], y=_chart_df["PM"],
            name="Preventivo", marker_color=_col["cm"], opacity=0.92,
            text=_pm_txt, textposition="inside", textfont=dict(color="#555555", size=12),
        ))
        _fig_main.add_trace(go.Bar(
            x=_chart_df["mes_lbl"], y=_chart_df["CM"],
            name="Correctivo (llamados)", marker_color=_col["pm"], opacity=0.92,
            text=_cm_txt, textposition="inside", textfont=dict(color="white", size=12),
        ))
        _fig_main.update_layout(
            barmode="stack",
            title=dict(text=f"Mantenciones {_cur_year} — {_col['label'] if _eds_sel_nombre == 'Todas' else _eds_sel_nombre}", font_size=14),
            height=340,
            margin=dict(l=0, r=0, t=40, b=0),
            legend=dict(orientation="h", y=-0.15),
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
            font_color=_t["text"],
            xaxis=dict(gridcolor=_t["border"]),
            yaxis=dict(gridcolor=_t["border"], title="N° mantenciones"),
        )
        st.plotly_chart(_fig_main, use_container_width=True, key=f"chart_main_{_ck}_{_eds_sel_nombre}")

        # ── SECCIÓN CONDICIONAL ──────────────────────────────────────────────
        if _eds_sel_nombre == "Todas":
            # ── TOP 5 EDS con más llamados ───────────────────────────────────
            st.markdown(
                f'<div style="font-size:1.0rem;font-weight:700;color:{_t["text"]};'
                f'margin:12px 0 8px 0;border-bottom:2px solid {_col["accent"]};'
                f'padding-bottom:4px;">📊 Top 5 EDS · Más llamados correctivos ({", ".join([_MES_ES[int(m.split("-")[1])] for m in _meses_activos])})</div>',
                unsafe_allow_html=True,
            )
            if not df_ll_f.empty and "eds_occim" in df_ll_f.columns:
                _top5_data = (
                    df_ll_f.groupby(["eds_occim","eds_nombre"], dropna=True)
                    .agg(
                        llamados=("n_llamado","count"),
                        ultimo_llamado=("fecha_llamado","max"),
                    )
                    .reset_index()
                    .sort_values("llamados", ascending=False)
                    .head(5)
                )

                _col_chart, _col_detail = st.columns([3, 2])
                with _col_chart:
                    _fig_top5 = go.Figure(go.Bar(
                        x=_top5_data["llamados"],
                        y=_top5_data["eds_nombre"],
                        orientation="h",
                        marker_color=_col["accent"],
                        text=_top5_data["llamados"],
                        textposition="outside",
                    ))
                    _fig_top5.update_layout(
                        height=280, margin=dict(l=0,r=40,t=10,b=0),
                        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                        font_color=_t["text"],
                        xaxis=dict(gridcolor=_t["border"]),
                        yaxis=dict(autorange="reversed"),
                    )
                    st.plotly_chart(_fig_top5, use_container_width=True, key=f"chart_top5_{_ck}")

                with _col_detail:
                    for _, _r5 in _top5_data.iterrows():
                        _ult5 = pd.to_datetime(_r5["ultimo_llamado"], errors="coerce")
                        _ult5_str = _ult5.strftime("%d/%m/%Y") if pd.notna(_ult5) else "—"
                        # Falla más frecuente para esa EDS desde df_wo
                        _fallas5 = pd.Series(dtype=str)
                        if not df_wo_c.empty and "failure_type" in df_wo_c.columns and "eds_occim" in df_wo_c.columns:
                            _fallas5 = df_wo_c[
                                (df_wo_c["eds_occim"] == _r5["eds_occim"]) &
                                (df_wo_c["failure_type"].str.strip() != "")
                            ]["failure_type"].value_counts()
                        _falla_top = _fallas5.index[0] if not _fallas5.empty else "—"
                        st.markdown(
                            f'<div style="background:{_t["card"]};border:1px solid {_t["border"]};'
                            f'border-radius:8px;padding:8px 10px;margin-bottom:6px;font-size:0.80rem;">'
                            f'<b style="color:{_col["accent"]};">{_r5["eds_nombre"]}</b><br>'
                            f'<span style="color:{_t["muted"]};">Llamados: </span><b>{int(_r5["llamados"])}</b> · '
                            f'<span style="color:{_t["muted"]};">Último: </span>{_ult5_str}<br>'
                            f'<span style="color:{_t["muted"]};">Falla frecuente: </span>{_falla_top}'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

                # ── Análisis de tipo de falla (company-level) — Dona ────────
                if not df_wo_c.empty and "failure_type" in df_wo_c.columns:
                    _df_fallas_c = df_wo_c[
                        (df_wo_c["maint_type"] == "Correctiva") &
                        (df_wo_c["failure_type"].str.strip() != "") &
                        (df_wo_c["mes_str"].isin(_meses_activos))
                    ]
                    st.markdown(
                        f'<div style="font-weight:700;font-size:0.95rem;margin:14px 0 8px 0;'
                        f'color:{_t["text"]};">🔩 Tipos de falla — correctivos</div>',
                        unsafe_allow_html=True,
                    )
                    _render_fallas_panel(_df_fallas_c, f"co_{_ck}")
            else:
                st.info("No hay datos de llamados para el período seleccionado.")

        else:
            # ── DETALLE EDS ESPECÍFICA ───────────────────────────────────────
            _eds_row = df_eds_activas[df_eds_activas["eds_occim"] == _eds_sel_code]
            _eds_info = _eds_row.iloc[0] if not _eds_row.empty else None

            if _eds_info is not None:
                st.markdown(
                    f'<div style="background:{_t["card"]};border-left:4px solid {_col["accent"]};'
                    f'border-radius:8px;padding:10px 14px;margin:8px 0 14px 0;font-size:0.85rem;">'
                    f'<b style="font-size:1.0rem;">{_eds_sel_nombre}</b>'
                    f'<span style="color:{_t["muted"]};"> · Cód: {_eds_sel_code}</span><br>'
                    f'<span style="color:{_t["muted"]};">Dirección: </span>{_eds_info.get("nombre","") or _eds_info.get("direccion","—")}'
                    f' · <span style="color:{_t["muted"]};">Comuna: </span>{_eds_info.get("comuna","—")}'
                    f' · <span style="color:{_t["muted"]};">Zona: </span>{_eds_info.get("zona_occim","—")}'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            # KPIs específicos EDS
            _sla_eds = "—"
            if not df_ll_f.empty and "cumplimiento" in df_ll_f.columns:
                _tot_e = (df_ll_f["cumplimiento"] != "SIN DATOS").sum()
                _ok_e  = (df_ll_f["cumplimiento"] == "CUMPLE").sum()
                _sla_eds = f"{_ok_e/_tot_e*100:.1f}%" if _tot_e > 0 else "—"

            _avg_resp = "—"
            if not df_ll_f.empty and "horas_resolucion" in df_ll_f.columns:
                _hr = pd.to_numeric(df_ll_f["horas_resolucion"], errors="coerce")
                if _hr.notna().any():
                    _avg_resp = f"{_hr.mean():.1f} h"

            # Regularidad: promedio días entre llamados
            _regularidad = "—"
            if not df_ll_f.empty and len(df_ll_f) > 1 and "fecha_llamado" in df_ll_f.columns:
                _fechas_ord = pd.to_datetime(df_ll_f["fecha_llamado"], errors="coerce").dropna().sort_values()
                if len(_fechas_ord) > 1:
                    _gaps = [(_fechas_ord.iloc[i+1] - _fechas_ord.iloc[i]).days for i in range(len(_fechas_ord)-1)]
                    _avg_gap = sum(_gaps) / len(_gaps)
                    _regularidad = f"c/ {_avg_gap:.0f} días"

            _dc1, _dc2, _dc3, _dc4 = st.columns(4)
            _dc1.metric("Llamados en período", len(df_ll_f))
            _dc2.metric("PMs en período",      df_pm_f["folio"].nunique() if not df_pm_f.empty else 0)
            _dc3.metric("SLA cumplimiento",     _sla_eds)
            _dc4.metric("T° resp. promedio",    _avg_resp)

            _dc5, _dc6 = st.columns(2)
            _dc5.metric("Regularidad llamados", _regularidad)
            if not df_ll_f.empty and "fecha_llamado" in df_ll_f.columns:
                _ult_ll = pd.to_datetime(df_ll_f["fecha_llamado"], errors="coerce").max()
                _dc6.metric("Último llamado", _ult_ll.strftime("%d/%m/%Y") if pd.notna(_ult_ll) else "—")

            # Timeline llamados para esta EDS
            if not df_ll_f.empty and "Mes" in df_ll_f.columns:
                _ll_mes = df_ll_f.groupby("Mes")["n_llamado"].count().reset_index()
                _all_months = pd.DataFrame({"Mes": list(range(1, _cur_month+1))})
                _ll_mes = _all_months.merge(_ll_mes, on="Mes", how="left").fillna(0)
                _ll_mes["lbl"] = _ll_mes["Mes"].map(_MES_ES)
                _fig_tl = go.Figure(go.Bar(
                    x=_ll_mes["lbl"], y=_ll_mes["n_llamado"],
                    marker_color=_col["cm"], name="Llamados",
                ))
                _fig_tl.update_layout(
                    title=dict(text="Llamados correctivos por mes", font_size=13),
                    height=240, margin=dict(l=0,r=0,t=35,b=0),
                    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                    font_color=_t["text"],
                    xaxis=dict(gridcolor=_t["border"]),
                    yaxis=dict(gridcolor=_t["border"]),
                )
                st.plotly_chart(_fig_tl, use_container_width=True, key=f"chart_tl_{_ck}_{_eds_sel_code}")

            # Últimas visitas
            st.markdown(
                f'<div style="font-weight:700;font-size:0.95rem;margin:12px 0 8px 0;'
                f'color:{_t["text"]};">🔧 Últimas atenciones</div>',
                unsafe_allow_html=True,
            )
            if not df_ll_f.empty:
                _ultimas = (
                    df_ll_f.sort_values("fecha_llamado", ascending=False)
                    [["fecha_llamado","tecnico_corto" if "tecnico_corto" in df_ll_f.columns else "tecnico",
                      "prioridad","cumplimiento","horas_resolucion"
                      if "horas_resolucion" in df_ll_f.columns else "cumplimiento"]]
                    .head(10)
                    .rename(columns={
                        "fecha_llamado":"Fecha",
                        "tecnico_corto":"Técnico","tecnico":"Técnico",
                        "prioridad":"Prioridad",
                        "cumplimiento":"SLA",
                        "horas_resolucion":"H. Resp.",
                    })
                )
                _ultimas["Fecha"] = pd.to_datetime(_ultimas["Fecha"], errors="coerce").dt.strftime("%d/%m/%Y")
                _show_df(_ultimas, use_container_width=True, hide_index=True)
            else:
                st.caption("Sin llamados en el período seleccionado.")

            # Análisis de fallas para EDS específica — Dona
            if not df_cm_f.empty and "failure_type" in df_cm_f.columns:
                _df_fallas_eds = df_cm_f[df_cm_f["failure_type"].str.strip() != ""]
                if not _df_fallas_eds.empty:
                    st.markdown(
                        f'<div style="font-weight:700;font-size:0.95rem;margin:14px 0 8px 0;'
                        f'color:{_t["text"]};">🔩 Tipos de falla registrados</div>',
                        unsafe_allow_html=True,
                    )
                    _render_fallas_panel(_df_fallas_eds, f"eds_{_ck}_{_eds_sel_code}")

            # ── Historial de numerales por equipo ────────────────────────────
            # Últimos 10 registros de cada lavadora/aspiradora de esta EDS, con
            # detección de anomalías DENTRO de cada OT (final − inicial):
            #   ✅ 1-50 fichas (prueba normal) · 🟡 50-100 (raro) ·
            #   🔴 >100, final<inicial o valor inválido.
            _df_num_hist_eds = build_numeral_historial(
                df_wo_c, eds_code=_eds_sel_code, n=10
            )
            if not _df_num_hist_eds.empty:
                _n_equipos = _df_num_hist_eds["equipment_code"].nunique()
                _n_alertas = int((_df_num_hist_eds["severidad"] == "alert").sum())
                _n_warn    = int((_df_num_hist_eds["severidad"] == "warn").sum())
                _alert_txt = (
                    f" · <span style='color:#EF4444;'>{_n_alertas} anomalía(s)</span>"
                    if _n_alertas else ""
                )
                _warn_txt = (
                    f" · <span style='color:#F59E0B;'>{_n_warn} a revisar</span>"
                    if _n_warn else ""
                )
                st.markdown(
                    f'<div style="font-weight:700;font-size:0.95rem;margin:18px 0 6px 0;'
                    f'color:{_t["text"]};">🔢 Historial de numerales — {_n_equipos} equipo(s)'
                    f'{_alert_txt}{_warn_txt}</div>',
                    unsafe_allow_html=True,
                )
                st.caption(
                    "Últimos 10 registros por equipo. **Fichas = N. Final − N. Inicial** "
                    "dentro de la misma OT (fichas que el técnico usó probando la máquina). "
                    "✅ 1–50 normal · 🟡 50–100 raro · 🔴 >100, Final<Inicial o valor inválido. "
                    "Saltos grandes *entre* fechas distintas son normales (venta de fichas)."
                )
                # Agrupar visualmente por equipo
                for _eq_code, _grp in _df_num_hist_eds.groupby("equipment_code", sort=False):
                    _eq_nombre = _grp["equipment"].iloc[0]
                    _eq_alertas = int((_grp["severidad"] == "alert").sum())
                    _badge = f"  🔴 {_eq_alertas}" if _eq_alertas else ""
                    with st.expander(f"{_eq_nombre}  ({_eq_code}){_badge}",
                                     expanded=bool(_eq_alertas)):
                        _disp = _grp.copy()
                        _disp["Fecha"] = pd.to_datetime(
                            _disp["fecha"], errors="coerce"
                        ).dt.strftime("%d/%m/%Y").fillna("—")
                        _cols_disp = ["Fecha","folio","technician","numeral_inicial","numeral_final","estado"]
                        _disp = _disp[[c for c in _cols_disp if c in _disp.columns]].rename(columns={
                            "folio":           "OT",
                            "technician":      "Técnico",
                            "numeral_inicial": "N. Inicial",
                            "numeral_final":   "N. Final",
                            "estado":          "Fichas (prueba)",
                        })
                        _show_df(_disp, use_container_width=True, hide_index=True,
                            column_config={
                                "Fecha":      st.column_config.TextColumn(width=95),
                                "OT":         st.column_config.TextColumn(width=95),
                                "Técnico":    st.column_config.TextColumn(width=165),
                                "N. Inicial": st.column_config.TextColumn(width=95),
                                "N. Final":   st.column_config.TextColumn(width=95),
                                "Fichas (prueba)": st.column_config.TextColumn(width=165,
                                    help="Final − Inicial dentro de la OT = fichas de prueba."),
                            })

        # ── Tabla de detalle EDS ─────────────────────────────────────────────
        # Título con contexto del período seleccionado (para dejar claro qué
        # rango temporal cubre la tabla — antes el usuario no sabía si estaba
        # viendo el año completo o el mes filtrado).
        _rango_lbl = (
            ", ".join([f"{_MES_ES[int(m.split('-')[1])]} {m.split('-')[0]}"
                       for m in _meses_activos])
            if _mes_sel else f"año {_cur_year} completo"
        )
        st.markdown(
            f'<div style="font-weight:700;font-size:0.95rem;margin:18px 0 8px 0;'
            f'color:{_t["text"]};">📋 Listado de estaciones — {_col["label"]}  '
            f'<span style="color:{_t["muted"]};font-weight:400;font-size:0.85rem;">'
            f'· período: {_rango_lbl}</span></div>',
            unsafe_allow_html=True,
        )

        # ── Construir tabla desde EDS master + df_ll_f (correctivas filtradas)
        #    + df_pm_f (preventivas filtradas). Ambos ya vienen filtrados por
        #    el multiselect de mes, así que la tabla ahora sí reacciona.
        _df_tbl = df_eds_c.copy()

        # Agregado de CORRECTIVAS por EDS (desde df_ll_f, ya filtrado por mes)
        if not df_ll_f.empty and "eds_occim" in df_ll_f.columns:
            _agg_corr = (df_ll_f.groupby("eds_occim", dropna=True).agg(
                correctivas=("n_llamado", "count"),
                p1=("prioridad", lambda x: (x.astype(str).str.upper() == "P1").sum()),
                ultima_correctiva=("fecha_llamado", "max"),
                ultimo_tecnico=("tecnico", lambda x: x.dropna().iloc[-1]
                                if len(x.dropna()) > 0 else None),
            ).reset_index())
        else:
            _agg_corr = pd.DataFrame(columns=[
                "eds_occim","correctivas","p1","ultima_correctiva","ultimo_tecnico"])

        # Agregado de PREVENTIVAS por EDS (desde df_pm_f, ya filtrado por mes)
        if not df_pm_f.empty and "eds_occim" in df_pm_f.columns:
            _agg_prev = (df_pm_f.groupby("eds_occim", dropna=True).agg(
                preventivas=("folio", "nunique"),
                ultima_preventiva=("creation_date", "max"),
            ).reset_index())
        else:
            _agg_prev = pd.DataFrame(columns=["eds_occim","preventivas","ultima_preventiva"])

        _df_tbl = _df_tbl.merge(_agg_corr, on="eds_occim", how="left") \
                          .merge(_agg_prev, on="eds_occim", how="left")

        # Rellenar contadores en 0 (EDS sin actividad en el período)
        for _c in ("correctivas","preventivas","p1"):
            if _c in _df_tbl.columns:
                _df_tbl[_c] = _df_tbl[_c].fillna(0).astype(int)

        # ORDEN TOTAL = correctivas + preventivas
        _df_tbl["ordenes_atendidas"] = _df_tbl["correctivas"] + _df_tbl["preventivas"]

        # Última atención con tipo: comparar ultima_correctiva vs ultima_preventiva
        # y quedarnos con la más reciente + etiqueta.
        _uc = pd.to_datetime(_df_tbl.get("ultima_correctiva"), errors="coerce", utc=True)
        _up = pd.to_datetime(_df_tbl.get("ultima_preventiva"), errors="coerce", utc=True)
        if hasattr(_uc, "dt") and _uc.dt.tz is not None:
            _uc = _uc.dt.tz_convert(None)
        if hasattr(_up, "dt") and _up.dt.tz is not None:
            _up = _up.dt.tz_convert(None)

        def _fmt_ult(uc, up):
            if pd.isna(uc) and pd.isna(up):
                return "—"
            if pd.isna(up) or (pd.notna(uc) and uc >= up):
                return f"{uc.strftime('%d/%m/%Y')} · Correctiva"
            return f"{up.strftime('%d/%m/%Y')} · Preventiva"
        _df_tbl["Última atención"] = [_fmt_ult(u, v) for u, v in zip(_uc, _up)]

        # Ratio "prev/corr" en formato compacto por 10 órdenes (redondeado)
        def _fmt_ratio(corr, prev):
            total = int(corr) + int(prev)
            if total == 0:
                return "—"
            p_share = round(int(prev) * 10 / total)
            c_share = 10 - p_share
            return f"{p_share}P / {c_share}C"
        _df_tbl["Ratio (P/C x10)"] = [_fmt_ratio(c, p)
            for c, p in zip(_df_tbl["correctivas"], _df_tbl["preventivas"])]

        # Columnas finales (quitamos Cód. Fracttal — siempre estaba vacío)
        _col_map = {
            "eds_occim":         "Cód. Occim",
            "_loc_code":         "LOC Fracttal",
            "nombre":            "Nombre / Dirección",
            "direccion":         "Dirección",
            "comuna":            "Comuna",
            "zona_occim":        "Zona",
            "region":            "Región",
            "ordenes_atendidas": "Órdenes atendidas",
            "correctivas":       "Correctivas",
            "preventivas":       "Preventivas",
            "p1":                "P1",
            "Última atención":   "Última atención",
            "Ratio (P/C x10)":   "Ratio (P/C x10)",
            "ultimo_tecnico":    "Último Técnico",
        }
        _cols_show = [c for c in _col_map if c in _df_tbl.columns]
        _df_display = _df_tbl[_cols_show].rename(columns=_col_map).copy()

        # Ordenar por Órdenes atendidas desc (donde más pasa antes)
        _sort_col = ("Órdenes atendidas" if "Órdenes atendidas" in _df_display.columns
                     else _df_display.columns[0])
        _df_display = _df_display.sort_values(_sort_col, ascending=False, na_position="last")
        _df_display = _df_display.fillna("—")
        _show_df(_df_display, use_container_width=True, hide_index=True,
            column_config={
                "Cód. Occim":        st.column_config.TextColumn(width=90),
                "LOC Fracttal":      st.column_config.TextColumn(width=95),
                "Nombre / Dirección":st.column_config.TextColumn(width=220),
                "Dirección":         st.column_config.TextColumn(width=200),
                "Comuna":            st.column_config.TextColumn(width=110),
                "Zona":              st.column_config.TextColumn(width=70),
                "Región":            st.column_config.TextColumn(width=70),
                "Órdenes atendidas": st.column_config.NumberColumn(format="%d", width=90,
                    help="Correctivas + Preventivas realizadas en el período seleccionado"),
                "Correctivas":       st.column_config.NumberColumn(format="%d", width=90,
                    help="Llamados correctivos (emergencia) atendidos en el período"),
                "Preventivas":       st.column_config.NumberColumn(format="%d", width=90,
                    help="Mantenciones preventivas realizadas en el período"),
                "P1":                st.column_config.NumberColumn(format="%d", width=60,
                    help="Correctivas de prioridad P1 (máquina detenida)"),
                "Última atención":   st.column_config.TextColumn(width=170,
                    help="Fecha de la última atención (correctiva o preventiva) + tipo"),
                "Ratio (P/C x10)":   st.column_config.TextColumn(width=110,
                    help="De cada 10 órdenes: cuántas fueron Preventivas / Correctivas"),
                "Último Técnico":    st.column_config.TextColumn(width=160),
            })

        # ── Registros relevantes de equipos y consumos (Shell) ──────────
        if company == "SHELL (Enex)":
            st.markdown(
                f'<div style="font-weight:700;font-size:0.95rem;margin:24px 0 8px 0;'
                f'color:{_t["text"]};">'
                f'📋 Registros relevantes de equipos y consumos</div>',
                unsafe_allow_html=True,
            )

            _shell_min_date = pd.Timestamp("2026-06-01", tz="UTC")
            _df_prev_all = (
                df_wo_c[
                    (df_wo_c["maint_type"] == "Preventiva")
                    & (df_wo_c["creation_date"] >= _shell_min_date)
                ]
                .sort_values("creation_date", ascending=False)
                .drop_duplicates(subset=["folio"], keep="first")
            )

            if _df_prev_all.empty:
                st.info("Sin registros de mantención preventiva Shell.")
            else:
                _eds_name_map = {}
                if not df_eds_c.empty:
                    _eds_name_map = dict(zip(
                        df_eds_c["eds_occim"].astype(str),
                        df_eds_c["nombre"],
                    ))

                _fc1, _fc2, _fc3 = st.columns(3)
                with _fc1:
                    _buscar_eds = st.text_input(
                        "Buscar por Código EDS",
                        placeholder="Ej: SH_211",
                        key=f"shell_reg_eds_{_ck}",
                    ).strip().upper()
                with _fc2:
                    _eds_codes = sorted(
                        _df_prev_all["eds_occim"].dropna().astype(str).unique().tolist()
                    )
                    _filter_opts = ["Todas las estaciones"] + [
                        f"{c} — {_eds_name_map.get(c, '')}" for c in _eds_codes
                    ]
                    _sel_filter = st.selectbox(
                        "Filtrar por estación",
                        _filter_opts,
                        key=f"shell_reg_filter_{_ck}",
                    )
                with _fc3:
                    _buscar_ot = st.text_input(
                        "Buscar por N° OT",
                        placeholder="Ej: OS-38200",
                        key=f"shell_reg_ot_{_ck}",
                    ).strip().upper()

                if _buscar_eds:
                    _df_prev_all = _df_prev_all[
                        _df_prev_all["eds_occim"].astype(str).str.upper().str.contains(_buscar_eds, na=False)
                    ]
                if _sel_filter != "Todas las estaciones":
                    _sel_code = _sel_filter.split(" — ")[0]
                    _df_prev_all = _df_prev_all[
                        _df_prev_all["eds_occim"].astype(str) == _sel_code
                    ]
                if _buscar_ot:
                    _df_prev_all = _df_prev_all[
                        _df_prev_all["folio"].str.upper().str.contains(_buscar_ot, na=False)
                    ]

                def _rpval(s, col, default="—"):
                    if s.empty or col not in s.columns:
                        return default
                    v = s.iloc[0][col]
                    if v is None or (isinstance(v, float) and pd.isna(v)):
                        return default
                    sv = str(v).strip()
                    return default if sv in ("", "None", "nan", "null") else sv

                def _rphora(df_sub, col):
                    if df_sub.empty or col not in df_sub.columns:
                        return "—"
                    vals = pd.to_datetime(df_sub[col], errors="coerce").dropna()
                    if vals.empty:
                        return "—"
                    ts = vals.min() if "inicio" in col else vals.max()
                    try:
                        if ts.tz is not None:
                            ts = ts.tz_convert("America/Santiago")
                        return ts.strftime("%H:%M")
                    except Exception:
                        return "—"

                # Pre-computar "N. Final anterior" por estación recorriendo
                # TODAS las preventivas Shell (sin filtros) cronológicamente.
                _df_prev_full = (
                    df_wo_c[df_wo_c["maint_type"] == "Preventiva"]
                    .sort_values("creation_date", ascending=True)
                    .drop_duplicates(subset=["folio"], keep="first")
                )
                _prev_final_lav: dict = {}
                _prev_final_asp: dict = {}
                _last_by_eds: dict = {}
                for _, _row in _df_prev_full.iterrows():
                    _f = _row["folio"]
                    _e = str(_row.get("eds_occim", ""))
                    _prev_final_lav[_f] = _last_by_eds.get((_e, "lavadora"), "—")
                    _prev_final_asp[_f] = _last_by_eds.get((_e, "aspiradora"), "—")
                    _sub_f = df_num_sub_eds[df_num_sub_eds["id_ot"] == _f]
                    _lav_f = _sub_f[_sub_f["tipo_activo"] == "lavadora"]
                    _asp_f = _sub_f[_sub_f["tipo_activo"] == "aspiradora"]
                    _lv = _rpval(_lav_f, "numeral_final", default=None)
                    _av = _rpval(_asp_f, "numeral_final", default=None)
                    if _lv is not None:
                        _last_by_eds[(_e, "lavadora")] = _lv
                    if _av is not None:
                        _last_by_eds[(_e, "aspiradora")] = _av

                _reg_rows = []
                for _, _ot in _df_prev_all.head(100).iterrows():
                    _fol = _ot["folio"]
                    _sub = df_num_sub_eds[df_num_sub_eds["id_ot"] == _fol]
                    _lav = _sub[_sub["tipo_activo"] == "lavadora"]
                    _asp = _sub[_sub["tipo_activo"] == "aspiradora"]

                    _cd = _ot["creation_date"]
                    _fecha = _cd.strftime("%d/%m/%Y") if pd.notna(_cd) else "—"
                    _tec = str(_ot.get("technician", "")).strip()
                    if _tec in ("", "nan", "None"):
                        _tec = "—"
                    _ec = str(_ot.get("eds_occim", "—"))

                    _reg_rows.append({
                        ("Datos Estación", "Código EDS"):               _ec,
                        ("Datos Estación", "Nombre Estación"):          _eds_name_map.get(_ec, "—"),
                        ("Datos OT", "N° OT"):                         _fol,
                        ("Datos OT", "Fecha"):                         _fecha,
                        ("Datos OT", "Hora inicio"):                   _rphora(_sub, "fecha_inicio_subtarea"),
                        ("Datos OT", "Hora término"):                  _rphora(_sub, "fecha_fin_subtarea"),
                        ("Datos OT", "Técnico"):                       _tec,
                        ("Numeral lavadora", "Anterior"):              _prev_final_lav.get(_fol, "—") or "—",
                        ("Numeral lavadora", "Actual"):                _rpval(_lav, "numeral_final"),
                        ("Numeral aspiradora", "Anterior"):            _prev_final_asp.get(_fol, "—") or "—",
                        ("Numeral aspiradora", "Actual"):              _rpval(_asp, "numeral_final"),
                        ("Fichas mantención", "Lavado"):               _rpval(_lav, "fichas_periodo"),
                        ("Fichas mantención", "Aspirado"):             _rpval(_asp, "fichas_periodo"),
                        ("Insumos", "Bomba dosificadora"):             _rpval(_lav, "bomba_dosificadora"),
                        ("Insumos", "Consumo (%)"):                    _rpval(_lav, "consumo_insumos"),
                        ("Tiempo fichas (seg)", "Lavado"):             _rpval(_lav, "tiempo_fichas_seg"),
                        ("Tiempo fichas (seg)", "Aspirado"):           _rpval(_asp, "tiempo_fichas_seg"),
                    })

                if _reg_rows:
                    _df_reg = pd.DataFrame(_reg_rows)
                    _df_reg.columns = pd.MultiIndex.from_tuples(_df_reg.columns)
                    _show_df(_df_reg, use_container_width=True, hide_index=True)
                else:
                    st.info("Sin registros para el filtro seleccionado.")

    # ════════════════════════════════════════════════════════════════════════
    # PESTAÑA: HISTORIAL DE NUMERALES (global, agrupado por EDS, con buscador)
    # ════════════════════════════════════════════════════════════════════════
    def _render_numeral_tab():
        st.caption(
            "Historial de numerales por EDS (**lavadoras y aspiradoras**) — los últimos "
            "10 registros por equipo. Validación en dos niveles: dentro de cada OT "
            "(Fichas = Final − Inicial) y entre OTs (el contador solo puede subir; "
            "si el inicial baja o salta ×5 respecto al final previo, es un error arrastrado)."
        )

        _hist = build_numeral_historial(df_wo_eds_full, eds_code=None, n=None)
        if _hist.empty:
            st.info("Aún no hay registros de numeral. Corre el sync de numerales para poblar los datos.")
            return
        _seq = analizar_secuencias(_hist, n=10)

        # ── Lookup de metadatos EDS (eds_occim → nombre, loc, macrozona) ─────────
        # Fuente primaria: zona_occim ya trae "Norte"/"Sur"/"Santiago" directamente.
        # Fallback: número romano de región para estaciones sin zona asignada (SHELL, RP, RT…).
        _ZONA_DIRECTA = {
            "Norte":          "Norte",
            "Centro III y IV": "Norte",   # Atacama (III) + Coquimbo (IV)
            "Sur":            "Sur",
            "Santiago":       "Santiago",
            "Centro":         "Santiago", # Mayoría en RM y V (Valparaíso)
        }
        _NORTE_REG = {"I", "II", "III", "IV", "XV"}
        _SUR_REG   = {"VII", "VIII", "IX", "X", "XI", "XII", "XIV", "XVI",
                      "LOS LAGOS", "LOS RIOS"}
        _eds_meta = {}
        if not df_eds.empty and "eds_occim" in df_eds.columns:
            for _, _row in df_eds.iterrows():
                _mk = str(_row.get("eds_occim", "") or "").strip()
                if _mk:
                    _zo = str(_row.get("zona_occim", "") or "").strip()
                    if _zo in _ZONA_DIRECTA:
                        _mz = _ZONA_DIRECTA[_zo]
                    else:
                        _reg = str(_row.get("region", "") or "").strip().upper()
                        if _reg in _NORTE_REG:
                            _mz = "Norte"
                        elif _reg in _SUR_REG:
                            _mz = "Sur"
                        else:
                            _mz = "Santiago"
                    _eds_meta[_mk] = {
                        "nombre":    str(_row.get("nombre", "") or ""),
                        "loc":       str(_row.get("_loc_code", "") or ""),
                        "macrozona": _mz,
                    }

        # ── Período ────────────────────────────────────────────────────────────
        _f = pd.to_datetime(_seq["fecha"], errors="coerce", utc=True).dt.tz_convert(None)
        _seq["_year"] = _f.dt.year
        _seq["_mes"]  = _f.dt.month
        _seq["_trim"] = _f.dt.quarter
        # Macrozona por equipo (via eds_occim → lookup)
        _seq["_zona"] = _seq["eds_occim"].apply(
            lambda eo: _eds_meta.get(str(eo or "").strip(), {}).get("macrozona", "")
        )

        # ── Filtros fila 1 ─────────────────────────────────────────────────────
        _fc1, _fc2, _fc3, _fc4, _fc5 = st.columns([0.8, 0.8, 1.2, 1.5, 1.0])
        with _fc1:
            _yrs = sorted([int(y) for y in _seq["_year"].dropna().unique()], reverse=True)
            _yr_sel = st.selectbox("Año", _yrs, index=0, key="num_tab_year")
        with _fc2:
            _trim_sel = st.selectbox("Trimestre", ["Todos","T1","T2","T3","T4"], key="num_tab_trim")
        with _fc3:
            _meses_en_yr = sorted(_seq[_seq["_year"] == _yr_sel]["_mes"].dropna().unique())
            _mes_sel = st.multiselect("Mes (opcional)", options=[int(m) for m in _meses_en_yr],
                format_func=lambda m: _MES_ES[int(m)], key="num_tab_mes")
        with _fc4:
            _cli_sel = st.multiselect("Cliente",
                options=["COPEC","Aramco (Esmax)","SHELL (Enex)"], key="num_tab_cli")
        with _fc5:
            _zona_sel = st.multiselect("Región",
                options=["Norte","Santiago","Sur"], key="num_tab_zona",
                help="Norte: Coquimbo y arriba · Santiago: RM + Valparaíso + O'Higgins · Sur: Maule y abajo")

        _sit_opts  = ["normal","raro","anomalo","incongruente","salto_seq"]
        _sit_label = {**CAT_LABEL, "salto_seq": "🟣 Salto de secuencia (entre OTs)"}
        _sit_sel   = st.multiselect("Situación a revisar", options=_sit_opts,
            format_func=lambda c: _sit_label.get(c, c),
            default=["raro","anomalo","incongruente","salto_seq"], key="num_tab_sit",
            help="Un equipo aparece si ALGUNO de sus últimos 10 registros cae en estas situaciones.")

        # ── Buscador de EDS ────────────────────────────────────────────────────
        _busq = st.text_input(
            "Buscar EDS",
            placeholder="Código (60079)  ·  LOC (LOC-074)  ·  Nombre (Costanera, Vitacura, Portugal…)",
            key="num_tab_busq",
        )

        # ── Scope período + cliente + región ──────────────────────────────────
        _scope = _seq[_seq["_year"] == _yr_sel].copy()
        if _trim_sel != "Todos":
            _scope = _scope[_scope["_trim"] == int(_trim_sel[1])]
        if _mes_sel:
            _scope = _scope[_scope["_mes"].isin(_mes_sel)]
        if _cli_sel:
            _scope = _scope[_scope["client"].isin(_cli_sel)]
        if _zona_sel:
            _scope = _scope[_scope["_zona"].isin(_zona_sel)]
        _codes_scope = set(_scope["equipment_code"].unique())

        _disp_seq = _seq[_seq["equipment_code"].isin(_codes_scope)].copy()
        if _disp_seq.empty:
            st.info("Sin equipos para los filtros seleccionados.")
            return

        _disp_seq["_match"] = (
            _disp_seq["categoria"].isin([c for c in _sit_sel if c != "salto_seq"]) |
            ((_disp_seq["seq_severidad"] == "alert") if "salto_seq" in _sit_sel else False)
        )

        # ── KPI cards ─────────────────────────────────────────────────────────
        _k1, _k2, _k3, _k4, _k5 = st.columns(5)
        _k1.metric("Equipos",         len(_codes_scope))
        _k2.metric("🟡 Raro",         int((_scope["categoria"] == "raro").sum()))
        _k3.metric("🔴 Anómalo",      int((_scope["categoria"] == "anomalo").sum()))
        _k4.metric("🟣 Incongruente", int((_scope["categoria"] == "incongruente").sum()))
        _k5.metric("🟣 Saltos sec.",  int((_disp_seq["seq_severidad"] == "alert").sum()))

        _codes_match = set(_disp_seq[_disp_seq["_match"]]["equipment_code"].unique())
        if not _codes_match:
            st.success("✅ Ningún equipo con registros en la situación seleccionada para estos filtros.")
            return

        # ── Agrupar equipos por EDS ────────────────────────────────────────────
        _eds_groups = {}
        for _code in _codes_match:
            _grp = _disp_seq[_disp_seq["equipment_code"] == _code]
            _eo  = str(_grp["eds_occim"].iloc[0] if "eds_occim" in _grp.columns else "").strip()
            _stn = str(_grp["station"].iloc[0])
            _ekey = _eo if _eo else _stn
            if _ekey not in _eds_groups:
                _meta = _eds_meta.get(_eo, {})
                _eds_groups[_ekey] = {
                    "nombre":    _meta.get("nombre") or _stn,
                    "loc":       _meta.get("loc", ""),
                    "macrozona": _meta.get("macrozona", ""),
                    "eds_occim": _eo,
                    "station":   _stn,
                    "client":    str(_grp["client"].iloc[0]),
                    "equipos":   [],
                    "n_alerta":  0,
                }
            _na = (
                int((_grp["seq_severidad"] == "alert").sum()) +
                int(_grp["categoria"].isin(["anomalo","incongruente"]).sum())
            )
            _eds_groups[_ekey]["equipos"].append((_na, _code))
            _eds_groups[_ekey]["n_alerta"] += _na

        # ── Aplicar buscador ───────────────────────────────────────────────────
        if _busq.strip():
            _q = _busq.strip().upper()
            _eds_groups = {
                k: v for k, v in _eds_groups.items()
                if (_q in str(v["eds_occim"]).upper()
                    or _q in str(v["loc"]).upper()
                    or _q in str(v["nombre"]).upper()
                    or _q in str(v["station"]).upper())
            }
        if not _eds_groups:
            st.info("Sin resultados. Prueba con otro código EDS, LOC o nombre de estación.")
            return

        _n_eds_total = len(_eds_groups)
        _n_eq_total  = sum(len(v["equipos"]) for v in _eds_groups.values())
        st.markdown(
            f'<div style="font-weight:700;font-size:0.92rem;margin:10px 0 4px 0;'
            f'color:{_t["text"]};">'
            f'{_n_eds_total} EDS con alertas · {_n_eq_total} equipo(s) a revisar</div>',
            unsafe_allow_html=True,
        )

        # ── Renderizar por EDS → equipos ──────────────────────────────────────
        _ZONA_COLOR = {"Norte": "#F59E0B", "Santiago": "#3B82F6", "Sur": "#10B981"}
        for _ekey, _einfo in sorted(_eds_groups.items(), key=lambda x: -x[1]["n_alerta"]):
            _loc_txt  = f"&nbsp;·&nbsp;{_einfo['loc']}"       if _einfo["loc"]      else ""
            _code_txt = f"&nbsp;·&nbsp;{_einfo['eds_occim']}" if _einfo["eds_occim"] else ""
            _n_eq_eds = len(_einfo["equipos"])
            _mz       = _einfo.get("macrozona", "")
            _mz_color = _ZONA_COLOR.get(_mz, "#888")
            _mz_txt   = (f'<span style="background:{_mz_color}22;color:{_mz_color};'
                         f'border-radius:4px;padding:1px 7px;font-size:0.75rem;'
                         f'font-weight:600;margin-left:8px;">{_mz}</span>') if _mz else ""
            st.markdown(
                f'<div style="background:{_t["card"]};border:1px solid {_t["border"]};'
                f'border-radius:8px;padding:8px 14px;margin:14px 0 2px 0;">'
                f'<span style="font-weight:700;font-size:0.95rem;">'
                f'{_einfo["nombre"] or _einfo["station"]}</span>'
                f'{_mz_txt}'
                f'<span style="font-size:0.80rem;color:#888;margin-left:8px;">'
                f'{_code_txt}{_loc_txt}&nbsp;·&nbsp;{_einfo["client"]}</span>'
                f'<span style="float:right;font-size:0.82rem;color:#888;">'
                f'{_einfo["n_alerta"]} alerta(s)&nbsp;·&nbsp;{_n_eq_eds} equipo(s)</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
            for _na, _code in sorted(_einfo["equipos"], key=lambda x: -x[0]):
                _grp = _disp_seq[_disp_seq["equipment_code"] == _code].sort_values("fecha", ascending=False)
                _eq_nombre = _grp["equipment"].iloc[0]
                _n_seq   = int((_grp["seq_severidad"] == "alert").sum())
                _n_intra = int(_grp["categoria"].isin(["anomalo","incongruente"]).sum())
                _badge = ""
                if _n_seq:   _badge += f"  🟣 {_n_seq} salto(s)"
                if _n_intra: _badge += f"  🔴 {_n_intra} intra-OT"
                with st.expander(f"{_eq_nombre}  ({_code}){_badge}",
                                 expanded=bool(_n_seq or _n_intra)):
                    _d = _grp.copy()
                    _d["Fecha"] = pd.to_datetime(_d["fecha"], errors="coerce").dt.strftime("%d/%m/%Y").fillna("—")
                    _d["Salto secuencia"] = _d["salto_seq"].replace("", "✅ OK")

                    # Detectar el primer registro donde comenzó el error (cronológico).
                    # IMPORTANTE: usar _grp["fecha"] (datetime), NO _d["Fecha"] (string DD/MM/YYYY
                    # que ordena lexicalmente mal: "03/04" < "17/03" por el día "03" < "17").
                    _origen_folio = None
                    for _, _or in _grp.sort_values("fecha").iterrows():
                        _seq_alert = _or.get("seq_severidad") == "alert"
                        _intra_mal = _or.get("categoria") in ("anomalo","incongruente")
                        if _seq_alert or _intra_mal:
                            _origen_folio = _or.get("folio")
                            break
                    _d["Origen"] = _d["folio"].apply(
                        lambda f: "🔴 Origen del error" if f == _origen_folio and _origen_folio else ""
                    )

                    if "comentario_tecnico" in _d.columns:
                        _d["comentario_tecnico"] = _d["comentario_tecnico"].fillna("").apply(_strip_comentario_headers)
                    _cols_d = ["Fecha","folio","technician","numeral_inicial","numeral_final",
                               "estado","Salto secuencia","Origen","comentario_tecnico"]
                    _d = _d[[c for c in _cols_d if c in _d.columns]].rename(columns={
                        "folio":           "OT",
                        "technician":      "Técnico",
                        "numeral_inicial": "N. Inicial",
                        "numeral_final":   "N. Final",
                        "estado":          "Fichas (intra-OT)",
                        "comentario_tecnico": "Comentario técnico / causa raíz",
                    })
                    _show_df(_d, use_container_width=True, hide_index=True,
                        column_config={
                            "Fecha":      st.column_config.TextColumn(width=85),
                            "OT":         st.column_config.TextColumn(width=90),
                            "Técnico":    st.column_config.TextColumn(width=130),
                            "N. Inicial": st.column_config.TextColumn(width=85),
                            "N. Final":   st.column_config.TextColumn(width=85),
                            "Fichas (intra-OT)": st.column_config.TextColumn(width=165),
                            "Salto secuencia":   st.column_config.TextColumn(width=210,
                                help="Compara el inicial de esta visita con el final de la anterior."),
                            "Origen": st.column_config.TextColumn(width=155,
                                help="Primera OT donde el contador dejó de ser coherente."),
                            "Comentario técnico / causa raíz": st.column_config.TextColumn(width=320,
                                help="Texto libre del técnico en Fracttal (falla, trabajo, observaciones). Vacío = no documentó."),
                        })

    # ── Renderizar tabs ──────────────────────────────────────────────────────
    with _tabs_eds[0]:
        _render_eds_tab("COPEC", 0)
    with _tabs_eds[1]:
        _render_eds_tab("Aramco (Esmax)", 1)
    with _tabs_eds[2]:
        _render_eds_tab("SHELL (Enex)", 2)
    with _tabs_eds[3]:
        _render_numeral_tab()



# ─────────────────────────────────────────────────────────────────────────────
# PÁGINA 4: UTILIZACIÓN DEL TIEMPO
# ─────────────────────────────────────────────────────────────────────────────
elif _page == _NAV_PAGES[4]:
    _hdr(_PAGE_TITLE[_NAV_PAGES[4]])
    # ── Sub-tabs ──────────────────────────────────────────────────────────────
    _util_sub_tab = st.radio(
        "",
        ["📅 Programación STO", "📅 Planificación Turnos", "📡 En Vivo", "📊 Utilización del tiempo"],
        horizontal=True,
        label_visibility="collapsed",
        key="util_sub_tab",
    )
    st.divider()

    # ─────────────────────────────────────────────────────────────────────
    # SUB-TAB: 📅 Programación STO — grid de programación diaria por
    # técnico leyendo directo el Excel '2026 UTILIZACIÓN DE TIEMPO.xlsx'
    # con colores replicados (verde=MP regional, rojo=vacaciones,
    # naranja=feriado, blanco=actividad normal).
    # ─────────────────────────────────────────────────────────────────────
    if _util_sub_tab == "📅 Programación STO":
        from openpyxl import load_workbook as _lwb
        import calendar as _cal_sto
        from datetime import date as _date_sto

        st.title("📅 Programación STO — Cronograma diario de técnicos")
        st.caption(
            "Cronograma leído directo del Excel `2026 UTILIZACIÓN DE TIEMPO.xlsx` en Google Drive. "
            "**🟢 verde** = MP fuera de Santiago · "
            "**🔴 rojo** = vacaciones / día libre · "
            "**🟠 naranja** = feriado · "
            "**⬜ blanco** = actividad normal (turno / oficina / MP local)."
        )

        _PATH_STO = ("G:/.shortcut-targets-by-id/15zHnoU5VZlkOwYc6EBziNcnS-sAAtwtk/"
                     "OPERACIONES/OPERACIONES/2026 UTILIZACIÓN DE TIEMPO.xlsx")

        @st.cache_data(ttl=300, show_spinner=False)
        def _load_grid_sto(mes_hoja: str):
            """Lee la hoja del Excel y devuelve (df, dict_colores).
            df: filas=días del mes, cols=técnicos, celdas=descripción de actividad
            dict_colores: {(fecha, tecnico): hex_color}"""
            try:
                _wb = _lwb(_PATH_STO, data_only=True)
            except Exception as e:
                return None, None, str(e)
            if mes_hoja not in _wb.sheetnames:
                return None, None, f"Hoja '{mes_hoja}' no existe"
            _ws = _wb[mes_hoja]
            # Fila 1: encabezados (col A=Fecha, B-R=técnicos)
            _tecnicos_cols = {}   # {col_letter_int: nombre_tecnico}
            for c in range(2, min(19, _ws.max_column + 1)):
                v = _ws.cell(1, c).value
                if v and str(v).strip():
                    _tecnicos_cols[c] = str(v).strip()
            # Filas de datos: buscar todas las filas con Fecha en col A
            _rows_data = []
            _colores = {}
            for r in range(2, _ws.max_row + 1):
                _fecha = _ws.cell(r, 1).value
                if not _fecha or not hasattr(_fecha, "year"):
                    continue
                _fila = {"_fecha": _fecha}
                for c, tec in _tecnicos_cols.items():
                    cell = _ws.cell(r, c)
                    _fila[tec] = cell.value or ""
                    if cell.fill and cell.fill.start_color:
                        try:
                            rgb = cell.fill.start_color.rgb
                            if rgb and isinstance(rgb, str) and rgb not in ("00000000","FFFFFFFF"):
                                _colores[(_fecha.date(), tec)] = "#" + rgb[-6:]
                        except Exception:
                            pass
                _rows_data.append(_fila)
            _df = pd.DataFrame(_rows_data)
            return _df, _colores, None

        # ── Filtros ────────────────────────────────────────────────────
        _MESES_STO = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
                      "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]
        _yr_actual = _date_sto.today().year
        _mes_actual = _date_sto.today().month
        _hojas_disp = [f"{_MESES_STO[m-1]} {_yr_actual}"
                       for m in range(_mes_actual, 0, -1)] + \
                      [f"{_MESES_STO[m-1]} {_yr_actual-1}"
                       for m in range(12, 0, -1)]

        _fc1, _fc2, _fc3, _fc4 = st.columns([2, 1.5, 1.5, 1])
        with _fc1:
            _mes_sel = st.selectbox("Mes", _hojas_disp, key="sto_mes_sel")
        with _fc2:
            _vista = st.selectbox("Vista", ["Mes completo","Semana","Día"],
                                  key="sto_vista")
        with _fc4:
            st.write("")
            if st.button("🔄 Refrescar", key="sto_refresh",
                         use_container_width=True):
                _load_grid_sto.clear()
                st.rerun()

        # Cargar datos
        with st.spinner(f"Leyendo {_mes_sel}…"):
            _df_sto, _colores_sto, _err_sto = _load_grid_sto(_mes_sel)

        if _err_sto:
            st.error(f"No se pudo leer el archivo: {_err_sto}")
        elif _df_sto is None or _df_sto.empty:
            st.warning(f"Sin datos en la hoja {_mes_sel}.")
        else:
            # Tercer filtro: técnico (dropdown poblado con los que existen)
            _tec_opts = ["Todos"] + [c for c in _df_sto.columns if c != "_fecha"]
            with _fc3:
                _tec_sel = st.selectbox("Técnico", _tec_opts, key="sto_tec_sel")

            _df_vista = _df_sto.copy()

            # Aplicar filtro de vista (semana / día)
            _hoy_norm = pd.Timestamp.today().normalize()
            if _vista == "Semana":
                # Selector de semana ISO
                _df_vista["_iso_wk"] = pd.to_datetime(_df_vista["_fecha"]).dt.isocalendar().week
                _semanas = sorted(_df_vista["_iso_wk"].unique().tolist())
                _sem_lbl = [f"Semana {w}" for w in _semanas]
                _sem_pick = st.selectbox("Semana", _sem_lbl, key="sto_sem_pick")
                _sem_n = int(_sem_pick.split()[1])
                _df_vista = _df_vista[_df_vista["_iso_wk"] == _sem_n].drop(columns="_iso_wk")
            elif _vista == "Día":
                _fechas_disp = pd.to_datetime(_df_vista["_fecha"]).dt.date.tolist()
                _dia_pick = st.selectbox("Día", _fechas_disp,
                    format_func=lambda d: d.strftime("%A %d/%m/%Y"),
                    key="sto_dia_pick")
                _df_vista = _df_vista[
                    pd.to_datetime(_df_vista["_fecha"]).dt.date == _dia_pick
                ]

            # Filtrar por técnico
            if _tec_sel != "Todos":
                _cols_keep = ["_fecha", _tec_sel]
                _df_vista = _df_vista[[c for c in _cols_keep if c in _df_vista.columns]]

            if _df_vista.empty:
                st.info("No hay registros para el filtro seleccionado.")
            else:
                # ── KPIs por técnico (solo si se ven varios) ──────────
                if _tec_sel == "Todos":
                    _n_dias = len(_df_vista)
                    _tecs = [c for c in _df_vista.columns if c != "_fecha"]
                    st.markdown(f"**{_n_dias} día(s) · {len(_tecs)} técnicos**")

                    # Contar por técnico: vacaciones, feriados, MP regional
                    _resumen = []
                    for _tec in _tecs:
                        _n_vac = _n_fer = _n_mp = _n_norm = 0
                        for _, _row in _df_vista.iterrows():
                            _d = pd.to_datetime(_row["_fecha"]).date()
                            _col = _colores_sto.get((_d, _tec), "").upper()
                            _val = str(_row.get(_tec, "") or "").upper()
                            if _col == "#FF0000" or "VACACIONES" in _val:
                                _n_vac += 1
                            elif _col == "#FFC000" or "FERIADO" in _val:
                                _n_fer += 1
                            elif _col == "#00FF00":
                                _n_mp += 1
                            elif _val.strip():
                                _n_norm += 1
                        _resumen.append({
                            "Técnico": _tec, "Días activos": _n_norm,
                            "🟢 MP regional": _n_mp,
                            "🔴 Vacaciones": _n_vac,
                            "🟠 Feriados": _n_fer,
                        })
                    _df_resumen = pd.DataFrame(_resumen)
                    with st.expander("📊 Resumen por técnico (conteos)", expanded=False):
                        _show_df(_df_resumen, hide_index=True, use_container_width=True)

                # ── Grid principal con colores ────────────────────────
                # Preparar DataFrame para display
                _df_disp = _df_vista.copy()
                _df_disp["Fecha"] = pd.to_datetime(_df_disp["_fecha"]).apply(
                    lambda d: f"{['LUN','MAR','MIÉ','JUE','VIE','SÁB','DOM'][d.weekday()]} "
                              f"{d.strftime('%d/%m')}"
                )
                _df_disp = _df_disp.drop(columns="_fecha")
                _cols_final_sto = ["Fecha"] + [c for c in _df_disp.columns if c != "Fecha"]
                _df_disp = _df_disp[_cols_final_sto]

                # Aplicar estilos con Styler
                def _estilo_celda(v, fecha_str, tec):
                    """Color de fondo basado en el dict de colores del Excel"""
                    # Extraer fecha del formato "LUN 30/06"
                    try:
                        _dd_mm = fecha_str.split(" ")[-1]
                        _mes_num_e = _MESES_STO.index(_mes_sel.split()[0]) + 1
                        _yr_e = int(_mes_sel.split()[1])
                        _d, _m = _dd_mm.split("/")
                        _fe = _date_sto(_yr_e, int(_m), int(_d))
                    except Exception:
                        return ""
                    _clr = _colores_sto.get((_fe, tec), "").upper()
                    _v_up = str(v or "").upper()
                    # Overrides por texto (si no hay color explícito)
                    if _clr == "#FF0000" or "VACACIONES" in _v_up:
                        return "background-color: #fecaca; color: #7f1d1d; font-weight:600"
                    if _clr == "#FFC000" or "FERIADO" in _v_up:
                        return "background-color: #fed7aa; color: #7c2d12; font-weight:600"
                    if _clr == "#00FF00":
                        return "background-color: #bbf7d0; color: #14532d; font-weight:500"
                    return ""

                def _apply_styles(row):
                    """Aplica estilos por celda leyendo la fecha de la col 'Fecha'"""
                    _fecha_str = row["Fecha"]
                    return [""] + [_estilo_celda(row[c], _fecha_str, c)
                                   for c in row.index if c != "Fecha"]

                _styled = _df_disp.style.apply(_apply_styles, axis=1)
                st.dataframe(_styled, use_container_width=True, hide_index=True,
                             height=min(600, 40 * len(_df_disp) + 50))

                st.caption(
                    f"📖 Fuente: `{_PATH_STO.split('/')[-1]}` hoja **{_mes_sel}** · "
                    f"cache 5 min · botón 🔄 Refrescar para forzar recarga."
                )

        st.stop()

    if _util_sub_tab == "📡 En Vivo":
        from datetime import datetime as _dt_vivo

        # ── Título + botón actualizar ─────────────────────────────────────────────
        _col_tit, _col_btn = st.columns([6, 1])
        with _col_tit:
            st.title("📡 En Vivo — Órdenes en Ejecución")
        with _col_btn:
            st.write("")
            if st.button("🔄 Actualizar", key="btn_envivo_refresh", use_container_width=True):
                load_ots_en_vivo_supabase.clear()
                st.rerun()

        _vivo_ts = _dt_vivo.now().strftime("%d/%m/%Y %H:%M")
        st.caption(
            f"Última actualización: **{_vivo_ts}** · "
            "caché renovado cada **2 min** automáticamente · "
            "incluye OTs *En Progreso*, *Por Validar* y *Por Iniciar*"
        )
        st.divider()

        # ── Carga de datos ────────────────────────────────────────────────────────
        with st.spinner("Cargando órdenes en curso…"):
            _raw_vivo = load_ots_en_vivo_supabase()

        if not _raw_vivo:
            st.info("✅ No hay órdenes activas en este momento.")
        else:
            _df_vivo = pd.DataFrame(_raw_vivo)

            # ── Helpers internos ─────────────────────────────────────────────────
            def _vivo_seg_fmt(s):
                try:
                    s = int(float(s or 0))
                    if s <= 0:
                        return "—"
                    h, rem = divmod(s, 3600)
                    m = rem // 60
                    return f"{h}h {m:02d}m" if h else f"{m}m"
                except Exception:
                    return "—"

            def _vivo_avance(row):
                try:
                    r = float(row.get("duracion_real_seg") or 0)
                    e = float(row.get("duracion_estim_seg") or 0)
                    if e <= 0:
                        return None
                    return min(round(r / e * 100, 1), 999.0)
                except Exception:
                    return None

            def _vivo_fmt_dt(x):
                if not x or str(x).strip() in ("", "None", "null"):
                    return "—"
                try:
                    t = pd.Timestamp(str(x))
                    if t.tzinfo:
                        t = t.tz_convert(None)
                    return t.strftime("%d/%m %H:%M")
                except Exception:
                    return "—"

            def _vivo_limpiar_ubi(s):
                """'// COPEC/ COPEC LAMPA/' → 'COPEC LAMPA'"""
                if not s or str(s).strip() in ("", "None"):
                    return "—"
                partes = [p.strip() for p in str(s).split("/") if p.strip()]
                return partes[-1] if partes else str(s)

            # ── Normalizar columnas ───────────────────────────────────────────────
            for _col in ["estado", "estado_tarea", "tipo_tarea", "responsable", "cliente"]:
                if _col in _df_vivo.columns:
                    _df_vivo[_col] = _df_vivo[_col].fillna("—")

            # Traducir valores en inglés de Fracttal → español
            _TAREA_MAP_VIVO = {
                "DONE":         "Finalizada",
                "NO_STARTED":   "No Iniciada",
                "IN_PROGRESS":  "En Proceso",
                "IN_REVIEW":    "En Revisión",
                "ON_HOLD":      "En Espera",
                "WAITING":      "En Espera",
            }
            if "estado_tarea" in _df_vivo.columns:
                _df_vivo["estado_tarea"] = _df_vivo["estado_tarea"].replace(_TAREA_MAP_VIVO)

            # T.Prog = duración planificada de la tarea (lo que Fracttal llama "duración")
            # T.Real de Fracttal = mismo valor que T.Prog (no hay tracking real por API)
            # → se elimina T.Real y se agrega T.Transcurrido calculado desde fecha_inicio
            _df_vivo["t_prog"]      = _df_vivo["duracion_estim_seg"].apply(_vivo_seg_fmt)

            def _vivo_transcurrido(row):
                """Tiempo transcurrido desde inicio hasta fin (o ahora si sigue abierta)."""
                try:
                    t_ini = pd.Timestamp(str(row.get("fecha_inicio") or ""))
                    if pd.isna(t_ini):
                        return "—"
                    t_fin_raw = row.get("fecha_finalizacion")
                    if t_fin_raw and str(t_fin_raw).strip() not in ("", "None", "null"):
                        t_fin = pd.Timestamp(str(t_fin_raw))
                    else:
                        t_fin = pd.Timestamp.utcnow()
                    if t_ini.tzinfo:
                        t_ini = t_ini.tz_localize(None)
                    if t_fin.tzinfo:
                        t_fin = t_fin.tz_localize(None)
                    secs = int((t_fin - t_ini).total_seconds())
                    if secs < 0:
                        return "—"
                    h, rem = divmod(secs, 3600)
                    m = rem // 60
                    return f"{h}h {m:02d}m" if h else f"{m}m"
                except Exception:
                    return "—"

            _df_vivo["t_transcurrido"] = _df_vivo.apply(_vivo_transcurrido, axis=1)

            def _vivo_avance_prog(row):
                """Avance = t_transcurrido / t_programado × 100."""
                try:
                    e = float(row.get("duracion_estim_seg") or 0)
                    if e <= 0:
                        return None
                    t_ini = pd.Timestamp(str(row.get("fecha_inicio") or ""))
                    if pd.isna(t_ini):
                        return None
                    t_fin_raw = row.get("fecha_finalizacion")
                    if t_fin_raw and str(t_fin_raw).strip() not in ("", "None", "null"):
                        t_fin = pd.Timestamp(str(t_fin_raw))
                    else:
                        t_fin = pd.Timestamp.utcnow()
                    if t_ini.tzinfo:
                        t_ini = t_ini.tz_localize(None)
                    if t_fin.tzinfo:
                        t_fin = t_fin.tz_localize(None)
                    r = max((t_fin - t_ini).total_seconds(), 0)
                    return min(round(r / e * 100, 1), 999.0)
                except Exception:
                    return None

            _df_vivo["avance_pct"]  = _df_vivo.apply(_vivo_avance_prog, axis=1)
            _df_vivo["inicio_fmt"]  = _df_vivo["fecha_inicio"].apply(_vivo_fmt_dt)
            _df_vivo["creacion_fmt"]= _df_vivo["fecha_creacion"].apply(_vivo_fmt_dt)
            _df_vivo["ubi_limpia"]  = _df_vivo["ubicacion"].apply(_vivo_limpiar_ubi)

            _ESTADO_T_ICON = {
                "En Proceso":  "🟢 En Proceso",
                "En Revisión": "🟡 En Revisión",
                "No Iniciada": "⚫ No Iniciada",
                "En Espera":   "🔵 En Espera",
                "Finalizada":  "✅ Finalizada",
            }
            _df_vivo["estado_tarea_lbl"] = _df_vivo["estado_tarea"].map(
                lambda x: _ESTADO_T_ICON.get(x, f"⚪ {x}")
            )

            # ── KPI cards ─────────────────────────────────────────────────────────
            _n_prog = int((_df_vivo["estado"] == "En Progreso").sum())
            _n_val  = int((_df_vivo["estado"] == "Por Validar").sum())
            _n_ini  = int((_df_vivo["estado"] == "Por Iniciar").sum())
            _n_tec  = int(
                _df_vivo.loc[_df_vivo["estado"] == "En Progreso", "responsable"]
                .replace("—", pd.NA).dropna().nunique()
            )
            _n_cm   = int(_df_vivo["tipo_tarea"].str.contains("CORREC", na=False, case=False).sum())
            _n_pm   = int(_df_vivo["tipo_tarea"].str.contains("PREVEN", na=False, case=False).sum())

            _kc1, _kc2, _kc3, _kc4, _kc5, _kc6 = st.columns(6)
            _kc1.metric("🔧 En Progreso", _n_prog)
            _kc2.metric("🔍 Por Validar", _n_val)
            _kc3.metric("⏳ Por Iniciar", _n_ini)
            _kc4.metric("👷 Técnicos activos", _n_tec)
            _kc5.metric("🚨 Correctivas", _n_cm)
            _kc6.metric("🛠️ Preventivas", _n_pm)

            st.divider()

            # ── Trabajando Ahora ──────────────────────────────────────────────────
            # Solo OTs cuya tarea está actualmente "En Proceso" en Fracttal
            _df_ahora = _df_vivo[_df_vivo["estado_tarea"] == "En Proceso"].copy()
            _n_ahora = len(_df_ahora)
            st.markdown(
                f'<div style="font-size:1.05rem;font-weight:700;'
                f'margin-bottom:8px;">🟢 Trabajando ahora — {_n_ahora} técnico(s)</div>',
                unsafe_allow_html=True,
            )
            if _df_ahora.empty:
                st.info(
                    "Ningún técnico tiene una tarea marcada como **En Proceso** en este momento. "
                    "Los técnicos aparecen aquí cuando inician una tarea en Fracttal "
                    "(Estado Tarea → En Proceso).",
                    icon="ℹ️",
                )
            else:
                _ahora_cols = ["id_ot", "responsable", "tipo_tarea", "cliente",
                               "ubi_limpia", "inicio_fmt", "t_transcurrido", "t_prog", "avance_pct"]
                _df_ahora_show = _df_ahora[
                    [c for c in _ahora_cols if c in _df_ahora.columns]
                ].rename(columns={
                    "id_ot": "OT", "responsable": "Técnico", "tipo_tarea": "Tipo",
                    "cliente": "Cliente", "ubi_limpia": "Ubicación",
                    "inicio_fmt": "Inicio", "t_transcurrido": "T. Transcurrido",
                    "t_prog": "T. Prog.", "avance_pct": "Avance %",
                })
                st.dataframe(
                    _df_ahora_show,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Avance %": st.column_config.ProgressColumn(
                            min_value=0, max_value=100, format="%.0f%%", width=100
                        ),
                    },
                )

            st.divider()

            # ── Agrupar tipo_tarea para el filtro ────────────────────────────────
            # Todas las variantes PREVENTIVA* → "PREVENTIVA" en el filtro
            # SOLICITUD COMERCIAL se excluye del filtro
            _TIPO_EXCLUIR_VIVO = {"SOLICITUD COMERCIAL", "ENTR. FACTURADA A CLIENTE"}
            _df_vivo["tipo_tarea_grp"] = _df_vivo["tipo_tarea"].apply(
                lambda x: "PREVENTIVA"
                if str(x).upper().startswith("PREVENTIVA") else x
            )

            # ── Filtros ───────────────────────────────────────────────────────────
            _fvc1, _fvc2, _fvc3, _fvc4 = st.columns(4)
            with _fvc1:
                _TIPO_PRIO = ["CORRECTIVA", "PREVENTIVA"]
                _ev_tipos_all = sorted(
                    t for t in _df_vivo["tipo_tarea_grp"].unique().tolist()
                    if t not in _TIPO_EXCLUIR_VIVO and t not in ("—",)
                )
                _ev_tipos_opts = (
                    [t for t in _TIPO_PRIO if t in _ev_tipos_all] +
                    [t for t in _ev_tipos_all if t not in _TIPO_PRIO]
                )
                _ev_sel_tipo = st.selectbox(
                    "Tipo OT", ["Todos"] + _ev_tipos_opts, key="ev_tipo"
                )
            with _fvc2:
                _ev_tecs = ["Todos"] + sorted(
                    t for t in _df_vivo["responsable"].unique().tolist() if t != "—"
                )
                _ev_sel_tec = st.selectbox("Técnico", _ev_tecs, key="ev_tec")
            with _fvc3:
                _ev_clis = ["Todos"] + sorted(
                    c for c in _df_vivo["cliente"].unique().tolist() if c != "—"
                )
                _ev_sel_cli = st.selectbox("Cliente", _ev_clis, key="ev_cli")
            with _fvc4:
                _ev_ests = ["Todos"] + sorted(_df_vivo["estado"].unique().tolist())
                _ev_sel_est = st.selectbox("Estado OT", _ev_ests, key="ev_est")

            _df_vf = _df_vivo.copy()
            if _ev_sel_tipo != "Todos":
                _df_vf = _df_vf[_df_vf["tipo_tarea_grp"] == _ev_sel_tipo]
            if _ev_sel_tec != "Todos":
                _df_vf = _df_vf[_df_vf["responsable"] == _ev_sel_tec]
            if _ev_sel_cli != "Todos":
                _df_vf = _df_vf[_df_vf["cliente"] == _ev_sel_cli]
            if _ev_sel_est != "Todos":
                _df_vf = _df_vf[_df_vf["estado"] == _ev_sel_est]

            # Sub-tipo preventiva: extrae MENSUAL, SEMESTRAL, etc.
            _df_vf = _df_vf.copy()
            _df_vf["sub_tipo_prev"] = _df_vf["tipo_tarea"].apply(
                lambda x: str(x).replace("PREVENTIVA", "").strip().title()
                if str(x).upper().startswith("PREVENTIVA") else ""
            )

            # ── Tabla principal ───────────────────────────────────────────────────
            st.subheader(f"Órdenes activas — {len(_df_vf):,} OT(s)")

            # Incluir columna Sub-tipo solo cuando se filtra por PREVENTIVA
            _mostrar_subtipo = (_ev_sel_tipo == "PREVENTIVA")
            _vivo_cols_disp = [
                "id_ot", "estado", "estado_tarea_lbl", "tipo_tarea_grp",
            ] + (["sub_tipo_prev"] if _mostrar_subtipo else []) + [
                "responsable", "cliente", "nombre_activo", "ubi_limpia",
                "inicio_fmt", "creacion_fmt", "t_transcurrido", "t_prog", "avance_pct",
            ]
            _df_vivo_show = _df_vf[[c for c in _vivo_cols_disp if c in _df_vf.columns]].copy()
            _df_vivo_show = _df_vivo_show.rename(columns={
                "id_ot":            "OT",
                "estado":           "Estado OT",
                "estado_tarea_lbl": "Estado Tarea",
                "tipo_tarea_grp":   "Tipo",
                "sub_tipo_prev":    "Sub-tipo",
                "responsable":      "Técnico",
                "cliente":          "Cliente",
                "nombre_activo":    "Activo / Equipo",
                "ubi_limpia":       "Ubicación",
                "inicio_fmt":       "Inicio",
                "creacion_fmt":     "Creación",
                "t_transcurrido":   "T. Transcurrido",
                "t_prog":           "T. Prog.",
                "avance_pct":       "Avance %",
            })

            st.dataframe(
                _df_vivo_show,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "OT":           st.column_config.TextColumn(width=100),
                    "Estado OT":    st.column_config.TextColumn(width=110),
                    "Estado Tarea": st.column_config.TextColumn(width=150),
                    "Tipo":         st.column_config.TextColumn(width=120),
                    "Sub-tipo":     st.column_config.TextColumn(width=110),
                    "Técnico":      st.column_config.TextColumn(width=180),
                    "Cliente":      st.column_config.TextColumn(width=95),
                    "Activo / Equipo": st.column_config.TextColumn(width=240),
                    "Ubicación":    st.column_config.TextColumn(width=200),
                    "Inicio":       st.column_config.TextColumn(width=90),
                    "Creación":     st.column_config.TextColumn(width=90),
                    "T. Transcurrido": st.column_config.TextColumn(width=110),
                    "T. Prog.":     st.column_config.TextColumn(width=75),
                    "Avance %":     st.column_config.ProgressColumn(
                        min_value=0, max_value=100, format="%.0f%%", width=100
                    ),
                },
            )

            # ── Desglose por técnico ──────────────────────────────────────────────
            if not _df_vf.empty:
                st.divider()

                # Calcular semanas/mes del periodo actual
                from datetime import datetime as _dt_c, timedelta as _td_c
                from zoneinfo import ZoneInfo as _ZIC
                _tz_c  = _ZIC("America/Santiago")
                _hoy_c = _dt_c.now(_tz_c)
                _s_act  = _hoy_c.isocalendar()[1]
                _s_prox = _s_act + 1
                _lun_act  = (_hoy_c - _td_c(days=_hoy_c.weekday())).replace(
                    hour=0, minute=0, second=0, microsecond=0)
                _dom_act  = _lun_act + _td_c(days=6, hours=23, minutes=59, seconds=59)
                _lun_prox = _lun_act + _td_c(days=7)
                _dom_prox = _lun_prox + _td_c(days=6, hours=23, minutes=59, seconds=59)

                # Header con selector de periodo alineado a la derecha
                _cc_tit, _cc_sel = st.columns([3, 1.2])
                with _cc_tit:
                    st.subheader("👷 Carga por técnico")
                with _cc_sel:
                    _periodo_opts = [
                        "Todo el periodo",
                        f"Semana actual  (S{_s_act})",
                        f"Próxima semana (S{_s_prox})",
                        f"Este mes ({_hoy_c.strftime('%b %y')})",
                    ]
                    _sel_periodo_c = st.selectbox(
                        "Periodo", _periodo_opts, key="carga_tec_periodo",
                        label_visibility="collapsed",
                    )

                # Filtrar por fecha_creacion según periodo elegido
                _df_carga = _df_vf.copy()
                if _sel_periodo_c != "Todo el periodo" and "fecha_creacion" in _df_carga.columns:
                    _df_carga["_fc_dt"] = pd.to_datetime(
                        _df_carga["fecha_creacion"], errors="coerce", utc=True
                    ).dt.tz_convert(_tz_c)
                    if "Semana actual" in _sel_periodo_c:
                        _df_carga = _df_carga[
                            (_df_carga["_fc_dt"] >= pd.Timestamp(_lun_act)) &
                            (_df_carga["_fc_dt"] <= pd.Timestamp(_dom_act))
                        ]
                    elif "Próxima" in _sel_periodo_c:
                        _df_carga = _df_carga[
                            (_df_carga["_fc_dt"] >= pd.Timestamp(_lun_prox)) &
                            (_df_carga["_fc_dt"] <= pd.Timestamp(_dom_prox))
                        ]
                    elif "Este mes" in _sel_periodo_c:
                        _df_carga = _df_carga[
                            _df_carga["_fc_dt"].dt.month == _hoy_c.month
                        ]

                st.caption(f"{len(_df_carga):,} OT(s) en el periodo seleccionado")

                if _df_carga.empty:
                    st.info("No hay OTs para el periodo seleccionado.", icon="📭")
                else:
                    _vivo_by_tec = (
                        _df_carga[_df_carga["responsable"] != "—"]
                        .groupby("responsable")
                        .agg(
                            total      =("id_ot",        "count"),
                            en_proceso =("estado_tarea",  lambda x: (x == "En Proceso").sum()),
                            correctivas=("tipo_tarea",
                                         lambda x: x.str.contains("CORREC", na=False, case=False).sum()),
                            preventivas=("tipo_tarea",
                                         lambda x: x.str.contains("PREVEN", na=False, case=False).sum()),
                            clientes   =("cliente",       lambda x: ", ".join(sorted(x[x != "—"].unique()))),
                        )
                        .reset_index()
                        .sort_values("total", ascending=False)
                        .rename(columns={
                            "responsable": "Técnico",
                            "total":       "Total OTs",
                            "en_proceso":  "🟢 En Proceso",
                            "correctivas": "🚨 Correctivas",
                            "preventivas": "🛠️ Preventivas",
                            "clientes":    "Clientes",
                        })
                    )
                    st.dataframe(
                        _vivo_by_tec,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Técnico":       st.column_config.TextColumn(width=200),
                            "Total OTs":     st.column_config.NumberColumn(width=85),
                            "🟢 En Proceso": st.column_config.NumberColumn(width=100),
                            "🚨 Correctivas":st.column_config.NumberColumn(width=100),
                            "🛠️ Preventivas":st.column_config.NumberColumn(width=100),
                            "Clientes":      st.column_config.TextColumn(width=200),
                        },
                    )

        st.stop()

    # ── Sub-tab: Planificación Turnos STO ─────────────────────────────────────
    if _util_sub_tab == "📅 Planificación Turnos":
        import json as _json_turnos
        from datetime import date as _date_turnos

        st.title("Planificación Turnos STO")

        _turnos_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "turnos_data.json")
        if not os.path.exists(_turnos_path):
            st.warning("No se encontró el archivo de turnos (turnos_data.json).")
            st.stop()

        with open(_turnos_path, "r", encoding="utf-8") as _f_t:
            _turnos_all = _json_turnos.load(_f_t)

        _weeks = _turnos_all.get("weeks", [])
        if not _weeks:
            st.info("Sin datos de turnos disponibles.")
            st.stop()

        _hoy = _date_turnos.today()
        _hoy_iso = _hoy.isoformat()
        _DIA_NOMBRE = {0: "LUN", 1: "MAR", 2: "MIÉ", 3: "JUE", 4: "VIE", 5: "SÁB", 6: "DOM"}
        _MES_LABEL_T = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
                        7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}
        _ZONE_LABEL_T = {"centro": "Zona Centro (Santiago)", "norte": "Zona Norte", "sur": "Zona Sur"}
        _ZONE_ICON_T  = {"centro": "🏙️", "norte": "🏔️", "sur": "🌲"}

        _sem_idx_actual = 0
        for _si, _w in enumerate(_weeks):
            if _hoy_iso in _w.get("dates", []):
                _sem_idx_actual = _si
                break

        _sem_labels = []
        for _wi, _w in enumerate(_weeks):
            _ds = _w.get("dates", [])
            if len(_ds) >= 7:
                _d0 = _date_turnos.fromisoformat(_ds[0])
                _d6 = _date_turnos.fromisoformat(_ds[6])
                _lbl = f"{_d0.day} {_MES_LABEL_T[_d0.month]} - {_d6.day} {_MES_LABEL_T[_d6.month]}"
                if _hoy_iso in _ds:
                    _lbl = f">>> {_lbl} (actual)"
                _sem_labels.append(_lbl)
            else:
                _sem_labels.append(f"Semana {_wi+1}")

        _sel_sem = st.selectbox(
            "Semana", _sem_labels, index=_sem_idx_actual, key="turno_semana"
        )
        _sel_idx = _sem_labels.index(_sel_sem)
        _w_data = _weeks[_sel_idx]
        _w_dates = _w_data.get("dates", [])
        _w_zones = _w_data.get("zones", {})

        _is_dark = _current_theme == "dark"
        _tcard   = "#1e293b" if _is_dark else "#ffffff"
        _ttxt    = "#e2e8f0" if _is_dark else "#1e293b"
        _tmuted  = "#94a3b8" if _is_dark else "#64748b"
        _taccent = "#3b82f6"
        _thoy_bg = "#1e3a5f" if _is_dark else "#dbeafe"
        _tlibre  = "#2d2d3f" if _is_dark else "#f1f5f9"
        _tborder = "#334155" if _is_dark else "#e2e8f0"
        _turno_c = ["#10b981", "#818cf8", "#f59e0b"]

        for _zk in ["centro", "norte", "sur"]:
            _zd = _w_zones.get(_zk)
            if not _zd:
                continue
            _equipo_turno = _zd.get("equipo", "")
            _turnos_list = _zd.get("turnos", [])

            st.markdown(
                f'<div style="background:{_tcard};border:1px solid {_tborder};border-radius:10px;'
                f'padding:14px 18px;margin:16px 0 6px 0;">'
                f'<span style="font-size:1.3rem;vertical-align:middle;">{_ZONE_ICON_T.get(_zk,"")}</span> '
                f'<span style="font-weight:700;font-size:1.0rem;color:{_ttxt};">{_ZONE_LABEL_T.get(_zk,_zk)}</span>'
                f'<span style="margin-left:12px;font-size:0.85rem;color:{_tmuted};">Equipo: </span>'
                f'<span style="font-weight:700;font-size:0.9rem;color:{_taccent};">{_equipo_turno}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )

            _hdr_html = (
                f'<th style="text-align:left;padding:10px 12px;min-width:180px;'
                f'font-size:0.8rem;color:{_tmuted};font-weight:600;'
                f'border-bottom:2px solid {_tborder};">Técnico / Turno</th>'
            )
            for _di, _dd in enumerate(_w_dates):
                _dp = _date_turnos.fromisoformat(_dd)
                _is_hoy = _dd == _hoy_iso
                _bg = _thoy_bg if _is_hoy else "transparent"
                _brd = f"border-left:2px solid {_taccent};border-right:2px solid {_taccent};border-top:2px solid {_taccent};" if _is_hoy else ""
                _hdr_html += (
                    f'<th style="text-align:center;padding:8px 2px;min-width:90px;'
                    f'background:{_bg};{_brd}border-bottom:2px solid {_tborder};">'
                    f'<div style="font-size:0.65rem;color:{_tmuted};font-weight:700;letter-spacing:0.05em;">'
                    f'{_DIA_NOMBRE[_di]}</div>'
                    f'<div style="font-size:0.95rem;font-weight:700;color:{_ttxt};">{_dp.day}</div>'
                    f'<div style="font-size:0.65rem;color:{_tmuted};">{_MES_LABEL_T[_dp.month]}</div>'
                    + (f'<div style="font-size:0.6rem;font-weight:800;color:{_taccent};'
                       f'margin-top:2px;">HOY</div>' if _is_hoy else '')
                    + '</th>'
                )

            _rows_html = ""
            for _ti, _tr in enumerate(_turnos_list):
                _tc = _turno_c[_ti % 3]
                _tec = _tr.get("tecnico", "—")
                _tnum = _tr.get("turno", _ti + 1)
                _hrs = _tr.get("horarios", [])
                _tec_obs = _tr.get("obs", "")
                _row_bg = "transparent" if _ti % 2 == 0 else (_tborder + "22")
                _obs_html = (
                    f' <span style="font-size:0.7rem;font-weight:700;color:#ef4444;'
                    f'margin-left:6px;">({_tec_obs})</span>' if _tec_obs else ""
                )
                _name_cell = (
                    f'<td style="padding:10px 12px;border-bottom:1px solid {_tborder};background:{_row_bg};">'
                    f'<div style="display:flex;align-items:center;gap:8px;">'
                    f'<span style="display:inline-block;width:10px;height:10px;border-radius:50%;'
                    f'background:{_tc};flex-shrink:0;"></span>'
                    f'<div>'
                    f'<div style="font-size:0.88rem;font-weight:700;color:{_ttxt};white-space:nowrap;">{_tec}{_obs_html}</div>'
                    f'<div style="font-size:0.68rem;color:{_tc};font-weight:600;">Turno {_tnum}</div>'
                    f'</div></div></td>'
                )
                _data_cells = ""
                for _di, _dd in enumerate(_w_dates):
                    _h = _hrs[_di] if _di < len(_hrs) else ""
                    _is_hoy = _dd == _hoy_iso
                    _cbg = _thoy_bg if _is_hoy else _row_bg
                    _brd = f"border-left:2px solid {_taccent};border-right:2px solid {_taccent};" if _is_hoy else ""
                    if _h.lower() == "libre":
                        _data_cells += (
                            f'<td style="text-align:center;padding:8px 4px;background:{_tlibre};'
                            f'{_brd}border-bottom:1px solid {_tborder};">'
                            f'<span style="color:{_tmuted};font-size:0.75rem;font-style:italic;">Libre</span></td>'
                        )
                    else:
                        _data_cells += (
                            f'<td style="text-align:center;padding:8px 4px;background:{_cbg};'
                            f'{_brd}border-bottom:1px solid {_tborder};">'
                            f'<span style="color:{_ttxt};font-size:0.8rem;font-weight:500;">{_h}</span></td>'
                        )
                _rows_html += f'<tr>{_name_cell}{_data_cells}</tr>'

            st.markdown(
                f'<div style="overflow-x:auto;margin-bottom:8px;">'
                f'<table style="width:100%;border-collapse:collapse;background:{_tcard};'
                f'border:1px solid {_tborder};border-radius:8px;overflow:hidden;">'
                f'<thead><tr>{_hdr_html}</tr></thead>'
                f'<tbody>{_rows_html}</tbody></table></div>',
                unsafe_allow_html=True,
            )

        st.caption(
            f"Datos desde planilla de turnos. "
            f"Generado: {_turnos_all.get('generated_at', '?')[:16].replace('T',' ')}"
        )
        st.stop()

    # ── Carga lazy: solo se ejecuta al visitar esta página ────────────────────
    with st.spinner("📂 Cargando planificación del tiempo…"):
        df_util, util_sheet, util_error = load_utilizacion_tiempo()
        util_sheets  = list_utilizacion_sheets()
        df_mtto_plan = load_mtto_realizados_planilla()

    # Colores de fondo por categoría (sincronizados con CATEGORY_COLORS_UTIL)
    # Adaptados al tema para evitar pasteles claros sobre fondo oscuro
    if _current_theme == "dark":
        _CAT_BG = {
            "Mant. Preventivo":   "#0d2e18",
            "Llamado Correctivo": "#2e0d0d",
            "Instalación":        "#0d1a2e",
            "Capacitación":       "#1e0d2e",
            "Reunión":            "#2e2200",
            "Feriado":            "#1a202c",
            "Inventario":         "#0a2030",
            "Oficina":            "#111f38",
        }
        _FUERA_STG_BG = "#0d2e1a"
    else:
        _CAT_BG = {
            "Mant. Preventivo":   "#dcfce7",
            "Llamado Correctivo": "#fee2e2",
            "Instalación":        "#dbeafe",
            "Capacitación":       "#f3e8ff",
            "Reunión":            "#fef3c7",
            "Feriado":            "#f1f5f9",
            "Inventario":         "#cffafe",
            "Oficina":            "#f8fafc",
        }
        _FUERA_STG_BG = "#d1fae5"

    # ── Selector de período y mes ─────────────────────────────────────────────
    _sheets_opts = util_sheets if util_sheets else ([util_sheet] if util_sheet else [])

    # Mapear nombre de hoja "MAYO 2026" → número de mes
    _MESES_FULL_NUM = {
        "ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,
        "JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"OCTUBRE":10,"NOVIEMBRE":11,"DICIEMBRE":12,
    }
    def _sheet_mes_num(s: str) -> int:
        su = str(s).upper()
        for nm, n in _MESES_FULL_NUM.items():
            if nm in su:
                return n
        return 0

    _TRIMESTRES_UTIL = {
        "T1 · Ene–Mar": [1,2,3], "T2 · Abr–Jun": [4,5,6],
        "T3 · Jul–Sep": [7,8,9], "T4 · Oct–Dic": [10,11,12],
    }
    _util_meses_nums = {_sheet_mes_num(s) for s in _sheets_opts}
    _trim_opts_util  = ["Todos"] + [
        k for k,v in _TRIMESTRES_UTIL.items() if any(m in v for m in _util_meses_nums)
    ]

    _uc1, _uc2 = st.columns([1.5, 2.5])
    with _uc1:
        _sel_trim_util = st.selectbox("Período", _trim_opts_util, key="util_trim")
    # Filtrar las hojas disponibles según el trimestre seleccionado
    if _sel_trim_util != "Todos":
        _sheets_filtrados = [s for s in _sheets_opts if _sheet_mes_num(s) in _TRIMESTRES_UTIL[_sel_trim_util]]
    else:
        _sheets_filtrados = _sheets_opts

    # Inicializar session_state con el mes auto-detectado
    if "util_sheet_sel" not in st.session_state or st.session_state["util_sheet_sel"] not in _sheets_filtrados:
        _def_sheet = util_sheet if util_sheet in _sheets_filtrados else (_sheets_filtrados[0] if _sheets_filtrados else "")
        st.session_state["util_sheet_sel"] = _def_sheet

    with _uc2:
        _sel_mes = st.selectbox(
            "📅  Mes",
            options=_sheets_filtrados,
            index=_sheets_filtrados.index(st.session_state["util_sheet_sel"]) if st.session_state["util_sheet_sel"] in _sheets_filtrados else 0,
            key=f"util_sheet_select_{_sel_trim_util}",   # key cambia con trimestre → reset automático
            help="Selecciona el mes a visualizar. Solo aparecen los meses disponibles en el Excel.",
        )
    st.session_state["util_sheet_sel"] = _sel_mes

    # Recargar datos si el mes seleccionado es distinto al auto-detectado
    if _sel_mes and _sel_mes != util_sheet:
        df_util, util_sheet, util_error = load_utilizacion_tiempo(sheet_name=_sel_mes)

    if util_error and df_util.empty:
        st.info(
            "📋 **Utilización del tiempo** — sección en proceso de migración a la nube.\n\n"
            "Esta vista lee un archivo Excel de Google Drive que solo está disponible en la versión "
            "local del dashboard. Estará disponible aquí una vez migrada a Supabase.",
            icon="🚧",
        )
        # ── Mostrar datos de Supabase que SÍ están disponibles ───────────────
        st.markdown("---")
        st.markdown("### 🛠️ Mantenciones Preventivas — estado actual")
        st.caption("Datos disponibles en la nube desde Supabase")
        with st.spinner("Cargando preventivas…"):
            _prev_plan = load_preventivas_supabase()
        if _prev_plan:
            _df_prev_plan = pd.DataFrame(_prev_plan)
            # Parsear fechas — quitar timezone si Supabase las devuelve con offset
            for _c in ["fecha_programada", "fecha_creacion", "fecha_finalizacion"]:
                if _c in _df_prev_plan.columns:
                    _tmp = pd.to_datetime(_df_prev_plan[_c], errors="coerce")
                    if _tmp.dt.tz is not None:
                        _tmp = _tmp.dt.tz_convert(None)
                    _df_prev_plan[_c] = _tmp
            # Filtrar próximas 30 días + pendientes
            _hoy = pd.Timestamp.today().normalize()
            _en_30 = _hoy + pd.Timedelta(days=30)
            if "fecha_programada" in _df_prev_plan.columns:
                _prox = _df_prev_plan[
                    (_df_prev_plan["fecha_programada"] >= _hoy) &
                    (_df_prev_plan["fecha_programada"] <= _en_30)
                ].copy()
            else:
                _prox = pd.DataFrame()
            # KPIs rápidos
            _pk1, _pk2, _pk3 = st.columns(3)
            _pk1.metric("Total OTs preventivas", f"{len(_df_prev_plan):,}")
            _pk2.metric("Próximas 30 días", f"{len(_prox):,}")
            _pend = len(_df_prev_plan[_df_prev_plan.get("estado", pd.Series()).str.lower().str.contains("inici", na=False)]) if "estado" in _df_prev_plan.columns else "—"
            _pk3.metric("No iniciadas", str(_pend))
            # Tabla próximas
            if not _prox.empty:
                st.markdown("**Próximas mantenciones (30 días)**")
                _cols_prox = [c for c in ["id_ot","nombre_tarea","responsable","fecha_programada","estado","estado_tarea"] if c in _prox.columns]
                _prox_disp = _prox[_cols_prox].copy()
                if "fecha_programada" in _prox_disp.columns:
                    _prox_disp["fecha_programada"] = _prox_disp["fecha_programada"].dt.strftime("%d/%m/%Y")
                _show_df(_prox_disp.sort_values("fecha_programada") if "fecha_programada" in _prox_disp.columns else _prox_disp,
                         use_container_width=True, hide_index=True)
            else:
                st.info("No hay mantenciones programadas en los próximos 30 días.")
        else:
            st.warning("No se pudieron cargar datos de Supabase.")
    else:
        if util_error:
            st.warning(f"Aviso: {util_error}")

        _miembros_activos = {m for g in GRUPOS_TERRENO.values() for m in g["miembros"]}
        u_techs = (
            df_util["tecnico"][df_util["tecnico"].isin(_miembros_activos)].nunique()
            if not df_util.empty else len(_miembros_activos)
        )
        u_labs  = int(df_util[~df_util["categoria"].isin(["Feriado"])]["fecha"].nunique()) if not df_util.empty else 0

        uk1, uk2 = st.columns(2)
        uk1.metric("Técnicos activos", str(u_techs))
        uk2.metric("Días laborales", str(u_labs))

        if df_util.empty:
            st.info("No hay datos de asignación para el mes en curso.")
        else:
            # Calcular semana del mes con semanas Domingo–Sábado
            # Semana 1 = desde el 1° del mes hasta el primer Sábado (puede ser 1-2 días)
            # Semana 2 = del siguiente Domingo al siguiente Sábado, etc.
            def _semana_del_mes(dt):
                d = pd.Timestamp(dt)
                first = d.replace(day=1)
                # Offset del 1° respecto a su Domingo (Dom=0, Lun=1, …, Sáb=6)
                first_sun_offset = (first.weekday() + 1) % 7
                elapsed = first_sun_offset + d.day - 1
                return f"Semana {elapsed // 7 + 1}"

            df_util = df_util.copy()
            df_util["semana_mes"] = df_util["fecha"].apply(_semana_del_mes)

            st.divider()

            # ════════════════════════════════════════════════════════════════
            # VISTA 1: Programación del día (filtro fecha + técnico)
            # ════════════════════════════════════════════════════════════════
            st.markdown('<div class="section-header">📅  Programación diaria</div>', unsafe_allow_html=True)

            dates_avail = sorted(df_util["fecha"].dt.normalize().unique())
            # Default = hoy si está en la lista, sino último día disponible
            today_ts = pd.Timestamp.now().normalize()
            default_idx = next(
                (i for i, d in enumerate(dates_avail) if pd.Timestamp(d) == today_ts),
                len(dates_avail) - 1,
            )
            _DIAS_ABBR_ES = {
                "Mon":"Lun","Tue":"Mar","Wed":"Mié",
                "Thu":"Jue","Fri":"Vie","Sat":"Sáb","Sun":"Dom",
            }
            date_labels = [
                f"{_DIAS_ABBR_ES.get(pd.Timestamp(d).strftime('%a'), pd.Timestamp(d).strftime('%a'))} "
                f"{pd.Timestamp(d).strftime('%d/%m/%Y')}"
                for d in dates_avail
            ]
            date_map    = dict(zip(date_labels, dates_avail))

            fd1, fd2 = st.columns([2, 2])
            with fd1:
                sel_fecha_lbl = st.selectbox(
                    "Fecha", date_labels, index=default_idx, key="util_fecha"
                )
            with fd2:
                all_techs = ["Todos los técnicos"] + sorted(df_util["tecnico"].unique().tolist())
                sel_tech_d = st.selectbox("Técnico", all_techs, key="util_tech_d")

            sel_fecha = date_map.get(sel_fecha_lbl)
            df_day = df_util[
                df_util["fecha"].dt.normalize() == pd.Timestamp(sel_fecha).normalize()
            ].copy() if sel_fecha is not None else pd.DataFrame()

            if sel_tech_d != "Todos los técnicos" and not df_day.empty:
                df_day = df_day[df_day["tecnico"] == sel_tech_d]

            if not df_day.empty:
                # Resumen por categorías ese día
                cat_resumen = df_day["categoria"].value_counts()
                fuera_n = df_day["fuera_santiago"].sum()
                pills = " ".join(
                    f'<span style="background:{CATEGORY_COLORS_UTIL.get(c,"#94a3b8")};'
                    f'color:white;padding:3px 10px;border-radius:12px;font-size:0.8rem;margin:2px;">'
                    f'{c} ({n})</span>'
                    for c, n in cat_resumen.items()
                )
                st.markdown(pills, unsafe_allow_html=True)
                if fuera_n:
                    st.caption(f"🟢 {int(fuera_n)} técnico(s) con ruta fuera de Santiago")

                # Tabla detallada
                def _color_row(row):
                    bg = _FUERA_STG_BG if row.get("Zona","").startswith("🟢") else _CAT_BG.get(row.get("Categoría",""), "")
                    return [f"background-color:{bg}"] * len(row) if bg else [""] * len(row)

                disp = pd.DataFrame([{
                    "Técnico":     r["tecnico"],
                    "Categoría":   r["categoria"],
                    "Zona":        "🟢 Fuera Stgo" if r["fuera_santiago"] else "⬜ Santiago",
                    "Programación del día": r["tareas"],
                } for _, r in df_day.iterrows()])

                _show_df(
                    disp.style.apply(_color_row, axis=1),
                    width="stretch", hide_index=True,
                    column_config={
                        "Zona": st.column_config.TextColumn(width=130),
                        "Categoría": st.column_config.TextColumn(width=160),
                        "Programación del día": st.column_config.TextColumn(),
                    },
                )
            else:
                st.info("Sin asignaciones para el día seleccionado.")

            st.divider()

            # ════════════════════════════════════════════════════════════════
            # VISTA 2: Semana del mes por técnico
            # ════════════════════════════════════════════════════════════════
            st.markdown('<div class="section-header">📆  Programación semanal por técnico</div>', unsafe_allow_html=True)

            sw1, sw2 = st.columns([2, 2])
            with sw1:
                semanas_disp = sorted(df_util["semana_mes"].unique())
                sel_semana = st.selectbox("Semana", semanas_disp, key="util_semana")
            with sw2:
                sel_tech_s = st.selectbox(
                    "Técnico", sorted(df_util["tecnico"].unique().tolist()), key="util_tech_s"
                )

            df_sem = df_util[
                (df_util["semana_mes"] == sel_semana) &
                (df_util["tecnico"] == sel_tech_s)
            ].sort_values("fecha")

            if not df_sem.empty:
                DIAS_ES = {"Monday":"Lunes","Tuesday":"Martes","Wednesday":"Miércoles",
                           "Thursday":"Jueves","Friday":"Viernes","Saturday":"Sábado","Sunday":"Domingo"}
                sem_rows = []
                for _, r in df_sem.iterrows():
                    ts = pd.Timestamp(r["fecha"])
                    dia_es = DIAS_ES.get(ts.strftime("%A"), ts.strftime("%A"))
                    sem_rows.append({
                        "Día":        f"{dia_es} {ts.strftime('%d/%m')}",
                        "Categoría":  r["categoria"],
                        "Zona":       "🟢 Fuera Stgo" if r["fuera_santiago"] else "⬜ Santiago",
                        "Programación": r["tareas"],
                    })
                df_sem_disp = pd.DataFrame(sem_rows)

                def _color_sem(row):
                    bg = _FUERA_STG_BG if row.get("Zona","").startswith("🟢") else _CAT_BG.get(row.get("Categoría",""), "")
                    return [f"background-color:{bg}"] * len(row) if bg else [""] * len(row)

                _show_df(
                    df_sem_disp.style.apply(_color_sem, axis=1),
                    width="stretch", hide_index=True,
                    column_config={
                        "Día": st.column_config.TextColumn(width=140),
                        "Categoría": st.column_config.TextColumn(width=160),
                        "Zona": st.column_config.TextColumn(width=130),
                        "Programación": st.column_config.TextColumn(),
                    },
                )
            else:
                st.info(f"No hay asignaciones para {sel_tech_s} en {sel_semana}.")

            st.divider()

            # ════════════════════════════════════════════════════════════════
            # VISTA 3: Distribución mensual + desglose por categoría
            # ════════════════════════════════════════════════════════════════
            st.markdown('<div class="section-header">📊  Distribución mensual por técnico</div>', unsafe_allow_html=True)

            cat_counts = (
                df_util[
                    (df_util["categoria"] != "Feriado") &
                    (df_util["tecnico"].isin(_miembros_activos))
                ]
                .groupby(["tecnico", "categoria"])
                .size()
                .reset_index(name="dias")
            )
            tech_order = (
                cat_counts.groupby("tecnico")["dias"].sum()
                .sort_values(ascending=False).index.tolist()
            )
            _util_k = f"_fig_util_{_current_theme}_{_sel_mes}"
            if _util_k not in st.session_state:
                fig_util = px.bar(
                    cat_counts, x="tecnico", y="dias", color="categoria",
                    color_discrete_map=CATEGORY_COLORS_UTIL,
                    title=f"Días por categoría — {util_sheet}",
                    barmode="stack",
                    category_orders={"tecnico": tech_order},
                    labels={"tecnico": "Técnico", "dias": "Días", "categoria": "Categoría"},
                )
                fig_util.update_layout(xaxis_tickangle=-30, legend_title="Categoría", height=400,
                                       margin=dict(t=50, b=80))
                _apply_plot_theme(fig_util)
                st.session_state[_util_k] = fig_util
            st.plotly_chart(st.session_state[_util_k], width="stretch")

            # Leyenda de colores
            leg_cols = st.columns(len(CATEGORY_COLORS_UTIL))
            for i, (cat, color) in enumerate(CATEGORY_COLORS_UTIL.items()):
                leg_cols[i].markdown(
                    f'<div style="display:flex;align-items:center;gap:4px;">'
                    f'<div style="width:12px;height:12px;background:{color};border-radius:3px;flex-shrink:0"></div>'
                    f'<small style="color:{_t["muted"]};line-height:1.2">{cat}</small></div>',
                    unsafe_allow_html=True,
                )

            st.divider()

            # ── Desglose por categoría ────────────────────────────────────
            st.markdown('<div class="section-header">🔍  Desglose por categoría</div>', unsafe_allow_html=True)
            st.caption("Selecciona una categoría para ver cada tarea individual del mes.")

            dc1, dc2 = st.columns([1, 3])
            with dc1:
                cat_sel = st.radio("Categoría", list(CATEGORY_COLORS_UTIL.keys()), key="util_cat_sel")
                n_cat = len(df_util[df_util["categoria"] == cat_sel])
                cat_color = CATEGORY_COLORS_UTIL.get(cat_sel, "#94a3b8")
                st.markdown(
                    f'<div style="margin-top:8px;padding:8px 12px;background:{cat_color}22;'
                    f'border-left:4px solid {cat_color};border-radius:4px;">'
                    f'<b style="color:{cat_color}">{cat_sel}</b><br>'
                    f'<small>{n_cat} días este mes</small></div>',
                    unsafe_allow_html=True,
                )
            with dc2:
                desglose_rows = []
                for _, r in df_util.iterrows():
                    for line in str(r["tareas"]).split(" | "):
                        line = line.strip()
                        if line and classify_task_line(line) == cat_sel:
                            desglose_rows.append({
                                "Fecha":   (
                                    f"{_DIAS_ABBR_ES.get(pd.Timestamp(r['fecha']).strftime('%a'), pd.Timestamp(r['fecha']).strftime('%a'))} "
                                    f"{pd.Timestamp(r['fecha']).strftime('%d/%m')}"
                                ),
                                "Técnico": r["tecnico"],
                                "Detalle": line,
                                "Zona":    "🟢 Fuera Stgo" if r["fuera_santiago"] else "Santiago",
                            })
                if desglose_rows:
                    df_des = pd.DataFrame(desglose_rows)
                    st.caption(f"**{cat_sel}** — {len(df_des)} tareas en {df_des['Técnico'].nunique()} técnicos")
                    _show_df(df_des, width="stretch", hide_index=True)
                else:
                    st.info(f"No hay tareas de '{cat_sel}' en {util_sheet}.")




# ─────────────────────────────────────────────────────────────────────────────
# PÁGINA 5: DESEMPEÑO TERRENO
# ─────────────────────────────────────────────────────────────────────────────
elif _page == _NAV_PAGES[0]:
    _hdr(_PAGE_TITLE[_NAV_PAGES[0]])
    st.divider()

    # ── Técnicos: Supabase o Excel ────────────────────────────────────────────
    if _USE_SUPABASE:
        df_tecnicos = load_tecnicos_supabase()
    else:
        df_tecnicos = load_base_tecnicos()
    _excel_to_full, _full_to_excel = build_tech_name_maps(df_tecnicos)

    # ── OTs: Supabase (rapido) o Fracttal API (lento) ─────────────────────────
    if _USE_SUPABASE:
        raw_wo = load_work_orders_supabase()
        # Si Supabase devuelve vacio (cache danada), limpiar y reintentar una vez
        if not raw_wo:
            load_work_orders_supabase.clear()
            raw_wo = load_work_orders_supabase()
        # Si sigue vacio, caer a Fracttal como fallback
        if not raw_wo:
            st.warning("Supabase no retorno OTs — usando Fracttal API como respaldo")
            raw_wo = _load_wo_con_progreso("ordenes de trabajo — Desempeno ST")
    else:
        raw_wo = _load_wo_con_progreso("ordenes de trabajo — Desempeno ST")
    _wo_sig = str(len(raw_wo))

    _ph_proc = st.empty()
    if st.session_state.get("_sc_sig_df_wo_base") != _wo_sig:
        with _ph_proc.container():
            with st.spinner(f"Procesando {len(raw_wo):,} ordenes de trabajo..."):
                df_wo = _sc("df_wo_base", _wo_sig, lambda: build_work_orders_df(raw_wo))
    else:
        df_wo = _sc("df_wo_base", _wo_sig, lambda: build_work_orders_df(raw_wo))
    _ph_proc.empty()

    # ── Utilidades compartidas ───────────────────────────────────────────────
    _SCORE_COLOR = {
        "verde":    "#22c55e",
        "amarillo": "#f59e0b",
        "rojo":     "#ef4444",
    }

    # ── Constantes de bono ────────────────────────────────────────────────────
    _BONO_TOTAL      = 500_000   # CLP/trimestre · pool equipo completo (seniors incluidos)
    _W_SLA           = 0.40      # 40 %
    _W_CAL           = 0.30      # 30 %
    _W_PREC          = 0.30      # 30 %  (trimestral, igual que SLA y Calidad)
    _MAX_SLA_CLP     = int(_BONO_TOTAL * _W_SLA)   # 200.000
    _MAX_CAL_CLP     = int(_BONO_TOTAL * _W_CAL)   # 150.000
    _MAX_PREC_TRM    = int(_BONO_TOTAL * _W_PREC)  # 150.000 (trimestral)

    def _score_level(score: float) -> tuple[str, str]:
        """Color/label informativo para score promedio (display, no bono)."""
        if score >= 90: return _SCORE_COLOR["verde"],    "✅ Excelente"
        if score >= 70: return _SCORE_COLOR["amarillo"], "⚠️ Regular"
        return               _SCORE_COLOR["rojo"],       "❌ Bajo"

    def _bono_sla(pct: float) -> tuple[int, str, str, int]:
        """
        Retorna (% bono, etiqueta, color, CLP/trimestre) según escala KPI Productividad SLA.
        Escala oficial — 40% bono = $200.000/trim pool (5 niveles):
          ≥ 95%        → 100% → $200.000
          93 a <95%    →  90% → $180.000
          90 a <93%    →  80% → $160.000
          85 a <90%    →  50% → $100.000
          < 85%        →   0% →       $0
        """
        m = _MAX_SLA_CLP  # 140.000
        if pct >= 95: return 100, "100%", _SCORE_COLOR["verde"],    m
        if pct >= 93: return  90,  "90%", "#16a34a",                int(m * .90)
        if pct >= 90: return  80,  "80%", "#4ade80",                int(m * .80)
        if pct >= 85: return  50,  "50%", _SCORE_COLOR["amarillo"], int(m * .50)
        return                  0, "Sin bono", _SCORE_COLOR["rojo"], 0

    def _bono_calidad(n_fallas: int, n_pms: int = 0) -> tuple[int, str, str, int]:
        """
        Retorna (% bono, etiqueta, color, CLP/trimestre) según exactitud porcentual.
        Exactitud = (PMs sin reincidencia) / total PMs × 100.
        Escala oficial Tasa de Reproceso — 30% bono = $150.000/trim pool (6 niveles):
          ≥ 98%        → 100% → $150.000
          96 a <98%    →  90% → $135.000
          94 a <96%    →  80% → $120.000
          92 a <94%    →  70% → $105.000
          90 a <92%    →  60% →  $90.000
          < 90%        →   0% →       $0
        """
        if n_pms > 0:
            exactitud = (1 - n_fallas / n_pms) * 100
        else:
            exactitud = 100.0 if n_fallas == 0 else 0.0

        m = _MAX_CAL_CLP   # 105.000
        if exactitud >= 98: return 100, f"{exactitud:.1f}% — 100%", _SCORE_COLOR["verde"],    m
        if exactitud >= 96: return  90, f"{exactitud:.1f}% —  90%", "#16a34a",                int(m * .90)
        if exactitud >= 94: return  80, f"{exactitud:.1f}% —  80%", "#4ade80",                int(m * .80)
        if exactitud >= 92: return  70, f"{exactitud:.1f}% —  70%", "#65a30d",                int(m * .70)
        if exactitud >= 90: return  60, f"{exactitud:.1f}% —  60%", _SCORE_COLOR["amarillo"], int(m * .60)
        return                       0, f"{exactitud:.1f}% —   0%", _SCORE_COLOR["rojo"],     0

    def _bono_prec(exactitud_pct: float) -> tuple[int, str, str, int]:
        """
        Retorna (% bono, etiqueta, color, CLP/trimestre) según cumplimiento de OTs correctas.
        Escala oficial Precisión Fracttal — 30% bono terreno = $150.000/trim pool (7 niveles):
          ≥ 95%  → 100% → $150.000/trim
          ≥ 90%  →  90% → $135.000/trim
          ≥ 85%  →  80% → $120.000/trim
          ≥ 80%  →  70% → $105.000/trim
          ≥ 75%  →  60% →  $90.000/trim
          ≥ 70%  →  50% →  $75.000/trim
          < 70%  →   0% →       $0
        """
        _m = _MAX_PREC_TRM
        if exactitud_pct >= 95: return 100, f"{exactitud_pct:.1f}% — ${_m:,.0f}/trim",        _SCORE_COLOR["verde"],    _m
        if exactitud_pct >= 90: return  90, f"{exactitud_pct:.1f}% — ${int(_m*.90):,}/trim",   "#16a34a",                int(_m * .90)
        if exactitud_pct >= 85: return  80, f"{exactitud_pct:.1f}% — ${int(_m*.80):,}/trim",   "#4ade80",                int(_m * .80)
        if exactitud_pct >= 80: return  70, f"{exactitud_pct:.1f}% — ${int(_m*.70):,}/trim",   "#65a30d",                int(_m * .70)
        if exactitud_pct >= 75: return  60, f"{exactitud_pct:.1f}% — ${int(_m*.60):,}/trim",   _SCORE_COLOR["amarillo"], int(_m * .60)
        if exactitud_pct >= 70: return  50, f"{exactitud_pct:.1f}% — ${int(_m*.50):,}/trim",   "#f97316",                int(_m * .50)
        return                           0, f"{exactitud_pct:.1f}% — $0",                       _SCORE_COLOR["rojo"],     0

    st.markdown(
        f'<div style="background:{_t["info_bg"]};border-left:4px solid #3b82f6;'
        f'border-radius:8px;padding:12px 16px;margin-bottom:12px;color:{_t["text"]};">'
        '<b>🏆 Desempeño Terreno</b> — Bono <b>$500.000 CLP/trimestre</b> (imponible) · '
        'pool compartido por todos los integrantes del equipo (seniors incluidos). '
        'Evaluado en 3 KPIs: '
        '<b>Productividad SLA</b> 40% ($200.000 pool) · '
        '<b>Efectividad MP</b> 30% ($150.000 pool) · '
        '<b>Precisión Fracttal</b> 30% ($150.000 pool). '
        f'<span style="font-size:0.82rem;color:{_t["muted"]};">'
        'Eq.1 Luis Pinto · Eq.2 Victor Bahamonde · Eq.3 Juan Gallardo · '
        'Eq.4 Carlos Avila · Eq.5 Luis Lopez</span></div>',
        unsafe_allow_html=True,
    )

    # ── Mapeo compartido: clave grupo → etiqueta display (usado en las 3 tabs) ──
    # Las claves de GRUPOS_TERRENO ya son el nombre del senior, por lo que
    # _EQUIPO_LABEL es una identidad — se mantiene para compatibilidad con el resto
    # del código sin necesidad de refactorizar cada uso.
    _EQUIPO_LABEL = {k: k for k in GRUPOS_TERRENO}
    # Etiquetas de display con paréntesis para los dos equipos de Carlos Avila
    _EQUIPO_LABEL["Carlos Avila Norte"] = "Carlos Avila (Norte)"
    _EQUIPO_LABEL["Carlos Avila Sur"]   = "Carlos Avila (Sur)"
    _LABEL_TO_GRUPO = {v: k for k, v in _EQUIPO_LABEL.items()}

    def _norm_n(s: str) -> str:
        return unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode().strip().lower()

    # Índice: TODAS las combinaciones de 2 palabras de cada short-name del grupo
    # Esto cubre nombres chilenos de 2, 3 y 4 partes (ej. "Iván Ignacio Vergara Ferrari" → "Ignacio Ferrari")
    _GRUPOS_NORM: dict[str, str] = {}
    for _grp_k, _grp_v in GRUPOS_TERRENO.items():
        for _mb in _grp_v["miembros"]:
            _GRUPOS_NORM[_norm_n(_mb)] = _grp_k
            _pts = _mb.split()
            for _i in range(len(_pts)):
                for _j in range(_i + 1, len(_pts)):
                    _GRUPOS_NORM[_norm_n(f"{_pts[_i]} {_pts[_j]}")] = _grp_k

    # Conjunto normalizado de exclusiones (No aplica / AUTEC)
    _NO_APLICA_NORM: set[str] = set()
    for _na in TECNICOS_NO_APLICA:
        _NO_APLICA_NORM.add(_norm_n(_na))
        _na_pts = _na.split()
        for _i in range(len(_na_pts)):
            for _j in range(_i + 1, len(_na_pts)):
                _NO_APLICA_NORM.add(_norm_n(f"{_na_pts[_i]} {_na_pts[_j]}"))

    def _get_equipo(t) -> str:
        """Nombre completo (API o Excel) → clave grupo. Prueba todas las combinaciones de 2 palabras."""
        if not isinstance(t, str) or not t.strip():
            return "Sin equipo"
        norm = _norm_n(t.strip())
        if norm in _GRUPOS_NORM:
            return _GRUPOS_NORM[norm]
        parts = t.strip().split()
        for _i in range(len(parts)):
            for _j in range(_i + 1, len(parts)):
                alias = _norm_n(f"{parts[_i]} {parts[_j]}")
                if alias in _GRUPOS_NORM:
                    return _GRUPOS_NORM[alias]
        return "Sin equipo"

    def _es_excluido(t) -> bool:
        """True si el técnico está en TECNICOS_NO_APLICA."""
        if not isinstance(t, str) or not t.strip():
            return False
        norm = _norm_n(t.strip())
        if norm in _NO_APLICA_NORM:
            return True
        parts = t.strip().split()
        for _i in range(len(parts)):
            for _j in range(_i + 1, len(parts)):
                if _norm_n(f"{parts[_i]} {parts[_j]}") in _NO_APLICA_NORM:
                    return True
        return False

    # ── Pre-computar columna "equipo" en df_wo (una sola vez por carga) ──────
    # Evita 6+ apply(_get_equipo) sobre las filas en cada rerun de filtros.
    # Optimización: _get_equipo solo se llama para los ~20 técnicos únicos,
    # luego se mapea vectorizado a las 20k filas → mismo resultado, ~1000× más rápido.
    def _build_wo_eq():
        _tmp = st.session_state["df_wo_base"].copy()
        _unique = _tmp["technician"].dropna().unique()
        _tech_equipo = {t: _get_equipo(t) for t in _unique}
        _tmp["equipo"] = _tmp["technician"].map(_tech_equipo).fillna("Sin equipo")
        return _tmp

    _ph_eq = st.empty()
    if st.session_state.get("_sc_sig_df_wo_eq") != _wo_sig + "_eq":
        with _ph_eq.container():
            with st.spinner("⚙️ Clasificando técnicos por equipo…"):
                df_wo = _sc("df_wo_eq", _wo_sig + "_eq", _build_wo_eq)
    else:
        df_wo = _sc("df_wo_eq", _wo_sig + "_eq", _build_wo_eq)
    _ph_eq.empty()

    # ── Helper compartido por los 3 tabs: semanas domingo→sábado ─────────────
    import calendar as _cal
    from datetime import date as _date, timedelta as _td

    def _semanas_del_mes(ym: str) -> list[tuple]:
        """Retorna lista de (label, start_date, end_date) para el mes 'YYYY-MM'."""
        y, m = int(ym[:4]), int(ym[5:7])
        first = _date(y, m, 1)
        last  = _date(y, m, _cal.monthrange(y, m)[1])
        semanas, num, cur = [], 1, first
        while cur <= last:
            wd = cur.weekday()           # Mon=0 … Sat=5, Sun=6
            days_to_sat = 0 if wd == 5 else (6 if wd == 6 else 5 - wd)
            end = min(cur + _td(days=days_to_sat), last)
            semanas.append((
                f"Semana {num}  ({cur.strftime('%d/%m')} – {end.strftime('%d/%m')})",
                cur, end,
            ))
            num += 1
            cur = end + _td(days=1)
        return semanas

    tab_sla, tab_cal, tab_prec, tab_bono = st.tabs([
        "📊  Desempeño SLA",
        "🎯  Efectividad MP",
        "📋  Precisión Fracttal",
        "💲  Resumen Bonos",
    ])

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 1 — PRODUCTIVIDAD (SLA)
    # ══════════════════════════════════════════════════════════════════════════
    with tab_sla:
        _tsla_desc, _tsla_ley = st.columns([3, 1])
        with _tsla_ley:
            st.markdown(
                f'<div style="background:{_t["card"]};border:1px solid {_t["border"]};'
                f'border-radius:8px;padding:10px 12px;font-size:0.76rem;color:{_t["text"]};'
                f'line-height:1.75;margin-top:2px;">'
                f'<div style="font-weight:700;color:{_t["muted"]};font-size:0.78rem;'
                f'margin-bottom:4px;letter-spacing:0.04em;">📊 ESCALA BONO SLA</div>'
                f'<span style="color:#22c55e;">≥ 95%</span> → <b>100%</b> · $200.000/trim<br>'
                f'<span style="color:#16a34a;">≥ 93%</span> → <b>90%</b> · $180.000/trim<br>'
                f'<span style="color:#4ade80;">≥ 90%</span> → <b>80%</b> · $160.000/trim<br>'
                f'<span style="color:#f59e0b;">≥ 85%</span> → <b>50%</b> · $100.000/trim<br>'
                f'<span style="color:#ef4444;">&lt; 85%</span> → <b>Sin bono</b>'
                f'</div>',
                unsafe_allow_html=True,
            )
        with _tsla_desc:
            st.markdown(
                f'<div style="background:{_t["warn_bg"]};border-left:4px solid #01798A;'
                f'border-radius:8px;padding:10px 16px;margin-bottom:12px;color:{_t["text"]};">'
                '<b>KPI 1.1 — Productividad (SLA)</b> — % de llamados correctivos cerrados '
                'dentro del tiempo comprometido. Binario por llamado: cumple ✅ o no cumple ❌.<br>'
                # ── 3 tablas compactas por cliente ────────────────────────────────────────
                f'<div style="display:flex;gap:8px;margin-top:10px;">'

                # ── COPEC ──────────────────────────────────────────────────────────────────
                f'<div style="flex:1;min-width:0;border-radius:6px;overflow:hidden;'
                f'border:1px solid {_t["border"]};">'
                f'<div style="background:#CC0000;color:#fff;font-size:0.72rem;font-weight:700;'
                f'padding:3px 8px;text-align:center;letter-spacing:0.04em;">🔴 COPEC</div>'
                f'<table style="font-size:0.72rem;border-collapse:collapse;width:100%;">'
                f'<tr style="background:#0C2540;color:#fff;">'
                f'<th style="padding:2px 7px;font-weight:600;">Prio</th>'
                f'<th style="padding:2px 7px;font-weight:600;text-align:center;">Stgo</th>'
                f'<th style="padding:2px 7px;font-weight:600;text-align:center;">Reg.</th></tr>'
                f'<tr style="background:{_t["card"]};color:{_t["text"]};">'
                f'<td style="padding:2px 7px;font-weight:700;">P1</td>'
                f'<td style="padding:2px 7px;text-align:center;">18 h</td>'
                f'<td style="padding:2px 7px;text-align:center;">24 h</td></tr>'
                f'<tr style="background:{_t["tbl_alt"]};color:{_t["text"]};">'
                f'<td style="padding:2px 7px;font-weight:700;">P2</td>'
                f'<td style="padding:2px 7px;text-align:center;">24 h</td>'
                f'<td style="padding:2px 7px;text-align:center;">48 h</td></tr>'
                f'<tr style="background:{_t["card"]};color:{_t["text"]};">'
                f'<td style="padding:2px 7px;font-weight:700;">P3</td>'
                f'<td style="padding:2px 7px;text-align:center;">36 h</td>'
                f'<td style="padding:2px 7px;text-align:center;">72 h</td></tr>'
                f'<tr style="background:{_t["tbl_alt"]};color:{_t["text"]};">'
                f'<td style="padding:2px 7px;font-weight:700;">P4</td>'
                f'<td style="padding:2px 7px;text-align:center;" colspan="2">96 h</td></tr>'
                f'</table></div>'

                # ── ARAMCO (Esmax) ─────────────────────────────────────────────────────────
                f'<div style="flex:1;min-width:0;border-radius:6px;overflow:hidden;'
                f'border:1px solid {_t["border"]};">'
                f'<div style="background:#16A34A;color:#fff;font-size:0.72rem;font-weight:700;'
                f'padding:3px 8px;text-align:center;letter-spacing:0.04em;">🟢 Aramco (Esmax)</div>'
                f'<table style="font-size:0.72rem;border-collapse:collapse;width:100%;">'
                f'<tr style="background:#0C2540;color:#fff;">'
                f'<th style="padding:2px 7px;font-weight:600;">Prio</th>'
                f'<th style="padding:2px 7px;font-weight:600;text-align:center;">Stgo</th>'
                f'<th style="padding:2px 7px;font-weight:600;text-align:center;">Reg.</th></tr>'
                f'<tr style="background:{_t["card"]};color:{_t["text"]};">'
                f'<td style="padding:2px 7px;font-weight:700;">P1</td>'
                f'<td style="padding:2px 7px;text-align:center;" colspan="2">24 h</td></tr>'
                f'<tr style="background:{_t["tbl_alt"]};color:{_t["text"]};">'
                f'<td style="padding:2px 7px;font-weight:700;">P2</td>'
                f'<td style="padding:2px 7px;text-align:center;" colspan="2">48 h</td></tr>'
                f'<tr style="background:{_t["card"]};color:{_t["text"]};">'
                f'<td style="padding:2px 7px;font-weight:700;">P3</td>'
                f'<td style="padding:2px 7px;text-align:center;" colspan="2">72 h</td></tr>'
                f'<tr style="background:{_t["tbl_alt"]};color:{_t["text"]};">'
                f'<td style="padding:2px 7px;font-weight:700;">P4</td>'
                f'<td style="padding:2px 7px;text-align:center;" colspan="2">100 h</td></tr>'
                f'</table></div>'

                # ── SHELL (Enex) ───────────────────────────────────────────────────────────
                f'<div style="flex:1;min-width:0;border-radius:6px;overflow:hidden;'
                f'border:1px solid {_t["border"]};">'
                f'<div style="background:#D4A800;color:#fff;font-size:0.72rem;font-weight:700;'
                f'padding:3px 8px;text-align:center;letter-spacing:0.04em;">🟡 Shell (Enex)</div>'
                f'<table style="font-size:0.72rem;border-collapse:collapse;width:100%;">'
                f'<tr style="background:#0C2540;color:#fff;">'
                f'<th style="padding:2px 7px;font-weight:600;">Prio</th>'
                f'<th style="padding:2px 7px;font-weight:600;text-align:center;">Stgo</th>'
                f'<th style="padding:2px 7px;font-weight:600;text-align:center;">Reg.</th></tr>'
                f'<tr style="background:{_t["card"]};color:{_t["text"]};">'
                f'<td style="padding:2px 7px;font-weight:700;">P1</td>'
                f'<td style="padding:2px 7px;text-align:center;" colspan="2">24 h</td></tr>'
                f'<tr style="background:{_t["tbl_alt"]};color:{_t["text"]};">'
                f'<td style="padding:2px 7px;font-weight:700;">P2</td>'
                f'<td style="padding:2px 7px;text-align:center;" colspan="2">48 h</td></tr>'
                f'<tr style="background:{_t["card"]};color:{_t["text"]};">'
                f'<td style="padding:2px 7px;font-weight:700;">P3</td>'
                f'<td style="padding:2px 7px;text-align:center;" colspan="2">72 h</td></tr>'
                f'<tr style="background:{_t["tbl_alt"]};color:{_t["text"]};">'
                f'<td style="padding:2px 7px;font-weight:700;">P4</td>'
                f'<td style="padding:2px 7px;text-align:center;" colspan="2">96 h</td></tr>'
                f'</table></div>'

                f'</div></div>',
                unsafe_allow_html=True,
            )

        # ── Umbrales SLA por zona — fallback generico (usar SLA_HOURS por cliente cuando sea posible)
        # Los valores reales por contrato estan en gdrive.SLA_HOURS y en Supabase.sla_umbrales_horas
        # Este dict se usa solo cuando no hay cumplimiento explicito en el dato fuente
        _SLA_ZONA = {
            ("P1", "Santiago"): 18, ("P1", "Regiones"): 24,   # COPEC mas estricto
            ("P2", "Santiago"): 24, ("P2", "Regiones"): 48,   # Correcto segun contrato
            ("P3", "Santiago"): 36, ("P3", "Regiones"): 72,   # Correcto segun contrato
            ("P4", "Santiago"): 96, ("P4", "Regiones"): 96,
        }
        _SLA_DEFAULT = {"P1": 24, "P2": 48, "P3": 72, "P4": 96}  # Valores genericos conservadores

        if df_llamados.empty:
            st.warning("No se pudieron cargar los llamados de Google Drive.")
        else:
            # ── Preprocessing cacheado en session_state ───────────────────────
            # Clave: tamaño de df_llamados; si cambian los datos se recalcula.
            # v3 — usa hora exacta (gdrive.py combina Fecha+Hora) y cumplimiento del Excel
            _sla_key = f"_sla_proc_v5_{len(df_llamados)}"
            if _sla_key not in st.session_state:
                _src = df_llamados[
                    df_llamados["fecha_atencion"].notna() &
                    df_llamados["fecha_llamado"].notna()
                ].copy()

                # ── Horas de resolución ─────────────────────────────────────────
                # Prioridad 1: tiempo_resp_real del Excel (TMPO.RESP.REAL) — valor exacto
                # Prioridad 2: diferencia entre fecha_atencion y fecha_llamado
                #              (ya incluye hora exacta tras la corrección en gdrive.py)
                if "tiempo_resp_real" in _src.columns:
                    _src["horas_resolucion"] = (
                        pd.to_timedelta(_src["tiempo_resp_real"], errors="coerce")
                        .dt.total_seconds() / 3600
                    ).clip(lower=0).round(2)
                    _miss = _src["horas_resolucion"].isna()
                    if _miss.any():
                        _src.loc[_miss, "horas_resolucion"] = (
                            (pd.to_datetime(_src.loc[_miss, "fecha_atencion"]) -
                             pd.to_datetime(_src.loc[_miss, "fecha_llamado"]))
                            .dt.total_seconds() / 3600
                        ).clip(lower=0).round(2)
                else:
                    _src["horas_resolucion"] = (
                        (pd.to_datetime(_src["fecha_atencion"]) -
                         pd.to_datetime(_src["fecha_llamado"]))
                        .dt.total_seconds() / 3600
                    ).clip(lower=0).round(2)

                _src["prioridad"] = _src["prioridad"].fillna("").str.strip().str.upper()

                # ── Zona normalizada (consistente con gdrive._compute_sla) ───────
                def _norm_zona_sla(z: str) -> str:
                    z = str(z).strip().upper()
                    if not z or z == "NAN":
                        return "Santiago"
                    # RM y variantes = Región Metropolitana → Santiago
                    if z in ("RM", "R.M.", "R.M") or any(k in z for k in ("SANTIAGO", "METRO")):
                        return "Santiago"
                    # Norte, Sur, Centro, Regiones explícitas → Regiones
                    return "Regiones"

                if "zona" in _src.columns:
                    _src["zona_norm"] = _src["zona"].fillna("").astype(str).apply(_norm_zona_sla)
                else:
                    _src["zona_norm"] = "Santiago"

                # ── Excepciones SLA justificadas (no modifican el Excel) ─────────
                # OTs que el Excel marca "NO CUMPLE" pero fueron justificadas
                # posteriormente por causas fuera de la responsabilidad de Occimiano.
                # Se registran aquí para no tocar el archivo fuente de Google Drive.
                # Formato: "OS-XXXXX" — el match usa normalización (sin espacios, sin puntos).
                _SLA_OVERRIDE_CUMPLE: set = {
                    # Mayo 2026 — ESMAX/Aramco — causas externas justificadas
                    "OS-37055",
                    "OS-37448",
                    "OS-37547",
                }
                # Normalizar los folios de override para comparación robusta
                def _norm_folio(s: str) -> str:
                    return str(s).strip().upper().replace(" ", "").replace(".", "")
                _override_norm = {_norm_folio(f) for f in _SLA_OVERRIDE_CUMPLE}

                # ── cumple_sla ──────────────────────────────────────────────────
                # Fuente primaria: columna "cumplimiento" del Excel (STATUS CUMPLIMIENTO /
                # Status Cumplimiento) — ya normalizada a "CUMPLE" / "NO CUMPLE" en gdrive.py.
                # Fallback: comparar horas_resolucion vs umbral SLA por prioridad y zona.
                if "cumplimiento" in _src.columns:
                    _src["cumple_sla"] = _src["cumplimiento"].map(
                        {"CUMPLE": True, "NO CUMPLE": False}
                    )
                    # Aplicar excepciones: buscar en os_fracttal Y en n_llamado como fallback
                    _override_mask = pd.Series(False, index=_src.index)
                    if "os_fracttal" in _src.columns:
                        _override_mask |= (
                            _src["os_fracttal"].astype(str).apply(_norm_folio).isin(_override_norm)
                        )
                    # Forzar cumple=True solo en los que el Excel dice NO CUMPLE
                    # (no toca los que ya eran CUMPLE)
                    _nc_mask = _src["cumple_sla"] == False
                    _src.loc[_override_mask & _nc_mask, "cumple_sla"] = True
                    # Completar filas sin cumplimiento explícito con cálculo
                    _fallback_mask = _src["cumple_sla"].isna()
                    if _fallback_mask.any():
                        for _pri in ("P1", "P2", "P3", "P4"):
                            for _zona in ("Santiago", "Regiones"):
                                _mask = (_fallback_mask &
                                         (_src["prioridad"] == _pri) &
                                         (_src["zona_norm"] == _zona))
                                if _mask.any():
                                    _umb = _SLA_ZONA.get((_pri, _zona), _SLA_DEFAULT.get(_pri, 24))
                                    _src.loc[_mask, "cumple_sla"] = (
                                        _src.loc[_mask, "horas_resolucion"] <= _umb
                                    )
                else:
                    _src["cumple_sla"] = pd.NA
                    for _pri in ("P1", "P2", "P3", "P4"):
                        for _zona in ("Santiago", "Regiones"):
                            _mask = (_src["prioridad"] == _pri) & (_src["zona_norm"] == _zona)
                            if _mask.any():
                                _umb = _SLA_ZONA.get((_pri, _zona), _SLA_DEFAULT.get(_pri, 24))
                                _src.loc[_mask, "cumple_sla"] = (
                                    _src.loc[_mask, "horas_resolucion"] <= _umb
                                )

                # Mes y fecha base
                _src["fecha_llamado_dt"] = pd.to_datetime(_src["fecha_llamado"], errors="coerce")
                _src["mes"] = _src["fecha_llamado_dt"].dt.to_period("M").astype(str)

                # Normalizar nombres técnicos cortos → nombres completos
                _src["tecnico"] = _src["tecnico"].apply(
                    lambda t: _excel_to_full.get(str(t).strip(), str(t).strip())
                    if isinstance(t, str) and t.strip() else t
                )

                # Mapear técnico → equipo y filtrar excluidos (cacheado aquí →
                # no se recalcula en cada cambio de filtro, solo al cargar datos nuevos)
                _src["equipo"] = _src["tecnico"].apply(_get_equipo)
                _src = _src[~_src["tecnico"].apply(_es_excluido)].copy()
                _src["equipo_label"] = _src["equipo"].map(_EQUIPO_LABEL).fillna(_src["equipo"])

                st.session_state[_sla_key] = _src

            df_sla_src = st.session_state[_sla_key].copy()

            # ── Excepciones SLA — forzar cumple=True para OTs justificadas ──
            # df_llamados ya tiene cumplimiento="CUMPLE" para estas OTs (override global).
            # Este bloque sincroniza cumple_sla en caso de que el caché v5 todavía
            # no se haya reconstruido con los datos corregidos.
            _SLA_EXC_NUM = {"140785", "143926", "145331"}
            if "cumple_sla" in df_sla_src.columns:
                _ov_mask = pd.Series(False, index=df_sla_src.index)
                for _col in ["n_llamado", "LLAMADO", "N° llamado"]:
                    if _col in df_sla_src.columns:
                        _ov_mask |= df_sla_src[_col].astype(str).str.strip().isin(_SLA_EXC_NUM)
                        break
                for _col in ["os_fracttal", "OS FRACTTAL"]:
                    if _col in df_sla_src.columns:
                        _ov_mask |= df_sla_src[_col].astype(str).str.strip().str.upper().isin(
                            {"OS-37055", "OS-37448", "OS-37547"})
                        break
                df_sla_src.loc[_ov_mask, "cumple_sla"] = True

            # ── Filtros ───────────────────────────────────────────────────────
            # Mapeo abreviatura → número de mes (para filtro trimestral)
            _MES_ABR_NUM = {"Ene":1,"Feb":2,"Mar":3,"Abr":4,"May":5,"Jun":6,
                            "Jul":7,"Ago":8,"Sep":9,"Oct":10,"Nov":11,"Dic":12}
            _TRIMESTRES_SLA = {
                "T1 · Ene–Mar": [1,2,3], "T2 · Abr–Jun": [4,5,6],
                "T3 · Jul–Sep": [7,8,9], "T4 · Oct–Dic": [10,11,12],
            }
            _meses_sla = sorted(df_sla_src["mes"].dropna().unique(), reverse=True)
            # mes está en formato "YYYY-MM" → extraer número de mes directamente
            _meses_sla_nums = {int(str(m).split("-")[1]) for m in _meses_sla if "-" in str(m)}
            _trim_opts_sla = ["Todos"] + [
                k for k,v in _TRIMESTRES_SLA.items() if any(m in v for m in _meses_sla_nums)
            ]

            _sf0, _sf1, _sf2, _sf3, _sf4 = st.columns([1.5, 1.4, 1.6, 2, 2])
            with _sf0:
                _trim_sla = st.selectbox("Período", _trim_opts_sla, key="sla_trim")
            with _sf1:
                # Filtrar meses al trimestre seleccionado
                if _trim_sla != "Todos":
                    _trim_m_sla = _TRIMESTRES_SLA[_trim_sla]
                    _meses_sla_disp = [m for m in _meses_sla
                                       if int(str(m).split("-")[1]) in _trim_m_sla]
                else:
                    _meses_sla_disp = _meses_sla
                _MESES_FULL_SLA = {
                    "01":"Enero","02":"Febrero","03":"Marzo","04":"Abril","05":"Mayo","06":"Junio",
                    "07":"Julio","08":"Agosto","09":"Septiembre","10":"Octubre","11":"Noviembre","12":"Diciembre",
                }
                def _ym_sla_lbl(ym):
                    p = str(ym).split("-")
                    return f"{_MESES_FULL_SLA.get(p[1], p[1])} {p[0][2:]}" if len(p) == 2 else ym
                _lbl_to_per_sla  = {_ym_sla_lbl(m): m for m in _meses_sla}
                _meses_sla_disp_lbl = [_ym_sla_lbl(m) for m in _meses_sla_disp]
                _mes_sla_lbl = st.multiselect("Mes", _meses_sla_disp_lbl, key="sla_mes",
                                              placeholder="Todos los meses")
                _mes_sla = [_lbl_to_per_sla[l] for l in _mes_sla_lbl if l in _lbl_to_per_sla]
            with _sf2:
                if len(_mes_sla) == 1:
                    _sems = _semanas_del_mes(_mes_sla[0])
                    _sem_opts = ["Todas"] + [s[0] for s in _sems]
                else:
                    _sem_opts = ["Todas"]
                _sem_sla = st.selectbox("Semana", _sem_opts, key="sla_semana")
            with _sf3:
                _equipos_con_datos = df_sla_src["equipo"].unique()
                _equipos_disp = ["Todos"] + [
                    lbl for grp, lbl in _EQUIPO_LABEL.items()
                    if grp in _equipos_con_datos
                ]
                _equipo_sla = st.selectbox("Equipo", _equipos_disp, key="sla_equipo")
            with _sf4:
                if _equipo_sla != "Todos":
                    _grp_key_filt = _LABEL_TO_GRUPO.get(_equipo_sla, "")
                    # Nombres desde el dato real (igual que Precisión) — incluye al senior
                    _tec_sla_opts = ["Todos"] + sorted(
                        t for t in df_sla_src[df_sla_src["equipo"] == _grp_key_filt]["tecnico"].dropna().unique()
                        if not _es_excluido(t)
                    )
                else:
                    _tec_sla_opts = ["Todos"] + sorted(
                        t for t in df_sla_src["tecnico"].dropna().unique()
                        if not _es_excluido(t) and _get_equipo(t) != "Sin equipo"
                    )
                _tec_sla_sel = st.selectbox("Técnico", _tec_sla_opts, key="sla_tecnico")

            _df_sla = df_sla_src.copy()
            # Período trimestral — mes está en formato "YYYY-MM", extraer número de mes
            if _trim_sla != "Todos":
                _trim_months_sla = _TRIMESTRES_SLA[_trim_sla]
                _df_sla = _df_sla[
                    _df_sla["mes"].apply(
                        lambda m: int(str(m).split("-")[1]) if "-" in str(m) else 0
                    ).isin(_trim_months_sla)
                ]
            if _mes_sla:  # lista no vacía = selección específica
                _df_sla = _df_sla[_df_sla["mes"].astype(str).isin([str(m) for m in _mes_sla])]
            if _sem_sla != "Todas" and len(_mes_sla) == 1:
                _sem_match = next((s for s in _sems if s[0] == _sem_sla), None)
                if _sem_match:
                    _sem_start, _sem_end = _sem_match[1], _sem_match[2]
                    _df_sla = _df_sla[
                        (_df_sla["fecha_llamado_dt"].dt.date >= _sem_start) &
                        (_df_sla["fecha_llamado_dt"].dt.date <= _sem_end)
                    ]
            if _equipo_sla != "Todos":
                _grp_sla = _LABEL_TO_GRUPO.get(_equipo_sla, _equipo_sla)
                _df_sla = _df_sla[_df_sla["equipo"] == _grp_sla]
            if _tec_sla_sel != "Todos":
                _df_sla = _df_sla[_df_sla["tecnico"] == _tec_sla_sel]

            st.divider()

            _df_con_pri = _df_sla[_df_sla["cumple_sla"].notna()].copy()
            _total_ll   = len(_df_sla)
            _con_pri    = len(_df_con_pri)
            _sin_pri    = _total_ll - _con_pri
            _pct_cumple = (float(_df_con_pri["cumple_sla"].sum()) / _con_pri * 100) if _con_pri > 0 else 0.0
            _bono_global, _bono_global_lbl, _, _bono_global_clp = _bono_sla(_pct_cumple)

            # ── Debug temporal: mostrar cuántas excepciones se aplicaron ────────
            _ov_aplicadas = int(_ov_mask.sum()) if "_ov_mask" in dir() else 0
            if _ov_aplicadas > 0:
                st.caption(f"ℹ️ {_ov_aplicadas} excepción(es) SLA aplicada(s) (causas externas justificadas)")

            _sm1, _sm2, _sm3, _sm4 = st.columns(4)
            _sm1.metric("Llamados cerrados", f"{_total_ll:,}")
            _sm2.metric("Con prioridad asignada", f"{_con_pri:,}")
            _cumple_abs = int(_df_con_pri["cumple_sla"].sum()) if not _df_con_pri.empty else 0
            _sm3.metric(
                "% Cumple SLA",
                f"{_pct_cumple:.1f}%",
                delta=f"{_cumple_abs:,} / {_con_pri:,} llamados",
                delta_color="off",
                help=f"Cálculo ponderado: {_cumple_abs:,} llamados que cumplen ÷ {_con_pri:,} total = {_pct_cumple:.1f}%. "
                     "No es promedio de % por cliente — cada llamado pesa igual independiente del cliente.",
            )
            _sm4.metric(
                "KPI Productividad (40%)",
                f"${_bono_global_clp:,.0f}",
                delta=f"{_bono_global_lbl} bono · max ${_MAX_SLA_CLP:,}",
                delta_color="off",
            )

            if _con_pri > 0:
                st.divider()
                # ── Tarjetas por equipo ───────────────────────────────────────
                st.markdown('<div class="section-header">👥 Cumplimiento SLA por equipo</div>',
                            unsafe_allow_html=True)
                _eq_sum = _df_con_pri[_df_con_pri["equipo"] != "Sin equipo"].groupby("equipo").agg(
                    total=("cumple_sla", "count"),
                    cumple=("cumple_sla", "sum"),
                    horas_prom=("horas_resolucion", "mean"),
                ).reset_index()
                _eq_sum["pct_sla"]      = (_eq_sum["cumple"] / _eq_sum["total"] * 100).round(1)
                _eq_sum["horas_prom"]   = _eq_sum["horas_prom"].round(1)
                _eq_sum["senior"]       = _eq_sum["equipo"].apply(
                    lambda g: GRUPOS_TERRENO.get(g, {}).get("senior", "")
                )
                _eq_sum["equipo_label"] = _eq_sum["equipo"].map(_EQUIPO_LABEL).fillna(_eq_sum["equipo"])
                # Respetar el orden canónico de GRUPOS_TERRENO (groupby ordena alfabético por defecto)
                _grp_order = {k: i for i, k in enumerate(GRUPOS_TERRENO)}
                _eq_sum["_order"] = _eq_sum["equipo"].map(_grp_order).fillna(99)
                _eq_sum = _eq_sum.sort_values("_order").drop(columns=["_order"]).reset_index(drop=True)

                _cols_eq = st.columns(max(1, len(_eq_sum)))
                for _col_ui, _row_eq in zip(_cols_eq, _eq_sum.itertuples()):
                    _bono_val, _bono_lbl, _col_eq, _bono_clp = _bono_sla(_row_eq.pct_sla)
                    # Título: técnico específico si está seleccionado y esta tarjeta tiene sus datos
                    if _tec_sla_sel != "Todos":
                        _sla_titulo    = _tec_sla_sel
                        _sla_subtitulo = f"Equipo: {_row_eq.equipo_label}"
                    else:
                        _sla_titulo    = _row_eq.equipo_label
                        _sla_subtitulo = f"Senior: {_row_eq.senior}"
                    _col_ui.markdown(
                        f'<div style="background:{_t["card"]};border:2px solid {_col_eq}33;'
                        f'border-radius:8px;padding:14px 16px;text-align:center;">'
                        f'<div style="font-weight:700;font-size:1rem;color:{_t["text"]};">'
                        f'{_sla_titulo}</div>'
                        f'<div style="font-size:0.8rem;color:{_t["muted"]};margin-bottom:8px;">'
                        f'{_sla_subtitulo}</div>'
                        f'<div style="font-size:1.6rem;font-weight:800;color:{_col_eq};">'
                        f'{_row_eq.pct_sla:.1f}% SLA</div>'
                        f'<div style="background:{_col_eq};color:#fff;border-radius:4px;'
                        f'padding:2px 8px;font-size:0.82rem;font-weight:700;'
                        f'margin:6px auto 4px;display:inline-block;">'
                        f'{_bono_lbl} → ${_bono_clp:,.0f}</div>'
                        f'<div style="font-size:0.78rem;color:{_t["muted"]};">'
                        f'{int(_row_eq.cumple)}/{int(_row_eq.total)} llamados '
                        f'· prom {_row_eq.horas_prom}h</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

                # ── Evolución del cumplimiento SLA ────────────────────
                st.divider()
                _sla_ev_by_week = len(_mes_sla) == 1
                _sla_ev_titulo = ("Evolución por semanas" if _sla_ev_by_week
                                  else "Evolución por mes seleccionado")
                st.markdown(f'<div class="section-header">📈 {_sla_ev_titulo} — Cumplimiento SLA</div>',
                            unsafe_allow_html=True)

                _MN_SLA = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
                           7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}
                def _sla_m2l(ym):
                    p = str(ym).split("-")
                    return f"{_MN_SLA.get(int(p[1]),p[1])} '{p[0][2:]}" if len(p)==2 else str(ym)

                _sla_hist = _df_con_pri.copy()
                if not _sla_hist.empty and "fecha_llamado" in _sla_hist.columns:
                    _sla_hist["_fecha_dt"] = pd.to_datetime(
                        _sla_hist["fecha_llamado"], errors="coerce"
                    )

                    if _sla_ev_by_week:
                        _sla_sems = _semanas_del_mes(_mes_sla[0])
                        _sla_rows = []
                        for _slbl, _sstart, _send in _sla_sems:
                            _sw = _sla_hist[
                                (_sla_hist["_fecha_dt"].dt.date >= _sstart) &
                                (_sla_hist["_fecha_dt"].dt.date <= _send)
                            ]
                            if _sw.empty:
                                continue
                            _sla_rows.append({
                                "bucket_lbl": _slbl,
                                "total": len(_sw),
                                "cumple": int(_sw["cumple_sla"].sum()),
                            })
                        _sla_trend = pd.DataFrame(_sla_rows)
                    else:
                        _sla_hist["_mes"] = _sla_hist["_fecha_dt"].dt.to_period("M").astype(str)
                        _sla_trend = (
                            _sla_hist.groupby("_mes")
                            .agg(total=("cumple_sla","count"), cumple=("cumple_sla","sum"))
                            .reset_index().sort_values("_mes")
                        )
                        _sla_trend["bucket_lbl"] = _sla_trend["_mes"].apply(_sla_m2l)

                    if not _sla_trend.empty:
                        _sla_trend = _sla_trend[_sla_trend["total"] > 0].copy()
                    if not _sla_trend.empty:
                        _sla_trend["pct_sla"] = (_sla_trend["cumple"] / _sla_trend["total"] * 100).round(1)
                        _sla_trend["pct_nc"]  = (100 - _sla_trend["pct_sla"]).round(1)

                    _sla_tot_all = int(_sla_trend["total"].sum())
                    _sla_ok_all  = int(_sla_trend["cumple"].sum())
                    _sla_pct_all = round(_sla_ok_all / _sla_tot_all * 100, 1) if _sla_tot_all else 0.0

                    _sc1, _sc2 = st.columns([1, 3])
                    with _sc1:
                        _sla_kpi_color = ("#22c55e" if _sla_pct_all >= 95 else
                                         ("#f59e0b" if _sla_pct_all >= 75 else "#ef4444"))
                        st.markdown(
                            f'<div style="background:{_t["card"]};border:2px solid {_sla_kpi_color}33;'
                            f'border-radius:10px;padding:20px;text-align:center;">'
                            f'<div style="font-size:2.2rem;font-weight:800;color:{_sla_kpi_color};">'
                            f'{_sla_pct_all:.1f}%</div>'
                            f'<div style="font-size:0.85rem;color:{_t["muted"]};margin-top:4px;">'
                            f'{_sla_ok_all:,} / {_sla_tot_all:,} llamados</div>'
                            f'<div style="font-size:0.8rem;color:{_t["muted"]};">cumplen SLA</div>'
                            f'<div style="font-size:0.75rem;color:{_t["muted"]};margin-top:8px;">'
                            f'Meta: ≥95%</div></div>',
                            unsafe_allow_html=True
                        )
                    with _sc2:
                        _sla_ev_sig = f"_fig_sla_ev_{_current_theme}_{_wo_sig}_{_equipo_sla}_{_tec_sla_sel}_{len(_df_con_pri)}_{'-'.join(str(m) for m in _mes_sla)}_{_sem_sla}"
                        if _sla_ev_sig not in st.session_state:
                            _fig_sla_ev = go.Figure()
                            _txt_ok_sla = _sla_trend.apply(lambda r:
                                f"<b>{r['pct_sla']:.1f}%</b><br><span style='font-size:10px;'>"
                                f"{int(r['cumple'])}/{int(r['total'])}</span>", axis=1).tolist()
                            _fig_sla_ev.add_trace(go.Bar(
                                x=_sla_trend["bucket_lbl"], y=_sla_trend["pct_sla"],
                                name="Cumple SLA", marker_color="#22c55e", opacity=0.95,
                                text=_txt_ok_sla, textposition="inside", insidetextanchor="middle",
                                textfont=dict(size=14, color="#ffffff",
                                              family="Inter, system-ui, sans-serif"),
                            ))
                            _txt_nc_sla = _sla_trend["pct_nc"].apply(
                                lambda v: f"<b>{v:.1f}%</b>" if v >= 10 else "").tolist()
                            _fig_sla_ev.add_trace(go.Bar(
                                x=_sla_trend["bucket_lbl"], y=_sla_trend["pct_nc"],
                                name="No cumple SLA", marker_color="#ef4444", opacity=0.92,
                                text=_txt_nc_sla, textposition="inside", insidetextanchor="middle",
                                textfont=dict(size=12, color="#ffffff"),
                            ))
                            for _, _r in _sla_trend.iterrows():
                                if _r["pct_nc"] < 10 and _r["pct_nc"] > 0:
                                    _fig_sla_ev.add_annotation(
                                        x=_r["bucket_lbl"], y=100, ax=0, ay=-35,
                                        xref="x", yref="y", axref="pixel", ayref="pixel",
                                        showarrow=True, arrowhead=2, arrowsize=1, arrowwidth=1.5,
                                        arrowcolor="#ef4444",
                                        text=f"<b style='color:#ef4444;'>{_r['pct_nc']:.1f}%</b>",
                                        font=dict(size=11, color="#ef4444"),
                                        bgcolor="rgba(255,255,255,0.95)",
                                        bordercolor="#ef4444", borderwidth=1, borderpad=3,
                                    )
                            _fig_sla_ev.add_hline(y=95, line_dash="dash", line_color="#22c55e",
                                                  annotation_text="Meta 95%",
                                                  annotation_position="top left", line_width=1.5)
                            _fig_sla_ev.update_layout(
                                barmode="stack", height=320,
                                margin=dict(t=40, b=30, l=10, r=20),
                                yaxis=dict(range=[0, 115], ticksuffix="%", title="% llamados"),
                                legend=dict(orientation="h", y=1.14, x=0), bargap=0.35,
                            )
                            _apply_plot_theme(_fig_sla_ev)
                            st.session_state[_sla_ev_sig] = _fig_sla_ev
                        st.plotly_chart(st.session_state[_sla_ev_sig], width="stretch")

                # ── Ranking de técnicos ───────────────────────────────────────
                st.divider()
                st.markdown('<div class="section-header">👷 Ranking de técnicos — Cumplimiento SLA</div>',
                            unsafe_allow_html=True)
                _tec_sla_rank = _df_con_pri.groupby(["tecnico", "equipo"]).agg(
                    llamados=("cumple_sla", "count"),
                    cumple=("cumple_sla",   "sum"),
                    horas_prom=("horas_resolucion", "mean"),
                ).reset_index()
                # ── Seniors: reemplazar su fila con el agregado del equipo completo ────
                # Usamos la columna "equipo" (misma fuente que las tarjetas equipo)
                # para evitar cualquier discrepancia de tildes/espacios en nombres completos.
                for _snr in SENIORS:
                    _snr_full = TECH_NAME_MAP.get(_snr, _snr)
                    _snr_idx = _tec_sla_rank.index[_tec_sla_rank["tecnico"] == _snr_full]
                    if len(_snr_idx) == 0:
                        continue
                    _snr_data = _df_con_pri[_df_con_pri["equipo"] == _snr]
                    if _snr_data.empty:
                        continue
                    _si = _snr_idx[0]
                    _tec_sla_rank.at[_si, "llamados"]   = len(_snr_data)
                    _tec_sla_rank.at[_si, "cumple"]     = int(_snr_data["cumple_sla"].sum())
                    _tec_sla_rank.at[_si, "horas_prom"] = _snr_data["horas_resolucion"].mean()
                _tec_sla_rank["pct_sla"]    = (_tec_sla_rank["cumple"] / _tec_sla_rank["llamados"] * 100).round(1)
                _tec_sla_rank["horas_prom"] = _tec_sla_rank["horas_prom"].round(1)
                _tec_sla_rank = _tec_sla_rank.sort_values("pct_sla", ascending=True)
                _tec_sla_rank["nivel"] = _tec_sla_rank["pct_sla"].apply(
                    lambda s: f"100% — ${_MAX_SLA_CLP:,} (≥95%)"         if s >= 95
                    else (f"90% — ${int(_MAX_SLA_CLP*.90):,} (93-95%)"    if s >= 93
                    else (f"80% — ${int(_MAX_SLA_CLP*.80):,} (90-93%)"    if s >= 90
                    else (f"50% — ${int(_MAX_SLA_CLP*.50):,} (85-90%)"    if s >= 85
                    else "Sin bono (<85%)")))
                )
                _tec_sla_rank["equipo_label"] = _tec_sla_rank["equipo"].map(_EQUIPO_LABEL).fillna(_tec_sla_rank["equipo"])
                _tec_sla_rank["label"] = (
                    _tec_sla_rank["tecnico"] + " (" + _tec_sla_rank["equipo_label"] + ")"
                )
                _tec_sla_rank["texto"] = _tec_sla_rank.apply(
                    lambda r: f"{r['pct_sla']:.1f}% → {_bono_sla(r['pct_sla'])[1]}", axis=1
                )

                _sla_rank_sig = f"{_wo_sig}_{_equipo_sla}_{_tec_sla_sel}_{_mes_sla}_{_trim_sla}_{_sem_sla}"
                _sla_rank_k   = f"_fig_sla_ranking_{_current_theme}_{_sla_rank_sig}"
                if _sla_rank_k not in st.session_state:
                    _fig_tec_sla = px.bar(
                        _tec_sla_rank,
                        x="pct_sla", y="label",
                        orientation="h",
                        color="nivel",
                        color_discrete_map={
                            f"100% — ${_MAX_SLA_CLP:,} (≥95%)":       "#1a6b2e",   # verde oscuro
                            f"90% — ${int(_MAX_SLA_CLP*.90):,} (93-95%)": "#27ae60",   # verde
                            f"80% — ${int(_MAX_SLA_CLP*.80):,} (90-93%)": "#7dbb1a",   # verde lima
                            f"50% — ${int(_MAX_SLA_CLP*.50):,} (85-90%)": "#f39c12",   # naranja
                            "Sin bono (<85%)":                              "#e74c3c",   # rojo
                        },
                        category_orders={"nivel": [
                            f"100% — ${_MAX_SLA_CLP:,} (≥95%)",
                            f"90% — ${int(_MAX_SLA_CLP*.90):,} (93-95%)",
                            f"80% — ${int(_MAX_SLA_CLP*.80):,} (90-93%)",
                            f"50% — ${int(_MAX_SLA_CLP*.50):,} (85-90%)",
                            "Sin bono (<85%)",
                        ]},
                        text="texto",
                        labels={"pct_sla": "% Cumplimiento SLA", "label": ""},
                    )
                    _fig_tec_sla.add_vline(x=95, line_dash="dash", line_color="#27ae60",
                                           annotation_text="100% bono (≥95%)",
                                           annotation_position="top right", line_width=1.5)
                    _fig_tec_sla.add_vline(x=85, line_dash="dot", line_color="#e74c3c",
                                           annotation_text="Mín. 85%",
                                           annotation_position="top left", line_width=1.5)
                    _fig_tec_sla.update_traces(textposition="outside", textfont=dict(size=12))
                    _fig_tec_sla.update_layout(
                        height=max(300, len(_tec_sla_rank) * 45 + 80),
                        margin=dict(t=20, b=20, l=10, r=160),
                        xaxis=dict(range=[0, 118]),
                        yaxis=dict(categoryorder="array",
                                   categoryarray=_tec_sla_rank["label"].tolist()),
                        legend_title="Bono SLA",
                    )
                    _apply_plot_theme(_fig_tec_sla)
                    st.session_state[_sla_rank_k] = _fig_tec_sla
                st.plotly_chart(st.session_state[_sla_rank_k], width="stretch")

                # ── Detalle SLA por técnico (tabla dinámica por mes / semana) ─
                st.divider()
                st.markdown('<div class="section-header">👤 Detalle SLA por técnico</div>',
                            unsafe_allow_html=True)

                _MES_NUM_ABR = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
                                7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}

                def _mes_lbl_tec(ym: str) -> str:
                    p = str(ym).split("-")
                    return f"{_MES_NUM_ABR.get(int(p[1]), p[1])} '{p[0][2:]}" if len(p)==2 else ym

                _df_tec = _df_con_pri[_df_con_pri["equipo"] != "Sin equipo"].copy()

                # ── Determinar granularidad ───────────────────────────────────
                _use_semanas = len(_mes_sla) == 1

                if _use_semanas:
                    # Una semana por columna (solo cuando hay exactamente 1 mes)
                    _sems_tec = _semanas_del_mes(_mes_sla[0])
                    def _get_sem(dt):
                        if pd.isna(dt): return None
                        d = dt.date() if hasattr(dt, "date") else dt
                        for _lbl, _s, _e in _sems_tec:
                            if _s <= d <= _e: return _lbl
                        return None
                    _df_tec["_periodo"] = _df_tec["fecha_llamado_dt"].apply(_get_sem)
                    _df_tec = _df_tec[_df_tec["_periodo"].notna()]
                    _cols_orden = [s[0] for s in _sems_tec]
                    _subttl = f"Vista semanal — {_mes_lbl_tec(_mes_sla[0])}"
                else:
                    # Un mes por columna
                    _df_tec["_periodo"] = _df_tec["mes"].astype(str)
                    _cols_orden = sorted(_df_tec["_periodo"].dropna().unique())
                    _subttl = "Vista mensual"

                # ── Agrupar por técnico + período ─────────────────────────────
                _tec_g = (
                    _df_tec.groupby(["equipo", "tecnico", "_periodo"])
                    .agg(total=("cumple_sla","count"), cumple=("cumple_sla","sum"))
                    .reset_index()
                )
                # Seniors: cada período usa el agregado del equipo
                for _snr in SENIORS:
                    _snr_full = TECH_NAME_MAP.get(_snr, _snr)
                    _snr_pers = _tec_g.loc[_tec_g["tecnico"] == _snr_full, "_periodo"].unique()
                    for _per in _snr_pers:
                        _si_g = _tec_g.index[(_tec_g["tecnico"] == _snr_full) & (_tec_g["_periodo"] == _per)]
                        if len(_si_g) == 0:
                            continue
                        _per_data = _df_tec[
                            (_df_tec["equipo"] == _snr) & (_df_tec["_periodo"] == _per)
                        ]
                        if _per_data.empty:
                            continue
                        _tec_g.at[_si_g[0], "total"]  = len(_per_data)
                        _tec_g.at[_si_g[0], "cumple"] = int(_per_data["cumple_sla"].sum())
                _tec_g["pct"] = (
                    _tec_g["cumple"] / _tec_g["total"] * 100
                ).round(1).where(_tec_g["total"] > 0)

                # Total global por técnico (última columna)
                _tec_tot = (
                    _df_tec.groupby(["equipo","tecnico"])
                    .agg(total=("cumple_sla","count"), cumple=("cumple_sla","sum"))
                    .reset_index()
                )
                # Seniors: total = agregado del equipo completo
                for _snr in SENIORS:
                    _snr_full = TECH_NAME_MAP.get(_snr, _snr)
                    _snr_idx = _tec_tot.index[_tec_tot["tecnico"] == _snr_full]
                    if len(_snr_idx) == 0:
                        continue
                    _snr_data = _df_tec[_df_tec["equipo"] == _snr]
                    if _snr_data.empty:
                        continue
                    _tec_tot.at[_snr_idx[0], "total"]  = len(_snr_data)
                    _tec_tot.at[_snr_idx[0], "cumple"] = int(_snr_data["cumple_sla"].sum())
                _tec_tot["pct"] = (_tec_tot["cumple"] / _tec_tot["total"] * 100).round(1)

                # ── Pivot ────────────────────────────────────────────────────
                if not _tec_g.empty:
                    _pivot = _tec_g.pivot_table(
                        index=["equipo","tecnico"], columns="_periodo",
                        values="pct", aggfunc="first"
                    ).reset_index()
                    _pivot.columns.name = None

                    # Unir total
                    _pivot = _pivot.merge(
                        _tec_tot[["equipo","tecnico","pct"]].rename(columns={"pct":"Total"}),
                        on=["equipo","tecnico"], how="left"
                    )

                    # Renombrar columnas de período
                    if not _use_semanas:
                        _pivot = _pivot.rename(columns={m: _mes_lbl_tec(m) for m in _cols_orden})
                        _cols_disp_tec = [_mes_lbl_tec(m) for m in _cols_orden
                                          if _mes_lbl_tec(m) in _pivot.columns]
                    else:
                        _cols_disp_tec = [c for c in _cols_orden if c in _pivot.columns]

                    # Añadir etiqueta de equipo
                    _pivot["equipo_lbl"] = _pivot["equipo"].map(_EQUIPO_LABEL).fillna(_pivot["equipo"])
                    _pivot = _pivot.rename(columns={"tecnico": "Técnico"})
                    _pivot = _pivot.sort_values(["equipo_lbl", "Técnico"])

                    _cols_final = ["equipo_lbl", "Técnico"] + _cols_disp_tec + ["Total"]
                    _cols_final = [c for c in _cols_final if c in _pivot.columns]

                    # column_config dinámico
                    _col_cfg_tec = {
                        "equipo_lbl": st.column_config.TextColumn("Equipo", width=130),
                        "Técnico":    st.column_config.TextColumn(width=180),
                        "Total":      st.column_config.ProgressColumn(
                            "Total", min_value=0, max_value=100, format="%.1f%%"),
                    }
                    for _pc in _cols_disp_tec:
                        _col_cfg_tec[_pc] = st.column_config.ProgressColumn(
                            _pc, min_value=0, max_value=100, format="%.1f%%"
                        )

                    st.caption(f"**{_subttl}** · % Cumplimiento SLA por técnico y período")
                    _show_df(
                        _pivot[_cols_final], hide_index=True, width="stretch",
                        column_config=_col_cfg_tec,
                    )
                else:
                    st.info("Sin datos suficientes para generar el detalle por técnico.")

                # ── Tabla detalle llamados ────────────────────────────────────
                st.divider()
                st.markdown('<div class="section-header">📋 Detalle de llamados</div>',
                            unsafe_allow_html=True)

                _df_sla_disp = _df_sla.copy()

                # a) Fecha exacta (dd/MM/yyyy)
                if "fecha_llamado_dt" in _df_sla_disp.columns:
                    _df_sla_disp["_fecha_exacta"] = (
                        _df_sla_disp["fecha_llamado_dt"]
                        .dt.strftime("%d/%m/%Y").fillna("—")
                    )
                    # b) Hora inicio SLA (HH:MM) desde la misma columna datetime
                    _df_sla_disp["_hora_inicio"] = (
                        _df_sla_disp["fecha_llamado_dt"]
                        .dt.strftime("%H:%M").fillna("—")
                    )
                elif "fecha_llamado" in _df_sla_disp.columns:
                    _fl = pd.to_datetime(_df_sla_disp["fecha_llamado"], errors="coerce")
                    _df_sla_disp["_fecha_exacta"] = _fl.dt.strftime("%d/%m/%Y").fillna("—")
                    _df_sla_disp["_hora_inicio"]  = _fl.dt.strftime("%H:%M").fillna("—")
                else:
                    _df_sla_disp["_fecha_exacta"] = "—"
                    _df_sla_disp["_hora_inicio"]  = "—"

                # b) Hora cierre (de fecha_atencion)
                if "fecha_atencion" in _df_sla_disp.columns:
                    _df_sla_disp["_hora_cierre"] = (
                        pd.to_datetime(_df_sla_disp["fecha_atencion"], errors="coerce")
                        .dt.strftime("%H:%M").fillna("—")
                    )
                elif "hora_fin" in _df_sla_disp.columns:
                    _df_sla_disp["_hora_cierre"] = (
                        _df_sla_disp["hora_fin"].apply(
                            lambda t: t.strftime("%H:%M") if hasattr(t, "strftime") else "—"
                        )
                    )
                else:
                    _df_sla_disp["_hora_cierre"] = "—"

                # c) Observación del técnico — texto libre del formulario Fracttal
                # (DESCRIPCIÓN DE LA FALLA + TRABAJO REALIZADO + OBSERVACIONES),
                # consolidado por sync_numerales.py en `comentario_tecnico`.
                # Antes usábamos `note`, que siempre viene vacío en Fracttal.
                _COMENTARIO_RESUMEN_MAX = 280   # tope para la columna del cuadro
                _note_map = {}
                if not df_wo.empty and "comentario_tecnico" in df_wo.columns:
                    _note_map = (
                        df_wo[df_wo["comentario_tecnico"].astype(str).str.strip() != ""]
                        .set_index("folio")["comentario_tecnico"]
                        .to_dict()
                    )

                def _resumen_comentario(folio: str) -> str:
                    txt = _note_map.get(str(folio).strip(), "")
                    if not txt:
                        return "—"
                    # Quitar los encabezados verbosos ("DESCRIPCIÓN DE LA FALLA: …")
                    txt = _strip_comentario_headers(txt)
                    if not txt or txt == "—":
                        return "—"
                    return txt if len(txt) <= _COMENTARIO_RESUMEN_MAX \
                              else txt[:_COMENTARIO_RESUMEN_MAX].rstrip() + "…"

                if "os_fracttal" in _df_sla_disp.columns and _note_map:
                    _df_sla_disp["_observacion"] = (
                        _df_sla_disp["os_fracttal"].apply(_resumen_comentario)
                    )
                else:
                    _df_sla_disp["_observacion"] = "—"

                # d) EDS código y OT
                # eds_occim ya existe; os_fracttal ya existe
                _df_sla_disp["cumple_sla"] = _df_sla_disp["cumple_sla"].apply(
                    lambda x: "✅ Sí" if x is True else ("❌ No" if x is False else "—")
                )

                # e) Barras Uso SLA / Exceso (mismo esquema que 'Cumplimiento SLA por OT')
                #    Uso = min(horas/umbral, 100%), Exceso = max(horas/umbral - 100%, 0)
                if "tiempo_resp_esp" in _df_sla_disp.columns and "horas_resolucion" in _df_sla_disp.columns:
                    def _pct_row(r):
                        h = r.get("horas_resolucion"); u = r.get("tiempo_resp_esp")
                        if pd.notna(h) and pd.notna(u) and float(u) > 0:
                            return round(float(h) / float(u) * 100, 1)
                        return None
                    _df_sla_disp["_pct"] = _df_sla_disp.apply(_pct_row, axis=1)
                    _df_sla_disp["_uso"] = _df_sla_disp["_pct"].apply(
                        lambda v: round(min(v, 100.0), 1) if pd.notna(v) else None)
                    _df_sla_disp["_exc"] = _df_sla_disp["_pct"].apply(
                        lambda v: round(max(v - 100.0, 0.0), 1) if pd.notna(v) else None)
                else:
                    _df_sla_disp["_uso"] = None
                    _df_sla_disp["_exc"] = None

                _cols_final = [c for c in [
                    "os_fracttal", "eds_occim",
                    "equipo_label", "tecnico", "cliente", "eds_nombre",
                    "_fecha_exacta", "_hora_inicio", "_hora_cierre",
                    "prioridad", "zona_norm",
                    "horas_resolucion", "_uso", "_exc", "cumple_sla",
                    "_observacion",
                ] if c in _df_sla_disp.columns]

                _df_sla_disp = _df_sla_disp[_cols_final].copy()
                _df_sla_disp.rename(columns={
                    "os_fracttal":     "OT (OS-XXXXX)",
                    "eds_occim":       "Cód. EDS",
                    "equipo_label":    "Equipo",
                    "tecnico":         "Técnico",
                    "cliente":         "Cliente",
                    "eds_nombre":      "Estación",
                    "_fecha_exacta":   "Fecha atención",
                    "_hora_inicio":    "Hora inicio SLA",
                    "_hora_cierre":    "Hora cierre OT",
                    "prioridad":       "Prioridad",
                    "zona_norm":       "Zona",
                    "horas_resolucion":"Horas resolución",
                    "_uso":            "Uso SLA",
                    "_exc":            "Exceso",
                    "cumple_sla":      "Cumple SLA",
                    "_observacion":    "Observación técnico",
                }, inplace=True)

                _show_df(
                    _df_sla_disp,
                    width="stretch",
                    hide_index=True,
                    column_config={
                        "OT (OS-XXXXX)":       st.column_config.TextColumn(width=120),
                        "Cód. EDS":            st.column_config.TextColumn(width=90),
                        "Fecha atención":      st.column_config.TextColumn(width=105),
                        "Hora inicio SLA":     st.column_config.TextColumn(width=110),
                        "Hora cierre OT":      st.column_config.TextColumn(width=105),
                        "Horas resolución":    st.column_config.NumberColumn(format="%.1f h"),
                        "Uso SLA":             st.column_config.ProgressColumn(
                            label="Uso SLA (0–100%)",
                            min_value=0, max_value=100, format="%.1f%%",
                            help="Porcentaje del umbral SLA consumido (tope 100%). "
                                 "Si excede, el sobretiempo va en la columna 'Exceso'."),
                        "Exceso":              st.column_config.ProgressColumn(
                            label="Exceso (>100%)",
                            min_value=0, max_value=100, format="%.1f%%",
                            help="Cuánto se pasó del umbral. 0/vacía = cumplió; "
                                 "llena = excedió >100% del umbral."),
                        "Observación técnico": st.column_config.TextColumn(width=380,
                            help="Resumen del texto libre del técnico en Fracttal: falla encontrada, trabajo realizado y observaciones del formulario de la OT."),
                        "Estación":            st.column_config.TextColumn(width=200),
                    },
                )


    # ══════════════════════════════════════════════════════════════════════════
    # TAB 2 — CALIDAD (REINCIDENCIAS)
    # ══════════════════════════════════════════════════════════════════════════
    with tab_cal:
        _tcal_desc, _tcal_ley = st.columns([3, 1])
        with _tcal_ley:
            st.markdown(
                f'<div style="background:{_t["card"]};border:1px solid {_t["border"]};'
                f'border-radius:8px;padding:10px 12px;font-size:0.76rem;color:{_t["text"]};'
                f'line-height:1.75;margin-top:2px;">'
                f'<div style="font-weight:700;color:{_t["muted"]};font-size:0.78rem;'
                f'margin-bottom:4px;letter-spacing:0.04em;">📊 ESCALA BONO EFECTIVIDAD MP</div>'
                f'<span style="color:#22c55e;">≥ 98%</span> → <b>100%</b> · $105.000/trim<br>'
                f'<span style="color:#16a34a;">≥ 96%</span> → <b>90%</b> · $94.500/trim<br>'
                f'<span style="color:#4ade80;">≥ 94%</span> → <b>80%</b> · $84.000/trim<br>'
                f'<span style="color:#65a30d;">≥ 92%</span> → <b>70%</b> · $73.500/trim<br>'
                f'<span style="color:#f59e0b;">≥ 90%</span> → <b>60%</b> · $63.000/trim<br>'
                f'<span style="color:#ef4444;">&lt; 90%</span> → <b>Sin bono</b>'
                f'</div>',
                unsafe_allow_html=True,
            )
        with _tcal_desc:
            st.markdown(
                f'<div style="background:{_t["warn_bg"]};border-left:4px solid #01798A;'
                f'border-radius:8px;padding:10px 16px;margin-bottom:8px;color:{_t["text"]};">'
                '<b>KPI 1.2 — Efectividad MP</b> — Fallas post-preventiva: correctivo generado '
                'dentro de los <b>5 días siguientes</b> a un mantenimiento preventivo en el '
                'mismo equipo. El error se imputa al técnico que realizó el <b>preventivo</b> '
                '(debió detectar o resolver el problema en la mantención). '
                'Excepción: causa del cliente → no aplica.</div>',
                unsafe_allow_html=True,
            )
        st.markdown(
            f'<div style="background:{_t["err_bg"]};border-left:4px solid #ef4444;'
            f'border-radius:8px;padding:10px 16px;margin-bottom:12px;font-size:0.9rem;color:{_t["text"]};">'
            '<b>⚠️ Criterio de atribución — F.A.O + F.N.A.O:</b> '
            'Todo correctivo generado dentro de los 5 días siguientes a un PM en la misma estación '
            'imputa KPI al técnico del preventivo, <b>independiente de la clasificación que registre '
            'el técnico del correctivo</b> (F.A.O o F.N.A.O). '
            'Declarar "F.N.A.O" no exime de responsabilidad: si el servicio falló tan rápido, '
            'el problema debió detectarse en el preventivo. '
            '<b>Única excepción:</b> causa raíz confirmada como daño externo del cliente '
            '(equipo roto por el cliente, falta de insumos como sal del ablandador, etc.).</div>',
            unsafe_allow_html=True,
        )

        # Cachear en session_state — no recomputa al cambiar filtros
        # .copy() evita mutar el objeto cacheado al añadir columnas (fecha_cm_dt, mes)
        # Clientes evaluados en reincidencias: todos los clientes con contrato de PM+SLA.
        # Abastible incluida: un correctivo en ≤5 días tras un PM en Abastible = error igual.
        _CLIENTES_SLA = {"COPEC", "Aramco (Esmax)", "ESMAX (Aramco)", "SHELL (Enex)", "ABASTIBLE"}
        _df_wo_reinc = df_wo[df_wo["client"].isin(_CLIENTES_SLA)].copy() \
                       if "client" in df_wo.columns else df_wo
        try:
            if st.session_state.get("_sc_sig_df_reinc") != _wo_sig:
                with st.spinner("⚙️ Calculando reincidencias…"):
                    df_reinc = _sc("df_reinc", _wo_sig,
                                    lambda: build_reincidencias(_df_wo_reinc, _excel_to_full)).copy()
            else:
                df_reinc = _sc("df_reinc", _wo_sig,
                                lambda: build_reincidencias(_df_wo_reinc, _excel_to_full)).copy()
        except Exception as _e_cal:
            st.error(f"⚠️ Error al cargar datos de reincidencias: {_e_cal}")
            st.exception(_e_cal)
            st.stop()

        # Filtro desde enero 2026 — KPI aplica solo al año en curso
        if not df_reinc.empty and "fecha_cm" in df_reinc.columns:
            df_reinc = df_reinc[
                pd.to_datetime(df_reinc["fecha_cm"], errors="coerce") >= pd.Timestamp("2026-01-01")
            ].copy()
        elif not df_reinc.empty and "fecha2" in df_reinc.columns:
            df_reinc = df_reinc[
                pd.to_datetime(df_reinc["fecha2"], errors="coerce") >= pd.Timestamp("2026-01-01")
            ].copy()

        # Compatibilidad: si llegara un df con columnas viejas (caché stale), renombrar
        if not df_reinc.empty and "fecha2" in df_reinc.columns:
            df_reinc = df_reinc.rename(columns={
                "folio1": "folio_pm", "fecha1": "fecha_pm",
                "folio2": "folio_cm", "fecha2": "fecha_cm",
                "tecnico2": "tecnico_cm",
            })

        # Recalcular es_reincidencia_tecnico con política actual (FAO+FNAO imputan).
        # Se hace aquí para no depender del caché que puede tener la política anterior
        # (solo Trabajos Especiales y causa=cliente quedan excluidos).
        if not df_reinc.empty and "falla_tipo" in df_reinc.columns and "causa_clasif" in df_reinc.columns:
            df_reinc["es_reincidencia_tecnico"] = (
                (df_reinc["falla_tipo"] != "especial") &
                (df_reinc["causa_clasif"] != "cliente")
            )

        # ── Filtros PRIMERO — necesarios para que las tarjetas reflejen el período ──
        if df_reinc.empty or "fecha_cm" not in df_reinc.columns:
            st.success("✅ Sin fallas post-preventiva detectadas en el período disponible.")
            _df_rc      = pd.DataFrame()
            _mes_rc     = []        # lista vacía → coherente con st.multiselect
            _trim_rc    = "Todos"
            _sem_rc     = "Todas"
            _eq_rc      = "Todos"
            _tec_rc_sel = "Todos"
            _atrib_rc   = "Todos"
            _n_bruto = _n_excl = _n_fnao = _n_sin_info = _n_sin_dato = _n_fao = _n_espec = 0
        else:
            df_reinc["fecha_cm_dt"] = pd.to_datetime(df_reinc["fecha_cm"], errors="coerce")
            df_reinc["mes"] = df_reinc["fecha_cm_dt"].dt.to_period("M").astype(str)

            _meses_rc = sorted(df_reinc["mes"].dropna().unique(), reverse=True)
            _meses_rc_nums = {int(str(m).split("-")[1]) for m in _meses_rc if "-" in str(m)}
            _TRIMESTRES_RC = {
                "T1 · Ene–Mar": [1,2,3], "T2 · Abr–Jun": [4,5,6],
                "T3 · Jul–Sep": [7,8,9], "T4 · Oct–Dic": [10,11,12],
            }
            _trim_opts_rc = ["Todos"] + [
                k for k, v in _TRIMESTRES_RC.items() if any(m in v for m in _meses_rc_nums)
            ]
            _grupos_con_datos = set(df_reinc["grupo_responsable"].dropna().unique())
            _rcA, _rc0, _rc1, _rc2, _rc3, _rc4 = st.columns([1.6, 1.4, 1.4, 1.4, 2, 2])
            with _rcA:
                _ATRIB_OPTS = ["Todos", "F.A.O", "F.N.A.O"]
                _atrib_rc = st.selectbox("Atribución", _ATRIB_OPTS, key="rc_atrib")
            with _rc0:
                _trim_rc = st.selectbox("Período", _trim_opts_rc, key="rc_trim")
            with _rc1:
                if _trim_rc != "Todos":
                    _trim_m_rc = _TRIMESTRES_RC[_trim_rc]
                    _meses_rc_disp = [m for m in _meses_rc
                                      if int(str(m).split("-")[1]) in _trim_m_rc]
                else:
                    _meses_rc_disp = _meses_rc
                _MESES_FULL_RC = {
                    "01":"Enero","02":"Febrero","03":"Marzo","04":"Abril","05":"Mayo","06":"Junio",
                    "07":"Julio","08":"Agosto","09":"Septiembre","10":"Octubre","11":"Noviembre","12":"Diciembre",
                }
                def _ym_rc_lbl(ym):
                    p = str(ym).split("-")
                    return f"{_MESES_FULL_RC.get(p[1], p[1])} {p[0][2:]}" if len(p) == 2 else ym
                _lbl_to_per_rc  = {_ym_rc_lbl(m): m for m in _meses_rc}
                _meses_rc_disp_lbl = [_ym_rc_lbl(m) for m in _meses_rc_disp]
                _mes_rc_lbl = st.multiselect("Mes", _meses_rc_disp_lbl, key="rc_mes",
                                             placeholder="Todos los meses")
                _mes_rc = [_lbl_to_per_rc[l] for l in _mes_rc_lbl if l in _lbl_to_per_rc]
            with _rc2:
                if len(_mes_rc) == 1:
                    _sems_rc = _semanas_del_mes(_mes_rc[0])
                    _sem_rc_opts = ["Todas"] + [s[0] for s in _sems_rc]
                else:
                    _sem_rc_opts = ["Todas"]
                _sem_rc = st.selectbox("Semana", _sem_rc_opts, key="rc_semana")
            with _rc3:
                _eq_rc_opts = ["Todos"] + [
                    _EQUIPO_LABEL[g]
                    for g in GRUPOS_TERRENO
                    if g in _grupos_con_datos and g in _EQUIPO_LABEL
                ]
                _eq_rc = st.selectbox("Equipo", _eq_rc_opts, key="rc_equipo")
            with _rc4:
                if _eq_rc != "Todos":
                    _grp_rc_k = _LABEL_TO_GRUPO.get(_eq_rc)
                    # Nombres desde el dato real (igual que Precisión) — incluye al senior
                    _tec_rc_opts = ["Todos"] + sorted(
                        t for t in df_reinc[df_reinc["grupo_responsable"] == _grp_rc_k]["tecnico_responsable"].dropna().unique()
                        if not _es_excluido(t)
                    )
                else:
                    _tec_rc_opts = ["Todos"] + sorted(
                        t for t in df_reinc["tecnico_responsable"].dropna().unique()
                        if not _es_excluido(t) and _get_equipo(t) != "Sin equipo"
                    )
                _tec_rc_sel = st.selectbox("Técnico", _tec_rc_opts, key="rc_tecnico")

            _df_rc = df_reinc.copy()
            if _trim_rc != "Todos":
                _trim_months_rc = _TRIMESTRES_RC[_trim_rc]
                _df_rc = _df_rc[
                    _df_rc["mes"].apply(
                        lambda m: int(str(m).split("-")[1]) if "-" in str(m) else 0
                    ).isin(_trim_months_rc)
                ]
            if _mes_rc:
                _df_rc = _df_rc[_df_rc["mes"].astype(str).isin([str(m) for m in _mes_rc])]
            if _sem_rc != "Todas" and len(_mes_rc) == 1:
                _sem_rc_match = next((s for s in _sems_rc if s[0] == _sem_rc), None)
                if _sem_rc_match:
                    _sem_rc_start, _sem_rc_end = _sem_rc_match[1], _sem_rc_match[2]
                    _df_rc = _df_rc[
                        (_df_rc["fecha_cm_dt"].dt.date >= _sem_rc_start) &
                        (_df_rc["fecha_cm_dt"].dt.date <= _sem_rc_end)
                    ]
            if _eq_rc != "Todos":
                _grp_rc_k = _LABEL_TO_GRUPO.get(_eq_rc)
                if _grp_rc_k:
                    _df_rc = _df_rc[_df_rc["grupo_responsable"] == _grp_rc_k]
            if _tec_rc_sel != "Todos":
                _df_rc = _df_rc[_df_rc["tecnico_responsable"] == _tec_rc_sel]

            # ── Trazabilidad: contar ANTES de excluir por tipo de falla ──────────
            # nunique(folio_cm) = OTs correctivas únicas → 1 OT = 1 error
            # (sin importar cuántas piezas o ítems tenga adentro esa OT)
            _n_bruto = _df_rc["folio_cm"].nunique() if "folio_cm" in _df_rc.columns else len(_df_rc)
            if "falla_tipo" in _df_rc.columns:
                _cnt_falla  = _df_rc["falla_tipo"].value_counts().to_dict()
                _n_fnao     = int(_cnt_falla.get("fnao",     0))
                _n_sin_info = int(_cnt_falla.get("sin_info", 0))
                _n_sin_dato = int(_cnt_falla.get("sin_dato", 0))
                _n_fao      = int(_cnt_falla.get("fao",      0))
                _n_espec    = int(_cnt_falla.get("especial", 0))
                _n_excl     = _n_espec   # Solo Trabajos Especiales excluidos (F.N.A.O ahora SÍ imputa)
                # Política actualizada: F.A.O + F.N.A.O ambas imputan KPI.
                # Trabajos Especiales (03.-): no son fallas post-PM → únicos excluidos.
                # F.N.A.O ya NO excluye: declarar "no atribuible" no exime de responsabilidad
                # si el correctivo nació dentro de los 5 días del PM en la misma estación.
                # Única excepción real: causa_clasif="cliente" (daño externo obvio del cliente).
                _df_rc = _df_rc[~_df_rc["falla_tipo"].isin(["especial"])]
            else:
                _n_fnao = _n_sin_info = _n_sin_dato = _n_fao = _n_espec = _n_excl = 0

            # ── Filtro de atribución (F.A.O / F.N.A.O) ──────────────────────────
            if _atrib_rc == "F.A.O":
                _df_rc = _df_rc[_df_rc["falla_tipo"] == "fao"] if "falla_tipo" in _df_rc.columns else _df_rc
            elif _atrib_rc == "F.N.A.O":
                _df_rc = _df_rc[_df_rc["falla_tipo"] == "fnao"] if "falla_tipo" in _df_rc.columns else _df_rc

        # ── KPIs globales: Total PMs y PMs sin reincidencia ──────────────────────
        # Solo se cuentan PMs de los mismos clientes que el numerador (COPEC/ESMAX/SHELL/Abastible).
        _df_pm_filt = df_wo[
            (df_wo["maint_type"] == "Preventiva") &
            (~df_wo["technician"].apply(_es_excluido)) &
            (df_wo["equipo"] != "Sin equipo") &
            (df_wo["client"].isin(_CLIENTES_SLA) if "client" in df_wo.columns else True)
        ].copy()
        # Columna auxiliar: nombre normalizado (sin acentos, sin espacios dobles, minúsculas).
        # Fracttal puede guardar el nombre con doble espacio o acentos distintos al TECH_NAME_MAP.
        # _get_equipo() ya usa NFD+combinaciones para asignar "equipo" correctamente,
        # pero la comparación por técnico individual necesita la misma tolerancia.
        _df_pm_filt["_tech_norm"] = _df_pm_filt["technician"].fillna("").apply(
            lambda s: " ".join(_norm_n(s).split())
        )
        # Aplicar mismo filtro de equipo/técnico que _df_rc
        if _eq_rc != "Todos":
            _grp_pm = _LABEL_TO_GRUPO.get(_eq_rc, _eq_rc)
            _df_pm_filt = _df_pm_filt[_df_pm_filt["equipo"] == _grp_pm]
        if _tec_rc_sel != "Todos":
            _full_tec_pm   = TECH_NAME_MAP.get(_tec_rc_sel, _tec_rc_sel)
            _full_tec_pm_n = " ".join(_norm_n(_full_tec_pm).split())
            _df_pm_filt = _df_pm_filt[_df_pm_filt["_tech_norm"] == _full_tec_pm_n]
        # Aplicar mismo filtro de período/mes que _df_rc
        if _df_pm_filt["creation_date"].dt.tz is not None:
            _pm_dates = _df_pm_filt["creation_date"].dt.tz_convert(None)
        else:
            _pm_dates = _df_pm_filt["creation_date"]
        if _trim_rc != "Todos":
            _df_pm_filt = _df_pm_filt[_pm_dates.dt.month.isin(_TRIMESTRES_RC[_trim_rc])]
            _pm_dates = _pm_dates[_df_pm_filt.index]
        if _mes_rc:
            _df_pm_filt = _df_pm_filt[
                _pm_dates.reindex(_df_pm_filt.index).dt.to_period("M").astype(str).isin(_mes_rc)
            ]
        # nunique("folio") = órdenes de trabajo únicas (1 OT puede tener N equipos)
        _n_pm_total    = _df_pm_filt["folio"].nunique() if "folio" in _df_pm_filt.columns else len(_df_pm_filt)
        _n_pm_con_reinc = (
            _df_rc["folio_pm"].nunique()
            if not _df_rc.empty and "folio_pm" in _df_rc.columns else 0
        )
        _n_pm_sin_reinc = max(0, _n_pm_total - _n_pm_con_reinc)

        # ── Tarjetas de bono por equipo — calculadas con el df ya filtrado ──────
        _EQUIPO_LABEL_CAL = _EQUIPO_LABEL
        _periodo_lbl = ", ".join(_mes_rc) if _mes_rc else "todos los datos disponibles"

        # Técnico específico → 1 tarjeta (su equipo)
        # Equipo específico  → 1 tarjeta (ese equipo)
        # Todos              → 5 tarjetas (todos los equipos)
        if _tec_rc_sel != "Todos":
            _tec_grp_key = _get_equipo(_tec_rc_sel)
            _eq_label_cal_iter = {
                k: v for k, v in _EQUIPO_LABEL_CAL.items() if k == _tec_grp_key
            } or _EQUIPO_LABEL_CAL
        elif _eq_rc != "Todos":
            _grp_rc_key = _LABEL_TO_GRUPO.get(_eq_rc)
            _eq_label_cal_iter = {
                k: v for k, v in _EQUIPO_LABEL_CAL.items() if k == _grp_rc_key
            } or _EQUIPO_LABEL_CAL
        else:
            _eq_label_cal_iter = _EQUIPO_LABEL_CAL

        # Clientes evaluados en reincidencias (mismos que numerador)
        _CLIENTES_SLA_RC = {"COPEC", "Aramco (Esmax)", "ESMAX (Aramco)", "SHELL (Enex)", "ABASTIBLE"}

        # ── Denominador de tarjetas: reusar _df_pm_filt (ya filtrado por
        #    excluidos, equipo, técnico, trimestre, mes) + filtro semana ──────
        _pm_tarjetas = _df_pm_filt.copy()
        if _sem_rc != "Todas" and len(_mes_rc) == 1:
            _sem_rc_match2 = next((s for s in _sems_rc if s[0] == _sem_rc), None)
            if _sem_rc_match2:
                _pm_t_dates = (
                    _pm_tarjetas["creation_date"].dt.tz_convert(None)
                    if _pm_tarjetas["creation_date"].dt.tz is not None
                    else _pm_tarjetas["creation_date"]
                )
                _pm_tarjetas = _pm_tarjetas[
                    (_pm_t_dates.dt.date >= _sem_rc_match2[1]) &
                    (_pm_t_dates.dt.date <= _sem_rc_match2[2])
                ]

        _cal_cols = st.columns(len(_eq_label_cal_iter))
        for _ci, (_gk, _gl) in enumerate(zip(_eq_label_cal_iter.keys(), _eq_label_cal_iter.values())):
            _senior_cal = GRUPOS_TERRENO.get(_gk, {}).get("senior", "")
            _pm_equipo = _pm_tarjetas[_pm_tarjetas["equipo"] == _gk]
            _n_pm = _pm_equipo["folio"].nunique() if "folio" in _pm_equipo.columns else len(_pm_equipo)
            _n_pm_sla = _n_pm  # ahora el denominador ya es solo clientes SLA
            # Fallas post-PM del período filtrado
            # nunique(folio_cm) = OTs correctivas únicas = 1 OT = 1 error
            if not _df_rc.empty and "folio_cm" in _df_rc.columns:
                _fallas_eq = _df_rc[_df_rc["grupo_responsable"] == _gk]["folio_cm"].nunique()
            else:
                _fallas_eq = 0
            _bpct, _blbl, _bcol, _bclp = _bono_calidad(int(_fallas_eq), int(_n_pm))
            # Ratio: cuántas PMs por cada correctiva generada
            if _fallas_eq > 0:
                _ratio = round(_n_pm / _fallas_eq, 1)
                _ratio_lbl = f"1 correctiva cada <b>{_ratio:,.0f}</b> PMs"
                _ratio_clr = "#22c55e" if _ratio >= 50 else ("#f59e0b" if _ratio >= 20 else "#ef4444")
            else:
                _ratio_lbl = "Sin correctivas ✅"
                _ratio_clr = "#22c55e"
            # Título: técnico si está filtrado (ya solo hay 1 tarjeta), equipo si es "Todos"
            if _tec_rc_sel != "Todos":
                _cal_titulo    = _tec_rc_sel
                _cal_subtitulo = f"Equipo: {_gl}"
            else:
                _cal_titulo    = _gl
                _cal_subtitulo = f"Senior: {_senior_cal}"

            # % exactitud para mostrar prominente
            _exactitud_cal = round((1 - _fallas_eq / _n_pm) * 100, 1) if _n_pm > 0 else 100.0

            # Nota de cobertura: si 100% y pocos PMs SLA, advertir
            if _exactitud_cal == 100.0 and _n_pm == 0:
                _cobertura_nota = (
                    f'<div style="font-size:0.65rem;color:#f59e0b;margin-top:3px;">'
                    f'⚠️ 0 PMs en COPEC/ESMAX/SHELL — sin cobertura de evaluación</div>'
                )
            elif _exactitud_cal == 100.0:
                _cobertura_nota = (
                    f'<div style="font-size:0.65rem;color:{_t["muted"]};margin-top:3px;">'
                    f'Verificados {_n_pm} PMs en COPEC/ESMAX/SHELL</div>'
                )
            else:
                _cobertura_nota = (
                    f'<div style="font-size:0.65rem;color:{_t["muted"]};margin-top:3px;">'
                    f'{_n_pm} PMs en COPEC/ESMAX/SHELL</div>'
                )

            _cal_cols[_ci].markdown(
                f'<div style="background:{_t["card"]};border:2px solid {_bcol}33;'
                f'border-radius:8px;padding:14px 16px;text-align:center;">'
                # Nombre equipo / técnico
                f'<div style="font-weight:700;font-size:0.95rem;color:{_t["text"]};">{_cal_titulo}</div>'
                f'<div style="font-size:0.78rem;color:{_t["muted"]};margin-bottom:8px;">{_cal_subtitulo}</div>'
                # % cumplimiento — dato principal
                f'<div style="font-size:2rem;font-weight:800;color:{_bcol};line-height:1.1;">'
                f'{_exactitud_cal:.1f}%</div>'
                f'<div style="font-size:0.72rem;color:{_t["muted"]};margin-bottom:6px;">exactitud</div>'
                # Chip bono
                f'<div style="background:{_bcol};color:#fff;border-radius:4px;'
                f'padding:2px 8px;font-size:0.80rem;font-weight:700;'
                f'margin:0 auto 6px;display:inline-block;">'
                f'{_blbl} → ${_bclp:,.0f}</div>'
                # Fallas y ratio — dato secundario
                f'<div style="font-size:0.73rem;color:{_t["muted"]};margin-top:4px;">'
                f'<b style="color:{_bcol if _fallas_eq == 0 else _t["text"]};">{_fallas_eq}</b> '
                f'fallas post-PM &nbsp;·&nbsp; {_n_pm} PMs</div>'
                f'<div style="font-size:0.72rem;color:{_ratio_clr};font-weight:600;margin-top:2px;">'
                f'{_ratio_lbl}</div>'
                # Cobertura de evaluación
                + _cobertura_nota +
                f'</div>',
                unsafe_allow_html=True,
            )

        st.caption(
            f"Exactitud = PMs sin reincidencia / total PMs × 100 — escala trimestral · "
            f"≥98%→${_MAX_CAL_CLP:,} · "
            f"≥96%→${int(_MAX_CAL_CLP*.90):,} · "
            f"≥94%→${int(_MAX_CAL_CLP*.80):,} · "
            f"≥92%→${int(_MAX_CAL_CLP*.70):,} · "
            f"≥90%→${int(_MAX_CAL_CLP*.60):,} · "
            f"<90%→$0 · Período: **{_periodo_lbl}**. "
            f"Reincidencias evaluadas solo en COPEC, ESMAX y SHELL."
        )

        # ── Expander de verificación de datos ─────────────────────────────────
        with st.expander("🔬 Verificar datos del algoritmo (diagnóstico)", expanded=False):
            st.markdown(
                f'<div style="font-size:0.82rem;color:{_t["muted"]};margin-bottom:8px;">'
                f'Esto permite confirmar que el algoritmo está evaluando correctamente. '
                f'Si un equipo muestra <b>0 PMs en COPEC/ESMAX/SHELL</b>, el 100% es por '
                f'falta de cobertura (ese equipo no atiende esos clientes), no por excelencia real.'
                f'</div>',
                unsafe_allow_html=True,
            )
            # Construir tabla diagnóstico por equipo
            _diag_rows = []
            _prev_sla = _df_wo_reinc[_df_wo_reinc["maint_type"] == "Preventiva"].copy() \
                if not _df_wo_reinc.empty else pd.DataFrame()
            _corr_sla = _df_wo_reinc[_df_wo_reinc["maint_type"] == "Correctiva"].copy() \
                if not _df_wo_reinc.empty else pd.DataFrame()

            _pm_sla_codes = set(_prev_sla["equipment_code"].dropna().unique()) \
                if not _prev_sla.empty else set()
            _cm_sla_codes = set(_corr_sla["equipment_code"].dropna().unique()) \
                if not _corr_sla.empty else set()
            _codigos_comunes = _pm_sla_codes & _cm_sla_codes

            for _gk2, _gl2 in _eq_label_cal_iter.items():
                _pm_eq2 = _prev_sla[_prev_sla["equipo"] == _gk2] \
                    if not _prev_sla.empty and "equipo" in _prev_sla.columns else pd.DataFrame()
                _n_pm_eq2    = len(_pm_eq2)
                _n_pm_eq2_ok = int((_pm_eq2["equipment_code"].str.strip() != "").sum()) \
                    if not _pm_eq2.empty else 0
                _fallas_diag = _df_rc[_df_rc["grupo_responsable"] == _gk2]["folio_cm"].nunique() \
                    if not _df_rc.empty and "folio_cm" in _df_rc.columns else 0
                _pm_codes_eq2 = set(_pm_eq2["equipment_code"].dropna().unique()) \
                    if not _pm_eq2.empty else set()
                _cms_matcheables = len(_corr_sla[_corr_sla["equipment_code"].isin(_pm_codes_eq2)]) \
                    if not _corr_sla.empty and _pm_codes_eq2 else 0
                _diag_rows.append({
                    "Equipo":                    _gl2,
                    "PMs COPEC/ESMAX/SHELL":     _n_pm_eq2,
                    "PMs con código activo":      _n_pm_eq2_ok,
                    "CMs matcheables":            _cms_matcheables,
                    "Reincidencias detectadas":   _fallas_diag,
                    "Cobertura real":             "✅" if _n_pm_eq2 > 0 else "⚠️ Sin evaluación",
                })
            _df_diag = pd.DataFrame(_diag_rows)

            # Resumen global
            _n_prev_total_sla = len(_prev_sla)
            _n_corr_total_sla = len(_corr_sla)
            _n_codigos_comunes = len(_codigos_comunes)

            _c1d, _c2d, _c3d = st.columns(3)
            _c1d.metric("PMs COPEC/ESMAX/SHELL", _n_prev_total_sla)
            _c2d.metric("CMs COPEC/ESMAX/SHELL", _n_corr_total_sla)
            _c3d.metric("Equipos con código común (matcheables)", _n_codigos_comunes)

            if _n_codigos_comunes == 0 and _n_prev_total_sla > 0 and _n_corr_total_sla > 0:
                st.error(
                    "⚠️ **Sin códigos comunes entre PMs y CMs** — el `equipment_code` "
                    "(campo `codigo_activo` en Supabase) no coincide entre preventivas y "
                    "correctivas. El algoritmo no puede detectar reincidencias aunque existan. "
                    "Revisar que `codigo_activo` esté poblado en ambos tipos de OT."
                )
            elif _n_prev_total_sla == 0:
                st.warning(
                    "⚠️ No se encontraron PMs de COPEC/ESMAX/SHELL en el período. "
                    "El 100% es por falta de datos, no por excelencia."
                )
            else:
                st.success(
                    f"✅ El algoritmo tiene cobertura: {_n_prev_total_sla} PMs y "
                    f"{_n_corr_total_sla} CMs de COPEC/ESMAX/SHELL, "
                    f"{_n_codigos_comunes} equipos matcheables."
                )

            st.dataframe(
                _df_diag,
                use_container_width=True,
                hide_index=True,
            )

            # Mostrar en qué meses hay reincidencias (si existen)
            if not df_reinc.empty and "fecha_cm" in df_reinc.columns:
                _df_reinc_mes = df_reinc.copy()
                _df_reinc_mes["mes"] = pd.to_datetime(
                    _df_reinc_mes["fecha_cm"], errors="coerce"
                ).dt.to_period("M").astype(str)
                _reinc_por_mes = _df_reinc_mes["mes"].value_counts().sort_index()
                if not _reinc_por_mes.empty:
                    st.markdown(
                        f'<div style="font-size:0.8rem;color:{_t["muted"]};margin-top:6px;">'
                        f'<b>Reincidencias detectadas por mes (todos los equipos):</b> '
                        + " · ".join(f"{m}: {n}" for m, n in _reinc_por_mes.items())
                        + f'</div>',
                        unsafe_allow_html=True,
                    )
            elif df_reinc.empty:
                st.markdown(
                    f'<div style="font-size:0.8rem;color:{_t["muted"]};margin-top:6px;">'
                    f'df_reinc vacío — no se detectó ninguna reincidencia en todo el período cargado.'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        # ── Gráfico: PMs realizados vs Correctivos ≤5d ───────────────────────
        st.divider()
        st.markdown('<div class="section-header">📊  PMs realizados vs Correctivos ≤ 5 días</div>', unsafe_allow_html=True)

        _cal_pm_sig = f"{_wo_sig}_{_eq_rc}_{_tec_rc_sel}_{_mes_rc}_{_trim_rc}_{_sem_rc}"
        _cal_pm_k   = f"_fig_cal_pm_rc_{_current_theme}_{_cal_pm_sig}"
        if _cal_pm_k not in st.session_state:
            # ── Construir tabla base según nivel de filtro ────────────────────
            if _tec_rc_sel != "Todos":
                _n_rc_chart = _df_rc["folio_cm"].nunique() if not _df_rc.empty and "folio_cm" in _df_rc.columns else 0
                _pm_base = [{"_x": _tec_rc_sel, "total": _n_pm_total, "errores": _n_rc_chart}]
                _pm_x_col = "Técnico"
                _pm_title = f"Exactitud preventivas — {_tec_rc_sel}"
            elif _eq_rc != "Todos":
                _grp_k_pm = _LABEL_TO_GRUPO.get(_eq_rc, _eq_rc)
                _mbs_pm   = GRUPOS_TERRENO.get(_grp_k_pm, {}).get("miembros", [])
                _pm_base  = []
                for _mb_pm in _mbs_pm:
                    _full_pm   = TECH_NAME_MAP.get(_mb_pm, _mb_pm)
                    _full_pm_n = " ".join(_norm_n(_full_pm).split())
                    _tmp_pm    = _df_pm_filt[_df_pm_filt["_tech_norm"] == _full_pm_n]
                    _cnt_pm  = _tmp_pm["folio"].nunique() if "folio" in _tmp_pm.columns else len(_tmp_pm)
                    _cnt_rc  = (_df_rc[_df_rc["tecnico_resp_short"] == _mb_pm]["folio_cm"].nunique()
                                if not _df_rc.empty and "tecnico_resp_short" in _df_rc.columns else 0)
                    _pm_base.append({"_x": _mb_pm, "total": _cnt_pm, "errores": _cnt_rc})
                _pm_x_col = "Técnico"
                _pm_title = f"Exactitud preventivas — {_eq_rc}"
            else:
                _pm_base = []
                for _gk_pm in _EQUIPO_LABEL.keys():
                    _tmp_pm2 = _df_pm_filt[_df_pm_filt["equipo"] == _gk_pm]
                    _cnt_pm  = _tmp_pm2["folio"].nunique() if "folio" in _tmp_pm2.columns else len(_tmp_pm2)
                    _cnt_rc = (_df_rc[_df_rc["grupo_responsable"] == _gk_pm]["folio_cm"].nunique()
                               if not _df_rc.empty and "grupo_responsable" in _df_rc.columns else 0)
                    _pm_base.append({"_x": _gk_pm, "total": _cnt_pm, "errores": _cnt_rc})
                _pm_x_col = "Equipo"
                _pm_title = "Exactitud preventivas — todos los equipos"

            if _pm_base:
                _stk_rows = []
                for _r in _pm_base:
                    _tot  = _r["total"]
                    _err  = _r["errores"]
                    _ok   = max(0, _tot - _err)
                    _pct_e = round(_err / _tot * 100, 2) if _tot > 0 else 0.0
                    _pct_o = round(100 - _pct_e, 2)
                    _stk_rows.append({
                        _pm_x_col: _r["_x"],
                        "_ok": _ok,
                        "_err": _err,
                        "_pct_err": _pct_e,
                        "_pct_ok":  _pct_o,
                        "_total":   _tot,
                    })
                _stk_df = pd.DataFrame(_stk_rows)

                _pm_tot_all = int(_stk_df["_total"].sum())
                _pm_err_all = int(_stk_df["_err"].sum())
                _pm_ok_all  = _pm_tot_all - _pm_err_all
                _pm_pct_all = round(_pm_ok_all / _pm_tot_all * 100, 1) if _pm_tot_all else 0.0

                _pc1, _pc2 = st.columns([1, 3])
                with _pc1:
                    _pm_kpi_color = ("#22c55e" if _pm_pct_all >= 95 else
                                    ("#f59e0b" if _pm_pct_all >= 75 else "#ef4444"))
                    st.markdown(
                        f'<div style="background:{_t["card"]};border:2px solid {_pm_kpi_color}33;'
                        f'border-radius:10px;padding:20px;text-align:center;">'
                        f'<div style="font-size:2.2rem;font-weight:800;color:{_pm_kpi_color};">'
                        f'{_pm_pct_all:.1f}%</div>'
                        f'<div style="font-size:0.85rem;color:{_t["muted"]};margin-top:4px;">'
                        f'{_pm_ok_all:,} / {_pm_tot_all:,} PMs</div>'
                        f'<div style="font-size:0.8rem;color:{_t["muted"]};">sin reincidencia ≤5d</div>'
                        f'<div style="font-size:0.75rem;color:{_t["muted"]};margin-top:8px;">'
                        f'Meta: ≥90%</div></div>',
                        unsafe_allow_html=True
                    )
                with _pc2:
                    if _cal_pm_k not in st.session_state:
                        _fig_pm_rc = go.Figure()
                        _txt_ok_pm = _stk_df.apply(lambda r:
                            f"<b>{r['_pct_ok']:.1f}%</b><br><span style='font-size:10px;'>"
                            f"{int(r['_ok'])}/{int(r['_total'])}</span>", axis=1).tolist()
                        _fig_pm_rc.add_trace(go.Bar(
                            x=_stk_df[_pm_x_col], y=_stk_df["_pct_ok"],
                            name="Sin reincidencia", marker_color="#22c55e", opacity=0.95,
                            text=_txt_ok_pm, textposition="inside", insidetextanchor="middle",
                            textfont=dict(size=14, color="#ffffff",
                                          family="Inter, system-ui, sans-serif"),
                        ))
                        _txt_err_pm = _stk_df["_pct_err"].apply(
                            lambda v: f"<b>{v:.1f}%</b>" if v >= 10 else "").tolist()
                        _fig_pm_rc.add_trace(go.Bar(
                            x=_stk_df[_pm_x_col], y=_stk_df["_pct_err"],
                            name="Correctivos ≤5d", marker_color="#ef4444", opacity=0.92,
                            text=_txt_err_pm, textposition="inside", insidetextanchor="middle",
                            textfont=dict(size=12, color="#ffffff"),
                        ))
                        for _, _row in _stk_df.iterrows():
                            if _row["_pct_err"] < 10 and _row["_pct_err"] > 0:
                                _fig_pm_rc.add_annotation(
                                    x=_row[_pm_x_col], y=100, ax=0, ay=-35,
                                    xref="x", yref="y", axref="pixel", ayref="pixel",
                                    showarrow=True, arrowhead=2, arrowsize=1, arrowwidth=1.5,
                                    arrowcolor="#ef4444",
                                    text=f"<b style='color:#ef4444;'>{_row['_pct_err']:.1f}%</b>",
                                    font=dict(size=11, color="#ef4444"),
                                    bgcolor="rgba(255,255,255,0.95)",
                                    bordercolor="#ef4444", borderwidth=1, borderpad=3,
                                )
                        _fig_pm_rc.update_layout(
                            barmode="stack", height=320,
                            margin=dict(t=40, b=30, l=10, r=20),
                            title=_pm_title,
                            yaxis=dict(range=[0, 115], ticksuffix="%", title="% preventivos"),
                            legend=dict(orientation="h", y=1.14, x=0), bargap=0.35,
                        )
                        _apply_plot_theme(_fig_pm_rc)
                        st.session_state[_cal_pm_k] = _fig_pm_rc
                    else:
                        _fig_pm_rc = st.session_state[_cal_pm_k]
                    st.plotly_chart(_fig_pm_rc, width="stretch")
            else:
                st.info("Sin datos para el período y filtros seleccionados.")

        if not _df_rc.empty or _n_bruto > 0:
            st.divider()

            _total_rc  = _df_rc["folio_cm"].nunique() if "folio_cm" in _df_rc.columns else len(_df_rc)
            _tec_rc    = int(_df_rc["es_reincidencia_tecnico"].sum()) if not _df_rc.empty else 0
            _espec_rc  = int((_df_rc["falla_tipo"] == "especial").sum()) if ("falla_tipo" in _df_rc.columns and not _df_rc.empty) else 0
            _sin_cl    = int((_df_rc["causa_clasif"] == "sin_clasificar").sum()) if not _df_rc.empty else 0

            # ── Fila 1: contexto general de PMs ──────────────────────────────
            _rp1, _rp2 = st.columns(2)
            _rp1.metric(
                "PMs realizados en el período",
                f"{_n_pm_total:,}",
                delta=f"Mantenimientos preventivos ejecutados",
                delta_color="off",
            )
            _rp2.metric(
                "PMs sin reincidencia ✅",
                f"{_n_pm_sin_reinc:,}",
                delta=f"{round(_n_pm_sin_reinc / _n_pm_total * 100, 1) if _n_pm_total > 0 else 0}% sin correctivo en 5 días",
                delta_color="off",
            )

            st.divider()

            # ── Fila 2: desglose de fallas ────────────────────────────────────
            _rm1, _rm2, _rm3, _rm4 = st.columns(4)
            _rm1.metric("Total fallas post-PM detectadas", f"{_n_bruto:,}",
                        delta=f"{_n_excl:,} Trab. Especiales excluidos — ver detalle ↓", delta_color="off")
            _rm2.metric("F.A.O + F.N.A.O — Error del técnico", f"{_tec_rc:,}",
                        delta="⚠️ Afecta KPI Calidad" if _tec_rc > 0 else "✅ Sin afectación",
                        delta_color="off")
            _rm3.metric("Trabajos Especiales", f"{_espec_rc:,}",
                        delta="No imputan KPI", delta_color="off")
            _rm4.metric("Sin causa registrada", f"{_sin_cl:,}",
                        delta="⚠️ Penaliza KPI Precisión" if _sin_cl > 0 else None,
                        delta_color="off")

            # ── Cuadro de trazabilidad completo ─────────────────────────────────
            with st.expander(
                f"🔍 Trazabilidad — desglose de las {_n_bruto:,} fallas post-PM detectadas",
                expanded=False,
            ):
                st.markdown(
                    f"""
| Clasificación (columna "Falla" en Fracttal) | Cant. | ¿Imputa KPI? | Criterio |
|---|---|---|---|
| 🔴 F.A.O — Falla Atribuible a Occimiano | **{_n_fao:,}** | ✅ Sí | Técnico confirmó error de Occimiano → imputa KPI |
| 🟠 F.N.A.O — No Atribuible a Occimiano | **{_n_fnao:,}** | ✅ Sí | Correctivo en ≤5 días = error del PM, independiente de clasificación |
| ⚫ Sin Información (opción "04.-") | **{_n_sin_info:,}** | ✅ Sí | Sin justificación → la duda recae en Occimiano |
| ⚫ Sin dato (campo completamente vacío) | **{_n_sin_dato:,}** | ✅ Sí | Sin justificación → la duda recae en Occimiano |
| 🔵 Trabajos Especiales | **{_n_espec:,}** | ⚠️ No | No es falla post-PM → no imputa KPI Calidad |
| **TOTAL detectadas** | **{_n_bruto:,}** | | |
| **TOTAL que imputan KPI** | **{_total_rc:,}** | | = F.A.O + F.N.A.O + Sin info + Sin dato (excl. Especiales y causa=cliente) |
| **TOTAL excluidos** | **{_n_excl:,}** | | = Solo Trabajos Especiales |
                    """,
                    unsafe_allow_html=False,
                )
                if _n_sin_dato > 0 or _n_sin_info > 0:
                    st.info(
                        f"ℹ️ {_n_sin_info + _n_sin_dato:,} OTs tienen el campo 'Falla' vacío o sin información. "
                        "Al no poder descartar responsabilidad de Occimiano, se imputan al técnico del PM. "
                        "Esto también penaliza el KPI de Precisión Fracttal."
                    )

            _df_rc_tec = _df_rc[_df_rc["es_reincidencia_tecnico"]] if not _df_rc.empty else pd.DataFrame()
            # ── Evolución mensual MC/MP — doble eje ──────────────────────────
            st.divider()
            _scope_lbl = (
                _tec_rc_sel if _tec_rc_sel != "Todos"
                else (_eq_rc if _eq_rc != "Todos" else "todos los equipos")
            )
            st.markdown(
                f'<div class="section-header">📈 Evolución mensual MC/MP ≤ 5 días — {_scope_lbl}</div>',
                unsafe_allow_html=True,
            )

            _cal_sig = f"{_wo_sig}_{_eq_rc}_{_tec_rc_sel}_{_mes_rc}_{_trim_rc}_{_sem_rc}"
            _cal_k   = f"_fig_cal_evol_{_current_theme}_{_cal_sig}"
            if _cal_k not in st.session_state:
                # PMs desde enero 2026 en adelante (KPI aplica a partir de 2026)
                _pm_m = df_wo[df_wo["maint_type"] == "Preventiva"].copy()
                _pm_m_dates = _pm_m["creation_date"].dt.tz_convert(None) \
                    if _pm_m["creation_date"].dt.tz is not None \
                    else _pm_m["creation_date"]
                _pm_m = _pm_m[_pm_m_dates >= pd.Timestamp("2026-01-01")]
                # Columna auxiliar para comparación tolerante de nombres de técnico
                _pm_m["_tech_norm"] = _pm_m["technician"].fillna("").apply(
                    lambda s: " ".join(_norm_n(s).split())
                )
                if _eq_rc != "Todos":
                    _grp_pm_evol = _LABEL_TO_GRUPO.get(_eq_rc, _eq_rc)
                    _pm_m = _pm_m[_pm_m["equipo"] == _grp_pm_evol]
                if _tec_rc_sel != "Todos":
                    _full_tec_evol   = TECH_NAME_MAP.get(_tec_rc_sel, _tec_rc_sel)
                    _full_tec_evol_n = " ".join(_norm_n(_full_tec_evol).split())
                    _pm_m = _pm_m[_pm_m["_tech_norm"] == _full_tec_evol_n]

                # ── Drill-down: 1 mes seleccionado → semanas; resto → meses ──────────
                _drill_weekly = (len(_mes_rc) == 1)

                if _drill_weekly:
                    # ── MODO SEMANAL ────────────────────────────────────────────────
                    _drm      = _mes_rc[0]                    # e.g. "2026-06"
                    _sems_ev  = _semanas_del_mes(_drm)        # [(label, start, end), ...]
                    _MESES_ES3 = {"01":"Ene","02":"Feb","03":"Mar","04":"Abr",
                                  "05":"May","06":"Jun","07":"Jul","08":"Ago",
                                  "09":"Sep","10":"Oct","11":"Nov","12":"Dic"}
                    _drm_p    = _drm.split("-")
                    _drm_lbl  = f"{_MESES_ES3.get(_drm_p[1], _drm_p[1])} '{_drm_p[0][2:]}"

                    # Filtrar PMs al mes seleccionado
                    _pm_dates3 = (_pm_m["creation_date"].dt.tz_convert(None)
                                  if _pm_m["creation_date"].dt.tz is not None
                                  else _pm_m["creation_date"])
                    _pm_m = _pm_m[_pm_dates3.dt.to_period("M").astype(str) == _drm].copy()
                    _pm_m["_fecha_d"] = (_pm_m["creation_date"].dt.tz_convert(None)
                                         if _pm_m["creation_date"].dt.tz is not None
                                         else _pm_m["creation_date"]).dt.date

                    # Asignador de semana
                    def _asigna_sem_ev(d):
                        for _sl, _ss, _se in _sems_ev:
                            if _ss <= d <= _se:
                                return _sl
                        return None

                    _pm_m["_slot"] = _pm_m["_fecha_d"].apply(_asigna_sem_ev)
                    _pm_by_slot = (_pm_m.dropna(subset=["_slot"])
                                   .groupby("_slot")["folio"].nunique()
                                   .reset_index(name="pms"))

                    # Errores por semana (usar _df_rc ya filtrado por equipo/tec/mes)
                    if (not _df_rc.empty and "fecha_cm_dt" in _df_rc.columns
                            and "folio_cm" in _df_rc.columns):
                        _rc_drill = _df_rc.copy()
                        _rc_drill["_fecha_d"] = pd.to_datetime(
                            _rc_drill["fecha_cm_dt"], errors="coerce"
                        ).dt.date
                        _rc_drill["_slot"] = _rc_drill["_fecha_d"].apply(_asigna_sem_ev)
                        _rc_by_slot = (_rc_drill.dropna(subset=["_slot"])
                                       .groupby("_slot")["folio_cm"].nunique()
                                       .reset_index(name="errores"))
                    else:
                        _rc_by_slot = pd.DataFrame(columns=["_slot", "errores"])

                    # Garantizar todas las semanas del mes (aunque tengan 0)
                    _all_slots = [_sl for _sl, _, _ in _sems_ev]
                    _ev = pd.DataFrame({"_slot": _all_slots})
                    _ev = _ev.merge(_pm_by_slot, on="_slot", how="left").fillna({"pms": 0})
                    _ev = _ev.merge(_rc_by_slot, on="_slot", how="left").fillna({"errores": 0})
                    _ev["pms"]     = _ev["pms"].astype(int)
                    _ev["errores"] = _ev["errores"].astype(int)
                    _ev["pct_err"] = (_ev["errores"] / _ev["pms"] * 100).round(2).where(_ev["pms"] > 0, 0)
                    _ev["pct_ok"]  = (100 - _ev["pct_err"]).round(2)
                    _ev["mes_lbl"] = _ev["_slot"]
                    _titulo_evol   = f"Evolución semanal — {_drm_lbl} · PMs ejecutados vs % error MC/MP"

                else:
                    # ── MODO MENSUAL ─────────────────────────────────────────────────
                    if _pm_m["creation_date"].dt.tz is not None:
                        _pm_m["_mes_pm"] = _pm_m["creation_date"].dt.tz_convert(None).dt.to_period("M").astype(str)
                    else:
                        _pm_m["_mes_pm"] = _pm_m["creation_date"].dt.to_period("M").astype(str)
                    # nunique("folio") = órdenes únicas por mes, no filas (1 OT puede tener N equipos)
                    _pm_by_mes = _pm_m.groupby("_mes_pm")["folio"].nunique().reset_index(name="pms")

                    # Errores únicos (folio_cm) por mes (desde _df_rc ya filtrado)
                    if not _df_rc.empty and "folio_cm" in _df_rc.columns and "mes" in _df_rc.columns:
                        _rc_by_mes = (
                            _df_rc.groupby("mes")["folio_cm"].nunique()
                            .reset_index(name="errores")
                            .rename(columns={"mes": "_mes_pm"})
                        )
                    else:
                        _rc_by_mes = pd.DataFrame(columns=["_mes_pm", "errores"])

                    # Unir por mes
                    _ev = _pm_by_mes.merge(_rc_by_mes, on="_mes_pm", how="left").fillna({"errores": 0})
                    _ev["errores"] = _ev["errores"].astype(int)
                    _ev["pct_err"] = (_ev["errores"] / _ev["pms"] * 100).round(2).where(_ev["pms"] > 0, 0)
                    _ev["pct_ok"]  = (100 - _ev["pct_err"]).round(2)
                    _ev = _ev.sort_values("_mes_pm")

                    # Etiquetas de mes en español
                    _MESES_ES2 = {"01":"Ene","02":"Feb","03":"Mar","04":"Abr","05":"May","06":"Jun",
                                  "07":"Jul","08":"Ago","09":"Sep","10":"Oct","11":"Nov","12":"Dic"}
                    def _mes_lbl2(ym):
                        p = str(ym).split("-")
                        return f"{_MESES_ES2.get(p[1], p[1])} '{p[0][2:]}" if len(p) == 2 else ym
                    _ev["mes_lbl"] = _ev["_mes_pm"].apply(_mes_lbl2)
                    _titulo_evol = "Evolución mensual — PMs ejecutados vs % error MC/MP"

                if not _ev.empty:
                    _fig_evol = make_subplots(specs=[[{"secondary_y": True}]])

                    # Barras: PMs realizados (eje izquierdo)
                    _fig_evol.add_trace(
                        go.Bar(
                            x=_ev["mes_lbl"], y=_ev["pms"],
                            name="PMs realizados",
                            marker_color="#94a3b8",
                            opacity=0.9,
                            text=_ev["pms"],
                            textposition="inside",
                            textfont=dict(size=13, color="#ffffff", family="Arial"),
                        ),
                        secondary_y=False,
                    )

                    # Línea tendencia PMs — amarillo
                    _fig_evol.add_trace(
                        go.Scatter(
                            x=_ev["mes_lbl"], y=_ev["pms"],
                            name="Tendencia PMs",
                            mode="lines+markers",
                            line=dict(color="#fbbf24", width=2.5, dash="dot"),
                            marker=dict(size=7, color="#fbbf24",
                                        line=dict(color="#ffffff", width=1)),
                        ),
                        secondary_y=False,
                    )

                    # Línea: % error (eje derecho) — rojo, sin texto (anotaciones separadas)
                    _fig_evol.add_trace(
                        go.Scatter(
                            x=_ev["mes_lbl"], y=_ev["pct_err"],
                            name="% Error MC/MP",
                            mode="lines+markers",
                            line=dict(color="#ef4444", width=3),
                            marker=dict(size=11, color="#ef4444",
                                        line=dict(color="#ffffff", width=2)),
                            customdata=list(zip(_ev["errores"], _ev["pms"])),
                            hovertemplate=(
                                "<b>%{x}</b><br>"
                                "% Error: %{y:.2f}%<br>"
                                "Errores: %{customdata[0]}<br>"
                                "PMs totales: %{customdata[1]}<br>"
                                "<extra></extra>"
                            ),
                        ),
                        secondary_y=True,
                    )
                    # Anotaciones con cuadro blanco bordeado en rojo — mismo patrón SLA
                    for _, _erow in _ev.iterrows():
                        _fig_evol.add_annotation(
                            x=_erow["mes_lbl"],
                            y=_erow["pct_err"],
                            yref="y2",
                            text=f"<b>{_erow['pct_err']:.2f}%</b>  {int(_erow['errores'])} err",
                            showarrow=False,
                            yanchor="bottom",
                            yshift=10,
                            font=dict(size=11, color="#b91c1c", family="Arial"),
                            bgcolor="rgba(255,255,255,0.88)",
                            bordercolor="#ef4444",
                            borderwidth=1.5,
                            borderpad=4,
                        )

                    # Línea: % exactitud (activable desde leyenda)
                    _fig_evol.add_trace(
                        go.Scatter(
                            x=_ev["mes_lbl"], y=_ev["pct_ok"],
                            name="% Exactitud",
                            mode="lines+markers",
                            line=dict(color="#22c55e", width=2, dash="dot"),
                            marker=dict(size=6, color="#22c55e"),
                            visible="legendonly",
                        ),
                        secondary_y=True,
                    )

                    _fig_evol.update_layout(
                        title=_titulo_evol,
                        height=430,
                        legend=dict(orientation="h", y=1.08, x=0),
                        margin=dict(t=60, b=20),
                        bargap=0.3,
                    )
                    _fig_evol.update_yaxes(title_text="PMs realizados", secondary_y=False)
                    _fig_evol.update_yaxes(
                        title_text="% Error MC/MP",
                        secondary_y=True,
                        tickformat=".2f",
                        ticksuffix="%",
                        range=[0, max(_ev["pct_err"].max() * 2, 1)],
                    )
                    _apply_plot_theme(_fig_evol)
                    st.session_state[_cal_k] = _fig_evol
                else:
                    st.session_state[_cal_k] = None

            _fig_evol_show = st.session_state.get(_cal_k)
            if _fig_evol_show is not None:
                st.plotly_chart(_fig_evol_show, width="stretch")
            else:
                st.info("Sin datos suficientes para mostrar la evolución.")

            st.divider()
            st.markdown('<div class="section-header">📋 Detalle de fallas post-preventiva</div>',
                        unsafe_allow_html=True)
            st.caption(
                "Cada fila = correctivo generado 1–5 días después de un preventivo en el mismo equipo. "
                "El técnico responsable es quien realizó el preventivo."
            )

            _rc_cols_base = [
                "equipment", "client", "station",
                "folio_pm", "folio_cm",
                "fecha_pm", "fecha_cm",
                "tecnico_resp_short", "tecnico_cm", "grupo_responsable",
                "dias_entre",
                "falla_raw", "falla_tipo",
                "causa_raiz", "causa_clasif",
                "es_reincidencia_tecnico",
                "comentario_tecnico_cm",
                "conclusion",
            ]
            # falla_raw / falla_tipo solo existen si el DF fue generado con la nueva lógica
            _rc_disp = _df_rc[[c for c in _rc_cols_base if c in _df_rc.columns]].copy()
            _rc_disp["fecha_pm"] = pd.to_datetime(_rc_disp["fecha_pm"]).dt.strftime("%d/%m/%Y")
            _rc_disp["fecha_cm"] = pd.to_datetime(_rc_disp["fecha_cm"]).dt.strftime("%d/%m/%Y")

            # ── Limpiar prefijo redundante del comentario STO ────────────────
            if "comentario_tecnico_cm" in _rc_disp.columns:
                import re as _re_sto
                _rc_disp["comentario_tecnico_cm"] = _rc_disp["comentario_tecnico_cm"].apply(
                    lambda x: _re_sto.sub(
                        r"^DESCRIPCI[OÓ]N DE LA FALLA ENCONTRADA:\s*", "", str(x or ""), flags=_re_sto.IGNORECASE
                    ).strip()
                )

            # ── Columna Conclusión — generada con valores raw de falla_tipo ────
            _CONCL_MAP = {
                "fao":      "Error del técnico del PM — falla confirmada como F.A.O",
                "sin_info": "⚠️ Atribuible por omisión: técnico seleccionó 'Sin Información'",
                "sin_dato": "⚠️ Atribuible por omisión: campo 'Falla' vacío en Fracttal",
                "especial": "Trabajo Especial — no imputa KPI Calidad",
                "fnao":     "Excluida — F.N.A.O declarada por el técnico",
            }
            if "falla_tipo" in _rc_disp.columns:
                _rc_disp["conclusion"] = _rc_disp["falla_tipo"].map(_CONCL_MAP).fillna("")

            _rc_disp["es_reincidencia_tecnico"] = _rc_disp["es_reincidencia_tecnico"].apply(
                lambda x: "🔴 Técnico" if x else "🟢 No técnico"
            )
            _rc_disp["causa_clasif"] = _rc_disp["causa_clasif"].apply(
                lambda x: "🔴 Sin causa" if x == "sin_clasificar"
                else ("🟡 Técnico" if x == "tecnico" else "🟢 Cliente")
            )
            if "falla_tipo" in _rc_disp.columns:
                _rc_disp["falla_tipo"] = _rc_disp["falla_tipo"].map({
                    "fao":      "🔴 F.A.O",
                    "fnao":     "🟢 F.N.A.O",
                    "sin_info": "⚫ Sin info",
                    "especial": "🔵 Esp.",
                    "sin_dato": "⚫ Sin dato",
                }).fillna(_rc_disp["falla_tipo"])
            # Mapear grupo interno → label de equipo (único incluso con mismo senior)
            _rc_disp["grupo_responsable"] = _rc_disp["grupo_responsable"].map(
                _EQUIPO_LABEL
            ).fillna(_rc_disp["grupo_responsable"])

            _rc_col_names = [
                "Equipo", "Cliente", "Estación",
                "OT MP", "OT MC",
                "Fecha MP", "Fecha MC",
                "Técnico MP", "Técnico MC", "T.Senior",
                "Días MP→MC",
            ]
            if "falla_raw" in _rc_disp.columns:
                _rc_col_names += ["Tipo Falla", "Clasif. Falla"]
            _rc_col_names += ["Causa raíz", "Clasif. causa", "Responsabilidad"]
            if "comentario_tecnico_cm" in _rc_disp.columns:
                _rc_col_names += ["Comentario STO"]
            if "conclusion" in _rc_disp.columns:
                _rc_col_names += ["Conclusión"]
            _rc_disp.columns = _rc_col_names

            _show_df(
                _rc_disp,
                width="stretch",
                hide_index=True,
                column_config={
                    "Días MP→MC":    st.column_config.NumberColumn(format="%d días"),
                    "Tipo Falla":    st.column_config.TextColumn(width=120),
                    "Clasif. Falla": st.column_config.TextColumn(width=100),
                    "Causa raíz":    st.column_config.TextColumn(width=220),
                    "Técnico MC":    st.column_config.TextColumn(width=160),
                    "Técnico MP":    st.column_config.TextColumn(width=160),
                    "Comentario STO": st.column_config.TextColumn(
                        width=340,
                        help="Comentario del técnico en el formulario del correctivo — descripción de la falla encontrada",
                    ),
                    "Conclusión":    st.column_config.TextColumn(
                        width=300,
                        help="Veredicto del sistema según la clasificación F.A.O / F.N.A.O registrada",
                    ),
                },
            )

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 3 — PRECISIÓN INFO OPERATIVA (KPI LLENADO)
    # ══════════════════════════════════════════════════════════════════════════
    with tab_prec:
        _tprec_desc, _tprec_ley = st.columns([3, 1])
        with _tprec_ley:
            st.markdown(
                f'<div style="background:{_t["card"]};border:1px solid {_t["border"]};'
                f'border-radius:8px;padding:10px 12px;font-size:0.76rem;color:{_t["text"]};'
                f'line-height:1.75;margin-top:2px;">'
                f'<div style="font-weight:700;color:{_t["muted"]};font-size:0.78rem;'
                f'margin-bottom:4px;letter-spacing:0.04em;">📊 ESCALA BONO PRECISIÓN</div>'
                f'<span style="color:#22c55e;">≥ 95%</span> → <b>100%</b> · $105.000/trim<br>'
                f'<span style="color:#16a34a;">≥ 90%</span> → <b>90%</b> · $94.500/trim<br>'
                f'<span style="color:#4ade80;">≥ 85%</span> → <b>80%</b> · $84.000/trim<br>'
                f'<span style="color:#65a30d;">≥ 80%</span> → <b>70%</b> · $73.500/trim<br>'
                f'<span style="color:#f59e0b;">≥ 75%</span> → <b>60%</b> · $63.000/trim<br>'
                f'<span style="color:#f97316;">≥ 70%</span> → <b>50%</b> · $52.500/trim<br>'
                f'<span style="color:#ef4444;">&lt; 70%</span> → <b>Sin bono</b><br>'
                f'<div style="border-top:1px solid {_t["border"]};margin:5px 0;padding-top:4px;'
                f'font-size:0.72rem;color:{_t["muted"]};">'
                f'3 componentes × 25 pts:<br>⏱ Tiempo · 🔍 Causa · 🔢 Numeral</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
        with _tprec_desc:
            st.markdown(
                f'<div style="background:{_t["warn_bg"]};border-left:4px solid #01798A;'
                f'border-radius:8px;padding:12px 16px;margin-bottom:12px;color:{_t["text"]};">'
                '<b>KPI Precisión Fracttal</b> — Representa el <b>30% del bono de desempeño</b> '
                '(<b>$105.000 bruto/trimestre</b> máximo, pago trimestral). '
                'Mide <b>3 componentes</b> por OT (25 pts c/u = 75 total): '
                '<b>Tiempo de ejecución</b> (solo MP), '
                '<b>Causa raíz</b> (solo MC — MP no responde a falla) y '
                '<b>Numeral registrado</b> (MC+MP en lavadoras/aspiradoras). '
                f'<span style="font-size:0.82rem;color:{_t["muted"]};">'
                'Una OT es "mala" si falla en <b>cualquiera</b> de los componentes que le aplican. '
                'El bono se mide por <b>% de OTs buenas</b>, no por suma de errores.</span></div>',
                unsafe_allow_html=True,
            )

        # ── Construir DataFrame de KPI (cacheado en session_state) ──────────────
        # Solo se recalcula cuando cambia raw_wo (nueva carga desde Fracttal).
        # Se añade columna "equipo" al cachear para evitar apply(_get_equipo) repetidos.
        # Filtrado de excluidos y columnas derivadas también se calculan aquí →
        # no se repiten en cada rerun (solo al cambiar raw_wo).
        def _build_kpi_raw_cached():
            _df = _cached_build_kpi_llenado(raw_wo)
            if not _df.empty:
                _df["equipo"] = _df["tecnico"].apply(_get_equipo)
                _df = _df[~_df["tecnico"].apply(_es_excluido)].copy()
                # Solo tipos de mantenimiento evaluables: Correctiva y Preventiva (cualquier variante).
                # Excluidos: Entrega de Insumos, Inspección, Pendiente, Garantía, etc.
                _tipo_upper = _df["maint_type"].str.upper()
                _df = _df[
                    _tipo_upper.str.contains("CORRECTIVA", na=False) |
                    _tipo_upper.str.contains("PREVENTIVA", na=False)
                ].copy()
                # Solo datos desde 2026 — el KPI de Precisión aplica a partir de ese año
                _df = _df[_df["creation_date"].dt.tz_convert(None).dt.year >= 2026].copy()
                _df["mes"] = (
                    _df["creation_date"].dt.tz_convert(None)
                    .dt.to_period("M").astype(str)
                )
                _df["creation_date_local"] = (
                    _df["creation_date"].dt.tz_convert(None).dt.date
                )
            return _df
        try:
            df_kpi_raw = _sc("df_kpi_raw_v3_equipo", _wo_sig, _build_kpi_raw_cached)
        except Exception as _e_prec:
            st.error(f"⚠️ Error al cargar datos de KPI de llenado: {_e_prec}")
            st.exception(_e_prec)
            st.stop()

        if df_kpi_raw.empty:
            st.error("No se pudieron cargar datos de work orders para el KPI.")
            st.stop()

        # ── Filtros de mes disponibles ────────────────────────────────────────────
        meses_disp = sorted(df_kpi_raw["mes"].dropna().unique(), reverse=True)

        if not meses_disp:
            st.warning("No hay OTs con fecha de creación disponibles para calcular el KPI.")
            st.stop()

        # ── Pre-computar scores de TODAS las OTs una sola vez por carga ──────────
        # Cada cambio de filtro (mes/semana/equipo/técnico) solo hace un slice
        # sobre este DataFrame ya calculado → órdenes de magnitud más rápido.
        # Cargar el desglose por subtarea (1 fila por OT × activo con numeral).
        # Vacío si la tabla aún no existe o no se corrió el sync.
        df_num_sub = _sc("df_num_sub_v1", _wo_sig, load_numerales_subtarea_supabase)

        _ESTADOS_NO_PUNTUAN = {
            "Canceladas", "Cancelado", "Cancelada",
            "ERROR DE INGRESO",
            "EQUIPO CON RECAMBIO",
            "DUPLICADO", "Duplicidad",
            "DE PRUEBA",
            "FUE REPETIDA EN OTRA OS",
            "PLAN INCOMPLETO",
        }

        @st.cache_data(ttl=1800, show_spinner=False)
        def _load_ot_estados() -> dict:
            """Consulta directa a Supabase: {id_ot: estado} para filtrar OTs no operativas."""
            from supabase_client import _query
            _rows = _query("ordenes_trabajo", "select=id_ot,estado&fecha_creacion=gte.2026-01-01", limit=20_000)
            return {r["id_ot"]: r.get("estado", "") for r in _rows if r.get("id_ot")}

        _ot_estados = _load_ot_estados()
        _folios_excluir = {f for f, e in _ot_estados.items() if e in _ESTADOS_NO_PUNTUAN}

        def _build_ot_all():
            _df = score_llenado_por_ot(df_kpi_raw)
            if not _df.empty:
                # Excluir OTs con estado no operativo (consulta directa a Supabase)
                if _folios_excluir and "folio" in _df.columns:
                    _n_antes = len(_df)
                    _df = _df[~_df["folio"].isin(_folios_excluir)].copy()
                # Agregar estado como columna para visualización
                _df["wo_status"] = _df["folio"].map(_ot_estados).fillna("")
                # Recalcular numeral_ok/motivo a partir de las subtareas
                # cuando estén disponibles. Una OT pasa solo si TODAS sus
                # subtareas-numeral pasan; el motivo es el más severo.
                _df = aplicar_numerales_subtarea(_df, df_num_sub)
                _df["score_numeral"] = _df["numeral_ok"].apply(lambda ok: 25 if ok else 0)
                # Recalcular score_total con la nueva columna
                _df["score_total"] = (
                    _df["score_tiempo"] + _df["score_causa"] + _df["score_numeral"]
                ).clip(upper=75).round(1)
                _df["equipo"] = _df["tecnico"].apply(_get_equipo)
                _df["mes"] = (
                    _df["creation_date"].dt.tz_convert(None)
                    .dt.to_period("M").astype(str)
                )
                _df["creation_date_local"] = (
                    _df["creation_date"].dt.tz_convert(None).dt.date
                )
            return _df
        df_ot_all = _sc("df_ot_all_scores_v5_equipo", _wo_sig, _build_ot_all)

        _meses_prec_nums = {int(str(m).split("-")[1]) for m in meses_disp if "-" in str(m)}
        _TRIMESTRES_PREC = {
            "T1 · Ene–Mar": [1,2,3], "T2 · Abr–Jun": [4,5,6],
            "T3 · Jul–Sep": [7,8,9], "T4 · Oct–Dic": [10,11,12],
        }
        _trim_opts_prec = ["Todos"] + [
            k for k, v in _TRIMESTRES_PREC.items() if any(m in v for m in _meses_prec_nums)
        ]
        kf0, kf1, kf2, kf3, kf4 = st.columns([1.5, 1.4, 1.5, 2, 2])
        with kf0:
            _trim_prec = st.selectbox("Período", _trim_opts_prec, key="kpi_trim")
        with kf1:
            if _trim_prec != "Todos":
                _trim_months_prec = _TRIMESTRES_PREC[_trim_prec]
                _meses_disp_filt = [
                    m for m in meses_disp
                    if "-" in str(m) and int(str(m).split("-")[1]) in _trim_months_prec
                ]
            else:
                _meses_disp_filt = meses_disp
            _meses_base = list(_meses_disp_filt if _meses_disp_filt else meses_disp)
            _MESES_FULL_PREC = {
                "01":"Enero","02":"Febrero","03":"Marzo","04":"Abril","05":"Mayo","06":"Junio",
                "07":"Julio","08":"Agosto","09":"Septiembre","10":"Octubre","11":"Noviembre","12":"Diciembre",
            }
            def _ym_prec_lbl(ym):
                p = str(ym).split("-")
                return f"{_MESES_FULL_PREC.get(p[1], p[1])} {p[0][2:]}" if len(p) == 2 else ym
            _lbl_to_per_prec = {_ym_prec_lbl(m): m for m in meses_disp}
            _meses_base_lbl  = [_ym_prec_lbl(m) for m in _meses_base]
            _meses_sel_lbl = st.multiselect(
                "Mes a evaluar",
                _meses_base_lbl,
                default=[],
                key="kpi_mes",
                placeholder="Todos los meses",
            )
            _meses_sel_raw = [_lbl_to_per_prec[l] for l in _meses_sel_lbl if l in _lbl_to_per_prec]
            # Lista de meses seleccionados; vacío = Todos
            _meses_prec = _meses_sel_raw if _meses_sel_raw else _meses_base
            mes_sel = _meses_prec[0] if len(_meses_prec) == 1 else "Todos"
            # Helper local para etiqueta de mes (p.ej. "2026-05" → "May '26")
            _MN_PREC = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
                        7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}
            def _prec_m2l(ym):
                p = str(ym).split("-")
                return f"{_MN_PREC.get(int(p[1]),p[1])} '{p[0][2:]}" if len(p)==2 else str(ym)
            # Etiqueta legible para display
            if not _meses_sel_raw:
                _mes_lbl_prec = "Todos los meses"
            elif len(_meses_sel_raw) <= 3:
                _mes_lbl_prec = ", ".join([_prec_m2l(m) for m in _meses_sel_raw])
            else:
                _mes_lbl_prec = f"{len(_meses_sel_raw)} meses seleccionados"

        with kf2:
            # Semana solo disponible cuando se selecciona 1 solo mes
            if len(_meses_sel_raw) == 1:
                _sems_prec = _semanas_del_mes(_meses_prec[0])
                _sem_prec_opts = ["Todas"] + [s[0] for s in _sems_prec]
            else:
                _sems_prec = []
                _sem_prec_opts = ["Todas"]
            _sem_prec = st.selectbox("Semana", _sem_prec_opts, key="kpi_semana",
                                     disabled=(len(_meses_sel_raw) != 1))
        with kf3:
            # Usar _EQUIPO_LABEL — labels: "Carlos Avila" (Coquimbo), "Luis Lopez" (Concepción)
            _eq_prec_opts = ["Todos"] + list(_EQUIPO_LABEL.values())
            equipo_kpi = st.selectbox("Equipo", _eq_prec_opts, key="kpi_equipo")
        with kf4:
            if equipo_kpi != "Todos":
                _grp_kpi = _LABEL_TO_GRUPO.get(equipo_kpi)
                # IMPORTANTE: usar nombres completos desde df_kpi_raw (misma fuente que el filtro).
                # Antes se usaban los nombres cortos de GRUPOS_TERRENO["miembros"] (ej "Jorge Rodriguez"),
                # pero df_ot_scores["tecnico"] contiene nombres completos ("Jorge Raúl  Rodríguez Fuentes")
                # → el filtro nunca coincidía → "No hay OTs cerradas".
                _tec_kpi_opts = ["Todos"] + sorted(
                    t for t in df_kpi_raw[df_kpi_raw["equipo"] == _grp_kpi]["tecnico"].dropna().unique()
                    if not _es_excluido(t)
                )
            else:
                _tec_kpi_opts = ["Todos"] + sorted(
                    t for t in df_kpi_raw["tecnico"].dropna().unique()
                    if not _es_excluido(t) and _get_equipo(t) != "Sin equipo"
                )
            tec_kpi_sel = st.selectbox("Técnico", _tec_kpi_opts, key="kpi_tecnico")

        # ── Filtrar scores pre-computados (sin recalcular) ───────────────────────
        # df_ot_all ya tiene scores calculados para todas las OTs.
        # Solo se hace un slice en memoria → cambios de filtro son instantáneos.
        _meses_prec_str = [str(m) for m in _meses_prec]
        df_ot_scores = df_ot_all[
            df_ot_all["mes"].astype(str).isin(_meses_prec_str)
        ].copy()
        if _sem_prec != "Todas":
            _sem_prec_match = next((s for s in _sems_prec if s[0] == _sem_prec), None)
            if _sem_prec_match:
                _sem_prec_start, _sem_prec_end = _sem_prec_match[1], _sem_prec_match[2]
                df_ot_scores = df_ot_scores[
                    (df_ot_scores["creation_date_local"] >= _sem_prec_start) &
                    (df_ot_scores["creation_date_local"] <= _sem_prec_end)
                ]
        if equipo_kpi != "Todos":
            _grp_kpi = _LABEL_TO_GRUPO.get(equipo_kpi)
            if _grp_kpi:
                df_ot_scores = df_ot_scores[df_ot_scores["equipo"] == _grp_kpi]
        # Snapshot pre-filtro técnico: para tarjeta equipo de seniors
        _df_ot_pre_tec = df_ot_scores
        if tec_kpi_sel != "Todos":
            df_ot_scores = df_ot_scores[df_ot_scores["tecnico"] == tec_kpi_sel]

        df_tec_scores = score_llenado_por_tecnico(df_ot_scores)

        # Añadir columna "equipo" a df_tec_scores para poder agrupar por equipo en overrides
        # (misma lógica robusta que las tarjetas equipo)
        if not df_tec_scores.empty:
            df_tec_scores["equipo"] = df_tec_scores["tecnico"].apply(_get_equipo)

        # ── df_tec_scores_rank: versión para rankings individuales donde los seniors
        #    muestran el promedio de su equipo completo (no solo sus propios casos).
        #    df_tec_scores (sin modificar) sigue usándose para tarjetas por equipo,
        #    donde el equipo ya agrega todos sus miembros correctamente.
        df_tec_scores_rank = df_tec_scores.copy()
        if not df_tec_scores_rank.empty:
            for _snr in SENIORS:
                _snr_full_r = TECH_NAME_MAP.get(_snr, _snr)
                _snr_idx_r = df_tec_scores_rank.index[df_tec_scores_rank["tecnico"] == _snr_full_r]
                if len(_snr_idx_r) == 0:
                    continue
                # Usar columna "equipo" para matching robusto (igual que tarjetas equipo)
                _team_rows_r = df_tec_scores[df_tec_scores["equipo"] == _snr]
                if len(_team_rows_r) <= 1:
                    continue  # solo el senior mismo, sin compañeros → no hay nada que agregar
                _si_r      = _snr_idx_r[0]
                _tot_ots_r = int(_team_rows_r["ots_evaluadas"].sum())
                _w_r       = _team_rows_r["ots_evaluadas"]  # pesos para promedios ponderados

                def _wsum_r(col):
                    return int(_team_rows_r[col].sum()) if col in _team_rows_r else 0

                def _wavg_r(col):
                    return round((_team_rows_r[col] * _w_r).sum() / _tot_ots_r, 1) \
                        if _tot_ots_r > 0 and col in _team_rows_r else 0.0

                _n_err_r = _wsum_r("n_errores")
                df_tec_scores_rank.at[_si_r, "ots_evaluadas"]        = _tot_ots_r
                df_tec_scores_rank.at[_si_r, "n_errores"]            = _n_err_r
                df_tec_scores_rank.at[_si_r, "ots_correctas"]        = _tot_ots_r - _n_err_r
                df_tec_scores_rank.at[_si_r, "correctivas"]          = _wsum_r("correctivas")
                df_tec_scores_rank.at[_si_r, "sin_causa"]            = _wsum_r("sin_causa")
                df_tec_scores_rank.at[_si_r, "causa_tecnico"]        = _wsum_r("causa_tecnico")
                df_tec_scores_rank.at[_si_r, "causa_cliente"]        = _wsum_r("causa_cliente")
                df_tec_scores_rank.at[_si_r, "err_tiempo"]           = _wsum_r("err_tiempo")
                df_tec_scores_rank.at[_si_r, "err_causa"]            = _wsum_r("err_causa")
                df_tec_scores_rank.at[_si_r, "err_numeral"]          = _wsum_r("err_numeral")
                df_tec_scores_rank.at[_si_r, "err_deteccion"]        = _wsum_r("err_deteccion")
                df_tec_scores_rank.at[_si_r, "err_total_dim"]        = (
                    _wsum_r("err_tiempo") + _wsum_r("err_causa") + _wsum_r("err_numeral")
                )
                df_tec_scores_rank.at[_si_r, "tiempo_ok_count"]      = _tot_ots_r - _wsum_r("err_tiempo")
                df_tec_scores_rank.at[_si_r, "causa_ok_count"]       = _tot_ots_r - _wsum_r("err_causa")
                df_tec_scores_rank.at[_si_r, "numeral_ok_count"]     = _tot_ots_r - _wsum_r("err_numeral")
                df_tec_scores_rank.at[_si_r, "deteccion_ok_count"]   = _tot_ots_r - _wsum_r("err_deteccion")
                df_tec_scores_rank.at[_si_r, "score_promedio"]       = _wavg_r("score_promedio")
                df_tec_scores_rank.at[_si_r, "score_tiempo_prom"]    = _wavg_r("score_tiempo_prom")
                df_tec_scores_rank.at[_si_r, "score_causa_prom"]     = _wavg_r("score_causa_prom")
                df_tec_scores_rank.at[_si_r, "score_numeral_prom"]   = _wavg_r("score_numeral_prom")
                df_tec_scores_rank.at[_si_r, "score_deteccion_prom"] = _wavg_r("score_deteccion_prom")
                df_tec_scores_rank.at[_si_r, "pct_tiempo_ok"]        = _wavg_r("pct_tiempo_ok")
                df_tec_scores_rank.at[_si_r, "pct_causa_ok"]         = _wavg_r("pct_causa_ok")
                df_tec_scores_rank.at[_si_r, "pct_numeral_ok"]       = _wavg_r("pct_numeral_ok")
                df_tec_scores_rank.at[_si_r, "pct_deteccion_ok"]     = _wavg_r("pct_deteccion_ok")
                _exactitud_r = round((1 - _n_err_r / _tot_ots_r) * 100, 1) if _tot_ots_r > 0 else 100.0
                df_tec_scores_rank.at[_si_r, "exactitud_pct"]        = _exactitud_r
                _bp_r, _bl_r, _, _bclp_r = _bono_prec(_exactitud_r)
                df_tec_scores_rank.at[_si_r, "bono_pct"]    = _bp_r
                df_tec_scores_rank.at[_si_r, "bono_semanal"] = _bclp_r
                df_tec_scores_rank.at[_si_r, "umbral_bono"] = _bp_r > 0

        st.divider()

        # ── Tarjetas de bono por equipo (sin título, justo bajo los filtros) ─────

        def _render_prec_card(col_st, titulo, subtitulo, tecs_df, theme):
            """Genera el HTML de una tarjeta de Precisión Fracttal."""
            _n = len(tecs_df)
            _llen  = int(tecs_df["ots_evaluadas"].sum()) if _n > 0 else 0
            _err   = int(tecs_df["n_errores"].sum())     if _n > 0 else 0
            _ok    = _llen - _err
            _pct   = (_ok / _llen * 100) if _llen > 0 else 100.0
            _etd = int(tecs_df["err_total_dim"].sum()) if "err_total_dim" in tecs_df.columns else _err
            _et = int(tecs_df["err_tiempo"].sum())    if "err_tiempo"    in tecs_df.columns else 0
            _ec = int(tecs_df["err_causa"].sum())     if "err_causa"     in tecs_df.columns else 0
            _en = int(tecs_df["err_numeral"].sum())   if "err_numeral"   in tecs_df.columns else 0
            _ed = int(tecs_df["err_deteccion"].sum()) if "err_deteccion" in tecs_df.columns else 0
            _bsum = int(tecs_df["bono_semanal"].sum()) if _n > 0 else 0
            _cbon = int(tecs_df["umbral_bono"].sum())  if _n > 0 else 0
            _bp, _bl, _bc, _bclp = _bono_prec(_pct)
            _bw = min(100, max(0, _pct))
            col_st.markdown(
                f'<div style="background:{theme["card"]};border:2px solid {_bc}33;'
                f'border-radius:8px;padding:12px 14px;text-align:center;">'
                f'<div style="font-weight:700;font-size:0.92rem;color:{theme["text"]};">{titulo}</div>'
                f'<div style="font-size:0.75rem;color:{theme["muted"]};margin-bottom:6px;">{subtitulo}</div>'
                f'<div style="font-size:2rem;font-weight:800;line-height:1.1;color:{_bc};">'
                f'{_pct:.1f}%</div>'
                f'<div style="font-size:0.68rem;font-weight:600;letter-spacing:0.05em;'
                f'text-transform:uppercase;color:{theme["muted"]};margin-bottom:6px;">cumplimiento</div>'
                f'<div style="background:{theme["prog_bg"]};border-radius:4px;height:8px;margin-bottom:6px;">'
                f'<div style="background:{_bc};width:{_bw:.0f}%;height:8px;border-radius:4px;"></div></div>'
                f'<div style="background:{_bc};color:#fff;border-radius:4px;'
                f'padding:3px 10px;font-size:0.82rem;font-weight:700;'
                f'margin:0 auto 8px auto;display:inline-block;">{_bl}</div>'
                f'<div style="border-top:1px solid {theme["muted"]}22;margin:6px 0;"></div>'
                f'<div style="display:flex;justify-content:center;gap:8px;margin-bottom:4px;">'
                f'<div style="text-align:center;">'
                f'<div style="font-size:1.0rem;font-weight:700;color:{theme["text"]};">{_llen}</div>'
                f'<div style="font-size:0.60rem;color:{theme["muted"]};">llenadas</div></div>'
                f'<div style="font-size:0.9rem;color:{theme["muted"]};padding-top:3px;">→</div>'
                f'<div style="text-align:center;">'
                f'<div style="font-size:1.0rem;font-weight:700;color:#ef4444;">{_err}</div>'
                f'<div style="font-size:0.60rem;color:{theme["muted"]};">con error</div></div>'
                f'<div style="font-size:0.9rem;color:{theme["muted"]};padding-top:3px;">+</div>'
                f'<div style="text-align:center;">'
                f'<div style="font-size:1.0rem;font-weight:700;color:#22c55e;">{_ok}</div>'
                f'<div style="font-size:0.60rem;color:{theme["muted"]};">correctas</div></div>'
                f'</div>'
                f'<div style="font-size:0.64rem;color:{theme["muted"]};margin-bottom:5px;">'
                f'Err.: {_etd} &nbsp;'
                f'(⏱{_et} 🔍{_ec} 🔢{_en} 🎯{_ed})</div>'
                f'<div style="font-size:0.68rem;color:{theme["muted"]};margin-top:2px;">'
                f'{_cbon}/{_n} con bono · ${_bsum:,.0f}/sem</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

        if not df_tec_scores.empty and not df_ot_scores.empty:
            # Técnico específico → 1 tarjeta (su equipo)
            # Equipo específico  → 1 tarjeta (ese equipo)
            # Todos              → 5 tarjetas (todos los equipos)
            if tec_kpi_sel != "Todos":
                _tec_grp_prec = _get_equipo(tec_kpi_sel)
                _EQUIPO_LABEL_PREC = {k: v for k, v in _EQUIPO_LABEL.items() if k == _tec_grp_prec} or _EQUIPO_LABEL
            elif equipo_kpi != "Todos":
                _grp_kpi_prec = _LABEL_TO_GRUPO.get(equipo_kpi)
                _EQUIPO_LABEL_PREC = {k: v for k, v in _EQUIPO_LABEL.items() if k == _grp_kpi_prec} or _EQUIPO_LABEL
            else:
                _EQUIPO_LABEL_PREC = _EQUIPO_LABEL

            # Detectar si el técnico seleccionado es un senior
            _tec_short_sel = {v: k for k, v in TECH_NAME_MAP.items()}.get(tec_kpi_sel, "")
            _es_senior_sel = tec_kpi_sel != "Todos" and _tec_short_sel in SENIORS

            if _es_senior_sel:
                # Senior seleccionado → 2 tarjetas: individual + promedio equipo
                _pgk_snr = _get_equipo(tec_kpi_sel)
                _pgl_snr = _EQUIPO_LABEL.get(_pgk_snr, _pgk_snr)
                _col_indiv, _col_equipo = st.columns(2)

                # Tarjeta 1: solo las OTs del senior
                _tec_indiv = df_tec_scores[df_tec_scores["tecnico"] == tec_kpi_sel]
                _render_prec_card(_col_indiv, tec_kpi_sel, "Registro individual", _tec_indiv, _t)

                # Tarjeta 2: equipo completo — usar datos PRE-filtro técnico
                _df_equipo_full = score_llenado_por_tecnico(
                    _df_ot_pre_tec[_df_ot_pre_tec["equipo"] == _pgk_snr]
                )
                _n_eq = len(_df_equipo_full)
                _render_prec_card(_col_equipo,
                    f"Equipo {_pgl_snr} ({_n_eq} téc.)",
                    "Promedio equipo · base bono senior",
                    _df_equipo_full, _t)
            else:
                # ── Banner GLOBAL (todos los equipos combinados) ───────────
                # Solo se muestra en la vista "Todos los equipos". Aparece
                # arriba en un banner distintivo (borde grueso + label GLOBAL)
                # para diferenciarlo visualmente de las tarjetas por equipo.
                if len(_EQUIPO_LABEL_PREC) > 1 and not df_tec_scores.empty:
                    _g_llen = int(df_tec_scores["ots_evaluadas"].sum())
                    _g_err  = int(df_tec_scores["n_errores"].sum())
                    _g_ok   = _g_llen - _g_err
                    _g_pct  = (_g_ok / _g_llen * 100) if _g_llen > 0 else 0.0
                    _g_et = int(df_tec_scores["err_tiempo"].sum())    if "err_tiempo"    in df_tec_scores.columns else 0
                    _g_ec = int(df_tec_scores["err_causa"].sum())     if "err_causa"     in df_tec_scores.columns else 0
                    _g_en = int(df_tec_scores["err_numeral"].sum())   if "err_numeral"   in df_tec_scores.columns else 0
                    _g_ed = int(df_tec_scores["err_deteccion"].sum()) if "err_deteccion" in df_tec_scores.columns else 0
                    _g_etd = _g_et + _g_ec + _g_en
                    _g_bp, _g_bl, _g_bc, _g_bclp = _bono_prec(_g_pct)
                    _g_bw = min(100, max(0, _g_pct))
                    _g_ntec = int(len(df_tec_scores))
                    _g_bsum = int(df_tec_scores["bono_semanal"].sum()) if "bono_semanal" in df_tec_scores.columns else 0
                    _g_cbon = int(df_tec_scores["umbral_bono"].sum())  if "umbral_bono"  in df_tec_scores.columns else 0

                    # Etiqueta de contexto del filtro activo
                    _g_ctx_bits = []
                    if _meses_prec_str:
                        _g_ctx_bits.append(f"{len(_meses_prec_str)} mes(es)")
                    if _sem_prec and _sem_prec != "Todas":
                        _g_ctx_bits.append(_sem_prec)
                    if _trim_prec and _trim_prec != "Todos":
                        _g_ctx_bits.append(_trim_prec)
                    _g_ctx = " · ".join(_g_ctx_bits) if _g_ctx_bits else "todo el período"

                    st.markdown(
                        f'<div style="background:linear-gradient(90deg, {_t["card"]} 0%, {_g_bc}11 100%);'
                        f'border:2px solid {_g_bc};border-left:6px solid {_g_bc};'
                        f'border-radius:8px;padding:12px 18px;margin-bottom:12px;'
                        f'display:flex;align-items:center;gap:20px;flex-wrap:wrap;">'
                        # Bloque 1: Label GLOBAL + %
                        f'<div style="flex:0 0 auto;text-align:left;">'
                        f'<div style="font-size:0.65rem;font-weight:800;letter-spacing:0.12em;'
                        f'text-transform:uppercase;color:{_g_bc};margin-bottom:2px;">🌐 GLOBAL</div>'
                        f'<div style="font-size:0.70rem;color:{_t["muted"]};line-height:1.1;">'
                        f'Todos los equipos<br>{_g_ntec} téc. · {_g_ctx}</div>'
                        f'</div>'
                        # Bloque 2: % Cumplimiento grande
                        f'<div style="flex:0 0 auto;text-align:center;border-left:1px solid {_t["muted"]}22;padding-left:20px;">'
                        f'<div style="font-size:2.2rem;font-weight:800;line-height:1;color:{_g_bc};">{_g_pct:.1f}%</div>'
                        f'<div style="font-size:0.62rem;font-weight:600;letter-spacing:0.05em;'
                        f'text-transform:uppercase;color:{_t["muted"]};">cumplimiento global</div>'
                        f'</div>'
                        # Bloque 3: Barra de progreso
                        f'<div style="flex:1 1 180px;min-width:150px;">'
                        f'<div style="background:{_t["prog_bg"]};border-radius:4px;height:10px;margin-bottom:4px;">'
                        f'<div style="background:{_g_bc};width:{_g_bw:.0f}%;height:10px;border-radius:4px;"></div></div>'
                        f'<div style="font-size:0.70rem;color:{_t["muted"]};">'
                        f'<b style="color:{_t["text"]};">{_g_llen:,}</b> llenadas · '
                        f'<b style="color:#ef4444;">{_g_err:,}</b> c/error · '
                        f'<b style="color:#22c55e;">{_g_ok:,}</b> correctas</div>'
                        f'</div>'
                        # Bloque 4: Desglose errores + bonos
                        f'<div style="flex:0 0 auto;text-align:right;border-left:1px solid {_t["muted"]}22;padding-left:20px;">'
                        f'<div style="background:{_g_bc};color:#fff;border-radius:4px;'
                        f'padding:3px 12px;font-size:0.78rem;font-weight:700;display:inline-block;margin-bottom:5px;">{_g_bl}</div>'
                        f'<div style="font-size:0.65rem;color:{_t["muted"]};line-height:1.4;">'
                        f'Err.: {_g_etd} (⏱{_g_et} 🔍{_g_ec} 🔢{_g_en} 🎯{_g_ed})<br>'
                        f'<b>{_g_cbon}/{_g_ntec}</b> con bono · <b>${_g_bsum:,.0f}</b>/sem</div>'
                        f'</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

                # Flujo normal: 1 tarjeta por equipo
                _prec_eq_cols = st.columns(len(_EQUIPO_LABEL_PREC))
                for _pi, (_pgk, _pgl) in enumerate(
                    zip(_EQUIPO_LABEL_PREC.keys(), _EQUIPO_LABEL_PREC.values())
                ):
                    _senior_prec = GRUPOS_TERRENO.get(_pgk, {}).get("senior", "")
                    _tecs_grp = df_tec_scores[
                        df_tec_scores["tecnico"].apply(_get_equipo) == _pgk
                    ] if not df_tec_scores.empty else pd.DataFrame()

                    if tec_kpi_sel != "Todos" and len(_tecs_grp) > 0:
                        _card_titulo    = tec_kpi_sel
                        _card_subtitulo = f"Equipo: {_pgl}"
                    else:
                        _card_titulo    = _pgl
                        _card_subtitulo = f"Senior: {_senior_prec}"

                    _render_prec_card(_prec_eq_cols[_pi], _card_titulo, _card_subtitulo, _tecs_grp, _t)
            st.caption(
                "**Llenadas** = OTs evaluadas en el período (base 100%)  ·  "
                "**Con error** = OTs con ≥1 componente malo  ·  "
                "**Cumplimiento** = OTs correctas / OTs llenadas  ·  "
                "**Err. individuales** = suma de fallos por dimensión (informativo, no mide el KPI)  ·  "
                "Escala (por técnico/trimestre): ≥95%→$105K · ≥90%→$94.5K · ≥85%→$84K · "
                "≥80%→$73.5K · ≥75%→$63K · ≥70%→$52.5K · <70%→$0"
            )

        st.divider()

        # ── KPIs del mes ──────────────────────────────────────────────────────────
        if df_ot_scores.empty:
            st.warning(f"No hay OTs cerradas en **{_mes_lbl_prec}** con los filtros aplicados.")
        else:
            total_ots_mes = len(df_ot_scores)
            score_global = df_ot_scores["score_total"].mean()
            # Tiempo: solo MP. Causa: solo MC (MP no responde a falla).
            # Numeral: toda lavadora/aspiradora (MC+MP) — el formulario lo exige en ambas.
            _df_mp_kpi = df_ot_scores[~df_ot_scores["es_correctiva"]] if "es_correctiva" in df_ot_scores.columns else df_ot_scores
            _df_mc_kpi = df_ot_scores[df_ot_scores["es_correctiva"]]  if "es_correctiva" in df_ot_scores.columns else pd.DataFrame()
            _df_lav_kpi = df_ot_scores[df_ot_scores["es_lavadora"]] if "es_lavadora" in df_ot_scores.columns else df_ot_scores
            pct_tiempo  = (_df_mp_kpi["score_tiempo"] >= 25).mean() * 100 if not _df_mp_kpi.empty else 0.0
            pct_causa   = _df_mc_kpi["causa_ok"].mean() * 100 if not _df_mc_kpi.empty else 0.0
            pct_numeral = _df_lav_kpi["numeral_ok"].mean() * 100  if not _df_lav_kpi.empty else 0.0
            tecnicos_con_bono = (df_tec_scores_rank["umbral_bono"]).sum() if not df_tec_scores_rank.empty else 0
            total_tecnicos = len(df_tec_scores_rank)

            color_global, lbl_global = _score_level(score_global)

            gk1, gk2, gk3, gk4, gk5 = st.columns(5)
            gk1.metric("Score global del mes", f"{score_global:.1f} / 100",
                       delta=lbl_global, delta_color="off")
            gk2.metric("OTs evaluadas", f"{total_ots_mes:,}")
            gk3.metric("Tiempo OK · solo MP (≥75%)", f"{pct_tiempo:.1f}%",
                       delta=f"{'✅' if pct_tiempo >= 75 else '⚠️'}")
            gk4.metric("Causa raíz OK · solo MC", f"{pct_causa:.1f}%",
                       delta=f"{'✅' if pct_causa >= 80 else '⚠️'}")
            gk5.metric("Técnicos con bono (≥90% exactitud)", f"{tecnicos_con_bono} / {total_tecnicos}")

            # ── Barra de puntaje global con desglose ──────────────────────────────
            st.divider()
            st.markdown('<div class="section-header">📊 Desglose del puntaje por dimensión</div>',
                        unsafe_allow_html=True)
            st.caption(
                f"Promedio mensual de **{total_ots_mes:,}** OTs  |  "
                f"Técnicos: **{total_tecnicos}**  |  Período: **{_mes_lbl_prec}**"
            )

            # Usar pct_* ya calculadas con el denominador correcto:
            #   pct_tiempo  → solo preventivas (MC tienen 25 auto, no cuentan)
            #   pct_causa   → solo correctivas (MP tienen 25 auto, no aplica)
            #   pct_numeral → toda lavadora/aspiradora (MC+MP) — el formulario lo exige
            dim_avg = {
                "⏱ Tiempo ejecución (25 pts)":   pct_tiempo  / 100 * 25,
                "🔍 Causa raíz (25 pts)":         pct_causa   / 100 * 25,
                "🔢 Numeral registrado (25 pts)": pct_numeral / 100 * 25,
            }
            dim_max = {k: 25 for k in dim_avg}
            dim_colors = ["#f59e0b", "#3b82f6", "#22c55e"]

            dim_df = pd.DataFrame([
                {
                    "Dimensión": k,
                    "Puntaje promedio": round(v, 1),
                    "Máximo": dim_max[k],
                    "% de máximo": round(v / dim_max[k] * 100, 1),
                }
                for k, v in dim_avg.items()
            ])

            _prec_sig = f"{_wo_sig}_{equipo_kpi}_{tec_kpi_sel}_{'|'.join(_meses_prec_str)}_{_sem_prec}_{_trim_prec}"
            dc1, dc2 = st.columns([3, 2])
            with dc1:
                _dim_k = f"_fig_prec_dim_{_current_theme}_{_prec_sig}"
                if _dim_k not in st.session_state:
                    fig_dim = px.bar(
                        dim_df, x="Puntaje promedio", y="Dimensión",
                        orientation="h",
                        color="Dimensión",
                        color_discrete_sequence=dim_colors,
                        text=dim_df["Puntaje promedio"].apply(lambda v: f"{v:.1f}"),
                        labels={"Puntaje promedio": "Puntaje promedio (escala de cada dimensión)"},
                    )
                    fig_dim.update_traces(textposition="outside")
                    fig_dim.update_layout(
                        showlegend=False, height=200,
                        margin=dict(t=10, b=10, l=10, r=60),
                        xaxis=dict(range=[0, 40]),
                        yaxis={"categoryorder": "array",
                               "categoryarray": list(reversed(list(dim_avg.keys())))},
                    )
                    _apply_plot_theme(fig_dim)
                    st.session_state[_dim_k] = fig_dim
                st.plotly_chart(st.session_state[_dim_k], width="stretch")
            with dc2:
                for row in dim_df.itertuples():
                    pct = row._4  # % de máximo
                    color = "#22c55e" if pct >= 99 else ("#f59e0b" if pct >= 90 else "#ef4444")
                    st.markdown(
                        f'<div style="margin-bottom:10px;">'
                        f'<div style="display:flex;justify-content:space-between;'
                        f'font-size:0.82rem;color:{_t["muted"]};margin-bottom:2px;">'
                        f'<span>{row.Dimensión}</span>'
                        f'<span style="color:{color};font-weight:700">{pct:.0f}%</span></div>'
                        f'<div style="background:{_t["prog_bg"]};border-radius:4px;height:10px;">'
                        f'<div style="background:{color};width:{min(pct,100):.0f}%;'
                        f'height:10px;border-radius:4px;"></div></div></div>',
                        unsafe_allow_html=True,
                    )

            # (Tarjetas de bono movidas arriba, justo bajo los filtros)

            # ══════════════════════════════════════════════════════════════════
            # SECCIÓN: CAUSA RAÍZ
            # ══════════════════════════════════════════════════════════════════
            st.divider()
            st.markdown('<div class="section-header">🔍  Causa Raíz — MC y MP</div>',
                        unsafe_allow_html=True)

            # ── Leyenda expandible ────────────────────────────────────────────
            with st.expander("📖  Leyenda de categorías válidas (F.N.A.O / F.A.O)", expanded=False):
                _leg1, _leg2 = st.columns(2)
                with _leg1:
                    st.markdown("**🟢 F.N.A.O — Falla No Atribuible a Occimiano**")
                    st.markdown("*Causas válidas (SI CORRESPONDE):*")
                    st.markdown("""
- `01.1.- DAÑO CAUSADO POR CLIENTE`
- `01.2.- MAL USO U OMISION EDS` *(sin sal, llave cerrada, etc.)*
- `01.3.- FICHERO (MOJADO/DAÑADO)`
- `02.3.- FICHERO / FALLA PROGRAMACION`
- `02.4.- REPUESTOS /OTROS` *(solo si la falla proviene del concesionario)*
- `03.1.- DAÑOS EN ESTRUCTURAS/GASFITERÍA/OOCC`
""")
                    st.markdown("*❌ No corresponde usar con F.N.A.O:*")
                    st.markdown("""
- `01.4.- REPUESTOS (DESGASTE)/OTROS` → debe ir en F.A.O
- `01.5.- ERROR 01 ELECTRICO` → debe ir en F.A.O
- `01.6.- ERROR 03 AGUA` → debe ir en F.A.O
- `01.7.- OTROS`, `02.7.- OTROS` → dato vago, no aceptado
- `02.2.- BY PASS / MOTORES` → debe ir en F.A.O
- `02.5.- ERROR 01 ELECTRICO` → debe ir en F.A.O
- `02.6.- ERROR 03 AGUA` → debe ir en F.A.O
- Números (147, 148, 150…) → no corresponde, dato no claro
""")
                with _leg2:
                    st.markdown("**🔴 F.A.O — Falla Atribuible a Occimiano**")
                    st.markdown("*Causas válidas para F.A.O:*")
                    st.markdown("""
- `01.4.- REPUESTOS (DESGASTE)/OTROS`
- `01.5.- ERROR 01 ELECTRICO`
- `01.6.- ERROR 03 AGUA`
- `02.2.- BY PASS / MOTORES`
- `02.5.- ERROR 01 ELECTRICO`
- `02.6.- ERROR 03 AGUA`
- `02.4.- REPUESTOS /OTROS` *(con falla comprobada)*
""")
                    st.markdown("**⚫ Siempre se considera ERROR:**")
                    st.markdown("""
- Campo **"Falla"** vacío o con `04.- SIN INFORMACION`
- Campo **"Causas de la Falla"** vacío o con `SIN CLASIFICAR`
- Causas con **solo números** (147, 148, 150, 160, etc.)
- Usar `01.7.- OTROS` o `02.7.- OTROS` sin especificar
""")

            # ── Helpers de mes compartidos por Causa Raíz y Tiempo ──────────
            import re as _re
            _MN2 = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
                    7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}
            def _m2l(ym):
                p = str(ym).split("-")
                return f"{_MN2.get(int(p[1]),p[1])} '{p[0][2:]}" if len(p)==2 else str(ym)

            # ── Agrupador que respeta la selección de meses/semanas ────────────
            # Regla: 1 mes seleccionado → desglose por semanas DENTRO de ese mes.
            #        ≥2 meses (o "Todos") → solo esos meses agrupados.
            def _agrupar_por_periodo(df_in, col_ok: str):
                """Devuelve DataFrame con columnas: bucket_lbl, ok, total, pct_ok, pct_err.
                df_in debe traer "mes" y "creation_date_local". col_ok = nombre de bool col."""
                if df_in.empty or col_ok not in df_in.columns:
                    return pd.DataFrame()
                df = df_in.copy()
                # ¿1 solo mes seleccionado? → semanas; si no, meses.
                un_mes = len(_meses_sel_raw) == 1
                if un_mes:
                    ym = _meses_sel_raw[0]
                    _sems = _semanas_del_mes(ym)
                    if "creation_date_local" not in df.columns:
                        return pd.DataFrame()
                    rows = []
                    for lbl, ini, fin in _sems:
                        m = ((df["creation_date_local"] >= ini) &
                             (df["creation_date_local"] <= fin))
                        sub = df[m]
                        tot = len(sub)
                        ok  = int(sub[col_ok].sum()) if tot else 0
                        # Etiqueta compacta "S1\n07-13/06" para el eje
                        n = lbl.split()[1]
                        rango = lbl.split("(")[1].rstrip(")")
                        rows.append(dict(bucket_lbl=f"S{n}<br>{rango}",
                                         ok=ok, total=tot))
                    g = pd.DataFrame(rows)
                else:
                    # Multi-mes: solo los meses que el usuario seleccionó
                    df = df[df["mes"].astype(str).isin(set(_meses_prec_str))]
                    if df.empty:
                        return pd.DataFrame()
                    g = (df.groupby("mes")
                           .agg(total=(col_ok,"count"), ok=(col_ok,"sum"))
                           .reset_index().sort_values("mes"))
                    g["bucket_lbl"] = g["mes"].apply(_m2l)
                # Excluir buckets sin datos (semanas futuras, meses sin OTs).
                # Sin esto, una semana con total=0 aparece como "100% error" porque
                # 100 - 0/1*100 = 100. Sólo mostrar buckets con al menos 1 OT.
                g = g[g["total"] > 0].copy()
                if g.empty:
                    return g
                g["pct_ok"]  = (g["ok"] / g["total"] * 100).round(1)
                g["pct_err"] = (100 - g["pct_ok"]).round(1)
                return g

            # ── Builder de gráfico apilado (mismo diseño que Numerales) ────────
            def _fig_apilada(g: pd.DataFrame, color_ok: str, label_ok: str,
                             label_err: str, meta: float = None,
                             titulo_y: str = "% OTs") -> "go.Figure":
                """Barras apiladas verde (OK) + rojo (Error). %OK CENTRADO en barra
                verde; %ERR fuera con flecha cuando la barra roja es pequeña.
                Si la barra roja es >=10% pone el % adentro."""
                fig = go.Figure()
                # Texto para barra verde: % centrado + fracción debajo
                txt_ok = g.apply(lambda r:
                    f"<b>{r['pct_ok']:.1f}%</b><br><span style='font-size:10px;'>"
                    f"{int(r['ok'])}/{int(r['total'])}</span>", axis=1).tolist()
                fig.add_trace(go.Bar(
                    x=g["bucket_lbl"], y=g["pct_ok"],
                    name=label_ok, marker_color=color_ok, opacity=0.95,
                    text=txt_ok, textposition="inside", insidetextanchor="middle",
                    textfont=dict(size=14, color="#ffffff",
                                  family="Inter, system-ui, sans-serif"),
                    customdata=g[["ok","total"]].values,
                    hovertemplate="%{x}<br>%{customdata[0]}/%{customdata[1]} OK"
                                  "<br>%{y:.1f}%<extra></extra>",
                ))
                # Barra roja: solo texto adentro si hay espacio (≥10%); si no, anotación externa con flecha
                txt_err_in = g["pct_err"].apply(
                    lambda v: f"<b>{v:.1f}%</b>" if v >= 10 else "").tolist()
                fig.add_trace(go.Bar(
                    x=g["bucket_lbl"], y=g["pct_err"],
                    name=label_err, marker_color="#ef4444", opacity=0.92,
                    text=txt_err_in, textposition="inside", insidetextanchor="middle",
                    textfont=dict(size=12, color="#ffffff"),
                    customdata=g[["ok","total","pct_err"]].values,
                    hovertemplate="%{x}<br>Error: %{customdata[2]:.1f}%<extra></extra>",
                ))
                # Flecha externa para errores pequeños (<10%)
                for _, r in g.iterrows():
                    if r["pct_err"] < 10 and r["pct_err"] > 0:
                        fig.add_annotation(
                            x=r["bucket_lbl"], y=100, ax=0, ay=-35,
                            xref="x", yref="y", axref="pixel", ayref="pixel",
                            showarrow=True, arrowhead=2, arrowsize=1, arrowwidth=1.5,
                            arrowcolor="#ef4444",
                            text=f"<b style='color:#ef4444;'>{r['pct_err']:.1f}% error</b>",
                            font=dict(size=11, color="#ef4444"),
                            bgcolor="rgba(255,255,255,0.95)",
                            bordercolor="#ef4444", borderwidth=1, borderpad=3,
                        )
                if meta is not None:
                    fig.add_hline(y=meta, line_dash="dash", line_color=color_ok,
                                  annotation_text=f"Meta {meta:.0f}%",
                                  annotation_position="top left", line_width=1.5)
                fig.update_layout(
                    barmode="stack", height=320,
                    margin=dict(t=40, b=30, l=10, r=20),
                    yaxis=dict(range=[0, 115], ticksuffix="%", title=titulo_y),
                    legend=dict(orientation="h", y=1.14, x=0), bargap=0.35,
                )
                _apply_plot_theme(fig)
                return fig

            # KPI card lateral (mismo diseño que el panel verde de Numerales)
            def _kpi_card(pct: float, ok: int, total: int, label: str,
                         meta_pct: float = 90.0) -> None:
                color = ("#22c55e" if pct >= meta_pct else
                         ("#f59e0b" if pct >= max(meta_pct - 20, 50) else "#ef4444"))
                st.markdown(
                    f'<div style="background:{_t["card"]};border:2px solid {color}33;'
                    f'border-radius:10px;padding:20px;text-align:center;">'
                    f'<div style="font-size:2.2rem;font-weight:800;color:{color};">'
                    f'{pct:.1f}%</div>'
                    f'<div style="font-size:0.85rem;color:{_t["muted"]};margin-top:4px;">'
                    f'{ok:,} / {total:,} OTs</div>'
                    f'<div style="font-size:0.8rem;color:{_t["muted"]};">{label}</div>'
                    f'<div style="font-size:0.75rem;color:{_t["muted"]};margin-top:8px;">'
                    f'Meta: ≥{meta_pct:.0f}%</div></div>',
                    unsafe_allow_html=True
                )

            # ── Gráfico Causa Raíz (diseño apilado igual que Numerales) ──────
            _periodo_lbl_cr = ("por semanas" if len(_meses_sel_raw) == 1
                               else "por mes seleccionado")
            st.markdown(f"**Evolución {_periodo_lbl_cr} — % correctivas con Causa Raíz correctamente llenada**")

            # Causa raíz solo aplica a correctivas. Las preventivas no responden
            # a una falla, así que se excluyen del análisis y del KPI.
            _df_cr_base = df_kpi_raw[df_kpi_raw["es_correctiva"]].copy() \
                if "es_correctiva" in df_kpi_raw.columns else df_kpi_raw.copy()

            if equipo_kpi != "Todos":
                _grp_cr = _LABEL_TO_GRUPO.get(equipo_kpi, equipo_kpi)
                _df_cr_base = _df_cr_base[_df_cr_base["equipo"] == _grp_cr]
            if tec_kpi_sel != "Todos":
                _df_cr_base = _df_cr_base[_df_cr_base["tecnico"] == tec_kpi_sel]

            if not _df_cr_base.empty and "causa_ok" in _df_cr_base.columns:
                _df_cr_base = _df_cr_base.copy()
                # Usar causa_ok ya calculada en data.py (MC: código Fracttal; MP: texto no vacío)
                _df_cr_base["_causa_ok"] = _df_cr_base["causa_ok"].fillna(False).astype(bool)
                # Subset al periodo seleccionado (para el KPI lateral y el gráfico)
                _df_cr_periodo = _df_cr_base[_df_cr_base["mes"].astype(str).isin(set(_meses_prec_str))].copy()
                if _sem_prec != "Todas":
                    _sem_match_cr = next((s for s in _sems_prec if s[0] == _sem_prec), None)
                    if _sem_match_cr and "creation_date_local" in _df_cr_periodo.columns:
                        _df_cr_periodo = _df_cr_periodo[
                            (_df_cr_periodo["creation_date_local"] >= _sem_match_cr[1]) &
                            (_df_cr_periodo["creation_date_local"] <= _sem_match_cr[2])
                        ]
                _cr_tot = len(_df_cr_periodo)
                _cr_ok  = int(_df_cr_periodo["_causa_ok"].sum()) if _cr_tot else 0
                _cr_pct = (_cr_ok / _cr_tot * 100) if _cr_tot else 0.0

                _g_cr = _agrupar_por_periodo(_df_cr_base, "_causa_ok")

                cc1, cc2 = st.columns([1, 3])
                with cc1:
                    _kpi_card(_cr_pct, _cr_ok, _cr_tot,
                              "con causa raíz correcta", meta_pct=80.0)
                with cc2:
                    if not _g_cr.empty:
                        _cr_sig = (f"_fig_cr_v2_{_current_theme}_{_wo_sig}_{equipo_kpi}"
                                   f"_{tec_kpi_sel}_{'-'.join(_meses_prec_str)}_{_sem_prec}")
                        if _cr_sig not in st.session_state:
                            st.session_state[_cr_sig] = _fig_apilada(
                                _g_cr, color_ok="#22c55e",
                                label_ok="Causa correcta", label_err="Sin causa / Vaga",
                                meta=80.0, titulo_y="% correctivas",
                            )
                        st.plotly_chart(st.session_state[_cr_sig], width="stretch")

                # ── Tabla detalle de Causa Raíz ───────────────────────────────
                _n_cr_total = len(_df_cr_base)
                with st.expander(f"📋 Detalle de OTs — Causa Raíz ({_n_cr_total:,} correctivas)", expanded=False):
                    _filtro_cr = _filtro_ot_input("kpi_filtro_ot_causa")
                    _det_cr = _df_cr_base[[c for c in
                        ["folio","equipment_code","eds_occim","tecnico","creation_date","maint_type",
                         "causa_raiz_raw","causa_clasif","comentario_tecnico","_causa_ok",
                         "es_correctiva"]
                        if c in _df_cr_base.columns]].copy()
                    _det_cr["creation_date"] = pd.to_datetime(_det_cr["creation_date"], errors="coerce")\
                        .dt.tz_convert(None).dt.strftime("%d/%m/%Y")
                    _det_cr["Estado"] = _det_cr["_causa_ok"].apply(
                        lambda v: "✅ Correcto" if v else "❌ Error")

                    # Diagnóstico breve del error de llenado (solo cuando _causa_ok=False)
                    # Explica por qué se imputa el error; vacío cuando está correcto.
                    import re as _re_diag
                    _DIAG_PREFIJO = _re_diag.compile(r"^0[1-4]\.\d", _re_diag.IGNORECASE)
                    def _diagnostico_causa(row) -> str:
                        if bool(row.get("_causa_ok", False)):
                            return ""
                        causa = str(row.get("causa_raiz_raw","") or "").strip()
                        com   = str(row.get("comentario_tecnico","") or "").strip()
                        com   = "" if com in ("—", "-") else com
                        es_mc = bool(row.get("es_correctiva", True))
                        tiene_com = bool(com) and len(com) > 5
                        causa_vacia = (not causa) or causa.upper() in ("SIN CLASIFICAR","N/A","NA")
                        # Caso PM: el form exige cualquier texto, igual quedó vacío
                        if not es_mc:
                            if causa_vacia:
                                return ("Preventiva sin observación de causa — el técnico cerró el form "
                                        "sin documentar qué encontró.")
                            return "Causa de preventiva no clasificable."
                        # Caso MC
                        if causa_vacia and tiene_com:
                            return ("Describió la falla en texto libre pero NO seleccionó la "
                                    "clasificación Fracttal — descuido de llenado.")
                        if causa_vacia and not tiene_com:
                            return ("No documentó NADA: ni clasificación Fracttal ni desglose en "
                                    "texto libre.")
                        # Tiene texto en la causa pero no clasifica → código no es Fracttal válido
                        if not _DIAG_PREFIJO.match(causa):
                            return (f"Valor en causa ('{causa[:40]}') no es código Fracttal válido "
                                    "(01.x–04.x).")
                        # Código tipo "OTROS" sin keyword reconocible
                        if "OTROS" in causa.upper():
                            extra = " El comentario aclara el detalle." if tiene_com else " Sin comentario que aclare."
                            return f"Código '{causa[:30]}' demasiado genérico.{extra}"
                        return "Clasificación no reconocida como técnico/cliente."

                    _det_cr["_diagnostico"] = _det_cr.apply(_diagnostico_causa, axis=1)
                    # Simplificar tipo: cualquier variante de preventiva (mensual, trimestral,
                    # por horas, etc.) → "PREVENTIVA". En esta tabla no aporta el subtipo.
                    if "maint_type" in _det_cr.columns:
                        _det_cr["maint_type"] = _det_cr["maint_type"].astype(str).str.upper().apply(
                            lambda t: "PREVENTIVA" if "PREVENTIVA" in t or "INSPECC" in t
                                      else ("CORRECTIVA" if "CORRECTIVA" in t else t.title()))
                    # Comentario del técnico (texto libre del PDF) — explica el "OTROS" / "SIN CLASIFICAR"
                    if "comentario_tecnico" in _det_cr.columns:
                        _det_cr["comentario_tecnico"] = (
                            _det_cr["comentario_tecnico"].fillna("").apply(_strip_comentario_headers))
                    # Causa raíz: SÍ mostrar Equipo. En Fracttal las correctivas
                    # siempre tienen 1 activo por OT (la falla es por equipo),
                    # así que cada fila corresponde inequívocamente a un equipo.
                    # Las preventivas multi-subtarea no tienen causa raíz registrada.
                    if "equipment_code" in _det_cr.columns:
                        _det_cr["equipment_code"] = _det_cr["equipment_code"].fillna("").replace("", "—")
                    if "eds_occim" in _det_cr.columns:
                        _det_cr["eds_occim"] = _det_cr["eds_occim"].fillna("").replace("", "—")
                    _det_cr = _det_cr.drop(
                        columns=["_causa_ok","es_correctiva"], errors="ignore"
                    ).rename(columns={
                        "folio":"OT","equipment_code":"Equipo","eds_occim":"EDS",
                        "tecnico":"Técnico","creation_date":"Fecha",
                        "maint_type":"Tipo","causa_raiz_raw":"Causa Raíz",
                        "causa_clasif":"Clasificación",
                        "comentario_tecnico":"Comentario técnico / qué hizo",
                        "_diagnostico":"Diagnóstico del error"
                    }).sort_values("Fecha", ascending=False)
                    # Orden final: OT - Equipo - EDS - resto, Comentario y Diagnóstico al final
                    _orden_cr = ["OT","Equipo","EDS","Técnico","Fecha","Tipo","Causa Raíz",
                                 "Clasificación","Estado",
                                 "Comentario técnico / qué hizo","Diagnóstico del error"]
                    _det_cr = _det_cr[[c for c in _orden_cr if c in _det_cr.columns]]
                    # Filtro por OT (si el usuario escribió algo)
                    _det_cr = _aplicar_filtro_ot(_det_cr, _filtro_cr, col="OT")
                    if _filtro_cr:
                        st.caption(f"Mostrando **{len(_det_cr):,}** de {_n_cr_total:,} OTs (filtro: `{_filtro_cr}`).")
                    _show_df(_det_cr, hide_index=True, width="stretch",
                        column_config={
                            "OT":            st.column_config.TextColumn(width=110),
                            "Equipo":        st.column_config.TextColumn(width=95,
                                help="Código del activo en Fracttal (ej. EQ-6249). En correctivas siempre hay 1 equipo por OT."),
                            "EDS":           st.column_config.TextColumn(width=85,
                                help="Código EDS Occimiano donde se realizó el MP/MC."),
                            "Técnico":       st.column_config.TextColumn(width=190),
                            "Fecha":         st.column_config.TextColumn(width=100),
                            "Tipo":          st.column_config.TextColumn(width=110),
                            "Causa Raíz":    st.column_config.TextColumn(width=220),
                            "Clasificación": st.column_config.TextColumn(width=110),
                            "Comentario técnico / qué hizo": st.column_config.TextColumn(width=340,
                                help="Texto libre del técnico en Fracttal (falla encontrada, trabajo realizado, observaciones). Explica el 'OTROS'/'SIN CLASIFICAR' o lo deja en evidencia si está vacío."),
                            "Diagnóstico del error": st.column_config.TextColumn(width=320,
                                help="Razón por la que el KPI imputa el error: qué llenó mal o qué omitió. Vacío si la fila está correcta."),
                            "Estado":        st.column_config.TextColumn(width=110),
                        })
            else:
                st.info("Sin datos de OTs correctivas para el filtro actual.")

            # ══════════════════════════════════════════════════════════════════
            # SECCIÓN: TIEMPO DE EJECUCIÓN
            # ══════════════════════════════════════════════════════════════════
            st.divider()
            st.markdown('<div class="section-header">⏱  Tiempo de Ejecución — Preventivos</div>',
                        unsafe_allow_html=True)

            # ── Leyenda expandible ────────────────────────────────────────────
            with st.expander("📖  Regla de cumplimiento de tiempo (banda 75% – 150% de la duración estimada)", expanded=False):
                st.markdown("""
**¿Cómo funciona?**

Cada mantenimiento preventivo tiene una **duración estimada** (programada en Fracttal).
El técnico debe ejecutar la tarea dentro de una **banda de tolerancia**:

- **Piso: 75% del tiempo estimado** — si se ejecuta más rápido, es sospecha de quick-tick.
- **Techo: 150% del tiempo estimado** — si se ejecuta más lento, es sobretiempo injustificado.

Ejemplo con MP de **1:30 h (90 min)** estimados:

| Duración estimada | Mínimo (75%) | Máximo (150%) | Ejemplo |
|---|---|---|---|
| 01:30 (90 min) | 01:07 (67 min) | 02:15 (135 min) | Ejecutó 1:20 (80 min) → ✅ Cumple |
| 01:30 (90 min) | 01:07 (67 min) | 02:15 (135 min) | Ejecutó 0:30 (30 min) → ❌ No cumple (déficit) |
| 01:30 (90 min) | 01:07 (67 min) | 02:15 (135 min) | Ejecutó 2:30 (150 min) → 🟣 Exceso |
| 00:40 (40 min) | 00:30 (30 min) | 01:00 (60 min) | Ejecutó 32 min → ✅ Cumple |
| 01:00 (60 min) | 00:45 (45 min) | 01:30 (90 min) | Ejecutó 50 min → ✅ Cumple |

**¿Qué se mide?**
- **Tiempo efectivo** = `max(tasks_duration, tiempo real por fechas OT)` vs **`Duración Estimada`**
- Si `75% × Estimada ≤ Tiempo Efectivo ≤ 150% × Estimada` → **✅ CUMPLE** (verde)
- Si `Tiempo Efectivo < 20% × Estimada` → **🔴 Injustificado** (quick-tick claro)
- Si `20% ≤ Tiempo Efectivo < 75% × Estimada` → **🟡 Déficit** (por debajo del mínimo)
- Si `Tiempo Efectivo > 150% × Estimada` → **🟣 Exceso** (sobretiempo injustificado)
- Si no hay duración estimada → **Sin datos** (no penaliza)

**¿Por qué la banda 75%–150%?**
El piso protege contra quick-ticks (tareas marcadas hechas sin ejecutar).
El techo protege contra pagos de sobretiempo no justificado o registros erróneos
(un MP que "duró" 3× lo estimado suele ser un olvido de cierre en Fracttal).

**¿Por qué usar max(tasks_duration, elapsed)?**
Si el técnico no llenó el campo de duración de tareas pero tuvo el OT abierto 90 min,
esos 90 min cuentan como tiempo real. Evita penalizar por campos sin llenar.
""")

            # ── Gráfico Tiempo de Ejecución (diseño apilado igual que Numerales) ─
            _periodo_lbl_te = ("por semanas" if len(_meses_sel_raw) == 1
                               else "por mes seleccionado")
            st.markdown(f"**Evolución {_periodo_lbl_te} — % preventivos con tiempo de ejecución correcto**")

            _df_te_base = df_kpi_raw[df_kpi_raw["maint_type"] != "CORRECTIVA"].copy() \
                          if "maint_type" in df_kpi_raw.columns else df_kpi_raw.copy()

            # Aplicar filtros de equipo/técnico
            if equipo_kpi != "Todos":
                _grp_te = _LABEL_TO_GRUPO.get(equipo_kpi, equipo_kpi)
                _df_te_base = _df_te_base[_df_te_base["equipo"] == _grp_te]
            if tec_kpi_sel != "Todos":
                _df_te_base = _df_te_base[_df_te_base["tecnico"] == tec_kpi_sel]

            if not _df_te_base.empty and "estimated_sec" in _df_te_base.columns:
                # Solo OTs con duración estimada > 0
                _df_te = _df_te_base[_df_te_base["estimated_sec"] > 0].copy()
                # Tiempo efectivo = max(tasks_duration, elapsed real por fechas).
                # Si el técnico no llenó tasks_duration (=0) pero el OT estuvo abierto
                # N minutos, ese tiempo cuenta. Consistente con _score_tiempo en data.py.
                _df_te["_effective_sec"] = _df_te[["duration_sec","elapsed_sec"]].fillna(0).max(axis=1)
                # CUMPLE si el tiempo efectivo está entre 75% y 150% del estimado.
                # <75%  = déficit (posible quick-tick / trabajo incompleto)
                # >150% = exceso (posible sobretiempo injustificado, holgazanería
                #                 o error de registro). Ambos extremos son "no cumple".
                _df_te["_te_ok"] = (
                    (_df_te["_effective_sec"] >= _df_te["estimated_sec"] * 0.75) &
                    (_df_te["_effective_sec"] <= _df_te["estimated_sec"] * 1.50)
                )
                # Una OT con múltiples equipos genera filas duplicadas con los
                # mismos tiempos neta; mantener solo una fila por OT.
                _df_te = _df_te.drop_duplicates(subset="folio", keep="first")

                # KPI del periodo seleccionado
                _df_te_periodo = _df_te[_df_te["mes"].astype(str).isin(set(_meses_prec_str))].copy()
                if _sem_prec != "Todas":
                    _sem_match_te2 = next((s for s in _sems_prec if s[0] == _sem_prec), None)
                    if _sem_match_te2 and "creation_date_local" in _df_te_periodo.columns:
                        _df_te_periodo = _df_te_periodo[
                            (_df_te_periodo["creation_date_local"] >= _sem_match_te2[1]) &
                            (_df_te_periodo["creation_date_local"] <= _sem_match_te2[2])
                        ]
                _te_tot = len(_df_te_periodo)
                _te_ok  = int(_df_te_periodo["_te_ok"].sum()) if _te_tot else 0
                _te_pct = (_te_ok / _te_tot * 100) if _te_tot else 0.0

                _g_te = _agrupar_por_periodo(_df_te, "_te_ok")

                tc1, tc2 = st.columns([1, 3])
                with tc1:
                    _kpi_card(_te_pct, _te_ok, _te_tot,
                              "preventivos con tiempo OK (75%–150%)", meta_pct=75.0)
                with tc2:
                    if not _g_te.empty:
                        _te_sig = (f"_fig_te_v2_{_current_theme}_{_wo_sig}_{equipo_kpi}"
                                   f"_{tec_kpi_sel}_{'-'.join(_meses_prec_str)}_{_sem_prec}")
                        if _te_sig not in st.session_state:
                            st.session_state[_te_sig] = _fig_apilada(
                                _g_te, color_ok="#3b82f6",
                                label_ok="Tiempo correcto (≥75%)",
                                label_err="Tiempo insuficiente",
                                meta=75.0, titulo_y="% preventivos",
                            )
                        st.plotly_chart(st.session_state[_te_sig], width="stretch")

                # ── Filtro de periodo para tablas y gráficos de dona/barras ──
                # (el gráfico de evolución mensual siempre muestra todos los meses)
                _df_te_p = _df_te[_df_te["mes"].astype(str).isin(set(_meses_prec_str))].copy()
                if _sem_prec != "Todas":
                    _sem_match_te = next((s for s in _sems_prec if s[0] == _sem_prec), None)
                    if _sem_match_te and "creation_date_local" in _df_te_p.columns:
                        _df_te_p = _df_te_p[
                            (_df_te_p["creation_date_local"] >= _sem_match_te[1]) &
                            (_df_te_p["creation_date_local"] <= _sem_match_te[2])
                        ]

                # ── Tabla detalle de Tiempo de Ejecución ─────────────────────
                def _fmt_seg(s):
                    if pd.isna(s) or s == 0: return "—"
                    s = int(s); h, m = s//3600, (s%3600)//60
                    return f"{h:02d}:{m:02d}"

                _det_te = _df_te_p.copy()
                _det_te["_minimo_sec"] = (_det_te["estimated_sec"] * 0.75).round(0)
                _det_te["_maximo_sec"] = (_det_te["estimated_sec"] * 1.50).round(0)
                _det_te["_pct_ej"]     = (_det_te["_effective_sec"] / _det_te["estimated_sec"] * 100).round(1)
                _det_te["_es_exceso"]  = _det_te["_effective_sec"] > _det_te["_maximo_sec"]
                _det_te_cd = pd.to_datetime(_det_te["creation_date"], errors="coerce")
                _det_te_cd = _det_te_cd.dt.tz_convert(None) if _det_te_cd.dt.tz is not None else _det_te_cd
                _det_te["creation_date"] = _det_te_cd.dt.strftime("%d/%m/%Y")

                _det_te_disp = _det_te[[c for c in
                    ["folio","eds_occim","tecnico","creation_date","maint_type",
                     "estimated_sec","_minimo_sec","_maximo_sec","_effective_sec",
                     "_pct_ej","_te_ok","_es_exceso"]
                    if c in _det_te.columns]].copy()
                _det_te_disp["T. Estimado"]   = _det_te_disp["estimated_sec"].apply(_fmt_seg)
                _det_te_disp["Mín. 75%"]       = _det_te_disp["_minimo_sec"].apply(_fmt_seg)
                _det_te_disp["Máx. 150%"]      = _det_te_disp["_maximo_sec"].apply(_fmt_seg)
                _det_te_disp["T. Ejecución"]   = _det_te_disp["_effective_sec"].apply(_fmt_seg)
                _det_te_disp["% Ejecutado"]    = _det_te_disp["_pct_ej"]
                def _estado_lbl(r):
                    if bool(r.get("_te_ok", False)):
                        return "✅ Cumple"
                    if bool(r.get("_es_exceso", False)):
                        return "⚠️ Exceso (>150%)"
                    return "❌ No cumple"
                _det_te_disp["Estado"] = _det_te_disp.apply(_estado_lbl, axis=1)

                # Diagnóstico de no cumplimiento del tiempo. Solo se llena cuando
                # _te_ok == False; cuando cumple queda vacío.
                def _fmt_hm(seg):
                    if seg is None or pd.isna(seg) or seg <= 0: return "0min"
                    seg = int(seg); h, m = seg//3600, (seg%3600)//60
                    return f"{h}h{m:02d}min" if h else f"{m}min"

                def _diag_tiempo(r):
                    if bool(r.get("_te_ok", False)):
                        return ""
                    estim = int(r.get("estimated_sec") or 0)
                    real  = int(r.get("_effective_sec") or 0)
                    minimo = int(r.get("_minimo_sec") or 0)
                    maximo = int(r.get("_maximo_sec") or 0)
                    pct = float(r.get("_pct_ej") or 0.0)
                    estim_h = _fmt_hm(estim)
                    real_h  = _fmt_hm(real)
                    min_h   = _fmt_hm(minimo)
                    max_h   = _fmt_hm(maximo)
                    # ── EXCESO (>150% del estimado) ─────────────────────────
                    if bool(r.get("_es_exceso", False)):
                        exceso = real - maximo
                        if pct > 250:
                            return (f"Exceso extremo: registró {real_h} cuando el estimado era {estim_h} "
                                    f"(tope máximo {max_h}). Más del doble del tope permitido; "
                                    f"revisar si fue error de registro o sobretiempo real.")
                        return (f"Sobretiempo: {real_h} ejecutado supera el tope de {max_h} "
                                f"(150% de {estim_h}). Excedió por {_fmt_hm(exceso)}.")
                    # ── DÉFICIT (<75% del estimado) ─────────────────────────
                    deficit = minimo - real
                    if real == 0:
                        return (f"Sin registro de tiempo: el técnico no documentó duración alguna. "
                                f"Estimado {estim_h}, mínimo aceptable {min_h}.")
                    if pct < 5:
                        return (f"Tiempo absurdo: registró {real_h} cuando el estimado era {estim_h} "
                                f"(mínimo {min_h}). Imposible ejecutar una preventiva en ese plazo.")
                    if pct < 20:
                        return (f"Tiempo injustificado: {real_h} es <20% del estimado ({estim_h}). "
                                f"Quick-tick probable — debió ser al menos {min_h}.")
                    if pct < 50:
                        return (f"Muy por debajo del mínimo: {real_h} ejecutado vs {min_h} requerido "
                                f"(75% de {estim_h}). Déficit de {_fmt_hm(deficit)}.")
                    # 50-75% → cerca pero no llega
                    return (f"El acumulado del tiempo no cumple ni siquiera el 75% mínimo de {estim_h}: "
                            f"registró {real_h} y se requerían {min_h} (faltaron {_fmt_hm(deficit)}).")

                _det_te_disp["Diagnóstico de no cumplimiento"] = _det_te.apply(_diag_tiempo, axis=1).values
                # Normalizar EDS para display (Equipo se omite: una OT puede tener
                # subtareas con activos distintos; lo que evalúa el KPI es la OT completa).
                if "eds_occim" in _det_te_disp.columns:
                    _det_te_disp["eds_occim"] = _det_te_disp["eds_occim"].fillna("").replace("", "—")
                _det_te_disp = _det_te_disp.drop(
                    columns=["estimated_sec","_minimo_sec","_maximo_sec","_effective_sec",
                             "_pct_ej","_te_ok","_es_exceso"],
                    errors="ignore"
                ).rename(columns={
                    "folio":"OT","eds_occim":"EDS",
                    "tecnico":"Técnico","creation_date":"Fecha","maint_type":"Tipo"
                }).sort_values("Fecha", ascending=False)
                # Orden: OT - EDS - Técnico - resto, Diagnóstico al final
                _orden_te = ["OT","EDS","Técnico","Fecha","Tipo",
                             "T. Estimado","Mín. 75%","Máx. 150%","T. Ejecución","% Ejecutado","Estado",
                             "Diagnóstico de no cumplimiento"]
                _det_te_disp = _det_te_disp[[c for c in _orden_te if c in _det_te_disp.columns]]

                _n_te_total = len(_det_te_disp)
                with st.expander(f"📋 Detalle de OTs — Tiempo de Ejecución ({_n_te_total:,} preventivos con estimado)", expanded=False):
                    _filtro_te = _filtro_ot_input("kpi_filtro_ot_tiempo")
                    _det_te_disp = _aplicar_filtro_ot(_det_te_disp, _filtro_te, col="OT")
                    if _filtro_te:
                        st.caption(f"Mostrando **{len(_det_te_disp):,}** de {_n_te_total:,} OTs (filtro: `{_filtro_te}`).")
                    _show_df(_det_te_disp, hide_index=True, width="stretch",
                        column_config={
                            "OT":          st.column_config.TextColumn(width=110),
                            "EDS":         st.column_config.TextColumn(width=85,
                                help="Código EDS Occimiano donde se realizó el MP/MC."),
                            "Técnico":     st.column_config.TextColumn(width=190),
                            "Fecha":       st.column_config.TextColumn(width=100),
                            "Tipo":        st.column_config.TextColumn(width=200),
                            "T. Estimado": st.column_config.TextColumn(width=100,
                                help="Duración programada en Fracttal (HH:MM)"),
                            "Mín. 75%":    st.column_config.TextColumn(width=90,
                                help="Tiempo mínimo aceptable = 75% del estimado"),
                            "Máx. 150%":   st.column_config.TextColumn(width=90,
                                help="Tope máximo aceptable = 150% del estimado. Superarlo cuenta como Exceso."),
                            "T. Ejecución":st.column_config.TextColumn(width=110,
                                help="Tiempo efectivo = max(tiempo tareas, tiempo real por fechas)"),
                            "% Ejecutado": st.column_config.ProgressColumn(
                                label="% Ejecutado", min_value=0, max_value=250, format="%.1f%%",
                                help="T.Efectivo / T.Estimado × 100. Cumple si está entre 75% y 150%."),
                            "Estado":      st.column_config.TextColumn(width=130),
                            "Diagnóstico de no cumplimiento": st.column_config.TextColumn(width=380,
                                help="Razón concreta por la que el tiempo no cumple: exceso, absurdo, sin registro, déficit vs mínimo. Vacío cuando cumple."),
                        })

                # ── Subsección: OTs con tiempo fuera de rango (déficit o exceso) ────
                st.markdown("---")
                st.markdown("**⚠️ OTs con tiempo de ejecución fuera de rango (< 20% ó > 150% del estimado)**")
                st.caption(
                    "Barras apiladas por equipo (o por técnico si filtras). "
                    "**Verde** = cumplen (75%–150%) · "
                    "**Amarillo** = déficit no razonable (20–75%) · "
                    "**Rojo** = déficit injustificado (< 20%) · "
                    "**Morado** = exceso injustificado (> 150%)"
                )

                # Marcar los 4 segmentos sobre el total de preventivos con estimado
                _df_te_p["_exceso"]    = _df_te_p["_effective_sec"] > _df_te_p["estimated_sec"] * 1.50
                _df_te_p["_absurdo"]   = (
                    (~_df_te_p["_te_ok"]) & (~_df_te_p["_exceso"]) &
                    (_df_te_p["_effective_sec"] < _df_te_p["estimated_sec"] * 0.20)
                )
                _df_te_p["_just_fail"] = (~_df_te_p["_te_ok"]) & (~_df_te_p["_absurdo"]) & (~_df_te_p["_exceso"])

                # Determinar agrupación según filtros
                if tec_kpi_sel != "Todos":
                    _grp_col = "tecnico"
                elif equipo_kpi != "Todos":
                    _grp_col = "tecnico"
                else:
                    _grp_col = "equipo"

                _te_grp = (
                    _df_te_p.groupby(_grp_col)
                    .agg(
                        ok=("_te_ok",     "sum"),
                        just_fail=("_just_fail", "sum"),
                        absurdo=("_absurdo",  "sum"),
                        exceso=("_exceso",   "sum"),
                        total=("_te_ok",  "count"),
                    )
                    .reset_index()
                )
                _te_grp["pct_ok"]      = (_te_grp["ok"]        / _te_grp["total"] * 100).round(1)
                _te_grp["pct_just"]    = (_te_grp["just_fail"]  / _te_grp["total"] * 100).round(1)
                _te_grp["pct_absurdo"] = (_te_grp["absurdo"]    / _te_grp["total"] * 100).round(1)
                _te_grp["pct_exceso"]  = (_te_grp["exceso"]     / _te_grp["total"] * 100).round(1)

                if _grp_col == "equipo":
                    _grp_order = {k: i for i, k in enumerate(GRUPOS_TERRENO)}
                    _te_grp["_order"] = _te_grp["equipo"].map(_grp_order).fillna(99)
                    _te_grp["_label"] = _te_grp["equipo"].map(_EQUIPO_LABEL).fillna(_te_grp["equipo"])
                    _te_grp = _te_grp.sort_values("_order").drop(columns=["_order"]).reset_index(drop=True)
                else:
                    _te_grp["_label"] = _te_grp["tecnico"]
                    _te_grp["_problemas"] = _te_grp["absurdo"] + _te_grp["exceso"]
                    _te_grp = _te_grp.sort_values("_problemas", ascending=False).drop(columns=["_problemas"]).reset_index(drop=True)

                _abs_sig = f"_fig_abs_{_current_theme}_{_wo_sig}_{equipo_kpi}_{tec_kpi_sel}_{'|'.join(_meses_prec_str)}_{_sem_prec}"
                if _abs_sig not in st.session_state:
                    _use_donut = (equipo_kpi == "Todos" and tec_kpi_sel == "Todos")

                    if _use_donut:
                        # ── Dos donas: vista general todos los equipos ───────────
                        _n_ok_d  = int(_te_grp["ok"].sum())
                        _n_jd    = int(_te_grp["just_fail"].sum())
                        _n_abd   = int(_te_grp["absurdo"].sum())
                        _n_ex_d  = int(_te_grp["exceso"].sum())
                        _n_no_cumple = _n_jd + _n_abd + _n_ex_d
                        _n_total_d   = _n_ok_d + _n_no_cumple

                        _fig_abs = make_subplots(
                            rows=1, cols=2,
                            specs=[[{"type": "domain"}, {"type": "domain"}]],
                            subplot_titles=[
                                "Cumplimiento general",
                                "Desglose no cumplen",
                            ],
                        )
                        # Dona 1: cumple vs no cumple (incluye exceso en amarillo)
                        _fig_abs.add_trace(go.Pie(
                            labels=["Cumplen (75%–150%)", "No cumplen"],
                            values=[_n_ok_d, _n_no_cumple],
                            hole=0.52,
                            marker=dict(colors=["#22c55e", "#f59e0b"],
                                        line=dict(color="#ffffff", width=2)),
                            textinfo="percent+value",
                            texttemplate="%{percent:.1%}<br>%{value:,} OTs",
                            hovertemplate="%{label}<br>%{value:,} OTs — %{percent:.1%}<extra></extra>",
                            direction="clockwise",
                            sort=False,
                        ), row=1, col=1)
                        # Dona 2: de los que no cumplen → 3 categorías (déficit 20-75%,
                        # injustificado <20%, exceso >150%)
                        _fig_abs.add_trace(go.Pie(
                            labels=["Déficit (20–75%)", "Injustificado (<20%)", "Exceso (>150%)"],
                            values=[_n_jd, _n_abd, _n_ex_d],
                            hole=0.52,
                            marker=dict(colors=["#f59e0b", "#ef4444", "#8b5cf6"],
                                        line=dict(color="#ffffff", width=2)),
                            textinfo="percent+value",
                            texttemplate="%{percent:.1%}<br>%{value:,} OTs",
                            hovertemplate="%{label}<br>%{value:,} OTs — %{percent:.1%}<extra></extra>",
                            direction="clockwise",
                            sort=False,
                        ), row=1, col=2)
                        _fig_abs.update_layout(
                            height=360,
                            margin=dict(t=40, b=10, l=10, r=10),
                            showlegend=True,
                            legend=dict(orientation="h", y=-0.05, x=0.5,
                                        xanchor="center", font=dict(size=11)),
                            annotations=[
                                dict(text=f"<b>{_n_total_d:,}</b><br>OTs",
                                     x=0.19, y=0.5, showarrow=False,
                                     font=dict(size=16)),
                                dict(text=f"<b>{_n_no_cumple:,}</b><br>No cumplen",
                                     x=0.81, y=0.5, showarrow=False,
                                     font=dict(size=14)),
                            ],
                        )
                    else:
                        # ── Barras apiladas: vista por técnico dentro del equipo ──
                        _x_lbl = _te_grp["_label"].tolist()
                        _fig_abs = go.Figure()
                        _fig_abs.add_trace(go.Bar(
                            name="Cumplen (75%–150%)",
                            x=_x_lbl,
                            y=_te_grp["ok"].tolist(),
                            marker_color="#22c55e",
                            text=[f"{int(v):,}<br>{p:.1f}%"
                                  for v, p in zip(_te_grp["ok"], _te_grp["pct_ok"])],
                            textposition="inside",
                            textfont=dict(size=11, color="#ffffff"),
                        ))
                        _fig_abs.add_trace(go.Bar(
                            name="Déficit (20–75%)",
                            x=_x_lbl,
                            y=_te_grp["just_fail"].tolist(),
                            marker_color="#f59e0b",
                            text=[f"{int(v):,}<br>{p:.1f}%" if v > 0 else ""
                                  for v, p in zip(_te_grp["just_fail"], _te_grp["pct_just"])],
                            textposition="inside",
                            textfont=dict(size=11, color="#ffffff"),
                        ))
                        _fig_abs.add_trace(go.Bar(
                            name="Injustificado (<20%)",
                            x=_x_lbl,
                            y=_te_grp["absurdo"].tolist(),
                            marker_color="#ef4444",
                            text=[f"{int(v):,}<br>{p:.1f}%" if v > 0 else ""
                                  for v, p in zip(_te_grp["absurdo"], _te_grp["pct_absurdo"])],
                            textposition="inside",
                            textfont=dict(size=11, color="#ffffff"),
                        ))
                        _fig_abs.add_trace(go.Bar(
                            name="Exceso (>150%)",
                            x=_x_lbl,
                            y=_te_grp["exceso"].tolist(),
                            marker_color="#8b5cf6",
                            text=[f"{int(v):,}<br>{p:.1f}%" if v > 0 else ""
                                  for v, p in zip(_te_grp["exceso"], _te_grp["pct_exceso"])],
                            textposition="inside",
                            textfont=dict(size=11, color="#ffffff"),
                        ))
                        _fig_abs.update_layout(
                            barmode="stack",
                            height=340,
                            margin=dict(t=30, b=20, l=10, r=20),
                            showlegend=True,
                            legend=dict(orientation="h", y=1.1, x=0),
                            yaxis=dict(title="OTs preventivas", tickformat="d"),
                            xaxis=dict(title=""),
                            bargap=0.35,
                        )

                    _apply_plot_theme(_fig_abs)
                    st.session_state[_abs_sig] = _fig_abs
                st.plotly_chart(st.session_state[_abs_sig], width="stretch")

                # ── Tabla detalle de OTs injustificadas ──────────────────────────
                _df_abs_only = _df_te_p[_df_te_p["_absurdo"]].copy()
                _n_absurdo   = len(_df_abs_only)

                if not _df_abs_only.empty:
                    _df_abs_only["_pct_ej_abs"] = (
                        _df_abs_only["_effective_sec"] / _df_abs_only["estimated_sec"] * 100
                    ).round(1)
                    _det_abs = _df_abs_only.copy()
                    _det_abs_cd = pd.to_datetime(_det_abs["creation_date"], errors="coerce")
                    _det_abs_cd = _det_abs_cd.dt.tz_convert(None) if _det_abs_cd.dt.tz is not None else _det_abs_cd
                    _det_abs["creation_date"] = _det_abs_cd.dt.strftime("%d/%m/%Y")
                    _det_abs["T. Estimado"]  = _det_abs["estimated_sec"].apply(_fmt_seg)
                    _det_abs["T. Ejecución"] = _det_abs["_effective_sec"].apply(_fmt_seg)
                    _det_abs["% Ejecutado"]  = _det_abs["_pct_ej_abs"]
                    if "eds_occim" in _det_abs.columns:
                        _det_abs["eds_occim"] = _det_abs["eds_occim"].fillna("").replace("", "—")
                    _det_abs_disp = _det_abs[[c for c in
                        ["folio","eds_occim","tecnico","creation_date","maint_type",
                         "T. Estimado","T. Ejecución","% Ejecutado"]
                        if c in _det_abs.columns]].rename(columns={
                            "folio":"OT","eds_occim":"EDS","tecnico":"Técnico",
                            "creation_date":"Fecha","maint_type":"Tipo",
                        }).sort_values("Fecha", ascending=False)

                    with st.expander(
                        f"📋 Detalle OTs con tiempo injustificado ({_n_absurdo:,} OTs)", expanded=False
                    ):
                        _filtro_abs = _filtro_ot_input("kpi_filtro_ot_injust")
                        _det_abs_disp = _aplicar_filtro_ot(_det_abs_disp, _filtro_abs, col="OT")
                        if _filtro_abs:
                            st.caption(f"Mostrando **{len(_det_abs_disp):,}** de {_n_absurdo:,} OTs (filtro: `{_filtro_abs}`).")
                        _show_df(_det_abs_disp, hide_index=True, width="stretch",
                            column_config={
                                "OT":          st.column_config.TextColumn(width=110),
                                "EDS":         st.column_config.TextColumn(width=85,
                                    help="Código EDS Occimiano donde se realizó el MP."),
                                "Técnico":     st.column_config.TextColumn(width=190),
                                "Fecha":       st.column_config.TextColumn(width=100),
                                "Tipo":        st.column_config.TextColumn(width=200),
                                "T. Estimado": st.column_config.TextColumn(width=100),
                                "T. Ejecución":st.column_config.TextColumn(width=110),
                                "% Ejecutado": st.column_config.ProgressColumn(
                                    min_value=0, max_value=50, format="%.1f%%",
                                    help="Todos en esta tabla están por debajo del 20% del estimado."),
                            })
                else:
                    st.success("✅ No hay preventivos con tiempo inferior al 20% del estimado en este período.")

                # ── Tabla detalle de OTs con EXCESO (> 150% del estimado) ────
                _df_ex_only = _df_te_p[_df_te_p["_exceso"]].copy()
                _n_exceso   = len(_df_ex_only)
                if not _df_ex_only.empty:
                    _df_ex_only["_pct_ej_ex"] = (
                        _df_ex_only["_effective_sec"] / _df_ex_only["estimated_sec"] * 100
                    ).round(1)
                    _det_ex = _df_ex_only.copy()
                    _det_ex_cd = pd.to_datetime(_det_ex["creation_date"], errors="coerce")
                    _det_ex_cd = _det_ex_cd.dt.tz_convert(None) if _det_ex_cd.dt.tz is not None else _det_ex_cd
                    _det_ex["creation_date"] = _det_ex_cd.dt.strftime("%d/%m/%Y")
                    _det_ex["T. Estimado"]  = _det_ex["estimated_sec"].apply(_fmt_seg)
                    _det_ex["T. Máximo"]    = (_det_ex["estimated_sec"] * 1.50).apply(_fmt_seg)
                    _det_ex["T. Ejecución"] = _det_ex["_effective_sec"].apply(_fmt_seg)
                    _det_ex["% Ejecutado"]  = _det_ex["_pct_ej_ex"]
                    if "eds_occim" in _det_ex.columns:
                        _det_ex["eds_occim"] = _det_ex["eds_occim"].fillna("").replace("", "—")
                    _det_ex_disp = _det_ex[[c for c in
                        ["folio","eds_occim","tecnico","creation_date","maint_type",
                         "T. Estimado","T. Máximo","T. Ejecución","% Ejecutado"]
                        if c in _det_ex.columns]].rename(columns={
                            "folio":"OT","eds_occim":"EDS","tecnico":"Técnico",
                            "creation_date":"Fecha","maint_type":"Tipo",
                        }).sort_values("% Ejecutado", ascending=False)

                    with st.expander(
                        f"🟣 Detalle OTs con exceso de tiempo — >150% ({_n_exceso:,} OTs)", expanded=False
                    ):
                        _filtro_ex = _filtro_ot_input("kpi_filtro_ot_exceso")
                        _det_ex_disp = _aplicar_filtro_ot(_det_ex_disp, _filtro_ex, col="OT")
                        if _filtro_ex:
                            st.caption(f"Mostrando **{len(_det_ex_disp):,}** de {_n_exceso:,} OTs (filtro: `{_filtro_ex}`).")
                        _pct_max = max(250, int(_det_ex_disp["% Ejecutado"].max()) + 20) if not _det_ex_disp.empty else 300
                        _show_df(_det_ex_disp, hide_index=True, width="stretch",
                            column_config={
                                "OT":          st.column_config.TextColumn(width=110),
                                "EDS":         st.column_config.TextColumn(width=85),
                                "Técnico":     st.column_config.TextColumn(width=190),
                                "Fecha":       st.column_config.TextColumn(width=100),
                                "Tipo":        st.column_config.TextColumn(width=200),
                                "T. Estimado": st.column_config.TextColumn(width=100),
                                "T. Máximo":   st.column_config.TextColumn(width=100,
                                    help="Tope aceptable = 150% del estimado."),
                                "T. Ejecución":st.column_config.TextColumn(width=110),
                                "% Ejecutado": st.column_config.ProgressColumn(
                                    min_value=150, max_value=_pct_max, format="%.1f%%",
                                    help="Todos en esta tabla superan el 150% del estimado."),
                            })

            else:
                st.info("Sin datos de duración estimada disponibles para el filtro actual.")

            # ══════════════════════════════════════════════════════════════════
            # SECCIÓN: REGISTRO DE NUMERALES
            # ══════════════════════════════════════════════════════════════════
            st.divider()
            st.markdown('<div class="section-header">🔢  Registro de Numerales — OTs con número de ficha</div>',
                        unsafe_allow_html=True)
            st.caption(
                "Aplica a **lavadoras y aspiradoras** en MC y MP. El formulario Fracttal "
                "exige el numeral en ambos tipos, así que un valor faltante o basura "
                "penaliza igual sea correctiva o preventiva."
            )

            # Numeral aplica a TODA lavadora/aspiradora (MC + MP) — se incluyen
            # correctivas y preventivas. Los equipos no-lavadora aportan score=25
            # auto (no_aplica), igual no distorsionan el indicador.
            _df_num_base = df_ot_scores.copy()
            if not _df_num_base.empty:
                _num_ok  = int(_df_num_base["numeral_ok"].sum())
                _num_tot = len(_df_num_base)
                _num_pct = _num_ok / _num_tot * 100 if _num_tot > 0 else 0.0

                # ── Evolución (respeta selección: semanas si 1 mes, meses si varios) ─
                # Incluye MC + MP (el numeral aplica a ambos en lavadora/aspiradora).
                _df_num_hist = df_ot_all.copy()
                if equipo_kpi != "Todos":
                    _grp_num = _LABEL_TO_GRUPO.get(equipo_kpi, equipo_kpi)
                    _df_num_hist = _df_num_hist[_df_num_hist["equipo"] == _grp_num]
                if tec_kpi_sel != "Todos":
                    _df_num_hist = _df_num_hist[_df_num_hist["tecnico"] == tec_kpi_sel]

                # _agrupar_por_periodo aplica la regla: 1 mes → semanas; varios → meses.
                _g_num = _agrupar_por_periodo(_df_num_hist, "numeral_ok")

                _periodo_lbl_num = ("por semanas" if len(_meses_sel_raw) == 1
                                    else "por mes seleccionado")

                nc1, nc2 = st.columns([1, 3])
                with nc1:
                    _kpi_card(_num_pct, _num_ok, _num_tot,
                              "con numeral registrado", meta_pct=90.0)
                with nc2:
                    if not _g_num.empty:
                        _num_sig = (f"_fig_num_v2_{_current_theme}_{_wo_sig}_{equipo_kpi}"
                                    f"_{tec_kpi_sel}_{'-'.join(_meses_prec_str)}_{_sem_prec}")
                        if _num_sig not in st.session_state:
                            st.session_state[_num_sig] = _fig_apilada(
                                _g_num, color_ok="#22c55e",
                                label_ok="Registró numeral", label_err="Sin numeral",
                                meta=90.0, titulo_y="% OTs",
                            )
                        st.plotly_chart(st.session_state[_num_sig], width="stretch")

                # ── Tabla detalle OTs numerales (cumple + no cumple) ──────────
                # Si tenemos el desglose por subtarea (numerales_subtarea), una
                # OT con varios activos se EXPANDE: 1 fila por (OT, activo) con
                # su numeral propio y motivo individual. La OT entera se considera
                # mala si CUALQUIER subtarea falla.
                _det_num = _df_num_base.copy()
                if df_num_sub is not None and not df_num_sub.empty:
                    _sub = df_num_sub.copy()
                    # Solo subtareas en activos donde aplica numeral
                    _sub = _sub[_sub["tipo_activo"].isin(("lavadora","aspiradora","lavainterior"))]
                    if not _sub.empty:
                        _sub = _sub.rename(columns={"id_ot": "folio"})
                        # Meta de la OT (técnico, fecha, cliente, EDS, tipo, comentario).
                        _meta_cols = [c for c in [
                            "folio","tecnico","creation_date","client","station","maint_type",
                            "eds_occim","comentario_tecnico","es_correctiva"]
                            if c in _det_num.columns]
                        _meta = _det_num[_meta_cols].drop_duplicates("folio")
                        # INNER merge: solo OTs en el scope actual (período/equipo/técnico).
                        # Antes era left desde _sub (todo 2026) → filas huérfanas con
                        # metadatos vacíos (None en técnico/fecha/cliente).
                        _exp = _meta.merge(_sub, on="folio", how="inner")
                        if not _exp.empty:
                            _exp["numeral_motivo"] = _exp["motivo"]
                            _exp["es_lavadora"]    = True
                            _exp["numeral_valor"]  = (
                                _exp["numeral_final"].fillna(_exp["numeral_inicial"]).fillna(""))
                            # Equipo = código del activo de ESTA subtarea (EQ-XXXX)
                            _exp["_equipo_sub"] = _exp["codigo_activo"].fillna("—")
                            _det_num = _exp

                # Columna Numeral y Estado con tres casos claros:
                #   🔵 No aplica  → equipo no es lavadora (no se exige numeral)
                #   ✅ 182740     → lavadora con numeral real encontrado en la nota
                #   ❌ Sin numeral → lavadora sin numeral registrado
                def _fmt_numeral_valor(row):
                    if not row.get("es_lavadora", True):
                        return "N/A"
                    v = str(row.get("numeral_valor", "") or "").strip()
                    return v if v else "—"

                def _fmt_numeral_estado(row):
                    if not row.get("es_lavadora", True):
                        return "🔵 No aplica"
                    # Motivo del veredicto de calidad (sin numeral / basura / exceso / etc.)
                    _motivo = str(row.get("numeral_motivo", "") or "")
                    if row.get("numeral_ok", False):
                        return "✅ Cumple"
                    # Dato MALO → mostrar la razón concreta (no solo "sin numeral")
                    return NUMERAL_MOTIVO_LABEL.get(_motivo, "❌ Dato inválido")

                def _fmt_numeral_ini_fin(row, campo):
                    if not row.get("es_lavadora", True):
                        return "—"
                    v = str(row.get(campo, "") or "").strip()
                    return v if v and v.lower() not in ("none", "null") else "—"

                def _fmt_fichas(row):
                    if not row.get("es_lavadora", True):
                        return "—"
                    fp = row.get("fichas_periodo", None)
                    if fp is None or (isinstance(fp, float) and pd.isna(fp)):
                        return "—"
                    try:
                        return f"{int(fp):,}"
                    except (ValueError, TypeError):
                        return "—"

                _det_num["_numeral"]  = _det_num.apply(_fmt_numeral_valor, axis=1)
                _det_num["_estado"]   = _det_num.apply(_fmt_numeral_estado, axis=1)
                _det_num["_n_ini"]    = _det_num.apply(lambda r: _fmt_numeral_ini_fin(r, "numeral_inicial"), axis=1)
                _det_num["_n_fin"]    = _det_num.apply(lambda r: _fmt_numeral_ini_fin(r, "numeral_final"), axis=1)
                _det_num["_fichas"]   = _det_num.apply(_fmt_fichas, axis=1)

                # Formatear fecha
                _det_num_cd = pd.to_datetime(_det_num["creation_date"], errors="coerce")
                _det_num_cd = _det_num_cd.dt.tz_convert(None) if _det_num_cd.dt.tz is not None else _det_num_cd
                _det_num["_fecha"] = _det_num_cd.dt.strftime("%d/%m/%Y")

                # Comentario / conclusión del técnico (texto libre del PDF de la OT).
                # Clave para rastrear la causa raíz cuando el numeral viene malo.
                # fillna("") evita que NaN llegue como "nan" al sanitizer.
                _serie_com = (_det_num["comentario_tecnico"].fillna("")
                              if "comentario_tecnico" in _det_num.columns
                              else pd.Series("", index=_det_num.index))
                _det_num["_comentario"] = _serie_com.apply(_strip_comentario_headers)

                # Equipo (código del activo + tipo legible). En vista expandida =
                # activo de la subtarea; en vista normal = no disponible a este nivel.
                _TIPO_LBL = {
                    "lavadora":     "Lavadora",
                    "aspiradora":   "Aspiradora",
                    "lavainterior": "Lavatapiz",
                }
                def _fmt_equipo(row):
                    code = str(row.get("_equipo_sub", "") or "").strip()
                    if not code or code == "—":
                        return "—"
                    tipo = _TIPO_LBL.get(str(row.get("tipo_activo", "") or "").strip(), "")
                    return f"{code} · {tipo}" if tipo else code
                if "_equipo_sub" in _det_num.columns:
                    _det_num["_equipo"] = _det_num.apply(_fmt_equipo, axis=1)
                else:
                    _det_num["_equipo"] = "—"

                # Construir df de display
                _cols_num = [c for c in
                    ["folio","_equipo","tecnico","_fecha","client","station","maint_type",
                     "_n_ini","_n_fin","_fichas","_estado","_comentario"]
                    if c in _det_num.columns]
                _det_num_disp = _det_num[_cols_num].rename(columns={
                    "folio":      "OT",
                    "_equipo":    "Equipo",
                    "tecnico":    "Técnico",
                    "_fecha":     "Fecha",
                    "client":     "Cliente",
                    "station":    "EDS",
                    "maint_type": "Tipo",
                    "_n_ini":     "N. Inicial",
                    "_n_fin":     "N. Final",
                    "_fichas":    "Fichas período",
                    "_estado":    "Estado",
                    "_comentario":"Comentario técnico / causa raíz",
                }).sort_values(["OT", "Estado"], ascending=[True, True])

                _n_sin      = int((~_df_num_base["numeral_ok"]).sum())
                _n_lav_ok   = int((_df_num_base["es_lavadora"] & _df_num_base["numeral_ok"]).sum()) \
                              if "es_lavadora" in _df_num_base.columns else int(_df_num_base["numeral_ok"].sum())
                _n_no_aplica = int((~_df_num_base["es_lavadora"]).sum()) \
                               if "es_lavadora" in _df_num_base.columns else 0
                with st.expander(
                    f"📋 Detalle OTs — numerales ({_n_lav_ok:,} registrados · {_n_sin:,} sin numeral · {_n_no_aplica:,} no aplica)",
                    expanded=False,
                ):
                    st.caption(
                        "Aplica a **lavadoras y aspiradoras**. Numerales **reales** "
                        "leídos del formulario de la tarea en Fracttal (no de la nota). "
                        "🔵 **No aplica** = equipo sin numeral (compresores, ablandadores, etc.). "
                        "✅ = numeral válido · ❌/🟣 = dato malo (sin numeral, basura, exceso de fichas, salto). "
                        "**Fichas período** = N. Final − N. Inicial (>20 en una OT = sospechoso). "
                        "**Comentario técnico** = lo que el técnico escribió en el formulario "
                        "(falla, trabajo realizado, observaciones) — la causa raíz real."
                    )
                    _n_num_total = len(_det_num_disp)
                    _filtro_num = _filtro_ot_input("kpi_filtro_ot_numeral")
                    _det_num_disp = _aplicar_filtro_ot(_det_num_disp, _filtro_num, col="OT")
                    if _filtro_num:
                        st.caption(f"Mostrando **{len(_det_num_disp):,}** de {_n_num_total:,} OTs (filtro: `{_filtro_num}`).")
                    _show_df(_det_num_disp, hide_index=True, width="stretch",
                        column_config={
                            "OT":            st.column_config.TextColumn(width=110),
                            "Equipo":        st.column_config.TextColumn(width=170,
                                help="Código del activo (EQ-XXXX) + tipo (Lavadora/Aspiradora/Lavatapiz). Una OT compuesta muestra una fila por activo con numeral."),
                            "Técnico":       st.column_config.TextColumn(width=180),
                            "Fecha":         st.column_config.TextColumn(width=95),
                            "Cliente":       st.column_config.TextColumn(width=85),
                            "EDS":           st.column_config.TextColumn(width=180),
                            "Tipo":          st.column_config.TextColumn(width=130),
                            "N. Inicial":    st.column_config.TextColumn(width=90,
                                help="Lectura inicial del contador (TOMA DE NUMERAL INICIAL)."),
                            "N. Final":      st.column_config.TextColumn(width=90,
                                help="Lectura final del contador (TOMA DE NUMERAL FINAL)."),
                            "Fichas período":st.column_config.TextColumn(width=100,
                                help="Diferencia Final − Inicial = fichas usadas en el período."),
                            "Estado":        st.column_config.TextColumn(width=150),
                            "Comentario técnico / causa raíz": st.column_config.TextColumn(width=340,
                                help="Texto libre del técnico en el formulario de Fracttal: falla encontrada, trabajo realizado y observaciones. Vacío = el técnico no documentó."),
                        })

            if not df_tec_scores_rank.empty:
                _tec_base = df_tec_scores_rank[df_tec_scores_rank["tecnico"].isin(TECNICOS_OCCIMIANO_FULL)].copy()

                # Guards para columnas nuevas (caches pre-migración)
                _guards = {
                    "err_total_dim":      lambda b: b["n_errores"],
                    "err_tiempo":         lambda b: 0,
                    "err_causa":          lambda b: 0,
                    "err_numeral":        lambda b: 0,
                    "err_deteccion":      lambda b: 0,
                    "ots_correctas":      lambda b: b["ots_evaluadas"] - b["n_errores"],
                    "tiempo_ok_count":    lambda b: (b["ots_evaluadas"] * b["pct_tiempo_ok"]    / 100).round(0).astype(int),
                    "causa_ok_count":     lambda b: (b["ots_evaluadas"] * b["pct_causa_ok"]     / 100).round(0).astype(int),
                    "numeral_ok_count":   lambda b: (b["ots_evaluadas"] * b["pct_numeral_ok"]   / 100).round(0).astype(int),
                    "deteccion_ok_count": lambda b: (b["ots_evaluadas"] * b.get("pct_deteccion_ok", pd.Series(0, index=b.index)) / 100).round(0).astype(int),
                    "pct_deteccion_ok":   lambda b: pd.Series(0.0, index=b.index),
                }
                for _cg, _fb in _guards.items():
                    if _cg not in _tec_base.columns:
                        try: _tec_base[_cg] = _fb(_tec_base)
                        except Exception: _tec_base[_cg] = 0
                if "exactitud_pct" not in _tec_base.columns:
                    _tec_base["exactitud_pct"] = (
                        (1 - _tec_base["n_errores"] / _tec_base["ots_evaluadas"].clip(lower=1)) * 100
                    ).round(1)

                # Formato X/Y (Z%) por dimensión — tolerante a NaN
                def _fmt_dim(ok, total, pct):
                    try:
                        ok_i    = int(ok)    if pd.notna(ok)    else 0
                        total_i = int(total) if pd.notna(total) else 0
                        pct_f   = float(pct) if pd.notna(pct)   else 0.0
                        return f"{ok_i}/{total_i} ({pct_f:.1f}%)"
                    except (ValueError, TypeError):
                        return "—"

                # Si _tec_base quedó vacío (p.ej. técnico seleccionado no está
                # en TECNICOS_OCCIMIANO_FULL) creamos columnas vacías manualmente
                # para no romper el layout con apply sobre DataFrame vacío.
                if _tec_base.empty:
                    for _c in ("col_tiempo", "col_causa", "col_numeral", "col_deteccion"):
                        _tec_base[_c] = pd.Series(dtype="object")
                else:
                    _tec_base["col_tiempo"]    = _tec_base.apply(lambda r: _fmt_dim(r["tiempo_ok_count"],    r["ots_evaluadas"], r["pct_tiempo_ok"]),    axis=1)
                    _tec_base["col_causa"]     = _tec_base.apply(lambda r: _fmt_dim(r["causa_ok_count"],     r["ots_evaluadas"], r["pct_causa_ok"]),     axis=1)
                    _tec_base["col_numeral"]   = _tec_base.apply(lambda r: _fmt_dim(r["numeral_ok_count"],   r["ots_evaluadas"], r["pct_numeral_ok"]),   axis=1)
                    _tec_base["col_deteccion"] = _tec_base.apply(lambda r: _fmt_dim(r["deteccion_ok_count"], r["ots_evaluadas"], r["pct_deteccion_ok"]), axis=1)

                st.markdown('<div class="section-header">📋 Resumen por técnico</div>',
                            unsafe_allow_html=True)
                tec_disp = _tec_base[[
                    "tecnico", "ots_evaluadas", "ots_correctas", "n_errores", "err_total_dim",
                    "col_tiempo", "col_causa", "col_numeral",
                    "exactitud_pct", "bono_label",
                ]].copy()

                tec_disp.columns = [
                    "Técnico", "OTs evaluadas", "OTs sin error", "OTs con error", "Errores individuales",
                    "⏱ Tiempo OK", "🔍 Causa OK", "🔢 Numeral OK",
                    "Exactitud %", "Bono semanal",
                ]

                _show_df(
                    tec_disp, width="stretch", hide_index=True,
                    column_config={
                        "OTs evaluadas":       st.column_config.NumberColumn(format="%d"),
                        "OTs sin error":       st.column_config.NumberColumn(
                            help="OTs donde los 3 componentes estuvieron correctos.", format="%d"),
                        "OTs con error":       st.column_config.NumberColumn(
                            help="OTs con al menos 1 componente incorrecto — estas cuentan para el KPI.", format="%d"),
                        "Errores individuales":st.column_config.NumberColumn(
                            help="Suma de fallos por dimensión (una OT puede aportar hasta 3).", format="%d"),
                        "⏱ Tiempo OK":         st.column_config.TextColumn(help="OTs con tiempo correcto / total (%)"),
                        "🔍 Causa OK":          st.column_config.TextColumn(help="OTs con causa raíz válida / total (%)"),
                        "🔢 Numeral OK":        st.column_config.TextColumn(help="OTs con numeral registrado / total (%)"),
                        "Exactitud %":         st.column_config.ProgressColumn(min_value=0, max_value=100, format="%.1f%%"),
                    },
                )

            # ── Evolución mensual (últimos 6 meses) ───────────────────────────────
            st.divider()
            st.markdown('<div class="section-header">📈 Evolución mensual del índice de llenado</div>',
                        unsafe_allow_html=True)

            ultimos_6 = meses_disp[:6]  # ya están ordenados desc
            # Usar df_ot_all pre-computado: solo slice, sin recalcular scores
            ot_hist = df_ot_all[df_ot_all["mes"].isin(ultimos_6)].copy()
            if equipo_kpi != "Todos":
                _grp_kpi_h = _LABEL_TO_GRUPO.get(equipo_kpi)
                if _grp_kpi_h:
                    ot_hist = ot_hist[ot_hist["equipo"] == _grp_kpi_h]
            if tec_kpi_sel != "Todos":
                ot_hist = ot_hist[ot_hist["tecnico"] == tec_kpi_sel]
            # df_ot_all ya trae columna "mes" — no hace falta recalcular
            if not ot_hist.empty:
                trend = (
                    ot_hist.groupby("mes")
                    .agg(
                        score_global= ("score_total",  "mean"),
                        pct_tiempo=   ("score_tiempo", lambda x: (x >= 34).mean() * 100),
                        pct_causa=    ("causa_ok",     lambda x: x.mean() * 100),
                        pct_numeral=  ("numeral_ok",   lambda x: x.mean() * 100),
                        ots=          ("folio",        "count"),
                    )
                    .reset_index()
                    .sort_values("mes")
                )
                trend = trend.round(1)

                _trend_k = f"_fig_prec_trend_{_current_theme}_{_wo_sig}_{equipo_kpi}_{tec_kpi_sel}"
                if _trend_k not in st.session_state:
                    fig_trend = px.line(
                        trend, x="mes", y="score_global",
                        markers=True, text="score_global",
                        labels={"mes": "Mes", "score_global": "Score global (0-100)"},
                        title="Score global de llenado por mes",
                    )
                    fig_trend.update_traces(
                        textposition="top center",
                        line_color="#3b82f6",
                        marker=dict(size=8),
                    )
                    fig_trend.add_hline(y=100, line_dash="dash", line_color=_SCORE_COLOR["verde"],
                                        annotation_text="Score perfecto (0 errores)")
                    fig_trend.update_layout(
                        yaxis=dict(range=[0, 105]),
                        height=320,
                        margin=dict(t=40, b=30),
                    )
                    _apply_plot_theme(fig_trend)
                    st.session_state[_trend_k] = fig_trend
                st.plotly_chart(st.session_state[_trend_k], width="stretch")

            # ── Drill-down: OTs individuales de un técnico ────────────────────────
            st.divider()
            st.markdown('<div class="section-header">🔍 Detalle de OTs por técnico</div>',
                        unsafe_allow_html=True)
            st.caption(
                "Selecciona uno o varios técnicos para ver el historial completo de sus OTs del periodo. "
                "Vacío = todos los técnicos del filtro actual (equipo seleccionado)."
            )

            if not df_tec_scores.empty:
                _tec_opts = df_tec_scores["tecnico"].tolist()
                tec_drill_list = st.multiselect(
                    "Técnico(s)",
                    _tec_opts,
                    default=[],
                    key="kpi_drill_tec",
                    placeholder="Todos los técnicos del equipo",
                )
                if tec_drill_list:
                    df_drill = df_ot_scores[df_ot_scores["tecnico"].isin(tec_drill_list)].copy()
                    _tec_drill_lbl = (", ".join(tec_drill_list) if len(tec_drill_list) <= 3
                                      else f"{len(tec_drill_list)} técnicos seleccionados")
                else:
                    # Sin selección explícita → todos los técnicos disponibles en df_ot_scores
                    df_drill = df_ot_scores[df_ot_scores["tecnico"].isin(_tec_opts)].copy()
                    _tec_drill_lbl = f"Todos ({df_drill['tecnico'].nunique()} técnicos)"
                df_drill = df_drill.sort_values(["tecnico","score_total"], ascending=[True, True])

                if not df_drill.empty:
                    avg_drill = df_drill["score_total"].mean()
                    color_d, lbl_d = _score_level(avg_drill)
                    st.markdown(
                        f'<div style="background:{color_d}18;border-left:4px solid {color_d};'
                        f'border-radius:8px;padding:10px 14px;margin-bottom:8px;">'
                        f'<b style="color:{color_d}">{_tec_drill_lbl}</b> — '
                        f'Score promedio: <b>{avg_drill:.1f}/100</b> &nbsp; {lbl_d} &nbsp; | &nbsp; '
                        f'<b>{len(df_drill)}</b> OTs en {_mes_lbl_prec}</div>',
                        unsafe_allow_html=True,
                    )

                    # ── Construir columnas de display una por una (sin riesgo de desajuste) ──

                    drill_disp = df_drill.copy()

                    # Scores por componente (con guard para cache viejo)
                    _s = drill_disp["score_tiempo"]  if "score_tiempo"    in drill_disp.columns else pd.Series(0, index=drill_disp.index)
                    _c = drill_disp["score_causa"]   if "score_causa"     in drill_disp.columns else pd.Series(0, index=drill_disp.index)
                    _n = drill_disp["score_numeral"] if "score_numeral"   in drill_disp.columns else pd.Series(0, index=drill_disp.index)
                    _d = drill_disp["score_deteccion"] if "score_deteccion" in drill_disp.columns else pd.Series(0, index=drill_disp.index)
                    # KPI: solo 3 componentes (tiempo, causa, numeral) — Modalidad excluida
                    _ok_bits = ((_s >= 25).astype(int) + (_c >= 25).astype(int) +
                                (_n >= 25).astype(int))

                    # Columna 1: Cumple
                    drill_disp["_cumple"] = _ok_bits.apply(lambda n: "✅ Cumple" if n == 3 else "❌ Con error")

                    # Columna 2: X/3
                    drill_disp["_x4"] = _ok_bits.apply(
                        lambda n: f"{n}/3 {'✅' if n == 3 else ('⚠️' if n >= 2 else '❌')}")

                    # Columna 3: Tipo (title case)
                    drill_disp["_tipo"] = drill_disp["maint_type"].fillna("").str.title()

                    # Columna 3a: Tipo(s) de equipo del OT — icono + label.
                    # Sirve para leer de un vistazo por qué el numeral aplica o no
                    # (lavadora sí aplica; ablandador/bomba no; OT mixta muestra
                    # todos los tipos separados por '+').
                    _EQUIPO_TIPOS = [
                        (r"HIDROLAVAD", "💦 Hidrolavadora"),
                        (r"LAVAINT",    "🧼 Lavainteriores"),
                        (r"LAVAD",      "🚿 Lavadora"),
                        (r"ASPIRA",     "🧹 Aspiradora"),
                        (r"ABLAND",     "💧 Ablandador"),
                        (r"HIDROPACK",  "🛢️ Hidropack"),
                        (r"BOMBA",      "⚙️ Bomba"),
                        (r"TERMO",      "🔥 Termo"),
                        (r"COMPRESOR",  "🌬️ Compresor"),
                        (r"TWISTER",    "🌀 Twister"),
                        (r"SKID",       "🛠️ Skid"),
                    ]
                    import re as _re_tipo
                    def _fmt_tipo_equipo(nom: str) -> str:
                        s = str(nom or "").upper()
                        if not s or s == "—":
                            return "—"
                        tipos = []
                        for pat, lbl in _EQUIPO_TIPOS:
                            if _re_tipo.search(pat, s) and lbl not in tipos:
                                tipos.append(lbl)
                        return " + ".join(tipos) if tipos else "🔧 Otro"
                    _eq_src = (drill_disp["equipo_nombre"]
                               if "equipo_nombre" in drill_disp.columns
                               else pd.Series("", index=drill_disp.index))
                    drill_disp["_tipo_equipo"] = _eq_src.fillna("").apply(_fmt_tipo_equipo)

                    # Columna 3b: Modalidad de atención (informativa, no KPI).
                    # Simplificamos '1.- ATENDIDO PRESENCIAL' → 'Presencial',
                    # '2.- ATENDIDO VÍA REMOTA' → 'Remoto', etc.
                    _MOD_MAP = {
                        "1.- ATENDIDO PRESENCIAL": "🚗 Presencial",
                        "2.- ATENDIDO VÍA REMOTA": "📞 Remoto",
                        "2.- ATENDIDO VIA REMOTA": "📞 Remoto",
                        "3.- ATENDIDO CON SU MP":  "🔧 Con MP",
                        "4.- LLAMADO DUPLICADO":   "⚠️ Duplicado",
                        "SIN CLASIFICAR":          "❔ Sin clasificar",
                        "":                        "—",
                    }
                    _mod_src = (drill_disp["deteccion_raw"]
                                if "deteccion_raw" in drill_disp.columns
                                else pd.Series("", index=drill_disp.index))
                    drill_disp["_modalidad"] = _mod_src.fillna("").apply(
                        lambda v: _MOD_MAP.get(str(v).strip().upper(), str(v).strip().title() or "—")
                    )

                    # Columna 4: ⏱ Tiempo — valor en minutos + ✅/❌
                    _em = drill_disp["elapsed_min"] if "elapsed_min" in drill_disp.columns else pd.Series(0.0, index=drill_disp.index)
                    drill_disp["_col_tiempo"] = drill_disp.apply(
                        lambda r: f"{'✅' if _s[r.name] >= 25 else '❌'} {_em[r.name]:.0f} min",
                        axis=1)

                    # Columna 5: 🔍 Causa raíz — texto + ✅/❌
                    def _fmt_causa(r):
                        es_corr = r.get("es_correctiva", True)
                        raw = str(r.get("causa_raiz_raw","") or "").strip()
                        ok  = _c[r.name] >= 25
                        if not es_corr:
                            return "✅ Preventiva (no aplica)"
                        if ok:
                            return f"✅ {raw[:38]}" if raw else "✅ Registrada"
                        # Descuido atribuible: no clasificó pero SÍ describió la falla
                        if r.get("causa_sin_clasif_con_desglose", False):
                            return "🔴 No clasificó (sí describió la falla)"
                        return f"❌ {raw[:35]}" if raw else "❌ Sin causa"
                    drill_disp["_col_causa"] = drill_disp.apply(_fmt_causa, axis=1)

                    # Columna 6: 🔢 Numeral — calidad: no aplica / registrado / motivo del dato malo
                    def _fmt_col_numeral(r):
                        if not r.get("es_lavadora", True):
                            return "🔵 No aplica"
                        if r.get("numeral_ok", False):
                            return "✅ Registrado"
                        return NUMERAL_MOTIVO_LABEL.get(
                            str(r.get("numeral_motivo", "") or ""), "❌ Dato inválido")
                    drill_disp["_col_numeral"] = drill_disp.apply(
                        _fmt_col_numeral,
                        axis=1)

                    # Columna 7: 💬 Observación — descripción de qué falló (3 componentes KPI)
                    _nombres_comp = {0: "Tiempo", 1: "Causa raíz", 2: "Numeral"}
                    _scores_comp  = [_s, _c, _n]
                    def _obs(r):
                        fallos = [_nombres_comp[i] for i, sc in enumerate(_scores_comp)
                                  if sc[r.name] < 25]
                        if not fallos:
                            return "✅ Registro perfecto"
                        return "⚠️ No cumple: " + ", ".join(fallos)
                    drill_disp["_obs"] = drill_disp.apply(_obs, axis=1)

                    # Columna 8: Fecha cierre
                    drill_disp["_fecha"] = pd.to_datetime(
                        drill_disp["final_date"], errors="coerce"
                    ).dt.strftime("%d/%m/%Y")

                    # EDS y Técnico (para identificar la fila en vista multi-técnico)
                    if "eds_occim" in drill_disp.columns:
                        drill_disp["_eds"] = drill_disp["eds_occim"].fillna("").replace("", "—")
                    else:
                        drill_disp["_eds"] = "—"

                    # Estado OT
                    if "wo_status" in drill_disp.columns:
                        drill_disp["_estado_ot"] = drill_disp["wo_status"].fillna("").replace("", "—")
                    else:
                        drill_disp["_estado_ot"] = "—"

                    # Selección ordenada — Fecha primero (más reciente arriba),
                    # luego Cumple, y Tipo Equipo justo antes de Numeral para
                    # leer causalmente por qué aplica o no.
                    drill_disp = drill_disp[[
                        "_fecha", "_cumple", "_x4", "folio", "_eds", "tecnico", "station",
                        "_tipo", "_modalidad", "_estado_ot", "score_total",
                        "_col_tiempo", "_col_causa", "_tipo_equipo", "_col_numeral",
                        "_obs",
                    ]].copy()

                    drill_disp.columns = [
                        "Fecha", "Cumple", "X/3", "OT", "EDS", "Técnico", "Estación",
                        "Tipo", "Modalidad", "Estado", "Score",
                        "⏱ Tiempo", "🔍 Causa raíz", "🔧 Equipo", "🔢 Numeral",
                        "💬 Observación",
                    ]

                    # Orden por Fecha desc: más recientes arriba (parseando dd/mm/yyyy)
                    _fecha_ord = pd.to_datetime(drill_disp["Fecha"],
                                                format="%d/%m/%Y", errors="coerce")
                    drill_disp = drill_disp.assign(_ord=_fecha_ord) \
                                           .sort_values("_ord", ascending=False,
                                                        na_position="last") \
                                           .drop(columns="_ord")

                    # Alertas rápidas
                    _qt_criticas  = int((_s < 1).sum())
                    _qt_rapidas   = int(((_s >= 1) & (_s < 25)).sum())
                    _sin_causa    = int((_c == 0).sum())

                    if _qt_criticas or _qt_rapidas:
                        st.markdown(
                            f'<div style="background:{_t["err_bg"]};border-left:4px solid #ef4444;'
                            f'border-radius:6px;padding:8px 14px;margin-bottom:8px;font-size:0.85rem;color:{_t["text"]};">'
                            f'⚠️ Tiempo insuficiente: <b>{_qt_criticas} OTs</b> con tiempo mínimo · '
                            f'<b>{_qt_rapidas} OTs</b> con tiempo parcial</div>',
                            unsafe_allow_html=True,
                        )
                    if _sin_causa:
                        st.markdown(
                            f'<div style="background:{_t["orange_bg"]};border-left:4px solid #01798A;'
                            f'border-radius:6px;padding:8px 14px;margin-bottom:8px;font-size:0.85rem;color:{_t["text"]};">'
                            f'📋 <b>{_sin_causa} correctiva(s) sin causa raíz</b> registrada</div>',
                            unsafe_allow_html=True,
                        )

                    _show_df(
                        drill_disp,
                        width="stretch",
                        hide_index=True,
                        column_config={
                            "Cumple":           st.column_config.TextColumn(width=100,
                                help="✅ = 3/3 componentes OK · ❌ = ≥1 falló"),
                            "X/3":              st.column_config.TextColumn(width=80,
                                help="Componentes correctos de los 3 del KPI"),
                            "OT":               st.column_config.TextColumn(width=110),
                            "EDS":              st.column_config.TextColumn(width=85,
                                help="Código EDS Occimiano."),
                            "Técnico":          st.column_config.TextColumn(width=180),
                            "Estación":         st.column_config.TextColumn(width=200),
                            "Tipo":             st.column_config.TextColumn(width=110),
                            "Modalidad":        st.column_config.TextColumn(width=140,
                                help="Método de detección de falla en Fracttal (informativo, no incide en el KPI). "
                                     "Presencial = técnico fue a la EDS · Remoto = resuelto vía telefónica/remota · "
                                     "Con MP = aprovechó una mantención preventiva."),
                            "Estado":           st.column_config.TextColumn(width=110,
                                help="Estado de la OT en Fracttal."),
                            "Score":            st.column_config.ProgressColumn(
                                min_value=0, max_value=75, format="%.1f"),
                            "⏱ Tiempo":         st.column_config.TextColumn(width=110,
                                help="Minutos con Fracttal abierto. ✅ cumple ≥75% del estimado neto · MC no aplica (auto-25)."),
                            "🔍 Causa raíz":    st.column_config.TextColumn(width=240,
                                help="Causa registrada por el técnico. Solo se evalúa en MC; MP siempre da 25 auto."),
                            "🔧 Equipo":        st.column_config.TextColumn(width=170,
                                help="Tipo(s) de equipo del OT. El numeral SOLO aplica a Lavadora, Aspiradora, "
                                     "Lavainteriores e Hidrolavadora. Ablandador/Bomba/Termo/etc. → 'No aplica'. "
                                     "OTs con varios tipos se muestran separados por '+'."),
                            "🔢 Numeral":       st.column_config.TextColumn(width=160,
                                help="Calidad del numeral en lavadora/aspiradora (MC+MP). Motivo del dato malo si no cumple."),
                            "💬 Observación":   st.column_config.TextColumn(width=260,
                                help="Resumen de cumplimiento de la OT (3 componentes)"),
                            "Fecha":            st.column_config.TextColumn(width=90),
                        },
                    )


    # ══════════════════════════════════════════════════════════════════════════
    # TAB 4 — RESUMEN BONOS
    # ══════════════════════════════════════════════════════════════════════════
    with tab_bono:
        _bi_info, _bi_escala = st.columns([3, 2])
        with _bi_info:
            st.markdown(
                f'<div style="background:{_t["info_bg"]};border-left:4px solid #f59e0b;'
                f'border-radius:8px;padding:12px 16px;margin-bottom:14px;color:{_t["text"]};">'
                f'<b>🏆 Resumen de Bono Trimestral</b> — Estimación consolidada por técnico y equipo.<br>'
                f'<span style="font-size:0.85rem;">Pool equipo: <b>$500.000 CLP/trimestre</b> · '
                f'Dividido entre todos los integrantes (seniors incluidos) → monto por persona varía según tamaño del equipo.<br>'
                f'50% por KPIs individuales · 50% por KPIs del equipo · '
                f'Ponderación: SLA <b>40%</b> · Efectividad MP <b>30%</b> · Precisión Fracttal <b>30%</b></span>'
                f'</div>',
                unsafe_allow_html=True,
            )
        with _bi_escala:
            st.markdown(
                f'<div style="background:{_t["card"]};border:1px solid {_t["border"]};'
                f'border-radius:8px;padding:12px 14px;font-size:0.78rem;color:{_t["text"]};'
                f'line-height:1.7;margin-bottom:14px;">'
                f'<div style="font-weight:700;font-size:0.82rem;margin-bottom:6px;'
                f'color:{_t["muted"]};letter-spacing:0.04em;">📊 ESCALA DE BONOS</div>'
                f'<div style="margin-bottom:5px;">'
                f'<span style="font-weight:600;">SLA</span> '
                f'<span style="color:{_t["muted"]};">40% · $200.000/trim (pool)</span><br>'
                f'<span style="color:#22c55e;">≥95%→100%</span> &nbsp;'
                f'<span style="color:#16a34a;">93→90%</span> &nbsp;'
                f'<span style="color:#4ade80;">90→80%</span> &nbsp;'
                f'<span style="color:#f59e0b;">85→50%</span> &nbsp;'
                f'<span style="color:#ef4444;">&lt;85%→0%</span>'
                f'</div>'
                f'<div style="margin-bottom:5px;">'
                f'<span style="font-weight:600;">Calidad MP</span> '
                f'<span style="color:{_t["muted"]};">30% · $150.000/trim (pool)</span><br>'
                f'<span style="color:#22c55e;">≥98%→100%</span> &nbsp;'
                f'<span style="color:#16a34a;">96→90%</span> &nbsp;'
                f'<span style="color:#4ade80;">94→80%</span> &nbsp;'
                f'<span style="color:#65a30d;">92→70%</span> &nbsp;'
                f'<span style="color:#f59e0b;">90→60%</span> &nbsp;'
                f'<span style="color:#ef4444;">&lt;90%→0%</span>'
                f'</div>'
                f'<div>'
                f'<span style="font-weight:600;">Precisión</span> '
                f'<span style="color:{_t["muted"]};">30% · $150.000/trim (pool)</span><br>'
                f'<span style="color:#22c55e;">≥95%→100%</span> &nbsp;'
                f'<span style="color:#16a34a;">90→90%</span> &nbsp;'
                f'<span style="color:#4ade80;">85→80%</span> &nbsp;'
                f'<span style="color:#65a30d;">80→70%</span> &nbsp;'
                f'<span style="color:#f59e0b;">75→60%</span> &nbsp;'
                f'<span style="color:#f97316;">70→50%</span> &nbsp;'
                f'<span style="color:#ef4444;">&lt;70%→0%</span>'
                f'</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

        # ── Período ───────────────────────────────────────────────────────────
        _TRIM_DICT_BONO = {
            "T1 · Ene–Mar": [1, 2, 3], "T2 · Abr–Jun": [4, 5, 6],
            "T3 · Jul–Sep": [7, 8, 9], "T4 · Oct–Dic": [10, 11, 12],
        }
        _current_month = 6   # Junio 2026
        _default_trim_bono = next(
            (k for k, v in _TRIM_DICT_BONO.items() if _current_month in v),
            "T2 · Abr–Jun",
        )
        _bf0, _bf1, _bf2, _bf3 = st.columns([2, 3, 2, 2])
        with _bf0:
            _trim_bono = st.selectbox(
                "Trimestre/Período",
                list(_TRIM_DICT_BONO.keys()),
                index=list(_TRIM_DICT_BONO.keys()).index(_default_trim_bono),
                key="bono_trim",
            )
        _trim_months_bono = _TRIM_DICT_BONO[_trim_bono]

        # Todos los meses disponibles en los datos para el filtro de drill-down
        _all_meses_bono = sorted(
            set(
                list(
                    pd.period_range(
                        f"2026-{_trim_months_bono[0]:02d}",
                        f"2026-{_trim_months_bono[-1]:02d}",
                        freq="M",
                    ).astype(str)
                )
            )
        )
        with _bf1:
            _meses_bono_sel = st.multiselect(
                "Mes específico (dejar vacío = todo el trimestre)",
                options=_all_meses_bono,
                default=[],
                key="bono_mes_drill",
            )
        _meses_bono_activos = _meses_bono_sel if _meses_bono_sel else _all_meses_bono

        # ── Filtros equipo / técnico ──────────────────────────────────────────
        _eq_opts_bono = ["Todos"] + [_EQUIPO_LABEL.get(k, k) for k in GRUPOS_TERRENO]
        with _bf2:
            _bono_eq_sel = st.selectbox("Equipo", _eq_opts_bono, key="bono_equipo")
        # Obtener clave del equipo seleccionado
        _bono_eq_key = next(
            (k for k in GRUPOS_TERRENO if _EQUIPO_LABEL.get(k, k) == _bono_eq_sel),
            None,
        ) if _bono_eq_sel != "Todos" else None
        # Opciones de técnico: miembros del equipo seleccionado (o todos si equipo = Todos)
        if _bono_eq_key:
            _tec_opts_bono = ["Todos"] + [
                m for m in GRUPOS_TERRENO[_bono_eq_key].get("miembros", [])
                if not _es_excluido(TECH_NAME_MAP.get(m, m))
            ]
        else:
            _tec_opts_bono = ["Todos"]
        with _bf3:
            _bono_tec_sel = st.selectbox(
                "Técnico",
                _tec_opts_bono,
                key="bono_tecnico",
                disabled=(_bono_eq_sel == "Todos"),
            )

        # ── Cargar datos de SLA ───────────────────────────────────────────────
        _sla_key_bono = f"_sla_proc_v5_{len(df_llamados)}"
        _df_sla_bono = st.session_state.get(_sla_key_bono, pd.DataFrame())
        if not _df_sla_bono.empty and "mes" in _df_sla_bono.columns:
            _df_sla_bono = _df_sla_bono[
                _df_sla_bono["mes"].astype(str).isin(_meses_bono_activos)
            ].copy()
        # Columna normalizada para matching robusto (igual que MP) — evita fallos por tildes
        if not _df_sla_bono.empty and "_tech_norm" not in _df_sla_bono.columns:
            _df_sla_bono["_tech_norm"] = _df_sla_bono["tecnico"].fillna("").apply(
                lambda s: " ".join(_norm_n(s).split())
            )

        # ── Cargar datos de Precisión ─────────────────────────────────────────
        # Usar df_ot_all_scores_v4_direct (con aplicar_numerales_subtarea) para
        # que los valores coincidan con la pestaña Precisión Fracttal.
        _df_ot_bono = st.session_state.get(
            "df_ot_all_scores_v4_direct", pd.DataFrame()
        ).copy()
        if not _df_ot_bono.empty:
            if "_tech_norm" not in _df_ot_bono.columns:
                _df_ot_bono["_tech_norm"] = _df_ot_bono["tecnico"].fillna("").apply(
                    lambda s: " ".join(_norm_n(s).split())
                )
        if not _df_ot_bono.empty and "mes" in _df_ot_bono.columns:
            _df_ot_bono_filt = _df_ot_bono[
                _df_ot_bono["mes"].astype(str).isin(_meses_bono_activos)
            ].copy()
        else:
            _df_ot_bono_filt = pd.DataFrame()

        # ── Cargar datos de MP (reincidencias) ────────────────────────────────
        _df_reinc_bono = st.session_state.get("df_reinc", pd.DataFrame()).copy()
        if not _df_reinc_bono.empty and "falla_tipo" in _df_reinc_bono.columns:
            # Solo F.A.O (Falla Atribuible a Occimiano) cuenta para el bono
            _df_reinc_bono = _df_reinc_bono[
                _df_reinc_bono["falla_tipo"] == "fao"
            ].copy()
        if not _df_reinc_bono.empty and "falla_tipo" in _df_reinc_bono.columns and "causa_clasif" in _df_reinc_bono.columns:
            _df_reinc_bono["es_reincidencia_tecnico"] = (
                (_df_reinc_bono["falla_tipo"] == "fao") &
                (_df_reinc_bono["causa_clasif"] != "cliente")
            )
        # Filtro de período
        if not _df_reinc_bono.empty and "fecha_cm" in _df_reinc_bono.columns:
            _reinc_fecha = pd.to_datetime(_df_reinc_bono["fecha_cm"], errors="coerce")
            _df_reinc_bono = _df_reinc_bono[
                _reinc_fecha.dt.month.isin(_trim_months_bono)
            ].copy()
            if _meses_bono_sel:
                if "mes" not in _df_reinc_bono.columns:
                    _df_reinc_bono["mes"] = (
                        pd.to_datetime(_df_reinc_bono["fecha_cm"], errors="coerce")
                        .dt.to_period("M").astype(str)
                    )
                _df_reinc_bono = _df_reinc_bono[
                    _df_reinc_bono["mes"].astype(str).isin(_meses_bono_activos)
                ].copy()

        # PMs del período para denominador (solo clientes SLA — mismos que el numerador)
        _CLIENTES_SLA_BONO = {"COPEC", "Aramco (Esmax)", "ESMAX (Aramco)", "SHELL (Enex)", "ABASTIBLE"}
        _df_pm_bono = df_wo[
            (df_wo["maint_type"] == "Preventiva") &
            (~df_wo["technician"].apply(_es_excluido)) &
            (df_wo["equipo"] != "Sin equipo") &
            (df_wo["client"].isin(_CLIENTES_SLA_BONO) if "client" in df_wo.columns else True)
        ].copy()
        if not _df_pm_bono.empty:
            _pm_dates_bono = (
                _df_pm_bono["creation_date"].dt.tz_convert(None)
                if _df_pm_bono["creation_date"].dt.tz is not None
                else _df_pm_bono["creation_date"]
            )
            _df_pm_bono = _df_pm_bono[_pm_dates_bono.dt.month.isin(_trim_months_bono)].copy()
            if _meses_bono_sel:
                _pm_mes = (
                    _df_pm_bono["creation_date"].dt.tz_convert(None)
                    if _df_pm_bono["creation_date"].dt.tz is not None
                    else _df_pm_bono["creation_date"]
                )
                _df_pm_bono = _df_pm_bono[
                    _pm_mes.dt.to_period("M").astype(str).isin(_meses_bono_activos)
                ].copy()
            _df_pm_bono["_tech_norm"] = _df_pm_bono["technician"].fillna("").apply(
                lambda s: " ".join(_norm_n(s).split())
            )

        # ── Función helper: calcular KPIs para un técnico específico ────────
        def _kpi_para_tecnico(tech_full: str, equipo_key: str):
            """
            Retorna dict con SLA, MP y Prec para un técnico dado.
            tech_full: nombre completo (API)
            Para seniors, retorna el agregado del equipo completo en lugar de datos individuales.
            """
            # Seniors: su KPI = promedio del equipo (igual que ranking y tabla SLA)
            _short_snr = next((k for k, v in TECH_NAME_MAP.items() if v == tech_full), None)
            if _short_snr in SENIORS:
                return _kpi_para_equipo(equipo_key)

            _tn = " ".join(_norm_n(tech_full).split())

            # SLA — matching normalizado (igual que MP) para tolerar diferencias de tildes
            if not _df_sla_bono.empty and "_tech_norm" in _df_sla_bono.columns:
                _sla_t = _df_sla_bono[_df_sla_bono["_tech_norm"] == _tn]
                _n_sla_total = len(_sla_t)
                _n_sla_ok = int(_sla_t["cumple_sla"].fillna(False).sum()) if "cumple_sla" in _sla_t.columns else 0
                _pct_sla = (_n_sla_ok / _n_sla_total * 100) if _n_sla_total > 0 else None
            else:
                _n_sla_total = _n_sla_ok = 0
                _pct_sla = None

            # MP
            _pm_t = _df_pm_bono[_df_pm_bono["_tech_norm"] == _tn] if not _df_pm_bono.empty else pd.DataFrame()
            _n_pm_t = _pm_t["folio"].nunique() if not _pm_t.empty and "folio" in _pm_t.columns else 0
            if not _df_reinc_bono.empty and "tecnico_resp_short" in _df_reinc_bono.columns:
                _short = next(
                    (k for k, v in TECH_NAME_MAP.items() if v == tech_full),
                    tech_full,
                )
                _fallas_t = _df_reinc_bono[
                    _df_reinc_bono["tecnico_resp_short"] == _short
                ]
                _n_fallas_t = _fallas_t["folio_cm"].nunique() if "folio_cm" in _fallas_t.columns else len(_fallas_t)
            else:
                _n_fallas_t = 0

            # Prec — misma fórmula que KPI screen: % OTs con 3/3 componentes correctos
            # (binario all-or-nothing: score_total >= 75 → correcta = 3/3 componentes OK)
            # matching normalizado para tolerar diferencias de tildes (igual que MP)
            if not _df_ot_bono_filt.empty and "_tech_norm" in _df_ot_bono_filt.columns:
                _prec_t = _df_ot_bono_filt[_df_ot_bono_filt["_tech_norm"] == _tn]
                _n_ots_prec = len(_prec_t)
                if _n_ots_prec > 0 and "score_total" in _prec_t.columns:
                    _n_correctas_t = int((_prec_t["score_total"] >= 75).sum())
                    _pct_prec = _n_correctas_t / _n_ots_prec * 100
                else:
                    _pct_prec = None
                    _n_ots_prec = 0
                    _n_correctas_t = 0
            else:
                _pct_prec = None
                _n_ots_prec = 0
                _n_correctas_t = 0

            return {
                "n_sla_ok": _n_sla_ok, "n_sla_total": _n_sla_total, "pct_sla": _pct_sla,
                "n_fallas": _n_fallas_t, "n_pm": _n_pm_t,
                "pct_prec": _pct_prec, "n_ots_prec": _n_ots_prec, "n_correctas_prec": _n_correctas_t,
            }

        def _kpi_para_equipo(equipo_key: str):
            """Agrega KPIs para todos los técnicos del equipo."""
            # SLA
            if not _df_sla_bono.empty and "equipo" in _df_sla_bono.columns:
                _sla_e = _df_sla_bono[_df_sla_bono["equipo"] == equipo_key]
                _n_sla_total = len(_sla_e)
                _n_sla_ok = int(_sla_e["cumple_sla"].fillna(False).sum()) if "cumple_sla" in _sla_e.columns else 0
                _pct_sla = (_n_sla_ok / _n_sla_total * 100) if _n_sla_total > 0 else None
            else:
                _n_sla_total = _n_sla_ok = 0
                _pct_sla = None

            # MP
            _pm_e = _df_pm_bono[_df_pm_bono["equipo"] == equipo_key] if not _df_pm_bono.empty else pd.DataFrame()
            _n_pm_e = _pm_e["folio"].nunique() if not _pm_e.empty and "folio" in _pm_e.columns else 0
            if not _df_reinc_bono.empty and "grupo_responsable" in _df_reinc_bono.columns:
                _fallas_e = _df_reinc_bono[_df_reinc_bono["grupo_responsable"] == equipo_key]
                _n_fallas_e = _fallas_e["folio_cm"].nunique() if "folio_cm" in _fallas_e.columns else len(_fallas_e)
            else:
                _n_fallas_e = 0

            # Prec — misma fórmula que KPI screen: % OTs con 3/3 componentes correctos (score >= 75)
            if not _df_ot_bono_filt.empty and "equipo" in _df_ot_bono_filt.columns:
                _prec_e = _df_ot_bono_filt[_df_ot_bono_filt["equipo"] == equipo_key]
                _n_ots_e = len(_prec_e)
                if _n_ots_e > 0 and "score_total" in _prec_e.columns:
                    _n_correctas_e = int((_prec_e["score_total"] >= 75).sum())
                    _pct_prec = _n_correctas_e / _n_ots_e * 100
                else:
                    _pct_prec = None
                    _n_ots_e = 0
                    _n_correctas_e = 0
            else:
                _pct_prec = None
                _n_ots_e = 0
                _n_correctas_e = 0

            return {
                "n_sla_ok": _n_sla_ok, "n_sla_total": _n_sla_total, "pct_sla": _pct_sla,
                "n_fallas": _n_fallas_e, "n_pm": _n_pm_e,
                "pct_prec": _pct_prec, "n_ots_prec": _n_ots_e, "n_correctas_prec": _n_correctas_e,
            }

        # ── Helpers de formato HTML para celdas ──────────────────────────────
        def _cel_sla(n_ok, n_tot, pct):
            if pct is None or n_tot == 0:
                return f'<span style="color:{_t["muted"]};font-size:0.78rem;">Sin datos</span>'
            nivel, lbl, col, _ = _bono_sla(pct)
            return (
                f'<span style="font-size:0.82rem;">{n_ok}/{n_tot} = {pct:.1f}%</span><br>'
                f'<span style="color:{col};font-weight:700;font-size:0.80rem;">→ {nivel}%</span>'
            )

        def _cel_mp(n_fallas, n_pm):
            if n_pm == 0:
                return f'<span style="color:{_t["muted"]};font-size:0.78rem;">Sin PMs</span>'
            exactitud = (1 - n_fallas / n_pm) * 100
            nivel, _, col, _ = _bono_calidad(n_fallas, n_pm)
            return (
                f'<span style="font-size:0.82rem;">{n_pm - n_fallas}/{n_pm} = {exactitud:.1f}%</span><br>'
                f'<span style="color:{col};font-weight:700;font-size:0.80rem;">→ {nivel}%</span>'
            )

        def _cel_prec(pct, n_correctas=None, n_ots=None):
            if pct is None:
                return f'<span style="color:{_t["muted"]};font-size:0.78rem;">Sin datos</span>'
            nivel, _, col, _ = _bono_prec(pct)
            if n_correctas is not None and n_ots is not None and n_ots > 0:
                lbl = f'{n_correctas}/{n_ots} = {pct:.1f}%'
            else:
                lbl = f'{pct:.1f}%'
            return (
                f'<span style="font-size:0.82rem;">{lbl}</span><br>'
                f'<span style="color:{col};font-weight:700;font-size:0.80rem;">→ {nivel}%</span>'
            )

        def _cel_pond(pond_pct, col):
            return (
                f'<span style="color:{col};font-weight:800;font-size:1.0rem;">{pond_pct:.1f}%</span>'
            )

        def _clp_fmt(v):
            return f'${v:,.0f}'.replace(',', '.')

        def _nivel_color(nivel_int):
            if nivel_int >= 90:
                return "#22c55e"
            if nivel_int >= 70:
                return "#4ade80"
            if nivel_int >= 50:
                return "#f59e0b"
            if nivel_int > 0:
                return "#f97316"
            return "#ef4444"

        # ── Mostrar tabla por equipo ──────────────────────────────────────────
        _no_data_global = (
            _df_sla_bono.empty and _df_ot_bono_filt.empty and _df_reinc_bono.empty
        )
        if _no_data_global and _df_pm_bono.empty:
            st.info(
                "Los datos del período aún no están disponibles. "
                "Visita las pestañas SLA, Efectividad MP y Precisión para que se carguen, "
                "luego vuelve aquí."
            )
        else:
            # ── Rotación callcenter Santiago ──────────────────────────────────
            # Ciclo fijo: Juan Gallardo → Luis Pinto → Victor Bahamonde → ...
            # Referencia: lunes 16/03/2026 → Juan Gallardo (índice 0)
            # Con esta referencia Q2 (Abr-Jun 2026) queda: Gallardo=5 sem, resto=4 sem
            from datetime import date as _date_cc
            _CC_STGO   = ["Juan Gallardo", "Luis Pinto", "Victor Bahamonde"]
            _CC_REF    = _date_cc(2026, 3, 16)  # lunes de referencia
            _semanas_cc_por_equipo: dict = {}
            for _ms_cc in _meses_bono_activos:
                for _lun_cc in pd.date_range(
                    start=pd.Period(_ms_cc, "M").start_time,
                    end=pd.Period(_ms_cc, "M").end_time,
                    freq="W-MON",
                ):
                    _idx_cc = ((_lun_cc.date() - _CC_REF).days // 7) % len(_CC_STGO)
                    _eq_cc  = _CC_STGO[_idx_cc]
                    _semanas_cc_por_equipo[_eq_cc] = (
                        _semanas_cc_por_equipo.get(_eq_cc, 0) + 1
                    )
            # ── Fin rotación ──────────────────────────────────────────────────

            for _grp_key, _grp_info in GRUPOS_TERRENO.items():
                # ── Filtro equipo ─────────────────────────────────────────────
                if _bono_eq_key and _grp_key != _bono_eq_key:
                    continue

                _senior = _grp_info.get("senior", _grp_key)
                _miembros_short = _grp_info.get("miembros", [])
                # Convertir short names → full names
                _miembros_full = [TECH_NAME_MAP.get(m, m) for m in _miembros_short]
                # Filtrar excluidos
                _miembros_full = [
                    t for t in _miembros_full if not _es_excluido(t)
                ]
                if not _miembros_full:
                    continue
                # Guardar tamaño real del equipo ANTES de filtro por técnico
                # (para dividir el bono callcenter correctamente)
                _n_equipo_real = len(_miembros_full)

                # ── Filtro técnico ────────────────────────────────────────────
                if _bono_tec_sel != "Todos":
                    _bono_tec_full = TECH_NAME_MAP.get(_bono_tec_sel, _bono_tec_sel)
                    _miembros_full = [t for t in _miembros_full if t == _bono_tec_full]
                    if not _miembros_full:
                        continue

                st.markdown(
                    f'<div style="font-size:1.05rem;font-weight:700;color:{_t["text"]};'
                    f'margin:18px 0 8px 0;border-bottom:2px solid {_t["border"]};'
                    f'padding-bottom:4px;">🌐 Equipo {_EQUIPO_LABEL.get(_grp_key, _grp_key)}'
                    f' <span style="font-size:0.82rem;color:{_t["muted"]};font-weight:400;">'
                    f'— Senior: {_senior}</span></div>',
                    unsafe_allow_html=True,
                )

                # Calcular KPIs por técnico y para el equipo
                _tec_kpis = {t: _kpi_para_tecnico(t, _grp_key) for t in _miembros_full}
                _eq_kpi = _kpi_para_equipo(_grp_key)

                # ── Bono por persona: pool / n_integrantes (seniors incluidos) ──
                # 50 % individual (KPIs propios) · 50 % equipo (KPIs agregados)
                # IMPORTANTE: siempre dividir por el equipo completo (_n_equipo_real),
                # no por _miembros_full que puede estar filtrado a 1 persona.
                _n_pool       = _n_equipo_real
                _pp_max       = int(_BONO_TOTAL / _n_pool) if _n_pool > 0 else _BONO_TOTAL
                _pp_ind       = int(_pp_max * 0.50)   # parte individual
                _pp_eq        = int(_pp_max * 0.50)   # parte equipo
                _MAX_IND_SLA  = int(_pp_ind * 0.40)
                _MAX_IND_MP   = int(_pp_ind * 0.30)
                _MAX_IND_PREC = int(_pp_ind * 0.30)
                _MAX_EQ_SLA   = int(_pp_eq  * 0.40)
                _MAX_EQ_MP    = int(_pp_eq  * 0.30)
                _MAX_EQ_PREC  = int(_pp_eq  * 0.30)
                # Callcenter: $100K/sem × semanas de turno del equipo (rotación stgo)
                # Norte/Sur no participan → 0 semanas
                _BONO_CC_SEMANAL = 100_000
                _n_semanas_cc = _semanas_cc_por_equipo.get(_grp_key, 0)
                _BONO_CC = int(_BONO_CC_SEMANAL * _n_semanas_cc / _n_equipo_real) if _n_equipo_real > 0 else 0
                _BONO_CC_EQ = _BONO_CC_SEMANAL * _n_semanas_cc  # total equipo por período

                # ── Construir tabla HTML ──────────────────────────────────────
                _hdr_teal = "#01798A"
                _hdr_text = "#ffffff"
                _row_bg_alt = _t["card"]
                _row_bg     = _t.get("prog_bg", "#f3f4f6") if _t["card"] != "#ffffff" else "#f9fafb"

                # Cabecera — CSS last-child da fondo teal a columna EQUIPO
                _html = (
                    f'<style>.bono-tbl tbody tr td:last-child{{'
                    f'background:rgba(1,121,138,0.11)!important;}}</style>'
                    f'<div style="overflow-x:auto;margin-bottom:16px;">'
                    f'<table class="bono-tbl" style="width:100%;border-collapse:collapse;'
                    f'font-size:0.83rem;color:{_t["text"]};">'
                    f'<thead><tr>'
                    f'<th style="background:{_hdr_teal};color:{_hdr_text};padding:8px 10px;'
                    f'text-align:left;min-width:160px;border-radius:6px 0 0 0;">KPI</th>'
                )
                for _tf in _miembros_full:
                    _short_name = next(
                        (k for k, v in TECH_NAME_MAP.items() if v == _tf), _tf
                    )
                    _html += (
                        f'<th style="background:{_hdr_teal};color:{_hdr_text};'
                        f'padding:8px 10px;text-align:center;min-width:130px;">'
                        f'{_short_name}</th>'
                    )
                _html += (
                    f'<th style="background:{_hdr_teal};color:{_hdr_text};'
                    f'padding:8px 10px;text-align:center;min-width:130px;'
                    f'border-radius:0 6px 0 0;font-style:italic;">EQUIPO</th>'
                    f'</tr></thead><tbody>'
                )

                def _tr_bg(i):
                    return _row_bg if i % 2 == 0 else _row_bg_alt

                # Fila 1: SLA
                _html += (
                    f'<tr style="background:{_tr_bg(0)};">'
                    f'<td style="padding:8px 10px;font-weight:600;border-bottom:1px solid {_t["border"]};">'
                    f'Desempeño SLA <span style="color:{_t["muted"]};font-size:0.76rem;">(40%)</span></td>'
                )
                for _tf in _miembros_full:
                    _k = _tec_kpis[_tf]
                    _html += (
                        f'<td style="padding:8px 10px;text-align:center;'
                        f'border-bottom:1px solid {_t["border"]};">'
                        f'{_cel_sla(_k["n_sla_ok"], _k["n_sla_total"], _k["pct_sla"])}</td>'
                    )
                _html += (
                    f'<td style="padding:8px 10px;text-align:center;'
                    f'border-bottom:1px solid {_t["border"]};font-style:italic;">'
                    f'{_cel_sla(_eq_kpi["n_sla_ok"], _eq_kpi["n_sla_total"], _eq_kpi["pct_sla"])}</td>'
                    f'</tr>'
                )

                # Fila 2: MP
                _html += (
                    f'<tr style="background:{_tr_bg(1)};">'
                    f'<td style="padding:8px 10px;font-weight:600;border-bottom:1px solid {_t["border"]};">'
                    f'Efectividad MP <span style="color:{_t["muted"]};font-size:0.76rem;">(30%)</span></td>'
                )
                for _tf in _miembros_full:
                    _k = _tec_kpis[_tf]
                    _html += (
                        f'<td style="padding:8px 10px;text-align:center;'
                        f'border-bottom:1px solid {_t["border"]};">'
                        f'{_cel_mp(_k["n_fallas"], _k["n_pm"])}</td>'
                    )
                _html += (
                    f'<td style="padding:8px 10px;text-align:center;'
                    f'border-bottom:1px solid {_t["border"]};font-style:italic;">'
                    f'{_cel_mp(_eq_kpi["n_fallas"], _eq_kpi["n_pm"])}</td>'
                    f'</tr>'
                )

                # Fila 3: Precisión
                _html += (
                    f'<tr style="background:{_tr_bg(2)};">'
                    f'<td style="padding:8px 10px;font-weight:600;border-bottom:1px solid {_t["border"]};">'
                    f'Precisión Fracttal <span style="color:{_t["muted"]};font-size:0.76rem;">(30%)</span></td>'
                )
                for _tf in _miembros_full:
                    _k = _tec_kpis[_tf]
                    _html += (
                        f'<td style="padding:8px 10px;text-align:center;'
                        f'border-bottom:1px solid {_t["border"]};">'
                        f'{_cel_prec(_k["pct_prec"], _k.get("n_correctas_prec"), _k.get("n_ots_prec"))}</td>'
                    )
                _html += (
                    f'<td style="padding:8px 10px;text-align:center;'
                    f'border-bottom:1px solid {_t["border"]};font-style:italic;">'
                    f'{_cel_prec(_eq_kpi.get("pct_prec"), _eq_kpi.get("n_correctas_prec"), _eq_kpi.get("n_ots_prec"))}</td>'
                    f'</tr>'
                )

                # Fila 4: Cumplimiento ponderado (highlight)
                _html += (
                    f'<tr style="background:{_t.get("info_bg", "#eff6ff")};">'
                    f'<td style="padding:8px 10px;font-weight:700;border-bottom:1px solid {_t["border"]};">'
                    f'Cumplimiento ponderado</td>'
                )

                def _pond_pct(kpis: dict) -> tuple:
                    _p = kpis.get("pct_sla")
                    _m_pct = ((1 - kpis["n_fallas"] / kpis["n_pm"]) * 100) if kpis["n_pm"] > 0 else None
                    _r = kpis.get("pct_prec")
                    _niv_sla  = _bono_sla(_p)[0]  if _p  is not None else 0
                    _niv_mp   = _bono_calidad(kpis["n_fallas"], kpis["n_pm"])[0] if kpis["n_pm"] > 0 else 0
                    _niv_prec = _bono_prec(_r)[0] if _r is not None else 0
                    _has_data = (_p is not None) or (kpis["n_pm"] > 0) or (_r is not None)
                    if not _has_data:
                        return None, "#888"
                    _w = 0.40 * _niv_sla + 0.30 * _niv_mp + 0.30 * _niv_prec
                    return _w, _nivel_color(int(_w))

                for _tf in _miembros_full:
                    _pond, _pcol = _pond_pct(_tec_kpis[_tf])
                    if _pond is None:
                        _pcell = f'<span style="color:{_t["muted"]};font-size:0.78rem;">Sin datos</span>'
                    else:
                        _pcell = _cel_pond(_pond, _pcol)
                    _html += (
                        f'<td style="padding:8px 10px;text-align:center;'
                        f'border-bottom:1px solid {_t["border"]};">{_pcell}</td>'
                    )
                _eq_pond, _eq_pcol = _pond_pct(_eq_kpi)
                if _eq_pond is None:
                    _eq_pcell = f'<span style="color:{_t["muted"]};font-size:0.78rem;">Sin datos</span>'
                else:
                    _eq_pcell = _cel_pond(_eq_pond, _eq_pcol)
                _html += (
                    f'<td style="padding:8px 10px;text-align:center;'
                    f'border-bottom:1px solid {_t["border"]};font-style:italic;">'
                    f'{_eq_pcell}</td></tr>'
                )

                # Fila 5: Bono individual estimado
                _html += (
                    f'<tr style="background:{_tr_bg(4)};">'
                    f'<td style="padding:8px 10px;font-weight:600;border-bottom:1px solid {_t["border"]};">'
                    f'Terreno individual <span style="color:{_t["muted"]};font-size:0.76rem;">'
                    f'({_clp_fmt(_pp_ind)} × cumpl. personal)</span></td>'
                )
                for _tf in _miembros_full:
                    _k = _tec_kpis[_tf]
                    _niv_sla_i  = _bono_sla(_k["pct_sla"])[0] if _k["pct_sla"] is not None else 0
                    _niv_mp_i   = _bono_calidad(_k["n_fallas"], _k["n_pm"])[0] if _k["n_pm"] > 0 else 0
                    _niv_prec_i = _bono_prec(_k["pct_prec"])[0] if _k["pct_prec"] is not None else 0
                    _bono_ind = int(
                        _MAX_IND_SLA  * _niv_sla_i  / 100 +
                        _MAX_IND_MP   * _niv_mp_i   / 100 +
                        _MAX_IND_PREC * _niv_prec_i / 100
                    )
                    _has_any = _k["pct_sla"] is not None or _k["n_pm"] > 0 or _k["pct_prec"] is not None
                    if not _has_any:
                        _bi_cell = f'<span style="color:{_t["muted"]};font-size:0.78rem;">Sin datos</span>'
                    else:
                        _bc = _nivel_color(int(0.40 * _niv_sla_i + 0.30 * _niv_mp_i + 0.30 * _niv_prec_i))
                        _bi_cell = f'<span style="color:{_bc};font-weight:700;">{_clp_fmt(_bono_ind)}</span>'
                    _html += (
                        f'<td style="padding:8px 10px;text-align:center;'
                        f'border-bottom:1px solid {_t["border"]};">{_bi_cell}</td>'
                    )
                _html += (
                    f'<td style="padding:8px 10px;text-align:center;'
                    f'border-bottom:1px solid {_t["border"]};color:{_t["muted"]};'
                    f'font-style:italic;font-size:0.80rem;">—</td></tr>'
                )

                # Fila 6: Bono equipo estimado (igual para todos)
                _eq_niv_sla  = _bono_sla(_eq_kpi["pct_sla"])[0] if _eq_kpi["pct_sla"] is not None else 0
                _eq_niv_mp   = _bono_calidad(_eq_kpi["n_fallas"], _eq_kpi["n_pm"])[0] if _eq_kpi["n_pm"] > 0 else 0
                _eq_niv_prec = _bono_prec(_eq_kpi.get("pct_prec"))[0] if _eq_kpi.get("pct_prec") is not None else 0
                _bono_eq_est = int(
                    _MAX_EQ_SLA  * _eq_niv_sla  / 100 +
                    _MAX_EQ_MP   * _eq_niv_mp   / 100 +
                    _MAX_EQ_PREC * _eq_niv_prec / 100
                )
                _eq_has_any = _eq_kpi["pct_sla"] is not None or _eq_kpi["n_pm"] > 0 or _eq_kpi.get("pct_prec") is not None
                if not _eq_has_any:
                    _be_val_cell = f'<span style="color:{_t["muted"]};font-size:0.78rem;">Sin datos</span>'
                else:
                    _eq_bc = _nivel_color(int(0.40 * _eq_niv_sla + 0.30 * _eq_niv_mp + 0.30 * _eq_niv_prec))
                    _be_val_cell = f'<span style="color:{_eq_bc};font-weight:700;">{_clp_fmt(_bono_eq_est)}</span>'

                _eq_pond_lbl = f'{_eq_pond:.0f}% eq.' if _eq_pond is not None else '% eq.'
                _html += (
                    f'<tr style="background:{_tr_bg(5)};">'
                    f'<td style="padding:8px 10px;font-weight:600;border-bottom:1px solid {_t["border"]};">'
                    f'Terreno colectivo <span style="color:{_t["muted"]};font-size:0.76rem;">'
                    f'({_clp_fmt(_pp_eq)} × {_eq_pond_lbl})</span></td>'
                )
                for _tf in _miembros_full:
                    _html += (
                        f'<td style="padding:8px 10px;text-align:center;'
                        f'border-bottom:1px solid {_t["border"]};">{_be_val_cell}</td>'
                    )
                _html += (
                    f'<td style="padding:8px 10px;text-align:center;'
                    f'border-bottom:1px solid {_t["border"]};font-style:italic;">'
                    f'{_be_val_cell}</td></tr>'
                )

                # Fila 6.5: Callcenter ($100K/equipo/sem ÷ N pers × 13 sem)
                if _n_semanas_cc > 0:
                    _cc_cell    = f'<span style="font-weight:700;">{_clp_fmt(_BONO_CC)}</span>'
                    _cc_eq_cell = f'<span style="font-weight:700;font-style:italic;">{_clp_fmt(_BONO_CC_EQ)}</span>'
                    _cc_lbl     = (f'$100K/sem &times; {_n_semanas_cc} sem de turno'
                                   f' &divide; {_n_equipo_real} pers')
                else:
                    _cc_cell    = f'<span style="color:{_t["muted"]};font-size:0.82rem;">—</span>'
                    _cc_eq_cell = f'<span style="color:{_t["muted"]};font-size:0.82rem;">—</span>'
                    _cc_lbl     = 'no participa en rotación'
                _html += (
                    f'<tr style="background:{_tr_bg(6)};">'
                    f'<td style="padding:8px 10px;font-weight:600;border-bottom:1px solid {_t["border"]};">'
                    f'Callcenter <span style="color:{_t["muted"]};font-size:0.76rem;">'
                    f'({_cc_lbl})</span></td>'
                )
                for _tf in _miembros_full:
                    _html += (
                        f'<td style="padding:8px 10px;text-align:center;'
                        f'border-bottom:1px solid {_t["border"]};">{_cc_cell}</td>'
                    )
                _html += (
                    f'<td style="padding:8px 10px;text-align:center;'
                    f'border-bottom:1px solid {_t["border"]};font-style:italic;">'
                    f'{_cc_eq_cell}</td></tr>'
                )

                # Fila 7: TOTAL trimestral (individual + equipo + callcenter)
                _totales_trim = {}
                _html += (
                    f'<tr style="background:{_t.get("info_bg", "#eff6ff")};">'
                    f'<td style="padding:9px 10px;font-weight:800;font-size:0.90rem;'
                    f'border-top:2px solid {_t["border"]};">TOTAL trimestral</td>'
                )
                for _tf in _miembros_full:
                    _k = _tec_kpis[_tf]
                    _niv_sla_i  = _bono_sla(_k["pct_sla"])[0] if _k["pct_sla"] is not None else 0
                    _niv_mp_i   = _bono_calidad(_k["n_fallas"], _k["n_pm"])[0] if _k["n_pm"] > 0 else 0
                    _niv_prec_i = _bono_prec(_k["pct_prec"])[0] if _k["pct_prec"] is not None else 0
                    _bono_ind_t = int(
                        _MAX_IND_SLA  * _niv_sla_i  / 100 +
                        _MAX_IND_MP   * _niv_mp_i   / 100 +
                        _MAX_IND_PREC * _niv_prec_i / 100
                    )
                    _total_t = _bono_ind_t + _bono_eq_est + _BONO_CC
                    _has_any_t = _k["pct_sla"] is not None or _k["n_pm"] > 0 or _k["pct_prec"] is not None
                    _totales_trim[_tf] = (_total_t, _has_any_t)
                    if not _has_any_t and not _eq_has_any:
                        _tot_cell = f'<span style="color:{_t["muted"]};font-size:0.78rem;">Sin datos</span>'
                    else:
                        _pond_tot = (
                            0.40 * _niv_sla_i + 0.30 * _niv_mp_i + 0.30 * _niv_prec_i
                            if _has_any_t else
                            0.40 * _eq_niv_sla + 0.30 * _eq_niv_mp + 0.30 * _eq_niv_prec
                        )
                        _tc = _nivel_color(int(_pond_tot))
                        _tot_cell = (
                            f'<span style="color:{_tc};font-weight:800;font-size:1.0rem;">'
                            f'{_clp_fmt(_total_t)}</span>'
                        )
                    _html += (
                        f'<td style="padding:9px 10px;text-align:center;'
                        f'border-top:2px solid {_t["border"]};">{_tot_cell}</td>'
                    )
                _html += (
                    f'<td style="padding:9px 10px;text-align:center;'
                    f'border-top:2px solid {_t["border"]};color:{_t["muted"]};'
                    f'font-style:italic;font-size:0.80rem;">—</td></tr>'
                )

                # Fila 8: Promedio mensual (trimestral ÷ 3)
                _html += (
                    f'<tr style="background:{_t.get("info_bg", "#eff6ff")};">'
                    f'<td style="padding:8px 10px;font-weight:700;font-size:0.87rem;">'
                    f'Promedio mensual <span style="color:{_t["muted"]};font-size:0.76rem;">(÷ 3)</span></td>'
                )
                for _tf in _miembros_full:
                    _tot_v, _has_v = _totales_trim[_tf]
                    if not _has_v and not _eq_has_any:
                        _mens_cell = f'<span style="color:{_t["muted"]};font-size:0.78rem;">—</span>'
                    else:
                        _mens_cell = (
                            f'<span style="font-weight:700;font-size:0.92rem;">'
                            f'{_clp_fmt(_tot_v // 3)}</span>'
                        )
                    _html += (
                        f'<td style="padding:8px 10px;text-align:center;">{_mens_cell}</td>'
                    )
                _html += (
                    f'<td style="padding:8px 10px;text-align:center;'
                    f'color:{_t["muted"]};font-style:italic;font-size:0.80rem;">—</td></tr>'
                )

                _html += '</tbody></table></div>'
                st.markdown(_html, unsafe_allow_html=True)

            # ── Disclaimer ────────────────────────────────────────────────────
            st.caption(
                "Estimación basada en los datos del período seleccionado. "
                "El bono equipo aplica igual para todos los miembros del equipo según el "
                "desempeño agregado. Los valores son orientativos y pueden diferir del "
                "cálculo oficial."
            )

            # ── Advertencia si faltan datos ───────────────────────────────────
            _missing = []
            if _df_sla_bono.empty:
                _missing.append("SLA (visita la pestaña 'Desempeño SLA')")
            if _df_reinc_bono.empty and _df_pm_bono.empty:
                _missing.append("Efectividad MP (visita la pestaña 'Efectividad MP')")
            if _df_ot_bono_filt.empty:
                _missing.append("Precisión Fracttal (visita la pestaña 'Precisión Fracttal')")
            if _missing:
                st.info(
                    "Datos no disponibles para: **" + "**, **".join(_missing) + "**. "
                    "Visita las pestañas correspondientes para que se carguen los datos."
                )

    # ══════════════════════════════════════════════════════════════════════════
    # EXPORTAR DATOS STO PARA APP MÓVIL (sto_data.json)
    # ══════════════════════════════════════════════════════════════════════════
    try:
        import json as _json_mod
        _sto_export_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sto_data.json")

        # 1. SLA — datos completos (sin filtro de trimestre)
        _sla_full = st.session_state.get(f"_sla_proc_v5_{len(df_llamados)}", pd.DataFrame())
        _sla_records = []
        if not _sla_full.empty:
            _sla_exp = _sla_full[["tecnico", "equipo", "cumple_sla", "mes"]].copy() if {"tecnico","equipo","cumple_sla","mes"}.issubset(_sla_full.columns) else pd.DataFrame()
            if not _sla_exp.empty:
                _sla_exp["mes_num"] = _sla_exp["mes"].apply(lambda m: int(str(m).split("-")[1]) if "-" in str(m) else int(m) if str(m).isdigit() else 0)
                _sla_exp["cumple"] = _sla_exp["cumple_sla"].fillna(False).astype(bool)
                _sla_g = _sla_exp.groupby(["tecnico", "equipo", "mes_num"]).agg(
                    cumple=("cumple", "sum"), total=("cumple", "count"),
                ).reset_index()
                _sla_g["tecnico"] = _sla_g["tecnico"].str.replace(r'\s+', ' ', regex=True).str.strip()
                _sla_records = _sla_g.to_dict(orient="records")

        # 2. Precisión — datos completos (sin filtro de trimestre)
        # Usar df_ot_all_scores_v4_direct (incluye aplicar_numerales_subtarea + filtro canceladas)
        # para que coincida con lo que muestra la pestaña Precisión Fracttal.
        _ot_full = st.session_state.get("df_ot_all_scores_v4_direct",
                   st.session_state.get("df_ot_bono_scores", pd.DataFrame()))
        _prec_records = []
        if not _ot_full.empty and "score_total" in _ot_full.columns:
            _prec_exp = _ot_full[["tecnico", "equipo", "score_total", "mes"]].copy() if {"tecnico","equipo","score_total","mes"}.issubset(_ot_full.columns) else pd.DataFrame()
            if not _prec_exp.empty:
                _prec_exp["mes_num"] = _prec_exp["mes"].apply(lambda m: int(str(m).split("-")[1]) if "-" in str(m) else int(m) if str(m).isdigit() else 0)
                _prec_exp["buena"] = _prec_exp["score_total"] >= 75
                _prec_g = _prec_exp.groupby(["tecnico", "equipo", "mes_num"]).agg(
                    buenas=("buena", "sum"), total=("buena", "count"),
                ).reset_index()
                _prec_g["tecnico"] = _prec_g["tecnico"].str.replace(r'\s+', ' ', regex=True).str.strip()
                _prec_records = _prec_g.to_dict(orient="records")

        # 3. Efectividad — reincidencias completas
        _reinc_full = st.session_state.get("df_reinc", pd.DataFrame())
        _reinc_records = []
        if not _reinc_full.empty and "falla_tipo" in _reinc_full.columns:
            _reinc_exp = _reinc_full[_reinc_full["falla_tipo"] == "fao"].copy()
            if "fecha_cm" in _reinc_exp.columns:
                _reinc_exp["mes_num"] = pd.to_datetime(_reinc_exp["fecha_cm"], errors="coerce").dt.month
            if "tecnico_resp_short" in _reinc_exp.columns and "grupo_responsable" in _reinc_exp.columns:
                _reinc_g = _reinc_exp.groupby(
                    ["tecnico_resp_short", "grupo_responsable", "mes_num"]
                ).agg(fallas=("folio_cm", "nunique") if "folio_cm" in _reinc_exp.columns else ("tecnico_resp_short", "count")).reset_index()
                _reinc_records = _reinc_g.rename(columns={"tecnico_resp_short": "tecnico_short", "grupo_responsable": "equipo"}).to_dict(orient="records")

        # 4. PMs — para denominador de efectividad
        _CLIENTES_SLA_EXP = {"COPEC", "Aramco (Esmax)", "ESMAX (Aramco)", "SHELL (Enex)", "ABASTIBLE"}
        _pm_records = []
        if not df_wo.empty:
            _pm_exp = df_wo[
                (df_wo["maint_type"] == "Preventiva") &
                (~df_wo["technician"].apply(_es_excluido)) &
                (df_wo["equipo"] != "Sin equipo") &
                (df_wo["client"].isin(_CLIENTES_SLA_EXP) if "client" in df_wo.columns else True)
            ].copy()
            if not _pm_exp.empty:
                _pm_dates = _pm_exp["creation_date"].dt.tz_convert(None) if _pm_exp["creation_date"].dt.tz is not None else _pm_exp["creation_date"]
                _pm_exp["mes_num"] = _pm_dates.dt.month
                _pm_exp["_tech_norm"] = _pm_exp["technician"].fillna("").apply(lambda s: " ".join(_norm_n(s).split()))
                _pm_g = _pm_exp.groupby(["technician", "equipo", "mes_num"]).agg(
                    pms=("folio", "nunique") if "folio" in _pm_exp.columns else ("technician", "count"),
                ).reset_index()
                _pm_g["technician"] = _pm_g["technician"].str.replace(r'\s+', ' ', regex=True).str.strip()
                _pm_records = _pm_g.rename(columns={"technician": "tecnico"}).to_dict(orient="records")

        # 5. Tech name maps y equipos
        _tech_map_exp = {k: v for k, v in TECH_NAME_MAP.items()}
        _equipos_exp = {}
        for _gk, _gv in GRUPOS_TERRENO.items():
            _miembros = [TECH_NAME_MAP.get(m, m) for m in _gv.get("miembros", [])]
            _equipos_exp[_gk] = {
                "label": _EQUIPO_LABEL.get(_gk, _gk),
                "senior": _gv.get("senior", ""),
                "miembros": _miembros,
            }

        _export_obj = {
            "updated_at": datetime.now().isoformat(),
            "sla": _sla_records,
            "precision": _prec_records,
            "reincidencias": _reinc_records,
            "pms": _pm_records,
            "tech_name_map": _tech_map_exp,
            "full_to_short": {v: k for k, v in TECH_NAME_MAP.items()},
            "equipos": _equipos_exp,
            "seniors": list(SENIORS),
        }

        # 6. Bono table pre-computed for mobile app
        try:
            from datetime import date as _dt_cc
            _yr = datetime.now().year
            _TRIMS_E = {"T1":[1,2,3],"T2":[4,5,6],"T3":[7,8,9],"T4":[10,11,12]}
            _CC_STGO_E = ["Juan Gallardo","Luis Pinto","Victor Bahamonde"]
            _CC_REF_E = _dt_cc(2026,3,16)
            _bono_tbl = {}
            for _tk,_tm in _TRIMS_E.items():
                _cc_w = {}
                for _mm in _tm:
                    _ps = f"{_yr}-{_mm:02d}"
                    for _lun in pd.date_range(pd.Period(_ps,"M").start_time,pd.Period(_ps,"M").end_time,freq="W-MON"):
                        _eq_r = _CC_STGO_E[(_lun.date()-_CC_REF_E).days//7 % 3]
                        _cc_w[_eq_r] = _cc_w.get(_eq_r,0)+1
                _sf = [r for r in _sla_records if r.get("mes_num") in _tm]
                _pf = [r for r in _prec_records if r.get("mes_num") in _tm]
                _mfr = [r for r in _pm_records if r.get("mes_num") in _tm]
                _rfr = [r for r in _reinc_records if r.get("mes_num") in _tm]
                _bt = []
                for _gk,_gv in GRUPOS_TERRENO.items():
                    _ms = [m for m in _gv.get("miembros",[]) if not _es_excluido(TECH_NAME_MAP.get(m,m))]
                    _mfl = [TECH_NAME_MAP.get(m,m) for m in _ms]; _n = len(_mfl)
                    if not _n: continue
                    eso = sum(r.get("cumple",0) for r in _sf if r.get("equipo")==_gk)
                    est = sum(r.get("total",0) for r in _sf if r.get("equipo")==_gk)
                    epb = sum(r.get("buenas",0) for r in _pf if r.get("equipo")==_gk)
                    ept = sum(r.get("total",0) for r in _pf if r.get("equipo")==_gk)
                    epm = sum(r.get("pms",0) for r in _mfr if r.get("equipo")==_gk)
                    efl = sum(r.get("fallas",0) for r in _rfr if r.get("equipo")==_gk)
                    esp = round(eso/est*100,1) if est else None
                    emp = round((1-efl/epm)*100,1) if epm else None
                    epp = round(epb/ept*100,1) if ept else None
                    ens = _bono_sla(esp)[0] if esp is not None else 0
                    enm = _bono_calidad(efl,epm)[0] if epm else 0
                    enp = _bono_prec(epp)[0] if epp is not None else 0
                    ec = round(.40*ens+.30*enm+.30*enp,1)
                    ppi = int(int(_BONO_TOTAL/_n)*.50); ppe = ppi
                    ncc = _cc_w.get(_gk,0); bcc = int(100000*ncc/_n) if _n else 0
                    be = int(ppe*.40*ens/100+ppe*.30*enm/100+ppe*.30*enp/100)
                    _trs = []
                    for tf in _mfl:
                        ts = next((k for k,v in TECH_NAME_MAP.items() if v==tf),tf)
                        iss = ts in SENIORS
                        if iss:
                            so,st2,pb,pt,fl,pm = eso,est,epb,ept,efl,epm
                        else:
                            so = sum(r.get("cumple",0) for r in _sf if r.get("tecnico")==tf)
                            st2 = sum(r.get("total",0) for r in _sf if r.get("tecnico")==tf)
                            pb = sum(r.get("buenas",0) for r in _pf if r.get("tecnico")==tf)
                            pt = sum(r.get("total",0) for r in _pf if r.get("tecnico")==tf)
                            pm = sum(r.get("pms",0) for r in _mfr if r.get("tecnico")==tf)
                            fl = sum(r.get("fallas",0) for r in _rfr if r.get("tecnico_short")==ts)
                        sp = round(so/st2*100,1) if st2 else None
                        mp2 = round((1-fl/pm)*100,1) if pm else None
                        pp2 = round(pb/pt*100,1) if pt else None
                        ns = _bono_sla(sp)[0] if sp is not None else 0
                        nm = _bono_calidad(fl,pm)[0] if pm else 0
                        np2 = _bono_prec(pp2)[0] if pp2 is not None else 0
                        c = round(.40*ns+.30*nm+.30*np2,1)
                        bi = int(ppi*.40*ns/100+ppi*.30*nm/100+ppi*.30*np2/100)
                        tot = bi+be+bcc
                        _trs.append({"short":ts,"is_senior":iss,
                            "sla_pct":sp,"sla_ok":so,"sla_tot":st2,"sla_niv":ns,
                            "mp_pct":mp2,"mp_f":fl,"mp_pm":pm,"mp_niv":nm,
                            "prec_pct":pp2,"prec_b":pb,"prec_t":pt,"prec_niv":np2,
                            "cumpl":c,"bono_ind":bi,"bono_eq":be,"bono_cc":bcc,
                            "total_trim":tot,"prom_mensual":tot//3})
                    _bt.append({"key":_gk,"label":_EQUIPO_LABEL.get(_gk,_gk),
                        "senior":_gv.get("senior",""),"n_eq":_n,
                        "pp_ind":ppi,"pp_eq":ppe,"n_semanas_cc":ncc,"bono_cc_eq":100000*ncc,
                        "tecs":_trs,
                        "eq":{"sla_pct":esp,"sla_ok":eso,"sla_tot":est,"sla_niv":ens,
                            "mp_pct":emp,"mp_f":efl,"mp_pm":epm,"mp_niv":enm,
                            "prec_pct":epp,"prec_b":epb,"prec_t":ept,"prec_niv":enp,
                            "cumpl":ec}})
                _bono_tbl[_tk] = _bt
            _export_obj["bono_table"] = _bono_tbl
        except Exception:
            pass

        with open(_sto_export_path, "w", encoding="utf-8") as _jf:
            _json_mod.dump(_export_obj, _jf, ensure_ascii=False, default=str)
        # Subir a Supabase para que la app móvil lo lea remotamente
        try:
            from supabase_client import _post as _post_sb
            _post_sb("sto_data_export", {
                "id": "latest",
                "data": _export_obj,
                "updated_at": datetime.now().isoformat(),
            })
        except Exception:
            pass
    except Exception:
        pass

# ─────────────────────────────────────────────────────────────────────────────
# PÁGINA 5: MANTENCIONES PREVENTIVAS
# ─────────────────────────────────────────────────────────────────────────────
elif _page == _NAV_PAGES[2]:
    _hdr(_PAGE_TITLE[_NAV_PAGES[2]], "Órdenes de mantenimiento preventivo — Fracttal One · 2026")
    st.divider()

    # ── Carga ─────────────────────────────────────────────────────────────
    with st.spinner("Cargando mantenciones preventivas…"):
        _raw_prev = load_preventivas_supabase()
    if not _raw_prev:
        st.warning("Sin datos de mantenciones preventivas. Verifica la sincronización con Fracttal.")
        st.stop()

    df_prev = pd.DataFrame(_raw_prev)
    df_prev["fecha_creacion"]    = pd.to_datetime(df_prev["fecha_creacion"],    errors="coerce", utc=True)
    df_prev["fecha_finalizacion"]= pd.to_datetime(df_prev["fecha_finalizacion"], errors="coerce", utc=True)
    _fc_prev  = df_prev["fecha_creacion"].dt.tz_convert(None)
    df_prev["_mes"]   = _fc_prev.dt.to_period("M").astype(str)
    df_prev["_month"] = _fc_prev.dt.month.astype("Int64")

    # ── Normalizar estado_tarea (Fracttal devuelve valores en inglés) ─────────
    _ESTADO_TAREA_MAP = {"DONE": "Finalizada", "NO_STARTED": "No Iniciada"}
    df_prev["estado_tarea"] = df_prev["estado_tarea"].replace(_ESTADO_TAREA_MAP)

    # ── Formatear duración (segundos → "HH:MM") ───────────────────────────────
    def _seg_a_hhmm(seg) -> str:
        try:
            s = int(seg)
            if s <= 0: return "—"
            h, m = divmod(s, 3600)
            return f"{h:02d}:{m // 60:02d}"
        except Exception:
            return "—"

    df_prev["dur_est_fmt"]  = df_prev["duracion_estim_seg"].apply(_seg_a_hhmm) \
        if "duracion_estim_seg" in df_prev.columns else "—"
    df_prev["dur_real_fmt"] = df_prev["duracion_real_seg"].apply(_seg_a_hhmm) \
        if "duracion_real_seg" in df_prev.columns else "—"

    # ── Fecha de inicio (local, sin tz) ──────────────────────────────────────
    if "fecha_inicio" in df_prev.columns:
        df_prev["fecha_inicio_fmt"] = (
            pd.to_datetime(df_prev["fecha_inicio"], errors="coerce", utc=True)
            .dt.tz_convert(None)
            .dt.strftime("%d/%m/%Y %H:%M")
        )

    # ── Rango fijo: Ene 2026 → hoy ────────────────────────────────────────
    _rango_inicio = pd.Timestamp("2026-01-01")
    _rango_fin    = pd.Timestamp.today().normalize()
    df_prev = df_prev[(_fc_prev >= _rango_inicio) & (_fc_prev <= _rango_fin)].copy()

    # ── Filtros fila 1 ────────────────────────────────────────────────────
    _pTRIM = {"T1 (Ene–Mar)":[1,2,3],"T2 (Abr–Jun)":[4,5,6],
              "T3 (Jul–Sep)":[7,8,9],"T4 (Oct–Dic)":[10,11,12]}
    _pf1, _pf2, _pf3, _pf4, _pf5 = st.columns([1.1, 1.3, 1.5, 1.5, 1.5])
    with _pf1:
        sel_ptrim = st.selectbox("Trimestre", ["Todos"] + list(_pTRIM.keys()), key="prev_trim")
    with _pf2:
        import calendar as _cal
        _pmeses_raw = sorted(df_prev["_mes"].dropna().unique(), reverse=True)
        def _pym_lbl(ym):
            try: y,m = ym.split("-"); return f"{_cal.month_abbr[int(m)]} {y}"
            except: return ym
        _pmes_map = {_pym_lbl(m): m for m in _pmeses_raw}
        sel_pmes = st.multiselect("Mes", list(_pmes_map.keys()), key="prev_mes",
                                  placeholder="Todos los meses")
    with _pf3:
        _ptipo_opts = ["Todos"] + sorted(df_prev["tipo_tarea"].dropna().unique().tolist())
        sel_ptipo = st.selectbox("Tipo de tarea", _ptipo_opts, key="prev_tipo")
    with _pf4:
        _pactiv_opts = ["Todos"] + sorted(df_prev["activador"].dropna().unique().tolist())
        sel_pactiv = st.selectbox("Activador", _pactiv_opts, key="prev_activ")
    with _pf5:
        _presp_opts = ["Todos"] + sorted(df_prev["responsable"].dropna().unique().tolist())
        sel_presp = st.selectbox("Responsable", _presp_opts, key="prev_resp")

    # ── Filtros fila 2 ────────────────────────────────────────────────────
    _pf6, _pf7, _pf8, _pf9, _pf10 = st.columns([1.1, 1.3, 1.5, 1.5, 1.5])
    with _pf6:
        sel_pestado = st.selectbox("Estado",
            ["Todos"] + sorted(df_prev["estado"].dropna().unique().tolist()), key="prev_estado")
    with _pf7:
        sel_pestarea = st.selectbox("Estado tarea",
            ["Todos"] + sorted(df_prev["estado_tarea"].dropna().unique().tolist()), key="prev_estarea")
    with _pf8:
        sel_pclasi = st.text_input("Clasificación 2", key="prev_clasi",
                                   placeholder="60198  ó  SH_736…")
    with _pf9:
        _ppl_opts = ["Todos"] + sorted(df_prev["plan_tareas"].dropna().unique().tolist()) \
            if "plan_tareas" in df_prev.columns else ["Todos"]
        sel_pplan = st.selectbox("Plan de tareas", _ppl_opts, key="prev_plan")

    # Helper: orden de clientes — primero los 3 principales, luego alfabético
    def _orden_clientes(series_clientes) -> list:
        _principales = ["COPEC", "Aramco (Esmax)", "SHELL (Enex)"]
        _unicos = set(series_clientes.dropna().unique().tolist())
        _ord = [c for c in _principales if c in _unicos]
        _resto = sorted(c for c in _unicos if c not in _principales)
        return ["Todos"] + _ord + _resto

    with _pf10:
        sel_pcli = st.selectbox(
            "Cliente",
            _orden_clientes(df_prev["cliente"]) if "cliente" in df_prev.columns else ["Todos"],
            key="prev_cliente",
        )

    # ── Aplicar filtros ───────────────────────────────────────────────────
    dfp = df_prev.copy()
    if sel_ptrim != "Todos":
        dfp = dfp[dfp["_month"].isin(_pTRIM[sel_ptrim])]
    if sel_pmes:
        _ppers = [_pmes_map[l] for l in sel_pmes if l in _pmes_map]
        if _ppers: dfp = dfp[dfp["_mes"].isin(_ppers)]
    if sel_ptipo   != "Todos": dfp = dfp[dfp["tipo_tarea"]  == sel_ptipo]
    if sel_pactiv  != "Todos": dfp = dfp[dfp["activador"]   == sel_pactiv]
    if sel_presp   != "Todos": dfp = dfp[dfp["responsable"] == sel_presp]
    if sel_pestado != "Todos": dfp = dfp[dfp["estado"]      == sel_pestado]
    if sel_pestarea!= "Todos": dfp = dfp[dfp["estado_tarea"]== sel_pestarea]
    if sel_pclasi.strip():
        dfp = dfp[dfp["clasificacion_2"].str.contains(sel_pclasi.strip(), case=False, na=False)]
    if sel_pplan != "Todos" and "plan_tareas" in dfp.columns:
        dfp = dfp[dfp["plan_tareas"] == sel_pplan]
    if sel_pcli != "Todos" and "cliente" in dfp.columns:
        dfp = dfp[dfp["cliente"] == sel_pcli]

    # ── df sin filtros de período (usado en tabs Planificación y EDS) ─────
    # Aplica tipo/activador/responsable/estado/clasif pero NO trimestre ni mes,
    # para que esos tabs puedan mostrar el rango completo de fecha_programada.
    _dfp_full = df_prev.copy()
    if sel_ptipo   != "Todos": _dfp_full = _dfp_full[_dfp_full["tipo_tarea"]  == sel_ptipo]
    if sel_pactiv  != "Todos": _dfp_full = _dfp_full[_dfp_full["activador"]   == sel_pactiv]
    if sel_presp   != "Todos": _dfp_full = _dfp_full[_dfp_full["responsable"] == sel_presp]
    if sel_pestado != "Todos": _dfp_full = _dfp_full[_dfp_full["estado"]      == sel_pestado]
    if sel_pestarea!= "Todos": _dfp_full = _dfp_full[_dfp_full["estado_tarea"]== sel_pestarea]
    if sel_pclasi.strip():
        _dfp_full = _dfp_full[_dfp_full["clasificacion_2"].str.contains(
            sel_pclasi.strip(), case=False, na=False)]
    if sel_pplan != "Todos" and "plan_tareas" in _dfp_full.columns:
        _dfp_full = _dfp_full[_dfp_full["plan_tareas"] == sel_pplan]
    if sel_pcli != "Todos" and "cliente" in _dfp_full.columns:
        _dfp_full = _dfp_full[_dfp_full["cliente"] == sel_pcli]

    # ── Universo válido (sin OTs anuladas) — usado en todos los KPIs ──────
    # OTs no operativas que NO representan trabajo real:
    #   • Cancelado: anulada explícitamente
    #   • ERROR DE INGRESO: creada por error de digitación
    #   • EQUIPO CON RECAMBIO: el equipo ya no aplica al plan
    # Las descontamos del total para que TODOS los KPIs (Total, Finalizadas,
    # velocímetro de cumplimiento, etc.) hablen del mismo universo.
    _ESTADOS_NO_CUENTAN = {
        "Cancelado", "Canceladas", "Cancelada",
        "ERROR DE INGRESO",
        "EQUIPO CON RECAMBIO",
    }
    _dfp_op = dfp[~dfp["estado"].isin(_ESTADOS_NO_CUENTAN)].copy()
    _anuladas_n = int(dfp["estado"].isin(_ESTADOS_NO_CUENTAN).sum())

    # ── KPIs ──────────────────────────────────────────────────────────────
    _ptot   = len(_dfp_op)
    _fin_mask = _dfp_op["estado_tarea"].isin(["Finalizada"]) | _dfp_op["estado"].isin(["Finalizadas"])
    _noi_mask = _dfp_op["estado_tarea"].isin(["No Iniciada"]) & ~_dfp_op["estado"].isin(["Finalizadas"])
    _pfin   = int(_fin_mask.sum())
    _pproc  = int((~_fin_mask & ~_noi_mask).sum())
    _pnoi   = int(_noi_mask.sum())
    _ppct   = round(_pfin / _ptot * 100, 1) if _ptot > 0 else 0.0

    # ── % Cumplimiento en tiempo y forma ──────────────────────────────────
    # Criterio Occimiano (v2 — con banda de anticipación):
    #   CUMPLE si la MP se ejecuta dentro de una ventana razonable respecto
    #   a la fecha programada, considerando el intervalo del plan.
    #
    #   Límite tarde:   fecha_finalización ≤ fecha_programada (mismo día o antes)
    #   Límite temprano: ejecución NO puede adelantarse más del 50% del
    #                    intervalo del plan. Ejemplo: MP mensual (30 días)
    #                    programada al día 30 y ejecutada el día 14 → adelanto
    #                    de 16 días = 53% del intervalo → INCUMPLE por anticipación
    #                    excesiva (la próxima quedará ~45 días después y genera reclamos).
    #
    # El intervalo se infiere del nombre del plan (plan_tareas):
    #   MENSUAL=30, BIMESTRAL=60, TRIMESTRAL=90, SEMESTRAL=180, ANUAL=365, etc.
    #   Si no se puede inferir, default = 30 días (mensual).
    #
    # Universo evaluado:
    #   • Solo OTs cuya fecha_programada ya pasó (las futuras no son atraso).
    #   • Anuladas ya excluidas vía _dfp_op arriba.
    # Inferir el intervalo de programación de forma EMPÍRICA desde el histórico:
    # para cada (codigo_activo, plan_tareas) miramos la mediana de días entre
    # fechas programadas consecutivas y hacemos snap a los valores estándar.
    # Esto es más robusto que parsear el nombre del plan (los planes Occimiano
    # se llaman "PLAN MTTO MSELF..." sin decir frecuencia).
    #
    # Fallback: si un (activo, plan) tiene solo una MP, usamos default = 30 días.
    _INTERVALOS_ESTANDAR = [7, 15, 30, 60, 90, 120, 180, 365, 730]
    _INTERVALO_DEFAULT   = 30

    def _snap_intervalo(dias: float) -> int:
        """Redondea el intervalo empírico al valor estándar más cercano en escala log."""
        if pd.isna(dias) or dias <= 0:
            return _INTERVALO_DEFAULT
        import math
        return min(_INTERVALOS_ESTANDAR,
                   key=lambda v: abs(math.log(v) - math.log(dias)))

    def _build_intervalos_por_activo_plan(df_full: pd.DataFrame) -> dict:
        """
        Retorna {(codigo_activo, plan_tareas): intervalo_dias} basado en la
        mediana de días entre fechas programadas consecutivas del mismo par.
        """
        if df_full.empty or "fecha_programada" not in df_full.columns:
            return {}
        _dfi = df_full[["codigo_activo","plan_tareas","fecha_programada"]].dropna(
            subset=["fecha_programada","codigo_activo"]
        ).copy()
        _dfi["plan_tareas"] = _dfi["plan_tareas"].fillna("(sin plan)")
        _dfi["_fp"] = pd.to_datetime(_dfi["fecha_programada"], errors="coerce", utc=True)
        _dfi = _dfi.dropna(subset=["_fp"]).sort_values(["codigo_activo","plan_tareas","_fp"])
        result: dict = {}
        for (act, plan), grp in _dfi.groupby(["codigo_activo","plan_tareas"], sort=False):
            fechas = grp["_fp"].tolist()
            if len(fechas) < 2:
                continue
            diffs = [(fechas[i+1] - fechas[i]).days for i in range(len(fechas)-1)
                     if (fechas[i+1] - fechas[i]).days > 0]
            if not diffs:
                continue
            med = sorted(diffs)[len(diffs)//2]  # mediana
            result[(act, plan)] = _snap_intervalo(med)
        return result

    # Construir el índice de intervalos usando TODAS las MPs (no solo las que
    # ya pasaron), para tener más datos históricos.
    _intv_idx = _build_intervalos_por_activo_plan(_dfp_op)

    def _lookup_intervalo(row) -> int:
        key = (row.get("codigo_activo"), str(row.get("plan_tareas") or "(sin plan)"))
        return _intv_idx.get(key, _INTERVALO_DEFAULT)

    _hoy_norm = pd.Timestamp.today().normalize()
    _df_cumpl = _dfp_op[_dfp_op["fecha_programada"].notna()].copy()
    _excluidas_n = _anuladas_n

    if not _df_cumpl.empty:
        # Convertir a hora local Chile antes de normalizar para no perder/ganar
        # un día en OTs nocturnas guardadas en UTC.
        _fp = pd.to_datetime(_df_cumpl["fecha_programada"],   errors="coerce", utc=True).dt.tz_convert("America/Santiago").dt.tz_localize(None).dt.normalize()
        _ff = pd.to_datetime(_df_cumpl["fecha_finalizacion"], errors="coerce", utc=True).dt.tz_convert("America/Santiago").dt.tz_localize(None).dt.normalize()
        _df_cumpl["_fp_n"] = _fp
        _df_cumpl["_ff_n"] = _ff
        # Solo OTs cuya fecha programada ya pasó
        _df_cumpl = _df_cumpl[_df_cumpl["_fp_n"] <= _hoy_norm]

        # Días de atraso:
        #   - Si está finalizada: ff - fp
        #   - Si NO está finalizada: hoy - fp (atraso indefinido, en curso)
        _df_cumpl["dias_atraso"] = (
            _df_cumpl["_ff_n"].fillna(_hoy_norm) - _df_cumpl["_fp_n"]
        ).dt.days

        # Intervalo de programación (días) inferido EMPÍRICAMENTE desde el
        # histórico de MPs del mismo (codigo_activo, plan_tareas).
        _df_cumpl["_intervalo_dias"] = _df_cumpl.apply(_lookup_intervalo, axis=1)

        # Anticipación: días entre ejecución y fecha programada (positivo = adelantada).
        # Solo se calcula para MPs finalizadas.
        _df_cumpl["dias_antelacion"] = (
            _df_cumpl["_fp_n"] - _df_cumpl["_ff_n"]
        ).dt.days
        _df_cumpl["pct_antelacion"] = (
            _df_cumpl["dias_antelacion"] / _df_cumpl["_intervalo_dias"] * 100
        ).round(1)
        # "Muy anticipada" = adelantada más del 50% del intervalo del plan.
        _df_cumpl["_muy_anticipada"] = (
            _df_cumpl["_ff_n"].notna() &
            (_df_cumpl["pct_antelacion"] > 50)
        )

        # Cumplimiento CRUDO (nuevo): no tarde AND no muy anticipada
        _df_cumpl["cumple"] = (
            _df_cumpl["_ff_n"].notna() &
            (_df_cumpl["dias_atraso"] <= 0) &
            (~_df_cumpl["_muy_anticipada"])
        )

        # Motivo del incumplimiento (para diagnóstico)
        def _motivo_incumpl(r):
            if bool(r.get("cumple", False)):
                return ""
            if pd.isna(r.get("_ff_n")):
                return f"⏳ Pendiente ({int(r['dias_atraso'])}d de atraso sin ejecutar)"
            if r["dias_atraso"] > 0:
                return f"⚠️ Ejecutada {int(r['dias_atraso'])} día(s) tarde"
            if bool(r.get("_muy_anticipada", False)):
                return (f"🔵 Anticipación excesiva: hecha {int(r['dias_antelacion'])}d "
                        f"antes ({r['pct_antelacion']:.0f}% del intervalo de "
                        f"{int(r['_intervalo_dias'])}d)")
            return ""
        _df_cumpl["_motivo"] = _df_cumpl.apply(_motivo_incumpl, axis=1)

        # Cumplimiento FLEXIBLE (≤ 5 días desde fecha programada, sin techo de anticipación)
        _df_cumpl["cumple_sem"] = (
            _df_cumpl["_ff_n"].notna() & (_df_cumpl["dias_atraso"] <= 5)
        ) | (
            _df_cumpl["_ff_n"].isna()
            & ((_hoy_norm - _df_cumpl["_fp_n"]).dt.days <= 5)
        )

        _cump_n  = int(_df_cumpl["cumple"].sum())
        _cump_tot= len(_df_cumpl)
        _cump_pct= round(_cump_n / _cump_tot * 100, 1) if _cump_tot > 0 else 0.0
        _atras_avg = round(_df_cumpl.loc[~_df_cumpl["cumple"], "dias_atraso"].mean(), 1) \
            if (~_df_cumpl["cumple"]).any() else 0.0

        # KPI de anticipación excesiva
        _antic_n   = int(_df_cumpl["_muy_anticipada"].sum())
        _antic_pct = round(_antic_n / _cump_tot * 100, 1) if _cump_tot > 0 else 0.0

        _cump_sem_n  = int(_df_cumpl["cumple_sem"].sum())
        _cump_sem_pct= round(_cump_sem_n / _cump_tot * 100, 1) if _cump_tot > 0 else 0.0
    else:
        _cump_n = _cump_tot = _cump_sem_n = _antic_n = 0
        _cump_pct = _cump_sem_pct = _antic_pct = 0.0
        _atras_avg = 0.0

    pk1,pk2,pk3,pk4,pk5 = st.columns(5)
    pk1.metric("Total OTs (operativas)", f"{_ptot:,}",
               help=f"Excluye {_anuladas_n} OTs anuladas (Cancelado / Error de ingreso / Equipo con recambio).")
    pk2.metric("Finalizadas",            f"{_pfin:,}",  delta=f"{_ppct}%")
    pk3.metric("En Proceso / Revisión",  f"{_pproc:,}")
    pk4.metric("No Iniciadas",           f"{_pnoi:,}")
    pk5.metric("% Completadas",          f"{_ppct}%")
    if _anuladas_n > 0:
        st.caption(
            f"ℹ️  Hay **{_anuladas_n} OTs anuladas** que no se cuentan en estos KPIs "
            f"(Cancelado / Error de ingreso / Equipo con recambio)."
        )

    # ── Velocímetros: cumplimiento diario y semanal ───────────────────────
    import plotly.graph_objects as _go

    def _gauge_color_of(pct: float) -> str:
        return "#10b981" if pct >= 90 else "#f59e0b" if pct >= 75 else "#ef4444"

    def _build_gauge(value: float, title_main: str, subtitle: str) -> "_go.Figure":
        color = _gauge_color_of(value)
        fig = _go.Figure(_go.Indicator(
            mode="gauge+number",
            value=value,
            number={"suffix": " %", "font": {"size": 32}},
            title={"text": f"<b>{title_main}</b><br>"
                           f"<span style='font-size:0.72rem;color:#94a3b8'>{subtitle}</span>",
                   "font": {"size": 13}},
            gauge={
                "axis":    {"range": [0, 100], "tickwidth": 1, "tickfont": {"size": 10}},
                "bar":     {"color": color, "thickness": 0.32},
                "bgcolor": "rgba(0,0,0,0)",
                "borderwidth": 0,
                "steps": [
                    {"range": [0,  75], "color": "rgba(239,68,68,0.18)"},
                    {"range": [75, 90], "color": "rgba(245,158,11,0.18)"},
                    {"range": [90,100], "color": "rgba(16,185,129,0.18)"},
                ],
            },
        ))
        fig.update_layout(height=260, margin=dict(l=10, r=10, t=60, b=10),
                          paper_bgcolor="rgba(0,0,0,0)")
        return fig

    # Gauge dedicado para anticipación indebida (color invertido: menos es mejor)
    def _antic_color_of(pct: float) -> str:
        # Anticipación: <5% bien, 5-15% revisar, >15% crítico
        return "#10b981" if pct <= 5 else "#f59e0b" if pct <= 15 else "#ef4444"

    def _build_antic_gauge(value: float, title_main: str, subtitle: str) -> "_go.Figure":
        color = _antic_color_of(value)
        fig = _go.Figure(_go.Indicator(
            mode="gauge+number",
            value=value,
            number={"suffix": " %", "font": {"size": 32}},
            title={"text": f"<b>{title_main}</b><br>"
                           f"<span style='font-size:0.72rem;color:#94a3b8'>{subtitle}</span>",
                   "font": {"size": 13}},
            gauge={
                "axis":    {"range": [0, 50], "tickwidth": 1, "tickfont": {"size": 10}},
                "bar":     {"color": color, "thickness": 0.32},
                "bgcolor": "rgba(0,0,0,0)",
                "borderwidth": 0,
                "steps": [
                    {"range": [0,   5], "color": "rgba(16,185,129,0.18)"},
                    {"range": [5,  15], "color": "rgba(245,158,11,0.18)"},
                    {"range": [15, 50], "color": "rgba(239,68,68,0.18)"},
                ],
            },
        ))
        fig.update_layout(height=260, margin=dict(l=10, r=10, t=60, b=10),
                          paper_bgcolor="rgba(0,0,0,0)")
        return fig

    _gc1, _gc2, _gc3 = st.columns(3)
    with _gc1:
        st.plotly_chart(
            _build_gauge(_cump_pct, "% Cumplimiento — Criterio Crudo",
                         "en ventana (no tarde y no >50% adelantada)"),
            use_container_width=True, key="prev_gauge_diario",
        )
    with _gc2:
        st.plotly_chart(
            _build_gauge(_cump_sem_pct, "% Cumplimiento — Criterio Flexible (≤ 5 días)",
                         "ejecutada ≤ 5 días desde fecha programada · por OT"),
            use_container_width=True, key="prev_gauge_semanal",
        )
    with _gc3:
        st.plotly_chart(
            _build_antic_gauge(_antic_pct, "% Anticipación indebida",
                               "ejecutadas >50% antes del intervalo del plan"),
            use_container_width=True, key="prev_gauge_antic",
        )

    # Panel de detalle unificado debajo de los velocímetros
    _gauge_color   = _gauge_color_of(_cump_pct)
    _gauge_color_s = _gauge_color_of(_cump_sem_pct)
    _gauge_color_a = _antic_color_of(_antic_pct)
    st.markdown(
        f"""<div style="padding:14px 18px;background:rgba(148,163,184,0.08);
             border-radius:10px;border-left:3px solid {_gauge_color};margin-top:-12px;">
          <div style="font-size:0.85rem;color:#94a3b8;letter-spacing:0.04em;
                      font-weight:600;text-transform:uppercase;margin-bottom:8px;">
            Detalle del cumplimiento</div>
          <table style="width:100%;border-collapse:collapse;font-size:0.92rem;">
            <tr style="border-bottom:1px solid rgba(148,163,184,0.25);">
              <td style="padding:6px 0;font-weight:600;color:var(--text-color, #0f172a);">Criterio Crudo (no tarde y no muy anticipada)</td>
              <td style="padding:6px 0;text-align:right;">
                <b style="color:{_gauge_color};font-size:1.05rem;">{_cump_n:,} / {_cump_tot:,}</b>
                <span style="color:var(--text-color, #475569);font-size:0.82rem;"> OTs OK · {_cump_pct}%</span>
              </td>
            </tr>
            <tr style="border-bottom:1px solid rgba(148,163,184,0.25);">
              <td style="padding:6px 0;font-weight:600;color:var(--text-color, #0f172a);">Criterio Flexible (≤ 5 días desde F.Prog.)</td>
              <td style="padding:6px 0;text-align:right;">
                <b style="color:{_gauge_color_s};font-size:1.05rem;">{_cump_sem_n:,} / {_cump_tot:,}</b>
                <span style="color:var(--text-color, #475569);font-size:0.82rem;"> OTs en plazo · {_cump_sem_pct}%</span>
              </td>
            </tr>
            <tr style="border-bottom:1px solid rgba(148,163,184,0.25);">
              <td style="padding:6px 0;font-weight:600;color:var(--text-color, #0f172a);">Anticipación indebida (&gt;50% del intervalo)</td>
              <td style="padding:6px 0;text-align:right;">
                <b style="color:{_gauge_color_a};font-size:1.05rem;">{_antic_n:,} / {_cump_tot:,}</b>
                <span style="color:var(--text-color, #475569);font-size:0.82rem;"> OTs muy antes · {_antic_pct}%</span>
              </td>
            </tr>
            <tr>
              <td style="padding:6px 0;font-size:0.84rem;color:var(--text-color, #475569);">Promedio de atraso (solo las tarde)</td>
              <td style="padding:6px 0;text-align:right;color:#ef4444;font-weight:600;">
                {_atras_avg} días
              </td>
            </tr>
          </table>
          <div style="color:#94a3b8;font-size:0.78rem;margin-top:10px;
                      padding-top:8px;border-top:1px solid rgba(148,163,184,0.18);">
            Solo OTs cuya fecha programada ya pasó.
            Se excluyeron <b>{_excluidas_n:,}</b> OTs anuladas (Cancelado / Error de ingreso).
            <b style="margin-left:8px;">Escala Cumplimiento:</b>
            <span style="color:#10b981">≥ 90% bien</span> ·
            <span style="color:#f59e0b">75–90% revisar</span> ·
            <span style="color:#ef4444">&lt; 75% crítico</span>
            <b style="margin-left:12px;">Anticipación:</b>
            <span style="color:#10b981">≤ 5% bien</span> ·
            <span style="color:#f59e0b">5–15% revisar</span> ·
            <span style="color:#ef4444">&gt; 15% crítico</span>
          </div>
        </div>""",
        unsafe_allow_html=True,
    )

    # ── Gráfico: % anticipación indebida por tipo de plan/programación ─────
    # Muestra si el problema de "hacer MPs demasiado antes" es sistémico
    # o solo afecta a ciertos tipos de programación (mensual / trimestral / etc.).
    if _cump_tot > 0:
        _df_a = _df_cumpl.copy()
        _MAP_INTV_LBL = {
            7:   "Semanal (7d)",     15:  "Quincenal (15d)",
            30:  "Mensual (30d)",    60:  "Bimestral (60d)",
            90:  "Trimestral (90d)", 120: "Cuatrimestral (120d)",
            180: "Semestral (180d)", 365: "Anual (365d)",
            730: "Bianual (730d)",
        }
        _df_a["_intv_lbl"] = _df_a["_intervalo_dias"].map(_MAP_INTV_LBL).fillna(
            _df_a["_intervalo_dias"].astype(str) + "d"
        )
        _agg_a = (_df_a.groupby(["_intervalo_dias","_intv_lbl"], as_index=False)
                       .agg(total=("cumple","size"),
                            antic=("_muy_anticipada","sum"),
                            tarde=("dias_atraso", lambda s: int((s > 0).sum())),
                            ok=("cumple","sum"))
                       .sort_values("_intervalo_dias"))
        _agg_a["pct_antic"] = (_agg_a["antic"] / _agg_a["total"] * 100).round(1)
        _agg_a["pct_tarde"] = (_agg_a["tarde"] / _agg_a["total"] * 100).round(1)
        _agg_a["pct_ok"]    = (_agg_a["ok"]    / _agg_a["total"] * 100).round(1)

        _fig_antic = _go.Figure()
        _fig_antic.add_trace(_go.Bar(
            name="✅ En ventana", x=_agg_a["_intv_lbl"], y=_agg_a["pct_ok"],
            marker_color="#10b981",
            text=[f"{p:.1f}%<br>{int(n)} OTs" for p, n in zip(_agg_a["pct_ok"], _agg_a["ok"])],
            textposition="inside", textfont=dict(color="white", size=11),
            hovertemplate="%{x}<br>En ventana: %{y:.1f}%<extra></extra>",
        ))
        _fig_antic.add_trace(_go.Bar(
            name="⚠️ Tarde", x=_agg_a["_intv_lbl"], y=_agg_a["pct_tarde"],
            marker_color="#f59e0b",
            text=[f"{p:.1f}%<br>{int(n)}" if p > 0 else "" for p, n in zip(_agg_a["pct_tarde"], _agg_a["tarde"])],
            textposition="inside", textfont=dict(color="white", size=11),
            hovertemplate="%{x}<br>Tarde: %{y:.1f}%<extra></extra>",
        ))
        _fig_antic.add_trace(_go.Bar(
            name="🔵 Muy anticipada (>50%)", x=_agg_a["_intv_lbl"], y=_agg_a["pct_antic"],
            marker_color="#3b82f6",
            text=[f"{p:.1f}%<br>{int(n)}" if p > 0 else "" for p, n in zip(_agg_a["pct_antic"], _agg_a["antic"])],
            textposition="inside", textfont=dict(color="white", size=11),
            hovertemplate="%{x}<br>Muy anticipada: %{y:.1f}%<extra></extra>",
        ))
        _fig_antic.update_layout(
            barmode="stack",
            title=dict(text="<b>Distribución por tipo de programación</b><br>"
                            "<span style='font-size:0.75rem;color:#94a3b8;font-weight:400'>"
                            "Del 100% de MPs de cada frecuencia, cuánto se hizo en ventana / tarde / muy anticipada</span>",
                       font=dict(size=13)),
            height=380,
            margin=dict(l=10, r=10, t=70, b=10),
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            legend=dict(orientation="h", y=1.08, x=0.5, xanchor="center"),
            yaxis=dict(title="% MPs", ticksuffix="%", range=[0, 100],
                       gridcolor="rgba(148,163,184,0.2)"),
            xaxis=dict(title=""),
            bargap=0.35,
        )
        st.plotly_chart(_fig_antic, use_container_width=True, key="prev_bar_antic_por_plan")

    # ── Desglose: tabla de OTs que NO cumplieron ─────────────────────────
    _df_incumpl = _df_cumpl[~_df_cumpl["cumple"]].copy() if _cump_tot > 0 else pd.DataFrame()
    with st.expander(
        f"🔍  Desglose del {_cump_pct}% — ver las {len(_df_incumpl):,} OTs que NO cumplieron",
        expanded=True,
    ):
        if _df_incumpl.empty:
            st.success("✅ Todas las OTs evaluadas cumplieron el plazo. Sin incumplimientos.")
        else:
            # Ordenar: primero las tarde (por atraso desc), luego las muy anticipadas
            _df_incumpl["_orden_prio"] = _df_incumpl["dias_atraso"].where(
                _df_incumpl["dias_atraso"] > 0, other=-_df_incumpl["dias_antelacion"].abs()
            )
            _df_incumpl = _df_incumpl.sort_values("_orden_prio", ascending=False)

            _det = pd.DataFrame({
                "OT":              _df_incumpl["id_ot"].values,
                "Estación":        _df_incumpl.get("estacion", _df_incumpl.get("ubicacion", "")).fillna("—").values,
                "Cód. EDS":        _df_incumpl.get("codigo_eds", _df_incumpl.get("clasificacion_2", "")).fillna("—").values,
                "Plan":            _df_incumpl.get("plan_tareas", pd.Series([""] * len(_df_incumpl))).fillna("—").values,
                "Intervalo (d)":   _df_incumpl["_intervalo_dias"].astype(int).values,
                "Equipo":          _df_incumpl["codigo_activo"].fillna("—").values,
                "F. Programada":   _df_incumpl["_fp_n"].dt.strftime("%d/%m/%Y").values,
                "F. Ejecución":    _df_incumpl["_ff_n"].dt.strftime("%d/%m/%Y").fillna("—").values,
                "Días atraso":     _df_incumpl["dias_atraso"].astype(int).values,
                "Días antes":      _df_incumpl["dias_antelacion"].fillna(0).astype(int).values,
                "% Antelación":    _df_incumpl["pct_antelacion"].fillna(0).values,
                "Motivo":          _df_incumpl["_motivo"].values,
                "¿Cumple semanal?": _df_incumpl["cumple_sem"].map({True: "✅ Sí", False: "❌ No"}).values,
                "Estado":          _df_incumpl["estado_tarea"].fillna("—").values,
                "Responsable":     _df_incumpl["responsable"].fillna("—").values,
            })

            _det = _det[["OT", "Estación", "Cód. EDS", "Plan", "Intervalo (d)", "Equipo",
                         "F. Programada", "F. Ejecución", "Días atraso", "Días antes",
                         "% Antelación", "Motivo", "¿Cumple semanal?", "Estado", "Responsable"]]

            _show_df(
                _det.reset_index(drop=True),
                hide_index=True,
                use_container_width=True,
                column_config={
                    "OT":               st.column_config.TextColumn(width=85),
                    "Estación":         st.column_config.TextColumn(width=180),
                    "Cód. EDS":         st.column_config.TextColumn(width=85),
                    "Plan":             st.column_config.TextColumn(width=180),
                    "Intervalo (d)":    st.column_config.NumberColumn(format="%d", width=85,
                        help="Intervalo del plan de mantención (días entre ejecuciones consecutivas)"),
                    "Equipo":           st.column_config.TextColumn(width=85),
                    "F. Programada":    st.column_config.TextColumn(width=95),
                    "F. Ejecución":     st.column_config.TextColumn(width=95),
                    "Días atraso":      st.column_config.NumberColumn(format="%d", width=80),
                    "Días antes":       st.column_config.NumberColumn(format="%d", width=80,
                        help="Días de anticipación respecto a la fecha programada (positivo = adelantada)"),
                    "% Antelación":     st.column_config.NumberColumn(format="%.1f%%", width=95,
                        help="Anticipación como % del intervalo del plan. >50% = anticipación excesiva."),
                    "Motivo":           st.column_config.TextColumn(width=300),
                    "¿Cumple semanal?": st.column_config.TextColumn(width=110),
                    "Estado":           st.column_config.TextColumn(width=100),
                    "Responsable":      st.column_config.TextColumn(width=150),
                },
            )
            # Métricas del desglose
            _sin_ejec   = (_det["F. Ejecución"] == "—").sum()
            _ejec_tarde = ((_det["F. Ejecución"] != "—") & (_det["Días atraso"] > 0)).sum()
            _ejec_antic = ((_det["F. Ejecución"] != "—") & (_det["% Antelación"] > 50)).sum()
            _ok_semana  = (_det["¿Cumple semanal?"] == "✅ Sí").sum()
            st.caption(
                f"Ordenadas por severidad · "
                f"{_sin_ejec:,} sin ejecutar · {_ejec_tarde:,} ejecutadas tarde · "
                f"**{_ejec_antic:,} muy anticipadas (>50% del intervalo)** · "
                f"de las tarde, {_ok_semana:,} rescatadas por el criterio semanal (≤5 días)."
            )

    st.divider()

    # ── Tabs ──────────────────────────────────────────────────────────────
    _ptab_plan, _ptab_cli, _ptab_tec, _ptab_eds, _ptab_uptime = st.tabs([
        "📅  Planificación", "🏢  Por Cliente", "👷  Por Técnico",
        "🏭  Por Activo/EDS", "⏱️  Uptime"
    ])

    # ── Tab 1: Por Cliente ────────────────────────────────────────────────
    with _ptab_cli:
        if dfp.empty:
            st.info("Sin registros para el filtro seleccionado.")
        else:
            st.caption(
                "Cumplimiento de mantenciones preventivas dividido por cliente. "
                "Se aplican los filtros de arriba (Mes, Trimestre, Plan, etc.). "
                "**Criterio Crudo**: fecha_finalización ≤ fecha_programada (mismo día o antes). "
                "Se excluyen OTs anuladas (Cancelado / Error de ingreso)."
            )

            # Preparar dataset evaluable por cliente
            _dfcli = dfp[
                dfp["fecha_programada"].notna()
                & ~dfp["estado"].isin(_ESTADOS_NO_CUENTAN)
            ].copy()
            _dfcli["_fp_n"] = pd.to_datetime(
                _dfcli["fecha_programada"],   errors="coerce", utc=True
            ).dt.tz_convert("America/Santiago").dt.tz_localize(None).dt.normalize()
            _dfcli["_ff_n"] = pd.to_datetime(
                _dfcli["fecha_finalizacion"], errors="coerce", utc=True
            ).dt.tz_convert("America/Santiago").dt.tz_localize(None).dt.normalize()
            _hoy_cli = pd.Timestamp.today().normalize()
            # Solo OTs cuya fecha programada ya pasó (universo evaluable)
            _dfcli_eval = _dfcli[_dfcli["_fp_n"] <= _hoy_cli].copy()
            _dfcli_eval["dias_atraso"] = (
                _dfcli_eval["_ff_n"].fillna(_hoy_cli) - _dfcli_eval["_fp_n"]
            ).dt.days
            _dfcli_eval["cumple"] = (
                _dfcli_eval["_ff_n"].notna() & (_dfcli_eval["dias_atraso"] <= 0)
            )

            _CLIENTES = ["COPEC", "Aramco (Esmax)", "SHELL (Enex)"]
            _ICON = {"COPEC": "🟢", "Aramco (Esmax)": "🟠", "SHELL (Enex)": "🟡"}

            # Resumen comparativo arriba
            _resumen = []
            for cli in _CLIENTES:
                _sub = _dfcli_eval[_dfcli_eval["cliente"] == cli]
                _tot = len(_sub)
                _cump = int(_sub["cumple"].sum()) if _tot > 0 else 0
                _no   = _tot - _cump
                _pct  = round(_cump / _tot * 100, 1) if _tot > 0 else 0.0
                _avg_atraso = round(_sub.loc[~_sub["cumple"], "dias_atraso"].mean(), 1) \
                    if (~_sub["cumple"]).any() else 0.0
                _resumen.append({
                    "Cliente": f"{_ICON.get(cli,'⚫')} {cli}",
                    "Total OTs evaluadas": _tot,
                    "✅ Cumplieron": _cump,
                    "❌ No cumplieron": _no,
                    "% Cumplimiento": f"{_pct}%",
                    "Atraso promedio (no cumplidas)": f"{_avg_atraso} días" if _no > 0 else "—",
                })
            _df_resumen = pd.DataFrame(_resumen)
            st.markdown(
                '<div class="section-header">📊  Resumen comparativo por cliente</div>',
                unsafe_allow_html=True,
            )
            _show_df(_df_resumen, hide_index=True, use_container_width=True,
                column_config={
                    "Cliente":                       st.column_config.TextColumn(width=180),
                    "Total OTs evaluadas":           st.column_config.NumberColumn(format="%d", width=160),
                    "✅ Cumplieron":                  st.column_config.NumberColumn(format="%d", width=130),
                    "❌ No cumplieron":               st.column_config.NumberColumn(format="%d", width=140),
                    "% Cumplimiento":                st.column_config.TextColumn(width=140),
                    "Atraso promedio (no cumplidas)":st.column_config.TextColumn(width=220),
                })
            st.divider()

            # Detalle por cliente
            for cli in _CLIENTES:
                _sub = _dfcli_eval[_dfcli_eval["cliente"] == cli]
                _tot = len(_sub)
                if _tot == 0:
                    st.markdown(f"### {_ICON.get(cli,'⚫')} {cli}")
                    st.info(f"Sin OTs evaluables para {cli} con los filtros actuales.")
                    st.divider()
                    continue

                _cump = int(_sub["cumple"].sum())
                _no   = _tot - _cump
                _pct  = round(_cump / _tot * 100, 1)
                _col_pct = "#10b981" if _pct >= 90 else "#f59e0b" if _pct >= 75 else "#ef4444"

                st.markdown(
                    f"""<div style="display:flex;align-items:center;gap:14px;margin-top:10px;">
                      <h3 style="margin:0;">{_ICON.get(cli,'⚫')} {cli}</h3>
                      <span style="background:{_col_pct};color:white;padding:4px 12px;
                                   border-radius:12px;font-size:0.9rem;font-weight:600;">
                        {_pct}% cumplimiento</span>
                      <span style="color:var(--text-color, #475569);font-size:0.88rem;">
                        {_cump:,} cumplieron · <b style="color:#ef4444;">{_no:,} no cumplieron</b>
                        de {_tot:,} evaluadas</span>
                    </div>""",
                    unsafe_allow_html=True,
                )

                # Dos columnas: cumplidas (resumen) y no cumplidas (tabla detalle)
                _c1, _c2 = st.columns([1, 2.5])
                with _c1:
                    st.markdown(
                        f"""<div style="padding:14px 18px;background:rgba(16,185,129,0.10);
                             border-radius:10px;border-left:3px solid #10b981;margin-top:8px;">
                          <div style="font-size:0.85rem;color:#94a3b8;letter-spacing:0.04em;
                                      font-weight:600;text-transform:uppercase;margin-bottom:6px;">
                            ✅ Cumplieron en plazo</div>
                          <div style="font-size:1.5rem;font-weight:700;color:#10b981;">
                            {_cump:,} OTs</div>
                          <div style="color:var(--text-color, #475569);font-size:0.82rem;margin-top:4px;">
                            {_pct}% del total evaluable
                          </div>
                        </div>
                        <div style="padding:14px 18px;background:rgba(239,68,68,0.10);
                             border-radius:10px;border-left:3px solid #ef4444;margin-top:10px;">
                          <div style="font-size:0.85rem;color:#94a3b8;letter-spacing:0.04em;
                                      font-weight:600;text-transform:uppercase;margin-bottom:6px;">
                            ❌ No cumplieron</div>
                          <div style="font-size:1.5rem;font-weight:700;color:#ef4444;">
                            {_no:,} OTs</div>
                          <div style="color:var(--text-color, #475569);font-size:0.82rem;margin-top:4px;">
                            atraso promedio: <b>{round(_sub.loc[~_sub['cumple'], 'dias_atraso'].mean(), 1) if _no > 0 else 0} días</b>
                          </div>
                        </div>""",
                        unsafe_allow_html=True,
                    )
                with _c2:
                    _no_sub = _sub[~_sub["cumple"]].copy().sort_values("dias_atraso", ascending=False).head(15)
                    if _no_sub.empty:
                        st.success("Todas las OTs de este cliente cumplieron el plazo.")
                    else:
                        _det_cli = pd.DataFrame({
                            "OT":            _no_sub["id_ot"].values,
                            "Estación":      _no_sub.get("estacion", _no_sub.get("ubicacion","")).fillna("—").values,
                            "Cód. EDS":      _no_sub.get("codigo_eds", _no_sub.get("clasificacion_2","")).fillna("—").values,
                            "Plan":          _no_sub.get("plan_tareas", pd.Series([""]*len(_no_sub))).fillna("—").values,
                            "F. Prog.":      _no_sub["_fp_n"].dt.strftime("%d/%m/%Y").values,
                            "F. Ejec.":      _no_sub["_ff_n"].dt.strftime("%d/%m/%Y").fillna("—").values,
                            "Días atraso":   _no_sub["dias_atraso"].astype(int).values,
                        })
                        _show_df(_det_cli.reset_index(drop=True), hide_index=True,
                            use_container_width=True,
                            column_config={
                                "OT":           st.column_config.TextColumn(width=85),
                                "Estación":     st.column_config.TextColumn(width=180),
                                "Cód. EDS":     st.column_config.TextColumn(width=85),
                                "Plan":         st.column_config.TextColumn(width=180),
                                "F. Prog.":     st.column_config.TextColumn(width=90),
                                "F. Ejec.":     st.column_config.TextColumn(width=90),
                                "Días atraso":  st.column_config.NumberColumn(format="%d d", width=95),
                            })
                        st.caption(
                            f"Top 15 OTs con mayor atraso · "
                            f"{_no - 15 if _no > 15 else 0} adicionales no mostradas"
                            if _no > 15 else
                            f"Mostrando las {_no} OTs que no cumplieron"
                        )
                st.divider()

    # ── Tab: Por Técnico ──────────────────────────────────────────────────
    with _ptab_tec:
        if dfp.empty:
            st.info("Sin datos para mostrar.")
        else:
            st.markdown('<div class="section-header">👷  Desempeño por responsable</div>',
                        unsafe_allow_html=True)
            _pt_resp = dfp.copy()
            # duracion_estim_seg / duracion_real_seg están en segundos → convertir a minutos
            _pt_resp["_dur_est_min"]  = pd.to_numeric(_pt_resp["duracion_estim_seg"],  errors="coerce") / 60 \
                if "duracion_estim_seg"  in _pt_resp.columns else None
            _pt_resp["_dur_real_min"] = pd.to_numeric(_pt_resp["duracion_real_seg"], errors="coerce") / 60 \
                if "duracion_real_seg" in _pt_resp.columns else None
            _pt_resp_g = _pt_resp.groupby("responsable").agg(
                Total      =("id_ot","count"),
                Finaliz    =("estado_tarea", lambda s: (s == "Finalizada").sum()),
                EnProceso  =("estado_tarea", lambda s: s.isin(["En Proceso","En Revisión"]).sum()),
                AvgEst_min =("_dur_est_min",  lambda s: round(s.dropna().mean(), 1) if len(s.dropna())>0 else None),
                AvgReal_min=("_dur_real_min", lambda s: round(s.dropna().mean(), 1) if len(s.dropna())>0 else None),
            ).reset_index().sort_values("Total", ascending=False)
            def _min_to_hhmm(mins):
                if mins is None or (isinstance(mins, float) and pd.isna(mins)):
                    return "—"
                try:
                    m = int(round(float(mins)))
                    return f"{m//60:02d}:{m%60:02d}"
                except Exception:
                    return "—"
            def _calc_efic(row):
                est  = row["AvgEst_min"]
                real = row["AvgReal_min"]
                if est and real and not pd.isna(est) and not pd.isna(real) and float(real) > 0:
                    return f"{round(float(est)/float(real)*100, 1)}%"
                return "—"
            _pt_resp_g["% Comp."]    = _pt_resp_g.apply(
                lambda r: f"{round(r['Finaliz']/r['Total']*100,1)}%" if r["Total"]>0 else "—", axis=1)
            _pt_resp_g["T.Est."]     = _pt_resp_g["AvgEst_min"].apply(_min_to_hhmm)
            _pt_resp_g["T.Real."]    = _pt_resp_g["AvgReal_min"].apply(_min_to_hhmm)
            _pt_resp_g["Eficiencia"] = _pt_resp_g.apply(_calc_efic, axis=1)
            _pt_resp_g = _pt_resp_g[["responsable","Total","Finaliz","EnProceso",
                                     "% Comp.","T.Est.","T.Real.","Eficiencia"]]
            _pt_resp_g.rename(columns={
                "responsable":"Responsable","Finaliz":"Finalizadas",
                "EnProceso":"En Proceso"}, inplace=True)
            _show_df(_pt_resp_g.reset_index(drop=True), hide_index=True, use_container_width=True,
                column_config={
                    "Responsable":  st.column_config.TextColumn(width=200),
                    "Total":        st.column_config.NumberColumn(format="%d", width=60),
                    "Finalizadas":  st.column_config.NumberColumn(format="%d", width=90),
                    "En Proceso":   st.column_config.NumberColumn(format="%d", width=90),
                    "% Comp.":      st.column_config.TextColumn(width=75),
                    "T.Est.":       st.column_config.TextColumn(width=70),
                    "T.Real.":      st.column_config.TextColumn(width=70),
                    "Eficiencia":   st.column_config.TextColumn(width=90),
                })
            st.caption("Eficiencia = T.Estimado / T.Real × 100%  (>100% = más rápido que lo estimado)")
            st.divider()

            # ── OTs activas por técnico ────────────────────────────────────
            st.markdown('<div class="section-header">⚙️  OTs en curso por responsable</div>',
                        unsafe_allow_html=True)
            _pt_activas = dfp[dfp["estado_tarea"].isin(["En Proceso","En Revisión","En Espera"])].copy()
            if _pt_activas.empty:
                st.info("No hay OTs activas en el período seleccionado.")
            else:
                _pt_activas_show = _pt_activas[[c for c in [
                    "id_ot","responsable","nombre_tarea","tipo_tarea",
                    "estado_tarea","fecha_programada","activador"
                ] if c in _pt_activas.columns]].copy()
                _pt_activas_show["fecha_programada"] = pd.to_datetime(
                    _pt_activas_show["fecha_programada"], errors="coerce").dt.strftime("%d/%m/%Y")
                _pt_activas_show.rename(columns={
                    "id_ot":"OT","responsable":"Responsable","nombre_tarea":"Tarea",
                    "tipo_tarea":"Tipo","estado_tarea":"Estado","fecha_programada":"F.Prog.",
                    "activador":"Activador"
                }, inplace=True)
                _show_df(_pt_activas_show.reset_index(drop=True), hide_index=True, use_container_width=True,
                    column_config={
                        "OT":          st.column_config.TextColumn(width=90),
                        "Responsable": st.column_config.TextColumn(width=180),
                        "Tarea":       st.column_config.TextColumn(width=200),
                        "Tipo":        st.column_config.TextColumn(width=170),
                        "Estado":      st.column_config.TextColumn(width=110),
                        "F.Prog.":     st.column_config.TextColumn(width=90),
                        "Activador":   st.column_config.TextColumn(width=100),
                    })
                st.caption(f"{len(_pt_activas_show):,} OTs activas")

    # ── Tab 4: Planificación ──────────────────────────────────────────────
    with _ptab_plan:
        # Parte de _dfp_full: respeta responsable/tipo/estado pero no trimestre/mes
        # IMPORTANTE: Supabase guarda fechas en UTC. Hay que convertirlas a hora
        # local de Chile (America/Santiago) ANTES de quitarles el tz, sino las
        # OTs nocturnas aparecen un día corrido.
        _dfplan = _dfp_full.copy()
        _dfplan["_fp_dt"] = pd.to_datetime(
            _dfplan["fecha_programada"], errors="coerce", utc=True
        ).dt.tz_convert("America/Santiago").dt.tz_localize(None)
        _dfplan["_ff_dt"] = pd.to_datetime(
            _dfplan["fecha_finalizacion"], errors="coerce", utc=True
        ).dt.tz_convert("America/Santiago").dt.tz_localize(None)
        _dfplan = _dfplan[_dfplan["_fp_dt"].notna()].copy()

        _hoy = pd.Timestamp.today().normalize()
        _lun_actual = _hoy - pd.Timedelta(days=_hoy.weekday())
        _dom_actual = _lun_actual + pd.Timedelta(days=6)
        _lun_prox   = _lun_actual + pd.Timedelta(weeks=1)
        _dom_prox   = _lun_prox   + pd.Timedelta(days=6)
        _semana_iso = _hoy.isocalendar().week

        # Encabezado con número de semana ISO
        st.markdown(
            f"""<div style="background:rgba(1,121,138,0.10);border-left:3px solid #01798A;
                 padding:10px 16px;border-radius:6px;margin-bottom:14px;">
              <span style="color:#01798A;font-weight:700;font-size:0.95rem;letter-spacing:0.04em;">
                📆  Semana ISO {_semana_iso}</span>
              <span style="color:#94a3b8;font-size:0.85rem;margin-left:12px;">
                Hoy: <b>{_hoy.strftime('%d/%m/%Y')}</b> · Rango actual:
                <b>{_lun_actual.strftime('%d/%m')} – {_dom_actual.strftime('%d/%m/%Y')}</b></span>
            </div>""",
            unsafe_allow_html=True,
        )

        _plan_sem_opts = {
            f"Semana actual (S{_semana_iso}): {_lun_actual.strftime('%d/%m')} – {_dom_actual.strftime('%d/%m/%Y')}":
                (_lun_actual, _dom_actual + pd.Timedelta(hours=23, minutes=59, seconds=59)),
            f"Próxima semana (S{_semana_iso+1}): {_lun_prox.strftime('%d/%m')} – {_dom_prox.strftime('%d/%m/%Y')}":
                (_lun_prox, _dom_prox + pd.Timedelta(hours=23, minutes=59, seconds=59)),
            "Últimas 2 semanas":
                (_lun_actual - pd.Timedelta(weeks=2), _hoy + pd.Timedelta(hours=23, minutes=59, seconds=59)),
            "Próximas 4 semanas":
                (_hoy, _hoy + pd.Timedelta(weeks=4) + pd.Timedelta(hours=23, minutes=59, seconds=59)),
        }
        _plan_sel = st.radio("Período", list(_plan_sem_opts.keys()),
                             horizontal=True, key="prev_plan_sem")
        _plan_ini, _plan_fin = _plan_sem_opts[_plan_sel]

        _df_semana = _dfplan[
            (_dfplan["_fp_dt"] >= _plan_ini) & (_dfplan["_fp_dt"] <= _plan_fin)
        ].copy()

        # ── Relojes: Tiempo de atención MP — dos métricas distintas ───────
        # 1) PROGRAMADA → EJECUCIÓN     (cumplimiento del plazo comprometido)
        # 2) DURACIÓN REAL DEL TRABAJO  (fecha_inicio → fecha_finalización =
        #    cuánto tarda el técnico ejecutando la OT desde que la abre)
        # Ambos calculados sobre OTs FINALIZADAS del filtro global, en horas.
        _dfp_for_avg = dfp[
            dfp["fecha_finalizacion"].notna()
            & dfp["fecha_programada"].notna()
            & ~dfp["estado"].isin(_ESTADOS_NO_CUENTAN)
        ].copy()

        def _fmt_h_d(hrs: float) -> str:
            """Formato 'X días Y horas' o 'Z horas' si < 24h."""
            neg = hrs < 0
            h_abs = abs(hrs)
            if h_abs < 24:
                txt = f"{h_abs:.1f} h"
            else:
                d = int(h_abs // 24)
                h = round(h_abs - d * 24, 1)
                txt = f"{d} día{'s' if d != 1 else ''} {h:.1f} h" if h > 0 else f"{d} día{'s' if d != 1 else ''}"
            return f"-{txt}" if neg else txt

        if not _dfp_for_avg.empty:
            # Sin normalize → precisión sub-día
            _fp_avg = pd.to_datetime(_dfp_for_avg["fecha_programada"], errors="coerce", utc=True).dt.tz_convert("America/Santiago").dt.tz_localize(None)
            _ff_avg = pd.to_datetime(_dfp_for_avg["fecha_finalizacion"], errors="coerce", utc=True).dt.tz_convert("America/Santiago").dt.tz_localize(None)
            _fi_avg = pd.to_datetime(_dfp_for_avg["fecha_inicio"],       errors="coerce", utc=True).dt.tz_convert("America/Santiago").dt.tz_localize(None)

            # Métrica 1: programada → ejecución
            _h_prog        = (_ff_avg - _fp_avg).dt.total_seconds() / 3600
            _avg_h_prog    = round(_h_prog.mean(), 1)
            _med_h_prog    = round(_h_prog.median(), 1)
            _pct_a_tiempo  = round((_h_prog <= 0).mean() * 100, 1)

            # Métrica 2: duración real del trabajo (inicio → finalización)
            # Solo cuenta cuando hay fecha_inicio (técnico abrió la OT)
            _h_dur        = (_ff_avg - _fi_avg).dt.total_seconds() / 3600
            _h_dur_valid  = _h_dur.where((_h_dur >= 0) & _fi_avg.notna())
            _avg_h_dur    = round(_h_dur_valid.mean(), 1)  if _h_dur_valid.notna().any() else 0.0
            _med_h_dur    = round(_h_dur_valid.median(), 1) if _h_dur_valid.notna().any() else 0.0
            _ot_dur_n     = int(_h_dur_valid.notna().sum())
            _ot_eval      = len(_dfp_for_avg)
        else:
            _avg_h_prog = _med_h_prog = _avg_h_dur = _med_h_dur = 0.0
            _pct_a_tiempo = 0
            _ot_eval = _ot_dur_n = 0

        _color_prog = "#10b981" if _avg_h_prog <= 0 else "#f59e0b" if _avg_h_prog <= 72 else "#ef4444"
        # Duración real: <1h excelente, <4h bueno, >4h investigar (puede ser
        # OT que el técnico dejó abierta sin cerrarla bien)
        _color_dur = "#10b981" if _avg_h_dur <= 1 else "#f59e0b" if _avg_h_dur <= 4 else "#ef4444"

        import plotly.graph_objects as _go2

        def _build_clock(value, color, title, subtitle, axis_range, tickvals, steps):
            fig = _go2.Figure(_go2.Indicator(
                mode="gauge+number",
                value=value,
                number={"suffix": " h", "font": {"size": 28}, "valueformat": ".1f"},
                title={"text": f"<b>{title}</b><br>"
                               f"<span style='font-size:0.72rem;color:#94a3b8'>{subtitle}</span>",
                       "font": {"size": 13}},
                gauge={
                    "axis":    {"range": axis_range, "tickwidth": 1, "tickfont": {"size": 9},
                                "tickvals": tickvals},
                    "bar":     {"color": color, "thickness": 0.32},
                    "bgcolor": "rgba(0,0,0,0)",
                    "borderwidth": 0,
                    "steps": steps,
                },
            ))
            fig.update_layout(height=250, margin=dict(l=10, r=10, t=60, b=10),
                              paper_bgcolor="rgba(0,0,0,0)")
            return fig

        _rc1, _rc2 = st.columns(2)
        with _rc1:
            # Eje del gauge programada→ejecución: -120 a +360h (-5d a +15d)
            _fig_prog = _build_clock(
                _avg_h_prog, _color_prog,
                "⏱ Programada → Ejecución",
                "horas entre fecha programada y ejecución",
                [-120, 360], [-120, -48, 0, 24, 72, 168, 360],
                [
                    {"range": [-120, 0], "color": "rgba(16,185,129,0.18)"},
                    {"range": [0,    72], "color": "rgba(245,158,11,0.18)"},
                    {"range": [72,  360], "color": "rgba(239,68,68,0.18)"},
                ],
            )
            # Línea de "fecha programada" (objetivo = 0)
            _fig_prog.update_traces(gauge_threshold={"line": {"color": "#0f172a", "width": 2},
                                                     "thickness": 0.85, "value": 0})
            st.plotly_chart(_fig_prog, use_container_width=True, key="prev_reloj_prog")
        with _rc2:
            # Duración real del trabajo: escala 0 a 8h (rango operativo razonable)
            _fig_dur = _build_clock(
                _avg_h_dur, _color_dur,
                "⏱ Duración real del trabajo",
                "horas entre inicio del técnico y cierre de la OT",
                [0, 8], [0, 0.5, 1, 2, 4, 6, 8],
                [
                    {"range": [0, 1], "color": "rgba(16,185,129,0.18)"},
                    {"range": [1, 4], "color": "rgba(245,158,11,0.18)"},
                    {"range": [4, 8], "color": "rgba(239,68,68,0.18)"},
                ],
            )
            st.plotly_chart(_fig_dur, use_container_width=True, key="prev_reloj_dur")

        # Panel de detalle unificado debajo
        st.markdown(
            f"""<div style="padding:14px 18px;background:rgba(148,163,184,0.08);
                 border-radius:10px;border-left:3px solid {_color_prog};margin-top:-8px;">
              <div style="font-size:0.85rem;color:#94a3b8;letter-spacing:0.04em;
                          font-weight:600;text-transform:uppercase;margin-bottom:8px;">
                Detalle del tiempo de atención
              </div>
              <table style="width:100%;border-collapse:collapse;font-size:0.92rem;">
                <tr style="border-bottom:1px solid rgba(148,163,184,0.25);">
                  <td style="padding:6px 0;font-weight:600;color:var(--text-color, #0f172a);">Programada → Ejecución</td>
                  <td style="padding:6px 0;text-align:right;color:{_color_prog};font-weight:600;font-size:1.05rem;">
                    {_fmt_h_d(_avg_h_prog)}
                  </td>
                  <td style="padding:6px 16px;color:var(--text-color, #475569);font-size:0.82rem;text-align:right;">
                    mediana {_fmt_h_d(_med_h_prog)}
                  </td>
                </tr>
                <tr style="border-bottom:1px solid rgba(148,163,184,0.25);">
                  <td style="padding:6px 0;font-weight:600;color:var(--text-color, #0f172a);">Duración real del trabajo</td>
                  <td style="padding:6px 0;text-align:right;color:{_color_dur};font-weight:600;font-size:1.05rem;">
                    {_fmt_h_d(_avg_h_dur)}
                  </td>
                  <td style="padding:6px 16px;color:var(--text-color, #475569);font-size:0.82rem;text-align:right;">
                    mediana {_fmt_h_d(_med_h_dur)} · {_ot_dur_n:,} OTs con inicio
                  </td>
                </tr>
                <tr style="border-bottom:1px solid rgba(148,163,184,0.25);">
                  <td style="padding:6px 0;font-weight:600;color:var(--text-color, #0f172a);">OTs evaluadas</td>
                  <td style="padding:6px 0;text-align:right;color:var(--text-color, #0f172a);font-weight:600;">{_ot_eval:,}</td>
                  <td></td>
                </tr>
                <tr>
                  <td style="padding:6px 0;font-weight:600;color:var(--text-color, #0f172a);">Atendidas a tiempo (criterio crudo)</td>
                  <td style="padding:6px 0;text-align:right;color:#10b981;font-weight:600;">{_pct_a_tiempo}%</td>
                  <td></td>
                </tr>
              </table>
              <div style="color:var(--text-color, #475569);font-size:0.78rem;margin-top:10px;
                          padding-top:8px;border-top:1px solid rgba(148,163,184,0.25);">
                Métrica 1 (<b>Programada → Ejecución</b>): cumplimiento del plazo
                comprometido. Valores negativos = ejecutada antes de lo programado.<br>
                Métrica 2 (<b>Duración real del trabajo</b>): tiempo entre que el
                técnico abre la OT en Fracttal (fecha_inicio) y la cierra
                (fecha_finalización). Indica cuánto tarda haciendo el trabajo real.
              </div>
            </div>""",
            unsafe_allow_html=True,
        )

        st.divider()

        # Tipo de equipo a partir del nombre (lavadora / aspiradora / etc.)
        def _tipo_eq(s):
            s = str(s or "").upper()
            if "LAVADORA" in s:        return "🚿 Lavadora"
            if "ASPIRA"   in s:        return "🌀 Aspiradora"
            if "LAVAINTER" in s or "INTERIOR" in s: return "🧽 Lava-interior"
            if "LAVABICI" in s:        return "🚲 Lava-bicicletas"
            if "TERMO"    in s:        return "♨️ Termo"
            if "BOMBA"    in s or "BBA" in s: return "🔧 Bomba"
            if "ABLAND"   in s:        return "💧 Ablandador"
            if "HIDROPACK" in s:       return "⚙️ Hidropack"
            if "TWISTER"  in s:        return "🌪️ Twister"
            if "SECADO"   in s or "DRY" in s: return "🌬️ Secador"
            return "🔩 Otro"

        # Plazo restante (en días) hasta la fecha programada
        _df_semana["_fp_norm"] = _df_semana["_fp_dt"].dt.normalize()
        _df_semana["_ff_norm"] = _df_semana["_ff_dt"].dt.normalize()
        _df_semana["dias_restantes"] = (_df_semana["_fp_norm"] - _hoy).dt.days
        # Atraso real cuando hay finalización: días entre ejecución y programada
        _df_semana["dias_ejec_vs_prog"] = (
            _df_semana["_ff_norm"] - _df_semana["_fp_norm"]
        ).dt.days
        # ¿Está finalizada? (mismo criterio que el resto de KPIs)
        _df_semana["_esta_fin"] = (
            _df_semana["estado_tarea"].isin(["Finalizada"])
            | _df_semana["estado"].isin(["Finalizadas"])
        )

        st.markdown(
            f'<div class="section-header">📅  OTs programadas en el período '
            f'<span style="color:#94a3b8;font-weight:500;font-size:0.85rem;">'
            f'· {len(_df_semana)} encontradas</span></div>',
            unsafe_allow_html=True,
        )

        if _df_semana.empty:
            st.warning(
                "No hay OTs preventivas programadas para este período en Supabase. "
                "Esto puede indicar que: (a) el sync aún no trajo OTs futuras desde Fracttal, "
                "o (b) las OTs de este período aún no se han generado en Fracttal."
            )
        else:
            _df_semana_sorted = _df_semana.sort_values("_fp_dt")

            def _plazo_label(row):
                # Si está finalizada → comparar fecha_ejecución vs programada
                if row["_esta_fin"] and pd.notna(row["_ff_norm"]):
                    d = int(row["dias_ejec_vs_prog"])
                    if d < 0:   return f"✅ Cumplida {-d}d antes"
                    if d == 0:  return "✅ Cumplida el día"
                    return f"⚠️ Cumplida {d}d tarde"
                # No finalizada → plazo hasta fecha programada
                d = int(row["dias_restantes"])
                if d > 0:   return f"⏳ en {d}d"
                if d == 0:  return "🟢 hoy"
                return f"⚠️ Vencida hace {-d}d"

            _df_show = pd.DataFrame({
                "F. Programada": _df_semana_sorted["_fp_dt"].dt.strftime("%d/%m/%Y").values,
                "F. Ejecución":  _df_semana_sorted["_ff_dt"].dt.strftime("%d/%m/%Y").fillna("—").values,
                "Plazo":         _df_semana_sorted.apply(_plazo_label, axis=1).values,
                "OT":            _df_semana_sorted["id_ot"].values,
                "Código":        _df_semana_sorted["codigo_activo"].fillna("—").values,
                "Tipo equipo":   _df_semana_sorted["nombre_activo"].apply(_tipo_eq).values,
                "Activo":        _df_semana_sorted["nombre_activo"].fillna("—").values,
                "Estación":      _df_semana_sorted.get("estacion", _df_semana_sorted.get("ubicacion","")).fillna("—").values,
                "Cód. EDS":      _df_semana_sorted.get("codigo_eds", _df_semana_sorted.get("clasificacion_2","")).fillna("—").values,
                "Plan":          _df_semana_sorted.get("plan_tareas", pd.Series([""]*len(_df_semana_sorted))).fillna("—").values,
                "Responsable":   _df_semana_sorted["responsable"].fillna("—").values,
                "Estado":        _df_semana_sorted["estado"].fillna("—").values,
                "Estado tarea":  _df_semana_sorted["estado_tarea"].fillna("—").values,
            })
            _show_df(_df_show.reset_index(drop=True), hide_index=True, use_container_width=True,
                column_config={
                    "F. Programada": st.column_config.TextColumn(width=100),
                    "F. Ejecución":  st.column_config.TextColumn(width=100),
                    "Plazo":         st.column_config.TextColumn(width=160),
                    "OT":            st.column_config.TextColumn(width=85),
                    "Código":        st.column_config.TextColumn(width=85),
                    "Tipo equipo":   st.column_config.TextColumn(width=130),
                    "Activo":        st.column_config.TextColumn(width=240),
                    "Estación":      st.column_config.TextColumn(width=200),
                    "Cód. EDS":      st.column_config.TextColumn(width=90),
                    "Plan":          st.column_config.TextColumn(width=200),
                    "Responsable":   st.column_config.TextColumn(width=160),
                    "Estado":        st.column_config.TextColumn(width=100),
                    "Estado tarea":  st.column_config.TextColumn(width=110),
                })
            _sfin_m  = _df_semana["_esta_fin"]
            _snoi_m  = _df_semana["estado_tarea"].isin(["No Iniciada"]) & ~_sfin_m
            _sfin    = int(_sfin_m.sum())
            _sproc   = int((~_sfin_m & ~_snoi_m & ~_df_semana["estado"].isin(["Cancelado"])).sum())
            _snoi    = int(_snoi_m.sum())
            # Solo cuenta como vencida si NO está finalizada y ya pasó la fecha
            _venc    = int(((_df_semana["dias_restantes"] < 0) & ~_sfin_m).sum())
            st.caption(f"{len(_df_show):,} OTs · {_sfin} finalizadas · {_sproc} en proceso · "
                       f"{_snoi} no iniciadas · {_venc} pendientes vencidas")
        st.divider()

        # ── OTs vencidas (programadas en pasado y no finalizadas) ─────────
        st.markdown('<div class="section-header">⚠️  OTs vencidas (no finalizadas)</div>',
                    unsafe_allow_html=True)
        # Excluir OTs anuladas — el mismo criterio que el velocímetro:
        # Cancelado / ERROR DE INGRESO / EQUIPO CON RECAMBIO no representan
        # mantenimientos pendientes, son OTs que se anularon por error.
        _df_venc = _dfplan[
            (_dfplan["_fp_dt"] < _hoy) &
            (~_dfplan["estado"].isin(_ESTADOS_NO_CUENTAN | {"Finalizadas"})) &
            (~_dfplan["estado_tarea"].isin(["Finalizada"]))   # excluir DONE normalizados
        ].copy()
        _df_venc["_atraso"] = (_hoy - _df_venc["_fp_dt"]).dt.days
        _df_venc = _df_venc.sort_values("_atraso", ascending=False)
        if _df_venc.empty:
            st.success("✅ No hay OTs preventivas vencidas.")
        else:
            _df_venc_show = _df_venc[[c for c in [
                "fecha_programada","_atraso","id_ot","responsable",
                "nombre_tarea","tipo_tarea","activador","estado","estado_tarea"
            ] if c in _df_venc.columns]].copy()
            _df_venc_show["fecha_programada"] = pd.to_datetime(
                _df_venc_show["fecha_programada"], errors="coerce").dt.strftime("%d/%m/%Y")
            _df_venc_show.rename(columns={
                "fecha_programada":"F.Prog.","_atraso":"Atraso (días)","id_ot":"OT",
                "responsable":"Responsable","nombre_tarea":"Tarea","tipo_tarea":"Tipo",
                "activador":"Activador","estado":"Estado","estado_tarea":"Estado Tarea"
            }, inplace=True)
            _show_df(_df_venc_show.reset_index(drop=True), hide_index=True, use_container_width=True,
                column_config={
                    "F.Prog.":       st.column_config.TextColumn(width=90),
                    "Atraso (días)": st.column_config.NumberColumn(format="%d días", width=105),
                    "OT":            st.column_config.TextColumn(width=90),
                    "Responsable":   st.column_config.TextColumn(width=180),
                    "Tarea":         st.column_config.TextColumn(width=200),
                    "Tipo":          st.column_config.TextColumn(width=170),
                    "Activador":     st.column_config.TextColumn(width=100),
                    "Estado":        st.column_config.TextColumn(width=100),
                    "Estado Tarea":  st.column_config.TextColumn(width=110),
                })
            st.caption(f"⚠️ {len(_df_venc_show):,} OTs preventivas vencidas sin finalizar")
        st.divider()

        # ── Próximas OTs futuras ───────────────────────────────────────────
        st.markdown('<div class="section-header">🔮  Próximas OTs programadas (hoy en adelante)</div>',
                    unsafe_allow_html=True)
        st.caption(
            "Mantenciones preventivas **pendientes** con fecha programada de hoy en "
            "adelante. Se excluyen OTs anuladas (Cancelado / Error de ingreso) y "
            "las que ya se ejecutaron anticipadamente — éstas se ven en el cuadro "
            "principal de la semana."
        )
        _df_fut = _dfplan[
            (_dfplan["_fp_dt"].dt.normalize() >= _hoy)
            & (~_dfplan["estado"].isin(_ESTADOS_NO_CUENTAN | {"Finalizadas"}))
            & (~_dfplan["estado_tarea"].isin(["Finalizada"]))
        ].copy().sort_values("_fp_dt")
        if _df_fut.empty:
            st.info("No hay OTs preventivas pendientes con fecha futura.")
        else:
            _df_fut["_dias_rest"] = (_df_fut["_fp_dt"].dt.normalize() - _hoy).dt.days
            _df_fut["_barra"] = _df_fut["_dias_rest"].clip(lower=0, upper=30)

            _fut_show = pd.DataFrame({
                "F. Programada": _df_fut["_fp_dt"].dt.strftime("%d/%m/%Y").values,
                "Plazo":         _df_fut["_dias_rest"].apply(
                    lambda d: "🟢 hoy" if d == 0
                    else (f"⏳ en {d} día{'s' if d != 1 else ''}")
                ).values,
                "⏱ Barra (días)": _df_fut["_barra"].values,
                "OT":            _df_fut["id_ot"].values,
                "Código":        _df_fut["codigo_activo"].fillna("—").values,
                "Tipo equipo":   _df_fut["nombre_activo"].apply(_tipo_eq).values,
                "Activo":        _df_fut["nombre_activo"].fillna("—").values,
                "Estación":      _df_fut.get("estacion", _df_fut.get("ubicacion","")).fillna("—").values,
                "Cód. EDS":      _df_fut.get("codigo_eds", _df_fut.get("clasificacion_2","")).fillna("—").values,
                "Plan":          _df_fut.get("plan_tareas", pd.Series([""]*len(_df_fut))).fillna("—").values,
                "Activador":     _df_fut["activador"].fillna("—").values,
                "Responsable":   _df_fut["responsable"].fillna("—").values,
                "Estado":        _df_fut["estado"].fillna("—").values,
            })
            _show_df(_fut_show.reset_index(drop=True), hide_index=True, use_container_width=True,
                column_config={
                    "F. Programada":  st.column_config.TextColumn(width=100),
                    "Plazo":          st.column_config.TextColumn(width=110),
                    "⏱ Barra (días)": st.column_config.ProgressColumn(
                        help="Días restantes hasta la fecha programada (0–30+)",
                        format="%d d", min_value=0, max_value=30, width=140,
                    ),
                    "OT":             st.column_config.TextColumn(width=85),
                    "Código":         st.column_config.TextColumn(width=85),
                    "Tipo equipo":    st.column_config.TextColumn(width=130),
                    "Activo":         st.column_config.TextColumn(width=240),
                    "Estación":       st.column_config.TextColumn(width=200),
                    "Cód. EDS":       st.column_config.TextColumn(width=90),
                    "Plan":           st.column_config.TextColumn(width=200),
                    "Activador":      st.column_config.TextColumn(width=100),
                    "Responsable":    st.column_config.TextColumn(width=160),
                    "Estado":         st.column_config.TextColumn(width=110),
                })
            _hoy_n = int((_df_fut["_dias_rest"] == 0).sum())
            _3d    = int(((_df_fut["_dias_rest"] > 0) & (_df_fut["_dias_rest"] <= 3)).sum())
            _7d    = int(((_df_fut["_dias_rest"] > 3) & (_df_fut["_dias_rest"] <= 7)).sum())
            _mas   = int((_df_fut["_dias_rest"] > 7).sum())
            st.caption(
                f"**{len(_fut_show):,} OTs pendientes** · "
                f"🟢 {_hoy_n} para hoy · ⏳ {_3d} en 1–3 días · {_7d} en 4–7 días · "
                f"{_mas} más allá de 7 días."
            )

    # ── Tab 5: Por Activo/EDS ─────────────────────────────────────────────
    with _ptab_eds:
        if _dfp_full.empty:
            st.info("Sin datos de mantenciones preventivas.")
        else:
            _dfeds = _dfp_full.copy()
            # Fechas en hora Chile (sino las nocturnas saltan un día)
            _dfeds["_fp_dt"] = pd.to_datetime(
                _dfeds["fecha_programada"],   errors="coerce", utc=True
            ).dt.tz_convert("America/Santiago").dt.tz_localize(None)
            _dfeds["_ff_dt"] = pd.to_datetime(
                _dfeds["fecha_finalizacion"], errors="coerce", utc=True
            ).dt.tz_convert("America/Santiago").dt.tz_localize(None)
            _hoy_eds = pd.Timestamp.today().normalize()

            # Excluir las OTs anuladas — no representan trabajos reales
            _dfeds = _dfeds[~_dfeds["estado"].isin(_ESTADOS_NO_CUENTAN)].copy()

            # Última PM finalizada por activo (fecha_finalizacion más reciente)
            _df_ult = (
                _dfeds[_dfeds["estado"] == "Finalizadas"]
                .sort_values("_ff_dt", ascending=False)
                .groupby(["codigo_activo","nombre_activo"], as_index=False)
                .first()
                [["codigo_activo","nombre_activo","ubicacion","clasificacion_2",
                  "codigo_eds","estacion","activador","_ff_dt","plan_tareas"]]
                .rename(columns={"_ff_dt":"_ultima_pm"})
            )

            # Próxima PM (OT real creada en Fracttal con fecha futura)
            _df_prx = (
                _dfeds[
                    (_dfeds["_fp_dt"] >= _hoy_eds) &
                    (~_dfeds["estado"].isin(["Finalizadas"]))
                ]
                .sort_values("_fp_dt", ascending=True)
                .groupby(["codigo_activo","nombre_activo"], as_index=False)
                .first()
                [["codigo_activo","nombre_activo","_fp_dt","responsable","tipo_tarea"]]
                .rename(columns={"_fp_dt":"_prox_pm","responsable":"resp_prox",
                                 "tipo_tarea":"tipo_prox"})
            )

            # Frecuencia histórica total por activo
            _df_freq_eds = (
                _dfeds.groupby(["codigo_activo","nombre_activo"], as_index=False)
                .agg(Total_PM=("id_ot","count"))
            )

            # Merge
            _df_eds = _df_ult.merge(_df_prx,      on=["codigo_activo","nombre_activo"], how="outer")
            _df_eds = _df_eds.merge(_df_freq_eds, on=["codigo_activo","nombre_activo"], how="left")

            # ── Cálculo: próxima MP teórica (cuando Fracttal aún no creó la OT) ──
            # Parseamos el activador a días: "Cada 1 mes" → 30, "Cada 6 meses" → 180, etc.
            import re as _re
            def _activador_a_dias(s):
                if pd.isna(s) or not s:
                    return None
                txt = str(s).lower()
                m = _re.search(r"(\d+)\s*(d[ií]a|semana|mes|a[ñn]o|hora)", txt)
                if not m:
                    return None
                n = int(m.group(1))
                unit = m.group(2)
                if   "hora"   in unit: return None  # depende de uso, no se puede proyectar
                if   "dia"    in unit or "día" in unit: return n
                if   "semana" in unit: return n * 7
                if   "mes"    in unit: return n * 30
                if   "año"    in unit or "ano" in unit: return n * 365
                return None
            _df_eds["_dias_freq"] = _df_eds["activador"].apply(_activador_a_dias)
            _df_eds["_prox_teorica"] = _df_eds["_ultima_pm"] + pd.to_timedelta(
                _df_eds["_dias_freq"], unit="D"
            )

            # Resolver "próxima a mostrar":
            #   1. Si hay OT futura creada → esa (📅 OT creada)
            #   2. Si no, pero podemos calcular teórica → esa (🔮 Proyección)
            #   3. Si no hay nada → —
            _df_eds["_es_real"] = _df_eds["_prox_pm"].notna()
            _df_eds["_prox_efectiva"] = _df_eds["_prox_pm"].fillna(_df_eds["_prox_teorica"])

            # Días calculados
            _df_eds["_dias_ult"]  = (_hoy_eds - _df_eds["_ultima_pm"]).dt.days
            _df_eds["_dias_prox"] = (_df_eds["_prox_efectiva"] - _hoy_eds).dt.days

            # Ordenar por proximidad
            _df_eds = _df_eds.sort_values("_dias_prox", na_position="last")

            # Origen visual
            _df_eds["Origen"] = _df_eds.apply(
                lambda r: "📅 OT en Fracttal" if r["_es_real"]
                else ("🔮 Proyección teórica" if pd.notna(r["_prox_efectiva"]) else "—"),
                axis=1,
            )

            # Formatear
            _df_eds_show = _df_eds.copy()
            _df_eds_show["Última PM"]        = _df_eds["_ultima_pm"].dt.strftime("%d/%m/%Y").fillna("—")
            _df_eds_show["Próxima PM"]       = _df_eds["_prox_efectiva"].dt.strftime("%d/%m/%Y").fillna("—")
            _df_eds_show["Días desde últ."]  = _df_eds["_dias_ult"].apply(
                lambda x: int(x) if pd.notna(x) else None)
            _df_eds_show["Días hasta prox."] = _df_eds["_dias_prox"].apply(
                lambda x: int(x) if pd.notna(x) else None)

            _df_eds_show = _df_eds_show[[
                "codigo_activo","nombre_activo","estacion","codigo_eds",
                "plan_tareas","Total_PM","activador",
                "Última PM","Días desde últ.",
                "Próxima PM","Días hasta prox.","Origen","resp_prox"
            ]].rename(columns={
                "codigo_activo":"Código","nombre_activo":"Activo",
                "estacion":"Estación","codigo_eds":"Cód. EDS",
                "plan_tareas":"Plan",
                "Total_PM":"Total PM","activador":"Activador",
                "resp_prox":"Próx. Resp."
            })

            st.markdown('<div class="section-header">🏭  Estado de mantenciones por activo / EDS</div>',
                        unsafe_allow_html=True)
            st.caption(
                "**Próxima PM**: si Fracttal ya creó la OT → 📅 OT en Fracttal · si "
                "no, se calcula 🔮 Proyección teórica = última PM + frecuencia del activador. "
                "Activadores por horas de funcionamiento no se proyectan (depende del uso)."
            )
            _show_df(_df_eds_show.reset_index(drop=True), hide_index=True, use_container_width=True,
                column_config={
                    "Código":           st.column_config.TextColumn(width=90),
                    "Activo":           st.column_config.TextColumn(width=220),
                    "Estación":         st.column_config.TextColumn(width=200),
                    "Cód. EDS":         st.column_config.TextColumn(width=85),
                    "Plan":             st.column_config.TextColumn(width=200),
                    "Total PM":         st.column_config.NumberColumn(format="%d", width=70),
                    "Activador":        st.column_config.TextColumn(width=110),
                    "Última PM":        st.column_config.TextColumn(width=90),
                    "Días desde últ.":  st.column_config.NumberColumn(format="%d", width=110),
                    "Próxima PM":       st.column_config.TextColumn(width=90),
                    "Días hasta prox.": st.column_config.NumberColumn(format="%d", width=115),
                    "Origen":           st.column_config.TextColumn(width=160),
                    "Próx. Resp.":      st.column_config.TextColumn(width=160),
                })
            _real = int(_df_eds["_es_real"].sum())
            _teor = int(_df_eds["_prox_teorica"].notna().sum() - _real)
            _sin  = int(len(_df_eds) - _real - _teor)
            st.caption(
                f"{len(_df_eds_show):,} activos únicos · "
                f"📅 {_real} con OT en Fracttal · 🔮 {_teor} con proyección teórica · "
                f"{_sin} sin datos para proyectar"
            )

    # ── Tab 6: Uptime ─────────────────────────────────────────────────────
    with _ptab_uptime:
        st.markdown(
            '<div class="section-header">⏱️  Uptime por mantención planificada</div>',
            unsafe_allow_html=True,
        )

        if "paro_equipo" not in dfp.columns or dfp["paro_equipo"].isna().all():
            st.info(
                "Sin datos de paro de equipo en este filtro. Verifica que el sync de "
                "Fracttal haya poblado los campos `paro_equipo` y `tiempo_paro_real_seg`."
            )
        else:
            _hoy = pd.Timestamp.today().normalize()
            _rango_dias = max((_hoy - _rango_inicio).days, 1)
            _seg_por_equipo = _rango_dias * 24 * 3600   # horas disponibles por equipo individual

            # Encabezado con período evaluado explícito
            st.markdown(
                f"""<div style="background:rgba(1,121,138,0.10);border-left:3px solid #01798A;
                     padding:10px 16px;border-radius:6px;margin-bottom:14px;">
                  <span style="color:#01798A;font-weight:700;font-size:0.95rem;">📅 Período evaluado:</span>
                  <span style="color:var(--text-color, #475569);font-size:0.9rem;margin-left:8px;">
                    <b>{_rango_inicio.strftime('%d/%m/%Y')}</b> → <b>{_hoy.strftime('%d/%m/%Y')}</b>
                    ({_rango_dias} días · {_rango_dias * 24:,} horas por equipo).</span>
                </div>""",
                unsafe_allow_html=True,
            )
            st.caption(
                "**Fórmula**: Uptime % = 1 − (Σ horas detenidas por MP) ÷ (N equipos × horas del período) · "
                "Solo cuenta como detención cuando la OT tiene **¿Paro de equipo? = SÍ**. "
                "El denominador es **horas-equipo disponibles** (cada equipo aporta sus 24h × días). "
                "Mide qué % del tiempo la flota completa estuvo operativa."
            )

            # ── Filtros: Cliente + Región ────────────────────────────────
            def _region_estacion(s):
                s = str(s or "").upper()
                NORTE = ["IQUIQUE","ARICA","ANTOFAGASTA","CALAMA","COPIAPÓ","COPIAPO",
                         "OVALLE","LA SERENA","COQUIMBO","VALLENAR","ILLAPEL","HUASCO",
                         "ALTO HOSPICIO","TOCOPILLA","MEJILLONES"]
                SUR   = ["CONCEPCIÓN","CONCEPCION","OSORNO","TEMUCO","VALDIVIA",
                         "PUERTO MONTT","CHILLÁN","CHILLAN","LOS ANGELES","LOS ÁNGELES",
                         "LINARES","TALCA","CURICÓ","CURICO","RANCAGUA","FRUTILLAR",
                         "VICTORIA","RENGO","SAN FERNANDO","PUERTO VARAS","ANCUD",
                         "CASTRO","COIHAIQUE","LA UNION","LA UNIÓN"]
                if any(c in s for c in NORTE): return "Norte"
                if any(c in s for c in SUR):   return "Sur"
                return "Santiago"

            _dfup = dfp.copy()
            _dfup["_region"] = _dfup["estacion"].fillna(_dfup.get("ubicacion","")).apply(_region_estacion)

            _fu1, _fu2 = st.columns(2)
            with _fu1:
                _cli_opts = _orden_clientes(_dfup["cliente"]) \
                    if "cliente" in _dfup.columns else ["Todos"]
                _up_cli = st.selectbox("Cliente", _cli_opts, key="up_cliente")
            with _fu2:
                _up_reg = st.selectbox("Región", ["Todas", "Santiago", "Norte", "Sur"], key="up_region")

            if _up_cli != "Todos":
                _dfup = _dfup[_dfup["cliente"] == _up_cli]
            if _up_reg != "Todas":
                _dfup = _dfup[_dfup["_region"] == _up_reg]

            # ── Cálculo principal sobre el subset filtrado ───────────────
            _con_paro = _dfup[_dfup["paro_equipo"] == True].copy()
            _paro_real = pd.to_numeric(_con_paro.get("tiempo_paro_real_seg"), errors="coerce").fillna(0)
            _paro_estim= pd.to_numeric(_con_paro.get("tiempo_paro_estim_seg"), errors="coerce").fillna(0)
            _paro_seg_serie = _paro_real.where(_paro_real > 0, _paro_estim)
            _paro_seg  = _paro_seg_serie.sum()
            if not _con_paro.empty:
                _con_paro = _con_paro.assign(_paro_calc_seg=_paro_seg_serie.values)

            # Universo de equipos en el filtro (codigo_activo puede venir
            # concatenado por comas si una OT cubre múltiples equipos).
            _eq_unicos = set()
            for _cod in _dfup.get("codigo_activo", pd.Series(dtype=str)).dropna():
                for _c in str(_cod).split(","):
                    _c = _c.strip()
                    if _c:
                        _eq_unicos.add(_c)
            _n_equipos = max(len(_eq_unicos), 1)

            # Denominador correcto: horas-equipo (cada equipo aporta sus 24h × días)
            _total_periodo_seg = _seg_por_equipo * _n_equipos

            _uk1, _uk2, _uk3, _uk4, _uk5 = st.columns(5)
            _uk1.metric("OTs con paro de equipo", f"{len(_con_paro):,}",
                        delta=f"de {len(_dfup):,} OTs en filtro")
            _uk2.metric("Tiempo total detenido",
                        f"{int(_paro_seg // 3600):,}h {int((_paro_seg % 3600) // 60):02d}m")
            _uk3.metric("Equipos evaluados", f"{_n_equipos:,}",
                        delta="únicos en el filtro")
            _uk4.metric("Días evaluados", f"{_rango_dias:,}",
                        delta="desde 01/01/2026")
            _uptime_pct_global = max(0.0, round((1 - _paro_seg / _total_periodo_seg) * 100, 3))
            _uk5.metric("Uptime global", f"{_uptime_pct_global}%",
                        help=f"({_n_equipos:,} equipos × {_rango_dias} días × 24h) − {int(_paro_seg//3600):,}h detenidas")

            st.divider()

            # ── Ranking: 5 EDS con mayor tiempo de paro ──────────────────
            st.markdown(
                '<div class="section-header">🏆  Ranking · 5 estaciones con mayor tiempo de paro</div>',
                unsafe_allow_html=True,
            )
            if _con_paro.empty:
                st.info("Ninguna OT del filtro tiene paro de equipo registrado.")
            else:
                _rank_eds = (_con_paro.groupby(
                    ["estacion", "codigo_eds"], dropna=False, as_index=False
                ).agg(
                    OTs_con_paro=("id_ot", "count"),
                    Tiempo_paro_seg=("_paro_calc_seg", "sum"),
                ))
                _rank_eds["Tiempo paro (h:mm)"] = _rank_eds["Tiempo_paro_seg"].apply(
                    lambda s: f"{int(s // 3600):,}h {int((s % 3600) // 60):02d}m" if s else "—"
                )
                _rank_eds["Tiempo paro (h)"] = (_rank_eds["Tiempo_paro_seg"] / 3600).round(1)
                _rank_eds = _rank_eds.sort_values("Tiempo_paro_seg", ascending=False).head(5)
                _rank_eds["#"] = ["🥇","🥈","🥉","4️⃣","5️⃣"][:len(_rank_eds)]

                _rank_show = _rank_eds[[
                    "#", "estacion", "codigo_eds", "OTs_con_paro",
                    "Tiempo paro (h:mm)", "Tiempo paro (h)"
                ]].rename(columns={
                    "estacion":      "Estación",
                    "codigo_eds":    "Cód. EDS",
                    "OTs_con_paro":  "OTs con paro",
                })
                _show_df(_rank_show.reset_index(drop=True), hide_index=True,
                    use_container_width=True,
                    column_config={
                        "#":                  st.column_config.TextColumn(width=50),
                        "Estación":           st.column_config.TextColumn(width=260),
                        "Cód. EDS":           st.column_config.TextColumn(width=100),
                        "OTs con paro":       st.column_config.NumberColumn(format="%d", width=120),
                        "Tiempo paro (h:mm)": st.column_config.TextColumn(width=140),
                        "Tiempo paro (h)":    st.column_config.ProgressColumn(
                            format="%.1f h",
                            min_value=0,
                            max_value=float(_rank_eds["Tiempo paro (h)"].max()) if not _rank_eds.empty else 1,
                            width=200,
                        ),
                    })

            st.divider()

            # ── Uptime por equipo ────────────────────────────────────────
            st.markdown(
                '<div class="section-header">🔧  Uptime por equipo</div>',
                unsafe_allow_html=True,
            )
            st.caption(
                "Cada fila es un equipo individual (o grupo si la OT tiene varios). "
                "**Uptime %** = % del tiempo del período en que el equipo NO estuvo "
                "detenido por mantención planificada. 100% = nunca paró por MP."
            )
            if _con_paro.empty:
                st.info("Ninguna OT del filtro tiene paro de equipo registrado.")
            else:
                # Último responsable por equipo (técnico que hizo la última MP)
                _ultimo_tec = (
                    _dfup[_dfup["fecha_finalizacion"].notna()]
                    .sort_values("fecha_finalizacion", ascending=False)
                    .groupby("codigo_activo", as_index=False)
                    .first()[["codigo_activo","responsable","fecha_finalizacion"]]
                    .rename(columns={"responsable":"_ultimo_tec","fecha_finalizacion":"_ultima_pm"})
                )
                _gp = (_con_paro.groupby(
                    ["codigo_activo", "nombre_activo", "estacion", "codigo_eds"],
                    dropna=False, as_index=False,
                ).agg(
                    OTs_con_paro=("id_ot", "count"),
                    Tiempo_paro_seg=("_paro_calc_seg", "sum"),
                ))
                _gp = _gp.merge(_ultimo_tec, on="codigo_activo", how="left")
                _gp["Última PM"] = pd.to_datetime(
                    _gp["_ultima_pm"], errors="coerce", utc=True
                ).dt.tz_convert("America/Santiago").dt.tz_localize(None).dt.strftime("%d/%m/%Y").fillna("—")
                _gp["Tiempo paro (h:mm)"] = _gp["Tiempo_paro_seg"].apply(
                    lambda s: f"{int(s // 3600):,}h {int((s % 3600) // 60):02d}m" if s else "—"
                )
                # Uptime por equipo individual: cada equipo dispone de _seg_por_equipo
                # horas en el período (no se multiplica por N: aquí cada fila ES un equipo).
                _gp["Uptime %"] = (1 - _gp["Tiempo_paro_seg"] / _seg_por_equipo).clip(lower=0) * 100
                _gp["Uptime %"] = _gp["Uptime %"].round(3)
                _gp = _gp.sort_values("Tiempo_paro_seg", ascending=False).head(30)

                _gp_show = _gp[[
                    "codigo_activo","nombre_activo","estacion","codigo_eds",
                    "OTs_con_paro","Tiempo paro (h:mm)","Última PM","_ultimo_tec","Uptime %"
                ]].rename(columns={
                    "codigo_activo": "Código",
                    "nombre_activo": "Activo",
                    "estacion":      "Estación",
                    "codigo_eds":    "Cód. EDS",
                    "OTs_con_paro":  "OTs c/paro",
                    "_ultimo_tec":   "Último técnico",
                })
                _show_df(_gp_show.reset_index(drop=True), hide_index=True,
                         use_container_width=True,
                         column_config={
                             "Código":             st.column_config.TextColumn(width=130),
                             "Activo":             st.column_config.TextColumn(width=260),
                             "Estación":           st.column_config.TextColumn(width=200),
                             "Cód. EDS":           st.column_config.TextColumn(width=90),
                             "OTs c/paro":         st.column_config.NumberColumn(format="%d", width=90),
                             "Tiempo paro (h:mm)": st.column_config.TextColumn(width=130),
                             "Última PM":          st.column_config.TextColumn(width=100),
                             "Último técnico":     st.column_config.TextColumn(width=180),
                             "Uptime %":           st.column_config.NumberColumn(format="%.3f%%", width=100),
                         })
                st.caption(
                    f"Top 30 equipos ordenados por tiempo total detenido · "
                    f"período evaluado: {_rango_inicio.strftime('%d/%m/%Y')} → {_hoy.strftime('%d/%m/%Y')}"
                )

# ══════════════════════════════════════════════════════════════════════════════
# PÁGINA ADMINISTRACIÓN (solo admins)
# ══════════════════════════════════════════════════════════════════════════════
elif _page == "🔐  Administración":
    if not _is_admin:
        st.error("Acceso restringido — solo administradores.")
        st.stop()
    _hdr("Administración", "Gestión de usuarios y analítica de uso")

    from supabase_client import (
        get_usuarios_admin, get_sesiones_admin,
        upsert_usuario_dashboard, _patch,
    )

    _adm_tab1, _adm_tab2 = st.tabs(["👥  Usuarios", "📊  Analítica de Uso"])

    # ── TAB 1: Gestión de usuarios ────────────────────────────────────────────
    with _adm_tab1:
        st.markdown("#### Usuarios autorizados")

        # Formulario agregar / editar usuario
        with st.expander("➕ Agregar o actualizar usuario", expanded=False):
            _ua_c1, _ua_c2, _ua_c3, _ua_c4 = st.columns([3, 2, 2, 1])
            with _ua_c1:
                _ua_email = st.text_input("Correo", placeholder="nombre@occimiano.cl",
                                          key="ua_email")
            with _ua_c2:
                _ua_nombre = st.text_input("Nombre", key="ua_nombre")
            with _ua_c3:
                _ua_rol = st.selectbox("Rol", ["usuario", "admin"], key="ua_rol")
            with _ua_c4:
                _ua_activo = st.checkbox("Activo", value=True, key="ua_activo")
            if st.button("💾 Guardar usuario", key="ua_save"):
                if not _ua_email.strip():
                    st.error("Ingresa un correo.")
                else:
                    ok = upsert_usuario_dashboard(
                        _ua_email, _ua_nombre, _ua_rol, _ua_activo
                    )
                    if ok:
                        st.success(f"✅ Usuario **{_ua_email}** guardado con rol **{_ua_rol}**.")
                        get_usuarios_admin.clear()
                        st.rerun()
                    else:
                        st.error("Error al guardar. Verifica los datos.")

        # ── Acciones de contraseña (invitar / forzar reset) ─────────────
        with st.expander("✉️ Enviar correo de invitación o reset de contraseña",
                          expanded=False):
            st.caption(
                "**Invitar usuario nuevo** — manda un correo para que defina "
                "su 1ª contraseña (link válido 24h).  \n"
                "**Forzar reset** — útil si un usuario olvidó su contraseña y "
                "no le llegó el correo automático (link válido 15min)."
            )
            _pc1, _pc2, _pc3 = st.columns([3, 2, 2])
            with _pc1:
                _pw_email = st.text_input("Correo del usuario",
                                          placeholder="nombre@occimiano.cl",
                                          key="pw_action_email")
            with _pc2:
                if st.button("📧 Enviar invitación", key="pw_invite_btn",
                              use_container_width=True):
                    ok, msg = request_password_reset(_pw_email, proposito="invite")
                    (st.success if ok else st.error)(msg)
            with _pc3:
                if st.button("🔄 Forzar reset", key="pw_reset_btn",
                              use_container_width=True):
                    ok, msg = request_password_reset(_pw_email, proposito="reset")
                    (st.success if ok else st.error)(msg)

        # Tabla de usuarios
        _df_usr = get_usuarios_admin()
        if _df_usr.empty:
            st.info("No hay usuarios en la tabla usuarios_dashboard.")
        else:
            _df_usr_show = _df_usr.copy()
            for _dc in ["creado_en", "ultimo_acceso"]:
                if _dc in _df_usr_show.columns:
                    _df_usr_show[_dc] = pd.to_datetime(
                        _df_usr_show[_dc], errors="coerce", utc=True
                    ).dt.tz_convert("America/Santiago").dt.strftime("%d/%m/%Y %H:%M").fillna("—")
            _df_usr_show = _df_usr_show.rename(columns={
                "email": "Correo", "nombre": "Nombre", "rol": "Rol",
                "activo": "Activo", "creado_en": "Creado", "ultimo_acceso": "Último acceso",
            })
            st.dataframe(
                _df_usr_show,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Correo":        st.column_config.TextColumn(width=250),
                    "Nombre":        st.column_config.TextColumn(width=180),
                    "Rol":           st.column_config.TextColumn(width=90),
                    "Activo":        st.column_config.CheckboxColumn(width=70),
                    "Creado":        st.column_config.TextColumn(width=130),
                    "Último acceso": st.column_config.TextColumn(width=130),
                },
            )
            st.caption(f"{len(_df_usr_show)} usuarios registrados")

    # ── TAB 2: Analítica de uso ───────────────────────────────────────────────
    with _adm_tab2:
        st.markdown("#### Sesiones del dashboard")
        _df_ses = get_sesiones_admin()
        # Excluir cuentas internas (desarrollo / administración) — su uso intensivo
        # distorsiona la analítica a nivel usuario operativo.
        _EMAILS_EXCLUIR_ANALITICA = {"jgavidia@occimiano.cl"}
        if not _df_ses.empty and "email" in _df_ses.columns:
            _df_ses = _df_ses[
                ~_df_ses["email"].str.strip().str.lower().isin(_EMAILS_EXCLUIR_ANALITICA)
            ].copy()
        if _df_ses.empty:
            st.info("No hay sesiones registradas todavía.")
        else:
            # Calcular duración de sesión
            for _dc in ["fecha_inicio", "ultima_actividad"]:
                if _dc in _df_ses.columns:
                    _df_ses[_dc] = pd.to_datetime(
                        _df_ses[_dc], errors="coerce", utc=True
                    )
            if "fecha_inicio" in _df_ses.columns and "ultima_actividad" in _df_ses.columns:
                _df_ses["duracion_min"] = (
                    (_df_ses["ultima_actividad"] - _df_ses["fecha_inicio"])
                    .dt.total_seconds() / 60
                ).round(1).clip(lower=0)
            else:
                _df_ses["duracion_min"] = None

            # Resumen por usuario
            st.markdown("##### Resumen por usuario")
            _df_res = (
                _df_ses.groupby("email")
                .agg(
                    sesiones      = ("id", "count"),
                    ultima_visita = ("fecha_inicio", "max"),
                    min_totales   = ("duracion_min", "sum"),
                )
                .reset_index()
                .sort_values("ultima_visita", ascending=False)
            )
            _df_res["ultima_visita"] = (
                _df_res["ultima_visita"]
                .dt.tz_convert("America/Santiago")
                .dt.strftime("%d/%m/%Y %H:%M")
            )
            _df_res["min_totales"] = _df_res["min_totales"].round(0).astype(int)
            _df_res = _df_res.rename(columns={
                "email": "Correo", "sesiones": "Sesiones",
                "ultima_visita": "Última visita", "min_totales": "Min. totales",
            })
            st.dataframe(_df_res, use_container_width=True, hide_index=True)

            # ── Gráfico de conectividad por usuario ───────────────────────
            if "duracion_min" in _df_ses.columns:
                _uso_grp = (
                    _df_ses.groupby("email")["duracion_min"]
                    .sum().round(1).reset_index()
                    .sort_values("duracion_min", ascending=True)
                )
                # Mostrar en horas si algún usuario supera 120 min
                _usar_horas = _uso_grp["duracion_min"].max() > 120
                if _usar_horas:
                    _uso_grp["_val"] = (_uso_grp["duracion_min"] / 60).round(1)
                    _unidad_lbl = "h"
                else:
                    _uso_grp["_val"] = _uso_grp["duracion_min"].round(1)
                    _unidad_lbl = "min"
                # Solo el nombre antes del @
                _uso_grp["_usr"] = _uso_grp["email"].str.split("@").str[0]

                _fig_uso = go.Figure(go.Bar(
                    y=_uso_grp["_usr"],
                    x=_uso_grp["_val"],
                    orientation="h",
                    marker=dict(
                        color=_uso_grp["_val"],
                        colorscale=[[0, "#93c5fd"], [1, "#1e3a5f"]],
                        showscale=False,
                    ),
                    text=[f"  {v:.1f} {_unidad_lbl}" for v in _uso_grp["_val"]],
                    textposition="outside",
                    textfont=dict(size=12, color="#334155", family="Arial"),
                    hovertemplate="<b>%{y}</b><br>Tiempo: %{x:.1f} " + _unidad_lbl + "<extra></extra>",
                ))
                _fig_uso.update_layout(
                    title=dict(
                        text=f"Tiempo total de uso por usuario ({_unidad_lbl})",
                        font=dict(size=15, color="#1e293b"),
                    ),
                    height=max(220, len(_uso_grp) * 56 + 100),
                    margin=dict(l=10, r=110, t=50, b=30),
                    xaxis=dict(
                        title=f"Tiempo ({_unidad_lbl})",
                        showgrid=True, gridcolor="#e2e8f0",
                        zeroline=False,
                    ),
                    yaxis=dict(showgrid=False, automargin=True),
                    plot_bgcolor="rgba(0,0,0,0)",
                    paper_bgcolor="rgba(0,0,0,0)",
                    font=dict(family="Arial"),
                )
                st.plotly_chart(_fig_uso, use_container_width=True)

            st.divider()
            st.markdown("##### Detalle de sesiones")
            _df_ses_show = _df_ses.copy()
            _df_ses_show["fecha_inicio"] = (
                _df_ses_show["fecha_inicio"]
                .dt.tz_convert("America/Santiago")
                .dt.strftime("%d/%m/%Y %H:%M")
            )
            _df_ses_show["ultima_actividad"] = (
                _df_ses_show["ultima_actividad"]
                .dt.tz_convert("America/Santiago")
                .dt.strftime("%d/%m/%Y %H:%M")
            )
            _df_ses_show = _df_ses_show[
                [c for c in ["email", "fecha_inicio", "ultima_actividad",
                              "duracion_min", "pagina_actual"] if c in _df_ses_show.columns]
            ].rename(columns={
                "email": "Correo", "fecha_inicio": "Inicio",
                "ultima_actividad": "Última actividad",
                "duracion_min": "Duración (min)", "pagina_actual": "Última página",
            })
            st.dataframe(
                _df_ses_show,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Correo":           st.column_config.TextColumn(width=230),
                    "Inicio":           st.column_config.TextColumn(width=130),
                    "Última actividad": st.column_config.TextColumn(width=130),
                    "Duración (min)":   st.column_config.NumberColumn(width=110, format="%.1f min"),
                    "Última página":    st.column_config.TextColumn(width=230),
                },
            )
            st.caption(f"{len(_df_ses_show)} sesiones registradas en total")


# ── Footer ────────────────────────────────────────────────────────────────────
st.divider()

# Timestamps de actualización de datos — todo en hora Chile (America/Santiago)
from zoneinfo import ZoneInfo as _ZoneInfo
_tz_stgo = _ZoneInfo("America/Santiago")

def _to_stgo(ts) -> "pd.Timestamp":
    """Convierte un Timestamp (con o sin tz) a hora Santiago."""
    t = pd.Timestamp(ts)
    if t.tzinfo is None:
        t = t.tz_localize("UTC")   # Supabase almacena en UTC
    return t.tz_convert("America/Santiago")

if "_session_start" not in st.session_state:
    st.session_state["_session_start"] = datetime.now(_tz_stgo)

_ultima_ot_str = "—"
_ultima_ll_str = "—"
try:
    # OT más reciente en Supabase (caché ya cargado, sin request adicional)
    _raw_check = load_work_orders_supabase()
    _cd_vals = [r.get("creation_date") for r in _raw_check if r.get("creation_date")]
    if _cd_vals:
        _ultima_ot_str = _to_stgo(max(_cd_vals)).strftime("%d/%m/%Y %H:%M")
except Exception:
    pass
try:
    # Llamado más reciente en Supabase (v_llamados_sla, caché vía _sc)
    _ll_check = st.session_state.get("_sc_df_llamados_supa")
    if _ll_check is None:
        _ll_check = load_all_llamados_supabase()
    if isinstance(_ll_check, pd.DataFrame) and not _ll_check.empty and "fecha_llamado" in _ll_check.columns:
        _max_ll = pd.to_datetime(_ll_check["fecha_llamado"], errors="coerce").max()
        if pd.notna(_max_ll):
            _ultima_ll_str = _to_stgo(_max_ll).strftime("%d/%m/%Y")
except Exception:
    pass

_session_ts = st.session_state["_session_start"]
if hasattr(_session_ts, "strftime"):
    _session_str = _session_ts.strftime("%d/%m/%Y %H:%M")
else:
    _session_str = "—"
_footer_col1, _footer_col2 = st.columns([5, 5])
with _footer_col1:
    st.caption("Occimiano — Indicadores Operacionales v1.2 | Fracttal One API + Supabase")
with _footer_col2:
    st.markdown(
        f'<div style="text-align:right;color:#94a3b8;font-size:0.72rem;line-height:1.6;">'
        f'📦 Última OT en Supabase: <b>{_ultima_ot_str}</b> &nbsp;·&nbsp; '
        f'📋 Último llamado: <b>{_ultima_ll_str}</b><br>'
        f'🔄 Dashboard cargado: <b>{_session_str}</b>'
        f'</div>',
        unsafe_allow_html=True,
    )
