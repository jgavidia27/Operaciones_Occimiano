"""
turnos_data.py - Parse shift schedules from 'Turnos Occim.xlsx' and export to JSON.

Reads 3 sheets (Centro, Norte, Sur), extracts weekly shift blocks,
normalizes technician names (fixing encoding issues and abbreviations),
and outputs a structured turnos_data.json grouped by week with all 3 zones.
"""

import json
import os
import re
from datetime import date, datetime, timedelta

import openpyxl

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

EXCEL_PATH = r"C:\Users\jgavi\Downloads\Turnos Occim.xlsx"
OUTPUT_DIR = r"C:\Users\jgavi\Documents\occimiano_dashboard"
OUTPUT_JSON = os.path.join(OUTPUT_DIR, "turnos_data.json")
YEAR = 2026

SHEET_CONFIG = {
    "TURNOS ZONA CENTRO": {
        "zone": "centro",
        "num_technicians": 3,
        "has_obs": True,
        "obs_col": 11,  # 1-based column K
    },
    "TURNOS ZONA NTE": {
        "zone": "norte",
        "num_technicians": 3,
        "has_obs": False,
    },
    "TURNO ZONA SUR": {
        "zone": "sur",
        "num_technicians": 2,
        "has_obs": False,
    },
}

# Month name mapping (Spanish abbreviations used in the Excel)
MONTH_MAP = {
    "ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12,
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5,
    "JUNIO": 6, "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9,
    "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}

# Abbreviation -> full name mapping (used in Norte zone for later weeks)
# Uses accented forms to match the canonical names from the Excel.
ABBREVIATION_MAP = {
    "CA": "Carlos Alberto Avila Palacios",
    "ER": "Erwin Maximiliano Rivera Talamilla",
    "EP": "Edson José Pérez Henríquez",  # accented to match Excel
}

# Name normalization: variant spellings -> canonical form.
# The Excel names already contain correct Unicode accents (e.g. U+00E9 for e-acute).
# This map handles spelling variants (e.g. middle name included in some weeks).
NAME_FIXES = {
    "Edison Jhon Carrasco Navarro": "Edison Carrasco Navarro",
}


def normalize_name(raw_name):
    """Normalize a technician name: expand abbreviations, fix spelling variants."""
    if raw_name is None:
        return "N/A"
    name = str(raw_name).strip()
    if not name:
        return "N/A"

    # Check abbreviation map first (short 2-3 char codes)
    if name.upper() in ABBREVIATION_MAP:
        return ABBREVIATION_MAP[name.upper()]

    # Check name fixes map (spelling variants)
    if name in NAME_FIXES:
        return NAME_FIXES[name]

    return name


def parse_month_header(month_str):
    """Parse month header string like 'JUN', 'JUN/ JUL', 'JUL / AGO'.
    Returns a list of month numbers (1-2 elements)."""
    if month_str is None:
        return []
    month_str = str(month_str).strip().upper()
    # Split on / and clean
    parts = [p.strip() for p in month_str.split("/")]
    months = []
    for part in parts:
        part = part.strip()
        if part in MONTH_MAP:
            months.append(MONTH_MAP[part])
    return months


def resolve_dates(months, day_numbers):
    """Given month(s) and 7 day numbers, resolve to actual dates.

    For a single month like 'JUN' with days [22,23,24,25,26,27,28] -> all June.
    For cross-month like 'JUN/JUL' with days [29,30,1,2,3,4,5] -> Jun 29-30, Jul 1-5.
    For single month header but cross-month days (e.g., 'JUL' with [27,28,29,30,31,1,2])
    -> detect the transition when days decrease.
    """
    dates = []
    if len(months) == 0:
        return dates

    primary_month = months[0]
    secondary_month = months[1] if len(months) > 1 else None

    # If no secondary month is specified but days wrap around (decrease),
    # infer the next month
    if secondary_month is None:
        for i in range(1, len(day_numbers)):
            if day_numbers[i] < day_numbers[i - 1]:
                # Day decreased -> month transition
                secondary_month = primary_month + 1
                if secondary_month > 12:
                    secondary_month = 1
                break

    current_month = primary_month
    prev_day = 0
    for day_num in day_numbers:
        day = int(day_num)
        # If day decreased, we've crossed into the next month
        if day < prev_day and secondary_month is not None:
            current_month = secondary_month
        try:
            year = YEAR
            # Handle year wrap (Dec -> Jan) if needed
            d = date(year, current_month, day)
            dates.append(d)
        except ValueError:
            # Invalid date - skip
            dates.append(None)
        prev_day = day

    return dates


def extract_turno_number(turno_str):
    """Extract turno number from string like 'Sem. Turno 1 ZC' -> 1."""
    if turno_str is None:
        return 0
    match = re.search(r"Turno\s+(\d+)", str(turno_str))
    if match:
        return int(match.group(1))
    return 0


def parse_sheet(ws, config):
    """Parse a single sheet and return a list of week records for this zone."""
    zone = config["zone"]
    num_techs = config["num_technicians"]
    has_obs = config.get("has_obs", False)
    obs_col = config.get("obs_col", 11)

    weeks = []
    row = 1
    max_row = ws.max_row

    while row <= max_row:
        # Look for a month header row: col 3 has month text, cols 4-10 have day numbers
        month_cell = ws.cell(row, 3).value
        if month_cell is None:
            row += 1
            continue

        month_str = str(month_cell).strip()

        # Check if this is a month header (has day numbers in cols 4-10)
        day_numbers = []
        for c in range(4, 11):
            v = ws.cell(row, c).value
            if v is not None:
                try:
                    day_numbers.append(float(v))
                except (ValueError, TypeError):
                    break

        if len(day_numbers) != 7:
            row += 1
            continue

        # This is a month/date header row
        months = parse_month_header(month_str)
        if not months:
            row += 1
            continue

        dates = resolve_dates(months, day_numbers)
        if len(dates) != 7 or any(d is None for d in dates):
            row += 1
            continue

        # Next row should be the column header row (Equipo, TEC, Sem.X, L, M, X, J, V, S, D)
        row += 1
        header_check = ws.cell(row, 1).value
        if header_check is None or str(header_check).strip() != "Equipo":
            continue

        # Next rows are the data rows (num_techs rows)
        row += 1
        equipo = None
        turnos = []
        obs_text = ""

        for t in range(num_techs):
            data_row = row + t
            if data_row > max_row:
                break

            # Team name only in first row of block
            eq_val = ws.cell(data_row, 1).value
            if eq_val is not None and str(eq_val).strip():
                equipo = str(eq_val).strip()

            tec_name = normalize_name(ws.cell(data_row, 2).value)
            turno_str = ws.cell(data_row, 3).value
            turno_num = extract_turno_number(turno_str)
            if turno_num == 0:
                turno_num = t + 1

            horarios = []
            for c in range(4, 11):
                v = ws.cell(data_row, c).value
                horarios.append(str(v).strip() if v is not None else "")

            turnos.append({
                "turno": turno_num,
                "tecnico": tec_name,
                "horarios": horarios,
            })

            # Check for obs in this row (centro only)
            if has_obs:
                obs_val = ws.cell(data_row, obs_col).value
                if obs_val is not None and str(obs_val).strip():
                    if obs_text:
                        obs_text += "; "
                    obs_text += str(obs_val).strip()

        row += num_techs

        # Build the week record
        week_dates = [d.isoformat() for d in dates]
        week_record = {
            "start_date": dates[0].isoformat(),
            "end_date": dates[6].isoformat(),
            "dates": week_dates,
            "zone": zone,
            "equipo": equipo or "",
            "turnos": turnos,
            "obs": obs_text,
        }
        weeks.append(week_record)

    return weeks


def group_weeks_by_date(all_zone_weeks):
    """Group week records from all zones by their date range (start_date-end_date).
    Returns list of combined week dicts with all 3 zones."""
    # Key: (start_date, end_date) -> {zone: data}
    grouped = {}
    for w in all_zone_weeks:
        key = (w["start_date"], w["end_date"])
        if key not in grouped:
            grouped[key] = {
                "start_date": w["start_date"],
                "end_date": w["end_date"],
                "dates": w["dates"],
                "zones": {},
            }
        grouped[key]["zones"][w["zone"]] = {
            "equipo": w["equipo"],
            "turnos": w["turnos"],
            "obs": w["obs"],
        }

    # Sort by start_date
    sorted_weeks = sorted(grouped.values(), key=lambda x: x["start_date"])
    return sorted_weeks


def get_turno_hoy(data=None):
    """Return the current week's shift data with today highlighted.

    If data is None, loads from turnos_data.json.
    Returns a dict with:
      - week: the current week's data (or None if not found)
      - today: today's ISO date string
      - today_index: 0-6 index for today in the week (Mon=0, Sun=6), or -1
      - today_schedules: dict per zone with each technician's schedule for today
    """
    if data is None:
        json_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "turnos_data.json"
        )
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

    today = date.today().isoformat()
    result = {
        "today": today,
        "week": None,
        "today_index": -1,
        "today_schedules": {},
    }

    for week in data["weeks"]:
        if week["start_date"] <= today <= week["end_date"]:
            result["week"] = week
            try:
                result["today_index"] = week["dates"].index(today)
            except ValueError:
                result["today_index"] = -1

            if result["today_index"] >= 0:
                idx = result["today_index"]
                for zone_name, zone_data in week["zones"].items():
                    schedules = []
                    for turno in zone_data["turnos"]:
                        schedules.append({
                            "turno": turno["turno"],
                            "tecnico": turno["tecnico"],
                            "horario": turno["horarios"][idx] if idx < len(turno["horarios"]) else "",
                        })
                    result["today_schedules"][zone_name] = {
                        "equipo": zone_data["equipo"],
                        "turnos": schedules,
                    }
            break

    return result


def main():
    """Parse the Excel file and generate turnos_data.json."""
    print("Loading workbook: %s" % EXCEL_PATH)
    wb = openpyxl.load_workbook(EXCEL_PATH)

    all_weeks = []
    for sheet_name, config in SHEET_CONFIG.items():
        print("  Parsing sheet: %s (%s)" % (sheet_name, config["zone"]))
        ws = wb[sheet_name]
        zone_weeks = parse_sheet(ws, config)
        print("    Found %d week blocks" % len(zone_weeks))
        all_weeks.extend(zone_weeks)

    print("Total zone-week records: %d" % len(all_weeks))

    # Group by date range
    grouped = group_weeks_by_date(all_weeks)
    print("Grouped into %d unique weeks" % len(grouped))

    # Build output
    output = {
        "generated_at": datetime.now().isoformat(),
        "weeks": grouped,
    }

    # Write JSON
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print("Written to: %s" % OUTPUT_JSON)

    # Print summary
    for week in grouped:
        zones_present = ", ".join(sorted(week["zones"].keys()))
        print("  %s to %s: zones=[%s]" % (
            week["start_date"], week["end_date"], zones_present
        ))
        for zn, zd in sorted(week["zones"].items()):
            techs = [t["tecnico"] for t in zd["turnos"]]
            print("    %s (%s): %s" % (zn, zd["equipo"], ", ".join(techs)))

    # Test get_turno_hoy
    print("\n--- Today's shifts ---")
    hoy = get_turno_hoy(output)
    print("Today: %s (index: %d)" % (hoy["today"], hoy["today_index"]))
    if hoy["week"]:
        print("Week: %s to %s" % (hoy["week"]["start_date"], hoy["week"]["end_date"]))
        for zn, zd in sorted(hoy["today_schedules"].items()):
            print("  %s (%s):" % (zn, zd["equipo"]))
            for t in zd["turnos"]:
                print("    Turno %d - %s: %s" % (t["turno"], t["tecnico"], t["horario"]))
    else:
        print("No shift data found for today.")


if __name__ == "__main__":
    main()
