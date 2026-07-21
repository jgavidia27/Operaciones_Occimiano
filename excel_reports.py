"""
excel_reports.py — Generador standalone del Excel Resumen mensual.
====================================================================

Extraído de app.py (_build_excel_resumen_shared) para poder ser llamado
desde:
  - El dashboard Streamlit (botón "Descargar Resumen")
  - Scripts de cron (envío semanal por email a seniors)

La función NO depende de streamlit ni de session_state. Recibe todos
los DataFrames y helpers como parámetros.

Uso desde Streamlit:
    from excel_reports import build_excel_resumen
    xlsx_bytes = build_excel_resumen(
        dl_mes="2026-07", dl_quien="Juan Gallardo", equipo_label="Juan Gallardo",
        tec_sel="Todos", sem_match=None,
        df_wo=df_wo, df_llamados=df_llamados, df_eds=df_eds,
        df_ot_scores=df_ot_scores,   # score_llenado_por_ot ya aplicado
        excel_to_full=_excel_to_full, label_to_grupo=_LABEL_TO_GRUPO,
        equipo_label_map=_EQUIPO_LABEL, es_excluido_fn=_es_excluido,
        get_equipo_fn=_get_equipo, norm_n_fn=_norm_n,
        strip_headers_fn=_strip_comentario_headers,
        build_eds_nombre_map_fn=_build_eds_nombre_map,
        numeral_motivo_label=NUMERAL_MOTIVO_LABEL,
    )

Sheets generados:
  1. Resumen           — SLA + Efectividad MP + Precisión Fracttal (KPIs)
  2. SLA               — Detalle por llamado
  3. Efectividad MP    — Reincidencias detalladas
  4. P. Fracttal       — 4 secciones (Detalle / Causa / Tiempo / Numerales)
"""

from __future__ import annotations

from io import BytesIO
from typing import Callable, Optional

import pandas as pd


def build_excel_resumen(
    dl_mes: str,
    dl_quien: str,
    equipo_label: str,
    tec_sel: str,
    sem_match: Optional[tuple],
    *,
    # ── DataFrames requeridos ────────────────────────────────────────────
    df_wo: pd.DataFrame,
    df_llamados: pd.DataFrame,
    df_eds: pd.DataFrame,
    df_ot_scores: pd.DataFrame,   # score_llenado_por_ot ya aplicado
    # ── Data opcional pre-calculada (se rebuildea si viene None) ─────────
    df_reinc: Optional[pd.DataFrame] = None,
    # ── Mapas / constantes de configuración ──────────────────────────────
    excel_to_full: dict,
    label_to_grupo: dict,
    equipo_label_map: dict,
    numeral_motivo_label: dict,
    # ── Funciones helper ─────────────────────────────────────────────────
    es_excluido_fn: Callable[[str], bool],
    get_equipo_fn: Callable[[str], str],
    norm_n_fn: Callable[[str], str],
    strip_headers_fn: Callable[[str], str],
    build_eds_nombre_map_fn: Callable[[pd.DataFrame], dict],
    # ── Contexto adicional ───────────────────────────────────────────────
    aviso_idx: Optional[dict] = None,
    build_reincidencias_fn: Optional[Callable] = None,
    score_llenado_por_tecnico_fn: Optional[Callable] = None,
    aplicar_transferencias_fn: Optional[Callable] = None,
    load_cotalker_index_fn: Optional[Callable] = None,
) -> bytes:
    """Retorna bytes del .xlsx con 4 sheets. Ver docstring del módulo."""
    buf = BytesIO()

    _LABEL_TO_GRUPO = label_to_grupo
    _EQUIPO_LABEL = equipo_label_map
    _excel_to_full = excel_to_full
    _es_excluido = es_excluido_fn
    _get_equipo = get_equipo_fn
    _norm_n = norm_n_fn
    _strip_comentario_headers = strip_headers_fn
    _build_eds_nombre_map = build_eds_nombre_map_fn
    NUMERAL_MOTIVO_LABEL = numeral_motivo_label

    # ── Precisión data: filtrar por mes/equipo/tec ──
    from data import _es_eds_excluida as _es_eds_excl_er
    if not df_ot_scores.empty and "mes" in df_ot_scores.columns:
        _df_ot_sc = df_ot_scores[df_ot_scores["mes"].astype(str) == str(dl_mes)].copy()
        # Excluir EDS internas (OCCIM-01) de Precision
        if "eds_occim" in _df_ot_sc.columns:
            _df_ot_sc = _df_ot_sc[~_df_ot_sc["eds_occim"].apply(_es_eds_excl_er)].copy()
        if sem_match is not None and "creation_date_local" in _df_ot_sc.columns:
            _df_ot_sc = _df_ot_sc[
                (_df_ot_sc["creation_date_local"] >= sem_match[1]) &
                (_df_ot_sc["creation_date_local"] <= sem_match[2])
            ]
        if equipo_label != "Todos" and "equipo" in _df_ot_sc.columns:
            _grp_kpi = _LABEL_TO_GRUPO.get(equipo_label, equipo_label)
            _df_ot_sc = _df_ot_sc[_df_ot_sc["equipo"] == _grp_kpi]
        if tec_sel != "Todos" and "tecnico" in _df_ot_sc.columns:
            _df_ot_sc = _df_ot_sc[_df_ot_sc["tecnico"] == tec_sel]
        _df_tec_sc = (score_llenado_por_tecnico_fn(_df_ot_sc)
                      if score_llenado_por_tecnico_fn else pd.DataFrame())
    else:
        _df_ot_sc = pd.DataFrame()
        _df_tec_sc = pd.DataFrame()

    # ── SLA data: preprocesar desde df_llamados ──
    _sla = pd.DataFrame()
    if not df_llamados.empty:
        _src = df_llamados[
            df_llamados["fecha_atencion"].notna() &
            df_llamados["fecha_llamado"].notna()
        ].copy()
        if "tiempo_resp_real" in _src.columns:
            _src["horas_resolucion"] = (
                pd.to_timedelta(_src["tiempo_resp_real"], errors="coerce")
                .dt.total_seconds() / 3600
            ).clip(lower=0).round(2)
            _miss = _src["horas_resolucion"].isna()
            if _miss.any():
                _src.loc[_miss, "horas_resolucion"] = (
                    (pd.to_datetime(_src.loc[_miss, "fecha_atencion"]) -
                     pd.to_datetime(_src.loc[_miss, "fecha_llamado"]))
                    .dt.total_seconds() / 3600
                ).clip(lower=0).round(2)
        else:
            _src["horas_resolucion"] = (
                (pd.to_datetime(_src["fecha_atencion"]) -
                 pd.to_datetime(_src["fecha_llamado"]))
                .dt.total_seconds() / 3600
            ).clip(lower=0).round(2)
        _src["prioridad"] = _src["prioridad"].fillna("").str.strip().str.upper()

        def _nz(z):
            z = str(z).strip().upper()
            if not z or z == "NAN":
                return "Santiago"
            if z in ("RM", "R.M.", "R.M") or any(k in z for k in ("SANTIAGO", "METRO")):
                return "Santiago"
            return "Regiones"

        _src["zona_norm"] = (_src["zona"].fillna("").astype(str).apply(_nz)
                              if "zona" in _src.columns else "Santiago")
        _SZ = {("P1","Santiago"):18,("P1","Regiones"):24,("P2","Santiago"):24,("P2","Regiones"):48,
               ("P3","Santiago"):36,("P3","Regiones"):72,("P4","Santiago"):96,("P4","Regiones"):96}
        _SD = {"P1":24,"P2":48,"P3":72,"P4":96}
        if "cumplimiento" in _src.columns:
            _src["cumple_sla"] = _src["cumplimiento"].map({"CUMPLE": True, "NO CUMPLE": False})
            _fb = _src["cumple_sla"].isna()
            if _fb.any():
                for _p in ("P1","P2","P3","P4"):
                    for _z in ("Santiago","Regiones"):
                        _m = _fb & (_src["prioridad"]==_p) & (_src["zona_norm"]==_z)
                        if _m.any():
                            _src.loc[_m,"cumple_sla"] = _src.loc[_m,"horas_resolucion"] <= _SZ.get((_p,_z),_SD.get(_p,24))
        else:
            _src["cumple_sla"] = pd.NA
            for _p in ("P1","P2","P3","P4"):
                for _z in ("Santiago","Regiones"):
                    _m = (_src["prioridad"]==_p) & (_src["zona_norm"]==_z)
                    if _m.any():
                        _src.loc[_m,"cumple_sla"] = _src.loc[_m,"horas_resolucion"] <= _SZ.get((_p,_z),_SD.get(_p,24))
        _src["fecha_llamado_dt"] = pd.to_datetime(_src["fecha_llamado"], errors="coerce")
        _src["mes"] = _src["fecha_llamado_dt"].dt.to_period("M").astype(str)
        _src["tecnico"] = _src["tecnico"].apply(
            lambda t: _excel_to_full.get(str(t).strip(), str(t).strip())
            if isinstance(t, str) and t.strip() else t)
        _src["equipo"] = _src["tecnico"].apply(_get_equipo)
        if aplicar_transferencias_fn:
            aplicar_transferencias_fn(_src, "fecha_llamado_dt", "equipo", "tecnico")
        _src = _src[~_src["tecnico"].apply(_es_excluido)].copy()
        # Excluir EDS internas (OCCIM-01) de SLA
        if "eds_occim" in _src.columns:
            _src = _src[~_src["eds_occim"].apply(_es_eds_excl_er)].copy()
        _src["equipo_label"] = _src["equipo"].map(_EQUIPO_LABEL).fillna(_src["equipo"])
        _sla = _src.copy()

    if not _sla.empty and "mes" in _sla.columns:
        _sla = _sla[_sla["mes"].astype(str) == str(dl_mes)]
        if sem_match is not None and "fecha_llamado_dt" in _sla.columns:
            _sla = _sla[
                (_sla["fecha_llamado_dt"].dt.date >= sem_match[1]) &
                (_sla["fecha_llamado_dt"].dt.date <= sem_match[2])
            ]
        if equipo_label != "Todos":
            _grp_dl = _LABEL_TO_GRUPO.get(equipo_label, equipo_label)
            _sla = _sla[_sla["equipo"] == _grp_dl]
        if tec_sel != "Todos":
            _sla = _sla[_sla["tecnico"] == tec_sel]

    # ── Reincidencia data ──
    _reinc = (df_reinc.copy() if (df_reinc is not None and not df_reinc.empty)
              else pd.DataFrame())
    if _reinc.empty and not df_wo.empty and build_reincidencias_fn is not None:
        _CL_SLA_DL = {"COPEC", "Aramco (Esmax)", "ESMAX (Aramco)", "SHELL (Enex)", "ABASTIBLE"}
        _wo_dl = df_wo[df_wo["client"].isin(_CL_SLA_DL)].copy() if "client" in df_wo.columns else df_wo
        # Excluir EDS internas (OCCIM-01) de reincidencias/Efectividad MP
        if "eds_occim" in _wo_dl.columns:
            _wo_dl = _wo_dl[~_wo_dl["eds_occim"].apply(_es_eds_excl_er)].copy()
        _reinc = build_reincidencias_fn(_wo_dl, _excel_to_full)
        if not _reinc.empty:
            if "fecha_cm" in _reinc.columns:
                _reinc = _reinc[pd.to_datetime(_reinc["fecha_cm"], errors="coerce") >= pd.Timestamp("2026-01-01")].copy()
            if "falla_tipo" in _reinc.columns and "causa_clasif" in _reinc.columns:
                _reinc["es_reincidencia_tecnico"] = (
                    (_reinc["falla_tipo"] != "especial") & (_reinc["causa_clasif"] != "cliente")
                )
            if "fecha_cm" in _reinc.columns:
                _reinc["fecha_cm_dt"] = pd.to_datetime(_reinc["fecha_cm"], errors="coerce")
                _reinc["mes"] = _reinc["fecha_cm_dt"].dt.to_period("M").astype(str)

    if not _reinc.empty and "fecha_cm" in _reinc.columns:
        if "fecha_cm_dt" not in _reinc.columns:
            _reinc["fecha_cm_dt"] = pd.to_datetime(_reinc["fecha_cm"], errors="coerce")
        if "mes" not in _reinc.columns:
            _reinc["mes"] = _reinc["fecha_cm_dt"].dt.to_period("M").astype(str)
        _reinc = _reinc[_reinc["mes"] == str(dl_mes)]
        if sem_match is not None:
            _reinc = _reinc[
                (_reinc["fecha_cm_dt"].dt.date >= sem_match[1]) &
                (_reinc["fecha_cm_dt"].dt.date <= sem_match[2])
            ]
        if equipo_label != "Todos":
            _grp_dl = _LABEL_TO_GRUPO.get(equipo_label, equipo_label)
            if "grupo_responsable" in _reinc.columns:
                _reinc = _reinc[_reinc["grupo_responsable"] == _grp_dl]
        if tec_sel != "Todos":
            if "tecnico_resp_short" in _reinc.columns:
                _norm_sel = _norm_n(tec_sel)
                _reinc = _reinc[
                    _reinc["tecnico_resp_short"].fillna("").apply(
                        lambda t: _norm_n(_excel_to_full.get(str(t).strip(), str(t).strip()))
                    ) == _norm_sel
                ]

    with pd.ExcelWriter(buf, engine="openpyxl") as wr:
        # ═══ Sheet 1: Resumen ═══
        rows_sla = []
        if not _sla.empty and "cumple_sla" in _sla.columns:
            _sg = _sla.groupby("tecnico").agg(
                cumple=("cumple_sla", lambda x: x.fillna(False).astype(bool).sum()),
                total=("cumple_sla", "count"),
            ).reset_index()
            for _, r in _sg.iterrows():
                p = round(r["cumple"] / r["total"] * 100, 1) if r["total"] > 0 else 0
                rows_sla.append((r["tecnico"], int(r["cumple"]), int(r["total"]), f"{p}%"))
            _ts = int(_sg["cumple"].sum()); _tt = int(_sg["total"].sum())
            _tp = round(_ts / _tt * 100, 1) if _tt > 0 else 0
            rows_sla.append(("Suma total", _ts, _tt, f"{_tp}%"))

        rows_mp = []
        _fallas_by_tec = {}
        if not _reinc.empty and "folio_cm" in _reinc.columns:
            _reinc_mp = _reinc.copy()
            if "falla_tipo" in _reinc_mp.columns:
                _reinc_mp = _reinc_mp[~_reinc_mp["falla_tipo"].isin(["especial"])]
            if "tecnico_responsable" in _reinc_mp.columns:
                _fg = _reinc_mp.groupby("tecnico_responsable")["folio_cm"].nunique()
                for t, n in _fg.items():
                    _fallas_by_tec[str(t).strip()] = int(n)
            elif "tecnico_resp_short" in _reinc_mp.columns:
                _fg = _reinc_mp.groupby("tecnico_resp_short")["folio_cm"].nunique()
                for t, n in _fg.items():
                    _fn = _excel_to_full.get(str(t).strip(), str(t).strip())
                    _fallas_by_tec[_fn] = int(n)

        _pm_dl = df_wo[
            (df_wo["maint_type"] == "Preventiva") &
            (~df_wo["technician"].apply(_es_excluido)) &
            (df_wo["equipo"] != "Sin equipo") &
            (~df_wo["eds_occim"].apply(_es_eds_excl_er) if "eds_occim" in df_wo.columns else True)
        ].copy()
        if not _pm_dl.empty and "creation_date" in _pm_dl.columns:
            _pmd = _pm_dl["creation_date"]
            if _pmd.dt.tz is not None:
                _pmd = _pmd.dt.tz_convert(None)
            _pm_dl = _pm_dl[_pmd.dt.to_period("M").astype(str) == str(dl_mes)]
        if equipo_label != "Todos" and "equipo" in _pm_dl.columns:
            _pm_dl = _pm_dl[_pm_dl["equipo"] == _LABEL_TO_GRUPO.get(equipo_label, equipo_label)]
        if tec_sel != "Todos" and "technician" in _pm_dl.columns:
            _pm_dl = _pm_dl[_pm_dl["technician"] == tec_sel]
        _pms_by_tec = {}
        if not _pm_dl.empty and "folio" in _pm_dl.columns:
            _pg = _pm_dl.groupby("technician")["folio"].nunique()
            for t, n in _pg.items():
                _pms_by_tec[str(t).strip()] = int(n)

        _all_mp = set(_fallas_by_tec.keys()) | set(_pms_by_tec.keys())
        _total_fallas = 0
        _total_pms = 0
        for t in sorted(_all_mp):
            _f = _fallas_by_tec.get(t, 0)
            _p = _pms_by_tec.get(t, 0)
            _ef = round((_p - _f) / _p * 100, 1) if _p > 0 else 100.0
            rows_mp.append((t, _f, f"{_ef}%"))
            _total_fallas += _f
            _total_pms += _p
        if rows_mp:
            _mef = round((_total_pms - _total_fallas) / _total_pms * 100, 1) if _total_pms > 0 else 100
            rows_mp.append(("Suma total", _total_fallas, f"{_mef}%"))

        _NCOLS_RES = 11

        def _pad(row, n=_NCOLS_RES):
            return row + [""] * (n - len(row))

        res = []
        res.append(_pad(["Cumplimiento SLA", "", "", "", "", "Efectividad MP", "", "",
                         "", "Explicativo de criterios"]))
        res.append(_pad(["Técnico", "Cumple SLA", "N° ordenes", "% Cumplimiento",
                         "", "Técnico", "Errores", "Efectividad",
                         "", "Cumplimiento SLA",
                         "% de llamados correctivos cerrados dentro del tiempo comprometido. "
                         "Binario por llamado: cumple ✅ o no cumple ❌."]))
        for i in range(max(len(rows_sla), len(rows_mp), 1)):
            row = list(rows_sla[i]) if i < len(rows_sla) else ["", "", "", ""]
            row.append("")
            row.extend(list(rows_mp[i]) if i < len(rows_mp) else ["", "", ""])
            res.append(_pad(row))
        res.append(_pad([]))
        res.append(_pad(["", "", "", "", "", "", "", "",
                         "", "Efectividad MP",
                         "Fallas post-preventiva: correctivo generado dentro de los 5 días "
                         "siguientes a un mantenimiento preventivo en el mismo equipo. "
                         "El error se imputa al técnico que realizó el preventivo."]))
        res.append(_pad([]))
        res.append(_pad([]))
        res.append(_pad(["Precisión Fracttal", "", "", "", "", "", "", "",
                         "", "Precisión Fracttal",
                         "Mide 3 componentes por OT (25 pts c/u = 75 total): "
                         "Tiempo de ejecución (solo MP), Causa raíz (solo MC) "
                         "y Numeral registrado (MP en lavadoras). "
                         "Una OT es \"mala\" si falla en cualquiera de los componentes que le aplican."]))
        res.append(_pad(["Técnico", "OTs evaluadas", "OTs sin error", "OTs con error",
                         "⏱ Tiempo OK", "\U0001f50d Causa OK", "\U0001f522 Numeral OK",
                         "Exactitud %"]))

        def _frac(ok, ap):
            if ap == 0:
                return "N/A"
            return f"{ok}/{ap} ({ok/ap*100:.1f}%)"

        if not _df_tec_sc.empty:
            for _, r in _df_tec_sc.iterrows():
                _ots = int(r.get("ots_evaluadas", 0))
                _nerr = int(r.get("n_errores", 0))
                _t_ok = int(r.get("tiempo_ok_count", 0))
                _t_ap = int(r.get("tiempo_aplica_count", 0))
                _c_ok = int(r.get("causa_ok_count", 0))
                _c_ap = int(r.get("causa_aplica_count", 0))
                _n_ok = int(r.get("numeral_ok_count", 0))
                _n_ap = int(r.get("numeral_aplica_count", 0))
                res.append(_pad([
                    r.get("tecnico", ""), _ots, _ots - _nerr, _nerr,
                    _frac(_t_ok, _t_ap),
                    _frac(_c_ok, _c_ap),
                    _frac(_n_ok, _n_ap),
                    f"{r.get('exactitud_pct', 0):.1f}%",
                ]))
        pd.DataFrame(res).to_excel(wr, sheet_name="Resumen", index=False, header=False)

        # ═══ Sheet 2: SLA ═══
        if not _sla.empty:
            _se = _sla.copy()
            _aviso_idx_xl = aviso_idx if aviso_idx is not None else (
                load_cotalker_index_fn() if load_cotalker_index_fn else {})
            if _aviso_idx_xl and "os_fracttal" in _se.columns:
                _se["n_llamado"] = (
                    _se["os_fracttal"]
                    .map(_aviso_idx_xl)
                    .apply(lambda v: str(v) if pd.notna(v) and str(v) not in ("", "nan") else "")
                    .where(lambda s: s != "", _se["n_llamado"].astype(str))
                )
            if "fecha_llamado_dt" in _se.columns:
                _se["Fecha atención"] = _se["fecha_llamado_dt"].dt.strftime("%d/%m/%Y").fillna("—")
                _se["Inicio SLA"] = _se["fecha_llamado_dt"].dt.strftime("%d/%m/%Y %H:%M").fillna("—")
            if "fecha_atencion" in _se.columns:
                _se["Fecha cierre OT Técnico"] = pd.to_datetime(
                    _se["fecha_atencion"], errors="coerce"
                ).dt.strftime("%d/%m/%Y %H:%M").fillna("—")
            if {"tiempo_resp_esp", "horas_resolucion"}.issubset(_se.columns):
                def _pct_dl(r):
                    h, u = r.get("horas_resolucion"), r.get("tiempo_resp_esp")
                    if pd.notna(h) and pd.notna(u) and float(u) > 0:
                        return round(float(h) / float(u) * 100, 1)
                    return None
                _se["_pct"] = _se.apply(_pct_dl, axis=1)
                _se["Uso SLA"] = _se["_pct"].apply(
                    lambda v: f"{min(v, 100):.1f}%" if pd.notna(v) else "—")
                _se["Exceso"] = _se["_pct"].apply(
                    lambda v: f"{max(v - 100, 0):.1f}%" if pd.notna(v) else "—")
            _se["Cumple SLA"] = _se["cumple_sla"].apply(
                lambda x: "Sí" if x is True else ("No" if x is False else "—"))
            _note_dl = {}
            if not df_wo.empty and "comentario_tecnico" in df_wo.columns:
                _note_dl = (
                    df_wo[df_wo["comentario_tecnico"].astype(str).str.strip() != ""]
                    .set_index("folio")["comentario_tecnico"].to_dict()
                )
            if "os_fracttal" in _se.columns and _note_dl:
                _se["Observación técnico"] = _se["os_fracttal"].apply(
                    lambda f: _strip_comentario_headers(
                        _note_dl.get(str(f).strip(), "—")))
            else:
                _se["Observación técnico"] = "—"
            _sla_cols = [c for c in [
                "os_fracttal", "n_llamado", "eds_occim", "equipo_label",
                "tecnico", "cliente", "eds_nombre",
                "Fecha atención", "Inicio SLA", "Fecha cierre OT Técnico",
                "prioridad", "zona_norm", "horas_resolucion",
                "Uso SLA", "Exceso", "Cumple SLA", "Observación técnico",
            ] if c in _se.columns]
            _se[_sla_cols].rename(columns={
                "os_fracttal": "N°Orden (OT)", "n_llamado": "N°Aviso",
                "eds_occim": "Cód. EDS", "equipo_label": "Equipo",
                "tecnico": "Técnico", "cliente": "Cliente",
                "eds_nombre": "Estación", "prioridad": "Prioridad",
                "zona_norm": "Zona", "horas_resolucion": "Horas resolución",
            }).to_excel(wr, sheet_name="SLA", index=False)
        else:
            pd.DataFrame({"Sin datos SLA": []}).to_excel(
                wr, sheet_name="SLA", index=False)

        # ═══ Sheet 3: Efectividad MP ═══
        if not _reinc.empty:
            _rc = _reinc.copy()
            for dc in ("fecha_pm", "fecha_cm"):
                if dc in _rc.columns:
                    _rc[dc] = pd.to_datetime(
                        _rc[dc], errors="coerce"
                    ).dt.strftime("%d/%m/%Y").fillna("—")
            _FT = {"fao": "F.A.O", "fnao": "F.N.A.O", "sin_info": "Sin info",
                   "especial": "Especial", "sin_dato": "Sin dato"}
            _CC = {"sin_causa": "Sin causa", "tecnico": "Técnico",
                   "cliente": "Cliente", "anulado": "Anulado por EDS"}
            if "falla_tipo" in _rc.columns:
                _rc["_ft"] = _rc["falla_tipo"].map(_FT).fillna(_rc["falla_tipo"])
                _CL = {"fao": "Falla atribuible", "fnao": "Falla no atribuible",
                       "sin_info": "Sin información", "especial": "Trabajo especial",
                       "sin_dato": "Sin dato"}
                _rc["_concl"] = _rc["falla_tipo"].map(_CL).fillna("—")
            if "causa_clasif" in _rc.columns:
                _rc["_cc"] = _rc["causa_clasif"].map(_CC).fillna(_rc["causa_clasif"])
            if "es_reincidencia_tecnico" in _rc.columns:
                _rc["_resp"] = _rc["es_reincidencia_tecnico"].apply(
                    lambda x: "Técnico" if x else "No técnico")
            if "grupo_responsable" in _rc.columns:
                _rc["_grp"] = _rc["grupo_responsable"].map(
                    _EQUIPO_LABEL).fillna(_rc["grupo_responsable"])
            _rc_map = [
                ("equipment", "Equipo"), ("client", "Cliente"),
                ("station", "Estación"), ("folio_pm", "OT MP"),
                ("folio_cm", "OT MC"), ("fecha_pm", "Fecha MP"),
                ("fecha_cm", "Fecha MC"),
                ("tecnico_resp_short", "Técnico MP"),
                ("tecnico_cm", "Técnico MC"), ("_grp", "T.Senior"),
                ("dias_entre", "Días MP→MC"),
                ("falla_raw", "Tipo Falla"), ("_ft", "Clasif. Falla"),
                ("causa_raiz", "Causa raíz"), ("_cc", "Clasif. causa"),
                ("_resp", "Responsabilidad"),
                ("comentario_tecnico_cm", "Comentario STO"),
                ("_concl", "Conclusión"),
            ]
            _rc_use = [(s, d) for s, d in _rc_map if s in _rc.columns]
            _rc[[s for s, _ in _rc_use]].rename(
                columns={s: d for s, d in _rc_use}
            ).to_excel(wr, sheet_name="Efectividad MP", index=False)
        else:
            pd.DataFrame({"Sin fallas post-preventiva": []}).to_excel(
                wr, sheet_name="Efectividad MP", index=False)

        # ═══ Sheet 4: P. Fracttal (4 secciones apiladas) ═══
        if not _df_ot_sc.empty:
            _pf = _df_ot_sc.copy()
            _eds_map_dl = _build_eds_nombre_map(df_eds)

            def _fmt_seg_dl(s):
                if pd.isna(s) or s == 0: return "—"
                s = int(s); h, m = s//3600, (s%3600)//60
                return f"{h:02d}:{m:02d}"

            _pf_cd = pd.to_datetime(_pf["creation_date"], errors="coerce")
            if _pf_cd.dt.tz is not None:
                _pf_cd = _pf_cd.dt.tz_convert(None)
            _pf["_fecha"] = _pf_cd.dt.strftime("%d/%m/%Y").fillna("—")

            _st = _pf["score_tiempo"].fillna(0) if "score_tiempo" in _pf.columns else pd.Series(0, index=_pf.index)
            _sc_c = _pf["score_causa"].fillna(0) if "score_causa" in _pf.columns else pd.Series(0, index=_pf.index)
            _sn = _pf["score_numeral"].fillna(0) if "score_numeral" in _pf.columns else pd.Series(0, index=_pf.index)
            _ok_bits = ((_st >= 25).astype(int) + (_sc_c >= 25).astype(int) + (_sn >= 25).astype(int))

            _pf["_cumple"] = _ok_bits.apply(lambda n: "✅ Cumple" if n == 3 else "❌ No cumple")
            _pf["_x3"] = _ok_bits.apply(lambda n: f"{n}/3")
            _pf["_tipo"] = _pf["maint_type"].fillna("").str.title() if "maint_type" in _pf.columns else "—"
            _pf["_modalidad"] = "—"
            if "deteccion_raw" in _pf.columns:
                _MOD_MAP_DL = {"1.- ATENDIDO PRESENCIAL": "Presencial",
                               "2.- ATENDIDO VÍA REMOTA": "Remoto", "2.- ATENDIDO VIA REMOTA": "Remoto",
                               "3.- ATENDIDO CON SU MP": "Con MP", "4.- LLAMADO DUPLICADO": "Duplicado"}
                _pf["_modalidad"] = _pf["deteccion_raw"].fillna("").apply(
                    lambda v: _MOD_MAP_DL.get(str(v).strip().upper(), str(v).strip().title() or "—"))
            _pf["_estado_ot"] = _pf["wo_status"].fillna("—") if "wo_status" in _pf.columns else "—"
            _em = _pf["elapsed_min"] if "elapsed_min" in _pf.columns else pd.Series(0.0, index=_pf.index)
            _pf["_col_tiempo"] = _pf.apply(
                lambda r: f"{'✅' if _st[r.name] >= 25 else '❌'} {_em[r.name]:.0f} min", axis=1)

            def _fmt_causa_dl(r):
                es_corr = r.get("es_correctiva", True)
                raw = str(r.get("causa_raiz_raw", "") or "").strip()
                ok = _sc_c[r.name] >= 25
                if not es_corr:
                    return "✅ Preventiva (no aplica)"
                if ok:
                    return f"✅ {raw[:38]}" if raw else "✅ Registrada"
                return f"❌ {raw[:35]}" if raw else "❌ Sin causa"

            _pf["_col_causa"] = _pf.apply(_fmt_causa_dl, axis=1)
            _eq_src = _pf["equipo_nombre"].fillna("") if "equipo_nombre" in _pf.columns else pd.Series("", index=_pf.index)
            _pf["_equipo"] = _eq_src.apply(lambda s: str(s).strip().title() if s else "—")

            def _fmt_num_dl(r):
                if not r.get("es_lavadora", True):
                    return "🔵 No aplica"
                if r.get("numeral_ok", False):
                    return "✅ Registrado"
                return NUMERAL_MOTIVO_LABEL.get(str(r.get("numeral_motivo", "") or ""), "❌ Dato inválido")

            _pf["_col_numeral"] = _pf.apply(_fmt_num_dl, axis=1)
            _nombres_comp_dl = {0: "Tiempo", 1: "Causa raíz", 2: "Numeral"}
            _scores_comp_dl = [_st, _sc_c, _sn]

            def _obs_dl(r):
                fallos = [_nombres_comp_dl[i] for i, sc in enumerate(_scores_comp_dl) if sc[r.name] < 25]
                if not fallos:
                    return "✅ Registro perfecto"
                return "⚠️ No cumple: " + ", ".join(fallos)

            _pf["_obs"] = _pf.apply(_obs_dl, axis=1)

            _det_completo = pd.DataFrame({
                "Fecha": _pf["_fecha"],
                "Cumple": _pf["_cumple"],
                "X/3": _pf["_x3"],
                "OT": _pf.get("folio", "—"),
                "EDS": _pf.get("eds_occim", "—"),
                "Técnico": _pf.get("tecnico", "—"),
                "Estación": _pf["station"].fillna("—").str.title() if "station" in _pf.columns else "—",
                "Tipo": _pf["_tipo"],
                "Modalidad": _pf["_modalidad"],
                "Estado": _pf["_estado_ot"],
                "Score": _pf["score_total"],
                "⏱ Tiempo": _pf["_col_tiempo"],
                "🔍 Causa raíz": _pf["_col_causa"],
                "🔧 Equipo": _pf["_equipo"],
                "🔢 Numeral": _pf["_col_numeral"],
                "💬 Observación": _pf["_obs"],
            })

            _cr = _pf[_pf["es_correctiva"].fillna(False).astype(bool)].copy() if "es_correctiva" in _pf.columns else pd.DataFrame()
            _det_causa = pd.DataFrame()
            if not _cr.empty:
                _cr["_estado_cr"] = "❌ No cumple"
                if "causa_ok" in _cr.columns:
                    _cr["_estado_cr"] = _cr["causa_ok"].fillna(False).apply(
                        lambda v: "✅ Cumple" if v else "❌ No cumple")
                if "eds_occim" in _cr.columns:
                    _cr["_eds_nombre"] = _cr["eds_occim"].map(_eds_map_dl).fillna("—").str.title()
                else:
                    _cr["_eds_nombre"] = "—"
                if "comentario_tecnico" in _cr.columns:
                    _cr["_comentario_cr"] = _cr["comentario_tecnico"].fillna("").apply(_strip_comentario_headers)
                else:
                    _cr["_comentario_cr"] = "—"
                _det_causa = pd.DataFrame({
                    "Fecha": _cr["_fecha"],
                    "Equipo": _cr["equipment_code"].fillna("—") if "equipment_code" in _cr.columns else "—",
                    "EDS": _cr.get("eds_occim", "—"),
                    "OT": _cr.get("folio", "—"),
                    "Tipo": _cr["maint_type"].fillna("").str.title() if "maint_type" in _cr.columns else "—",
                    "Técnico": _cr.get("tecnico", "—"),
                    "Nombre EDS": _cr["_eds_nombre"],
                    "Causa Raíz": _cr["causa_raiz_raw"].fillna("—") if "causa_raiz_raw" in _cr.columns else "—",
                    "Clasificación": _cr["causa_clasif"].fillna("—") if "causa_clasif" in _cr.columns else "—",
                    "Estado": _cr["_estado_cr"],
                    "Comentario técnico / qué hizo": _cr["_comentario_cr"],
                })

            _te = _pf[_pf["maint_type"].str.upper().str.contains("PREVENTIVA", na=False)].copy() if "maint_type" in _pf.columns else pd.DataFrame()
            _det_tiempo = pd.DataFrame()
            _te_estim_col = "estimated_sec" if "estimated_sec" in _te.columns else ("estim_sec_sum" if not _te.empty and "estim_sec_sum" in _te.columns else None)
            if not _te.empty and _te_estim_col:
                if _te_estim_col != "estimated_sec":
                    _te["estimated_sec"] = _te[_te_estim_col]
                if "duration_sec" not in _te.columns and "exec_sec_sum" in _te.columns:
                    _te["duration_sec"] = _te["exec_sec_sum"]
                if "elapsed_sec" not in _te.columns and "max_elapsed" in _te.columns:
                    _te["elapsed_sec"] = _te["max_elapsed"]
                _te = _te[_te["estimated_sec"].fillna(0) > 0].copy()
                if not _te.empty:
                    _te["_effective_sec"] = _te[["duration_sec", "elapsed_sec"]].fillna(0).max(axis=1)
                    _te["_piso_pct"] = _te["client"].apply(
                        lambda c: 0.50 if c == "SHELL (Enex)" else 0.70)
                    _te["_te_ok"] = _te["_effective_sec"] >= _te["estimated_sec"] * _te["_piso_pct"]
                    _te["_te_sobre"] = _te["_effective_sec"] > _te["estimated_sec"] * 1.50
                    _te = _te.drop_duplicates(subset="folio", keep="first")
                    _te["_minimo_sec"] = (_te["estimated_sec"] * _te["_piso_pct"]).round(0)
                    _te["_maximo_sec"] = (_te["estimated_sec"] * 1.50).round(0)
                    _te["_pct_ej"] = (_te["_effective_sec"] / _te["estimated_sec"] * 100).round(1)

                    def _estado_te_dl(r):
                        if bool(r.get("_te_ok", False)):
                            if bool(r.get("_te_sobre", False)):
                                return "✅ Cumple · Sobretiempo"
                            return "✅ Cumple"
                        return "❌ No cumple"

                    _te["_estado_te"] = _te.apply(_estado_te_dl, axis=1)
                    if "eds_occim" in _te.columns:
                        _te["_eds_nombre_te"] = _te["eds_occim"].map(_eds_map_dl).fillna("—").str.title()
                    else:
                        _te["_eds_nombre_te"] = "—"

                    def _diag_te_dl(r):
                        estim = int(r.get("estimated_sec") or 0)
                        real = int(r.get("_effective_sec") or 0)
                        minimo = int(r.get("_minimo_sec") or 0)
                        pct = float(r.get("_pct_ej") or 0.0)

                        def _hm(s):
                            if s <= 0: return "0min"
                            h, m = int(s)//3600, (int(s)%3600)//60
                            return f"{h}h{m:02d}min" if h else f"{m}min"

                        if bool(r.get("_te_ok", False)):
                            return ""
                        if real == 0:
                            return f"Sin registro de tiempo. Estimado {_hm(estim)}, mínimo {_hm(minimo)}."
                        if pct < 20:
                            return f"Tiempo insuficiente: {_hm(real)} vs mínimo {_hm(minimo)}."
                        return f"Por debajo del mínimo: {_hm(real)} vs {_hm(minimo)} requerido."

                    _te["_diag"] = _te.apply(_diag_te_dl, axis=1)
                    _det_tiempo = pd.DataFrame({
                        "Fecha": _te["_fecha"],
                        "Estado": _te["_estado_te"],
                        "EDS": _te.get("eds_occim", "—"),
                        "OT": _te.get("folio", "—"),
                        "Tipo": _te["maint_type"].fillna("").str.title() if "maint_type" in _te.columns else "—",
                        "Técnico": _te.get("tecnico", "—"),
                        "Nombre EDS": _te["_eds_nombre_te"],
                        "T. Estimado": _te["estimated_sec"].apply(_fmt_seg_dl),
                        "T. Mínimo": _te["_minimo_sec"].apply(_fmt_seg_dl),
                        "Máx. 150%": _te["_maximo_sec"].apply(_fmt_seg_dl),
                        "T. Ejecución": _te["_effective_sec"].apply(_fmt_seg_dl),
                        "% Ejecutado": _te["_pct_ej"],
                        "Diagnóstico": _te["_diag"],
                    })

            _det_numeral = pd.DataFrame()
            if not _pf.empty:
                _nm = _pf.copy()

                def _est_num_dl(r):
                    if not r.get("es_lavadora", True):
                        return "🔵 No aplica"
                    if r.get("numeral_ok", False):
                        return "✅ Cumple"
                    return NUMERAL_MOTIVO_LABEL.get(str(r.get("numeral_motivo", "") or ""), "❌ Dato inválido")

                _nm["_estado_num"] = _nm.apply(_est_num_dl, axis=1)
                _nm["_n_ini"] = _nm.apply(lambda r: str(r.get("numeral_inicial", "") or "").strip() or "—", axis=1)
                _nm["_n_fin"] = _nm.apply(lambda r: str(r.get("numeral_final", "") or "").strip() or "—", axis=1)

                def _fichas_dl(r):
                    fp = r.get("fichas_periodo", None)
                    if fp is None or (isinstance(fp, float) and pd.isna(fp)):
                        return "—"
                    try:
                        return f"{int(fp):,}"
                    except (ValueError, TypeError):
                        return "—"

                _nm["_fichas"] = _nm.apply(_fichas_dl, axis=1)
                if "comentario_tecnico" in _nm.columns:
                    _nm["_com_num"] = _nm["comentario_tecnico"].fillna("").apply(_strip_comentario_headers)
                else:
                    _nm["_com_num"] = "—"
                _det_numeral = pd.DataFrame({
                    "Fecha": _nm["_fecha"],
                    "Equipo": _nm["equipment_code"].fillna("—") if "equipment_code" in _nm.columns else "—",
                    "Cliente": _nm["client"].fillna("—").str.title() if "client" in _nm.columns else "—",
                    "OT": _nm.get("folio", "—"),
                    "Tipo": _nm["maint_type"].fillna("").str.title() if "maint_type" in _nm.columns else "—",
                    "Técnico": _nm.get("tecnico", "—"),
                    "EDS": _nm.get("eds_occim", "—"),
                    "N. Inicial": _nm["_n_ini"],
                    "N. Final": _nm["_n_fin"],
                    "Fichas período": _nm["_fichas"],
                    "Estado": _nm["_estado_num"],
                    "Comentario técnico / causa raíz": _nm["_com_num"],
                })

            _ws_name = "P. Fracttal"
            _row_cursor = 0

            _sec1_title = pd.DataFrame([["DETALLE COMPLETO"] + [""] * (_det_completo.shape[1] - 1)])
            _sec1_title.to_excel(wr, sheet_name=_ws_name, index=False, header=False, startrow=_row_cursor)
            _row_cursor += 1
            _det_completo.to_excel(wr, sheet_name=_ws_name, index=False, startrow=_row_cursor)
            _row_cursor += len(_det_completo) + 1 + 1

            if not _det_causa.empty:
                _sec2_title = pd.DataFrame([["RESUMEN CAUSA RAIZ"] + [""] * (_det_causa.shape[1] - 1)])
                _sec2_title.to_excel(wr, sheet_name=_ws_name, index=False, header=False, startrow=_row_cursor)
                _row_cursor += 1
                _det_causa.to_excel(wr, sheet_name=_ws_name, index=False, startrow=_row_cursor)
                _row_cursor += len(_det_causa) + 1 + 1

            if not _det_tiempo.empty:
                _sec3_title = pd.DataFrame([["RESUMEN TIEMPO"] + [""] * (_det_tiempo.shape[1] - 1)])
                _sec3_title.to_excel(wr, sheet_name=_ws_name, index=False, header=False, startrow=_row_cursor)
                _row_cursor += 1
                _det_tiempo.to_excel(wr, sheet_name=_ws_name, index=False, startrow=_row_cursor)
                _row_cursor += len(_det_tiempo) + 1 + 1

            if not _det_numeral.empty:
                _sec4_title = pd.DataFrame([["RESUMEN NUMERALES"] + [""] * (_det_numeral.shape[1] - 1)])
                _sec4_title.to_excel(wr, sheet_name=_ws_name, index=False, header=False, startrow=_row_cursor)
                _row_cursor += 1
                _det_numeral.to_excel(wr, sheet_name=_ws_name, index=False, startrow=_row_cursor)
        else:
            pd.DataFrame({"Sin OTs evaluadas": []}).to_excel(
                wr, sheet_name="P. Fracttal", index=False)

    # ═══ Post-proceso: estilar encabezados de todas las hojas ═══
    buf.seek(0)
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _wb = load_workbook(buf)

    _HDR_FILL  = PatternFill("solid", fgColor="1F4E78")
    _HDR_FONT  = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    _HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _thin = Side(style="thin", color="D9D9D9")
    _BORDER    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _SECTION_FILL = PatternFill("solid", fgColor="F2F2F2")
    _SECTION_FONT = Font(name="Arial", size=11, bold=True, color="1F4E78")

    # Palabras que identifican filas de "sección" (encabezados internos apilados en P. Fracttal)
    _SECTION_LABELS = {
        "DESEMPEÑO SLA", "EFECTIVIDAD MP", "PRECISIÓN FRACTTAL",
        "RESUMEN TIEMPO", "RESUMEN CAUSA RAÍZ", "RESUMEN NUMERALES",
        "PRECISION FRACTTAL", "RESUMEN CAUSA RAIZ",
        "DETALLE COMPLETO",
    }

    def _es_fila_seccion(ws, row_idx: int) -> bool:
        v = ws.cell(row=row_idx, column=1).value
        if not isinstance(v, str):
            return False
        return v.strip().upper() in _SECTION_LABELS

    for _ws in _wb.worksheets:
        if _ws.max_row < 1:
            continue
        # Estilar cada fila que sea encabezado real de tabla (fila 1) o sección apilada
        # Heurística: fila 1 SIEMPRE es encabezado; luego cualquier fila que empiece
        # con un texto de _SECTION_LABELS marca inicio de sección — la fila SIGUIENTE
        # (row+1) es su encabezado de tabla.
        _header_rows = {1}
        for r in range(1, _ws.max_row + 1):
            if _es_fila_seccion(_ws, r):
                # La fila de sección misma se estila diferente (banda gris + texto azul)
                for c in range(1, _ws.max_column + 1):
                    cell = _ws.cell(row=r, column=c)
                    cell.fill = _SECTION_FILL
                    cell.font = _SECTION_FONT
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                _ws.row_dimensions[r].height = 22
                # La fila siguiente es el encabezado de esa sub-tabla
                if r + 1 <= _ws.max_row:
                    _header_rows.add(r + 1)

        for hr in _header_rows:
            for c in range(1, _ws.max_column + 1):
                cell = _ws.cell(row=hr, column=c)
                if cell.value is None or cell.value == "":
                    continue
                cell.fill = _HDR_FILL
                cell.font = _HDR_FONT
                cell.alignment = _HDR_ALIGN
                cell.border = _BORDER
            _ws.row_dimensions[hr].height = 32

        # Congelar primera fila (solo si es encabezado real de la hoja)
        _ws.freeze_panes = "A2"

    _out = BytesIO()
    _wb.save(_out)
    _out.seek(0)
    return _out.getvalue()
