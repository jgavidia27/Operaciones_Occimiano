"""
gdrive.py — Lee archivos de Google Drive (sincronizado en G:) para el Dashboard Occimiano.
Lee en modo lectura únicamente. NO modifica ningún archivo.
"""
import os
import re
import warnings
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles.fills import PatternFill

warnings.filterwarnings("ignore", category=UserWarning)

# ── Rutas base ────────────────────────────────────────────────────────────────
_DRIVE_ROOT = "G:/.shortcut-targets-by-id/15zHnoU5VZlkOwYc6EBziNcnS-sAAtwtk/OPERACIONES/OPERACIONES"
_SLA = f"{_DRIVE_ROOT}/SLA OPERACIONES"
_KPI_ARAMCO = "G:/Unidades compartidas/OPERACIONES/KPI/Aramco"

FILES = {
    "listado_eds":     f"{_DRIVE_ROOT}/Listado de EDS.xlsx",
    "llamados_copec":  f"{_SLA}/Llamados Correctivos COPEC 2024 V2.0.xlsx",   # archivo maestro COPEC (abr 2024 → hoy)
    # llamados_esmax: path resuelto dinámicamente por _find_latest_esmax()
    "llamados_shell":  f"{_SLA}/Llamados correctivos Shell.xlsx",
    "utilizacion":     f"{_DRIVE_ROOT}/2026 UTILIZACIÓN DE TIEMPO.xlsx",
    # base_tecnicos: primero busca en Google Drive, si no existe cae al Desktop
    "base_tecnicos":   f"{_DRIVE_ROOT}/base tecnicos.xlsx",
    "base_tecnicos_fallback": r"C:\Users\jgavi\Desktop\base tecnicos.xlsx",
}


def _find_latest_esmax() -> str:
    """
    Retorna la ruta del archivo ESMAX más reciente en la carpeta SLA OPERACIONES.
    Ignora archivos temporales (prefijo ~$). Si no encuentra ninguno, retorna "".
    Así el dashboard siempre usa el archivo activo sin importar el año en el nombre.
    """
    import glob
    candidates = [
        f for f in glob.glob(os.path.join(_SLA, "*ESMAX*.xlsx"))
        if not os.path.basename(f).startswith("~$")
    ]
    if not candidates:
        return ""
    return max(candidates, key=os.path.getmtime)

# ── Utilización del tiempo — configuración ────────────────────────────────────
_MESES_ES = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO",  6: "JUNIO",   7: "JULIO", 8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
}

# 8 categorías — nombres tomados directamente del vocabulario del Excel
CATEGORY_COLORS_UTIL = {
    "Mant. Preventivo":   "#22c55e",   # MP xxx / PULSE xxx
    "Llamado Correctivo": "#ef4444",   # MC xxx / LLAMADOS
    "Instalación":        "#3b82f6",   # INSTALACION / INTALACION / INSPECCION
    "Capacitación":       "#a855f7",   # CAP. / CAPACITACION / BOOK / CUMPLE (reunión masiva)
    "Reunión":            "#f59e0b",   # REU / REUNION
    "Feriado":            "#94a3b8",   # FERIADO / IRRENUNCIABLE / VACACIONES
    "Inventario":         "#06b6d4",   # INVENTARIO / ENT. FICHAS
    "Oficina":            "#64748b",   # OFICINA / REV. TECNICA / PREP. / resto
}

# Prioridad dominante (mayor índice = gana si hay empate en el día)
_CAT_PRIORITY = [
    "Oficina", "Inventario", "Reunión", "Capacitación",
    "Llamado Correctivo", "Mant. Preventivo", "Instalación", "Feriado",
]


def classify_task_line(line: str) -> str:
    """Clasifica una línea de tarea en una de las 8 categorías reconocibles del Excel."""
    t = line.strip().upper()
    if not t:
        return None

    # Ignorar líneas que son artefactos de celdas fusionadas (solo números o abrev.)
    if re.match(r"^[\d]+$", t) or t in ("A.M.", "/", ""):
        return None

    # ── Feriado / No laboral ───────────────────────────────────────────────
    if any(k in t for k in ["FERIADO", "IRRENUNCIABLE", "VACACION", "LICENCIA"]):
        return "Feriado"

    # ── Mant. Preventivo — MP xxx y PULSE xxx (trabajo de terreno preventivo) ─
    if re.match(r"^MP\b", t) or re.match(r"^PULSE\b", t):
        return "Mant. Preventivo"

    # ── Llamado Correctivo — MC xxx (intervención concreta) ───────────────
    if re.match(r"^MC\b", t) or re.match(r"^MC\+", t) or t == "LLAMADOS":
        return "Llamado Correctivo"

    # ── Instalación / Desinstalación / Inspección ─────────────────────────
    if any(k in t for k in ["INSTALACION", "INTALACION", "INSTAL", "DESINTALAR",
                             "DESINST", "INSPECCION", "INSPECCIÓN"]):
        return "Instalación"

    # ── Capacitación — reuniones masivas tipo CUMPLE./BOOK/CAP. ──────────
    if any(k in t for k in ["CAP.", "CAPACITACION", "BOOK"]):
        return "Capacitación"
    # "09:00 CUMPLE" = reunión de cumplimiento/compliance (evento de todo el equipo)
    if re.match(r"^\d{1,2}:\d{2}\s*CUMPLE", t):
        return "Capacitación"

    # ── Reunión ───────────────────────────────────────────────────────────
    if re.match(r"^REU\b", t) or "REUNION" in t or "REUNIÓN" in t:
        return "Reunión"

    # ── Inventario ────────────────────────────────────────────────────────
    if any(k in t for k in ["INVENTARIO", "REP. STOCK", "STOCK", "ENT. FICHAS",
                             "ENT. EQUIPO", "FICHAS", "BODEGA"]):
        return "Inventario"

    # ── Oficina (todo lo que queda: OFICINA, REV. TECNICA, PREP., etc.) ──
    return "Oficina"


def _dominant_category(lines: list) -> str:
    """Retorna la categoría dominante de una lista de líneas de tareas."""
    cats = {classify_task_line(l) for l in lines if classify_task_line(l)}
    cats.discard(None)
    if not cats:
        return "Oficina"
    for p in reversed(_CAT_PRIORITY):
        if p in cats:
            return p
    return "Oficina"

# SLA por cliente, prioridad y zona (horas) — Fuente: tabla SLA para Clientes
# Confirmado por operaciones 2026-05-29
SLA_HOURS: dict = {
    "COPEC": {
        "P1": {"Santiago": 18,  "Regiones": 24},
        "P2": {"Santiago": 24,  "Regiones": 48},
        "P3": {"Santiago": 36,  "Regiones": 72},
        "P4": {"Santiago": 96,  "Regiones": 96},
    },
    "ESMAX (Aramco)": {
        "P1": {"Santiago": 24,  "Regiones": 24},
        "P2": {"Santiago": 48,  "Regiones": 48},
        "P3": {"Santiago": 72,  "Regiones": 72},
        "P4": {"Santiago": 100, "Regiones": 100},
    },
    "SHELL (Enex)": {
        "P1": {"Santiago": 24, "Regiones": 24},
        "P2": {"Santiago": 48, "Regiones": 48},
        "P3": {"Santiago": 72, "Regiones": 72},
        "P4": {"Santiago": 96, "Regiones": 96},  # actualizado según documento SLA may-2026
    },
}
# Fallback genérico si el cliente no está en el catálogo
SLA_DEFAULT = {
    "P1": {"Santiago": 24,  "Regiones": 24},
    "P2": {"Santiago": 48,  "Regiones": 48},
    "P3": {"Santiago": 72,  "Regiones": 72},
    "P4": {"Santiago": 100, "Regiones": 100},
}
_SLA_DEFAULT = SLA_DEFAULT  # alias interno (compatibilidad)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _is_red_row(ws, row_idx: int) -> bool:
    """Check if any cell in a row has a red/orange fill (inactive EDS marker)."""
    for cell in ws[row_idx]:
        fill = cell.fill
        if fill and fill.fgColor:
            rgb = fill.fgColor.rgb
            if isinstance(rgb, str) and len(rgb) == 8:
                r = int(rgb[2:4], 16)
                g = int(rgb[4:6], 16)
                b = int(rgb[6:8], 16)
                # Red: r > 180, g < 100, b < 100
                if r > 180 and g < 120 and b < 120:
                    return True
    return False


def _read_llamados(path: str, sheet: str, header_row: int, date_col: str = "Fecha") -> pd.DataFrame:
    """Read a llamados correctivos sheet with the correct header row."""
    if not os.path.exists(path):
        return pd.DataFrame()
    df = pd.read_excel(path, sheet_name=sheet, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    if date_col in df.columns:
        df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
    return df


# ── Public loaders ────────────────────────────────────────────────────────────

@st.cache_data(ttl=1800, show_spinner=False, persist="disk")
def load_listado_eds() -> pd.DataFrame:
    """
    Returns the Listado de EDS with active/inactive flag.
    Uses openpyxl to detect red-highlighted rows (inactive stations).
    """
    path = FILES["listado_eds"]
    if not os.path.exists(path):
        return pd.DataFrame()

    # Step 1: detect red rows via openpyxl
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb["EDS"]
    red_rows = set()
    for row in ws.iter_rows(min_row=2):  # row 1 = header
        row_idx = row[0].row
        if _is_red_row(ws, row_idx):
            red_rows.add(row_idx - 2)  # convert to 0-indexed pandas row
    wb.close()

    # Step 2: read with pandas
    df = pd.read_excel(path, sheet_name="EDS")

    # Keep the useful columns
    col_map = {
        "EDS":             "eds_occim",
        # Keep "EDS OCCIM" as a separate column — it holds the Fracttal-internal code
        # (EE_S038 for ESMAX, SH_36 for Shell) needed to join llamados.
        # It MUST NOT duplicate "eds_occim" or it gets dropped by deduplication.
        "EDS OCCIM":       "eds_occim_raw",
        "EDS CLIENTE":     "eds_cliente",
        "Dirección ":      "direccion",
        "DIRECCIÓN":       "direccion",
        "COMUNA":          "comuna",
        "COMUNA.1":        "comuna",
        "Facturación":     "cliente",
        "CLIENTE":         "cliente",
        "Zona":            "zona_occim",
        "ZONA DE CLIENTE": "zona_cliente",
        "Zona de Atención":"zona_atencion",
        "ZONA DE ATENCIÓN":"zona_atencion",
        "REGIÓN ":         "region",
    }
    df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})

    # Deduplicate columns after rename
    df = df.loc[:, ~df.columns.duplicated()]

    # Add active flag
    df["activa"] = True
    df.loc[list(red_rows), "activa"] = False

    # Normalize client names (PETROBRAS = ESMAX/Aramco historically)
    if "cliente" in df.columns:
        df["cliente"] = df["cliente"].astype(str).str.strip().str.upper().replace({
            "PETROBRAS": "ESMAX (Aramco)",
            "ESMAX": "ESMAX (Aramco)",
            "ARAMCO": "ESMAX (Aramco)",
            "SHELL": "SHELL (Enex)",
            "ENEX": "SHELL (Enex)",
            "ABASTIBLE": "ABASTIBLE",
            "COPEC": "COPEC",
        })

    # Keep only rows with a valid EDS number
    if "eds_occim" in df.columns:
        df = df[df["eds_occim"].notna()]
        df["eds_occim"] = df["eds_occim"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)

    return df


@st.cache_data(ttl=1800, show_spinner=False, persist="disk")
def _file_mtime(path: str) -> int:
    """Retorna mtime entero del archivo, o 0 si no existe. Usado como clave de caché."""
    try:
        return int(os.path.getmtime(path))
    except OSError:
        return 0


@st.cache_data(ttl=1800, show_spinner=False, persist="disk")
def load_llamados_copec(desde: str = "2025-01-01", _mtime: int = 0) -> pd.DataFrame:
    """
    Carga llamados COPEC. _mtime se pasa desde app.py con la fecha de modificación
    del archivo — esto invalida el caché automáticamente cuando el Excel cambia.
    """
    path = FILES["llamados_copec"]
    if not os.path.exists(path):
        return pd.DataFrame()
    df = _read_llamados(path, sheet="LLAMADOS DD", header_row=3)
    return _enrich_llamados(df, "COPEC", desde)


@st.cache_data(ttl=1800, show_spinner=False, persist="disk")
def load_llamados_esmax(desde: str = "2025-01-01", _mtime: int = 0) -> pd.DataFrame:
    path = _find_latest_esmax()
    if not path or not os.path.exists(path):
        return pd.DataFrame()
    df = _read_llamados(path, sheet="LLAMADOS DD", header_row=3)
    return _enrich_llamados(df, "ESMAX (Aramco)", desde)


@st.cache_data(ttl=1800, show_spinner=False, persist="disk")
def load_llamados_shell(desde: str = "2025-01-01", _mtime: int = 0) -> pd.DataFrame:
    path = FILES["llamados_shell"]
    if not os.path.exists(path):
        return pd.DataFrame()
    df = _read_llamados(path, sheet="Detalle", header_row=1)
    if "File" in df.columns and "Codigo EDS" not in df.columns:
        df = df.rename(columns={"File": "Codigo EDS"})
    return _enrich_llamados(df, "SHELL (Enex)", desde)


@st.cache_data(ttl=3600 * 8, show_spinner=False, persist="disk")
def load_llamados_fracttal(
    clientes: tuple = ("COPEC", "SHELL (Enex)"),
    desde: str = "2026-01-01",
) -> pd.DataFrame:
    """
    Carga llamados correctivos de emergencia (COPEC + SHELL) directamente desde Fracttal.

    Estrategia de identificación (dos fuentes combinadas, sin duplicados):

    FUENTE PRINCIPAL — Solicitudes de Trabajo (/api/work_requests/):
      Cada llamado de emergencia entra a Fracttal como una Solicitud de Trabajo
      con types_description = "LLAMADO DE EMERGENCIA" (o similar).
      La solicitud queda vinculada a la OT correctiva mediante wo_folio.
      → T₀ = date_incident de la solicitud (o date como fallback)
      → folio → OT para obtener técnico, final_date, prioridad

    FUENTE SECUNDARIA — Work Orders (/api/work_orders/):
      Fallback para OTs correctivas que no tienen solicitud previa o cuyo
      description contiene "LLAMADO" (creadas directamente sin solicitud).

    Lógica de SLA:
    - T₀     = fecha_incidente de la solicitud > creation_date de la OT
    - T_final = final_date del técnico (no fecha administrativa de cierre)
    - Reloj 24/7 corrido, sin pausas ni festivos
    - Zona     = derivada del catálogo EDS (RM/Santiago vs Regiones)
    - Prioridad: VERY_HIGH→P1 · HIGH→P2 · MEDIUM→P3 · LOW/VERY_LOW→P4
    """
    from api import load_work_orders, load_work_requests as _load_wr
    from data import _parse_hierarchy

    raw_wo   = load_work_orders()
    raw_req  = _load_wr(months_back=12)
    desde_dt = pd.Timestamp(desde, tz="UTC")

    # ── Índice work_orders por folio (lookup O(1)) ────────────────────────────
    wo_idx: dict = {str(r.get("wo_folio", "")): r for r in raw_wo if r.get("wo_folio")}

    # ── Catálogo EDS ──────────────────────────────────────────────────────────
    try:
        df_eds_cat = load_listado_eds()
        if not df_eds_cat.empty and "eds_occim" in df_eds_cat.columns:
            _eds_zona   = df_eds_cat.set_index("eds_occim").get("region",   pd.Series(dtype=str)).to_dict()
            _eds_nombre = df_eds_cat.set_index("eds_occim").get("direccion", pd.Series(dtype=str)).to_dict()
            _eds_comuna = df_eds_cat.set_index("eds_occim").get("comuna",   pd.Series(dtype=str)).to_dict()
        else:
            _eds_zona = _eds_nombre = _eds_comuna = {}
    except Exception:
        _eds_zona = _eds_nombre = _eds_comuna = {}

    def _zona(eds_code: str) -> str:
        region = str(_eds_zona.get(eds_code, "")).upper()
        if any(k in region for k in ["METROPOLITANA", "SANTIAGO", "RM"]):
            return "Santiago"
        return "Regiones"

    def _map_prio(p: str):
        p = str(p or "").upper().strip().replace(" ", "_")
        if "VERY_HIGH" in p or "CRITIC" in p: return "P1"
        if p == "HIGH":                        return "P2"
        if "MEDIUM" in p:                      return "P3"
        if "LOW" in p:                         return "P4"
        return None

    # ── FUENTE 1: Solicitudes de Trabajo con tipo "LLAMADO" ───────────────────
    # Estas son las más confiables: el campo types_description identifica
    # explícitamente los llamados de emergencia.
    llamado_folios: dict = {}   # folio → t0_raw
    for r in raw_req:
        tipo = str(r.get("types_description") or "").upper()
        if not any(kw in tipo for kw in ["LLAMADO", "EMERGENCIA"]):
            continue
        folio = str(r.get("wo_folio") or "").strip()
        if not folio:
            continue
        t0_val = r.get("date_incident") or r.get("date")
        if folio not in llamado_folios:
            llamado_folios[folio] = t0_val

    # ── FUENTE 2: Work Orders con "LLAMADO" en description (fallback) ─────────
    # Cubre OTs creadas directamente en Fracttal sin solicitud previa.
    for r in raw_wo:
        if "LLAMADO" in str(r.get("description") or "").upper():
            folio = str(r.get("wo_folio") or "").strip()
            if folio and folio not in llamado_folios:
                llamado_folios[folio] = None   # T0 vendrá de creation_date

    if not llamado_folios:
        return pd.DataFrame()

    # ── Construir filas ───────────────────────────────────────────────────────
    rows = []
    for folio, t0_req in llamado_folios.items():
        r = wo_idx.get(folio)
        if not r:
            continue   # solicitud sin OT vinculada todavía

        # Filtro de fecha (por creation_date de la OT)
        creation = pd.to_datetime(r.get("creation_date"), utc=True, errors="coerce")
        if pd.isna(creation) or creation < desde_dt:
            continue

        # Filtro de cliente
        client, station = _parse_hierarchy(r.get("parent_description") or "")
        if client not in clientes:
            continue

        # Prioridad (desde la OT)
        prio = _map_prio(r.get("priorities_description", ""))
        if not prio:
            continue

        eds_code = str(r.get("groups_2_description") or "").strip()
        zona     = _zona(eds_code)

        # T₀: solicitud > creation_date OT
        t0_raw  = t0_req or r.get("creation_date")
        t0      = pd.to_datetime(t0_raw, utc=True, errors="coerce")
        t_final = pd.to_datetime(
            r.get("final_date") or r.get("date_maintenance"), utc=True, errors="coerce"
        )

        # Horas de resolución y cumplimiento SLA
        if pd.notna(t0) and pd.notna(t_final):
            horas_res   = max(0.0, (t_final - t0).total_seconds() / 3600)
            tiempo_real = t_final - t0
        else:
            horas_res   = None
            tiempo_real = None

        umbral = SLA_HOURS.get(client, SLA_DEFAULT).get(prio, {}).get(zona, None)
        if horas_res is not None and umbral is not None:
            cumplimiento = "CUMPLE" if horas_res <= umbral else "NO CUMPLE"
        else:
            cumplimiento = "SIN DATOS"

        t0_local     = t0.tz_convert(None)     if pd.notna(t0)     else None
        tfinal_local = t_final.tz_convert(None) if pd.notna(t_final) else None

        wo_close_raw   = r.get("wo_final_date")
        wo_close_local = (
            pd.to_datetime(wo_close_raw, utc=True, errors="coerce").tz_convert(None)
            if wo_close_raw else None
        )

        rows.append({
            "os_fracttal":      folio,
            "n_llamado":        r.get("id_work_order", ""),
            "fecha_llamado":    t0_local,
            "fecha_atencion":   tfinal_local,
            "wo_cierre_ot":     wo_close_local,
            "hora_llamado":     t0_local.time()     if t0_local     is not None else None,
            "hora_fin":         tfinal_local.time() if tfinal_local is not None else None,
            "eds_occim":        eds_code,
            "eds_nombre":       _eds_nombre.get(eds_code) or str(r.get("items_log_description") or station),
            "commune":          _eds_comuna.get(eds_code, ""),
            "cliente":          client,
            "tecnico":          str(r.get("personnel_description") or "").strip(),
            "prioridad":        prio,
            "zona":             zona,
            "cumplimiento":     cumplimiento,
            "tiempo_resp_real": tiempo_real,
            "estado_atencion":  "ATENDIDO",
            "facturacion":      client,
        })

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df["fecha_llamado"]  = pd.to_datetime(df["fecha_llamado"],  errors="coerce")
    df["fecha_atencion"] = pd.to_datetime(df["fecha_atencion"], errors="coerce")
    df["cliente"]        = df["cliente"].replace({"ENEX": "SHELL (Enex)"})
    return df


@st.cache_data(ttl=1800, show_spinner=False, persist="disk")
def load_all_llamados(desde: str = "2026-01-01") -> pd.DataFrame:
    """
    Carga llamados correctivos de los 3 clientes desde sus Excel en Google Drive.
    Los 3 clientes usan Excel porque los work_requests de Fracttal no tienen
    el tipo 'LLAMADO DE EMERGENCIA' de forma consistente, lo que causaba registros
    faltantes cuando se leía desde la API.

    Fuentes:
    - COPEC       → Excel "Llamados Correctivos COPEC 2024 V2.0.xlsx"
    - SHELL (Enex)→ Excel "Llamados correctivos Shell.xlsx"
    - ESMAX       → Excel dinámico (último archivo ESMAX/Aramco en Google Drive)
    """
    frames = []

    # COPEC desde Excel
    _copec_path = FILES.get("llamados_copec", "")
    df_copec = load_llamados_copec(desde, _mtime=_file_mtime(_copec_path))
    if not df_copec.empty:
        frames.append(df_copec)

    # SHELL desde Excel
    _shell_path = FILES.get("llamados_shell", "")
    df_shell = load_llamados_shell(desde, _mtime=_file_mtime(_shell_path))
    if not df_shell.empty:
        frames.append(df_shell)

    # ESMAX desde Excel
    _esmax_path = _find_latest_esmax()
    df_esmax = load_llamados_esmax(desde, _mtime=_file_mtime(_esmax_path))
    if not df_esmax.empty:
        frames.append(df_esmax)

    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


# ── Private enrichment ────────────────────────────────────────────────────────

def _enrich_llamados(df: pd.DataFrame, client_label: str, desde: str) -> pd.DataFrame:
    if df.empty:
        return df

    # Normalize column names
    df.columns = [str(c).strip() for c in df.columns]

    # Standardize key column names
    rename = {
        "Codigo EDS": "eds_occim",
        "Codigo Eds": "eds_occim",
        "EDS ": "eds_nombre",
        "EDS": "eds_nombre",
        "Mecanico": "tecnico",
        "Técnico": "tecnico",
        "Fecha": "fecha_llamado",
        "Fecha de atencion": "fecha_atencion",
        "Hora": "hora_llamado",
        "Hora de Atencion(FIN)": "hora_fin",
        "P1-P2-P3-P4": "prioridad",
        "PRIORIDAD": "prioridad",
        "STATUS CUMPLIMIENTO": "cumplimiento",
        "Status Cumplimiento": "cumplimiento",
        "TMPO.RESP.REAL": "tiempo_resp_real",
        "TMPO.RESP.ESP": "tiempo_resp_esp",
        "ZONA": "zona",
        "Zona": "zona",
        "OS FRACTTAL": "os_fracttal",
        "N° llamado": "n_llamado",
        "Atencion": "estado_atencion",
        "Facturación": "facturacion",
        "COMUNA": "comuna",
        "VALIDACIÓN DE COBRO": "cobro",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})

    # Add client label
    df["cliente"] = client_label

    # Parse dates
    for dc in ["fecha_llamado", "fecha_atencion"]:
        if dc in df.columns:
            df[dc] = pd.to_datetime(df[dc], errors="coerce")

    # Combine fecha + hora → datetimes con hora exacta
    # El Excel guarda la fecha (Fecha / Fecha de atencion) como date-only y la hora
    # en una columna separada (Hora / Hora de Atencion(FIN)).  Sin combinarlas el
    # cálculo de SLA queda en múltiplos de 24 h → resultados incorrectos.
    def _time_to_td(t) -> "pd.Timedelta":
        try:
            return pd.Timedelta(
                hours=t.hour, minutes=t.minute,
                seconds=getattr(t, "second", 0)
            )
        except Exception:
            return pd.Timedelta(0)

    if "hora_llamado" in df.columns and "fecha_llamado" in df.columns:
        df["fecha_llamado"] = df["fecha_llamado"] + df["hora_llamado"].apply(_time_to_td)

    if "hora_fin" in df.columns and "fecha_atencion" in df.columns:
        df["fecha_atencion"] = df["fecha_atencion"] + df["hora_fin"].apply(_time_to_td)

    # Filter by date
    if "fecha_llamado" in df.columns:
        df = df[df["fecha_llamado"] >= pd.Timestamp(desde)]

    # Normalize EDS code
    if "eds_occim" in df.columns:
        df["eds_occim"] = (
            df["eds_occim"]
            .astype(str)
            .str.strip()
            .str.replace(r"\.0$", "", regex=True)
        )

    # Extract priority (P1/P2/P3/P4) from Asunto if not in prioridad col
    if "prioridad" not in df.columns or df.get("prioridad", pd.Series()).isna().all():
        if "Asunto" in df.columns:
            df["prioridad"] = df["Asunto"].str.extract(r"\b(P[1-4])\b", expand=False)

    # Compute SLA compliance if not present
    if "cumplimiento" not in df.columns:
        df["cumplimiento"] = _compute_sla(df, client_label)

    # Normalize cumplimiento values to CUMPLE / NO CUMPLE
    if "cumplimiento" in df.columns:
        df["cumplimiento"] = df["cumplimiento"].astype(str).str.upper().str.strip()
        df["cumplimiento"] = df["cumplimiento"].replace({
            "SI CUMPLE PLAZO": "CUMPLE",
            "SÍ CUMPLE PLAZO": "CUMPLE",
            "SI CUMPLE": "CUMPLE",
            "SÍ CUMPLE": "CUMPLE",
            "NO CUMPLE PLAZO": "NO CUMPLE",
        })

    # Drop fully empty rows
    df = df.dropna(how="all")

    return df


def _compute_sla(df: pd.DataFrame, client_label: str = "") -> pd.Series:
    """
    Calcula cumplimiento SLA por fila usando el catálogo por cliente.
    client_label: "COPEC" | "ESMAX (Aramco)" | "SHELL (Enex)" | ""
    Si el cliente no está en el catálogo usa _SLA_DEFAULT.
    """
    result = pd.Series(["Sin datos"] * len(df), index=df.index)

    if "prioridad" not in df.columns or "fecha_llamado" not in df.columns or "fecha_atencion" not in df.columns:
        return result

    sla_cliente = SLA_HOURS.get(client_label, _SLA_DEFAULT)
    zona_col = df.get("zona", pd.Series(["Santiago"] * len(df), index=df.index))

    for idx, row in df.iterrows():
        p = str(row.get("prioridad", "")).upper()
        if p not in sla_cliente:
            continue
        zona = str(zona_col.get(idx, "Santiago")).upper().strip()
        # Zonas RM/Metropolitana → Santiago; NORTE/SUR/CENTRO/REGION → Regiones
        _is_stgo = zona in ("RM", "R.M.") or any(k in zona for k in ["SANTIAGO", "METRO"])
        _is_reg  = any(k in zona for k in ["NORTE", "SUR", "CENTRO", "REGION", "REGIONES", "RG"])
        sla_key = "Santiago" if _is_stgo else ("Regiones" if _is_reg else "Santiago")
        sla_hrs = sla_cliente[p][sla_key]

        fl = row.get("fecha_llamado")
        fa = row.get("fecha_atencion")
        if pd.isna(fl) or pd.isna(fa):
            continue
        diff_hrs = (fa - fl).total_seconds() / 3600
        result[idx] = "CUMPLE" if diff_hrs <= sla_hrs else "NO CUMPLE"

    return result


# ── Code resolution ───────────────────────────────────────────────────────────

def _eds_numeric_id(code_str) -> int:
    """Extract the numeric station ID from an EDS OCCIM code.

    Examples:
      'EE_S038' → 38   (ESMAX Fracttal code)
      'SH_11'   → 11   (Shell station code)
      'SH_2A'   → 2    (strips trailing letter — two machines, same station)
      'SH_36'   → 36
      'PART-10' → 10
    Returns -1 if no numeric ID can be extracted.
    """
    s = str(code_str).strip()
    s = re.sub(r'^(EE_S|SH_|PART-)', '', s, flags=re.IGNORECASE)
    s = re.sub(r'[A-Za-z]+$', '', s)   # strip trailing letters (A, B …)
    try:
        return int(s)
    except (ValueError, TypeError):
        return -1


def resolve_llamados_eds_codes(df_llamados: pd.DataFrame, df_eds: pd.DataFrame) -> pd.DataFrame:
    """
    ESMAX and Shell llamados identify stations with a short numeric code:
      • ESMAX  — "Codigo EDS" column holds a number like 38, 263, 45.
                 In the EDS listado, the "EDS OCCIM" column stores these same
                 numbers with an EE_S prefix: EE_S038, EE_S263, EE_S045.
      • Shell  — "File" column holds a number like 11, 36, 279.
                 In the EDS listado, the "EDS OCCIM" column stores these as
                 SH_11, SH_36, SH_279 (or SH_2A / PART-10 for shared stations).

    This function translates those numeric IDs to the Occimiano eds_occim key
    (PBR-04, SH_11, PART-10 …) so every llamado can join the master EDS list.
    COPEC llamados already use the Occimiano key and are left untouched.

    Requires that load_listado_eds() was called with the updated col_map that
    saves the raw "EDS OCCIM" Excel column as 'eds_occim_raw'.
    """
    if df_llamados.empty or df_eds.empty:
        return df_llamados
    if "eds_occim" not in df_llamados.columns or "eds_occim" not in df_eds.columns:
        return df_llamados

    # The lookup key column — prefer eds_occim_raw (raw EDS OCCIM from Excel)
    # because it carries EE_S038 / SH_36 style codes needed for numeric matching.
    raw_col = "eds_occim_raw" if "eds_occim_raw" in df_eds.columns else "eds_cliente"

    df = df_llamados.copy()

    for client_label, search_key in [("Aramco (Esmax)", "ESMAX"), ("ESMAX (Aramco)", "ESMAX"), ("SHELL (Enex)", "SHELL")]:
        mask = df.get("cliente", pd.Series(dtype=str)) == client_label
        if not mask.any():
            continue

        client_eds = df_eds[
            df_eds["cliente"].str.upper().str.contains(search_key, na=False)
        ].copy()
        if client_eds.empty or raw_col not in client_eds.columns:
            continue

        # Build: numeric_id → eds_occim (first match wins for A/B duplicate stations)
        client_eds["_num"] = client_eds[raw_col].apply(_eds_numeric_id)
        code_map: dict[int, str] = {}
        for _, row in client_eds.iterrows():
            nid = row["_num"]
            if nid > 0 and nid not in code_map:
                code_map[nid] = str(row["eds_occim"])

        def _translate(x: str) -> str:
            # Acepta tanto "38" (legacy) como "EE_S038" / "SH_36" (Supabase)
            nid = _eds_numeric_id(x)
            if nid > 0:
                return code_map.get(nid, x)
            return x

        df.loc[mask, "eds_occim"] = df.loc[mask, "eds_occim"].apply(_translate)

    return df


# ── Base de técnicos ─────────────────────────────────────────────────────────

# Mapeo COMPLETO nombre_corto_Excel → nombre_completo_Fracttal (Libro4.xlsx 2026)
# Incluye tanto los casos especiales como los estándar, para garantizar consolidación
# aunque base_tecnicos no esté disponible o tenga nombres distintos.
_TECH_SPECIAL_CASES = {
    # ── Luis Pinto (RM) ───────────────────────────────────────────────────────
    "Luis Pinto":       "Luis Alberto Pinto Jofre",
    "Juan Francisco":   "Juan Francisco Toro Jimenez",
    "Jorge Rodriguez":  "Jorge Raúl Rodríguez Fuentes",
    "Breyans Toledo":   "Breyans Andrés Toledo Quintana",
    # ── Victor Bahamonde (RM) ─────────────────────────────────────────────────
    "Victor Bahamonde": "Victor Hugo Bahamonde Bustamante",
    "Martin Flores":    "Martín Ignacio Flores Galaz",
    "Eduardo Toro":     "Eduardo Toro Ramos",
    # ── Juan Gallardo (RM) ────────────────────────────────────────────────────
    "Juan Gallardo":    "Juan Antonio Gallardo Romero",
    "Javier Hein":      "Javier Hein Pacheco",
    "Edison Carrasco":  "Edison Jhon Carrasco Navarro",
    "Ignacio Ferrari":  "Iván Ignacio Vergara Ferrari",
    # ── Carlos Avila (Coquimbo) ───────────────────────────────────────────────
    "Carlos Avila":     "Carlos Alberto Avila Palacios",
    "Edson Perez":      "Edson José Pérez Henríquez",
    "Erwin Rivera":     "Erwin Maximiliano Rivera Talamilla",
    # ── Luis Lopez (Concepción) ───────────────────────────────────────────────
    "Gaston Fuller":    "Gastón Eduardo Fuller Quilodrán",
    "Luis Lopez":       "Luis Joel Lopez Isla",
    # ── AUTEC — empresa subcontratista (consolidar variantes) ─────────────────
    "AUTEC":            "AUTEC LTDA",
    "Autec":            "AUTEC LTDA",
    "AUTEC IQUIQUE":    "AUTEC LTDA",
    "AUTEC Francisco":  "AUTEC LTDA",
}

# Nombres completos (Fracttal) de técnicos Occimiano propios — excluye subcontratistas
TECNICOS_OCCIMIANO_FULL: frozenset = frozenset(
    v for k, v in _TECH_SPECIAL_CASES.items()
    if not str(k).upper().startswith("AUTEC")
)

# Mapeo nombre_corto_Excel → nombre_completo_Fracttal (incluye AUTEC para consolidación)
TECH_NAME_MAP: dict = dict(_TECH_SPECIAL_CASES)


@st.cache_data(ttl=3600, show_spinner=False, persist="disk")
def load_base_tecnicos() -> pd.DataFrame:
    """
    Carga la base de técnicos Occimiano.
    Retorna todos los registros del Excel con columna 'short_name'
    (primer_nombre + primer_apellido) para hacer match con el Excel
    de Utilización del Tiempo.
    """
    path = FILES.get("base_tecnicos", "")
    if not os.path.exists(path):
        path = FILES.get("base_tecnicos_fallback", "")
    if not os.path.exists(path):
        return pd.DataFrame()
    df = pd.read_excel(path)
    df.columns = [str(c).strip() for c in df.columns]

    if "full_name" not in df.columns:
        nombres   = df.get("Nombres",   pd.Series(dtype=str)).fillna("").astype(str).str.strip()
        apellidos = df.get("Apellidos", pd.Series(dtype=str)).fillna("").astype(str).str.strip()
        df["full_name"] = (nombres + " " + apellidos).str.strip()

    def _short(row) -> str:
        n = str(row.get("Nombres",   "") or "").strip().split()
        a = str(row.get("Apellidos", "") or "").strip().split()
        return f"{n[0] if n else ''} {a[0] if a else ''}".strip()

    df["short_name"] = df.apply(_short, axis=1)
    return df


def build_tech_name_maps(df_tecnicos: pd.DataFrame) -> tuple[dict, dict]:
    """
    Construye dos diccionarios de mapeo de nombres:
      - excel_to_full: {nombre_en_excel → nombre_completo_en_fracttal}
      - full_to_excel: {nombre_completo → nombre_corto_excel}

    Combina:
      1. Mapeo automático (short_name de base_tecnicos)
      2. Casos especiales (_TECH_SPECIAL_CASES)
      3. Normalización de acentos para mayor tolerancia
    """
    excel_to_full: dict = {}
    full_to_excel: dict = {}

    if not df_tecnicos.empty:
        for _, row in df_tecnicos.iterrows():
            full  = str(row.get("full_name",  "") or "").strip()
            short = str(row.get("short_name", "") or "").strip()
            if full and short:
                excel_to_full[short] = full
                full_to_excel[full]  = short

    # Casos especiales (sobreescriben si coinciden)
    for excel_name, full_name in _TECH_SPECIAL_CASES.items():
        excel_to_full[excel_name] = full_name
        full_to_excel[full_name]  = excel_name

    # Variantes sin tilde para tolerancia
    def _strip_accents(s: str) -> str:
        import unicodedata
        return "".join(
            c for c in unicodedata.normalize("NFD", s)
            if unicodedata.category(c) != "Mn"
        )

    extras: dict = {}
    for k, v in list(excel_to_full.items()):
        k_norm = _strip_accents(k)
        if k_norm != k:
            extras[k_norm] = v
    excel_to_full.update(extras)

    return excel_to_full, full_to_excel


# ── KPIs ──────────────────────────────────────────────────────────────────────

def kpis_por_eds(df: pd.DataFrame) -> pd.DataFrame:
    """Group llamados by EDS and compute key KPIs."""
    if df.empty:
        return pd.DataFrame()

    grp = df.groupby("eds_occim").agg(
        total_llamados=("eds_occim", "count"),
        ultimo_llamado=("fecha_llamado", "max"),
        ultimo_tecnico=("tecnico", lambda x: x.dropna().iloc[-1] if len(x.dropna()) > 0 else None),
        p1=("prioridad", lambda x: (x.str.upper() == "P1").sum()),
        p2=("prioridad", lambda x: (x.str.upper() == "P2").sum()),
        p3=("prioridad", lambda x: (x.str.upper() == "P3").sum()),
        p4=("prioridad", lambda x: (x.str.upper() == "P4").sum()),
        cumple=("cumplimiento", lambda x: (x == "CUMPLE").sum()),
        no_cumple=("cumplimiento", lambda x: (x == "NO CUMPLE").sum()),
        cliente=("cliente", "first"),
    ).reset_index()

    total_con_sla = grp["cumple"] + grp["no_cumple"]
    grp["pct_cumplimiento"] = (
        (grp["cumple"] / total_con_sla * 100)
        .where(total_con_sla > 0, 0)
        .astype(float)
        .round(1)
    )

    return grp


# ── Utilización del tiempo ────────────────────────────────────────────────────

@st.cache_data(ttl=1800, show_spinner=False, persist="disk")
def list_utilizacion_sheets() -> list:
    """
    Retorna la lista de hojas de meses disponibles en el Excel de
    Utilización del Tiempo (hojas cuyo nombre contiene un mes en español).
    """
    path = FILES.get("utilizacion", "")
    if not os.path.exists(path):
        return []
    try:
        wb = load_workbook(path, read_only=True)
        validas = [
            s for s in wb.sheetnames
            if any(mes in s.upper() for mes in _MESES_ES.values())
        ]
        wb.close()
        return validas
    except Exception:
        return []


@st.cache_data(ttl=1800, show_spinner=False, persist="disk")
def load_utilizacion_tiempo(sheet_name: str = ""):
    """
    Lee el archivo '2026 UTILIZACIÓN DE TIEMPO.xlsx' y devuelve el cronograma
    de técnicos del mes indicado.

    Parámetros:
      sheet_name — nombre exacto de la hoja (ej. "MAYO 2026").
                   Si se omite o está vacío, auto-detecta el mes actual
                   (o el anterior si el actual no existe).

    Retorna (df, sheet_name_usado, error_msg).
    df columnas: fecha, tecnico, categoria, tareas, task_details, fuera_santiago, es_feriado
    """
    path = FILES.get("utilizacion", "")
    if not os.path.exists(path):
        return pd.DataFrame(), "", f"Archivo no encontrado: {path}"

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        return pd.DataFrame(), sheet_name, str(e)

    # ── Resolver qué hoja usar ────────────────────────────────────────────────
    if not sheet_name:
        # Auto-detectar: mes actual → mes anterior
        now = pd.Timestamp.now()
        sheet_name = _MESES_ES.get(now.month, "") + " " + str(now.year)
        if sheet_name not in wb.sheetnames:
            prev = now - pd.DateOffset(months=1)
            sheet_name = _MESES_ES.get(prev.month, "") + " " + str(prev.year)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return pd.DataFrame(), sheet_name, f"Hoja '{sheet_name}' no encontrada en el archivo"
    else:
        if sheet_name not in wb.sheetnames:
            wb.close()
            return pd.DataFrame(), sheet_name, f"Hoja '{sheet_name}' no encontrada en el archivo"

    ws = wb[sheet_name]

    # Leer nombres de técnicos desde fila 1, columnas B en adelante (hasta col 19 aprox)
    techs = []
    tech_cols = []
    for i, cell in enumerate(ws[1]):
        col_num = i + 1
        if col_num == 1:
            continue  # columna Fecha
        v = cell.value
        if not v or not str(v).strip():
            continue
        v_clean = str(v).strip()
        # Parar en las columnas de mantenciones (col U en adelante)
        if v_clean in ("FECHA REAL", "FECHA PROGR.", "EE_S212"):
            break
        techs.append(v_clean)
        tech_cols.append(col_num)

    records = []
    for row in ws.iter_rows(min_row=2, max_row=100):
        fecha_val = row[0].value
        if fecha_val is None:
            continue
        try:
            fecha = pd.Timestamp(fecha_val).normalize()
        except Exception:
            continue

        for tech, col_idx in zip(techs, tech_cols):
            cell = row[col_idx - 1]
            val = cell.value

            # Detectar color de celda
            try:
                fg = str(cell.fill.fgColor.rgb)
            except Exception:
                fg = "FFFFFFFF"

            is_outside_stgo = (fg == "FF00FF00")    # verde = fuera de Santiago
            is_feriado_finde = (fg == "FFFFC000")   # amarillo = feriado/fin de semana

            # Parsear líneas de tareas
            lines = [l.strip() for l in str(val).split("\n") if l.strip()] if val else []

            # Celdas vacías = descanso legal (finde / turno libre) → ignorar siempre
            if not lines:
                continue

            # Determinar categoría dominante
            cat = _dominant_category(lines)
            # Solo forzar Feriado si el texto lo indica explícitamente
            # (el color amarillo SOLO no es suficiente; puede ser día libre sin texto)

            # Construir lista de categorías individuales por tarea (para el desglose)
            task_details = []
            for line in lines:
                lcat = classify_task_line(line)
                if lcat:
                    task_details.append({"cat": lcat, "tarea": line})

            records.append({
                "fecha":          fecha,
                "tecnico":        tech,
                "categoria":      cat,
                "tareas":         " | ".join(lines),
                "task_details":   task_details,
                "fuera_santiago": is_outside_stgo,
                "es_feriado":     is_feriado_finde,
            })

    wb.close()
    return pd.DataFrame(records) if records else pd.DataFrame(), sheet_name, None


@st.cache_data(ttl=1800, show_spinner=False, persist="disk")
def load_mtto_realizados_planilla():
    """
    Lee los registros de mantenciones realizadas desde las columnas U-AI
    del archivo de Utilización del Tiempo del mes en curso.
    Retorna un DataFrame con: eds_code, fecha_real, fecha_prog, descripcion,
    tipo_semana, ult_mtto, provincia, comuna, tipo_mp, dias_trans, cumple
    """
    path = FILES.get("utilizacion", "")
    if not os.path.exists(path):
        return pd.DataFrame()

    now = pd.Timestamp.now()
    sheet_name = _MESES_ES.get(now.month, "") + " " + str(now.year)

    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return pd.DataFrame()

    if sheet_name not in wb.sheetnames:
        prev = now - pd.DateOffset(months=1)
        sheet_name = _MESES_ES.get(prev.month, "") + " " + str(prev.year)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return pd.DataFrame()

    ws = wb[sheet_name]

    # col U=21, V=22, W=23, X=24, Y=25, Z=26, AA=27, AB=28, AC=29, AD=30, AE=31, AH=34, AI=35
    records = []
    for row in ws.iter_rows(min_row=2, max_row=600, min_col=21, max_col=35, values_only=True):
        if all(v is None for v in row):
            continue
        # Saltar sub-encabezados repetidos
        if str(row[0] or "").strip().upper() in ("FECHA REAL", "FECHA REAL "):
            continue

        eds_code = row[2]
        if eds_code is None:
            continue
        eds_code = str(eds_code).strip().replace(".0", "")
        if not eds_code or eds_code in ("N° EDS", "EE_S212"):
            continue

        def _ts(v):
            if v is None:
                return pd.NaT
            try:
                return pd.Timestamp(v)
            except Exception:
                return pd.NaT

        records.append({
            "eds_code":   eds_code,
            "fecha_real": _ts(row[0]),
            "fecha_prog": _ts(row[1]),
            "descripcion": str(row[4] or "").strip(),
            "tipo_semana": str(row[5] or "").strip(),
            "ult_mtto":   _ts(row[6]),
            "provincia":  str(row[8] or "").strip(),
            "comuna":     str(row[9] or "").strip(),
            "tipo_mp":    str(row[10] or "").strip(),
            "dias_trans": row[13],
            "cumple":     str(row[14] or "").strip(),
        })

    wb.close()
    return pd.DataFrame(records) if records else pd.DataFrame()
